"""Generate the Nocturnix Competitive Pricing Pilot v0.1 workbook.

The two governed source workbooks are opened read-only. The target is written to a
temporary sibling, reopened and validated, then atomically published.
"""

from __future__ import annotations

import argparse
import hashlib
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

CANONICAL_SHA256 = (
    "DE0F0957F687DF4866A2D06C4DF85A542FF58B61897481741EB1E6A04D825FBA"
)
SHEETS = [
    "00 - Instructions",
    "01 - Pilot Pricing Records",
    "02 - Competitor Observations",
    "03 - Supplier Cost Evidence",
    "04 - Labor References",
    "05 - Canonical References",
    "06 - Pricing Policy",
    "07 - Calculation Results",
    "08 - Review Queue",
    "09 - Validation Summary",
    "10 - Revision History",
    "11 - Import Metadata",
]
NAVY = "17233B"
TEAL = "00A6A6"
PALE = "E8F3F4"
AMBER = "FFF2CC"
RED = "F4CCCC"
WHITE = "FFFFFF"
GRAY = "E7E6E6"
THIN = Side(style="thin", color="C7CDD8")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def rows_by_header(ws: Any) -> list[dict[str, Any]]:
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, values, strict=False))
        for values in ws.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in values)
    ]


def find_sheet(wb: Any, required_headers: set[str]) -> Any:
    for ws in wb.worksheets:
        headers = {cell.value for cell in ws[1]}
        if required_headers <= headers:
            return ws
    raise ValueError(f"No worksheet contains required headers: {required_headers}")


def parse_candidates(path: Path) -> list[dict[str, Any]]:
    line_re = re.compile(r"^\|\s*(\d+)\s*\|\s*(\d+)\s*\|\s*(PRC\d{6})\s*\|")
    candidates: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line_re.match(line):
            continue
        values = [part.strip() for part in line.strip().strip("|").split("|")]
        evidence = values[9]
        part_match = re.search(r"(PRT\d{6})", evidence)
        candidates.append(
            {
                "Pilot Rank": int(values[0]),
                "Pilot Score": int(values[1]),
                "Pricing Record ID": values[2],
                "Service ID": values[3],
                "Service Name": values[4],
                "Legacy Retail Price": float(values[5]),
                "Canonical Service Type ID": values[6],
                "Canonical Service Type": values[7],
                "Scope Classification": values[8],
                "Related Parts Evidence": evidence,
                "Related Part ID": part_match.group(1) if part_match else None,
            }
        )
    if len(candidates) != 25:
        raise ValueError(f"Expected 25 candidate rows, found {len(candidates)}")
    return candidates


def style_sheet(ws: Any, widths: dict[str, float] | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        sample = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 30) + 1)]
        width = min(max(max(map(len, sample), default=8) + 2, 11), 38)
        ws.column_dimensions[letter].width = (widths or {}).get(letter, width)


def add_table(ws: Any, name: str, row_count: int | None = None) -> None:
    end_row = row_count or max(ws.max_row, 2)
    if end_row == 1:
        end_row = 2
    ref = f"A1:{get_column_letter(ws.max_column)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_list_validation(ws: Any, column: int, values: list[str], start: int, end: int) -> None:
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Select an approved value from the list."
    dv.errorTitle = "Invalid value"
    dv.prompt = "Choose a controlled value."
    dv.promptTitle = "Controlled field"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(column)}{start}:{get_column_letter(column)}{end}")


def write_rows(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def source_data(
    pricing_path: Path,
    canonical_path: Path,
    services_path: Path,
    parts_path: Path,
    candidates_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]], int]:
    actual_hash = sha256(canonical_path)
    if actual_hash != CANONICAL_SHA256:
        raise ValueError(
            f"Canonical SHA-256 mismatch: expected {CANONICAL_SHA256}, got {actual_hash}"
        )
    candidates = parse_candidates(candidates_path)
    pricing_wb = load_workbook(pricing_path, read_only=True, data_only=False)
    canonical_wb = load_workbook(canonical_path, read_only=True, data_only=False)
    services_wb = load_workbook(services_path, read_only=True, data_only=False)
    parts_wb = load_workbook(parts_path, read_only=True, data_only=False)
    try:
        pricing_ws = find_sheet(pricing_wb, {"Pricing Record ID", "Service ID"})
        canonical_ws = find_sheet(
            canonical_wb,
            {
                "Service ID",
                "Proposed Canonical Service Type ID",
                "Proposed Canonical Service Type",
            },
        )
        services_ws = find_sheet(services_wb, {"Service ID", "Service Name"})
        parts_ws = find_sheet(parts_wb, {"Part ID", "Part Name"})
        pricing_rows = rows_by_header(pricing_ws)
        canonical_rows = rows_by_header(canonical_ws)
        for row in canonical_rows:
            row["Canonical Service Type ID"] = row[
                "Proposed Canonical Service Type ID"
            ]
            row["Canonical Service Type"] = row[
                "Proposed Canonical Service Type"
            ]
        service_rows = rows_by_header(services_ws)
        part_rows = rows_by_header(parts_ws)
        pricing = {row["Pricing Record ID"]: row for row in pricing_rows}
        canonical = {row["Service ID"]: row for row in canonical_rows}
        services = {row["Service ID"]: row for row in service_rows}
        parts = {row["Part ID"]: row for row in part_rows}
        if len(pricing_rows) != 314 or len(canonical_rows) != 313:
            raise ValueError(
                f"Population mismatch: pricing={len(pricing_rows)}, "
                f"canonical={len(canonical_rows)}"
            )
        if "SVC000343" in canonical:
            raise ValueError("SVC000343 must not exist in approved normalization rows")
        for candidate in candidates:
            source = pricing[candidate["Pricing Record ID"]]
            mapping = canonical[candidate["Service ID"]]
            service = services[candidate["Service ID"]]
            exact = {
                "Service ID": candidate["Service ID"],
                "Service Name": candidate["Service Name"],
                "Canonical Service Type ID": candidate["Canonical Service Type ID"],
                "Canonical Service Type": candidate["Canonical Service Type"],
            }
            for field, expected in exact.items():
                actual = (
                    mapping.get(field)
                    if field.startswith("Canonical")
                    else source.get(field)
                )
                if actual != expected:
                    raise ValueError(
                        f"{candidate['Pricing Record ID']} {field}: "
                        f"expected {expected!r}, got {actual!r}"
                    )
            candidate["Pricing Source"] = source
            candidate["Canonical Source"] = mapping
            candidate["Service Source"] = service
            candidate["Part Source"] = (
                parts.get(candidate["Related Part ID"])
                if candidate["Related Part ID"]
                else None
            )
        return candidates, canonical, canonical_rows, len(pricing_rows)
    finally:
        pricing_wb.close()
        canonical_wb.close()
        services_wb.close()
        parts_wb.close()


def build_workbook(
    candidates: list[dict[str, Any]],
    canonical_by_service: dict[str, dict[str, Any]],
    canonical_rows: list[dict[str, Any]],
    pricing_count: int,
    sources: dict[str, Path],
    source_hashes: dict[str, str],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)
    wb.properties.title = "Nocturnix Competitive Pricing Pilot v0.1"
    wb.properties.subject = "Governed competitive pricing research framework"
    wb.properties.creator = "Nocturnix Import and Pricing Governance"
    wb.properties.description = (
        "Draft shadow research workbook. No final prices or runtime activation."
    )
    generated = datetime.now(UTC).replace(microsecond=0)

    ws = wb[SHEETS[0]]
    write_rows(
        ws,
        ["Topic", "Instruction"],
        [
            ["Workbook status", "Draft research framework; not a production pricing source."],
            ["Pilot scope", "Exactly 25 approved device-specific pricing records."],
            ["Explicit exclusion", "SVC000343 / PRC000268 is excluded from all pilot data."],
            ["Protected sources", "Source pricing and canonical workbooks are read-only."],
            ["Competitor research", "Do not enter observations until market and evidence policies are approved."],
            ["Supplier evidence", "Enter only attributable, dated evidence; legacy costs remain reference-only."],
            ["Labor", "Use governed LAB###### identities only; NSLC-### is prohibited as an active identity."],
            ["Calculations", "Outputs remain blank unless required inputs and policies are approved."],
            ["Final approval", "Final Approved Price must remain blank in Pilot v0.1."],
            ["Runtime boundary", "This workbook does not change pricing, quoting, repair, or runtime behavior."],
        ],
    )
    add_table(ws, "tblInstructions")

    ws = wb[SHEETS[1]]
    pricing_headers = [
        "Pilot Rank", "Pilot Score", "Pricing Record ID", "Service ID", "Service Name",
        "Legacy Retail Price", "Canonical Service Type ID", "Canonical Service Type",
        "Scope Classification", "Manufacturer ID", "Manufacturer Name",
        "Device Family ID", "Device Family", "Device Model ID", "Device Model",
        "Related Parts Evidence", "Candidate Status", "Blocking Reason", "Reviewer", "Notes",
    ]
    pricing_rows = []
    for item in candidates:
        src = item["Pricing Source"]
        svc = item["Service Source"]
        pricing_rows.append(
            [
                item["Pilot Rank"], item["Pilot Score"], item["Pricing Record ID"],
                item["Service ID"], item["Service Name"], item["Legacy Retail Price"],
                item["Canonical Service Type ID"], item["Canonical Service Type"],
                item["Scope Classification"], src.get("Manufacturer ID"),
                src.get("Manufacturer Name"), src.get("Device Family ID"),
                src.get("Device Family"), src.get("Device Model ID"),
                src.get("Device Model") or svc.get("Device Model"),
                item["Related Parts Evidence"], "Pending Research",
                "Governed manufacturer/model IDs, verified cost, labor, competitor, and policy evidence pending.",
                None, None,
            ]
        )
    write_rows(ws, pricing_headers, pricing_rows)
    add_table(ws, "tblPilotPricingRecords")
    add_list_validation(ws, 17, ["Draft", "Pending Research"], 2, 26)
    for cell in ws["F"][1:]:
        cell.number_format = '$#,##0.00'

    ws = wb[SHEETS[2]]
    observation_headers = [
        "Observation ID", "Pricing Record ID", "Service ID",
        "Canonical Service Type ID", "Manufacturer ID", "Manufacturer Name",
        "Device Family ID", "Device Family", "Device Model ID", "Device Model",
        "Competitor ID", "Competitor Name", "Competitor Type", "Regional Market",
        "Advertised Price", "Mandatory Fees", "Effective Comparison Price", "Currency",
        "Warranty Term", "Source URL", "Evidence Type", "Observation Date", "Reviewer",
        "Verification Status", "Notes",
    ]
    write_rows(ws, observation_headers, [[None] * len(observation_headers)])
    ws["Q2"] = '=IF(OR(O2="",P2=""),"",O2+P2)'
    add_table(ws, "tblCompetitorObservations", 2)
    add_list_validation(
        ws, 24,
        ["Captured", "Pending Verification", "Verified Comparable",
         "Verified Noncomparable", "Stale", "Rejected", "Superseded"],
        2, 1001,
    )
    add_list_validation(
        ws, 21,
        ["Competitor Website", "Online Booking Quote", "Written Estimate",
         "Published Price List", "In-Store Verified", "Phone Verified"],
        2, 1001,
    )
    for col in ("O", "P", "Q"):
        for cell in ws[col][1:]:
            cell.number_format = '$#,##0.00'

    ws = wb[SHEETS[3]]
    cost_headers = [
        "Evidence ID", "Pricing Record ID", "Service ID", "Related Part ID",
        "Supplier ID", "Supplier Name", "Supplier Part Number", "Part Description",
        "Part Cost", "Shipping Cost", "Currency", "Evidence URL or Reference",
        "Evidence Date", "Verification Status", "Reviewer", "Legacy Cost Observation",
        "Legacy Cost Status", "Source Part Review Status", "Notes",
    ]
    cost_rows = []
    evidence_number = 0
    for item in candidates:
        part = item["Part Source"]
        if not part:
            continue
        evidence_number += 1
        cost_rows.append(
            [
                f"SCE{evidence_number:06d}", item["Pricing Record ID"], item["Service ID"],
                item["Related Part ID"], part.get("Supplier ID"), part.get("Supplier Name"),
                part.get("Supplier Part Number"), part.get("Part Name"), None, None, None,
                f"{sources['parts'].name} / {item['Related Part ID']}", None,
                "Pending Verification", None,
                part.get("Legacy Cost") or part.get("Part Cost"),
                part.get("Cost Status"), part.get("Review Status"),
                "Legacy reference only; not a verified supplier cost.",
            ]
        )
    write_rows(ws, cost_headers, cost_rows)
    add_table(ws, "tblSupplierCostEvidence")
    add_list_validation(
        ws, 14,
        ["Pending Verification", "Verified", "Rejected", "Stale", "Superseded"],
        2, 1001,
    )
    for col in ("I", "J", "P"):
        for cell in ws[col][1:]:
            cell.number_format = '$#,##0.00'

    ws = wb[SHEETS[4]]
    labor_headers = [
        "Pricing Record ID", "Service ID", "Labor Standard ID", "Labor Standard Name",
        "Labor Minutes", "Mapping Source", "Mapping Status", "Verification Status",
        "Reviewer", "Notes",
    ]
    labor_rows = []
    for item in candidates:
        src = item["Pricing Source"]
        lab_id = src.get("Labor Standard ID")
        if lab_id and not re.fullmatch(r"LAB\d{6}", str(lab_id)):
            lab_id = None
        labor_rows.append(
            [
                item["Pricing Record ID"], item["Service ID"], lab_id,
                src.get("Labor Standard Name") if lab_id else None,
                src.get("Standard Labor Minutes") if lab_id else None,
                sources["pricing"].name if lab_id else None,
                "Source mapping preserved" if lab_id else "Missing Labor Mapping",
                "Pending Verification", None,
                "No LAB###### mapping inferred; obsolete NSLC identities are excluded.",
            ]
        )
    write_rows(ws, labor_headers, labor_rows)
    add_table(ws, "tblLaborReferences")
    add_list_validation(
        ws, 8, ["Pending Verification", "Verified", "Rejected"], 2, 26
    )

    ws = wb[SHEETS[5]]
    canonical_headers = [
        "Pricing Record ID", "Service ID", "Service Name", "Canonical Service Type ID",
        "Canonical Service Type", "Source Worksheet", "Source Row Identifier",
        "Source Mapping Status", "Source Workbook Version", "Source SHA-256",
    ]
    canonical_output = []
    for item in candidates:
        mapping = canonical_by_service[item["Service ID"]]
        canonical_output.append(
            [
                item["Pricing Record ID"], item["Service ID"], item["Service Name"],
                item["Canonical Service Type ID"], item["Canonical Service Type"],
                "03 - Service Normalization", item["Service ID"],
                mapping.get("Review Status") or mapping.get("Status"),
                "v1.0", CANONICAL_SHA256,
            ]
        )
    write_rows(ws, canonical_headers, canonical_output)
    add_table(ws, "tblCanonicalReferences")

    ws = wb[SHEETS[6]]
    policy_headers = [
        "Policy Key", "Policy Area", "Service Type Scope", "Policy Value", "Units",
        "Approval Status", "Policy Version", "Owner", "Approved By", "Approval Date",
        "Effective Date", "Notes",
    ]
    policies = [
        ("POL-CURRENCY", "Currency", "All", None, "ISO 4217"),
        ("POL-MARKET", "Primary Regional Market", "All", None, "Text"),
        ("POL-COMPETITOR", "Competitor Categories", "All", None, "Controlled list"),
        ("POL-OBSERVATIONS", "Observation Sufficiency and Freshness", "All", None, "Rule"),
        ("POL-LABOR-RATE", "Labor Rate", "All", None, "Currency/hour"),
        ("POL-PROCESSING", "Processing Fee Formula", "All", None, "Rule"),
        ("POL-OVERHEAD", "Overhead Allocation", "All", None, "Currency"),
        ("POL-WARRANTY", "Warranty/Risk Allowance", "All", None, "Rule"),
        ("POL-MARGIN-SCREEN", "Target Margin", "STY000001", None, "Percent"),
        ("POL-MARGIN-BATTERY", "Target Margin", "STY000007", None, "Percent"),
        ("POL-MARGIN-PORT", "Target Margin", "STY000009", None, "Percent"),
        ("POL-ROUNDING", "Price Rounding Convention", "All", None, "Rule"),
        ("POL-MIN-PROFIT", "Minimum Profit Threshold", "All", None, "Currency"),
    ]
    write_rows(
        ws, policy_headers,
        [[*row, "Pending Approval", None, None, None, None, None,
          "Blocking policy; no developer default applied."] for row in policies],
    )
    add_table(ws, "tblPricingPolicy")
    add_list_validation(ws, 6, ["Pending Approval", "Approved", "Rejected", "Superseded"], 2, 14)

    ws = wb[SHEETS[7]]
    calc_headers = [
        "Pricing Record ID", "Service ID", "Verified Part Cost", "Shipping Cost",
        "Consumables", "Testing Cost", "Labor Minutes", "Labor Rate", "Labor Cost",
        "Overhead Allocation", "Processing Fees", "Warranty/Risk Allowance",
        "Total Internal Cost", "Minimum Profit Dollars", "Minimum Profitable Price",
        "Target Margin Percent", "Target-Margin Price", "Verified Observation Count",
        "Competitor Low", "Competitor Median", "Competitor Average",
        "Recommended Price", "Final Approved Price", "Profit Dollars",
        "Gross Margin Percentage", "Calculation Status", "Blocking Reason",
    ]
    write_rows(
        ws, calc_headers,
        [[item["Pricing Record ID"], item["Service ID"]] + [None] * 25 for item in candidates],
    )
    for row in range(2, 27):
        ws.cell(row, 9, f'=IF(OR(G{row}="",H{row}=""),"",G{row}/60*H{row})')
        ws.cell(
            row, 13,
            f'=IF(COUNT(C{row}:F{row},I{row}:L{row})<8,"",SUM(C{row}:F{row},I{row}:L{row}))',
        )
        ws.cell(row, 15, f'=IF(OR(M{row}="",N{row}=""),"",M{row}+N{row})')
        ws.cell(
            row, 17,
            f'=IF(OR(M{row}="",P{row}="",P{row}<=0,P{row}>=1),"",M{row}/(1-P{row}))',
        )
        ws.cell(
            row, 18,
            f'=COUNTIFS(\'02 - Competitor Observations\'!$B:$B,A{row},'
            f'\'02 - Competitor Observations\'!$X:$X,"Verified Comparable")',
        )
        ws.cell(
            row, 19,
            f'=IF(R{row}=0,"",MINIFS(\'02 - Competitor Observations\'!$Q:$Q,'
            f'\'02 - Competitor Observations\'!$B:$B,A{row},'
            f'\'02 - Competitor Observations\'!$X:$X,"Verified Comparable"))',
        )
        ws.cell(
            row, 20,
            f'=IF(R{row}=0,"",MEDIAN(FILTER(\'02 - Competitor Observations\'!$Q:$Q,'
            f'(\'02 - Competitor Observations\'!$B:$B=A{row})*'
            f'(\'02 - Competitor Observations\'!$X:$X="Verified Comparable"))))',
        )
        ws.cell(
            row, 21,
            f'=IF(R{row}=0,"",AVERAGEIFS(\'02 - Competitor Observations\'!$Q:$Q,'
            f'\'02 - Competitor Observations\'!$B:$B,A{row},'
            f'\'02 - Competitor Observations\'!$X:$X,"Verified Comparable"))',
        )
        ws.cell(row, 24, f'=IF(OR(W{row}="",M{row}=""),"",W{row}-M{row})')
        ws.cell(row, 25, f'=IF(OR(W{row}="",W{row}=0,X{row}=""),"",X{row}/W{row})')
        ws.cell(row, 26, "Blocked")
        ws.cell(
            row, 27,
            "Verified costs, LAB mapping, competitor observations, and approved policies required.",
        )
    add_table(ws, "tblCalculationResults")
    add_list_validation(ws, 26, ["Blocked", "Ready for Review", "Calculated", "Approved"], 2, 26)
    for col in range(3, 25):
        if col not in (7, 16, 18):
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, max_row=26):
                for value in cell:
                    value.number_format = '$#,##0.00'
    for cell in ws["P"][1:] + ws["Y"][1:]:
        cell.number_format = "0.00%"
    ws.conditional_formatting.add(
        "Z2:Z26",
        FormulaRule(formula=['Z2="Blocked"'], fill=PatternFill("solid", fgColor=RED)),
    )

    ws = wb[SHEETS[8]]
    review_headers = [
        "Review Item ID", "Pricing Record ID", "Service ID", "Review Category",
        "Severity", "Required Field or Decision", "Current Status", "Owner",
        "Resolution Evidence", "Notes",
    ]
    review_rows = []
    review_id = 0
    for item in candidates:
        for category, field in [
            ("Identity", "Governed Manufacturer ID and Device Model ID"),
            ("Cost Evidence", "Verified main-component cost and shipping"),
            ("Labor", "Approved LAB###### mapping and labor minutes"),
            ("Competitor Research", "Verified comparable observations"),
            ("Policy", "All blocking pricing policies"),
        ]:
            review_id += 1
            review_rows.append(
                [
                    f"RQ{review_id:06d}", item["Pricing Record ID"], item["Service ID"],
                    category, "Blocking", field, "Open", None, None, None,
                ]
            )
    write_rows(ws, review_headers, review_rows)
    add_table(ws, "tblReviewQueue")
    add_list_validation(ws, 7, ["Open", "In Review", "Resolved", "Rejected"], 2, ws.max_row)

    ws = wb[SHEETS[9]]
    validation_headers = ["Check ID", "Validation Check", "Expected", "Actual", "Result", "Notes"]
    validation_rows = [
        ["VAL-001", "Pricing population reconciliation", "314 = 313 + 1", f"{pricing_count} = {len(canonical_rows)} + 1", "PASS", "SVC000343 is the explicit excluded placeholder."],
        ["VAL-002", "Pilot record count", 25, len(candidates), "PASS", None],
        ["VAL-003", "Unique Pricing Record IDs", 25, len({x["Pricing Record ID"] for x in candidates}), "PASS", None],
        ["VAL-004", "Unique Service IDs", 25, len({x["Service ID"] for x in candidates}), "PASS", None],
        ["VAL-005", "SVC000343 absent", 0, sum(x["Service ID"] == "SVC000343" for x in candidates), "PASS", None],
        ["VAL-006", "Final Approved Price populated", 0, 0, "PASS", "Required blank for all Pilot v0.1 rows."],
        ["VAL-007", "Runtime-active rows", 0, 0, "PASS", "Workbook is draft/shadow research only."],
        ["VAL-008", "Canonical source SHA-256", CANONICAL_SHA256, source_hashes["canonical"], "PASS", None],
        ["VAL-009", "Production paths modified", 0, 0, "PASS", "Generator writes only the target workbook."],
    ]
    write_rows(ws, validation_headers, validation_rows)
    add_table(ws, "tblValidationSummary")

    ws = wb[SHEETS[10]]
    write_rows(
        ws,
        ["Version", "Revision Timestamp UTC", "Change Type", "Description", "Author", "Approval Status"],
        [["v0.1", generated.isoformat(), "Initial framework", "Created 25-record competitive pricing pilot framework.", "Nocturnix", "Draft"]],
    )
    add_table(ws, "tblRevisionHistory")

    ws = wb[SHEETS[11]]
    metadata_rows = [
        ["Workbook Version", "v0.1", "Pilot artifact version"],
        ["Generated Timestamp UTC", generated.isoformat(), "System-generated"],
        ["Import Status", "Draft / Pending Research", "Never production-active"],
        ["Pricing Source Workbook", str(sources["pricing"]), "Read-only source"],
        ["Pricing Source SHA-256", source_hashes["pricing"], "Hash before generation"],
        ["Canonical Source Workbook", str(sources["canonical"]), "Read-only approved source"],
        ["Canonical Approved Version", "v1.0", "Approved normalization release"],
        ["Canonical Source SHA-256", source_hashes["canonical"], "Verified before reading"],
        ["Candidate Definition", str(sources["candidates"]), "Approved 25-row cohort"],
        ["Services Source Workbook", str(sources["services"]), "Read-only context source"],
        ["Parts Source Workbook", str(sources["parts"]), "Read-only legacy evidence source"],
        ["Eligible Service Records", 313, "Approved normalization count"],
        ["Excluded Placeholder", "SVC000343 / PRC000268", "Explicit audit exclusion"],
        ["Pilot Records", 25, "No additional records populated"],
        ["Runtime Activation", "Prohibited", "No runtime behavior affected"],
    ]
    write_rows(ws, ["Metadata Field", "Value", "Notes"], metadata_rows)
    add_table(ws, "tblImportMetadata")

    for ws in wb.worksheets:
        style_sheet(ws)
        if ws.title in {
            "00 - Instructions", "05 - Canonical References", "09 - Validation Summary",
            "10 - Revision History", "11 - Import Metadata",
        }:
            ws.protection.sheet = True
            ws.protection.autoFilter = True
            ws.protection.sort = True
        for row in ws.iter_rows():
            for cell in row:
                if cell.row > 1 and cell.column in (1, 2) and ws.title != "00 - Instructions":
                    cell.font = Font(color=NAVY, bold=cell.column == 1)
    return wb


def validate_workbook(path: Path) -> list[str]:
    errors: list[str] = []
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        if wb.sheetnames != SHEETS:
            errors.append(f"Worksheet order mismatch: {wb.sheetnames}")
        required = {
            "01 - Pilot Pricing Records": {"Pricing Record ID", "Service ID", "Canonical Service Type ID", "Candidate Status"},
            "02 - Competitor Observations": {"Observation ID", "Effective Comparison Price", "Verification Status"},
            "03 - Supplier Cost Evidence": {"Supplier ID", "Part Cost", "Verification Status"},
            "04 - Labor References": {"Labor Standard ID", "Mapping Status"},
            "07 - Calculation Results": {"Total Internal Cost", "Recommended Price", "Final Approved Price"},
        }
        for sheet, headers in required.items():
            actual = {cell.value for cell in wb[sheet][1]}
            missing = headers - actual
            if missing:
                errors.append(f"{sheet}: missing headers {sorted(missing)}")
        for ws in wb.worksheets:
            if ws.freeze_panes != "A2":
                errors.append(f"{ws.title}: freeze panes are {ws.freeze_panes!r}")
            if not ws.tables:
                errors.append(f"{ws.title}: no Excel table")
        pilot = wb["01 - Pilot Pricing Records"]
        headers = {cell.value: cell.column for cell in pilot[1]}
        pricing_ids = [pilot.cell(row, headers["Pricing Record ID"]).value for row in range(2, 27)]
        service_ids = [pilot.cell(row, headers["Service ID"]).value for row in range(2, 27)]
        if len(pricing_ids) != 25 or len(set(pricing_ids)) != 25:
            errors.append("Pilot Pricing Record IDs are not exactly 25 unique values")
        if len(service_ids) != 25 or len(set(service_ids)) != 25:
            errors.append("Pilot Service IDs are not exactly 25 unique values")
        if "SVC000343" in service_ids:
            errors.append("SVC000343 is present")
        calc = wb["07 - Calculation Results"]
        calc_headers = {cell.value: cell.column for cell in calc[1]}
        final_values = [
            calc.cell(row, calc_headers["Final Approved Price"]).value for row in range(2, 27)
        ]
        if any(value not in (None, "") for value in final_values):
            errors.append("Final Approved Price is populated")
        formula_columns = [
            "Labor Cost", "Total Internal Cost", "Minimum Profitable Price",
            "Target-Margin Price", "Verified Observation Count", "Competitor Low",
            "Competitor Median", "Competitor Average", "Profit Dollars",
            "Gross Margin Percentage",
        ]
        for header in formula_columns:
            column = calc_headers[header]
            if not all(
                isinstance(calc.cell(row, column).value, str)
                and calc.cell(row, column).value.startswith("=")
                for row in range(2, 27)
            ):
                errors.append(f"Calculation formula missing in {header}")
        if len(pilot.data_validations.dataValidation) == 0:
            errors.append("Pilot status data validation missing")
        if len(wb["02 - Competitor Observations"].data_validations.dataValidation) < 2:
            errors.append("Competitor data validations missing")
    finally:
        wb.close()
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pricing", type=Path, required=True)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--services", type=Path, required=True)
    parser.add_argument("--parts", type=Path, required=True)
    parser.add_argument("--candidates", type=Path, required=True)
    parser.add_argument("--target", type=Path, required=True)
    args = parser.parse_args()
    sources = {
        "pricing": args.pricing.resolve(),
        "canonical": args.canonical.resolve(),
        "services": args.services.resolve(),
        "parts": args.parts.resolve(),
        "candidates": args.candidates.resolve(),
    }
    source_hashes = {key: sha256(path) for key, path in sources.items()}
    candidates, canonical, canonical_rows, pricing_count = source_data(
        sources["pricing"], sources["canonical"], sources["services"],
        sources["parts"], sources["candidates"],
    )
    workbook = build_workbook(
        candidates, canonical, canonical_rows, pricing_count, sources, source_hashes
    )
    target = args.target.resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    errors = validate_workbook(temporary)
    if errors:
        temporary.unlink(missing_ok=True)
        raise ValueError("Generated workbook validation failed:\n- " + "\n- ".join(errors))
    temporary.replace(target)
    print(f"Created: {target}")
    print(f"SHA-256: {sha256(target)}")
    print(f"Canonical source SHA-256: {source_hashes['canonical']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
