import hashlib
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass(frozen=True)
class AuditSummary:
    staging_hash: str
    raw_hash: str
    canonical_hash: str
    raw_hash_after: str
    staging_hash_after: str
    canonical_hash_after: str
    hash_unchanged: bool
    row_count: int
    old_counts: Counter
    staging_counts: Counter
    changed_rows: list[dict]
    manual_reason_counts: Counter
    primary_dest_counts: Counter
    secondary_inventory_count: int
    unique_sku_rows: int
    duplicate_sku_group_count: int
    duplicate_sku_rows: int
    exact_duplicate_patterns: int
    exact_duplicate_rows: int
    exact_duplicate_excess: int
    conflicting_sku_rows: int
    multi_supplier_rows: int
    multi_condition_rows: int
    missing_sku_rows: int
    classification_counts: Counter
    comparison_lines: list[str]


STAGING_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Legacy_Catalog_Staging_Preview_v1.xlsx"
)
RAW_PATH = Path(r"D:\Business Portal\300_Pricing\Legacy\Raw Import Data.xlsx")
CANONICAL_PATH = Path(
    r"D:\Projects\Nocturnix Repair Platform\Data\Nocturnix_Master_Database.xlsm"
)
OUTPUT_PATH = Path("Output") / "legacy_staging_verification_report.txt"

PHYSICAL_CATEGORIES = {"Part", "Device", "Tool", "Accessory"}
REPAIR_CATEGORY = "Repair"


def compute_sha256(path: Path) -> str:
    data = path.read_bytes()
    return hashlib.sha256(data).hexdigest()


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_sku(value: object) -> str:
    return normalize(value)


def normalize_condition(value: object) -> str:
    return normalize(value).title()


def normalize_price(value: object) -> object:
    if value is None or normalize(value) == "":
        return None
    text = normalize(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return text


def normalize_cost(value: object) -> object:
    if value is None or normalize(value) == "":
        return None
    text = normalize(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return text


def classify_sku_group(sku: str, group: list[dict]) -> str:
    if sku == "":
        return "Missing SKU"
    if len(group) == 1:
        return "Unique SKU"
    attrs = {
        (
            normalize(rec["Record Category"]),
            normalize(rec["Legacy Name"]),
            normalize(rec["Legacy Manufacturer"]),
            normalize(rec["Legacy Type"]),
        )
        for rec in group
    }
    suppliers = {
        normalize(rec["Legacy Supplier"])
        for rec in group
        if normalize(rec["Legacy Supplier"]) != ""
    }
    conditions = {
        normalize_condition(rec["Legacy Condition"])
        for rec in group
        if normalize_condition(rec["Legacy Condition"]) != ""
    }
    if len(attrs) == 1:
        if len(suppliers) > 1:
            return "Multi-Supplier Variant"
        if len(conditions) > 1:
            return "Multi-Condition Variant"
        return "Same SKU / Same Item"
    return "Same SKU / Conflicting Item"


def old_disposition(record: dict) -> str:
    price = normalize_price(record["Legacy Retail Price"])
    cost = normalize_cost(record["Legacy Cost"])
    missing_manufacturer = normalize(record["Legacy Manufacturer"]) == ""
    missing_supplier = normalize(record["Legacy Supplier"]) == ""
    both_zero = (
        isinstance(price, float)
        and price == 0.0
        and isinstance(cost, float)
        and cost == 0.0
    )
    if both_zero:
        return "Requires Manual Review"
    if missing_manufacturer or missing_supplier:
        return "Mappable After Lookup Enrichment"
    return "Automatically Mappable"


def reasons_for_record(record: dict, classification: str) -> list[str]:
    reasons = []
    price = normalize_price(record["Legacy Retail Price"])
    cost = normalize_cost(record["Legacy Cost"])
    if classification == "Same SKU / Conflicting Item":
        reasons.append("conflicting SKU")
    if classification == "Multi-Supplier Variant":
        reasons.append("multi-supplier variant")
    if classification == "Multi-Condition Variant":
        reasons.append("multi-condition variant")
    if classification == "Missing SKU":
        reasons.append("missing SKU")
    if (
        isinstance(price, float)
        and price == 0.0
        and isinstance(cost, float)
        and cost == 0.0
    ):
        reasons.append("both Price and Cost zero")
    if normalize(record["Legacy Manufacturer"]) == "":
        reasons.append("missing Manufacturer")
    supplier_blank = normalize(record["Legacy Supplier"]) == ""
    if supplier_blank and normalize(record["Record Category"]) in PHYSICAL_CATEGORIES:
        reasons.append("missing Supplier")
    if isinstance(price, str) or isinstance(cost, str):
        reasons.append("invalid monetary type")
    if normalize(record["Record Category"]) not in {
        *PHYSICAL_CATEGORIES,
        REPAIR_CATEGORY,
    }:
        reasons.append("unsupported category")
    if not reasons:
        reasons.append("no explicit reason")
    return reasons


def load_records(path: Path, sheet_name: str) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} missing from {path}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize(value) for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def load_raw_exact_duplicates(path: Path) -> tuple[int, int, int]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    raw_rows = [tuple(row) for row in rows[1:]]
    raw_group = Counter(raw_rows)
    exact_patterns = sum(1 for count in raw_group.values() if count > 1)
    exact_rows = sum(count for count in raw_group.values() if count > 1)
    excess = sum(count - 1 for count in raw_group.values() if count > 1)
    return exact_patterns, exact_rows, excess


def summarize() -> AuditSummary:
    staging_hash = compute_sha256(STAGING_PATH)
    raw_hash = compute_sha256(RAW_PATH)
    canonical_hash = compute_sha256(CANONICAL_PATH)

    records = load_records(STAGING_PATH, "01 - All Staging Records")
    if not records:
        raise RuntimeError("No records found in staging sheet")

    sku_groups: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        sku_groups[normalize_sku(rec["Legacy SKU"])].append(rec)

    classification_by_sku = {
        sku: classify_sku_group(sku, group)
        for sku, group in sku_groups.items()
    }
    classification_counts = Counter(
        classification_by_sku[normalize_sku(rec["Legacy SKU"])]
        for rec in records
    )

    old_counts = Counter(old_disposition(rec) for rec in records)
    staging_counts = Counter(normalize(rec["Review Status"]) for rec in records)

    duplicate_sku_groups = {
        sku: group
        for sku, group in sku_groups.items()
        if sku != "" and len(group) > 1
    }
    duplicate_sku_group_count = len(duplicate_sku_groups)
    duplicate_sku_rows = sum(len(group) for group in duplicate_sku_groups.values())

    missing_sku_rows = sum(1 for rec in records if normalize_sku(rec["Legacy SKU"]) == "")
    unique_sku_rows = classification_counts["Unique SKU"]
    conflicting_sku_rows = classification_counts["Same SKU / Conflicting Item"]
    multi_supplier_rows = classification_counts["Multi-Supplier Variant"]
    multi_condition_rows = classification_counts["Multi-Condition Variant"]

    exact_duplicate_patterns, exact_duplicate_rows, exact_duplicate_excess = (
        load_raw_exact_duplicates(RAW_PATH)
    )

    changed_rows = []
    for rec in records:
        original = old_disposition(rec)
        current = normalize(rec["Review Status"])
        if original != current:
            classification = classification_by_sku[normalize_sku(rec["Legacy SKU"])]
            reasons = reasons_for_record(rec, classification)
            changed_rows.append(
                {
                    "Source Row Number": normalize(rec.get("Source Row Number")),
                    "Legacy SKU": normalize(rec["Legacy SKU"]),
                    "Category": normalize(rec["Record Category"]),
                    "Name": normalize(rec["Legacy Name"]),
                    "Duplicate Classification": classification,
                    "Old Disposition": original,
                    "New Disposition": current,
                    "Reasons": reasons,
                }
            )

    manual_rows = [rec for rec in records if normalize(rec["Review Status"]) == "Requires Manual Review"]
    manual_reason_counts = Counter()
    for rec in manual_rows:
        classification = classification_by_sku[normalize_sku(rec["Legacy SKU"])]
        for reason in reasons_for_record(rec, classification):
            manual_reason_counts[reason] += 1

    primary_dest_counts = Counter(normalize(rec["Destination Dataset"]) for rec in records)
    secondary_inventory_rows = [
        rec
        for rec in records
        if normalize(rec["Record Category"]) in PHYSICAL_CATEGORIES
        and (normalize(rec["Legacy Serial Number"]) or normalize(rec["Legacy Bin"]))
    ]
    secondary_inventory_count = len(secondary_inventory_rows)

    staging_hash_after = compute_sha256(STAGING_PATH)
    raw_hash_after = compute_sha256(RAW_PATH)
    canonical_hash_after = compute_sha256(CANONICAL_PATH)
    hash_unchanged = (
        staging_hash == staging_hash_after
        and raw_hash == raw_hash_after
        and canonical_hash == canonical_hash_after
    )

    return AuditSummary(
        staging_hash=staging_hash,
        raw_hash=raw_hash,
        canonical_hash=canonical_hash,
        staging_hash_after=staging_hash_after,
        raw_hash_after=raw_hash_after,
        canonical_hash_after=canonical_hash_after,
        hash_unchanged=hash_unchanged,
        row_count=len(records),
        old_counts=old_counts,
        staging_counts=staging_counts,
        changed_rows=changed_rows,
        manual_reason_counts=manual_reason_counts,
        primary_dest_counts=primary_dest_counts,
        secondary_inventory_count=secondary_inventory_count,
        unique_sku_rows=unique_sku_rows,
        duplicate_sku_group_count=duplicate_sku_group_count,
        duplicate_sku_rows=duplicate_sku_rows,
        exact_duplicate_patterns=exact_duplicate_patterns,
        exact_duplicate_rows=exact_duplicate_rows,
        exact_duplicate_excess=exact_duplicate_excess,
        conflicting_sku_rows=conflicting_sku_rows,
        multi_supplier_rows=multi_supplier_rows,
        multi_condition_rows=multi_condition_rows,
        missing_sku_rows=missing_sku_rows,
        classification_counts=classification_counts,
        comparison_lines=[],
    )


def render_report(summary: AuditSummary) -> str:
    lines: list[str] = []
    lines.append("Legacy Staging Preview Verification Report")
    lines.append("=====================================")
    lines.append("")
    lines.append("Hashes before execution:")
    lines.append(f"  Raw:       {summary.raw_hash}")
    lines.append(f"  Staging:   {summary.staging_hash}")
    lines.append(f"  Canonical: {summary.canonical_hash}")
    lines.append("")
    lines.append("Hashes after execution:")
    lines.append(f"  Raw:       {summary.raw_hash_after}")
    lines.append(f"  Staging:   {summary.staging_hash_after}")
    lines.append(f"  Canonical: {summary.canonical_hash_after}")
    lines.append("")
    lines.append(f"Hashes unchanged: {summary.hash_unchanged}")
    lines.append("")
    lines.append(f"Total staging rows: {summary.row_count}")
    lines.append("")
    lines.append("Disposition counts:")
    for key, value in summary.old_counts.items():
        lines.append(f"  Old disposition {key}: {value}")
    for key, value in summary.staging_counts.items():
        lines.append(f"  Staging disposition {key}: {value}")
    lines.append(f"  Changed rows: {len(summary.changed_rows)}")
    lines.append("")
    lines.append("Changed rows with reasons:")
    for row in summary.changed_rows:
        lines.append(
            f"  Source {row['Source Row Number']} SKU={row['Legacy SKU']} category={row['Category']} "
            f"old={row['Old Disposition']} new={row['New Disposition']} classification={row['Duplicate Classification']} "
            f"reasons={row['Reasons']}"
        )
    lines.append("")
    lines.append("Manual review reason counts:")
    for key, value in summary.manual_reason_counts.items():
        lines.append(f"  {key}: {value}")
    lines.append("")
    lines.append("Primary destination counts:")
    for key, value in summary.primary_dest_counts.items():
        lines.append(f"  {key}: {value}")
    lines.append("")
    lines.append(f"Secondary inventory candidate rows: {summary.secondary_inventory_count}")
    lines.append("")
    lines.append("SKU classification totals:")
    lines.append(f"  Unique SKU rows: {summary.unique_sku_rows}")
    lines.append(f"  Duplicate SKU groups: {summary.duplicate_sku_group_count}")
    lines.append(f"  Rows in duplicate SKU groups: {summary.duplicate_sku_rows}")
    lines.append(f"  Exact duplicate patterns: {summary.exact_duplicate_patterns}")
    lines.append(f"  Rows in exact duplicate patterns: {summary.exact_duplicate_rows}")
    lines.append(f"  Excess exact duplicates: {summary.exact_duplicate_excess}")
    lines.append(f"  Conflicting SKU rows: {summary.conflicting_sku_rows}")
    lines.append(f"  Multi-supplier rows: {summary.multi_supplier_rows}")
    lines.append(f"  Multi-condition rows: {summary.multi_condition_rows}")
    lines.append(f"  Missing SKU rows: {summary.missing_sku_rows}")
    lines.append("")
    lines.append("Classification counts:")
    for key, value in summary.classification_counts.items():
        lines.append(f"  {key}: {value}")
    lines.append("")
    lines.append("Report saved to Output/legacy_staging_verification_report.txt")
    return "\n".join(lines)


def main() -> int:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    summary = summarize()
    report = render_report(summary)
    OUTPUT_PATH.write_text(report, encoding="utf-8")
    print(report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
