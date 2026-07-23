"""Generate the standalone Master Compatibility Catalog v1 review workbook.

Inputs are opened read-only. The output is written to a temporary sibling,
validated after reopening, and atomically published. This script never writes
to the canonical database or another source workbook.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
DEVICES_PATH = WORKING_DIR / "Nocturnix_Master_Devices_Catalog_v1.xlsx"
SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
PARTS_PATH = WORKING_DIR / "Nocturnix_Master_Parts_Catalog_v1.xlsx"
PRICING_PATH = WORKING_DIR / "Nocturnix_Master_Pricing_Catalog_v1.xlsx"
PROPOSAL_PATH = (
    WORKING_DIR / "Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = WORKING_DIR / "Nocturnix_Master_Compatibility_Catalog_v1.xlsx"
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(
    f"{OUTPUT_PATH.stem}.tmp{OUTPUT_PATH.suffix}"
)

PROTECTED_PATHS = (
    DEVICES_PATH,
    SERVICES_PATH,
    PARTS_PATH,
    PRICING_PATH,
    PROPOSAL_PATH,
    CANONICAL_PATH,
)
CANONICAL_SHEET = "35 Compatibility Matrix"
IMPORT_BATCH_ID = "MASTER-COMPATIBILITY-V1-REVIEW"
NAMESPACE_AUTHORITY = "ADR-010"
COMPATIBILITY_ID_PATTERN = re.compile(r"^CMP\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Compatibility Records",
    "02 - Relationship Types",
    "03 - Compatibility Levels",
    "04 - Evidence Types",
    "05 - Devices",
    "06 - Services",
    "07 - Parts",
    "08 - Family Relationships",
    "09 - Model Relationships",
    "10 - Variant Relationships",
    "11 - Unresolved Review",
    "12 - Validation Summary",
    "13 - Revision History",
    "14 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterCompatibilityInstructions",
    "01 - Compatibility Records": "tblMasterCompatibilityCatalog",
    "02 - Relationship Types": "tblCompatibilityRelationshipTypes",
    "03 - Compatibility Levels": "tblCompatibilityLevels",
    "04 - Evidence Types": "tblCompatibilityEvidenceTypes",
    "05 - Devices": "tblCompatibilityDevices",
    "06 - Services": "tblCompatibilityServices",
    "07 - Parts": "tblCompatibilityParts",
    "08 - Family Relationships": "tblCompatibilityFamilyReview",
    "09 - Model Relationships": "tblCompatibilityModelReview",
    "10 - Variant Relationships": "tblCompatibilityVariantReview",
    "11 - Unresolved Review": "tblCompatibilityUnresolvedReview",
    "12 - Validation Summary": "tblMasterCompatibilityValidation",
    "13 - Revision History": "tblMasterCompatibilityRevisionHistory",
    "14 - Import Metadata": "tblMasterCompatibilityImportMetadata",
}
COMPATIBILITY_HEADERS = [
    "Compatibility ID",
    "Relationship Type",
    "Device ID",
    "Device Family Code",
    "Device Variant",
    "Service ID",
    "Part ID",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Name",
    "Service Name",
    "Part Name",
    "Compatibility Level",
    "Compatibility Status",
    "Evidence Type",
    "Evidence Source",
    "Evidence Detail",
    "Confidence",
    "Requires Manual Review",
    "Active",
    "Effective Date",
    "Expiration Date",
    "Review Status",
    "Reviewer",
    "Reviewer Notes",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Created At",
    "Updated At",
]
UNRESOLVED_HEADERS = [
    "Candidate Type",
    "Device ID",
    "Device Family Code",
    "Device Name",
    "Service ID",
    "Service Name",
    "Part ID",
    "Part Name",
    "Missing Evidence",
    "Ambiguity Reason",
    "Required Action",
    "Review Priority",
    "Review Status",
    "Reviewer Notes",
]
RELATIONSHIP_TYPES = [
    "Device to Service",
    "Device to Part",
    "Device Family to Service",
    "Device Family to Part",
    "Device Variant to Service",
    "Device Variant to Part",
]
COMPATIBILITY_LEVELS = [
    "Family Level",
    "Model Level",
    "Variant Level",
    "Universal",
    "Not Applicable",
    "Unresolved",
]
EVIDENCE_TYPES = [
    "Explicit Source Match",
    "Canonical Relationship",
    "Exact Model Match",
    "Exact Manufacturer and Model",
    "Family-Level Evidence",
    "Legacy Name Evidence",
    "Manual Research Required",
    "No Reliable Evidence",
]
COMPATIBILITY_STATUSES = [
    "Pending Review",
    "Proposed",
    "Confirmed",
    "Rejected",
    "Archived",
]
REVIEW_STATUSES = [
    "Pending Review",
    "Pending Evidence Review",
    "Pending Device Review",
    "Pending Service Review",
    "Pending Part Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
CONFIDENCE_VALUES = ["Unassessed", "Low", "Medium", "High"]
YES_NO_VALUES = ["Yes", "No"]
MODEL_NOISE = {
    "assembly",
    "battery",
    "camera",
    "charging",
    "digitizer",
    "front",
    "glass",
    "lcd",
    "oled",
    "port",
    "repair",
    "replacement",
    "screen",
    "service",
}
PLACEHOLDER_MANUFACTURERS = {
    "",
    "n a",
    "na",
    "none",
    "unknown",
    "tbd",
}
GENERIC_PRECISE_NAMES = {
    "battery",
    "camera",
    "charging port",
    "screen adhesive",
}
DEFINED_NAME_BY_HEADER = {
    "Relationship Type": "DV_RelationshipTypes",
    "Device ID": "DV_DeviceIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Service ID": "DV_ServiceIDs",
    "Part ID": "DV_PartIDs",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Compatibility Level": "DV_CompatibilityLevels",
    "Compatibility Status": "DV_CompatibilityStatuses",
    "Evidence Type": "DV_EvidenceTypes",
    "Confidence": "DV_ConfidenceValues",
    "Requires Manual Review": "DV_YesNo",
    "Active": "DV_YesNo",
    "Review Status": "DV_ReviewStatuses",
}
DEFINED_NAME_SPECS = {
    "DV_RelationshipTypes": ("02 - Relationship Types", 1),
    "DV_CompatibilityLevels": ("03 - Compatibility Levels", 1),
    "DV_CompatibilityStatuses": ("03 - Compatibility Levels", 2),
    "DV_ReviewStatuses": ("03 - Compatibility Levels", 3),
    "DV_ConfidenceValues": ("03 - Compatibility Levels", 4),
    "DV_YesNo": ("03 - Compatibility Levels", 5),
    "DV_DeviceFamilyCodes": ("03 - Compatibility Levels", 6),
    "DV_ManufacturerIDs": ("03 - Compatibility Levels", 7),
    "DV_EvidenceTypes": ("04 - Evidence Types", 1),
    "DV_DeviceIDs": ("05 - Devices", 1),
    "DV_ServiceIDs": ("06 - Services", 1),
    "DV_PartIDs": ("07 - Parts", 1),
}
PROHIBITED_HEADERS = {
    "Price",
    "Cost",
    "Supplier Cost",
    "Stock",
    "Quantity",
    "Inventory",
    "Final Approval",
    "Final Customer Price",
}


class CompatibilityCatalogError(RuntimeError):
    """Raised when the compatibility review artifact violates its contract."""


def text(value: Any) -> str:
    """Return stripped text, treating None as blank."""
    return "" if value is None else str(value).strip()


def ascii_value(value: Any) -> Any:
    """Make generated text ASCII-safe without changing typed values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
    }
    for source, replacement in replacements.items():
        value = value.replace(source, replacement)
    return value.encode("ascii", "replace").decode("ascii")


def excel_safe_value(value: Any) -> Any:
    """Normalize timezone-aware values for Excel persistence."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def file_hash(path: Path) -> str:
    """Return a SHA-256 digest without changing the file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require nonempty protected inputs."""
    for path in paths:
        if not path.is_file():
            raise CompatibilityCatalogError(f"Required input is missing: {path}")
        if path.stat().st_size == 0:
            raise CompatibilityCatalogError(f"Required input is empty: {path}")


def read_records(
    path: Path,
    sheet_name: str,
    *,
    header_label: str | None = None,
    keep_vba: bool = False,
) -> list[dict[str, Any]]:
    """Read a worksheet as records while locating an optional header label."""
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=keep_vba,
    )
    try:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header_index = 0
    if header_label is not None:
        header_index = next(
            (
                index
                for index, row in enumerate(rows)
                if header_label in {text(value) for value in row}
            ),
            -1,
        )
        if header_index < 0:
            raise CompatibilityCatalogError(
                f"{path.name}/{sheet_name} lacks {header_label!r}"
            )
    headers = [text(value) for value in rows[header_index]]
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(text(value) for value in row):
            continue
        records.append(
            {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
        )
    return records


def require_headers(
    records: Sequence[dict[str, Any]],
    headers: Iterable[str],
    context: str,
) -> None:
    """Require source fields even when the source has no data rows."""
    if not records:
        raise CompatibilityCatalogError(f"{context} contains no records")
    missing = set(headers) - set(records[0])
    if missing:
        raise CompatibilityCatalogError(
            f"{context} lacks required fields: {sorted(missing)}"
        )


def read_existing_compatibility_ids() -> tuple[set[str], list[str], int]:
    """Inspect the canonical namespace without modifying the workbook."""
    records = read_records(
        CANONICAL_PATH,
        CANONICAL_SHEET,
        header_label="Compatibility ID",
        keep_vba=True,
    )
    populated = [
        text(record.get("Compatibility ID"))
        for record in records
        if text(record.get("Compatibility ID"))
    ]
    valid = [value for value in populated if COMPATIBILITY_ID_PATTERN.fullmatch(value)]
    duplicates = sorted(
        value for value, count in Counter(valid).items() if count > 1
    )
    if duplicates:
        raise CompatibilityCatalogError(
            f"Duplicate canonical Compatibility IDs: {duplicates}"
        )
    malformed = sorted(set(populated) - set(valid))
    highest = max((int(value[3:]) for value in valid), default=0)
    return set(valid), malformed, highest


def normalized(value: Any) -> str:
    """Return conservative lowercase alphanumeric text."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).casefold()))


def model_signature(value: Any) -> tuple[str, ...]:
    """Return model tokens after removing only known repair-context noise."""
    return tuple(
        token
        for token in re.findall(r"[a-z0-9]+", text(value).casefold())
        if token not in MODEL_NOISE and token != "for"
    )


def manufacturer_agrees(
    device: dict[str, Any],
    target: dict[str, Any],
) -> bool:
    """Require an explicit shared manufacturer ID or nonplaceholder name."""
    device_id = normalized(device.get("Manufacturer ID"))
    target_id = normalized(target.get("Manufacturer ID"))
    if device_id and target_id:
        return device_id == target_id
    device_names = {
        normalized(device.get("Manufacturer Name")),
        normalized(device.get("Product Line")),
    } - PLACEHOLDER_MANUFACTURERS
    target_name = normalized(target.get("Manufacturer Name"))
    return bool(target_name and target_name not in PLACEHOLDER_MANUFACTURERS) and (
        target_name in device_names
    )


def relationship_key(record: dict[str, Any]) -> tuple[str, ...]:
    """Return the governed relationship uniqueness key."""
    target_id = text(record["Service ID"]) or text(record["Part ID"])
    endpoint = text(record["Device ID"]) or text(record["Device Family Code"])
    return (
        text(record["Relationship Type"]),
        endpoint,
        target_id,
        text(record["Device Variant"]),
    )


def unresolved_row(
    target_kind: str,
    target: dict[str, Any],
    *,
    reason: str,
    missing: str,
    device: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Create one unresolved review row without inventing a result."""
    device = device or {}
    is_service = target_kind == "Service"
    return {
        "Candidate Type": f"Device to {target_kind}",
        "Device ID": text(device.get("Device ID")),
        "Device Family Code": text(target.get("Device Family Code")),
        "Device Name": text(device.get("Device Name")),
        "Service ID": text(target.get("Service ID")) if is_service else "",
        "Service Name": text(target.get("Service Name")) if is_service else "",
        "Part ID": text(target.get("Part ID")) if not is_service else "",
        "Part Name": text(target.get("Part Name")) if not is_service else "",
        "Missing Evidence": missing,
        "Ambiguity Reason": reason,
        "Required Action": f"Research and review {target_kind.lower()} applicability",
        "Review Priority": "High" if "ambiguous" in reason.casefold() else "Normal",
        "Review Status": "Pending Evidence Review",
        "Reviewer Notes": "",
    }


def base_relationship(
    target_kind: str,
    target: dict[str, Any],
    *,
    relationship_type: str,
    level: str,
    evidence_type: str,
    evidence_detail: str,
    confidence: str,
    device: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build a proposed relationship row from preserved source evidence."""
    device = device or {}
    is_service = target_kind == "Service"
    source_name = SERVICES_PATH.name if is_service else PARTS_PATH.name
    manufacturer_id = text(device.get("Manufacturer ID")) or text(
        target.get("Manufacturer ID")
    )
    manufacturer_name = text(device.get("Manufacturer Name")) or text(
        target.get("Manufacturer Name")
    )
    return {
        "Compatibility ID": "",
        "Relationship Type": relationship_type,
        "Device ID": text(device.get("Device ID")),
        "Device Family Code": text(target.get("Device Family Code"))
        or text(device.get("Device Family Code")),
        "Device Variant": text(device.get("Variant"))
        if level == "Variant Level"
        else "",
        "Service ID": text(target.get("Service ID")) if is_service else "",
        "Part ID": text(target.get("Part ID")) if not is_service else "",
        "Manufacturer ID": manufacturer_id,
        "Manufacturer Name": manufacturer_name,
        "Device Name": text(device.get("Device Name")),
        "Service Name": text(target.get("Service Name")) if is_service else "",
        "Part Name": text(target.get("Part Name")) if not is_service else "",
        "Compatibility Level": level,
        "Compatibility Status": "Proposed",
        "Evidence Type": evidence_type,
        "Evidence Source": f"{source_name}/{target.get('_sheet', '')}",
        "Evidence Detail": evidence_detail,
        "Confidence": confidence,
        "Requires Manual Review": "Yes",
        "Active": "No",
        "Effective Date": "",
        "Expiration Date": "",
        "Review Status": "Pending Review",
        "Reviewer": "",
        "Reviewer Notes": "",
        "Source Record Number": target.get("Source Record Number"),
        "Source Workbook": target.get("Source Workbook") or source_name,
        "Source Worksheet": target.get("Source Worksheet")
        or target.get("_sheet", ""),
        "Import Batch ID": IMPORT_BATCH_ID,
        "Created At": target.get("Created At"),
        "Updated At": target.get("Updated At"),
    }


def target_candidates(
    target_kind: str,
    target: dict[str, Any],
    devices: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Select the strongest conservative relationship tier for one target."""
    target_id_field = f"{target_kind} ID"
    if not text(target.get(target_id_field)):
        return [], [
            unresolved_row(
                target_kind,
                target,
                reason=f"Missing {target_id_field}",
                missing=target_id_field,
            )
        ]
    explicit_device_id = text(target.get("Device ID"))
    if explicit_device_id:
        matches = [
            device
            for device in devices
            if text(device.get("Device ID")) == explicit_device_id
        ]
        if len(matches) != 1:
            return [], [
                unresolved_row(
                    target_kind,
                    target,
                    reason="Explicit Device ID is missing or ambiguous",
                    missing="Resolvable Device ID",
                )
            ]
        device = matches[0]
        level = "Variant Level" if text(device.get("Variant")) else "Model Level"
        relationship_type = (
            f"Device Variant to {target_kind}"
            if level == "Variant Level"
            else f"Device to {target_kind}"
        )
        return [
            base_relationship(
                target_kind,
                target,
                relationship_type=relationship_type,
                level=level,
                evidence_type="Explicit Source Match",
                evidence_detail="Source contains an exact Master Device ID.",
                confidence="High",
                device=device,
            )
        ], []

    model_value = (
        target.get("Device Model")
        or target.get("Device Name")
        or target.get("Model Number")
    )
    signature = model_signature(model_value)
    exact_matches = [
        device
        for device in devices
        if signature
        and signature
        in {
            model_signature(device.get("Device Name")),
            model_signature(device.get("Device Display Name")),
            model_signature(device.get("Model Number")),
        }
        and text(target.get("Device Family Code"))
        == text(device.get("Device Family Code"))
        and manufacturer_agrees(device, target)
    ]
    if len(exact_matches) == 1:
        device = exact_matches[0]
        level = "Variant Level" if text(device.get("Variant")) else "Model Level"
        relationship_type = (
            f"Device Variant to {target_kind}"
            if level == "Variant Level"
            else f"Device to {target_kind}"
        )
        return [
            base_relationship(
                target_kind,
                target,
                relationship_type=relationship_type,
                level=level,
                evidence_type="Exact Manufacturer and Model",
                evidence_detail=(
                    "Exact normalized model tokens, manufacturer, and family agree."
                ),
                confidence="High",
                device=device,
            )
        ], []
    if len(exact_matches) > 1:
        return [], [
            unresolved_row(
                target_kind,
                target,
                reason="Exact evidence is ambiguous across multiple devices",
                missing="Unique exact Device match",
            )
        ]

    family = text(target.get("Device Family Code"))
    if family:
        return [
            base_relationship(
                target_kind,
                target,
                relationship_type=f"Device Family to {target_kind}",
                level="Family Level",
                evidence_type="Family-Level Evidence",
                evidence_detail=(
                    "Source explicitly designates the device family; model evidence "
                    "is absent or insufficient."
                ),
                confidence="Medium",
            )
        ], []
    return [], [
        unresolved_row(
            target_kind,
            target,
            reason="No reliable device or family evidence",
            missing="Device ID, exact model evidence, or explicit family",
        )
    ]


def build_population(
    devices: Sequence[dict[str, Any]],
    services: Sequence[dict[str, Any]],
    parts: Sequence[dict[str, Any]],
    excluded_rows: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Build deterministic proposed relationships and unresolved candidates."""
    relationships: list[dict[str, Any]] = []
    unresolved: list[dict[str, Any]] = []
    for target_kind, targets in (("Service", services), ("Part", parts)):
        for target in targets:
            status = text(target.get("Status"))
            source_number = text(target.get("Source Record Number"))
            if status in {"Rejected", "Archived"} or source_number in excluded_rows:
                unresolved.append(
                    unresolved_row(
                        target_kind,
                        target,
                        reason="Source record is rejected, archived, or excluded",
                        missing="Eligible source record",
                    )
                )
                continue
            proposed, pending = target_candidates(target_kind, target, devices)
            relationships.extend(proposed)
            unresolved.extend(pending)
    relationships.sort(
        key=lambda row: (
            text(row["Relationship Type"]),
            text(row["Device ID"]),
            text(row["Service ID"]) or text(row["Part ID"]),
            text(row["Source Record Number"]).zfill(12),
            text(row["Device Family Code"]),
        )
    )
    keys = [relationship_key(row) for row in relationships]
    duplicate_keys = sorted(key for key, count in Counter(keys).items() if count > 1)
    if duplicate_keys:
        raise CompatibilityCatalogError(
            f"Duplicate relationship candidates: {duplicate_keys}"
        )
    unresolved.sort(
        key=lambda row: (
            text(row["Candidate Type"]),
            text(row["Service ID"]) or text(row["Part ID"]),
            text(row["Device ID"]),
        )
    )
    return relationships, unresolved


def allocate_ids(
    records: Sequence[dict[str, Any]],
    existing: set[str],
    highest: int,
) -> None:
    """Assign a continuous canonical-safe sequence in existing record order."""
    for offset, record in enumerate(records, start=1):
        value = f"CMP{highest + offset:06d}"
        if value in existing:
            raise CompatibilityCatalogError(f"Generated ID overlaps canonical: {value}")
        record["Compatibility ID"] = value


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write one formatted table, including a blank placeholder data row."""
    worksheet.append(list(headers))
    materialized = list(rows) or [{header: "" for header in headers}]
    for record in materialized:
        worksheet.append(
            [
                excel_safe_value(ascii_value(record.get(header, "")))
                for header in headers
            ]
        )
    end_column = get_column_letter(len(headers))
    table = Table(
        displayName=table_name,
        ref=f"A1:{end_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = table.ref
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
    for index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 38)
        worksheet.column_dimensions[get_column_letter(index)].width = width
        if header in {"Effective Date", "Expiration Date"}:
            for cell in worksheet.iter_cols(
                min_col=index,
                max_col=index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                cell[0].number_format = "yyyy-mm-dd"
        if header in {"Created At", "Updated At", "Revision Date"}:
            for cell in worksheet.iter_cols(
                min_col=index,
                max_col=index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                cell[0].number_format = "yyyy-mm-dd hh:mm:ss"


def control_rows(
    headers: Sequence[str],
    values: Sequence[Sequence[str]],
) -> list[dict[str, str]]:
    """Align controlled lists into one lookup table."""
    return [
        {
            header: group[index] if index < len(group) else ""
            for header, group in zip(headers, values, strict=True)
        }
        for index in range(max(map(len, values)))
    ]


def lookup_rows(
    source: Sequence[dict[str, Any]],
    headers: Sequence[str],
) -> list[dict[str, Any]]:
    """Create a deterministic source snapshot with selected fields."""
    return [
        {header: row.get(header, "") for header in headers}
        for row in sorted(
            source,
            key=lambda row: tuple(text(row.get(header)) for header in headers),
        )
    ]


def add_defined_name(
    workbook: Workbook,
    name: str,
    sheet: str,
    column: int,
) -> None:
    """Add one workbook-scoped name over nonblank lookup values."""
    worksheet = workbook[sheet]
    last_row = max(
        (
            row
            for row in range(2, worksheet.max_row + 1)
            if text(worksheet.cell(row, column).value)
        ),
        default=2,
    )
    letter = get_column_letter(column)
    workbook.defined_names.add(
        DefinedName(name, attr_text=f"'{sheet}'!${letter}$2:${letter}${last_row}")
    )


def add_defined_names(workbook: Workbook) -> None:
    """Create every approved cross-sheet validation name."""
    for name, (sheet, column) in DEFINED_NAME_SPECS.items():
        add_defined_name(workbook, name, sheet, column)


def add_data_validations(workbook: Workbook) -> None:
    """Apply named-list validation to governed primary fields."""
    worksheet = workbook["01 - Compatibility Records"]
    for header, name in DEFINED_NAME_BY_HEADER.items():
        column = COMPATIBILITY_HEADERS.index(header) + 1
        validation = DataValidation(
            type="list",
            formula1=f"={name}",
            allow_blank=True,
        )
        validation.error = f"Select a value from {name}."
        validation.errorTitle = "Invalid controlled value"
        validation.showErrorMessage = True
        worksheet.add_data_validation(validation)
        letter = get_column_letter(column)
        validation.add(f"{letter}2:{letter}{worksheet.max_row}")


def metadata_rows(
    records: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
    hashes: dict[Path, str],
    existing: set[str],
    malformed: Sequence[str],
    highest: int,
) -> list[dict[str, str]]:
    """Create review lineage and protected-source metadata."""
    first_id = text(records[0]["Compatibility ID"]) if records else ""
    final_id = text(records[-1]["Compatibility ID"]) if records else ""
    rows = [
        {"Metadata Field": "Import Batch ID", "Value": IMPORT_BATCH_ID},
        {"Metadata Field": "Namespace Authority", "Value": NAMESPACE_AUTHORITY},
        {"Metadata Field": "Schema Columns", "Value": str(len(COMPATIBILITY_HEADERS))},
        {"Metadata Field": "Canonical Valid ID Count", "Value": str(len(existing))},
        {
            "Metadata Field": "Highest Canonical Compatibility ID",
            "Value": f"CMP{highest:06d}" if highest else "",
        },
        {"Metadata Field": "Malformed Canonical IDs", "Value": "; ".join(malformed)},
        {"Metadata Field": "First Generated ID", "Value": first_id},
        {"Metadata Field": "Final Generated ID", "Value": final_id},
        {"Metadata Field": "Generated Relationship Count", "Value": str(len(records))},
        {"Metadata Field": "Unresolved Candidate Count", "Value": str(len(unresolved))},
        {"Metadata Field": "Canonical Import Authorized", "Value": "No"},
    ]
    for path, digest in hashes.items():
        rows.extend(
            [
                {
                    "Metadata Field": f"Protected Input Path: {path.name}",
                    "Value": str(path),
                },
                {
                    "Metadata Field": f"SHA-256: {path.name}",
                    "Value": digest,
                },
            ]
        )
    return rows


def validation_rows(
    records: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
    highest: int,
    malformed: Sequence[str],
) -> list[dict[str, str]]:
    """Create a visible runtime validation summary."""
    counts = Counter(text(row["Relationship Type"]) for row in records)
    levels = Counter(text(row["Compatibility Level"]) for row in records)
    rows = [
        {
            "Validation": "Primary schema",
            "Result": "PASS",
            "Details": f"{len(COMPATIBILITY_HEADERS)} columns",
        },
        {
            "Validation": "Worksheet contract",
            "Result": "PASS",
            "Details": f"{len(SHEET_NAMES)} worksheets",
        },
        {
            "Validation": "Canonical namespace",
            "Result": "PASS",
            "Details": (
                f"Highest valid CMP{highest:06d}; "
                f"malformed values {len(malformed)}"
            ),
        },
        {
            "Validation": "Relationship population",
            "Result": "PASS",
            "Details": str(len(records)),
        },
        {
            "Validation": "Unresolved candidates",
            "Result": "PASS",
            "Details": str(len(unresolved)),
        },
        {
            "Validation": "Generated approvals",
            "Result": "PASS",
            "Details": "None",
        },
    ]
    rows.extend(
        {
            "Validation": f"Relationship type: {name}",
            "Result": "PASS",
            "Details": str(counts.get(name, 0)),
        }
        for name in RELATIONSHIP_TYPES
    )
    rows.extend(
        {
            "Validation": f"Compatibility level: {name}",
            "Result": "PASS",
            "Details": str(levels.get(name, 0)),
        }
        for name in COMPATIBILITY_LEVELS
    )
    return rows


def build_workbook(
    records: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
    devices: Sequence[dict[str, Any]],
    services: Sequence[dict[str, Any]],
    parts: Sequence[dict[str, Any]],
    families: Sequence[dict[str, Any]],
    manufacturers: Sequence[dict[str, Any]],
    hashes: dict[Path, str],
    existing: set[str],
    malformed: Sequence[str],
    highest: int,
) -> Workbook:
    """Construct the complete review workbook in memory."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in SHEET_NAMES:
        workbook.create_sheet(sheet_name)
    content: dict[str, tuple[Sequence[str], Sequence[dict[str, Any]]]] = {
        "00 - Instructions": (
            ["Topic", "Instruction"],
            [
                {
                    "Topic": "Purpose",
                    "Instruction": "Review proposed compatibility relationships.",
                },
                {
                    "Topic": "Boundary",
                    "Instruction": (
                        "No canonical, pricing, inventory, or catalog writes."
                    ),
                },
                {
                    "Topic": "Approval",
                    "Instruction": "Generated records are inactive proposals.",
                },
            ],
        ),
        "01 - Compatibility Records": (COMPATIBILITY_HEADERS, records),
        "02 - Relationship Types": (
            ["Relationship Type", "Target Type"],
            [
                {
                    "Relationship Type": value,
                    "Target Type": "Service" if value.endswith("Service") else "Part",
                }
                for value in RELATIONSHIP_TYPES
            ],
        ),
        "03 - Compatibility Levels": (
            [
                "Compatibility Level",
                "Compatibility Status",
                "Review Status",
                "Confidence",
                "Yes/No",
                "Device Family Code",
                "Manufacturer ID",
            ],
            control_rows(
                [
                    "Compatibility Level",
                    "Compatibility Status",
                    "Review Status",
                    "Confidence",
                    "Yes/No",
                    "Device Family Code",
                    "Manufacturer ID",
                ],
                [
                    COMPATIBILITY_LEVELS,
                    COMPATIBILITY_STATUSES,
                    REVIEW_STATUSES,
                    CONFIDENCE_VALUES,
                    YES_NO_VALUES,
                    [
                        text(row.get("Device Family Code"))
                        for row in families
                        if text(row.get("Device Family Code"))
                    ],
                    [
                        text(row.get("Manufacturer ID"))
                        for row in manufacturers
                        if text(row.get("Manufacturer ID"))
                    ],
                ],
            ),
        ),
        "04 - Evidence Types": (
            ["Evidence Type", "Evidence Rank"],
            [
                {"Evidence Type": value, "Evidence Rank": index}
                for index, value in enumerate(EVIDENCE_TYPES, start=1)
            ],
        ),
        "05 - Devices": (
            [
                "Device ID",
                "Device Family Code",
                "Device Name",
                "Variant",
                "Manufacturer ID",
                "Manufacturer Name",
                "Product Line",
            ],
            lookup_rows(
                devices,
                [
                    "Device ID",
                    "Device Family Code",
                    "Device Name",
                    "Variant",
                    "Manufacturer ID",
                    "Manufacturer Name",
                    "Product Line",
                ],
            ),
        ),
        "06 - Services": (
            [
                "Service ID",
                "Service Name",
                "Device Family Code",
                "Device Model",
                "Manufacturer ID",
                "Manufacturer Name",
                "Repair Type",
            ],
            lookup_rows(
                services,
                [
                    "Service ID",
                    "Service Name",
                    "Device Family Code",
                    "Device Model",
                    "Manufacturer ID",
                    "Manufacturer Name",
                    "Repair Type",
                ],
            ),
        ),
        "07 - Parts": (
            [
                "Part ID",
                "Part Name",
                "Device Family Code",
                "Device ID",
                "Device Name",
                "Manufacturer ID",
                "Manufacturer Name",
                "Compatibility Scope",
                "Part Type",
            ],
            lookup_rows(
                parts,
                [
                    "Part ID",
                    "Part Name",
                    "Device Family Code",
                    "Device ID",
                    "Device Name",
                    "Manufacturer ID",
                    "Manufacturer Name",
                    "Compatibility Scope",
                    "Part Type",
                ],
            ),
        ),
        "08 - Family Relationships": (
            COMPATIBILITY_HEADERS,
            [
                row
                for row in records
                if text(row["Compatibility Level"]) == "Family Level"
            ],
        ),
        "09 - Model Relationships": (
            COMPATIBILITY_HEADERS,
            [
                row
                for row in records
                if text(row["Compatibility Level"]) == "Model Level"
            ],
        ),
        "10 - Variant Relationships": (
            COMPATIBILITY_HEADERS,
            [
                row
                for row in records
                if text(row["Compatibility Level"]) == "Variant Level"
            ],
        ),
        "11 - Unresolved Review": (UNRESOLVED_HEADERS, unresolved),
        "12 - Validation Summary": (
            ["Validation", "Result", "Details"],
            validation_rows(records, unresolved, highest, malformed),
        ),
        "13 - Revision History": (
            ["Version", "Revision Date", "Authority", "Change"],
            [
                {
                    "Version": "1.0",
                    "Revision Date": date(2026, 7, 23),
                    "Authority": NAMESPACE_AUTHORITY,
                    "Change": "Initial compatibility review framework.",
                }
            ],
        ),
        "14 - Import Metadata": (
            ["Metadata Field", "Value"],
            metadata_rows(
                records,
                unresolved,
                hashes,
                existing,
                malformed,
                highest,
            ),
        ),
    }
    for sheet_name in SHEET_NAMES:
        headers, rows = content[sheet_name]
        append_table(
            workbook[sheet_name],
            headers,
            rows,
            TABLE_NAMES[sheet_name],
        )
    add_defined_names(workbook)
    add_data_validations(workbook)
    pending_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet_name in (
        "08 - Family Relationships",
        "09 - Model Relationships",
        "10 - Variant Relationships",
        "11 - Unresolved Review",
    ):
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = pending_fill
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read populated table rows from an open workbook."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    records = []
    for row in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def require_excel_archive(path: Path) -> None:
    """Require the minimum OOXML workbook members."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        missing = required - set(archive.namelist())
    if missing:
        raise CompatibilityCatalogError(
            f"Temporary workbook lacks OOXML members: {sorted(missing)}"
        )


def validate_reopened(
    path: Path,
    expected: Sequence[dict[str, Any]],
    existing: set[str],
    highest: int,
) -> None:
    """Validate the temporary artifact before atomic publication."""
    require_excel_archive(path)
    workbook = load_workbook(path, data_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise CompatibilityCatalogError("Reopened worksheet order differs")
        if len(set(workbook.sheetnames)) != len(SHEET_NAMES):
            raise CompatibilityCatalogError("Reopened worksheet names are duplicated")
        all_tables: list[str] = []
        for sheet_name in SHEET_NAMES:
            worksheet = workbook[sheet_name]
            expected_table = TABLE_NAMES[sheet_name]
            if expected_table not in worksheet.tables:
                raise CompatibilityCatalogError(
                    f"{sheet_name} lacks table {expected_table}"
                )
            if worksheet.freeze_panes != "A2":
                raise CompatibilityCatalogError(f"{sheet_name} header is not frozen")
            all_tables.extend(worksheet.tables)
        if len(all_tables) != len(set(all_tables)):
            raise CompatibilityCatalogError("Table names are not globally unique")
        required_names = set(DEFINED_NAME_BY_HEADER.values())
        if not required_names <= set(workbook.defined_names):
            missing = required_names - set(workbook.defined_names)
            raise CompatibilityCatalogError(f"Defined names missing: {sorted(missing)}")
        for name, (sheet_name, column) in DEFINED_NAME_SPECS.items():
            destinations = list(workbook.defined_names[name].destinations)
            if len(destinations) != 1:
                raise CompatibilityCatalogError(
                    f"Defined name {name} has multiple destinations"
                )
            destination_sheet, coordinates = destinations[0]
            if destination_sheet != sheet_name:
                raise CompatibilityCatalogError(
                    f"Defined name {name} targets the wrong worksheet"
                )
            min_column, min_row, max_column, _ = range_boundaries(coordinates)
            if min_column != column or max_column != column or min_row != 2:
                raise CompatibilityCatalogError(
                    f"Defined name {name} targets the wrong column or start row"
                )
        primary = workbook["01 - Compatibility Records"]
        actual = table_records(primary, TABLE_NAMES[primary.title])
        if len(actual) != len(expected):
            raise CompatibilityCatalogError("Reopened relationship count differs")
        headers = [
            text(cell.value)
            for cell in primary[1][: len(COMPATIBILITY_HEADERS)]
        ]
        if headers != COMPATIBILITY_HEADERS:
            raise CompatibilityCatalogError("Reopened primary schema differs")
        ids = [text(row["Compatibility ID"]) for row in actual]
        if any(not COMPATIBILITY_ID_PATTERN.fullmatch(value) for value in ids):
            raise CompatibilityCatalogError("Reopened Compatibility ID is malformed")
        if len(ids) != len(set(ids)) or set(ids) & existing:
            raise CompatibilityCatalogError(
                "Reopened Compatibility IDs overlap or duplicate"
            )
        expected_ids = [
            f"CMP{highest + offset:06d}"
            for offset in range(1, len(actual) + 1)
        ]
        if ids != expected_ids:
            raise CompatibilityCatalogError(
                "Reopened Compatibility IDs are not continuous"
            )
        if [relationship_key(row) for row in actual] != [
            relationship_key(row) for row in expected
        ]:
            raise CompatibilityCatalogError(
                "Reopened deterministic relationship order differs"
            )
        for row in actual:
            if text(row["Compatibility Status"]) != "Proposed":
                raise CompatibilityCatalogError(
                    "Generated relationship is not Proposed"
                )
            if text(row["Review Status"]) in {"Approved", "Ready for Approval"}:
                raise CompatibilityCatalogError("Generated relationship is approved")
            if text(row["Active"]) != "No":
                raise CompatibilityCatalogError("Generated relationship is active")
            if not text(row["Evidence Type"]) or not text(row["Evidence Detail"]):
                raise CompatibilityCatalogError("Generated evidence is incomplete")
        queue_ids: list[str] = []
        for sheet_name in (
            "08 - Family Relationships",
            "09 - Model Relationships",
            "10 - Variant Relationships",
        ):
            queue_ids.extend(
                text(row["Compatibility ID"])
                for row in table_records(
                    workbook[sheet_name],
                    TABLE_NAMES[sheet_name],
                )
            )
        if sorted(queue_ids) != sorted(ids):
            raise CompatibilityCatalogError(
                "Relationship review queues do not reconcile"
            )
    finally:
        workbook.close()


def load_sources() -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    set[str],
]:
    """Load and validate the three protected master catalogs."""
    devices = read_records(DEVICES_PATH, "01 - Master Devices")
    services = read_records(SERVICES_PATH, "01 - Master Services")
    parts = read_records(PARTS_PATH, "01 - Master Parts")
    families = read_records(DEVICES_PATH, "03 - Device Families")
    manufacturers = read_records(DEVICES_PATH, "02 - Manufacturers")
    exclusions = read_records(PROPOSAL_PATH, "02 - Duplicate Exclusions")
    require_headers(
        devices,
        {
            "Device ID",
            "Device Family Code",
            "Device Name",
            "Variant",
            "Manufacturer ID",
            "Manufacturer Name",
        },
        "Master Devices",
    )
    require_headers(
        services,
        {
            "Service ID",
            "Service Name",
            "Device Family Code",
            "Device Model",
            "Manufacturer ID",
            "Manufacturer Name",
            "Source Record Number",
        },
        "Master Services",
    )
    require_headers(
        parts,
        {
            "Part ID",
            "Part Name",
            "Device Family Code",
            "Device ID",
            "Device Name",
            "Manufacturer ID",
            "Manufacturer Name",
            "Source Record Number",
        },
        "Master Parts",
    )
    for record in services:
        record["_sheet"] = "01 - Master Services"
    for record in parts:
        record["_sheet"] = "01 - Master Parts"
    excluded_rows = {
        text(record.get("Source Row Number"))
        for record in exclusions
        if text(record.get("Source Row Number"))
    }
    return devices, services, parts, families, manufacturers, excluded_rows


def main() -> int:
    """Generate, validate, and atomically publish the review workbook."""
    workbook: Workbook | None = None
    try:
        require_files(PROTECTED_PATHS)
        before_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        (
            devices,
            services,
            parts,
            families,
            manufacturers,
            excluded_rows,
        ) = load_sources()
        existing, malformed, highest = read_existing_compatibility_ids()
        records, unresolved = build_population(
            devices,
            services,
            parts,
            excluded_rows,
        )
        allocate_ids(records, existing, highest)
        workbook = build_workbook(
            records,
            unresolved,
            devices,
            services,
            parts,
            families,
            manufacturers,
            before_hashes,
            existing,
            malformed,
            highest,
        )
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        if TEMP_OUTPUT_PATH.exists():
            TEMP_OUTPUT_PATH.unlink()
        workbook.save(TEMP_OUTPUT_PATH)
        workbook.close()
        workbook = None
        validate_reopened(TEMP_OUTPUT_PATH, records, existing, highest)
        after_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        if after_hashes != before_hashes:
            raise CompatibilityCatalogError(
                "A protected source workbook changed during generation"
            )
        os.replace(TEMP_OUTPUT_PATH, OUTPUT_PATH)
        type_counts = Counter(text(row["Relationship Type"]) for row in records)
        level_counts = Counter(text(row["Compatibility Level"]) for row in records)
        print("Master Compatibility Catalog V1 validation: PASS")
        print(f"Generated workbook: {OUTPUT_PATH}")
        print(f"Schema columns: {len(COMPATIBILITY_HEADERS)}")
        print(f"Generated relationships: {len(records)}")
        print(f"Unresolved candidates: {len(unresolved)}")
        print(f"Relationship counts: {dict(sorted(type_counts.items()))}")
        print(f"Level counts: {dict(sorted(level_counts.items()))}")
        if records:
            print(
                "Compatibility ID range: "
                f"{records[0]['Compatibility ID']} through "
                f"{records[-1]['Compatibility ID']}"
            )
        print("Protected input hashes: PASS")
        print("Canonical, pricing, inventory, and catalog writes: NONE")
        return 0
    except (
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        CompatibilityCatalogError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError as close_error:
                print(f"ERROR: Failed to close workbook: {close_error}", file=sys.stderr)
        try:
            if TEMP_OUTPUT_PATH.exists():
                TEMP_OUTPUT_PATH.unlink()
        except OSError as cleanup_error:
            print(
                f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}",
                file=sys.stderr,
            )


if __name__ == "__main__":
    raise SystemExit(main())
