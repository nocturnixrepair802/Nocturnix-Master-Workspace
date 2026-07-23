"""Generate the proposed Nocturnix Master Services Catalog v1 workbook.

The output is a standalone review artifact. This script reads protected inputs,
creates a new workbook, and never imports records into the canonical database.
"""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_PATH = Path(r"D:\Business Portal\300_Pricing\Legacy\Raw Import Data.xlsx")
STAGING_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Legacy_Catalog_Staging_Preview_v1.xlsx"
)
PROPOSAL_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
LABOR_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
SERVICE_ID_SOURCE_PATH = CANONICAL_PATH
SERVICE_ID_SOURCE_SHEET = "34 Master Services"
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Services_Catalog_v1.xlsx"
)

EXPECTED_SERVICE_ROWS = 314
SHEET_NAMES = [
    "00 - Instructions",
    "01 - Master Services",
    "02 - Service Categories",
    "03 - Repair Types",
    "04 - Device Families",
    "05 - Manufacturers",
    "06 - Labor Standards",
    "07 - Labor Tiers",
    "08 - Difficulty Levels",
    "09 - Skill Levels",
    "10 - Turnaround Times",
    "11 - Warranty Options",
    "12 - Status Values",
    "13 - Validation Summary",
    "14 - Revision History",
    "15 - Import Metadata",
]
TABLE_NAMES_BY_SHEET = {
    "00 - Instructions": {"tblMasterServicesInstructions"},
    "01 - Master Services": {"tblMasterServicesCatalog"},
    "02 - Service Categories": {"tblServiceCategories"},
    "03 - Repair Types": {"tblRepairTypes"},
    "04 - Device Families": {"tblDeviceFamilies"},
    "05 - Manufacturers": {"tblManufacturers"},
    "06 - Labor Standards": {"tblLaborStandardsLookup"},
    "07 - Labor Tiers": {"tblLaborTiersLookup"},
    "08 - Difficulty Levels": {"tblDifficultyLevels"},
    "09 - Skill Levels": {"tblSkillLevels"},
    "10 - Turnaround Times": {"tblTurnaroundTimes"},
    "11 - Warranty Options": {"tblWarrantyOptions"},
    "12 - Status Values": {"tblServiceStatusValues"},
    "13 - Validation Summary": {
        "tblMasterServicesValidationSummary",
        "tblLaborMatchAudit",
    },
    "14 - Revision History": {"tblMasterServicesRevisionHistory"},
    "15 - Import Metadata": {"tblMasterServicesImportMetadata"},
}
MASTER_HEADERS = [
    "Service ID",
    "Legacy Service SKU",
    "Active",
    "Status",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Device Series",
    "Device Model",
    "Service Category ID",
    "Service Category",
    "Repair Type ID",
    "Repair Type",
    "Service Name",
    "Service Display Name",
    "Service Description",
    "Labor Standard ID",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Labor Tier",
    "Repair Difficulty",
    "Skill Level",
    "Turnaround Time",
    "Requires Parts",
    "Requires Labor",
    "Diagnostic Required",
    "Warranty Eligible",
    "Default Warranty",
    "Mobile Service Eligible",
    "Mail-In Eligible",
    "Pricing Status",
    "Legacy Retail Price",
    "Legacy Cost",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Review Status",
    "Reviewer Notes",
    "Effective Date",
    "Last Reviewed",
    "Created At",
    "Updated At",
]
LABOR_AUDIT_HEADERS = [
    "Source Record Number",
    "Service ID",
    "Legacy Service Name",
    "Labor Standard ID",
    "Match Score",
    "Second Best Score",
    "Score Margin",
    "Match Evidence",
    "Mapping Result",
    "Mapped Minutes",
    "Mapped Labor Tier",
    "Mapped Difficulty",
    "Mapped Skill Level",
]
PRICING_STATUSES = [
    "Pending Pricing Review",
    "Legacy Price Review",
    "No Pricing Exceptions",
    "Archive Candidate",
]
SERVICE_STATUSES = ["Active", "Planned", "Future", "Draft", "Archived"]
REVIEW_STATUSES = [
    "Pending Review",
    "Pending Labor Mapping",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
DEFINED_NAME_BY_HEADER = {
    "Active": "DV_YesNo",
    "Status": "DV_ServiceStatuses",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Service Category ID": "DV_ServiceCategoryIDs",
    "Repair Type ID": "DV_RepairTypeIDs",
    "Labor Standard ID": "DV_LaborStandardIDs",
    "Labor Tier": "DV_LaborTiers",
    "Repair Difficulty": "DV_DifficultyLevels",
    "Skill Level": "DV_SkillLevels",
    "Turnaround Time": "DV_TurnaroundTimes",
    "Requires Parts": "DV_YesNo",
    "Requires Labor": "DV_YesNo",
    "Diagnostic Required": "DV_YesNo",
    "Warranty Eligible": "DV_YesNo",
    "Default Warranty": "DV_WarrantyOptions",
    "Mobile Service Eligible": "DV_YesNo",
    "Mail-In Eligible": "DV_YesNo",
    "Pricing Status": "DV_PricingStatuses",
    "Review Status": "DV_ReviewStatuses",
}
REQUIRED_DEFINED_NAMES = set(DEFINED_NAME_BY_HEADER.values())
LABOR_MATCH_THRESHOLD = 0.82
LABOR_TIE_MARGIN = 0.03
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
PERSISTED_SCORE_FIELDS = {"Match Score", "Second Best Score", "Score Margin"}
PERSISTED_MINUTE_FIELDS = {
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Mapped Minutes",
    "Matched Standard Minutes",
    "Matched Minimum Minutes",
    "Matched Maximum Minutes",
}
PERSISTED_DATE_FIELDS = {
    "Effective Date",
    "Last Reviewed",
    "Created At",
    "Updated At",
    "Generated At UTC",
    "Decision Date",
}
PERSISTED_SCORE_TOLERANCE = Decimal("1e-12")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


class CatalogError(RuntimeError):
    """Raised when the proposed catalog cannot be generated safely."""


@dataclass(frozen=True)
class LaborMatch:
    """A reliable labor-standard match or a deliberately blank result."""

    labor_id: str = ""
    standard_minutes: Any = ""
    minimum_minutes: Any = ""
    maximum_minutes: Any = ""
    labor_tier: str = ""
    difficulty: str = ""
    skill_level: str = ""
    match_score: float = 0.0
    second_best_score: float = 0.0
    score_margin: float = 0.0
    evidence: str = ""
    mapping_result: str = "Pending Labor Mapping"


def ascii_value(value: Any) -> Any:
    """Return ASCII-safe strings while preserving typed non-string values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return value.encode("ascii", "replace").decode("ascii")


def text(value: Any) -> str:
    """Return stripped, ASCII-safe text."""
    return "" if value is None else str(ascii_value(value)).strip()


def normalized_text(value: Any) -> str:
    """Normalize persisted text without changing substantive characters."""
    return "" if value is None else str(value).strip()


def normalized_number(value: Any) -> Decimal | None:
    """Normalize a persisted numeric value and reject populated nonnumbers."""
    if value is None or normalized_text(value) == "":
        return None
    if isinstance(value, bool):
        raise ValueError("Boolean is not a persisted numeric value")
    try:
        number = Decimal(normalized_text(value))
    except InvalidOperation as exc:
        raise ValueError(f"Nonnumeric persisted value: {value!r}") from exc
    if not number.is_finite():
        raise ValueError(f"Nonfinite persisted value: {value!r}")
    return number


def normalized_datetime(value: Any) -> str:
    """Normalize persisted dates and equivalent timezone representations."""
    if value is None or normalized_text(value) == "":
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(UTC).replace(tzinfo=None)
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = normalized_text(value)
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        try:
            return date.fromisoformat(raw).isoformat()
        except ValueError:
            return raw
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)
    return parsed.isoformat()


def normalized_persisted_value(field_name: str, value: Any) -> Any:
    """Normalize a reopened workbook value according to its field semantics."""
    if field_name in PERSISTED_SCORE_FIELDS | PERSISTED_MINUTE_FIELDS:
        return normalized_number(value)
    if field_name in PERSISTED_DATE_FIELDS:
        return normalized_datetime(value)
    return normalized_text(value)


def persisted_values_equal(field_name: str, expected: Any, actual: Any) -> bool:
    """Compare values while tolerating only equivalent Excel representations."""
    try:
        normalized_expected = normalized_persisted_value(field_name, expected)
        normalized_actual = normalized_persisted_value(field_name, actual)
    except ValueError:
        return False
    if field_name in PERSISTED_SCORE_FIELDS:
        if normalized_expected is None or normalized_actual is None:
            return normalized_expected is normalized_actual
        return (
            abs(normalized_expected - normalized_actual)
            <= PERSISTED_SCORE_TOLERANCE
        )
    return normalized_expected == normalized_actual


def persisted_value_token(field_name: str, value: Any) -> str:
    """Return a stable digest token for a normalized persisted value."""
    try:
        normalized = normalized_persisted_value(field_name, value)
    except ValueError:
        return f"!invalid!{normalized_text(value)}"
    if normalized is None:
        return ""
    if field_name in PERSISTED_SCORE_FIELDS:
        return f"{normalized:.12f}"
    if field_name in PERSISTED_MINUTE_FIELDS:
        return format(normalized.normalize(), "f")
    return str(normalized)


def assert_persisted_comparison_policy() -> None:
    """Exercise the persisted-value equivalence boundary without workbook I/O."""
    assert persisted_values_equal("Labor Standard ID", "", None)
    assert persisted_values_equal("Mapped Minutes", 45, 45.0)
    assert persisted_values_equal("Match Score", 0.82, Decimal("0.820000000000"))
    assert persisted_values_equal("Match Evidence", " evidence ", "evidence")
    assert not persisted_values_equal("Labor Standard ID", "LAB001", "LAB002")
    assert not persisted_values_equal(
        "Match Score", Decimal("0.82"), Decimal("0.820000000002")
    )


def normalized(value: Any) -> str:
    """Return a comparison key containing lowercase ASCII alphanumerics."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).lower()))


def money(value: Any) -> Decimal | None:
    """Parse a source monetary observation without calculating a replacement."""
    if value is None or text(value) == "" or isinstance(value, bool):
        return None
    try:
        parsed = Decimal(text(value).replace("$", "").replace(",", ""))
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def sha256_file(path: Path) -> str:
    """Calculate a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def labor_audit_digest(records: Sequence[dict[str, Any]]) -> str:
    """Return a deterministic digest covering every labor-audit value."""
    digest = hashlib.sha256()
    for record in records:
        values = [
            persisted_value_token(header, record.get(header))
            for header in LABOR_AUDIT_HEADERS
        ]
        digest.update("\x1f".join(values).encode("ascii"))
        digest.update(b"\n")
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Fail before generation if an input is unavailable."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise CatalogError(f"Required file(s) missing: {', '.join(missing)}")


def validate_sheet_names() -> None:
    """Enforce Excel worksheet-name constraints."""
    if len(set(SHEET_NAMES)) != len(SHEET_NAMES):
        raise CatalogError("Worksheet names must be unique")
    overlong = [name for name in SHEET_NAMES if len(name) > 31]
    if overlong:
        raise CatalogError(f"Worksheet names exceed 31 characters: {overlong}")


def read_sheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read a worksheet as dictionaries without saving its workbook."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise CatalogError(f"Missing worksheet {sheet_name!r} in {path}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        try:
            headers = [text(value) for value in next(rows)]
        except StopIteration as exc:
            raise CatalogError(f"Worksheet {sheet_name!r} is empty") from exc
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    finally:
        workbook.close()


def read_existing_service_ids(path: Path) -> tuple[set[str], list[str]]:
    """Read valid and malformed canonical Service IDs without changing the source."""
    if not path.is_file():
        raise CatalogError(f"Canonical Service ID source does not exist: {path}")
    if path.stat().st_size <= 0:
        raise CatalogError(f"Canonical Service ID source is empty: {path}")
    if not zipfile.is_zipfile(path):
        raise CatalogError(
            f"Canonical Service ID source is not a valid Excel ZIP container: {path}"
        )
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_vba=True,
        )
    except Exception as exc:
        raise CatalogError(
            f"Unable to open canonical Service ID source {path}: {exc}"
        ) from exc
    try:
        if SERVICE_ID_SOURCE_SHEET not in workbook.sheetnames:
            raise CatalogError(
                f"Missing worksheet {SERVICE_ID_SOURCE_SHEET!r} in {path}"
            )
        worksheet = workbook[SERVICE_ID_SOURCE_SHEET]
        header_row = 0
        service_id_column = 0
        for row in worksheet.iter_rows(min_row=1, max_row=min(50, worksheet.max_row)):
            for cell in row:
                if text(cell.value) == "Service ID":
                    header_row = cell.row
                    service_id_column = cell.column
                    break
            if service_id_column:
                break
        if not service_id_column:
            raise CatalogError(
                f"Service ID header not found in {SERVICE_ID_SOURCE_SHEET!r}"
            )
        valid: set[str] = set()
        malformed: list[str] = []
        for (value,) in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            min_col=service_id_column,
            max_col=service_id_column,
            values_only=True,
        ):
            identifier = text(value)
            if not identifier:
                continue
            if SERVICE_ID_PATTERN.fullmatch(identifier):
                valid.add(identifier)
            else:
                malformed.append(identifier)
        if not valid:
            raise CatalogError(
                "No valid canonical Service IDs were found; next ID cannot be guessed"
            )
        return valid, malformed
    finally:
        workbook.close()


def next_service_number(existing_ids: set[str], generated_count: int) -> int:
    """Return the next numeric suffix and reject six-digit sequence overflow."""
    highest = max(int(identifier[3:]) for identifier in existing_ids)
    first = highest + 1
    final = first + generated_count - 1
    if final > 999999:
        raise CatalogError("Generated Service ID sequence exceeds six digits")
    return first


def read_table(path: Path, table_name: str) -> list[dict[str, Any]]:
    """Read a named Excel Table from a protected workbook."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        for worksheet in workbook.worksheets:
            if table_name not in worksheet.tables:
                continue
            table = worksheet.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            rows = list(
                worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )
            headers = [text(value) for value in rows[0]]
            return [dict(zip(headers, row, strict=False)) for row in rows[1:]]
    finally:
        workbook.close()
    return []


def first_value(record: dict[str, Any], names: Sequence[str]) -> Any:
    """Return the first populated value from candidate field names."""
    for name in names:
        value = record.get(name)
        if value is not None and text(value) != "":
            return value
    return ""


def proposed_lookup(
    existing: Sequence[dict[str, Any]],
    legacy_names: Iterable[str],
    id_candidates: Sequence[str],
    name_candidates: Sequence[str],
    proposed_prefix: str,
) -> tuple[list[dict[str, Any]], dict[str, tuple[str, str]]]:
    """Combine canonical and legacy lookup values without changing either source."""
    rows: list[dict[str, Any]] = []
    by_name: dict[str, tuple[str, str]] = {}
    for record in existing:
        name = text(first_value(record, name_candidates))
        identifier = text(first_value(record, id_candidates))
        if not name or not identifier or normalized(name) in by_name:
            continue
        by_name[normalized(name)] = (identifier, name)
        rows.append(
            {"ID": identifier, "Name": name, "Source": "Canonical", "Active": "Yes"}
        )
    missing = sorted(
        {text(name) for name in legacy_names if text(name)} - {row["Name"] for row in rows},
        key=str.casefold,
    )
    for sequence, name in enumerate(missing, start=1):
        key = normalized(name)
        if key in by_name:
            continue
        identifier = f"{proposed_prefix}-{sequence:03d}"
        by_name[key] = (identifier, name)
        rows.append(
            {
                "ID": identifier,
                "Name": name,
                "Source": "Legacy Proposed",
                "Active": "Yes",
            }
        )
    return rows, by_name


def device_category(record: dict[str, Any]) -> str:
    """Map legacy type/name text to the labor workbook device category."""
    source = normalized(
        " ".join(
            text(record.get(name))
            for name in ("Legacy Type", "Legacy Group", "Legacy Name")
        )
    )
    mappings = [
        (("tablet", "ipad"), "Tablets"),
        (("computer", "laptop", "desktop", "macbook", "chromebook"), "Computers"),
        (("console", "xbox", "playstation", "switch", "controller"), "Gaming"),
        (("watch", "wearable"), "Wearables"),
        (("phone", "iphone", "android", "samsung"), "Phones"),
    ]
    for terms, category in mappings:
        if any(term in source for term in terms):
            return category
    return ""


def family_name(category: str) -> str:
    """Return a singular proposed device-family name."""
    return {
        "Phones": "Phone",
        "Tablets": "Tablet",
        "Computers": "Computer",
        "Gaming": "Gaming Device",
        "Wearables": "Wearable",
    }.get(category, "")


def service_similarity(record: dict[str, Any], labor_service: Any) -> float:
    """Score labor service text against legacy group, name, and type fields."""
    target = normalized(labor_service)
    if not target:
        return 0.0
    candidates = [
        normalized(record.get("Legacy Group")),
        normalized(record.get("Legacy Name")),
        normalized(record.get("Legacy Type")),
    ]
    scores = []
    for candidate in candidates:
        if not candidate:
            continue
        if target in candidate or candidate in target:
            scores.append(1.0)
        else:
            scores.append(SequenceMatcher(None, candidate, target).ratio())
    return max(scores, default=0.0)


def labor_match(
    service: dict[str, Any], labor_rows: Sequence[dict[str, Any]]
) -> LaborMatch:
    """Select a reliable labor match using documented categorical evidence."""
    expected_category = device_category(service)
    legacy_manufacturer = normalized(service.get("Legacy Manufacturer"))
    candidates: list[tuple[float, float, float, dict[str, Any]]] = []
    for labor in labor_rows:
        labor_category = text(labor.get("Device Category"))
        if expected_category and labor_category != expected_category:
            continue
        similarity = service_similarity(service, labor.get("Service"))
        labor_manufacturer = normalized(labor.get("Manufacturer"))
        manufacturer_score = 0.0
        if legacy_manufacturer and labor_manufacturer == legacy_manufacturer:
            manufacturer_score = 0.2
        elif legacy_manufacturer and labor_manufacturer and (
            legacy_manufacturer in labor_manufacturer
            or labor_manufacturer in legacy_manufacturer
        ):
            manufacturer_score = 0.1
        category_score = 0.1 if expected_category else 0.0
        score = similarity + manufacturer_score + category_score
        candidates.append((score, similarity, manufacturer_score, labor))
    candidates.sort(key=lambda item: item[0], reverse=True)
    if not candidates:
        return LaborMatch(evidence="No labor candidates matched device category")
    best_score, similarity, manufacturer_score, best = candidates[0]
    second_score = candidates[1][0] if len(candidates) > 1 else 0.0
    margin = best_score - second_score
    best_id = text(first_value(best, ("Labor ID", "Labor Standard ID")))
    second_id = (
        text(first_value(candidates[1][3], ("Labor ID", "Labor Standard ID")))
        if len(candidates) > 1
        else ""
    )
    evidence = (
        f"device_category={expected_category or 'unknown'}; "
        f"labor_service={text(best.get('Service'))}; "
        f"service_similarity={similarity:.3f}; "
        f"manufacturer_support={manufacturer_score:.3f}"
    )
    if not best_id or best_score <= LABOR_MATCH_THRESHOLD:
        return LaborMatch(
            match_score=best_score,
            second_best_score=second_score,
            score_margin=margin,
            evidence=evidence,
            mapping_result="Pending Labor Mapping",
        )
    if second_id and best_id != second_id and margin <= LABOR_TIE_MARGIN:
        return LaborMatch(
            match_score=best_score,
            second_best_score=second_score,
            score_margin=margin,
            evidence=evidence,
            mapping_result="Ambiguous",
        )
    return LaborMatch(
        labor_id=best_id,
        standard_minutes=best.get("Standard Minutes", ""),
        minimum_minutes=best.get("Minimum Minutes", ""),
        maximum_minutes=best.get("Maximum Minutes", ""),
        labor_tier=text(first_value(best, ("Labor Rate Tier", "Labor Tier"))),
        difficulty=text(best.get("Repair Difficulty")),
        skill_level=text(best.get("Skill Level")),
        match_score=best_score,
        second_best_score=second_score,
        score_margin=margin,
        evidence=evidence,
        mapping_result="Mapped",
    )


def pricing_status(price_value: Any, cost_value: Any) -> str:
    """Classify preserved Repair pricing observations without calculating values."""
    price = money(price_value)
    cost = money(cost_value)
    if price is None or cost is None:
        return "Pending Pricing Review"
    if price > 0:
        return "Legacy Price Review"
    return "Pending Pricing Review"


def create_repair_types(
    repairs: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, tuple[str, str]]]:
    """Create deterministic proposed repair-type lookups from legacy classifications."""
    names = sorted(
        {
            text(first_value(row, ("Legacy Group", "Legacy Type", "Legacy Name")))
            for row in repairs
        }
        - {""},
        key=str.casefold,
    )
    lookup: dict[str, tuple[str, str]] = {}
    rows = []
    for sequence, name in enumerate(names, start=1):
        identifier = f"RT-{sequence:04d}"
        lookup[normalized(name)] = (identifier, name)
        rows.append(
            {
                "Repair Type ID": identifier,
                "Repair Type": name,
                "Service Category ID": "SC-REPAIR",
                "Active": "Yes",
            }
        )
    return rows, lookup


def create_device_families(
    canonical_rows: Sequence[dict[str, Any]],
    repairs: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, tuple[str, str]]]:
    """Prefer canonical family codes and add deterministic proposed gaps."""
    rows: list[dict[str, Any]] = []
    lookup: dict[str, tuple[str, str]] = {}
    for record in canonical_rows:
        code = text(first_value(record, ("Device Family Code", "Family Code")))
        name = text(first_value(record, ("Device Family", "Device Family Name")))
        if not code or not name or normalized(name) in lookup:
            continue
        lookup[normalized(name)] = (code, name)
        rows.append(
            {
                "Device Family Code": code,
                "Device Family Name": name,
                "Description": text(record.get("Description")),
                "Active": text(record.get("Active")) or "Yes",
            }
        )
    proposed_codes = {
        "Phone": "DF-PHONE",
        "Tablet": "DF-TABLET",
        "Computer": "DF-COMPUTER",
        "Gaming Device": "DF-GAMING",
        "Wearable": "DF-WEARABLE",
    }
    required_names = sorted(
        {family_name(device_category(record)) for record in repairs} - {""},
        key=str.casefold,
    )
    used_codes = {row["Device Family Code"] for row in rows}
    for name in required_names:
        key = normalized(name)
        if key in lookup:
            continue
        base_code = proposed_codes[name]
        code = base_code
        suffix = 1
        while code in used_codes:
            suffix += 1
            code = f"{base_code}-{suffix}"
        used_codes.add(code)
        lookup[key] = (code, name)
        rows.append(
            {
                "Device Family Code": code,
                "Device Family Name": name,
                "Description": "Proposed legacy-derived device family.",
                "Active": "Yes",
            }
        )
    return rows, lookup


def build_service_rows(
    repairs: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
    manufacturers: dict[str, tuple[str, str]],
    repair_types: dict[str, tuple[str, str]],
    device_families: dict[str, tuple[str, str]],
    conflict_decisions: dict[int, dict[str, Any]],
    first_service_number: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Transform retained Repair records into proposed canonical service rows."""
    ordered = sorted(repairs, key=lambda row: int(row["Source Row Number"]))
    output = []
    labor_audit = []
    for sequence, source in enumerate(ordered, start=1):
        service_id = f"SVC{first_service_number + sequence - 1:06d}"
        source_row = int(source["Source Row Number"])
        labor = labor_match(source, labor_rows)
        category = device_category(source)
        device_family = family_name(category)
        family_code, device_family = device_families.get(
            normalized(device_family), ("", device_family)
        )
        type_source = text(first_value(source, ("Legacy Group", "Legacy Type", "Legacy Name")))
        repair_type_id, repair_type_name = repair_types.get(
            normalized(type_source), ("", type_source)
        )
        legacy_manufacturer_name = text(source.get("Legacy Manufacturer"))
        manufacturer_id, manufacturer_name = manufacturers.get(
            normalized(legacy_manufacturer_name), ("", legacy_manufacturer_name)
        )
        conflict = conflict_decisions.get(source_row, {})
        conflict_status = text(conflict.get("Approval Status"))
        unresolved_conflict = bool(conflict) and conflict_status != "Approved"
        if unresolved_conflict or not manufacturer_id or not family_code:
            review_status = "Pending Review"
        elif not labor.labor_id:
            review_status = "Pending Labor Mapping"
        else:
            review_status = "Ready for Approval"
        legacy_name = text(source.get("Legacy Name"))
        legacy_type = text(source.get("Legacy Type"))
        description = text(source.get("Legacy Note")) or (
            f"Proposed canonical service derived from legacy record: {legacy_name}."
        )
        output.append(
            {
                "Service ID": service_id,
                "Legacy Service SKU": text(source.get("Legacy SKU")),
                "Active": "Yes",
                "Status": "Draft",
                "Manufacturer ID": manufacturer_id,
                "Manufacturer Name": manufacturer_name,
                "Device Family Code": family_code,
                "Device Family Name": device_family,
                "Device Series": text(source.get("Legacy Group")),
                "Device Model": legacy_name,
                "Service Category ID": "SC-REPAIR",
                "Service Category": "Repair",
                "Repair Type ID": repair_type_id,
                "Repair Type": repair_type_name or legacy_type,
                "Service Name": legacy_name,
                "Service Display Name": legacy_name,
                "Service Description": description,
                "Labor Standard ID": labor.labor_id,
                "Standard Minutes": labor.standard_minutes,
                "Minimum Minutes": labor.minimum_minutes,
                "Maximum Minutes": labor.maximum_minutes,
                "Labor Tier": labor.labor_tier,
                "Repair Difficulty": labor.difficulty,
                "Skill Level": labor.skill_level,
                "Turnaround Time": "To Be Determined",
                "Requires Parts": "Yes",
                "Requires Labor": "Yes",
                "Diagnostic Required": "No",
                "Warranty Eligible": "Yes",
                "Default Warranty": "N/A",
                "Mobile Service Eligible": "No",
                "Mail-In Eligible": "No",
                "Pricing Status": pricing_status(
                    source.get("Legacy Retail Price"), source.get("Legacy Cost")
                ),
                "Legacy Retail Price": source.get("Legacy Retail Price"),
                "Legacy Cost": source.get("Legacy Cost"),
                "Source Record Number": source_row,
                "Source Workbook": PROPOSAL_PATH.name,
                "Source Worksheet": "01 - Retained",
                "Import Batch ID": "MASTER-SERVICES-V1-DRAFT",
                "Review Status": review_status,
                "Reviewer Notes": text(conflict.get("Reviewer Notes")),
                "Effective Date": "",
                "Last Reviewed": "",
                "Created At": source.get("Legacy Created At", ""),
                "Updated At": source.get("Legacy Updated At", ""),
            }
        )
        labor_audit.append(
            {
                "Source Record Number": source_row,
                "Service ID": service_id,
                "Legacy Service Name": legacy_name,
                "Labor Standard ID": labor.labor_id,
                "Match Score": labor.match_score,
                "Second Best Score": labor.second_best_score,
                "Score Margin": labor.score_margin,
                "Match Evidence": labor.evidence,
                "Mapping Result": labor.mapping_result,
                "Mapped Minutes": labor.standard_minutes,
                "Mapped Labor Tier": labor.labor_tier,
                "Mapped Difficulty": labor.difficulty,
                "Mapped Skill Level": labor.skill_level,
            }
        )
    return output, labor_audit


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    records: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write an ASCII-safe, styled Excel Table and freeze its header."""
    worksheet.append([ascii_value(header) for header in headers])
    for record in records:
        worksheet.append([ascii_value(record.get(header, "")) for header in headers])
    if not records:
        worksheet.append(["" for _ in headers])
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    format_columns(worksheet, headers)


def append_table_at(
    worksheet: Worksheet,
    start_row: int,
    headers: Sequence[str],
    records: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Append a second documented table below an existing worksheet table."""
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=start_row, column=column, value=ascii_value(header))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_offset, record in enumerate(records, start=1):
        for column, header in enumerate(headers, start=1):
            worksheet.cell(
                row=start_row + row_offset,
                column=column,
                value=ascii_value(record.get(header, "")),
            )
    end_row = start_row + max(len(records), 1)
    end_column = get_column_letter(len(headers))
    table = Table(
        displayName=table_name,
        ref=f"A{start_row}:{end_column}{end_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    for column, header in enumerate(headers, start=1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = max(
            worksheet.column_dimensions[letter].width or 0,
            38 if header == "Match Evidence" else min(24, len(header) + 3),
        )


def format_columns(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Apply semantic number formats and bounded readable widths."""
    for column, header in enumerate(headers, start=1):
        letter = get_column_letter(column)
        width = max(12, min(38, len(header) + 3))
        if any(token in header for token in ("Description", "Notes", "Workbook")):
            width = 38
        worksheet.column_dimensions[letter].width = width
        if header in {"Legacy Retail Price", "Legacy Cost", "Hourly Rate"}:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = CURRENCY_FORMAT
        elif header in {"Standard Minutes", "Minimum Minutes", "Maximum Minutes"}:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = "0"
        elif "Date" in header or header in {"Last Reviewed", "Created At", "Updated At"}:
            number_format = DATETIME_FORMAT if header.endswith(" At") else DATE_FORMAT
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = number_format


def add_list_validation(
    worksheet: Worksheet,
    header: str,
    formula: str,
    last_row: int,
) -> None:
    """Apply a strict drop-down validation to a Master Services column."""
    headers = [cell.value for cell in worksheet[1]]
    column = headers.index(header) + 1
    letter = get_column_letter(column)
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = f"Select a controlled value for {header}."
    validation.errorTitle = "Invalid value"
    worksheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{last_row}")


def add_validation_defined_names(
    workbook: Workbook, lookup_lengths: dict[str, int]
) -> None:
    """Create workbook-defined names for every cross-sheet validation list."""
    defined_ranges = {
        "DV_YesNo": "'12 - Status Values'!$B$2:$B$3",
        "DV_ServiceStatuses": "'12 - Status Values'!$B$4:$B$8",
        "DV_PricingStatuses": "'12 - Status Values'!$B$9:$B$12",
        "DV_ReviewStatuses": "'12 - Status Values'!$B$13:$B$18",
        "DV_ManufacturerIDs": (
            f"'05 - Manufacturers'!$A$2:$A${lookup_lengths['manufacturers'] + 1}"
        ),
        "DV_DeviceFamilyCodes": (
            f"'04 - Device Families'!$A$2:$A${lookup_lengths['families'] + 1}"
        ),
        "DV_ServiceCategoryIDs": (
            f"'02 - Service Categories'!$A$2:$A${lookup_lengths['categories'] + 1}"
        ),
        "DV_RepairTypeIDs": (
            f"'03 - Repair Types'!$A$2:$A${lookup_lengths['repair_types'] + 1}"
        ),
        "DV_LaborStandardIDs": (
            f"'06 - Labor Standards'!$A$2:$A${lookup_lengths['labor'] + 1}"
        ),
        "DV_LaborTiers": (
            f"'07 - Labor Tiers'!$A$2:$A${lookup_lengths['tiers'] + 1}"
        ),
        "DV_DifficultyLevels": (
            f"'08 - Difficulty Levels'!$A$2:$A${lookup_lengths['difficulty'] + 1}"
        ),
        "DV_SkillLevels": (
            f"'09 - Skill Levels'!$A$2:$A${lookup_lengths['skills'] + 1}"
        ),
        "DV_TurnaroundTimes": (
            f"'10 - Turnaround Times'!$A$2:$A${lookup_lengths['turnaround'] + 1}"
        ),
        "DV_WarrantyOptions": (
            f"'11 - Warranty Options'!$A$2:$A${lookup_lengths['warranty'] + 1}"
        ),
    }
    for name, attr_text in defined_ranges.items():
        workbook.defined_names.add(DefinedName(name, attr_text=attr_text))


def add_master_validations(workbook: Workbook, lookup_lengths: dict[str, int]) -> None:
    """Add all requested Master Services drop-down validations."""
    worksheet = workbook[SHEET_NAMES[1]]
    end_row = max(worksheet.max_row, 2)
    add_validation_defined_names(workbook, lookup_lengths)
    for header, defined_name in DEFINED_NAME_BY_HEADER.items():
        add_list_validation(worksheet, header, f"={defined_name}", end_row)


def add_master_conditional_formatting(worksheet: Worksheet) -> None:
    """Highlight requested review states and missing lookup mappings."""
    headers = [cell.value for cell in worksheet[1]]
    end_row = worksheet.max_row
    for header, value, fill in (
        ("Review Status", "Pending Review", YELLOW_FILL),
        ("Review Status", "Pending Labor Mapping", ORANGE_FILL),
        ("Status", "Draft", YELLOW_FILL),
        ("Pricing Status", "Pending Pricing Review", YELLOW_FILL),
        ("Pricing Status", "Legacy Price Review", ORANGE_FILL),
    ):
        column = headers.index(header) + 1
        letter = get_column_letter(column)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{end_row}",
            CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill),
        )
    for header in ("Manufacturer ID", "Device Family Code", "Labor Standard ID"):
        column = headers.index(header) + 1
        letter = get_column_letter(column)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{end_row}",
            FormulaRule(formula=[f'LEN(TRIM(${letter}2))=0'], fill=RED_FILL),
        )


def build_workbook(
    service_rows: list[dict[str, Any]],
    labor_audit_rows: list[dict[str, Any]],
    repair_type_rows: list[dict[str, Any]],
    manufacturer_rows: list[dict[str, Any]],
    family_rows: list[dict[str, Any]],
    labor_rows: list[dict[str, Any]],
    tier_rows: list[dict[str, Any]],
    hashes: dict[Path, str],
    existing_service_ids: set[str],
    malformed_existing_ids: Sequence[str],
    generated_at: datetime,
) -> Workbook:
    """Build the complete 16-sheet review workbook in memory."""
    validate_sheet_names()
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}

    instructions = [
        {"Topic": "Purpose", "Guidance": "Proposed canonical Master Services review catalog."},
        {"Topic": "Scope", "Guidance": "Retained legacy Repair records only; no database import."},
        {"Topic": "Sources", "Guidance": f"{PROPOSAL_PATH}; {LABOR_PATH}; {CANONICAL_PATH}"},
        {"Topic": "Generation Date", "Guidance": generated_at},
        {"Topic": "Status", "Guidance": "1.0 Draft"},
        {"Topic": "Warning", "Guidance": "This is a proposed canonical catalog."},
        {"Topic": "Pricing", "Guidance": "No final pricing has been approved."},
        {"Topic": "Import Boundary", "Guidance": "Canonical workbook import requires separate approval."},
        {"Topic": "Review Workflow", "Guidance": "Resolve identity, lookup, labor, pricing, then approve import separately."},
        {"Topic": "Regeneration", "Guidance": "Run the generator from the project virtual environment after source review."},
    ]
    append_table(sheets[SHEET_NAMES[0]], ["Topic", "Guidance"], instructions, "tblMasterServicesInstructions")
    append_table(sheets[SHEET_NAMES[1]], MASTER_HEADERS, service_rows, "tblMasterServicesCatalog")

    category_rows = [
        {"Service Category ID": "SC-REPAIR", "Service Category": "Repair", "Description": "Device repair and related service work.", "Active": "Yes"}
    ]
    append_table(sheets[SHEET_NAMES[2]], ["Service Category ID", "Service Category", "Description", "Active"], category_rows, "tblServiceCategories")
    append_table(sheets[SHEET_NAMES[3]], ["Repair Type ID", "Repair Type", "Service Category ID", "Active"], repair_type_rows, "tblRepairTypes")
    append_table(sheets[SHEET_NAMES[4]], ["Device Family Code", "Device Family Name", "Description", "Active"], family_rows, "tblDeviceFamilies")
    manufacturer_output = [{"Manufacturer ID": row["ID"], "Manufacturer Name": row["Name"], "Source": row["Source"], "Active": row["Active"]} for row in manufacturer_rows]
    append_table(sheets[SHEET_NAMES[5]], ["Manufacturer ID", "Manufacturer Name", "Source", "Active"], manufacturer_output, "tblManufacturers")

    labor_headers = list(labor_rows[0]) if labor_rows else ["Labor ID"]
    tier_headers = list(tier_rows[0]) if tier_rows else ["Labor Rate Tier"]
    append_table(sheets[SHEET_NAMES[6]], labor_headers, labor_rows, "tblLaborStandardsLookup")
    append_table(sheets[SHEET_NAMES[7]], tier_headers, tier_rows, "tblLaborTiersLookup")
    difficulty = sorted({text(row.get("Repair Difficulty")) for row in labor_rows if text(row.get("Repair Difficulty"))})
    skills = sorted({text(row.get("Skill Level")) for row in labor_rows if text(row.get("Skill Level"))})
    turnaround = ["Same Day", "1-2 Business Days", "3-5 Business Days", "To Be Determined"]
    warranties = sorted(
        {"N/A"}
        | {
            text(first_value(row, ("Recommended Warranty", "Warranty")))
            for row in labor_rows
            if text(first_value(row, ("Recommended Warranty", "Warranty")))
        }
    )
    append_table(sheets[SHEET_NAMES[8]], ["Difficulty Level", "Description", "Active"], [{"Difficulty Level": value, "Description": "Labor catalog controlled value.", "Active": "Yes"} for value in difficulty], "tblDifficultyLevels")
    append_table(sheets[SHEET_NAMES[9]], ["Skill Level", "Description", "Active"], [{"Skill Level": value, "Description": "Labor catalog controlled value.", "Active": "Yes"} for value in skills], "tblSkillLevels")
    append_table(sheets[SHEET_NAMES[10]], ["Turnaround Time", "Description", "Active"], [{"Turnaround Time": value, "Description": "Proposed service turnaround option.", "Active": "Yes"} for value in turnaround], "tblTurnaroundTimes")
    append_table(sheets[SHEET_NAMES[11]], ["Warranty Option", "Description", "Active"], [{"Warranty Option": value, "Description": "Labor catalog warranty option.", "Active": "Yes"} for value in warranties], "tblWarrantyOptions")
    status_rows = ([{"Value Type": "Yes/No", "Value": value, "Active": "Yes"} for value in ("Yes", "No")] + [{"Value Type": "Service Status", "Value": value, "Active": "Yes"} for value in SERVICE_STATUSES] + [{"Value Type": "Pricing Status", "Value": value, "Active": "Yes"} for value in PRICING_STATUSES] + [{"Value Type": "Review Status", "Value": value, "Active": "Yes"} for value in REVIEW_STATUSES])
    append_table(sheets[SHEET_NAMES[12]], ["Value Type", "Value", "Active"], status_rows, "tblServiceStatusValues")

    labor_mapped = sum(1 for row in service_rows if row["Labor Standard ID"])
    validation_rows = [
        {"Validation Check": "Expected Repair population", "Result": "PASS" if len(service_rows) == EXPECTED_SERVICE_ROWS else "FAIL", "Evidence": len(service_rows)},
        {"Validation Check": "Unique Service IDs", "Result": "PASS" if len({row["Service ID"] for row in service_rows}) == len(service_rows) else "FAIL", "Evidence": len(service_rows)},
        {"Validation Check": "Labor mappings", "Result": "PASS", "Evidence": f"{labor_mapped} mapped; {len(service_rows) - labor_mapped} pending"},
        {"Validation Check": "Final pricing absent", "Result": "PASS", "Evidence": "Legacy observations only"},
        {"Validation Check": "Canonical import", "Result": "PASS", "Evidence": "Not performed"},
    ]
    append_table(sheets[SHEET_NAMES[13]], ["Validation Check", "Result", "Evidence"], validation_rows, "tblMasterServicesValidationSummary")
    append_table_at(
        sheets[SHEET_NAMES[13]],
        start_row=len(validation_rows) + 4,
        headers=LABOR_AUDIT_HEADERS,
        records=labor_audit_rows,
        table_name="tblLaborMatchAudit",
    )
    revision = [{"Version": "1.0 Draft", "Date": generated_at.date(), "Change Type": "Initial Creation", "Description": "Generated proposed canonical Master Services Catalog from reviewed retained legacy Repair records", "Prepared By": "Tamara Grandoit", "Approved By": "", "Status": "Draft", "Notes": "No canonical database import performed"}]
    append_table(sheets[SHEET_NAMES[14]], ["Version", "Date", "Change Type", "Description", "Prepared By", "Approved By", "Status", "Notes"], revision, "tblMasterServicesRevisionHistory")
    highest_existing = max(
        existing_service_ids, key=lambda identifier: int(identifier[3:])
    )
    metadata = [{"Metadata Field": "Artifact Type", "Value": "Proposed canonical Master Services Catalog"}, {"Metadata Field": "Generated At UTC", "Value": generated_at}, {"Metadata Field": "Expected Service Rows", "Value": EXPECTED_SERVICE_ROWS}, {"Metadata Field": "Highest Existing Service ID", "Value": highest_existing}, {"Metadata Field": "First Generated Service ID", "Value": service_rows[0]["Service ID"]}, {"Metadata Field": "Final Generated Service ID", "Value": service_rows[-1]["Service ID"]}, {"Metadata Field": "Malformed Existing Service IDs", "Value": "; ".join(malformed_existing_ids)}, {"Metadata Field": "Service ID Source Worksheet", "Value": SERVICE_ID_SOURCE_SHEET}, {"Metadata Field": "Labor Match Audit SHA-256", "Value": labor_audit_digest(labor_audit_rows)}, {"Metadata Field": "Output", "Value": str(OUTPUT_PATH)}]
    for path, digest in hashes.items():
        metadata.extend(({"Metadata Field": f"Protected Input: {path.name}", "Value": str(path)}, {"Metadata Field": f"SHA-256: {path.name}", "Value": digest}))
    append_table(sheets[SHEET_NAMES[15]], ["Metadata Field", "Value"], metadata, "tblMasterServicesImportMetadata")

    lengths = {"manufacturers": len(manufacturer_rows), "families": len(family_rows), "categories": len(category_rows), "repair_types": len(repair_type_rows), "labor": len(labor_rows), "tiers": len(tier_rows), "difficulty": len(difficulty), "skills": len(skills), "turnaround": len(turnaround), "warranty": len(warranties)}
    add_master_validations(workbook, lengths)
    add_master_conditional_formatting(sheets[SHEET_NAMES[1]])
    return workbook


def worksheet_records(worksheet: Worksheet) -> list[dict[str, Any]]:
    """Read a generated worksheet's primary header-row dataset."""
    headers = [text(cell.value) for cell in worksheet[1]]
    return [
        dict(zip(headers, values, strict=False))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None and text(value) != "" for value in values)
    ]


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read records from a named table that may not begin in row one."""
    if table_name not in worksheet.tables:
        raise CatalogError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = [text(value) for value in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:] if row[0]]


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Return the exact header sequence for a named table."""
    if table_name not in worksheet.tables:
        raise CatalogError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, _max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    return [
        text(worksheet.cell(row=min_row, column=column).value)
        for column in range(min_col, max_col + 1)
    ]


def validate_defined_names_and_validations(
    workbook: Workbook, master: Worksheet
) -> None:
    """Validate named-list definitions and every Master Services drop-down."""
    actual_names = set(workbook.defined_names.keys())
    missing_names = sorted(REQUIRED_DEFINED_NAMES - actual_names)
    if missing_names:
        raise CatalogError(f"Required defined names missing: {missing_names}")
    headers = [text(cell.value) for cell in master[1]]
    actual_by_header: dict[str, str] = {}
    for validation in master.data_validations.dataValidation:
        if validation.type != "list":
            continue
        formula = text(validation.formula1)
        if "!" in formula:
            raise CatalogError(
                f"Direct cross-sheet validation formula is prohibited: {formula}"
            )
        ranges = str(validation.sqref).split()
        if not ranges:
            raise CatalogError("List validation has no target range")
        min_col, _min_row, _max_col, _max_row = range_boundaries(ranges[0])
        header = headers[min_col - 1]
        actual_by_header[header] = formula
    for header, defined_name in DEFINED_NAME_BY_HEADER.items():
        expected = f"={defined_name}"
        if actual_by_header.get(header) != expected:
            raise CatalogError(
                f"Validation for {header!r} must reference {expected}, found "
                f"{actual_by_header.get(header)!r}"
            )


def id_name_lookup(
    worksheet: Worksheet, id_header: str, name_header: str
) -> dict[str, str]:
    """Return an ID-to-name map and reject duplicate inconsistent IDs."""
    result: dict[str, str] = {}
    for record in worksheet_records(worksheet):
        identifier = text(record.get(id_header))
        name = text(record.get(name_header))
        if not identifier:
            continue
        if identifier in result and result[identifier] != name:
            raise CatalogError(f"Lookup ID {identifier!r} resolves to multiple names")
        result[identifier] = name
    return result


def validate_service_relationships(workbook: Workbook, records: Sequence[dict[str, Any]]) -> None:
    """Validate conditional blanks and all ID/name relationships."""
    relationship_rules = {
        ("Manufacturer ID", "Manufacturer Name"): id_name_lookup(
            workbook[SHEET_NAMES[5]], "Manufacturer ID", "Manufacturer Name"
        ),
        ("Device Family Code", "Device Family Name"): id_name_lookup(
            workbook[SHEET_NAMES[4]], "Device Family Code", "Device Family Name"
        ),
        ("Repair Type ID", "Repair Type"): id_name_lookup(
            workbook[SHEET_NAMES[3]], "Repair Type ID", "Repair Type"
        ),
        ("Service Category ID", "Service Category"): id_name_lookup(
            workbook[SHEET_NAMES[2]], "Service Category ID", "Service Category"
        ),
    }
    for row_number, record in enumerate(records, start=2):
        review_status = text(record.get("Review Status"))
        manufacturer_id = text(record.get("Manufacturer ID"))
        family_code = text(record.get("Device Family Code"))
        labor_id = text(record.get("Labor Standard ID"))
        if not manufacturer_id and review_status != "Pending Review":
            raise CatalogError(
                f"Blank Manufacturer ID requires Pending Review in row {row_number}"
            )
        if not family_code and review_status != "Pending Review":
            raise CatalogError(
                f"Blank Device Family Code requires Pending Review in row {row_number}"
            )
        if not labor_id and review_status not in {
            "Pending Labor Mapping",
            "Pending Review",
        }:
            raise CatalogError(
                f"Blank Labor Standard ID has invalid status in row {row_number}"
            )
        if review_status == "Ready for Approval" and not all(
            (manufacturer_id, family_code, labor_id)
        ):
            raise CatalogError(
                f"Ready for Approval row {row_number} has a blank relationship"
            )
        for (id_field, name_field), lookup in relationship_rules.items():
            identifier = text(record.get(id_field))
            name = text(record.get(name_field))
            if identifier and lookup.get(identifier) != name:
                raise CatalogError(
                    f"{id_field}/{name_field} mismatch in row {row_number}: "
                    f"{identifier!r}/{name!r}"
                )


def validate_service_id_sequence(
    generated_ids: Sequence[str], existing_ids: set[str]
) -> tuple[str, str]:
    """Validate format, uniqueness, non-reuse, and continuous allocation."""
    if any(not SERVICE_ID_PATTERN.fullmatch(identifier) for identifier in generated_ids):
        raise CatalogError("A generated Service ID does not match ^SVC\\d{6}$")
    if len(generated_ids) != len(set(generated_ids)):
        raise CatalogError("Generated Service IDs are not unique")
    overlap = set(generated_ids) & existing_ids
    if overlap:
        raise CatalogError(f"Generated Service IDs already exist: {sorted(overlap)[:10]}")
    generated_numbers = [int(identifier[3:]) for identifier in generated_ids]
    expected_first = max(int(identifier[3:]) for identifier in existing_ids) + 1
    expected_numbers = list(
        range(expected_first, expected_first + len(generated_numbers))
    )
    if generated_numbers != expected_numbers:
        raise CatalogError("Generated Service IDs are not one continuous sequence")
    expected_final = expected_first + len(generated_numbers) - 1
    if generated_numbers[0] != expected_first:
        raise CatalogError("First generated Service ID is not highest existing plus one")
    if generated_numbers[-1] != expected_final:
        raise CatalogError("Final generated Service ID does not reconcile to row count")
    return f"SVC{expected_first:06d}", f"SVC{expected_final:06d}"


def validate_generated_workbook(
    path: Path,
    hashes_before: dict[Path, str],
    existing_ids: set[str],
    malformed_existing_ids: Sequence[str],
    expected_labor_audit_rows: Sequence[dict[str, Any]],
) -> list[str]:
    """Reopen the generated workbook and run essential structural checks."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    messages = []
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise CatalogError("Generated worksheet names/order are invalid")
        if len(set(workbook.sheetnames)) != 16 or any(len(name) > 31 for name in workbook.sheetnames):
            raise CatalogError("Generated worksheet names violate Excel constraints")
        for sheet_name, expected_tables in TABLE_NAMES_BY_SHEET.items():
            worksheet = workbook[sheet_name]
            if set(worksheet.tables) != expected_tables:
                raise CatalogError(
                    f"Excel Table contract is invalid on {sheet_name!r}"
                )
            if worksheet.freeze_panes != "A2":
                raise CatalogError(f"Header row is not frozen on {sheet_name!r}")
            if not worksheet.auto_filter.ref:
                raise CatalogError(f"Filters are not enabled on {sheet_name!r}")
        master = workbook[SHEET_NAMES[1]]
        if "tblMasterServicesCatalog" not in master.tables:
            raise CatalogError("tblMasterServicesCatalog is missing")
        headers = [cell.value for cell in master[1]]
        if headers != MASTER_HEADERS:
            raise CatalogError("Master Services schema is invalid")
        records = list(master.iter_rows(min_row=2, values_only=True))
        records = [row for row in records if row[0]]
        if len(records) != EXPECTED_SERVICE_ROWS:
            raise CatalogError(f"Expected {EXPECTED_SERVICE_ROWS} services, found {len(records)}")
        ids = [text(row[0]) for row in records]
        first_id, final_id = validate_service_id_sequence(ids, existing_ids)
        if any(header in headers for header in ("Final Customer Price", "Final Price", "Final Cost")):
            raise CatalogError("A prohibited final pricing field exists")
        service_records = worksheet_records(master)
        validate_defined_names_and_validations(workbook, master)
        warranty_values = {
            text(record.get("Warranty Option"))
            for record in worksheet_records(workbook[SHEET_NAMES[11]])
        }
        if "N/A" not in warranty_values:
            raise CatalogError("Warranty Options must contain N/A")
        if any(
            text(record.get("Default Warranty")) not in warranty_values
            for record in service_records
        ):
            raise CatalogError("A service Default Warranty is not a valid lookup")
        validate_service_relationships(workbook, service_records)
        labor_audit = table_records(
            workbook[SHEET_NAMES[13]], "tblLaborMatchAudit"
        )
        if (
            table_headers(workbook[SHEET_NAMES[13]], "tblLaborMatchAudit")
            != LABOR_AUDIT_HEADERS
        ):
            raise CatalogError("Labor match audit schema is invalid")
        if len(labor_audit) != EXPECTED_SERVICE_ROWS:
            raise CatalogError("Labor match audit does not cover all service rows")
        for audit_row_number, (expected, reopened) in enumerate(
            zip(expected_labor_audit_rows, labor_audit, strict=True),
            start=2,
        ):
            service_id = normalized_text(
                reopened.get("Service ID") or expected.get("Service ID")
            )
            for field in LABOR_AUDIT_HEADERS:
                expected_value = expected.get(field)
                reopened_value = reopened.get(field)
                if not persisted_values_equal(
                    field,
                    expected_value,
                    reopened_value,
                ):
                    raise CatalogError(
                        f"Audit row {audit_row_number} ({service_id}) changed "
                        f"{field}: expected {expected_value!r} "
                        f"({type(expected_value).__name__}), reopened "
                        f"{reopened_value!r} "
                        f"({type(reopened_value).__name__})"
                    )
        service_ids = [text(record.get("Service ID")) for record in service_records]
        audit_ids = [text(record.get("Service ID")) for record in labor_audit]
        if audit_ids != service_ids:
            raise CatalogError(
                "Labor match audit rows do not preserve service order and identity"
            )
        if len(audit_ids) != len(set(audit_ids)):
            raise CatalogError("Labor match audit contains duplicate Service IDs")
        if any(not SERVICE_ID_PATTERN.fullmatch(identifier) for identifier in audit_ids):
            raise CatalogError("Labor match audit contains an invalid Service ID")
        services_by_id = {
            text(record.get("Service ID")): record for record in service_records
        }
        labor_by_id = {
            text(record.get("Labor ID")): record
            for record in worksheet_records(workbook[SHEET_NAMES[6]])
            if text(record.get("Labor ID"))
        }
        for record in labor_audit:
            service_id = text(record.get("Service ID"))
            service = services_by_id[service_id]
            result = text(record.get("Mapping Result"))
            if result not in {"Mapped", "Pending Labor Mapping", "Ambiguous"}:
                raise CatalogError(
                    "Labor match audit contains an invalid Mapping Result"
                )
            if record.get("Source Record Number") != service.get(
                "Source Record Number"
            ):
                raise CatalogError(
                    f"Labor audit source row changed for {service_id}"
                )
            if text(record.get("Legacy Service Name")) != text(
                service.get("Service Name")
            ):
                raise CatalogError(
                    f"Labor audit legacy service name changed for {service_id}"
                )
            audit_labor_id = text(record.get("Labor Standard ID"))
            service_labor_id = text(service.get("Labor Standard ID"))
            match_score = float(record.get("Match Score") or 0)
            second_score = float(record.get("Second Best Score") or 0)
            score_margin = float(record.get("Score Margin") or 0)
            if abs((match_score - second_score) - score_margin) > 1e-9:
                raise CatalogError(
                    f"Labor audit score margin is inconsistent for {service_id}"
                )
            if not text(record.get("Match Evidence")):
                raise CatalogError(
                    f"Labor audit evidence is blank for {service_id}"
                )
            if result == "Mapped" and (
                not audit_labor_id or audit_labor_id != service_labor_id
            ):
                raise CatalogError("Mapped labor audit row does not reconcile")
            if result != "Mapped" and (audit_labor_id or service_labor_id):
                raise CatalogError("Unresolved labor audit row populated a labor ID")
            if result == "Mapped" and (
                match_score <= LABOR_MATCH_THRESHOLD
                or score_margin <= LABOR_TIE_MARGIN
            ):
                raise CatalogError("Mapped labor audit row is below or tied at threshold")
            if result == "Ambiguous" and (
                match_score <= LABOR_MATCH_THRESHOLD
                or score_margin > LABOR_TIE_MARGIN
            ):
                raise CatalogError("Ambiguous labor audit scores are inconsistent")
            if result == "Pending Labor Mapping" and (
                match_score > LABOR_MATCH_THRESHOLD
                and score_margin > LABOR_TIE_MARGIN
            ):
                raise CatalogError(
                    "Eligible labor audit row was left Pending Labor Mapping"
                )
            mapped_fields = {
                "Mapped Minutes": service.get("Standard Minutes", ""),
                "Mapped Labor Tier": text(service.get("Labor Tier")),
                "Mapped Difficulty": text(service.get("Repair Difficulty")),
                "Mapped Skill Level": text(service.get("Skill Level")),
            }
            if result == "Mapped":
                labor = labor_by_id.get(audit_labor_id)
                if labor is None:
                    raise CatalogError(
                        f"Labor audit references unknown Labor Standard ID "
                        f"{audit_labor_id!r}"
                    )
                expected_service_labor_fields = {
                    "Standard Minutes": labor.get("Standard Minutes", ""),
                    "Minimum Minutes": labor.get("Minimum Minutes", ""),
                    "Maximum Minutes": labor.get("Maximum Minutes", ""),
                    "Labor Tier": text(labor.get("Labor Rate Tier")),
                    "Repair Difficulty": text(labor.get("Repair Difficulty")),
                    "Skill Level": text(labor.get("Skill Level")),
                }
                for field, expected_value in expected_service_labor_fields.items():
                    actual_value = service.get(field)
                    if not persisted_values_equal(
                        field,
                        expected_value,
                        actual_value,
                    ):
                        raise CatalogError(
                            f"Audit row for {service_id} changed {field}: expected "
                            f"{expected_value!r} ({type(expected_value).__name__}), "
                            f"reopened {actual_value!r} "
                            f"({type(actual_value).__name__})"
                        )
                expected_labor_fields = {
                    "Mapped Minutes": labor.get("Standard Minutes", ""),
                    "Mapped Labor Tier": text(labor.get("Labor Rate Tier")),
                    "Mapped Difficulty": text(labor.get("Repair Difficulty")),
                    "Mapped Skill Level": text(labor.get("Skill Level")),
                }
                for comparisons in (mapped_fields, expected_labor_fields):
                    for field, expected_value in comparisons.items():
                        actual_value = record.get(field)
                        if not persisted_values_equal(
                            field,
                            expected_value,
                            actual_value,
                        ):
                            raise CatalogError(
                                f"Audit row for {service_id} changed {field}: "
                                f"expected {expected_value!r} "
                                f"({type(expected_value).__name__}), reopened "
                                f"{actual_value!r} "
                                f"({type(actual_value).__name__})"
                            )
            elif any(text(record.get(field)) for field in mapped_fields):
                raise CatalogError(
                    f"Unresolved labor audit row has mapped values for {service_id}"
                )
        metadata_values = {
            text(record.get("Metadata Field")): text(record.get("Value"))
            for record in worksheet_records(workbook[SHEET_NAMES[15]])
        }
        expected_audit_digest = metadata_values.get("Labor Match Audit SHA-256")
        if not expected_audit_digest:
            raise CatalogError("Labor match audit digest metadata is missing")
        if labor_audit_digest(labor_audit) != expected_audit_digest:
            raise CatalogError("Labor match audit values changed after generation")
        table_names = [name for sheet in workbook.worksheets for name in sheet.tables]
        if len(table_names) != len(set(table_names)):
            raise CatalogError("Excel Table names are not unique")
        messages.append(f"Workbook structure: PASS ({len(records)} services, 16 sheets)")
        messages.append(
            f"Service ID sequence: PASS ({first_id} through {final_id})"
        )
        messages.append(
            "Malformed existing Service IDs excluded from allocation: "
            f"{len(malformed_existing_ids)}"
        )
        messages.append("Defined names and list validations: PASS")
        messages.append("Warranty and ID/name relationships: PASS")
        messages.append("Labor match audit: PASS (314 rows)")
        messages.append(f"Excel Tables: PASS ({len(table_names)} unique names)")
    finally:
        workbook.close()
    hashes_after = {input_path: sha256_file(input_path) for input_path in hashes_before}
    changed = [str(input_path) for input_path in hashes_before if hashes_before[input_path] != hashes_after[input_path]]
    if changed:
        raise CatalogError(f"Protected input hash changed: {', '.join(changed)}")
    messages.append("Protected input hashes: PASS (unchanged)")
    return messages


def main() -> int:
    """Generate and validate the standalone proposed Master Services workbook."""
    try:
        assert_persisted_comparison_policy()
        protected = (
            RAW_PATH,
            STAGING_PATH,
            PROPOSAL_PATH,
            LABOR_PATH,
            CANONICAL_PATH,
        )
        existing_service_ids, malformed_existing_ids = read_existing_service_ids(
            SERVICE_ID_SOURCE_PATH
        )
        require_files(protected)
        hashes = {path: sha256_file(path) for path in protected}
        retained = read_sheet_records(PROPOSAL_PATH, "01 - Retained")
        repairs = [row for row in retained if text(row.get("Record Category")) == "Repair"]
        if len(repairs) != EXPECTED_SERVICE_ROWS:
            raise CatalogError(f"Expected {EXPECTED_SERVICE_ROWS} retained Repair rows, found {len(repairs)}")
        conflicts = read_sheet_records(PROPOSAL_PATH, "03 - SKU Conflicts")
        conflict_decisions = {int(row["Source Row Number"]): row for row in conflicts if text(row.get("Source Row Number"))}
        labor_rows = read_sheet_records(LABOR_PATH, "01 - Labor Standards")
        tier_rows = read_sheet_records(LABOR_PATH, "02 - Labor Rate Tiers")
        canonical_manufacturers = read_table(CANONICAL_PATH, "tblManufacturerCatalog")
        canonical_families = read_table(CANONICAL_PATH, "tblDeviceCatalog")
        manufacturer_rows, manufacturer_lookup = proposed_lookup(canonical_manufacturers, (text(row.get("Legacy Manufacturer")) for row in repairs), ("Manufacturer ID", "Manufacturer Code"), ("Manufacturer Name", "Manufacturer"), "MFR-LEGACY")
        repair_type_rows, repair_type_lookup = create_repair_types(repairs)
        family_rows, family_lookup = create_device_families(
            canonical_families, repairs
        )
        first_service_number = next_service_number(
            existing_service_ids, len(repairs)
        )
        generated_at = datetime.now(UTC).replace(tzinfo=None)
        service_rows, labor_audit_rows = build_service_rows(
            repairs,
            labor_rows,
            manufacturer_lookup,
            repair_type_lookup,
            family_lookup,
            conflict_decisions,
            first_service_number,
        )
        workbook = build_workbook(
            service_rows,
            labor_audit_rows,
            repair_type_rows,
            manufacturer_rows,
            family_rows,
            labor_rows,
            tier_rows,
            hashes,
            existing_service_ids,
            malformed_existing_ids,
            generated_at,
        )
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(OUTPUT_PATH)
        workbook.close()
        messages = validate_generated_workbook(
            OUTPUT_PATH,
            hashes,
            existing_service_ids,
            malformed_existing_ids,
            labor_audit_rows,
        )
        print(f"Generated: {OUTPUT_PATH}")
        print(f"Master Services rows: {len(service_rows)}")
        print(
            f"Service ID range: {service_rows[0]['Service ID']} through "
            f"{service_rows[-1]['Service ID']}"
        )
        if malformed_existing_ids:
            print(
                "Malformed existing Service IDs excluded: "
                f"{'; '.join(malformed_existing_ids)}"
            )
        print(f"Labor mapped: {sum(1 for row in service_rows if row['Labor Standard ID'])}")
        print(f"Pending labor mapping: {sum(1 for row in service_rows if row['Review Status'] == 'Pending Labor Mapping')}")
        for message in messages:
            print(message)
        return 0
    except (CatalogError, OSError, ValueError, KeyError, IndexError) as exc:
        print(f"ERROR: {ascii_value(str(exc))}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
