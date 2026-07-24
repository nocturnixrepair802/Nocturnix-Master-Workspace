"""Read-only, profile-driven QA validation for Nocturnix review workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

Severity = Literal["ERROR", "WARNING", "INFO"]
Status = Literal["PASS", "PASS WITH WARNINGS", "FAIL"]
Category = Literal["CONFIGURATION", "WORKBOOK", "BUSINESS_RULE"]

DEFAULT_THRESHOLDS = {
    "max_errors_for_pass": 0,
    "max_warnings_for_pass": 0,
    "max_errors_for_warning_status": 0,
}
HIDDEN_CHARACTER_PATTERN = re.compile(
    "[\u0000-\u0008\u000b\u000c\u000e-\u001f\u007f\u200b-\u200f"
    "\u202a-\u202e\u2060\ufeff]"
)


@dataclass(frozen=True)
class Finding:
    """One machine-readable QA finding."""

    severity: Severity
    category: Category
    code: str
    message: str
    sheet: str | None = None
    cell: str | None = None
    value: Any = None
    suggestion: str | None = None


@dataclass(frozen=True)
class QAResult:
    """Complete QA result suitable for Markdown and JSON output."""

    workbook: str
    generated_at_utc: str
    profile: str
    status: Status
    counts: dict[str, int]
    category_counts: dict[str, int]
    workbook_sha256: str
    findings: list[Finding]
    metrics: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable result."""
        data = asdict(self)
        data["findings"] = [asdict(finding) for finding in self.findings]
        return data


def text(value: Any) -> str:
    """Return display text without changing the underlying workbook value."""
    return "" if value is None else str(value)


def stripped(value: Any) -> str:
    """Return trimmed text for validation comparisons."""
    return text(value).strip()


def file_hash(path: Path) -> str:
    """Return a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_config(path: Path) -> dict[str, Any]:
    """Load and minimally validate a JSON QA profile."""
    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON profile: {exc}") from exc
    if not isinstance(config, dict):
        raise ValueError("QA profile root must be a JSON object")
    if not isinstance(config.get("sheets", {}), dict):
        raise ValueError("QA profile 'sheets' must be an object")
    return config


def add(
    findings: list[Finding],
    severity: Severity,
    code: str,
    message: str,
    *,
    category: Category = "WORKBOOK",
    sheet: str | None = None,
    cell: str | None = None,
    value: Any = None,
    suggestion: str | None = None,
) -> None:
    """Append a finding using a compact call site."""
    findings.append(
        Finding(
            severity,
            category,
            code,
            message,
            sheet,
            cell,
            value,
            suggestion,
        )
    )


def check_enabled(sheet_config: Mapping[str, Any], name: str) -> bool:
    """Return whether one worksheet-specific validation family is enabled."""
    checks = sheet_config.get("checks", {})
    return bool(checks.get(name, True))


def validate_profile(config: Mapping[str, Any], findings: list[Finding]) -> None:
    """Report invalid profile settings separately from workbook defects."""
    valid_checks = {
        "required_fields",
        "headers",
        "row_count",
        "duplicates",
        "business_rules",
        "freeze_panes",
        "data_validations",
    }
    for sheet_name, sheet_config in config.get("sheets", {}).items():
        if not isinstance(sheet_config, dict):
            add(
                findings,
                "ERROR",
                "CONFIG_SHEET_PROFILE",
                "Worksheet profile must be a JSON object.",
                category="CONFIGURATION",
                sheet=sheet_name,
            )
            continue
        checks = sheet_config.get("checks", {})
        if not isinstance(checks, dict):
            add(
                findings,
                "ERROR",
                "CONFIG_CHECKS",
                "Worksheet 'checks' must be a JSON object.",
                category="CONFIGURATION",
                sheet=sheet_name,
            )
        else:
            unknown = sorted(set(checks) - valid_checks)
            non_boolean = sorted(
                key for key, value in checks.items() if not isinstance(value, bool)
            )
            if unknown:
                add(
                    findings,
                    "ERROR",
                    "CONFIG_UNKNOWN_CHECK",
                    "Worksheet profile contains unknown check switches.",
                    category="CONFIGURATION",
                    sheet=sheet_name,
                    value=unknown,
                )
            if non_boolean:
                add(
                    findings,
                    "ERROR",
                    "CONFIG_CHECK_TYPE",
                    "Worksheet check switches must be true or false.",
                    category="CONFIGURATION",
                    sheet=sheet_name,
                    value=non_boolean,
                )
        if sheet_config.get("data_scope", "table") not in {"table", "worksheet"}:
            add(
                findings,
                "ERROR",
                "CONFIG_DATA_SCOPE",
                "data_scope must be 'table' or 'worksheet'.",
                category="CONFIGURATION",
                sheet=sheet_name,
                value=sheet_config.get("data_scope"),
            )


def validate_package(path: Path, findings: list[Finding]) -> None:
    """Validate basic OOXML container integrity without extracting it."""
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        add(
            findings,
            "ERROR",
            "INTEGRITY_EXTENSION",
            "Workbook must be an .xlsx or .xlsm OOXML package.",
            value=path.suffix,
        )
        return
    required_members = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    try:
        with zipfile.ZipFile(path) as archive:
            members = set(archive.namelist())
            corrupt_member = archive.testzip()
    except zipfile.BadZipFile:
        add(
            findings,
            "ERROR",
            "INTEGRITY_ZIP",
            "File is not a valid OOXML ZIP package.",
        )
        return
    missing = sorted(required_members - members)
    if missing:
        add(
            findings,
            "ERROR",
            "INTEGRITY_MEMBERS",
            "Required OOXML package members are missing.",
            value=missing,
        )
    if corrupt_member:
        add(
            findings,
            "ERROR",
            "INTEGRITY_CORRUPT_MEMBER",
            "OOXML package contains a corrupt member.",
            value=corrupt_member,
        )


def header_map(
    worksheet: Worksheet,
    header_row: int,
    min_column: int,
    max_column: int,
    findings: list[Finding],
    *,
    validate_headers: bool,
) -> tuple[list[str], dict[str, int]]:
    """Read one header row and return exact headers plus unique indexes."""
    headers = [
        stripped(worksheet.cell(header_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    while headers and not headers[-1]:
        headers.pop()
    positions: dict[str, int] = {}
    counts = Counter(header for header in headers if header)
    for column, header in enumerate(headers, start=min_column):
        if not header and validate_headers:
            add(
                findings,
                "WARNING",
                "HEADER_BLANK",
                "Blank header found inside the populated header range.",
                sheet=worksheet.title,
                cell=f"{get_column_letter(column)}{header_row}",
                suggestion="Name or remove the blank column before import.",
            )
        elif counts[header] > 1 and validate_headers:
            add(
                findings,
                "ERROR",
                "HEADER_DUPLICATE",
                f"Duplicate header: {header}",
                sheet=worksheet.title,
                cell=f"{get_column_letter(column)}{header_row}",
                value=header,
            )
        elif header and counts[header] == 1:
            positions[header] = column
    return headers, positions


def iter_data_rows(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> Iterable[tuple[int, tuple[Any, ...]]]:
    """Yield source rows once in deterministic worksheet order."""
    rows = worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
        values_only=True,
    )
    for row_number, values in enumerate(rows, start=min_row):
        if any(stripped(value) for value in values):
            yield row_number, values


def data_region(
    worksheet: Worksheet,
    sheet_config: Mapping[str, Any],
    findings: list[Finding],
) -> tuple[int, int, int, int, int] | None:
    """Resolve header and data bounds, preferring an Excel Table body."""
    if sheet_config.get("data_scope", "table") == "worksheet":
        header_row = int(sheet_config.get("header_row", 1))
        return (
            header_row,
            header_row + 1,
            worksheet.max_row,
            1,
            worksheet.max_column,
        )
    table_name = sheet_config.get("table_name")
    if table_name:
        if table_name not in worksheet.tables:
            add(
                findings,
                "ERROR",
                "CONFIG_TABLE_NOT_FOUND",
                "Configured Excel Table was not found.",
                category="CONFIGURATION",
                sheet=worksheet.title,
                value=table_name,
            )
            return None
        table = worksheet.tables[table_name]
    else:
        tables = list(worksheet.tables.values())
        if len(tables) != 1:
            add(
                findings,
                "ERROR",
                "CONFIG_TABLE_SELECTION",
                "Table-scoped validation requires exactly one table or table_name.",
                category="CONFIGURATION",
                sheet=worksheet.title,
                value=list(worksheet.tables),
            )
            return None
        table = tables[0]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    return min_row, min_row + 1, max_row, min_column, max_column


def validate_sheet_structure(
    worksheet: Worksheet,
    sheet_config: Mapping[str, Any],
    findings: list[Finding],
) -> tuple[list[str], dict[str, int], tuple[int, int, int, int, int] | None]:
    """Validate headers, merges, visibility, filters, and freeze panes."""
    region = data_region(worksheet, sheet_config, findings)
    if region is None:
        return [], {}, None
    header_row, _, _, min_column, max_column = region
    if header_row < 1 or header_row > worksheet.max_row:
        add(
            findings,
            "ERROR",
            "HEADER_ROW_MISSING",
            f"Configured header row {header_row} is outside the used range.",
            sheet=worksheet.title,
        )
        return [], {}, region
    headers, positions = header_map(
        worksheet,
        header_row,
        min_column,
        max_column,
        findings,
        validate_headers=check_enabled(sheet_config, "headers"),
    )
    if check_enabled(sheet_config, "headers"):
        expected_headers = list(sheet_config.get("expected_headers", []))
        required_headers = set(
            sheet_config.get("required_headers", expected_headers)
        )
        missing = sorted(required_headers - positions.keys())
        if missing:
            add(
                findings,
                "ERROR",
                "HEADERS_REQUIRED_MISSING",
                "Required headers are missing.",
                sheet=worksheet.title,
                value=missing,
            )
        if expected_headers and headers != expected_headers:
            add(
                findings,
                "ERROR",
                "HEADERS_INCONSISTENT",
                "Header names or order differ from the configured contract.",
                sheet=worksheet.title,
                value={"expected": expected_headers, "actual": headers},
            )
    merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
    merge_policy = sheet_config.get("merged_cells", "warn")
    if merged_ranges and merge_policy != "allow":
        add(
            findings,
            "ERROR" if merge_policy == "error" else "WARNING",
            "STRUCTURE_MERGED_CELLS",
            "Merged cells may prevent deterministic tabular import.",
            sheet=worksheet.title,
            value=merged_ranges,
            suggestion="Unmerge cells in import-oriented worksheets.",
        )
    hidden_rows = [
        index
        for index, dimension in worksheet.row_dimensions.items()
        if dimension.hidden
    ]
    hidden_columns = [
        index
        for index, dimension in worksheet.column_dimensions.items()
        if dimension.hidden
    ]
    hidden_policy = sheet_config.get("hidden", "warn")
    if hidden_rows or hidden_columns:
        add(
            findings,
            "ERROR" if hidden_policy == "error" else "WARNING",
            "STRUCTURE_HIDDEN",
            "Hidden rows or columns can conceal import data.",
            sheet=worksheet.title,
            value={"rows": hidden_rows, "columns": hidden_columns},
        )
    if sheet_config.get("require_filter", False) and not (
        worksheet.auto_filter.ref or worksheet.tables
    ):
        add(
            findings,
            "WARNING",
            "USABILITY_FILTER_MISSING",
            "No worksheet filter or Excel Table filter is present.",
            sheet=worksheet.title,
            suggestion="Enable filtering for reviewer usability.",
        )
    expected_freeze = sheet_config.get("freeze_panes")
    actual_freeze = (
        worksheet.freeze_panes.coordinate
        if hasattr(worksheet.freeze_panes, "coordinate")
        else worksheet.freeze_panes
    )
    if (
        check_enabled(sheet_config, "freeze_panes")
        and expected_freeze is not None
        and actual_freeze != expected_freeze
    ):
        add(
            findings,
            "WARNING",
            "USABILITY_FREEZE_PANES",
            "Freeze panes differ from the configured review layout.",
            sheet=worksheet.title,
            value={"expected": expected_freeze, "actual": actual_freeze},
        )
    return headers, positions, region


def validate_text_quality(
    worksheet: Worksheet,
    row_number: int,
    values: Sequence[Any],
    findings: list[Finding],
    min_column: int,
) -> None:
    """Detect surrounding whitespace and invisible control characters."""
    for column, value in enumerate(values, start=min_column):
        if not isinstance(value, str) or not value:
            continue
        cell = f"{get_column_letter(column)}{row_number}"
        if value != value.strip():
            add(
                findings,
                "WARNING",
                "TEXT_SURROUNDING_WHITESPACE",
                "Text contains leading or trailing whitespace.",
                sheet=worksheet.title,
                cell=cell,
                value=repr(value),
                suggestion="Trim the value before import.",
            )
        hidden = sorted(set(HIDDEN_CHARACTER_PATTERN.findall(value)))
        if hidden:
            add(
                findings,
                "WARNING",
                "TEXT_HIDDEN_CHARACTER",
                "Text contains control or invisible Unicode characters.",
                sheet=worksheet.title,
                cell=cell,
                value=[f"U+{ord(character):04X}" for character in hidden],
                suggestion="Remove invisible characters before import.",
            )


def duplicate_key(
    values: Sequence[Any],
    positions: Mapping[str, int],
    fields: Sequence[str],
) -> tuple[str, ...] | None:
    """Return a normalized duplicate key, or None for an all-blank key."""
    key = tuple(stripped(values[positions[field] - 1]).casefold() for field in fields)
    return key if any(key) else None


def validate_rule(
    worksheet: Worksheet,
    row_number: int,
    values: Sequence[Any],
    positions: Mapping[str, int],
    min_column: int,
    rule: Mapping[str, Any],
    findings: list[Finding],
) -> None:
    """Apply one configured row-level business rule."""
    rule_type = str(rule.get("type", ""))
    severity: Severity = rule.get("severity", "ERROR")
    code = str(rule.get("code", "BUSINESS_RULE"))
    fields = list(rule.get("fields", []))
    missing_fields = [field for field in fields if field not in positions]
    if missing_fields:
        return
    row_values = {
        field: values[positions[field] - 1]
        for field in fields
    }
    failure = False
    if rule_type == "allowed_values":
        allowed = {str(value).casefold() for value in rule.get("values", [])}
        failure = any(
            stripped(value).casefold() not in allowed
            for value in row_values.values()
            if stripped(value)
        )
    elif rule_type == "regex":
        pattern = re.compile(str(rule.get("pattern", "")))
        failure = any(
            stripped(value) and not pattern.fullmatch(stripped(value))
            for value in row_values.values()
        )
    elif rule_type == "paired_fields":
        failure = len({bool(stripped(value)) for value in row_values.values()}) > 1
    elif rule_type == "forbidden_values":
        forbidden = {
            str(value).casefold() for value in rule.get("values", [])
        }
        failure = any(
            stripped(value).casefold() in forbidden
            for value in row_values.values()
        )
    if failure:
        first_field = fields[0]
        add(
            findings,
            severity,
            code,
            str(rule.get("message", "Configured business rule failed.")),
            category="BUSINESS_RULE",
            sheet=worksheet.title,
            cell=(
                f"{get_column_letter(positions[first_field] + min_column - 1)}"
                f"{row_number}"
            ),
            value={field: text(value) for field, value in row_values.items()},
            suggestion=rule.get("suggestion"),
        )


def validate_sheet_data(
    worksheet: Worksheet,
    sheet_config: Mapping[str, Any],
    headers: Sequence[str],
    positions: Mapping[str, int],
    region: tuple[int, int, int, int, int] | None,
    findings: list[Finding],
) -> int:
    """Validate required values, duplicates, text quality, and business rules."""
    if not headers or region is None:
        return 0
    _, min_row, max_row, min_column, max_column = region
    row_positions = {
        field: column - min_column + 1
        for field, column in positions.items()
    }
    required_fields = [
        field
        for field in sheet_config.get("required_fields", [])
        if field in positions
    ] if check_enabled(sheet_config, "required_fields") else []
    duplicate_keys = [
        list(fields)
        for fields in sheet_config.get("duplicate_keys", [])
        if all(field in positions for field in fields)
    ] if check_enabled(sheet_config, "duplicates") else []
    seen: list[dict[tuple[str, ...], int]] = [
        {} for _ in duplicate_keys
    ]
    row_count = 0
    for row_number, values in iter_data_rows(
        worksheet,
        min_row,
        max_row,
        min_column,
        max_column,
    ):
        row_count += 1
        validate_text_quality(
            worksheet,
            row_number,
            values,
            findings,
            min_column,
        )
        for field in required_fields:
            absolute_column = positions[field]
            value_index = absolute_column - min_column
            if not stripped(values[value_index]):
                add(
                    findings,
                    "ERROR",
                    "DATA_REQUIRED_BLANK",
                    f"Required field is blank: {field}",
                    sheet=worksheet.title,
                    cell=f"{get_column_letter(absolute_column)}{row_number}",
                )
        for index, fields in enumerate(duplicate_keys):
            key = duplicate_key(values, row_positions, fields)
            if key is None:
                continue
            if key in seen[index]:
                add(
                    findings,
                    "ERROR",
                    "DATA_DUPLICATE_MAPPING",
                    "Duplicate mapping key found.",
                    sheet=worksheet.title,
                    cell=f"A{row_number}",
                    value={
                        "fields": fields,
                        "key": list(key),
                        "first_row": seen[index][key],
                        "duplicate_row": row_number,
                    },
                )
            else:
                seen[index][key] = row_number
        if check_enabled(sheet_config, "business_rules"):
            for rule in sheet_config.get("business_rules", []):
                validate_rule(
                    worksheet,
                    row_number,
                    values,
                    row_positions,
                    min_column,
                    rule,
                    findings,
                )
    expected_rows = sheet_config.get("expected_data_rows")
    if (
        check_enabled(sheet_config, "row_count")
        and expected_rows is not None
        and row_count != int(expected_rows)
    ):
        add(
            findings,
            "ERROR",
            "DATA_ROW_COUNT",
            "Populated data-row count differs from the configured contract.",
            sheet=worksheet.title,
            value={"expected": int(expected_rows), "actual": row_count},
        )
    return row_count


def validate_data_validation(
    worksheet: Worksheet,
    sheet_config: Mapping[str, Any],
    positions: Mapping[str, int],
    findings: list[Finding],
) -> None:
    """Check that configured editable fields have list validation."""
    if not check_enabled(sheet_config, "data_validations"):
        return
    validation_fields = [
        field
        for field in sheet_config.get("validation_fields", [])
        if field in positions
    ]
    if not validation_fields:
        return
    covered_columns: set[int] = set()
    for validation in worksheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        for cell_range in validation.ranges.ranges:
            covered_columns.update(
                range(cell_range.min_col, cell_range.max_col + 1)
            )
    for field in validation_fields:
        if positions[field] not in covered_columns:
            add(
                findings,
                "WARNING",
                "VALIDATION_LIST_MISSING",
                f"Editable categorical field lacks list validation: {field}",
                sheet=worksheet.title,
                cell=f"{get_column_letter(positions[field])}1",
                suggestion="Add a bounded list validation before reviewer use.",
            )


def determine_status(
    counts: Mapping[str, int],
    thresholds: Mapping[str, Any],
) -> Status:
    """Apply configurable thresholds to produce the final status."""
    errors = counts.get("ERROR", 0)
    warnings = counts.get("WARNING", 0)
    max_errors_for_warning = int(
        thresholds.get("max_errors_for_warning_status", 0)
    )
    max_errors_for_pass = int(thresholds.get("max_errors_for_pass", 0))
    max_warnings_for_pass = int(thresholds.get("max_warnings_for_pass", 0))
    if errors > max_errors_for_warning:
        return "FAIL"
    if errors > max_errors_for_pass or warnings > max_warnings_for_pass:
        return "PASS WITH WARNINGS"
    return "PASS"


def validate_workbook(path: Path, config: Mapping[str, Any]) -> QAResult:
    """Run a complete read-only QA validation."""
    findings: list[Finding] = []
    validate_profile(config, findings)
    validate_package(path, findings)
    before_hash = file_hash(path)
    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=True,
    )
    metrics: dict[str, Any] = {"sheets": {}}
    try:
        expected_sheets = list(config.get("expected_sheets", []))
        actual_sheets = workbook.sheetnames
        if expected_sheets and actual_sheets != expected_sheets:
            add(
                findings,
                "ERROR",
                "WORKBOOK_SHEETS",
                "Worksheet names or order differ from the configured contract.",
                value={"expected": expected_sheets, "actual": actual_sheets},
            )
        configured_sheets = config.get("sheets", {})
        for sheet_name, sheet_config in configured_sheets.items():
            if sheet_name not in workbook.sheetnames:
                add(
                    findings,
                    "ERROR",
                    "WORKSHEET_MISSING",
                    f"Required worksheet is missing: {sheet_name}",
                )
                continue
            worksheet = workbook[sheet_name]
            headers, positions, region = validate_sheet_structure(
                worksheet,
                sheet_config,
                findings,
            )
            row_count = validate_sheet_data(
                worksheet,
                sheet_config,
                headers,
                positions,
                region,
                findings,
            )
            validate_data_validation(
                worksheet,
                sheet_config,
                positions,
                findings,
            )
            metrics["sheets"][sheet_name] = {
                "rows": row_count,
                "columns": len(headers),
                "tables": list(worksheet.tables),
                "data_validations": len(
                    worksheet.data_validations.dataValidation
                ),
            }
    finally:
        workbook.close()
    after_hash = file_hash(path)
    if after_hash != before_hash:
        add(
            findings,
            "ERROR",
            "READ_ONLY_VIOLATION",
            "Workbook bytes changed during validation.",
        )
    counts = Counter(finding.severity for finding in findings)
    count_map = {
        severity: counts.get(severity, 0)
        for severity in ("ERROR", "WARNING", "INFO")
    }
    categories = Counter(finding.category for finding in findings)
    category_map = {
        category: categories.get(category, 0)
        for category in ("CONFIGURATION", "WORKBOOK", "BUSINESS_RULE")
    }
    thresholds = {
        **DEFAULT_THRESHOLDS,
        **config.get("thresholds", {}),
    }
    return QAResult(
        workbook=str(path.resolve()),
        generated_at_utc=datetime.now(UTC).isoformat(),
        profile=str(config.get("profile_name", "Unnamed profile")),
        status=determine_status(count_map, thresholds),
        counts=count_map,
        category_counts=category_map,
        workbook_sha256=before_hash,
        findings=findings,
        metrics=metrics,
    )


def markdown_report(result: QAResult) -> str:
    """Render a detailed Markdown QA report."""
    lines = [
        "# Workbook QA Report",
        "",
        f"- Workbook: `{result.workbook}`",
        f"- Profile: {result.profile}",
        f"- Generated UTC: {result.generated_at_utc}",
        f"- SHA-256: `{result.workbook_sha256}`",
        f"- Final status: **{result.status}**",
        (
            "- Findings: "
            f"{result.counts['ERROR']} error(s), "
            f"{result.counts['WARNING']} warning(s), "
            f"{result.counts['INFO']} informational"
        ),
        "",
        "## Worksheet metrics",
        "",
        "| Worksheet | Rows | Columns | Tables | Validations |",
        "|---|---:|---:|---:|---:|",
    ]
    for sheet, metrics in result.metrics.get("sheets", {}).items():
        lines.append(
            f"| {sheet} | {metrics['rows']} | {metrics['columns']} | "
            f"{len(metrics['tables'])} | {metrics['data_validations']} |"
        )
    for category, heading in (
        ("CONFIGURATION", "Configuration issues"),
        ("WORKBOOK", "Workbook issues"),
        ("BUSINESS_RULE", "Business-rule issues"),
    ):
        lines.extend(["", f"## {heading}", ""])
        matching = [
            finding
            for finding in result.findings
            if finding.category == category
        ]
        if not matching:
            lines.append("None.")
            continue
        for finding in matching:
            location = " / ".join(
                part for part in (finding.sheet, finding.cell) if part
            )
            suffix = f" ({location})" if location else ""
            lines.append(
                f"- **{finding.severity} · {finding.code}**"
                f"{suffix}: {finding.message}"
            )
            if finding.value is not None:
                lines.append(
                    f"  - Value: `{json.dumps(finding.value, default=str)}`"
                )
            if finding.suggestion:
                lines.append(f"  - Suggested action: {finding.suggestion}")
    suggestions = sorted(
        {
            finding.suggestion
            for finding in result.findings
            if finding.suggestion
        }
    )
    lines.extend(["", "## Suggested improvements", ""])
    if suggestions:
        lines.extend(f"- {suggestion}" for suggestion in suggestions)
    else:
        lines.append("None.")
    lines.extend(["", f"## Final status: {result.status}", ""])
    return "\n".join(lines)


def print_summary(result: QAResult) -> None:
    """Print the concise console result."""
    print(f"Workbook: {result.workbook}")
    print(f"Profile: {result.profile}")
    print(f"Errors: {result.counts['ERROR']}")
    print(f"Warnings: {result.counts['WARNING']}")
    print(f"Info: {result.counts['INFO']}")
    print(
        "Configuration issues: "
        f"{result.category_counts['CONFIGURATION']}"
    )
    print(f"Workbook issues: {result.category_counts['WORKBOOK']}")
    print(
        "Business-rule issues: "
        f"{result.category_counts['BUSINESS_RULE']}"
    )
    print(f"Final status: {result.status}")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Read-only QA validation for Nocturnix Excel review workbooks."
    )
    parser.add_argument("workbook", type=Path, help="Workbook to validate")
    parser.add_argument(
        "--config",
        type=Path,
        required=True,
        help="JSON validation profile",
    )
    parser.add_argument(
        "--markdown",
        type=Path,
        help="Markdown report path (default: <workbook>.qa.md)",
    )
    parser.add_argument(
        "--json",
        dest="json_path",
        type=Path,
        help="JSON report path (default: <workbook>.qa.json)",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    """Run QA and write Markdown/JSON automation artifacts."""
    args = parse_args(argv)
    if not args.workbook.is_file():
        print(f"ERROR: Workbook not found: {args.workbook}", file=sys.stderr)
        return 2
    if not args.config.is_file():
        print(f"ERROR: QA profile not found: {args.config}", file=sys.stderr)
        return 2
    markdown_path = args.markdown or args.workbook.with_suffix(".qa.md")
    json_path = args.json_path or args.workbook.with_suffix(".qa.json")
    try:
        result = validate_workbook(args.workbook, load_config(args.config))
        markdown_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        markdown_path.write_text(markdown_report(result), encoding="utf-8")
        json_path.write_text(
            json.dumps(result.to_dict(), indent=2, default=str) + "\n",
            encoding="utf-8",
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    print_summary(result)
    print(f"Markdown report: {markdown_path}")
    print(f"JSON report: {json_path}")
    return 1 if result.status == "FAIL" else 0


if __name__ == "__main__":
    raise SystemExit(main())
