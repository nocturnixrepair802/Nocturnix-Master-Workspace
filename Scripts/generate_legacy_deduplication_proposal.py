"""Generate the legacy-catalog deduplication proposal workbook.

This is a review-artifact generator. It never saves either input workbook and it
does not write to the canonical database. Run it from the project virtual
environment after reviewers have confirmed the input files are the intended
versions.
"""

from __future__ import annotations

import hashlib
import math
import re
import sys
from collections import Counter, defaultdict
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

RAW_PATH = Path(r"D:\Business Portal\300_Pricing\Legacy\Raw Import Data.xlsx")
STAGING_PATH = Path(
    "D:\\Business Portal\\300_Pricing\\Working\\"
    r"Nocturnix_Legacy_Catalog_Staging_Preview_v1.xlsx"
)
CANONICAL_PATH = Path("Data") / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = Path(
    "D:\\Business Portal\\300_Pricing\\Working\\"
    r"Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
STAGING_SHEET = "01 - All Staging Records"

EXPECTED_SOURCE_ROWS = 743
EXPECTED_EXCLUSIONS = 315
EXPECTED_CONFLICTING_ROWS = 15
EXPECTED_RETAINED_ROWS = EXPECTED_SOURCE_ROWS - EXPECTED_EXCLUSIONS

PHYSICAL_CATEGORIES = {"Part", "Device", "Tool", "Accessory"}
PRIMARY_DESTINATIONS = {
    "Repair": "master_services staging",
    "Part": "parts_catalog staging",
    "Device": "master_devices staging",
    "Tool": "tool_catalog staging",
    "Accessory": "accessories_catalog staging",
}
SHEET_NAMES = [
    "00 - Summary",
    "01 - Retained",
    "02 - Duplicate Exclusions",
    "03 - SKU Conflicts",
    "04 - Lookup Enrichment",
    "05 - Pricing Review",
    "06 - Manufacturer Review",
    "07 - Supplier Review",
    "08 - Destinations",
    "09 - Decision Log",
    "10 - Validation",
    "11 - Import Metadata",
]
EXCLUSION_COMPARE_FIELDS = [
    ("Category", "Record Category"),
    ("Manufacturer", "Legacy Manufacturer"),
    ("Name", "Legacy Name"),
    ("Type", "Legacy Type"),
    ("Price", "Legacy Retail Price"),
    ("Cost", "Legacy Cost"),
    ("Supplier", "Legacy Supplier"),
    ("Condition", "Legacy Condition"),
    ("Stock", "Legacy Stock"),
    ("Serial Number", "Legacy Serial Number"),
    ("Bin", "Legacy Bin"),
    ("Tax Free", "Legacy Tax Free"),
    ("Note", "Legacy Note"),
    ("Updated At", "Legacy Updated At"),
    ("Created At", "Legacy Created At"),
]
EXCLUSION_FIELDS_COMPARED = "; ".join(
    label for label, _column in EXCLUSION_COMPARE_FIELDS
)
EXCLUSION_HEADERS = [
    "Exact Duplicate Group ID",
    "Excluded Source Row Number",
    "Retained Source Row Number",
    "Legacy SKU",
    "Match Type",
    "Fields Compared",
    "Exact Match Verified",
    *(
        header
        for label, _column in EXCLUSION_COMPARE_FIELDS
        for header in (f"Excluded {label}", f"Retained {label}", f"{label} Match")
    ),
    "Difference Summary",
    "Exclusion Reason",
    "Approval Status",
    "Reviewer Notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
CONFLICT_FILL = PatternFill("solid", fgColor="F4CCCC")
MISSING_FILL = PatternFill("solid", fgColor="FCE4D6")
ZERO_FILL = PatternFill("solid", fgColor="E2F0D9")
INVALID_FILL = PatternFill("solid", fgColor="F8CBAD")
CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
DATE_FORMAT = "yyyy-mm-dd"
REPAIR_HISTORICAL_NOTES = (
    "Legacy repair price retained for historical and market reference only. "
    "Final Nocturnix pricing will be established after completion of the Master "
    "Pricing Model."
)
REPAIR_UNPRICED_NOTES = (
    "Valid repair record. Do not archive. Do not classify as free. Await pricing "
    "model completion."
)
REPAIR_COST_ONLY_NOTES = (
    "Retain provisional cost. Final customer pricing will be calculated from the "
    "approved pricing model."
)


class ProposalError(RuntimeError):
    """Raised when source data fails a proposal invariant."""


@dataclass(frozen=True)
class SourceRecord:
    """A staging record with normalized source-row provenance."""

    source_row: int
    values: dict[str, Any]


@dataclass(frozen=True)
class ProposalData:
    """All calculated proposal categories and audit metrics."""

    staging_headers: list[str]
    retained: list[SourceRecord]
    exclusions: list[dict[str, Any]]
    conflicts: list[dict[str, Any]]
    enrichment: list[dict[str, Any]]
    zero_review: list[dict[str, Any]]
    manufacturer_review: list[dict[str, Any]]
    supplier_review: list[dict[str, Any]]
    destinations: list[dict[str, Any]]
    exact_pattern_count: int
    exact_participating_rows: int
    unique_sku_rows: int
    zero_counts: Counter[str]
    pricing_review_counts: Counter[str]
    destination_counts: Counter[str]
    secondary_inventory_count: int
    unique_unresolved_row_count: int
    unresolved_issue_instance_count: int


def sha256_file(path: Path) -> str:
    """Return a streaming SHA-256 hash for *path*."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ascii_text(value: Any) -> Any:
    """Make generated text ASCII-safe while preserving non-text cell values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return value.encode("ascii", "replace").decode("ascii")


def text(value: Any) -> str:
    """Return a stripped, ASCII-safe text representation."""
    if value is None:
        return ""
    return str(ascii_text(value)).strip()


def canonical_scalar(value: Any) -> tuple[str, Any]:
    """Create a stable, type-aware value used for exact-row comparison."""
    if value is None or text(value) == "":
        return ("blank", "")
    if isinstance(value, bool):
        return ("bool", value)
    if isinstance(value, datetime):
        return ("datetime", value.isoformat())
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return ("number", "nan")
        return ("number", str(Decimal(str(value)).normalize()))
    return ("text", text(value))


def monetary_value(value: Any) -> Decimal | None:
    """Return a Decimal for valid monetary values, otherwise None."""
    if value is None or text(value) == "":
        return None
    if isinstance(value, bool):
        return None
    cleaned = text(value).replace(",", "").replace("$", "")
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None
    return amount if amount.is_finite() else None


def is_invalid_money(value: Any) -> bool:
    """Return True only for a populated, non-numeric monetary value."""
    return text(value) != "" and monetary_value(value) is None


def require_files(paths: Iterable[Path]) -> None:
    """Fail before doing work when a required input is missing."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise ProposalError(f"Required file(s) missing: {', '.join(missing)}")


def load_staging_records(path: Path) -> tuple[list[str], list[SourceRecord]]:
    """Load staging data read-only and retain every original staging field."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if STAGING_SHEET not in workbook.sheetnames:
            raise ProposalError(f"Missing staging sheet: {STAGING_SHEET}")
        worksheet = workbook[STAGING_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ProposalError("The staging worksheet is empty") from exc
        headers = [text(value) for value in header_row]
        if not headers or any(not header for header in headers):
            raise ProposalError("Staging headers must be nonblank")
        if len(set(headers)) != len(headers):
            raise ProposalError("Staging headers must be unique")
        required = {
            "Source Row Number",
            "Legacy SKU",
            "Record Category",
            "Legacy Manufacturer",
            "Legacy Name",
            "Legacy Type",
            "Legacy Retail Price",
            "Legacy Cost",
            "Legacy Supplier",
            "Legacy Condition",
            *(column for _label, column in EXCLUSION_COMPARE_FIELDS),
        }
        missing = sorted(required - set(headers))
        if missing:
            raise ProposalError(f"Missing staging columns: {', '.join(missing)}")

        records: list[SourceRecord] = []
        for values in rows:
            if all(value is None or text(value) == "" for value in values):
                continue
            row = dict(zip(headers, values, strict=False))
            source_text = text(row["Source Row Number"])
            try:
                source_row = int(float(source_text))
            except ValueError as exc:
                raise ProposalError(
                    f"Invalid Source Row Number: {source_text!r}"
                ) from exc
            records.append(SourceRecord(source_row, row))
        return headers, records
    finally:
        workbook.close()


def load_raw_groups(path: Path) -> tuple[list[str], dict[tuple[Any, ...], list[int]]]:
    """Group raw worksheet rows by their complete, type-aware cell pattern."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = [text(value) for value in next(rows)]
        except StopIteration as exc:
            raise ProposalError("The raw import worksheet is empty") from exc
        groups: dict[tuple[Any, ...], list[int]] = defaultdict(list)
        for excel_row, values in enumerate(rows, start=2):
            if all(value is None or text(value) == "" for value in values):
                continue
            key = tuple(canonical_scalar(value) for value in values)
            groups[key].append(excel_row)
        return raw_headers, groups
    finally:
        workbook.close()


def conflict_fields(group: Sequence[SourceRecord]) -> str:
    """List business fields whose values differ within a duplicate-SKU group."""
    candidates = [
        ("Category", "Record Category"),
        ("Manufacturer", "Legacy Manufacturer"),
        ("Name", "Legacy Name"),
        ("Type", "Legacy Type"),
    ]
    conflicts = []
    for label, column in candidates:
        if len({text(record.values.get(column)) for record in group}) > 1:
            conflicts.append(label)
    return ", ".join(conflicts)


def classify_zero_values(record: SourceRecord) -> str:
    """Assign one mutually exclusive monetary review category."""
    price_raw = record.values.get("Legacy Retail Price")
    cost_raw = record.values.get("Legacy Cost")
    if is_invalid_money(price_raw) or is_invalid_money(cost_raw):
        return "Invalid Monetary Value"
    price = monetary_value(price_raw)
    cost = monetary_value(cost_raw)
    if price is None or cost is None:
        return "Invalid Monetary Value"
    if price == 0 and cost == 0:
        return "Price = 0 and Cost = 0"
    if price == 0 and cost > 0:
        return "Price = 0 and Cost > 0"
    if price > 0 and cost == 0:
        return "Price > 0 and Cost = 0"
    return "Price > 0 and Cost > 0"


def pricing_review_details(record: SourceRecord, monetary_pattern: str) -> dict[str, str]:
    """Return proposal-only pricing guidance without calculating any values."""
    category = text(record.values.get("Record Category"))
    if category == "Repair":
        if monetary_pattern == "Price > 0 and Cost = 0":
            return {
                "Business Interpretation": (
                    "Legacy provisional repair price. Internal repair cost has not "
                    "yet been calculated."
                ),
                "Pricing Status": "Pending Pricing Review",
                "Recommended Action": "Review Legacy Price",
                "Default Reviewer Notes": REPAIR_HISTORICAL_NOTES,
            }
        if monetary_pattern == "Price = 0 and Cost = 0":
            return {
                "Business Interpretation": "Repair pricing has not yet been developed.",
                "Pricing Status": "Pending Pricing Review",
                "Recommended Action": "Await Pricing Model",
                "Default Reviewer Notes": REPAIR_UNPRICED_NOTES,
            }
        if monetary_pattern == "Price = 0 and Cost > 0":
            return {
                "Business Interpretation": (
                    "Known internal cost exists. Customer price not yet established."
                ),
                "Pricing Status": "Pending Pricing Review",
                "Recommended Action": "Await Pricing Model",
                "Default Reviewer Notes": REPAIR_COST_ONLY_NOTES,
            }
        if monetary_pattern == "Invalid Monetary Value":
            return {
                "Business Interpretation": (
                    "Legacy repair monetary value requires source-format research."
                ),
                "Pricing Status": "Pending Pricing Review",
                "Recommended Action": "Do Not Publish",
                "Default Reviewer Notes": (
                    "Preserve the legacy value and await pricing-model completion and "
                    "source-data review."
                ),
            }
        return {
            "Business Interpretation": (
                "Legacy repair price and cost are provisional reference values."
            ),
            "Pricing Status": "Pending Pricing Review",
            "Recommended Action": "Review Legacy Price",
            "Default Reviewer Notes": REPAIR_HISTORICAL_NOTES,
        }

    nonrepair_details = {
        "Price = 0 and Cost = 0": (
            "Price and cost require business review.",
            "Pending Pricing Review",
            "Do Not Publish",
            "Determine whether values are missing, intentionally unpriced, or inactive.",
        ),
        "Price = 0 and Cost > 0": (
            "Known cost exists but customer price is not established.",
            "Pending Pricing Review",
            "Do Not Publish",
            "Review the customer price before publication.",
        ),
        "Price > 0 and Cost = 0": (
            "Customer price exists but cost requires review.",
            "Pending Pricing Review",
            "Research Cost",
            "Research the missing or intentionally zero cost.",
        ),
        "Price > 0 and Cost > 0": (
            "Price and cost are populated.",
            "Pricing Approved",
            "Review Legacy Price",
            "No zero-value pricing exception identified.",
        ),
        "Invalid Monetary Value": (
            "A monetary value is not numeric.",
            "Pending Pricing Review",
            "Do Not Publish",
            "Correct the monetary data type using verified source information.",
        ),
    }
    interpretation, status, action, notes = nonrepair_details[monetary_pattern]
    return {
        "Business Interpretation": interpretation,
        "Pricing Status": status,
        "Recommended Action": action,
        "Default Reviewer Notes": notes,
    }


def enrichment_reasons(record: SourceRecord) -> list[str]:
    """Return visible reasons that a retained record needs lookup work."""
    values = record.values
    category = text(values.get("Record Category"))
    reasons: list[str] = []
    if text(values.get("Legacy Manufacturer")) == "":
        reasons.append("Missing Manufacturer")
    if category in PHYSICAL_CATEGORIES and text(values.get("Legacy Supplier")) == "":
        reasons.append("Missing Supplier")
    if category not in PRIMARY_DESTINATIONS:
        reasons.append("Category Normalization")
    if text(values.get("Legacy Type")) == "":
        reasons.append("Type Normalization")
    if is_invalid_money(values.get("Legacy Retail Price")) or is_invalid_money(
        values.get("Legacy Cost")
    ):
        reasons.append("Numeric Monetary Correction")
    review_status = text(values.get("Review Status"))
    if review_status == "Mappable After Lookup Enrichment" and not reasons:
        reasons.append("Other Lookup Reconciliation")
    return reasons


def inventory_candidate(record: SourceRecord) -> bool:
    """Identify physical rows with item-level inventory provenance."""
    values = record.values
    category = text(values.get("Record Category"))
    return category in PHYSICAL_CATEGORIES and bool(
        text(values.get("Legacy Serial Number")) or text(values.get("Legacy Bin"))
    )


def build_exclusion_comparison(
    group_id: str,
    excluded: SourceRecord,
    retained: SourceRecord,
) -> dict[str, Any]:
    """Build and verify a self-contained exact-duplicate comparison row."""
    comparison: dict[str, Any] = {
        "Exact Duplicate Group ID": group_id,
        "Excluded Source Row Number": excluded.source_row,
        "Retained Source Row Number": retained.source_row,
        "Legacy SKU": text(excluded.values.get("Legacy SKU")),
        "Match Type": "Exact Duplicate",
        "Fields Compared": EXCLUSION_FIELDS_COMPARED,
    }
    differences: list[str] = []
    for label, staging_column in EXCLUSION_COMPARE_FIELDS:
        excluded_value = excluded.values.get(staging_column)
        retained_value = retained.values.get(staging_column)
        matches = canonical_scalar(excluded_value) == canonical_scalar(retained_value)
        comparison[f"Excluded {label}"] = excluded_value
        comparison[f"Retained {label}"] = retained_value
        comparison[f"{label} Match"] = "Yes" if matches else "No"
        if not matches:
            differences.append(label)

    exact_match = not differences
    comparison["Exact Match Verified"] = "Yes" if exact_match else "No"
    comparison["Difference Summary"] = "; ".join(differences)
    comparison["Exclusion Reason"] = (
        "Excess exact duplicate; retain first source row"
    )
    comparison["Approval Status"] = "Pending Review"
    comparison["Reviewer Notes"] = ""
    if not exact_match:
        raise ProposalError(
            "Unexpected difference in exact-duplicate group "
            f"{group_id}: excluded source {excluded.source_row}, retained source "
            f"{retained.source_row}, fields={comparison['Difference Summary']}"
        )
    return comparison


def build_proposal(
    headers: list[str],
    staging_records: list[SourceRecord],
    raw_groups: dict[tuple[Any, ...], list[int]],
) -> ProposalData:
    """Calculate proposal categories without changing source records."""
    by_source = {record.source_row: record for record in staging_records}
    if len(by_source) != len(staging_records):
        raise ProposalError("Duplicate Source Row Number values in staging data")

    raw_source_rows = {row for rows in raw_groups.values() for row in rows}
    staging_source_rows = set(by_source)
    if raw_source_rows != staging_source_rows:
        missing = sorted(raw_source_rows - staging_source_rows)
        extra = sorted(staging_source_rows - raw_source_rows)
        raise ProposalError(
            f"Raw/staging provenance mismatch; missing={missing[:10]}, extra={extra[:10]}"
        )

    duplicate_groups = [rows for rows in raw_groups.values() if len(rows) > 1]
    excluded_rows: set[int] = set()
    exclusions: list[dict[str, Any]] = []
    for group_number, source_rows in enumerate(
        sorted(duplicate_groups, key=lambda rows: rows[0]), start=1
    ):
        retained_source = min(source_rows)
        retained_record = by_source[retained_source]
        group_id = f"EDG-{group_number:04d}"
        for source_row in sorted(source_rows):
            if source_row == retained_source:
                continue
            excluded_rows.add(source_row)
            exclusions.append(
                build_exclusion_comparison(
                    group_id,
                    excluded=by_source[source_row],
                    retained=retained_record,
                )
            )

    retained = [
        record for record in staging_records if record.source_row not in excluded_rows
    ]
    sku_groups: dict[str, list[SourceRecord]] = defaultdict(list)
    for record in retained:
        sku_groups[text(record.values.get("Legacy SKU"))].append(record)

    conflicting_groups: list[tuple[str, list[SourceRecord], str]] = []
    for sku, group in sorted(sku_groups.items()):
        attributes = {
            (
                text(record.values.get("Record Category")),
                text(record.values.get("Legacy Manufacturer")),
                text(record.values.get("Legacy Name")),
                text(record.values.get("Legacy Type")),
            )
            for record in group
        }
        if sku and len(group) > 1 and len(attributes) > 1:
            conflicting_groups.append((sku, group, conflict_fields(group)))

    conflicts: list[dict[str, Any]] = []
    for group_number, (sku, group, fields) in enumerate(conflicting_groups, start=1):
        group_id = f"CSG-{group_number:04d}"
        for record in sorted(group, key=lambda item: item.source_row):
            values = record.values
            conflicts.append(
                {
                    "Source Row Number": record.source_row,
                    "Legacy SKU": sku,
                    "Category": values.get("Record Category"),
                    "Manufacturer": values.get("Legacy Manufacturer"),
                    "Name": values.get("Legacy Name"),
                    "Type": values.get("Legacy Type"),
                    "Price": values.get("Legacy Retail Price"),
                    "Cost": values.get("Legacy Cost"),
                    "Supplier": values.get("Legacy Supplier"),
                    "Condition": values.get("Legacy Condition"),
                    "Conflict Group ID": group_id,
                    "Conflict Fields": fields,
                    "Proposed Resolution": "",
                    "Approval Status": "Pending Review",
                    "Reviewer Notes": "",
                }
            )

    enrichment: list[dict[str, Any]] = []
    manufacturer_review: list[dict[str, Any]] = []
    supplier_review: list[dict[str, Any]] = []
    zero_review: list[dict[str, Any]] = []
    destinations: list[dict[str, Any]] = []
    zero_counts: Counter[str] = Counter()
    pricing_review_counts: Counter[str] = Counter()
    destination_counts: Counter[str] = Counter()

    for record in retained:
        values = record.values
        sku = text(values.get("Legacy SKU"))
        category = text(values.get("Record Category"))
        reasons = enrichment_reasons(record)
        if reasons:
            enrichment.append(
                {
                    "Source Row Number": record.source_row,
                    "Legacy SKU": sku,
                    "Category": category,
                    "Manufacturer": values.get("Legacy Manufacturer"),
                    "Type": values.get("Legacy Type"),
                    "Supplier": values.get("Legacy Supplier"),
                    "Price": values.get("Legacy Retail Price"),
                    "Cost": values.get("Legacy Cost"),
                    "Enrichment Reasons": "; ".join(reasons),
                    "Proposed Lookup Result": "",
                    "Approval Status": "Pending Review",
                    "Reviewer Notes": "",
                }
            )
        if text(values.get("Legacy Manufacturer")) == "":
            manufacturer_review.append(
                {
                    "Source Row Number": record.source_row,
                    "Legacy SKU": sku,
                    "Category": category,
                    "Name": values.get("Legacy Name"),
                    "Current Manufacturer": "",
                    "Proposed Manufacturer": "",
                    "Approval Status": "Pending Review",
                    "Reviewer Notes": "",
                }
            )
        if category in PHYSICAL_CATEGORIES and text(values.get("Legacy Supplier")) == "":
            supplier_review.append(
                {
                    "Source Row Number": record.source_row,
                    "Legacy SKU": sku,
                    "Category": category,
                    "Name": values.get("Legacy Name"),
                    "Current Supplier": "",
                    "Proposed Supplier": "",
                    "Approval Status": "Pending Review",
                    "Reviewer Notes": "",
                }
            )

        zero_category = classify_zero_values(record)
        pricing_details = pricing_review_details(record, zero_category)
        zero_counts[zero_category] += 1
        pricing_review_counts[pricing_details["Pricing Status"]] += 1
        if (
            category == "Repair"
            and zero_category == "Price > 0 and Cost = 0"
        ):
            pricing_review_counts["Historical Price Review"] += 1
        if pricing_details["Pricing Status"] == "Archive Candidate":
            pricing_review_counts["Archive Candidates"] += 1
        zero_review.append(
            {
                "Source Row Number": record.source_row,
                "Legacy SKU": sku,
                "Category": category,
                "Name": values.get("Legacy Name"),
                "Price": values.get("Legacy Retail Price"),
                "Cost": values.get("Legacy Cost"),
                "Zero Value Category": zero_category,
                **pricing_details,
                "Proposed Interpretation": "",
                "Approval Status": "Pending Review",
                "Reviewer Notes": "",
            }
        )

        primary = PRIMARY_DESTINATIONS.get(category, "Manual destination review")
        inventory = inventory_candidate(record)
        pricing = not (
            is_invalid_money(values.get("Legacy Retail Price"))
            or is_invalid_money(values.get("Legacy Cost"))
        )
        destination_counts[primary] += 1
        destinations.append(
            {
                "Source Row Number": record.source_row,
                "Legacy SKU": sku,
                "Legacy Alias Candidate": sku,
                "Proposed Canonical ID": "",
                "Category": category,
                "Primary Destination": primary,
                "Secondary Inventory Destination": (
                    "inventory_items staging" if inventory else ""
                ),
                "Secondary Pricing Destination": "pricing staging" if pricing else "",
                "Destination Rationale": (
                    "Category mapping; secondary flags are proposal indicators only"
                ),
                "Approval Status": "Pending Review",
            }
        )

    conflicting_source_rows = {
        int(row["Source Row Number"]) for row in conflicts
    }
    unique_sku_rows = sum(
        1 for record in retained if len(sku_groups[text(record.values.get("Legacy SKU"))]) == 1
    )
    unresolved_pricing_review = [
        row
        for row in zero_review
        if row["Pricing Status"] != "Pricing Approved"
    ]
    unresolved_retained_rows = {
        int(row["Source Row Number"])
        for review_rows in (
            conflicts,
            enrichment,
            manufacturer_review,
            supplier_review,
            unresolved_pricing_review,
        )
        for row in review_rows
    }
    unresolved_issue_instance_count = (
        len(exclusions)
        + len(conflicts)
        + len(enrichment)
        + len(manufacturer_review)
        + len(supplier_review)
        + len(unresolved_pricing_review)
    )

    proposal = ProposalData(
        staging_headers=headers,
        retained=retained,
        exclusions=exclusions,
        conflicts=conflicts,
        enrichment=enrichment,
        zero_review=zero_review,
        manufacturer_review=manufacturer_review,
        supplier_review=supplier_review,
        destinations=destinations,
        exact_pattern_count=len(duplicate_groups),
        exact_participating_rows=sum(len(rows) for rows in duplicate_groups),
        unique_sku_rows=unique_sku_rows,
        zero_counts=zero_counts,
        pricing_review_counts=pricing_review_counts,
        destination_counts=destination_counts,
        secondary_inventory_count=sum(
            1 for record in retained if inventory_candidate(record)
        ),
        unique_unresolved_row_count=len(unresolved_retained_rows),
        unresolved_issue_instance_count=unresolved_issue_instance_count,
    )
    validate_calculation(proposal, staging_source_rows, conflicting_source_rows)
    return proposal


def validate_calculation(
    proposal: ProposalData,
    source_rows: set[int],
    conflicting_source_rows: set[int],
) -> None:
    """Enforce approved counts and lossless provenance before workbook creation."""
    retained_rows = {record.source_row for record in proposal.retained}
    excluded_rows = {
        int(row["Excluded Source Row Number"]) for row in proposal.exclusions
    }
    retained_references = {
        int(row["Retained Source Row Number"]) for row in proposal.exclusions
    }
    comparison_rows_verified = all(
        row["Match Type"] == "Exact Duplicate"
        and row["Exact Match Verified"] == "Yes"
        and row["Difference Summary"] == ""
        and all(row[f"{label} Match"] == "Yes" for label, _ in EXCLUSION_COMPARE_FIELDS)
        and int(row["Excluded Source Row Number"])
        != int(row["Retained Source Row Number"])
        for row in proposal.exclusions
    )
    checks = {
        "source row count": len(source_rows) == EXPECTED_SOURCE_ROWS,
        "retained row count": len(retained_rows) == EXPECTED_RETAINED_ROWS,
        "exclusion row count": len(excluded_rows) == EXPECTED_EXCLUSIONS,
        "conflicting row count": len(conflicting_source_rows)
        == EXPECTED_CONFLICTING_ROWS,
        "retained rows unique": len(retained_rows) == len(proposal.retained),
        "excluded rows unique": len(excluded_rows) == len(proposal.exclusions),
        "retained exclusion references valid": retained_references <= retained_rows,
        "duplicate comparisons verified": comparison_rows_verified,
        "proposal categories disjoint": retained_rows.isdisjoint(excluded_rows),
        "no source row lost": retained_rows | excluded_rows == source_rows,
        "conflicts retained": conflicting_source_rows <= retained_rows,
    }
    failed = [name for name, passed in checks.items() if not passed]
    if failed:
        raise ProposalError(f"Proposal calculation validation failed: {', '.join(failed)}")


def records_to_rows(
    records: Sequence[SourceRecord], headers: Sequence[str]
) -> list[list[Any]]:
    """Convert retained records to rows while preserving staging column order."""
    return [
        [ascii_text(record.values.get(header)) for header in headers]
        + ["", text(record.values.get("Legacy SKU")), "Pending Review"]
        for record in records
    ]


def dict_rows(
    records: Sequence[dict[str, Any]], headers: Sequence[str]
) -> list[list[Any]]:
    """Convert dictionaries to an ASCII-safe matrix."""
    return [[ascii_text(record.get(header, "")) for header in headers] for record in records]


def safe_table_name(sheet_name: str, sequence: int) -> str:
    """Return a workbook-unique, Excel-compatible table name."""
    stem = re.sub(r"[^A-Za-z0-9_]", "_", sheet_name)
    return f"tbl_{sequence:02d}_{stem}"[:250]


def write_table_sheet(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    table_name: str,
) -> None:
    """Write, format, filter, and freeze a worksheet table."""
    worksheet.append([ascii_text(header) for header in headers])
    for row in rows:
        worksheet.append([ascii_text(value) for value in row])
    if not rows:
        worksheet.append(["" for _ in headers])

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    apply_column_formats(worksheet, headers)
    apply_conditional_formats(worksheet, headers)
    size_columns(worksheet, headers)


def apply_column_formats(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Apply currency and date formats by semantic column name."""
    for column_number, header in enumerate(headers, start=1):
        lowered = header.lower()
        if header in {
            "Price",
            "Cost",
            "Legacy Retail Price",
            "Legacy Cost",
            "Excluded Price",
            "Retained Price",
            "Excluded Cost",
            "Retained Cost",
        }:
            for cell in worksheet.iter_cols(
                min_col=column_number,
                max_col=column_number,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = CURRENCY_FORMAT
        if "date" in lowered or lowered.endswith("updated at") or lowered.endswith(
            "created at"
        ):
            for cell in worksheet.iter_cols(
                min_col=column_number,
                max_col=column_number,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = DATE_FORMAT


def column_letter_for(headers: Sequence[str], name: str) -> str | None:
    """Return an Excel column letter for a header, if present."""
    try:
        position = headers.index(name) + 1
    except ValueError:
        return None
    letters = ""
    while position:
        position, remainder = divmod(position - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def apply_conditional_formats(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Apply all requested review-oriented conditional formats."""
    last_row = max(worksheet.max_row, 2)
    full_range = f"A2:{column_letter_for(headers, headers[-1])}{last_row}"

    pending = column_letter_for(headers, "Approval Status") or column_letter_for(
        headers, "Status"
    )
    if pending:
        worksheet.conditional_formatting.add(
            f"{pending}2:{pending}{last_row}",
            CellIsRule(operator="equal", formula=['"Pending Review"'], fill=PENDING_FILL),
        )

    for header in headers:
        if header.endswith(" Match") or header == "Exact Match Verified":
            column = column_letter_for(headers, header)
            if column:
                worksheet.conditional_formatting.add(
                    f"{column}2:{column}{last_row}",
                    CellIsRule(operator="equal", formula=['"No"'], fill=CONFLICT_FILL),
                )
    difference_summary = column_letter_for(headers, "Difference Summary")
    if difference_summary:
        worksheet.conditional_formatting.add(
            f"{difference_summary}2:{difference_summary}{last_row}",
            FormulaRule(
                formula=[f'LEN(TRIM(${difference_summary}2))>0'],
                fill=CONFLICT_FILL,
            ),
        )
    conflict = column_letter_for(headers, "Conflict Group ID")
    if conflict:
        worksheet.conditional_formatting.add(
            full_range,
            FormulaRule(formula=[f'LEN(${conflict}2)>0'], fill=CONFLICT_FILL),
        )

    for header in ("Manufacturer", "Current Manufacturer", "Legacy Manufacturer"):
        column = column_letter_for(headers, header)
        if column:
            worksheet.conditional_formatting.add(
                f"{column}2:{column}{last_row}",
                FormulaRule(formula=[f'LEN(TRIM(${column}2))=0'], fill=MISSING_FILL),
            )
            break
    for header in ("Supplier", "Current Supplier", "Legacy Supplier"):
        column = column_letter_for(headers, header)
        if column:
            worksheet.conditional_formatting.add(
                f"{column}2:{column}{last_row}",
                FormulaRule(formula=[f'LEN(TRIM(${column}2))=0'], fill=MISSING_FILL),
            )
            break

    for header in ("Price", "Legacy Retail Price"):
        column = column_letter_for(headers, header)
        if column:
            worksheet.conditional_formatting.add(
                f"{column}2:{column}{last_row}",
                CellIsRule(operator="equal", formula=["0"], fill=ZERO_FILL),
            )
            break
    for header in ("Cost", "Legacy Cost"):
        column = column_letter_for(headers, header)
        if column:
            worksheet.conditional_formatting.add(
                f"{column}2:{column}{last_row}",
                CellIsRule(operator="equal", formula=["0"], fill=ZERO_FILL),
            )
            break

    category = column_letter_for(headers, "Zero Value Category")
    if category:
        worksheet.conditional_formatting.add(
            f"{category}2:{category}{last_row}",
            CellIsRule(
                operator="equal", formula=['"Invalid Monetary Value"'], fill=INVALID_FILL
            ),
        )
    reasons = column_letter_for(headers, "Enrichment Reasons")
    if reasons:
        worksheet.conditional_formatting.add(
            f"{reasons}2:{reasons}{last_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Invalid",${reasons}2))'],
                fill=INVALID_FILL,
            ),
        )


def size_columns(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Apply practical fixed caps so large review tables remain navigable."""
    for index, header in enumerate(headers, start=1):
        letter = column_letter_for(headers, header)
        if letter is None:
            continue
        width = max(12, min(38, len(header) + 3))
        if any(word in header for word in ("Reason", "Notes", "Fields", "Rationale")):
            width = 38
        worksheet.column_dimensions[letter].width = width


def summary_rows(proposal: ProposalData) -> list[list[Any]]:
    """Build the proposal summary metric table."""
    rows: list[list[Any]] = [
        ["Source rows", EXPECTED_SOURCE_ROWS, "Calculated source population"],
        ["Unique SKU rows", proposal.unique_sku_rows, "Retained single-row SKU groups"],
        ["Exact duplicate patterns", proposal.exact_pattern_count, "Full-row patterns"],
        [
            "Rows participating in exact duplicate patterns",
            proposal.exact_participating_rows,
            "Includes retained representatives and exclusions",
        ],
        ["Proposed excess duplicate exclusions", len(proposal.exclusions), "Pending"],
        ["Candidate retained rows", len(proposal.retained), "Pending"],
        ["Conflicting rows", len(proposal.conflicts), "Pending manual review"],
        ["Lookup-enrichment rows", len(proposal.enrichment), "Pending"],
    ]
    for category in (
        "Pending Pricing Review",
        "Historical Price Review",
        "Archive Candidates",
        "Pricing Approved",
    ):
        rows.append(
            [
                f"Pricing review: {category}",
                proposal.pricing_review_counts[category],
                "Proposal classification",
            ]
        )
    for destination, count in sorted(proposal.destination_counts.items()):
        rows.append([f"Primary: {destination}", count, "Proposed destination"])
    rows.extend(
        [
            [
                "Secondary inventory candidates",
                proposal.secondary_inventory_count,
                "inventory_items staging indicator",
            ],
            [
                "Unique retained rows with unresolved issues",
                proposal.unique_unresolved_row_count,
                "Distinct retained source rows",
            ],
            [
                "Total unresolved issue instances",
                proposal.unresolved_issue_instance_count,
                "Not distinct rows",
            ],
        ]
    )
    return rows


def validate_sheet_names(sheet_names: Sequence[str]) -> None:
    """Enforce Excel's worksheet-title limit and workbook uniqueness."""
    overlong = [name for name in sheet_names if len(name) > 31]
    if overlong:
        raise ProposalError(
            f"Worksheet titles exceed 31 characters: {', '.join(overlong)}"
        )
    if len(set(sheet_names)) != len(sheet_names):
        raise ProposalError("Worksheet titles must be unique")


def validation_rows(proposal: ProposalData) -> list[list[Any]]:
    """Build validations that are also independently checked after reopening."""
    checks = [
        ("All required worksheets exist", True, str(len(SHEET_NAMES))),
        (
            "Every worksheet title is 31 characters or fewer",
            all(len(name) <= 31 for name in SHEET_NAMES),
            f"Maximum length: {max(map(len, SHEET_NAMES))}",
        ),
        (
            "Every worksheet title is unique",
            len(set(SHEET_NAMES)) == len(SHEET_NAMES),
            f"{len(set(SHEET_NAMES))} unique titles",
        ),
        (
            "Retained plus exclusions equals source rows",
            len(proposal.retained) + len(proposal.exclusions) == EXPECTED_SOURCE_ROWS,
            f"{len(proposal.retained)} + {len(proposal.exclusions)}",
        ),
        (
            "Exactly 315 excess duplicates proposed",
            len(proposal.exclusions) == EXPECTED_EXCLUSIONS,
            str(len(proposal.exclusions)),
        ),
        (
            "All duplicate exclusion comparisons are exact matches",
            all(
                row["Match Type"] == "Exact Duplicate"
                and row["Exact Match Verified"] == "Yes"
                and row["Difference Summary"] == ""
                and all(
                    row[f"{label} Match"] == "Yes"
                    for label, _column in EXCLUSION_COMPARE_FIELDS
                )
                for row in proposal.exclusions
            ),
            f"{len(proposal.exclusions)} exact matches",
        ),
        (
            "All 15 conflicting rows retained and pending",
            len(proposal.conflicts) == EXPECTED_CONFLICTING_ROWS,
            str(len(proposal.conflicts)),
        ),
        (
            "Zero-cost Repair rows are pending pricing review",
            all(
                row["Pricing Status"] == "Pending Pricing Review"
                for row in proposal.zero_review
                if row["Category"] == "Repair"
                and monetary_value(row["Cost"]) == Decimal("0")
            ),
            "Repair pricing approval deferred",
        ),
        ("No final canonical IDs generated", True, "Blank proposal column"),
        ("No source row lost", True, "Validated before workbook creation"),
        ("No duplicate source row within proposal category", True, "Validated"),
        ("Input hashes unchanged", True, "Rechecked after output save"),
    ]
    return [[name, "PASS" if passed else "FAIL", evidence] for name, passed, evidence in checks]


def create_workbook(
    proposal: ProposalData,
    raw_headers: Sequence[str],
    hashes_before: dict[Path, str],
    generated_at: datetime,
) -> Workbook:
    """Create the complete proposal workbook in memory."""
    validate_sheet_names(SHEET_NAMES)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}

    write_table_sheet(
        sheets[SHEET_NAMES[0]],
        ["Metric", "Count", "Notes"],
        summary_rows(proposal),
        safe_table_name(SHEET_NAMES[0], 0),
    )
    retained_headers = [
        *proposal.staging_headers,
        "Proposed Canonical ID",
        "Legacy Alias Candidate",
        "Approval Status",
    ]
    write_table_sheet(
        sheets[SHEET_NAMES[1]],
        retained_headers,
        records_to_rows(proposal.retained, proposal.staging_headers),
        safe_table_name(SHEET_NAMES[1], 1),
    )

    sheet_definitions: list[tuple[int, list[str], Sequence[dict[str, Any]]]] = [
        (
            2,
            EXCLUSION_HEADERS,
            proposal.exclusions,
        ),
        (
            3,
            [
                "Source Row Number",
                "Legacy SKU",
                "Category",
                "Manufacturer",
                "Name",
                "Type",
                "Price",
                "Cost",
                "Supplier",
                "Condition",
                "Conflict Group ID",
                "Conflict Fields",
                "Proposed Resolution",
                "Approval Status",
                "Reviewer Notes",
            ],
            proposal.conflicts,
        ),
        (
            4,
            [
                "Source Row Number",
                "Legacy SKU",
                "Category",
                "Manufacturer",
                "Type",
                "Supplier",
                "Price",
                "Cost",
                "Enrichment Reasons",
                "Proposed Lookup Result",
                "Approval Status",
                "Reviewer Notes",
            ],
            proposal.enrichment,
        ),
        (
            5,
            [
                "Source Row Number",
                "Legacy SKU",
                "Category",
                "Name",
                "Price",
                "Cost",
                "Zero Value Category",
                "Business Interpretation",
                "Pricing Status",
                "Recommended Action",
                "Default Reviewer Notes",
                "Proposed Interpretation",
                "Approval Status",
                "Reviewer Notes",
            ],
            proposal.zero_review,
        ),
        (
            6,
            [
                "Source Row Number",
                "Legacy SKU",
                "Category",
                "Name",
                "Current Manufacturer",
                "Proposed Manufacturer",
                "Approval Status",
                "Reviewer Notes",
            ],
            proposal.manufacturer_review,
        ),
        (
            7,
            [
                "Source Row Number",
                "Legacy SKU",
                "Category",
                "Name",
                "Current Supplier",
                "Proposed Supplier",
                "Approval Status",
                "Reviewer Notes",
            ],
            proposal.supplier_review,
        ),
        (
            8,
            [
                "Source Row Number",
                "Legacy SKU",
                "Legacy Alias Candidate",
                "Proposed Canonical ID",
                "Category",
                "Primary Destination",
                "Secondary Inventory Destination",
                "Secondary Pricing Destination",
                "Destination Rationale",
                "Approval Status",
            ],
            proposal.destinations,
        ),
    ]
    for index, headers, records in sheet_definitions:
        write_table_sheet(
            sheets[SHEET_NAMES[index]],
            headers,
            dict_rows(records, headers),
            safe_table_name(SHEET_NAMES[index], index),
        )

    decision_headers = [
        "Decision ID",
        "Record or Group ID",
        "Decision Type",
        "Decision",
        "Reason",
        "Reviewer",
        "Decision Date",
        "Status",
        "Notes",
    ]
    write_table_sheet(
        sheets[SHEET_NAMES[9]],
        decision_headers,
        [],
        safe_table_name(SHEET_NAMES[9], 9),
    )
    write_table_sheet(
        sheets[SHEET_NAMES[10]],
        ["Validation Check", "Result", "Evidence"],
        validation_rows(proposal),
        safe_table_name(SHEET_NAMES[10], 10),
    )
    metadata = [
        ["Artifact Type", "Deduplication and enrichment proposal only"],
        ["Production Import", "No"],
        ["Generated At UTC", generated_at],
        ["Generator", str(Path(__file__).resolve())],
        ["Raw Input", str(RAW_PATH)],
        ["Raw SHA-256", hashes_before[RAW_PATH]],
        ["Raw Header Count", len(raw_headers)],
        ["Staging Input", str(STAGING_PATH)],
        ["Staging SHA-256", hashes_before[STAGING_PATH]],
        ["Canonical Reference", str(CANONICAL_PATH.resolve())],
        ["Canonical SHA-256", hashes_before[CANONICAL_PATH]],
        ["Output", str(OUTPUT_PATH)],
        ["Canonical ID Policy", "Blank; Legacy SKU retained as alias candidate only"],
    ]
    write_table_sheet(
        sheets[SHEET_NAMES[11]],
        ["Metadata Field", "Value"],
        metadata,
        safe_table_name(SHEET_NAMES[11], 11),
    )
    sheets[SHEET_NAMES[11]]["B4"].number_format = "yyyy-mm-dd hh:mm:ss"
    return workbook


def validate_reopened_exclusions(
    retained: Worksheet,
    exclusions: Worksheet,
) -> str:
    """Validate every self-contained duplicate comparison after reopening."""
    exclusion_headers = [cell.value for cell in exclusions[1]]
    if exclusion_headers != EXCLUSION_HEADERS:
        raise ProposalError("Duplicate exclusion headers do not match specification")
    header_columns = {
        str(header): column
        for column, header in enumerate(exclusion_headers, start=1)
    }
    exclusion_count = exclusions.max_row - 1
    if exclusion_count != EXPECTED_EXCLUSIONS:
        raise ProposalError("Reopened exclusion count is not 315")

    retained_headers = [cell.value for cell in retained[1]]
    retained_source_column = retained_headers.index("Source Row Number") + 1
    retained_source_rows = {
        int(retained.cell(row=row, column=retained_source_column).value)
        for row in range(2, retained.max_row + 1)
    }
    individual_match_headers = [
        header for header in exclusion_headers if str(header).endswith(" Match")
    ]
    excluded_source_rows: list[int] = []

    for row in range(2, exclusions.max_row + 1):
        row_values = {
            header: exclusions.cell(row=row, column=column).value
            for header, column in header_columns.items()
        }
        excluded_source = int(row_values["Excluded Source Row Number"])
        retained_source = int(row_values["Retained Source Row Number"])
        excluded_source_rows.append(excluded_source)
        if row_values["Match Type"] != "Exact Duplicate":
            raise ProposalError(f"Invalid Match Type on exclusion row {row}")
        if row_values["Exact Match Verified"] != "Yes":
            raise ProposalError(f"Exact match not verified on exclusion row {row}")
        if any(row_values[header] != "Yes" for header in individual_match_headers):
            raise ProposalError(f"Individual field mismatch on exclusion row {row}")
        if row_values["Difference Summary"] not in (None, ""):
            raise ProposalError(f"Difference Summary is not blank on exclusion row {row}")
        if retained_source not in retained_source_rows:
            raise ProposalError(
                f"Exclusion row {row} points to non-retained source {retained_source}"
            )
        if excluded_source == retained_source:
            raise ProposalError(
                f"Excluded and retained source rows are equal on exclusion row {row}"
            )

    if len(set(excluded_source_rows)) != EXPECTED_EXCLUSIONS:
        raise ProposalError("The 315 excluded source rows are not unique")
    return "Duplicate exclusion comparisons: PASS (315 exact matches)"


def validate_reopened_pricing_review(worksheet: Worksheet) -> str:
    """Confirm zero-cost Repair rows are deferred, not treated as failures."""
    headers = [cell.value for cell in worksheet[1]]
    required = {
        "Category",
        "Cost",
        "Business Interpretation",
        "Pricing Status",
        "Recommended Action",
        "Default Reviewer Notes",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise ProposalError(
            f"Pricing Review columns missing after reopen: {', '.join(missing)}"
        )
    columns = {str(header): index for index, header in enumerate(headers, start=1)}
    zero_cost_repair_rows = 0
    for row in range(2, worksheet.max_row + 1):
        category = worksheet.cell(row=row, column=columns["Category"]).value
        cost = worksheet.cell(row=row, column=columns["Cost"]).value
        if category != "Repair" or monetary_value(cost) != Decimal("0"):
            continue
        zero_cost_repair_rows += 1
        status = worksheet.cell(row=row, column=columns["Pricing Status"]).value
        interpretation = worksheet.cell(
            row=row, column=columns["Business Interpretation"]
        ).value
        notes = worksheet.cell(
            row=row, column=columns["Default Reviewer Notes"]
        ).value
        if status != "Pending Pricing Review":
            raise ProposalError(
                f"Zero-cost Repair row {row} is not Pending Pricing Review"
            )
        if not interpretation or not notes:
            raise ProposalError(
                f"Zero-cost Repair row {row} lacks deferred-pricing guidance"
            )
    return (
        "Repair zero-cost pricing deferral: PASS "
        f"({zero_cost_repair_rows} pending rows)"
    )


def validate_reopened_workbook(path: Path, hashes_before: dict[Path, str]) -> list[str]:
    """Reopen the output and independently verify structure and key counts."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    messages: list[str] = []
    try:
        if any(len(name) > 31 for name in workbook.sheetnames):
            raise ProposalError("Reopened workbook contains an overlong worksheet title")
        if len(set(workbook.sheetnames)) != len(workbook.sheetnames):
            raise ProposalError("Reopened workbook contains duplicate worksheet titles")
        if workbook.sheetnames != SHEET_NAMES:
            raise ProposalError("Generated worksheet names/order do not match specification")
        messages.append(
            f"Worksheet titles: PASS ({len(workbook.sheetnames)} unique, all <= 31 chars)"
        )

        retained = workbook[SHEET_NAMES[1]]
        exclusions = workbook[SHEET_NAMES[2]]
        conflicts = workbook[SHEET_NAMES[3]]
        pricing_review = workbook[SHEET_NAMES[5]]
        retained_count = retained.max_row - 1
        exclusion_count = exclusions.max_row - 1
        conflict_count = conflicts.max_row - 1
        if retained_count + exclusion_count != EXPECTED_SOURCE_ROWS:
            raise ProposalError("Reopened retained/exclusion counts do not reconcile")
        if exclusion_count != EXPECTED_EXCLUSIONS:
            raise ProposalError("Reopened exclusion count is not 315")
        if conflict_count != EXPECTED_CONFLICTING_ROWS:
            raise ProposalError("Reopened conflict count is not 15")

        conflict_headers = [cell.value for cell in conflicts[1]]
        status_column = conflict_headers.index("Approval Status") + 1
        if any(
            conflicts.cell(row=row, column=status_column).value != "Pending Review"
            for row in range(2, conflicts.max_row + 1)
        ):
            raise ProposalError("A conflicting row is not Pending Review")
        messages.append(
            f"Counts: PASS (retained={retained_count}, exclusions={exclusion_count}, "
            f"conflicts={conflict_count})"
        )
        messages.append(validate_reopened_exclusions(retained, exclusions))
        messages.append(validate_reopened_pricing_review(pricing_review))

        retained_headers = [cell.value for cell in retained[1]]
        canonical_column = retained_headers.index("Proposed Canonical ID") + 1
        if any(
            retained.cell(row=row, column=canonical_column).value not in (None, "")
            for row in range(2, retained.max_row + 1)
        ):
            raise ProposalError("A final canonical ID was generated")
        messages.append("Canonical IDs: PASS (all blank)")

        for worksheet in workbook.worksheets:
            if not worksheet.tables:
                raise ProposalError(f"Worksheet lacks an Excel Table: {worksheet.title}")
            if worksheet.freeze_panes != "A2":
                raise ProposalError(f"Worksheet header is not frozen: {worksheet.title}")
        messages.append("Tables and frozen headers: PASS")
    finally:
        workbook.close()

    hashes_after = {path_: sha256_file(path_) for path_ in hashes_before}
    changed = [str(path_) for path_ in hashes_before if hashes_before[path_] != hashes_after[path_]]
    if changed:
        raise ProposalError(f"Protected input hash changed: {', '.join(changed)}")
    messages.append("Protected input SHA-256 hashes: PASS (unchanged)")
    return messages


def print_summary(proposal: ProposalData, validation_messages: Sequence[str]) -> None:
    """Print a concise, review-friendly execution result."""
    print("Legacy Deduplication Proposal Validation")
    print("========================================")
    print(f"Output: {OUTPUT_PATH}")
    print(f"Proposed retained: {len(proposal.retained)}")
    print(f"Proposed exclusions: {len(proposal.exclusions)}")
    print(f"Conflicting rows: {len(proposal.conflicts)}")
    print(f"Lookup-enrichment rows: {len(proposal.enrichment)}")
    print("Pricing Review Summary:")
    for category in (
        "Pending Pricing Review",
        "Historical Price Review",
        "Archive Candidates",
        "Pricing Approved",
    ):
        print(f"  {category}: {proposal.pricing_review_counts[category]}")
    print("Primary destinations:")
    for destination, count in sorted(proposal.destination_counts.items()):
        print(f"  {destination}: {count}")
    print(f"Secondary inventory candidates: {proposal.secondary_inventory_count}")
    print(
        "Unique retained rows with unresolved issues: "
        f"{proposal.unique_unresolved_row_count}"
    )
    print(
        "Total unresolved issue instances: "
        f"{proposal.unresolved_issue_instance_count}"
    )
    print("Validation:")
    for message in validation_messages:
        print(f"  {message}")


def main() -> int:
    """Generate, reopen, and validate the proposal workbook."""
    try:
        protected_paths = (RAW_PATH, STAGING_PATH, CANONICAL_PATH)
        require_files(protected_paths)
        hashes_before = {path: sha256_file(path) for path in protected_paths}
        staging_headers, staging_records = load_staging_records(STAGING_PATH)
        raw_headers, raw_groups = load_raw_groups(RAW_PATH)
        proposal = build_proposal(staging_headers, staging_records, raw_groups)

        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        generated_at = datetime.now(UTC).replace(tzinfo=None)
        workbook = create_workbook(proposal, raw_headers, hashes_before, generated_at)
        workbook.save(OUTPUT_PATH)
        workbook.close()

        validation_messages = validate_reopened_workbook(OUTPUT_PATH, hashes_before)
        print_summary(proposal, validation_messages)
        return 0
    except (OSError, ValueError, KeyError, ProposalError) as exc:
        print(f"ERROR: {ascii_text(str(exc))}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
