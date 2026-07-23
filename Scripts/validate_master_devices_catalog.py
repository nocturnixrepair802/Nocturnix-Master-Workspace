"""Independently validate the Nocturnix Master Devices Catalog v1 workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
PROPOSAL_PATH = WORKING_DIR / "Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
MASTER_SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
MASTER_PARTS_PATH = WORKING_DIR / "Nocturnix_Master_Parts_Catalog_v1.xlsx"
MASTER_PRICING_PATH = WORKING_DIR / "Nocturnix_Master_Pricing_Catalog_v1.xlsx"
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = WORKING_DIR / "Nocturnix_Master_Devices_Catalog_v1.xlsx"
PROTECTED_PATHS = (
    PROPOSAL_PATH,
    MASTER_SERVICES_PATH,
    MASTER_PARTS_PATH,
    MASTER_PRICING_PATH,
    CANONICAL_PATH,
)

EXPECTED_DEVICE_ROWS = 46
IMPORT_BATCH_ID = "MASTER-DEVICES-V1-REVIEW"
DEVICE_ID_PATTERN = re.compile(r"^DEV\d{6}$")
SHEET_NAMES = [
    "00 - Instructions",
    "01 - Master Devices",
    "02 - Manufacturers",
    "03 - Device Families",
    "04 - Product Lines",
    "05 - Device Series",
    "06 - Form Factors",
    "07 - Operating Systems",
    "08 - Network Types",
    "09 - Device Statuses",
    "10 - Identity Review",
    "11 - Relationship Review",
    "12 - Validation Summary",
    "13 - Revision History",
    "14 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterDevicesInstructions",
    "01 - Master Devices": "tblMasterDevicesCatalog",
    "02 - Manufacturers": "tblDeviceManufacturers",
    "03 - Device Families": "tblDeviceFamilies",
    "04 - Product Lines": "tblProductLines",
    "05 - Device Series": "tblDeviceSeries",
    "06 - Form Factors": "tblFormFactors",
    "07 - Operating Systems": "tblOperatingSystems",
    "08 - Network Types": "tblNetworkTypes",
    "09 - Device Statuses": "tblDeviceStatuses",
    "10 - Identity Review": "tblDeviceIdentityReview",
    "11 - Relationship Review": "tblDeviceRelationshipReview",
    "12 - Validation Summary": "tblMasterDevicesValidation",
    "13 - Revision History": "tblMasterDevicesRevisionHistory",
    "14 - Import Metadata": "tblMasterDevicesImportMetadata",
}
DEVICE_HEADERS = [
    "Device ID",
    "Legacy Device SKU",
    "Active",
    "Status",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Product Line",
    "Device Series",
    "Device Name",
    "Device Display Name",
    "Device Description",
    "Model Number",
    "Variant",
    "Generation",
    "Release Year",
    "Form Factor",
    "Operating System Family",
    "Network Type",
    "Storage Capacity",
    "Memory Capacity",
    "Color",
    "Region",
    "Carrier",
    "Wi-Fi Only",
    "Cellular Capable",
    "Repair Supported",
    "Parts Supported",
    "Mail-In Eligible",
    "Mobile Service Eligible",
    "Compatibility Status",
    "Service Mapping Status",
    "Parts Mapping Status",
    "Review Status",
    "Legacy Retail Price",
    "Legacy Cost",
    "Currency",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Reviewer",
    "Reviewer Notes",
    "Effective Date",
    "Last Reviewed",
    "Created At",
    "Updated At",
]
IDENTITY_HEADERS = [
    "Device ID",
    "Legacy Device SKU",
    "Manufacturer Name",
    "Device Family Name",
    "Product Line",
    "Device Series",
    "Device Name",
    "Model Number",
    "Variant",
    "Missing Identity Inputs",
    "Identity Concern",
    "Required Action",
    "Review Status",
    "Reviewer Notes",
]
RELATIONSHIP_HEADERS = [
    "Device ID",
    "Device Name",
    "Manufacturer ID",
    "Device Family Code",
    "Compatibility Status",
    "Service Mapping Status",
    "Parts Mapping Status",
    "Missing Relationships",
    "Required Action",
    "Review Status",
    "Reviewer Notes",
]
DEVICE_STATUSES = {"Draft", "Active", "Planned", "Future", "Discontinued", "Archived", "Rejected"}
REVIEW_STATUSES = {
    "Pending Review",
    "Pending Manufacturer Review",
    "Pending Family Review",
    "Pending Identity Review",
    "Pending Relationship Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
}
COMPATIBILITY_STATUSES = {
    "Pending Compatibility Review",
    "Family-Level Only",
    "Model-Level Proposed",
    "Variant-Level Proposed",
    "Not Applicable",
    "Approved",
    "Rejected",
}
SERVICE_MAPPING_STATUSES = {
    "Pending Service Mapping",
    "Family Services Available",
    "Model Services Proposed",
    "Complete",
    "Not Applicable",
    "Approved",
    "Rejected",
}
PARTS_MAPPING_STATUSES = {
    "Pending Parts Mapping",
    "Family Parts Available",
    "Model Parts Proposed",
    "Complete",
    "Not Applicable",
    "Approved",
    "Rejected",
}
YES_NO_FIELDS = {
    "Active",
    "Wi-Fi Only",
    "Cellular Capable",
    "Repair Supported",
    "Parts Supported",
    "Mail-In Eligible",
    "Mobile Service Eligible",
}
GENERATED_BLANK_FIELDS = {
    "Device Series",
    "Model Number",
    "Variant",
    "Generation",
    "Release Year",
    "Form Factor",
    "Operating System Family",
    "Network Type",
    "Storage Capacity",
    "Memory Capacity",
    "Color",
    "Region",
    "Carrier",
    "Currency",
    "Reviewer",
    "Effective Date",
    "Last Reviewed",
}
DEFINED_NAME_BY_HEADER = {
    "Active": "DV_YesNo",
    "Status": "DV_DeviceStatuses",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Product Line": "DV_ProductLines",
    "Device Series": "DV_DeviceSeries",
    "Form Factor": "DV_FormFactors",
    "Operating System Family": "DV_OperatingSystemFamilies",
    "Network Type": "DV_NetworkTypes",
    "Wi-Fi Only": "DV_YesNo",
    "Cellular Capable": "DV_YesNo",
    "Repair Supported": "DV_YesNo",
    "Parts Supported": "DV_YesNo",
    "Mail-In Eligible": "DV_YesNo",
    "Mobile Service Eligible": "DV_YesNo",
    "Compatibility Status": "DV_CompatibilityStatuses",
    "Service Mapping Status": "DV_ServiceMappingStatuses",
    "Parts Mapping Status": "DV_PartsMappingStatuses",
    "Review Status": "DV_ReviewStatuses",
    "Currency": "DV_Currencies",
}
DEFINED_NAME_SHEET = {
    "DV_YesNo": "09 - Device Statuses",
    "DV_DeviceStatuses": "09 - Device Statuses",
    "DV_ReviewStatuses": "09 - Device Statuses",
    "DV_CompatibilityStatuses": "09 - Device Statuses",
    "DV_ServiceMappingStatuses": "09 - Device Statuses",
    "DV_PartsMappingStatuses": "09 - Device Statuses",
    "DV_Currencies": "09 - Device Statuses",
    "DV_ManufacturerIDs": "02 - Manufacturers",
    "DV_DeviceFamilyCodes": "03 - Device Families",
    "DV_ProductLines": "04 - Product Lines",
    "DV_DeviceSeries": "05 - Device Series",
    "DV_FormFactors": "06 - Form Factors",
    "DV_OperatingSystemFamilies": "07 - Operating Systems",
    "DV_NetworkTypes": "08 - Network Types",
}
PROHIBITED_HEADERS = {
    "Serial Number",
    "IMEI",
    "Stock",
    "Stock Quantity",
    "Bin",
    "Location",
    "Inventory Quantity",
    "Final Cost",
    "Markup",
    "Margin",
    "Final Customer Price",
}


class DevicesValidationError(RuntimeError):
    """Raised when the Master Devices workbook violates its contract."""


def text(value: Any) -> str:
    """Return stripped text, treating None as blank."""
    return "" if value is None else str(value).strip()


def decimal_value(value: Any, field: str) -> Decimal | None:
    """Parse a monetary value with blank-versus-zero semantics."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        result = Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise DevicesValidationError(f"{field} is not numeric: {value!r}") from exc
    if not result.is_finite() or result < 0:
        raise DevicesValidationError(f"{field} must be finite and nonnegative: {value!r}")
    return result


def persisted_equal(expected: Any, actual: Any) -> bool:
    """Compare source values after normal Excel persistence normalization."""
    if expected in (None, "") and actual in (None, ""):
        return True
    if isinstance(expected, (date, datetime)) or isinstance(actual, (date, datetime)):
        return expected == actual
    return text(expected) == text(actual)


def file_hash(path: Path) -> str:
    """Return SHA-256 for a file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Validate protected source workbooks."""
    for path in paths:
        if not path.exists():
            raise DevicesValidationError(f"Required source workbook does not exist: {path}")
        if path.stat().st_size <= 0:
            raise DevicesValidationError(f"Required source workbook is empty: {path}")
        if not zipfile.is_zipfile(path):
            raise DevicesValidationError(f"Required source workbook is not a valid ZIP-based Office file: {path}")


def require_generated_workbook(path: Path) -> None:
    """Validate the independent validator input before openpyxl reads it."""
    if not path.exists() or path.stat().st_size <= 0 or not zipfile.is_zipfile(path):
        raise DevicesValidationError("Generated devices workbook is missing or invalid; rerun the generator successfully.")
    with zipfile.ZipFile(path) as archive:
        required_members = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
        missing = sorted(required_members - set(archive.namelist()))
    if missing:
        raise DevicesValidationError(
            f"Generated devices workbook is missing required OOXML members: {missing}"
        )


def read_records(path: Path, sheet: str, header_row: int, *, keep_vba: bool = False) -> list[dict[str, Any]]:
    """Read one worksheet into dictionaries."""
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    try:
        if sheet not in workbook.sheetnames:
            raise DevicesValidationError(f"Required worksheet {sheet!r} is missing from {path}")
        worksheet = workbook[sheet]
        headers = [text(cell.value) for cell in worksheet[header_row]]
        if not any(headers):
            raise DevicesValidationError(f"Header row {header_row} is blank in {path}!{sheet}")
        records = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            record = {header: value for header, value in zip_longest(headers, values, fillvalue=None) if header}
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Return an Excel Table's exact header order."""
    if table_name not in worksheet.tables:
        raise DevicesValidationError(f"Missing table {table_name} on {worksheet.title}")
    min_column, min_row, max_column, _max_row = range_boundaries(worksheet.tables[table_name].ref)
    return [text(worksheet.cell(min_row, column).value) for column in range(min_column, max_column + 1)]


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Return an Excel Table's populated records."""
    headers = table_headers(worksheet, table_name)
    min_column, min_row, max_column, max_row = range_boundaries(worksheet.tables[table_name].ref)
    return [
        {
            header: worksheet.cell(row, column).value
            for header, column in zip(headers, range(min_column, max_column + 1), strict=True)
        }
        for row in range(min_row + 1, max_row + 1)
        if any(
            worksheet.cell(row, column).value not in (None, "")
            for column in range(min_column, max_column + 1)
        )
    ]


def source_devices() -> list[dict[str, Any]]:
    """Read and validate the retained Device population."""
    records = read_records(PROPOSAL_PATH, "01 - Retained", 1)
    required = {
        "Record Category",
        "Source Row Number",
        "Legacy SKU",
        "Legacy Group",
        "Legacy Type",
        "Legacy Manufacturer",
        "Legacy Name",
        "Legacy Retail Price",
        "Legacy Cost",
        "Legacy Note",
        "Source Workbook",
        "Source Worksheet",
        "Source Created At",
        "Source Updated At",
    }
    missing = required - set(records[0]) if records else required
    if missing:
        raise DevicesValidationError(f"Proposal retained table is missing headers: {sorted(missing)}")
    devices = [row for row in records if text(row["Record Category"]) == "Device"]
    if len(devices) != EXPECTED_DEVICE_ROWS:
        raise DevicesValidationError(f"Expected {EXPECTED_DEVICE_ROWS} source Devices; found {len(devices)}")
    devices.sort(key=lambda row: int(row["Source Row Number"]))
    numbers = [int(row["Source Row Number"]) for row in devices]
    if len(numbers) != len(set(numbers)):
        raise DevicesValidationError("Source Device row numbers are not unique")
    exclusions = read_records(PROPOSAL_PATH, "02 - Duplicate Exclusions", 1)
    excluded = {
        int(row["Excluded Source Row Number"])
        for row in exclusions
        if row.get("Excluded Source Row Number") not in (None, "")
    }
    if set(numbers) & excluded:
        raise DevicesValidationError("A generated source Device also occurs in duplicate exclusions")
    return devices


def canonical_identity() -> tuple[set[str], int, list[str], dict[str, str], dict[str, str]]:
    """Read canonical Device IDs, manufacturers, and families."""
    workbook = load_workbook(CANONICAL_PATH, read_only=True, data_only=True, keep_vba=True)
    try:
        if "32 Devices" not in workbook.sheetnames:
            raise DevicesValidationError("Canonical worksheet '32 Devices' is missing")
        devices = workbook["32 Devices"]
        headers = [text(cell.value) for cell in devices[1]]
        if "Device ID" not in headers:
            raise DevicesValidationError("Canonical '32 Devices' has no Device ID header")
        column = headers.index("Device ID")
        populated = [
            str(row[column])
            for row in devices.iter_rows(min_row=2, values_only=True)
            if row[column] is not None and str(row[column]).strip()
        ]
        valid_list = [value for value in populated if DEVICE_ID_PATTERN.fullmatch(value)]
        malformed = [value for value in populated if not DEVICE_ID_PATTERN.fullmatch(value)]
        duplicates = sorted(value for value, count in Counter(valid_list).items() if count > 1)
        if duplicates:
            raise DevicesValidationError(f"Canonical Device IDs are duplicated: {duplicates}")

        manufacturers_sheet = workbook["30 Manufacturers"]
        manufacturer_headers = [text(cell.value) for cell in manufacturers_sheet[3]]
        manufacturer_id_column = manufacturer_headers.index("Manufacturer ID")
        manufacturer_name_column = manufacturer_headers.index("Manufacturer")
        manufacturers = {
            text(row[manufacturer_id_column]): text(row[manufacturer_name_column])
            for row in manufacturers_sheet.iter_rows(min_row=4, values_only=True)
            if text(row[manufacturer_id_column]) and text(row[manufacturer_name_column])
        }
        families_sheet = workbook["31 Device Families"]
        family_headers = [text(cell.value) for cell in families_sheet[3]]
        family_id_column = family_headers.index("Device Family Code")
        family_name_column = family_headers.index("Device Family")
        families = {
            text(row[family_id_column]): text(row[family_name_column])
            for row in families_sheet.iter_rows(min_row=4, values_only=True)
            if text(row[family_id_column]) and text(row[family_name_column])
        }
    finally:
        workbook.close()
    return (
        set(valid_list),
        max((int(value[3:]) for value in valid_list), default=0),
        malformed,
        manufacturers,
        families,
    )


def validate_structure(workbook: Any) -> None:
    """Validate exact worksheet, table, and schema contracts."""
    if workbook.sheetnames != SHEET_NAMES:
        raise DevicesValidationError(f"Worksheet order differs from contract: {workbook.sheetnames!r}")
    if len(set(SHEET_NAMES)) != len(SHEET_NAMES) or any(len(name) > 31 for name in SHEET_NAMES):
        raise DevicesValidationError("Configured worksheet names are duplicate or exceed 31 characters")
    table_names = []
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        expected_table = TABLE_NAMES[sheet_name]
        if expected_table not in worksheet.tables:
            raise DevicesValidationError(f"Missing table {expected_table} on {sheet_name}")
        if worksheet.freeze_panes != "A2":
            raise DevicesValidationError(f"{sheet_name} does not freeze the header row")
        if worksheet.merged_cells.ranges:
            raise DevicesValidationError(f"{sheet_name} contains merged cells")
        table = worksheet.tables[expected_table]
        if not table.autoFilter:
            raise DevicesValidationError(f"{expected_table} does not have a filter")
        table_names.append(table.name)
    if len(table_names) != len(set(table_names)):
        raise DevicesValidationError("Excel Table names are not unique")
    if table_headers(workbook["01 - Master Devices"], TABLE_NAMES["01 - Master Devices"]) != DEVICE_HEADERS:
        raise DevicesValidationError("Master Devices schema is not the approved 48-column schema")
    if table_headers(workbook["10 - Identity Review"], TABLE_NAMES["10 - Identity Review"]) != IDENTITY_HEADERS:
        raise DevicesValidationError("Identity Review schema differs from the contract")
    if table_headers(workbook["11 - Relationship Review"], TABLE_NAMES["11 - Relationship Review"]) != RELATIONSHIP_HEADERS:
        raise DevicesValidationError("Relationship Review schema differs from the contract")
    if PROHIBITED_HEADERS & set(DEVICE_HEADERS):
        raise DevicesValidationError("Master Devices contains prohibited inventory or final-pricing fields")


def validate_names_and_validations(workbook: Any, row_count: int) -> None:
    """Validate defined names and field-level list validation coverage."""
    missing = sorted(set(DEFINED_NAME_SHEET) - set(workbook.defined_names))
    if missing:
        raise DevicesValidationError(f"Required defined names are missing: {missing}")
    for name, expected_sheet in DEFINED_NAME_SHEET.items():
        destinations = list(workbook.defined_names[name].destinations)
        if len(destinations) != 1 or destinations[0][0] != expected_sheet:
            raise DevicesValidationError(f"Defined name {name} does not resolve to {expected_sheet}")
        _sheet, reference = destinations[0]
        try:
            range_boundaries(reference.replace("$", ""))
        except ValueError as exc:
            raise DevicesValidationError(f"Defined name {name} has invalid range {reference}") from exc
    worksheet = workbook["01 - Master Devices"]
    validations = list(worksheet.data_validations.dataValidation)
    formula_counts = Counter(text(validation.formula1) for validation in validations)
    expected_counts = Counter(f"={name}" for name in DEFINED_NAME_BY_HEADER.values())
    if formula_counts != expected_counts:
        raise DevicesValidationError(
            f"List-validation formulas differ; expected {expected_counts}, found {formula_counts}"
        )
    for validation in validations:
        if validation.type != "list" or "!" in text(validation.formula1):
            raise DevicesValidationError("A validation is not a named list or uses a direct cross-sheet reference")
        ranges = list(validation.ranges.ranges)
        if len(ranges) != 1:
            raise DevicesValidationError(f"Validation {validation.formula1} does not have one range")
        min_column, min_row, max_column, max_row = range_boundaries(str(ranges[0]))
        if min_column != max_column or min_row != 2 or max_row != row_count + 1:
            raise DevicesValidationError(f"Validation {validation.formula1} has incomplete row coverage")
        header = DEVICE_HEADERS[min_column - 1]
        expected_formula = f"={DEFINED_NAME_BY_HEADER[header]}"
        if text(validation.formula1) != expected_formula:
            raise DevicesValidationError(f"{header} uses {validation.formula1}, expected {expected_formula}")


def validate_lookup_snapshots(
    workbook: Any,
    canonical_manufacturers: dict[str, str],
    canonical_families: dict[str, str],
    device_rows: Sequence[dict[str, Any]],
    sources: Sequence[dict[str, Any]],
) -> None:
    """Reconcile canonical lookup snapshots and controlled-list values."""
    manufacturer_rows = table_records(workbook["02 - Manufacturers"], "tblDeviceManufacturers")
    actual_manufacturers = {
        text(row["Manufacturer ID"]): text(row["Manufacturer Name"])
        for row in manufacturer_rows
        if text(row["Manufacturer ID"])
    }
    if actual_manufacturers != canonical_manufacturers:
        raise DevicesValidationError("Manufacturer lookup snapshot differs from canonical source")
    family_rows = table_records(workbook["03 - Device Families"], "tblDeviceFamilies")
    actual_families = {
        text(row["Device Family Code"]): text(row["Device Family Name"])
        for row in family_rows
        if text(row["Device Family Code"])
    }
    if actual_families != canonical_families:
        raise DevicesValidationError("Device Family lookup snapshot differs from canonical source")
    status_rows = table_records(workbook["09 - Device Statuses"], "tblDeviceStatuses")
    expected_sets = {
        "Device Status": DEVICE_STATUSES,
        "Review Status": REVIEW_STATUSES,
        "Compatibility Status": COMPATIBILITY_STATUSES,
        "Service Mapping Status": SERVICE_MAPPING_STATUSES,
        "Parts Mapping Status": PARTS_MAPPING_STATUSES,
        "Yes/No": {"Yes", "No"},
        "Currency": {"USD"},
    }
    for header, expected in expected_sets.items():
        actual = {text(row[header]) for row in status_rows if text(row[header])}
        if actual != expected:
            raise DevicesValidationError(f"{header} lookup differs: {actual}")
    product_lines = {
        text(row["Product Line"])
        for row in table_records(workbook["04 - Product Lines"], "tblProductLines")
        if text(row["Product Line"])
    }
    expected_product_lines = {
        text(source["Legacy Group"]) for source in sources if text(source["Legacy Group"])
    }
    if product_lines != expected_product_lines:
        raise DevicesValidationError("Product Line lookup does not reconcile to source observations")
    controlled_lookups = {
        "Device Series": (
            "05 - Device Series",
            "tblDeviceSeries",
        ),
        "Form Factor": (
            "06 - Form Factors",
            "tblFormFactors",
        ),
        "Operating System Family": (
            "07 - Operating Systems",
            "tblOperatingSystems",
        ),
        "Network Type": (
            "08 - Network Types",
            "tblNetworkTypes",
        ),
    }
    for field, (sheet, table_name) in controlled_lookups.items():
        allowed = {
            text(row[field])
            for row in table_records(workbook[sheet], table_name)
            if text(row[field])
        }
        invalid = sorted(
            {text(row[field]) for row in device_rows if text(row[field])} - allowed
        )
        if invalid:
            raise DevicesValidationError(f"{field} contains values outside its lookup: {invalid}")
    if any(text(row["Product Line"]) not in product_lines for row in device_rows):
        raise DevicesValidationError("A generated Product Line is absent from its lookup")
    if any(text(row["Currency"]) not in {"", "USD"} for row in device_rows):
        raise DevicesValidationError("A generated Currency is absent from its lookup")


def expected_family(source_type: Any, canonical_families: dict[str, str]) -> tuple[str, str]:
    """Return the only allowed explicit source-type family mapping."""
    value = text(source_type)
    if not value.casefold().startswith("device - "):
        return "", ""
    candidate = value.split("-", 1)[1].strip()
    matches = [(code, name) for code, name in canonical_families.items() if name.casefold() == candidate.casefold()]
    return matches[0] if len(matches) == 1 else ("", "")


def validate_device_rows(
    rows: Sequence[dict[str, Any]],
    sources: Sequence[dict[str, Any]],
    existing_ids: set[str],
    highest_existing: int,
    canonical_manufacturers: dict[str, str],
    canonical_families: dict[str, str],
) -> None:
    """Validate identity, source mapping, statuses, and monetary observations."""
    if len(rows) != EXPECTED_DEVICE_ROWS:
        raise DevicesValidationError(f"Expected {EXPECTED_DEVICE_ROWS} Device rows; found {len(rows)}")
    ids = [text(row["Device ID"]) for row in rows]
    if any(not DEVICE_ID_PATTERN.fullmatch(value) for value in ids):
        raise DevicesValidationError("A generated Device ID does not match ^DEV\\d{6}$")
    if len(ids) != len(set(ids)):
        raise DevicesValidationError("Generated Device IDs are duplicated")
    if set(ids) & existing_ids:
        raise DevicesValidationError("Generated Device IDs collide with canonical IDs")
    expected_numbers = list(range(highest_existing + 1, highest_existing + 1 + len(rows)))
    if [int(value[3:]) for value in ids] != expected_numbers:
        raise DevicesValidationError("Generated Device IDs are not the required continuous sequence")

    manufacturer_name_to_ids: dict[str, list[str]] = {}
    for identifier, name in canonical_manufacturers.items():
        manufacturer_name_to_ids.setdefault(name.casefold(), []).append(identifier)

    for worksheet_row, (row, source) in enumerate(zip(rows, sources, strict=True), start=2):
        source_number = int(source["Source Row Number"])
        if int(row["Source Record Number"]) != source_number:
            raise DevicesValidationError(f"Row {worksheet_row} source record does not preserve ordering")
        expected_text = {
            "Legacy Device SKU": source["Legacy SKU"],
            "Manufacturer Name": source["Legacy Manufacturer"],
            "Product Line": source["Legacy Group"],
            "Device Name": source["Legacy Name"],
            "Device Display Name": source["Legacy Name"],
            "Device Description": source["Legacy Note"],
            "Source Workbook": source["Source Workbook"],
            "Source Worksheet": source["Source Worksheet"],
            "Reviewer Notes": source["Legacy Note"],
        }
        for field, expected in expected_text.items():
            if not persisted_equal(expected, row[field]):
                raise DevicesValidationError(
                    f"Row {worksheet_row} ({ids[worksheet_row - 2]}) changed {field}: "
                    f"{expected!r} -> {row[field]!r}"
                )
        if text(row["Import Batch ID"]) != IMPORT_BATCH_ID:
            raise DevicesValidationError(f"Row {worksheet_row} has the wrong Import Batch ID")
        if not persisted_equal(source["Source Created At"], row["Created At"]):
            raise DevicesValidationError(f"Row {worksheet_row} changed Created At")
        if not persisted_equal(source["Source Updated At"], row["Updated At"]):
            raise DevicesValidationError(f"Row {worksheet_row} changed Updated At")
        if decimal_value(source["Legacy Retail Price"], "source retail price") != decimal_value(row["Legacy Retail Price"], "retail price"):
            raise DevicesValidationError(f"Row {worksheet_row} changed Legacy Retail Price")
        if decimal_value(source["Legacy Cost"], "source cost") != decimal_value(row["Legacy Cost"], "cost"):
            raise DevicesValidationError(f"Row {worksheet_row} changed Legacy Cost")

        manufacturer_name = text(row["Manufacturer Name"])
        exact_ids = manufacturer_name_to_ids.get(manufacturer_name.casefold(), [])
        expected_manufacturer_id = exact_ids[0] if len(exact_ids) == 1 else ""
        if text(row["Manufacturer ID"]) != expected_manufacturer_id:
            raise DevicesValidationError(f"Row {worksheet_row} has an unsupported Manufacturer ID mapping")
        family_code, family_name = expected_family(source["Legacy Type"], canonical_families)
        if text(row["Device Family Code"]) != family_code or text(row["Device Family Name"]) != family_name:
            raise DevicesValidationError(f"Row {worksheet_row} has an unsupported Device Family mapping")
        if text(row["Manufacturer ID"]) and canonical_manufacturers[text(row["Manufacturer ID"])] != manufacturer_name:
            raise DevicesValidationError(f"Row {worksheet_row} Manufacturer ID/name do not agree")
        if family_code and canonical_families[family_code] != text(row["Device Family Name"]):
            raise DevicesValidationError(f"Row {worksheet_row} Device Family code/name do not agree")

        if not expected_manufacturer_id:
            expected_review = "Pending Manufacturer Review"
        elif not family_code:
            expected_review = "Pending Family Review"
        else:
            expected_review = "Pending Identity Review"
        expected_defaults = {
            "Active": "Yes",
            "Status": "Draft",
            "Wi-Fi Only": "No",
            "Cellular Capable": "No",
            "Repair Supported": "No",
            "Parts Supported": "No",
            "Mail-In Eligible": "No",
            "Mobile Service Eligible": "No",
            "Compatibility Status": "Pending Compatibility Review",
            "Service Mapping Status": "Pending Service Mapping",
            "Parts Mapping Status": "Pending Parts Mapping",
            "Review Status": expected_review,
        }
        for field, expected in expected_defaults.items():
            if text(row[field]) != expected:
                raise DevicesValidationError(f"Row {worksheet_row} has invalid {field}: {row[field]!r}")
        if any(text(row[field]) for field in GENERATED_BLANK_FIELDS):
            raise DevicesValidationError(f"Row {worksheet_row} contains an inferred V1-only blank field")
        if any(text(row[field]) not in {"Yes", "No"} for field in YES_NO_FIELDS):
            raise DevicesValidationError(f"Row {worksheet_row} has a non-Yes/No boolean")
        if text(row["Status"]) not in DEVICE_STATUSES or text(row["Review Status"]) not in REVIEW_STATUSES:
            raise DevicesValidationError(f"Row {worksheet_row} contains an uncontrolled status")


def validate_review_tables(workbook: Any, device_rows: Sequence[dict[str, Any]]) -> None:
    """Validate complete, ordered identity and relationship queues."""
    device_ids = [text(row["Device ID"]) for row in device_rows]
    identity = table_records(workbook["10 - Identity Review"], "tblDeviceIdentityReview")
    relationships = table_records(workbook["11 - Relationship Review"], "tblDeviceRelationshipReview")
    for label, rows in (("Identity Review", identity), ("Relationship Review", relationships)):
        ids = [text(row["Device ID"]) for row in rows]
        if ids != device_ids or len(ids) != len(set(ids)):
            raise DevicesValidationError(f"{label} rows are missing, duplicated, or reordered")
    for row in identity:
        if not text(row["Missing Identity Inputs"]) or not text(row["Required Action"]):
            raise DevicesValidationError(f"Identity Review {row['Device ID']} lacks review instructions")
    for row in relationships:
        if text(row["Review Status"]) != "Pending Relationship Review":
            raise DevicesValidationError(f"Relationship Review {row['Device ID']} has an invalid status")
        if text(row["Missing Relationships"]) != "Compatibility; Service mapping; Parts mapping":
            raise DevicesValidationError(f"Relationship Review {row['Device ID']} omits required relationships")


def validate_metadata(
    workbook: Any,
    protected_hashes: dict[Path, str],
    highest_existing: int,
    malformed_ids: Sequence[str],
) -> None:
    """Validate metadata, provenance hashes, and visible validation results."""
    metadata_rows = table_records(workbook["14 - Import Metadata"], "tblMasterDevicesImportMetadata")
    metadata = {text(row["Metadata Key"]): row["Metadata Value"] for row in metadata_rows}
    expected = {
        "Import Batch ID": IMPORT_BATCH_ID,
        "Namespace Authority": "ADR-009",
        "Expected Device Rows": str(EXPECTED_DEVICE_ROWS),
        "Schema Columns": str(len(DEVICE_HEADERS)),
        "Highest Existing Device ID": (
            f"DEV{highest_existing:06d}"
            if highest_existing
            else "None - ADR-009 empty namespace"
        ),
        "First Proposed Device ID": f"DEV{highest_existing + 1:06d}",
        "Final Proposed Device ID": f"DEV{highest_existing + EXPECTED_DEVICE_ROWS:06d}",
        "Malformed Existing Device IDs": "; ".join(malformed_ids),
        "Canonical Write Performed": "No",
    }
    for key, value in expected.items():
        if text(metadata.get(key)) != value:
            raise DevicesValidationError(f"Metadata {key!r} differs: {metadata.get(key)!r}")
    for path, digest in protected_hashes.items():
        key = f"SHA-256 {path.name}"
        if text(metadata.get(key)) != digest:
            raise DevicesValidationError(f"Metadata hash for {path} differs from the current protected input")
    validations = table_records(workbook["12 - Validation Summary"], "tblMasterDevicesValidation")
    if not validations or any(text(row["Result"]) != "PASS" for row in validations):
        raise DevicesValidationError("Validation Summary contains a non-PASS result")


def main() -> int:
    """Run the independent validation workflow."""
    workbook: Any | None = None
    try:
        require_files(PROTECTED_PATHS)
        require_generated_workbook(OUTPUT_PATH)
        protected_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        sources = source_devices()
        existing_ids, highest_existing, malformed_ids, manufacturers, families = canonical_identity()
        workbook = load_workbook(OUTPUT_PATH, read_only=False, data_only=False)
        validate_structure(workbook)
        device_rows = table_records(workbook["01 - Master Devices"], "tblMasterDevicesCatalog")
        validate_names_and_validations(workbook, len(device_rows))
        validate_lookup_snapshots(
            workbook,
            manufacturers,
            families,
            device_rows,
            sources,
        )
        validate_device_rows(
            device_rows,
            sources,
            existing_ids,
            highest_existing,
            manufacturers,
            families,
        )
        validate_review_tables(workbook, device_rows)
        validate_metadata(workbook, protected_hashes, highest_existing, malformed_ids)
        after_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        if protected_hashes != after_hashes:
            raise DevicesValidationError("A protected source workbook hash changed during validation")
        print("Master Devices Catalog V1 independent validation: PASS")
        print(f"Validated workbook: {OUTPUT_PATH}")
        print(f"Device rows: {len(device_rows)}")
        print(f"Schema columns: {len(DEVICE_HEADERS)}")
        print(f"Device ID range: {device_rows[0]['Device ID']} through {device_rows[-1]['Device ID']}")
        print(f"Canonical malformed Device IDs excluded: {len(malformed_ids)}")
        print("Protected input hashes: PASS")
        print("Canonical database writes: NONE")
        return 0
    except (
        AssertionError,
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        DevicesValidationError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
