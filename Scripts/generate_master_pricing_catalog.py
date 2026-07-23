"""Generate the standalone Nocturnix Master Pricing Catalog v1 review workbook.

ADR-007 authorizes PRC000001 when the canonical pricing namespace is empty.
The script never saves or modifies an input workbook.
"""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
MASTER_SERVICES_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Services_Catalog_v1.xlsx"
)
LABOR_CATALOG_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards"
    r"\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
PROPOSAL_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
PRICING_ID_SOURCE_PATH = CANONICAL_PATH
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Pricing_Catalog_v1.xlsx"
)
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(
    f"{OUTPUT_PATH.stem}.tmp{OUTPUT_PATH.suffix}"
)

MASTER_SERVICES_SHEET = "01 - Master Services"
LABOR_SHEET = "01 - Labor Standards"
EXCLUSIONS_SHEET = "02 - Duplicate Exclusions"
PRICING_ID_SOURCE_SHEET = "50 Pricing"
EXPECTED_ROWS = 314
IMPORT_BATCH_ID = "MASTER-PRICING-V1-REVIEW"
PRICING_ID_PATTERN = re.compile(r"^PRC\d{6}$")
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
NAMESPACE_AUTHORITY = "ADR-007"

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Pricing Records",
    "02 - Pricing Statuses",
    "03 - Cost Components",
    "04 - Labor References",
    "05 - Service References",
    "06 - Margin Targets",
    "07 - Regional Markets",
    "08 - Pricing Methods",
    "09 - Review Queue",
    "10 - Validation Summary",
    "11 - Revision History",
    "12 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterPricingInstructions",
    "01 - Pricing Records": "tblMasterPricingCatalog",
    "02 - Pricing Statuses": "tblPricingStatuses",
    "03 - Cost Components": "tblCostComponents",
    "04 - Labor References": "tblPricingLaborReferences",
    "05 - Service References": "tblPricingServiceReferences",
    "06 - Margin Targets": "tblMarginTargets",
    "07 - Regional Markets": "tblRegionalMarkets",
    "08 - Pricing Methods": "tblPricingMethods",
    "09 - Review Queue": "tblPricingReviewQueue",
    "10 - Validation Summary": "tblMasterPricingValidation",
    "11 - Revision History": "tblMasterPricingRevisionHistory",
    "12 - Import Metadata": "tblMasterPricingImportMetadata",
}
PRICING_HEADERS = [
    "Pricing Record ID",
    "Service ID",
    "Legacy Service SKU",
    "Service Name",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Legacy Pricing Status",
    "Pricing Status",
    "Pricing Method",
    "Currency",
    "Legacy Retail Price",
    "Legacy Cost",
    "Part Cost",
    "Shipping Cost",
    "Consumables Cost",
    "Testing Cost",
    "Labor Standard ID",
    "Standard Labor Minutes",
    "Labor Rate Tier",
    "Labor Rate",
    "Labor Cost",
    "Overhead Allocation",
    "Processing Fees",
    "Warranty Allowance",
    "Risk Allowance",
    "Total Internal Cost",
    "Target Margin Percent",
    "Market Adjustment",
    "Recommended Price",
    "Minimum Approved Price",
    "Maximum Approved Price",
    "Final Customer Price",
    "Regional Market",
    "Effective Date",
    "Expiration Date",
    "Review Status",
    "Pricing Confidence",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Reviewer",
    "Reviewer Notes",
    "Created At",
    "Updated At",
]
PRICING_STATUSES = [
    "Pending Pricing Review",
    "Legacy Price Review",
    "Cost Research Required",
    "Labor Mapping Required",
    "Market Research Required",
    "Ready for Pricing Calculation",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
PRICING_METHODS = [
    "Cost Plus",
    "Market Aligned",
    "Fixed Service Fee",
    "Diagnostic Fee",
    "Labor Only",
    "Parts and Labor",
    "Manual Review",
    "Not Yet Determined",
]
REVIEW_STATUSES = [
    "Pending Review",
    "In Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
CONFIDENCE_VALUES = ["Unassessed", "Low", "Medium", "High"]
COST_FIELDS = [
    "Part Cost",
    "Shipping Cost",
    "Consumables Cost",
    "Testing Cost",
    "Labor Rate",
    "Labor Cost",
    "Overhead Allocation",
    "Processing Fees",
    "Warranty Allowance",
    "Risk Allowance",
]
MONEY_FIELDS = {
    "Legacy Retail Price",
    "Legacy Cost",
    *COST_FIELDS,
    "Total Internal Cost",
    "Market Adjustment",
    "Recommended Price",
    "Minimum Approved Price",
    "Maximum Approved Price",
    "Final Customer Price",
}
CALCULATED_FIELDS = {
    "Labor Cost",
    "Total Internal Cost",
    "Recommended Price",
    "Minimum Approved Price",
    "Maximum Approved Price",
    "Final Customer Price",
}
DEFINED_NAME_BY_HEADER = {
    "Pricing Status": "DV_PricingStatuses",
    "Pricing Method": "DV_PricingMethods",
    "Currency": "DV_Currencies",
    "Service ID": "DV_ServiceIDs",
    "Labor Standard ID": "DV_LaborStandardIDs",
    "Labor Rate Tier": "DV_LaborRateTiers",
    "Target Margin Percent": "DV_MarginTargets",
    "Regional Market": "DV_RegionalMarkets",
    "Review Status": "DV_ReviewStatuses",
    "Pricing Confidence": "DV_PricingConfidenceValues",
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
PERCENT_FORMAT = "0.00%"
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


class PricingCatalogError(RuntimeError):
    """Raised when the pricing review workbook cannot be generated safely."""


def text(value: Any) -> str:
    """Return stripped text, treating None as blank."""
    return "" if value is None else str(value).strip()


def ascii_value(value: Any) -> Any:
    """Return ASCII-safe text while preserving non-text values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return value.encode("ascii", "replace").decode("ascii")


def excel_safe_value(value: Any) -> Any:
    """Normalize workbook-bound date/time values for Excel persistence."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def assert_excel_safe_value_contract() -> None:
    """Exercise the Excel persistence normalization boundary."""
    aware_utc = datetime(2026, 7, 23, 12, 30, tzinfo=UTC)
    assert excel_safe_value(aware_utc) == datetime(2026, 7, 23, 12, 30)

    eastern = timezone(timedelta(hours=-4))
    aware_eastern = datetime(2026, 7, 23, 8, 30, tzinfo=eastern)
    assert excel_safe_value(aware_eastern) == datetime(2026, 7, 23, 12, 30)

    naive = datetime(2026, 7, 23, 12, 30)
    assert excel_safe_value(naive) is naive

    calendar_date = date(2026, 7, 23)
    assert excel_safe_value(calendar_date) is calendar_date

    aware_clock = time(8, 30, tzinfo=eastern)
    normalized_clock = excel_safe_value(aware_clock)
    assert normalized_clock == time(8, 30)
    assert normalized_clock.tzinfo is None

    assert excel_safe_value("") == ""
    assert excel_safe_value(None) is None
    timestamp_text = "2026-07-23T12:30:00Z"
    assert excel_safe_value(timestamp_text) == timestamp_text


def decimal_value(value: Any, field: str, *, allow_text: bool = False) -> Decimal | None:
    """Parse a finite Decimal without conflating blank and zero."""
    if value is None or text(value) == "":
        return None
    if isinstance(value, bool):
        raise PricingCatalogError(f"{field} contains Boolean value {value!r}")
    try:
        result = Decimal(text(value))
    except InvalidOperation:
        if allow_text:
            return None
        raise PricingCatalogError(f"{field} is not numeric: {value!r}") from None
    if not result.is_finite():
        raise PricingCatalogError(f"{field} is not finite: {value!r}")
    return result


def file_hash(path: Path) -> str:
    """Calculate SHA-256 without changing the file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require readable, nonempty ZIP-based Excel inputs."""
    seen: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved in seen:
            raise PricingCatalogError(f"Protected input configured twice: {path}")
        seen.add(resolved)
        if not path.is_file():
            raise PricingCatalogError(f"Required workbook does not exist: {path}")
        if path.stat().st_size <= 0:
            raise PricingCatalogError(f"Required workbook is empty: {path}")
        if not zipfile.is_zipfile(path):
            raise PricingCatalogError(f"Required workbook is not a valid Excel ZIP: {path}")


def read_records(
    path: Path,
    sheet_name: str,
    *,
    keep_vba: bool = False,
) -> list[dict[str, Any]]:
    """Read a header-row worksheet without saving its workbook."""
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_vba=keep_vba,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise PricingCatalogError(f"{path.name} lacks worksheet {sheet_name!r}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [text(value) for value in next(rows)]
        except StopIteration as exc:
            raise PricingCatalogError(f"{path.name}:{sheet_name} is empty") from exc
        if not any(headers):
            raise PricingCatalogError(f"{path.name}:{sheet_name} has no header row")
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    finally:
        workbook.close()


def read_existing_pricing_ids() -> tuple[set[str], list[str]]:
    """Read canonical PRC IDs, permitting ADR-007's approved empty namespace."""
    workbook = load_workbook(
        PRICING_ID_SOURCE_PATH,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )
    try:
        if PRICING_ID_SOURCE_SHEET not in workbook.sheetnames:
            raise PricingCatalogError(
                f"Canonical workbook lacks worksheet {PRICING_ID_SOURCE_SHEET!r}"
            )
        worksheet = workbook[PRICING_ID_SOURCE_SHEET]
        header_cell = None
        for row in worksheet.iter_rows():
            for cell in row:
                if text(cell.value) == "Pricing Record ID":
                    header_cell = cell
                    break
            if header_cell is not None:
                break
        if header_cell is None:
            return set(), []
        valid: set[str] = set()
        malformed: list[str] = []
        for row in range(header_cell.row + 1, worksheet.max_row + 1):
            candidate = text(worksheet.cell(row=row, column=header_cell.column).value)
            if not candidate:
                continue
            if PRICING_ID_PATTERN.fullmatch(candidate):
                if candidate in valid:
                    raise PricingCatalogError(
                        "Duplicate canonical Pricing Record ID: "
                        f"{candidate} in worksheet {PRICING_ID_SOURCE_SHEET!r}"
                    )
                valid.add(candidate)
            else:
                malformed.append(candidate)
        return valid, malformed
    finally:
        workbook.close()


def allocate_pricing_ids(existing: set[str], count: int) -> list[str]:
    """Allocate and validate the ADR-007 continuous PRC sequence."""
    highest = max((int(identifier[3:]) for identifier in existing), default=0)
    generated = [f"PRC{number:06d}" for number in range(highest + 1, highest + count + 1)]
    if any(not PRICING_ID_PATTERN.fullmatch(identifier) for identifier in generated):
        raise PricingCatalogError("Generated Pricing Record ID has invalid format")
    if len(generated) != len(set(generated)):
        raise PricingCatalogError("Generated Pricing Record IDs are not unique")
    if existing.intersection(generated):
        raise PricingCatalogError("Generated Pricing Record ID reuses an existing ID")
    numeric_ids = [int(identifier[3:]) for identifier in generated]
    expected_numbers = list(range(highest + 1, highest + count + 1))
    if numeric_ids != expected_numbers:
        raise PricingCatalogError("Generated Pricing Record IDs are not continuous")
    if generated:
        expected_first = (
            f"PRC{highest + 1:06d}" if existing else "PRC000001"
        )
        if generated[0] != expected_first:
            raise PricingCatalogError(
                f"First generated Pricing Record ID must be {expected_first}"
            )
        if int(generated[-1][3:]) != int(generated[0][3:]) + count - 1:
            raise PricingCatalogError(
                "Final Pricing Record ID does not reconcile to generated row count"
            )
    return generated


def require_headers(
    rows: Sequence[dict[str, Any]],
    required: Iterable[str],
    label: str,
) -> None:
    """Require source columns even when the source has no data rows."""
    if not rows:
        raise PricingCatalogError(f"{label} contains no records")
    missing = sorted(set(required) - set(rows[0]))
    if missing:
        raise PricingCatalogError(f"{label} missing columns: {', '.join(missing)}")


def labor_lookup(rows: Sequence[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Index protected labor rows by unique Labor ID."""
    lookup: dict[str, dict[str, Any]] = {}
    for row in rows:
        identifier = text(row.get("Labor ID"))
        if not identifier:
            continue
        if identifier in lookup:
            raise PricingCatalogError(f"Duplicate Labor ID in source: {identifier}")
        lookup[identifier] = row
    return lookup


def validate_sources(
    services: list[dict[str, Any]],
    labor: list[dict[str, Any]],
    exclusions: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Validate the 314-row service population and its protected relationships."""
    require_headers(
        services,
        {
            "Service ID",
            "Legacy Service SKU",
            "Service Name",
            "Manufacturer ID",
            "Manufacturer Name",
            "Device Family Code",
            "Device Family Name",
            "Labor Standard ID",
            "Standard Minutes",
            "Labor Tier",
            "Pricing Status",
            "Legacy Retail Price",
            "Legacy Cost",
            "Source Record Number",
            "Source Workbook",
            "Source Worksheet",
            "Status",
            "Review Status",
            "Created At",
            "Updated At",
        },
        "Master Services",
    )
    require_headers(
        labor,
        {"Labor ID", "Standard Minutes", "Labor Rate Tier"},
        "Labor Standards",
    )
    if len(services) != EXPECTED_ROWS:
        raise PricingCatalogError(
            f"Expected {EXPECTED_ROWS} Master Services rows; found {len(services)}"
        )
    service_ids = [text(row["Service ID"]) for row in services]
    if len(set(service_ids)) != len(service_ids):
        raise PricingCatalogError("Master Services contains duplicate Service IDs")
    invalid = [value for value in service_ids if not SERVICE_ID_PATTERN.fullmatch(value)]
    if invalid:
        raise PricingCatalogError(f"Invalid Service IDs: {', '.join(invalid[:10])}")
    source_rows = [text(row["Source Record Number"]) for row in services]
    if "" in source_rows or len(set(source_rows)) != len(source_rows):
        raise PricingCatalogError("Source Record Numbers must be nonblank and unique")
    try:
        numeric_source_rows = [int(value) for value in source_rows]
    except ValueError as exc:
        raise PricingCatalogError(
            "Source Record Numbers must be integer values"
        ) from exc
    if numeric_source_rows != sorted(numeric_source_rows):
        raise PricingCatalogError(
            "Master Services must be in ascending Source Record Number order"
        )
    prohibited = [
        row["Service ID"]
        for row in services
        if text(row.get("Status")) in {"Rejected", "Archived"}
        or text(row.get("Review Status")) in {"Rejected", "Archived"}
    ]
    if prohibited:
        raise PricingCatalogError(
            f"Rejected or archived services are not eligible: {prohibited[:10]}"
        )
    excluded_rows = {
        text(row.get("Excluded Source Row Number") or row.get("Source Row Number"))
        for row in exclusions
    }
    overlap = sorted(set(source_rows).intersection(excluded_rows))
    if overlap:
        raise PricingCatalogError(
            f"Master Services reintroduces duplicate-exclusion rows: {overlap[:10]}"
        )
    lookup = labor_lookup(labor)
    for row in services:
        labor_id = text(row["Labor Standard ID"])
        if not labor_id:
            continue
        if labor_id not in lookup:
            raise PricingCatalogError(
                f"Service {row['Service ID']} references unknown Labor ID {labor_id}"
            )
        source = lookup[labor_id]
        if row["Standard Minutes"] != source["Standard Minutes"]:
            raise PricingCatalogError(
                f"Service {row['Service ID']} changed Standard Minutes for {labor_id}"
            )
        if text(row["Labor Tier"]) != text(source["Labor Rate Tier"]):
            raise PricingCatalogError(
                f"Service {row['Service ID']} changed Labor Rate Tier for {labor_id}"
            )
    return lookup


def initial_pricing_status(service: dict[str, Any]) -> str:
    """Apply the approved initial-status precedence."""
    if not text(service["Labor Standard ID"]):
        return "Labor Mapping Required"
    return "Cost Research Required"


def build_pricing_rows(
    services: Sequence[dict[str, Any]],
    pricing_ids: Sequence[str],
) -> list[dict[str, Any]]:
    """Create one deliberately unresolved pricing record per service."""
    records: list[dict[str, Any]] = []
    for pricing_id, service in zip(pricing_ids, services, strict=True):
        row = {header: "" for header in PRICING_HEADERS}
        row.update(
            {
                "Pricing Record ID": pricing_id,
                "Service ID": service["Service ID"],
                "Legacy Service SKU": service["Legacy Service SKU"],
                "Service Name": service["Service Name"],
                "Manufacturer ID": service["Manufacturer ID"],
                "Manufacturer Name": service["Manufacturer Name"],
                "Device Family Code": service["Device Family Code"],
                "Device Family Name": service["Device Family Name"],
                "Legacy Pricing Status": service["Pricing Status"],
                "Pricing Status": initial_pricing_status(service),
                "Pricing Method": "Not Yet Determined",
                "Legacy Retail Price": service["Legacy Retail Price"],
                "Legacy Cost": service["Legacy Cost"],
                "Labor Standard ID": service["Labor Standard ID"],
                "Standard Labor Minutes": service["Standard Minutes"],
                "Labor Rate Tier": service["Labor Tier"],
                "Review Status": "Pending Review",
                "Pricing Confidence": "Unassessed",
                "Source Record Number": service["Source Record Number"],
                "Source Workbook": service["Source Workbook"],
                "Source Worksheet": service["Source Worksheet"],
                "Import Batch ID": IMPORT_BATCH_ID,
                "Created At": service["Created At"],
                "Updated At": service["Updated At"],
            }
        )
        records.append(row)
    return records


def build_review_queue(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Route unresolved inputs without proposing resolutions."""
    queue: list[dict[str, Any]] = []
    for row in records:
        missing = []
        if not text(row["Labor Standard ID"]):
            missing.append("Labor Mapping")
        missing.extend(
            [
                "Part Cost Research",
                "Shipping Cost Research",
                "Consumables Review",
                "Testing Cost Review",
                "Labor Rate Review",
                "Overhead Review",
                "Fee Review",
                "Warranty/Risk Review",
                "Margin Review",
                "Regional Market Review",
                "Legacy Price Review",
            ]
        )
        queue.append(
            {
                "Pricing Record ID": row["Pricing Record ID"],
                "Service ID": row["Service ID"],
                "Service Name": row["Service Name"],
                "Pricing Status": row["Pricing Status"],
                "Missing Inputs": "; ".join(missing),
                "Required Action": "Complete documented research and review",
                "Review Priority": (
                    "High" if row["Pricing Status"] == "Labor Mapping Required" else "Medium"
                ),
                "Review Status": "Pending Review",
                "Reviewer Notes": "",
            }
        )
    return queue


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    records: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Append an Excel Table with a required placeholder row when empty."""
    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    if records:
        for record in records:
            worksheet.append(
                [
                    excel_safe_value(ascii_value(record.get(header, "")))
                    for header in headers
                ]
            )
    else:
        worksheet.append(["" for _ in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"
    table = Table(
        displayName=table_name,
        ref=f"A1:{get_column_letter(len(headers))}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    format_columns(worksheet, headers)


def format_columns(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Apply readable widths and semantic number formats."""
    for column, header in enumerate(headers, 1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = (
            42 if any(word in header for word in ("Notes", "Inputs", "Guidance")) else 22
        )
        if header in MONEY_FIELDS:
            number_format = CURRENCY_FORMAT
        elif header == "Target Margin Percent":
            number_format = PERCENT_FORMAT
        elif header in {"Effective Date", "Expiration Date", "Date"}:
            number_format = DATE_FORMAT
        elif header in {"Created At", "Updated At", "Generated At UTC"}:
            number_format = DATETIME_FORMAT
        else:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = number_format


def add_defined_names(workbook: Workbook, lengths: dict[str, int]) -> None:
    """Create every workbook-scoped validation name."""
    ranges = {
        "DV_PricingStatuses": "'02 - Pricing Statuses'!$B$2:$B$11",
        "DV_ReviewStatuses": "'02 - Pricing Statuses'!$B$12:$B$17",
        "DV_PricingConfidenceValues": "'02 - Pricing Statuses'!$B$18:$B$21",
        "DV_Currencies": "'02 - Pricing Statuses'!$B$22:$B$22",
        "DV_PricingMethods": "'08 - Pricing Methods'!$A$2:$A$9",
        "DV_ServiceIDs": f"'05 - Service References'!$A$2:$A${lengths['services'] + 1}",
        "DV_LaborStandardIDs": (
            f"'04 - Labor References'!$A$2:$A${lengths['labor'] + 1}"
        ),
        "DV_LaborRateTiers": (
            f"'04 - Labor References'!$C$2:$C${lengths['labor'] + 1}"
        ),
        "DV_MarginTargets": "'06 - Margin Targets'!$B$2:$B$2",
        "DV_RegionalMarkets": "'07 - Regional Markets'!$A$2:$A$2",
    }
    for name, reference in ranges.items():
        workbook.defined_names.add(DefinedName(name, attr_text=reference))


def add_data_validations(workbook: Workbook) -> None:
    """Apply only defined-name list validations to Pricing Records."""
    worksheet = workbook["01 - Pricing Records"]
    headers = [text(cell.value) for cell in worksheet[1]]
    for header, name in DEFINED_NAME_BY_HEADER.items():
        column = headers.index(header) + 1
        letter = get_column_letter(column)
        validation = DataValidation(type="list", formula1=f"={name}", allow_blank=True)
        validation.error = f"Select a controlled value for {header}."
        validation.errorTitle = "Invalid value"
        worksheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{worksheet.max_row}")


def add_conditional_formatting(worksheet: Worksheet) -> None:
    """Highlight unresolved and prohibited status states."""
    headers = [text(cell.value) for cell in worksheet[1]]
    end_row = worksheet.max_row
    for header in ("Pricing Status", "Review Status"):
        letter = get_column_letter(headers.index(header) + 1)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{end_row}",
            FormulaRule(
                formula=[
                    f'OR(ISNUMBER(SEARCH("Required",{letter}2)),'
                    f'ISNUMBER(SEARCH("Pending",{letter}2)))'
                ],
                fill=YELLOW_FILL,
            ),
        )
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{end_row}",
            CellIsRule(operator="equal", formula=['"Approved"'], fill=RED_FILL),
        )
    for header in COST_FIELDS:
        letter = get_column_letter(headers.index(header) + 1)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{end_row}",
            FormulaRule(formula=[f'LEN(TRIM({letter}2))=0'], fill=YELLOW_FILL),
        )


def add_queue_conditional_formatting(worksheet: Worksheet) -> None:
    """Highlight every unresolved review queue row."""
    headers = [text(cell.value) for cell in worksheet[1]]
    status_letter = get_column_letter(headers.index("Review Status") + 1)
    missing_letter = get_column_letter(headers.index("Missing Inputs") + 1)
    worksheet.conditional_formatting.add(
        f"{status_letter}2:{status_letter}{worksheet.max_row}",
        CellIsRule(operator="equal", formula=['"Pending Review"'], fill=YELLOW_FILL),
    )
    worksheet.conditional_formatting.add(
        f"{missing_letter}2:{missing_letter}{worksheet.max_row}",
        FormulaRule(
            formula=[f"LEN(TRIM({missing_letter}2))>0"],
            fill=YELLOW_FILL,
        ),
    )


def build_workbook(
    pricing_rows: list[dict[str, Any]],
    services: list[dict[str, Any]],
    labor: list[dict[str, Any]],
    review_queue: list[dict[str, Any]],
    hashes: dict[Path, str],
    existing_ids: set[str],
    malformed_ids: Sequence[str],
) -> Workbook:
    """Build the complete 13-sheet review artifact in memory."""
    if len(SHEET_NAMES) != len(set(SHEET_NAMES)) or any(
        len(name) > 31 for name in SHEET_NAMES
    ):
        raise PricingCatalogError("Worksheet names must be unique and <= 31 characters")
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}

    instructions = [
        {"Topic": "Purpose", "Guidance": "Pricing review framework only."},
        {"Topic": "Population", "Guidance": "One row per reviewed Master Service."},
        {"Topic": "Legacy Values", "Guidance": "Historical/provisional observations only."},
        {"Topic": "Calculation", "Guidance": "No recommended or final price is calculated."},
        {"Topic": "Approval", "Guidance": "No generated row is approved."},
        {"Topic": "Import", "Guidance": "No canonical import is performed."},
    ]
    append_table(sheets[SHEET_NAMES[0]], ["Topic", "Guidance"], instructions, TABLE_NAMES[SHEET_NAMES[0]])
    append_table(sheets[SHEET_NAMES[1]], PRICING_HEADERS, pricing_rows, TABLE_NAMES[SHEET_NAMES[1]])

    status_rows = (
        [{"Value Type": "Pricing Status", "Value": value, "Description": "Controlled lifecycle value"} for value in PRICING_STATUSES]
        + [{"Value Type": "Review Status", "Value": value, "Description": "Controlled review value"} for value in REVIEW_STATUSES]
        + [{"Value Type": "Pricing Confidence", "Value": value, "Description": "Controlled evidence confidence"} for value in CONFIDENCE_VALUES]
        + [{"Value Type": "Currency", "Value": "USD", "Description": "Candidate only; not defaulted"}]
    )
    append_table(sheets[SHEET_NAMES[2]], ["Value Type", "Value", "Description"], status_rows, TABLE_NAMES[SHEET_NAMES[2]])

    cost_rows = [
        {
            "Cost Component": field,
            "Required for Calculation": "Yes",
            "Initial Status": "Unresolved",
            "Owner": "Pricing Governance",
            "Notes": "Do not invent or convert a legacy zero into a verified cost.",
        }
        for field in COST_FIELDS
    ]
    append_table(sheets[SHEET_NAMES[3]], list(cost_rows[0]), cost_rows, TABLE_NAMES[SHEET_NAMES[3]])

    labor_references = [
        {
            "Labor Standard ID": row["Labor ID"],
            "Standard Labor Minutes": row["Standard Minutes"],
            "Labor Rate Tier": row["Labor Rate Tier"],
            "Repair Difficulty": row.get("Repair Difficulty", ""),
            "Skill Level": row.get("Skill Level", ""),
            "Status": row.get("Status", ""),
        }
        for row in labor
        if text(row.get("Labor ID"))
    ]
    append_table(
        sheets[SHEET_NAMES[4]],
        ["Labor Standard ID", "Standard Labor Minutes", "Labor Rate Tier", "Repair Difficulty", "Skill Level", "Status"],
        labor_references,
        TABLE_NAMES[SHEET_NAMES[4]],
    )
    service_references = [
        {
            "Service ID": row["Service ID"],
            "Service Name": row["Service Name"],
            "Manufacturer ID": row["Manufacturer ID"],
            "Manufacturer Name": row["Manufacturer Name"],
            "Device Family Code": row["Device Family Code"],
            "Device Family Name": row["Device Family Name"],
            "Source Record Number": row["Source Record Number"],
        }
        for row in services
    ]
    append_table(sheets[SHEET_NAMES[5]], list(service_references[0]), service_references, TABLE_NAMES[SHEET_NAMES[5]])
    append_table(
        sheets[SHEET_NAMES[6]],
        ["Margin Target ID", "Target Margin Percent", "Status", "Notes"],
        [{"Margin Target ID": "UNRESOLVED", "Target Margin Percent": "", "Status": "Pending Governance", "Notes": "No target approved."}],
        TABLE_NAMES[SHEET_NAMES[6]],
    )
    append_table(
        sheets[SHEET_NAMES[7]],
        ["Regional Market", "Status", "Notes"],
        [{"Regional Market": "", "Status": "Pending Research", "Notes": "No region approved."}],
        TABLE_NAMES[SHEET_NAMES[7]],
    )
    method_rows = [
        {"Pricing Method": value, "Description": "Controlled proposed method", "Active": "Yes"}
        for value in PRICING_METHODS
    ]
    append_table(sheets[SHEET_NAMES[8]], list(method_rows[0]), method_rows, TABLE_NAMES[SHEET_NAMES[8]])
    queue_headers = [
        "Pricing Record ID", "Service ID", "Service Name", "Pricing Status",
        "Missing Inputs", "Required Action", "Review Priority", "Review Status",
        "Reviewer Notes",
    ]
    append_table(sheets[SHEET_NAMES[9]], queue_headers, review_queue, TABLE_NAMES[SHEET_NAMES[9]])
    validation_rows = [
        {"Validation Check": "Pricing row count", "Result": "PASS", "Evidence": len(pricing_rows)},
        {"Validation Check": "Final customer prices blank", "Result": "PASS", "Evidence": len(pricing_rows)},
        {"Validation Check": "Protected inputs", "Result": "PASS", "Evidence": len(hashes)},
        {"Validation Check": "Canonical import", "Result": "PASS", "Evidence": "Not performed"},
    ]
    append_table(sheets[SHEET_NAMES[10]], list(validation_rows[0]), validation_rows, TABLE_NAMES[SHEET_NAMES[10]])
    revision = [{
        "Version": "1.0 Draft",
        "Date": datetime.now(UTC).date(),
        "Description": "Initial Master Pricing review framework",
        "Status": "Draft",
        "Approved By": "",
        "Notes": "No pricing approval or canonical import.",
    }]
    append_table(sheets[SHEET_NAMES[11]], list(revision[0]), revision, TABLE_NAMES[SHEET_NAMES[11]])
    highest = (
        max(existing_ids, key=lambda value: int(value[3:]))
        if existing_ids
        else ""
    )
    metadata = [
        {"Metadata Field": "Artifact", "Value": "Master Pricing Catalog V1 Review"},
        {"Metadata Field": "Generated At UTC", "Value": datetime.now(UTC)},
        {"Metadata Field": "Import Batch ID", "Value": IMPORT_BATCH_ID},
        {"Metadata Field": "Pricing Row Count", "Value": len(pricing_rows)},
        {"Metadata Field": "Pricing ID Source", "Value": str(PRICING_ID_SOURCE_PATH)},
        {"Metadata Field": "Pricing ID Worksheet", "Value": PRICING_ID_SOURCE_SHEET},
        {"Metadata Field": "Namespace Authority", "Value": NAMESPACE_AUTHORITY},
        {"Metadata Field": "Existing Pricing Record Count", "Value": len(existing_ids)},
        {"Metadata Field": "Highest Existing Pricing Record ID", "Value": highest},
        {"Metadata Field": "First Generated Pricing Record ID", "Value": pricing_rows[0]["Pricing Record ID"]},
        {"Metadata Field": "Final Generated Pricing Record ID", "Value": pricing_rows[-1]["Pricing Record ID"]},
        {"Metadata Field": "Generated Pricing Record Count", "Value": len(pricing_rows)},
        {"Metadata Field": "Malformed Existing Pricing IDs", "Value": "; ".join(malformed_ids)},
        {"Metadata Field": "Schema Columns", "Value": len(PRICING_HEADERS)},
        {"Metadata Field": "Output", "Value": str(OUTPUT_PATH)},
    ]
    for path, digest in hashes.items():
        metadata.extend(
            [
                {"Metadata Field": f"Protected Input Path: {path.name}", "Value": str(path)},
                {"Metadata Field": f"SHA-256: {path.name}", "Value": digest},
            ]
        )
    append_table(sheets[SHEET_NAMES[12]], ["Metadata Field", "Value"], metadata, TABLE_NAMES[SHEET_NAMES[12]])

    add_defined_names(
        workbook,
        {"services": len(service_references), "labor": len(labor_references)},
    )
    add_data_validations(workbook)
    add_conditional_formatting(sheets[SHEET_NAMES[1]])
    add_queue_conditional_formatting(sheets[SHEET_NAMES[9]])
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read records from a named generated table."""
    if table_name not in worksheet.tables:
        raise PricingCatalogError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = [text(value) for value in rows[0]]
    return [
        dict(zip(headers, row, strict=False))
        for row in rows[1:]
        if any(value is not None and text(value) != "" for value in row)
    ]


def validate_reopened(
    output_path: Path,
    expected_rows: Sequence[dict[str, Any]],
    expected_hashes: dict[Path, str],
) -> None:
    """Reopen and validate the generated artifact before reporting success."""
    if (
        not output_path.is_file()
        or output_path.stat().st_size <= 0
        or not zipfile.is_zipfile(output_path)
    ):
        raise PricingCatalogError(
            "Temporary generated pricing workbook is missing or invalid"
        )
    with zipfile.ZipFile(output_path) as archive:
        required_members = {"[Content_Types].xml", "xl/workbook.xml"}
        if (
            not required_members.issubset(archive.namelist())
            or archive.testzip() is not None
        ):
            raise PricingCatalogError(
                "Temporary generated pricing workbook is missing Excel content"
            )
    workbook = load_workbook(output_path, data_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise PricingCatalogError(
                f"Reopened worksheet order changed: {workbook.sheetnames}"
            )
        actual_tables: list[str] = []
        for name in SHEET_NAMES:
            worksheet = workbook[name]
            if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
                raise PricingCatalogError(f"{name} lacks frozen header or filter")
            expected_table = TABLE_NAMES[name]
            if expected_table not in worksheet.tables:
                raise PricingCatalogError(f"{name} lacks table {expected_table}")
            actual_tables.extend(worksheet.tables)
        if len(actual_tables) != len(set(actual_tables)):
            raise PricingCatalogError("Excel Table names are not unique")
        records = table_records(
            workbook["01 - Pricing Records"],
            "tblMasterPricingCatalog",
        )
        if len(records) != EXPECTED_ROWS or len(records) != len(expected_rows):
            raise PricingCatalogError("Reopened pricing row count changed")
        headers = [
            text(value)
            for value in next(
                workbook["01 - Pricing Records"].iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                )
            )
        ]
        if headers != PRICING_HEADERS:
            raise PricingCatalogError("Reopened Pricing Records schema changed")
        for index, (expected, actual) in enumerate(
            zip(expected_rows, records, strict=True),
            2,
        ):
            for field in PRICING_HEADERS:
                expected_value = excel_safe_value(expected[field])
                actual_value = excel_safe_value(actual[field])
                if text(expected_value) != text(actual_value):
                    raise PricingCatalogError(
                        f"Reopened row {index} changed {field}: "
                        f"{expected[field]!r} -> {actual[field]!r}"
                    )
        names = set(workbook.defined_names)
        missing_names = set(DEFINED_NAME_BY_HEADER.values()) - names
        if missing_names:
            raise PricingCatalogError(
                f"Reopened workbook missing defined names: {sorted(missing_names)}"
            )
        validations = workbook["01 - Pricing Records"].data_validations.dataValidation
        formulas = {text(item.formula1) for item in validations if item.type == "list"}
        required_formulas = {f"={name}" for name in DEFINED_NAME_BY_HEADER.values()}
        if formulas != required_formulas or any("!" in formula for formula in formulas):
            raise PricingCatalogError(
                f"Unexpected list-validation formulas: {sorted(formulas)}"
            )
        for row in records:
            if any(text(row[field]) for field in {*COST_FIELDS, *CALCULATED_FIELDS}):
                raise PricingCatalogError(
                    f"V1 cost/calculated field populated for {row['Service ID']}"
                )
            expected_status = (
                "Labor Mapping Required"
                if not text(row["Labor Standard ID"])
                else "Cost Research Required"
            )
            if text(row["Pricing Status"]) != expected_status:
                raise PricingCatalogError(
                    f"Reopened {row['Service ID']} has Pricing Status "
                    f"{row['Pricing Status']!r}; expected {expected_status!r}"
                )
            if (
                text(row["Pricing Method"]) != "Not Yet Determined"
                or text(row["Review Status"]) != "Pending Review"
                or text(row["Pricing Confidence"]) != "Unassessed"
                or text(row["Currency"])
            ):
                raise PricingCatalogError(
                    f"Reopened {row['Service ID']} changed a V1 review default"
                )
            for field in MONEY_FIELDS:
                amount = decimal_value(row[field], field, allow_text=field.startswith("Legacy"))
                if amount is not None and amount < 0:
                    raise PricingCatalogError(
                        f"Negative {field} for {row['Service ID']}: {row[field]!r}"
                    )
        queue = table_records(
            workbook["09 - Review Queue"],
            "tblPricingReviewQueue",
        )
        if len(queue) != len(records):
            raise PricingCatalogError("Reopened Review Queue row count changed")
        for record, item in zip(records, queue, strict=True):
            if any(
                text(record[field]) != text(item[field])
                for field in (
                    "Pricing Record ID",
                    "Service ID",
                    "Service Name",
                    "Pricing Status",
                )
            ):
                raise PricingCatalogError(
                    f"Reopened Review Queue changed {record['Service ID']}"
                )
    finally:
        workbook.close()
    current_hashes = {path: file_hash(path) for path in expected_hashes}
    if current_hashes != expected_hashes:
        changed = [str(path) for path in expected_hashes if current_hashes[path] != expected_hashes[path]]
        raise PricingCatalogError(f"Protected input hash changed: {changed}")


def main() -> int:
    """Generate and validate the standalone pricing review workbook."""
    protected = (
        MASTER_SERVICES_PATH,
        LABOR_CATALOG_PATH,
        PROPOSAL_PATH,
        CANONICAL_PATH,
    )
    workbook: Workbook | None = None
    published = False
    try:
        assert_excel_safe_value_contract()
        require_files(protected)
        hashes = {path: file_hash(path) for path in protected}
        existing_ids, malformed_ids = read_existing_pricing_ids()
        services = read_records(MASTER_SERVICES_PATH, MASTER_SERVICES_SHEET)
        labor = read_records(LABOR_CATALOG_PATH, LABOR_SHEET)
        exclusions = read_records(PROPOSAL_PATH, EXCLUSIONS_SHEET)
        validate_sources(services, labor, exclusions)
        pricing_ids = allocate_pricing_ids(existing_ids, len(services))
        pricing_rows = build_pricing_rows(services, pricing_ids)
        review_queue = build_review_queue(pricing_rows)
        workbook = build_workbook(
            pricing_rows,
            services,
            labor,
            review_queue,
            hashes,
            existing_ids,
            malformed_ids,
        )
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        if TEMP_OUTPUT_PATH.exists():
            TEMP_OUTPUT_PATH.unlink()
        workbook.save(TEMP_OUTPUT_PATH)
        workbook.close()
        workbook = None
        validate_reopened(TEMP_OUTPUT_PATH, pricing_rows, hashes)
        TEMP_OUTPUT_PATH.replace(OUTPUT_PATH)
        published = True
    except (
        AssertionError,
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        PricingCatalogError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError as close_error:
                print(
                    f"ERROR: Failed to close temporary workbook: {close_error}",
                    file=sys.stderr,
                )
        if not published:
            try:
                if TEMP_OUTPUT_PATH.exists():
                    TEMP_OUTPUT_PATH.unlink()
            except OSError as cleanup_error:
                print(
                    f"ERROR: Failed to remove temporary workbook "
                    f"{TEMP_OUTPUT_PATH}: {cleanup_error}",
                    file=sys.stderr,
                )

    print(f"Generated: {OUTPUT_PATH}")
    print(f"Namespace authority: {NAMESPACE_AUTHORITY}")
    print(f"Existing Pricing Record IDs: {len(existing_ids)}")
    print(
        "Generated Pricing Record ID range: "
        f"{pricing_rows[0]['Pricing Record ID']} through "
        f"{pricing_rows[-1]['Pricing Record ID']}"
    )
    print(f"Malformed existing Pricing Record IDs: {len(malformed_ids)}")
    print(f"Pricing records: {len(pricing_rows)}")
    print(f"Review queue rows: {len(review_queue)}")
    print("Final customer prices: PASS (all blank)")
    print("Protected input hashes: PASS")
    print("Canonical import: NOT PERFORMED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
