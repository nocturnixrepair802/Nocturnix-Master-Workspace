"""Validate the standalone Nocturnix Master Services Catalog v1 workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_PATH = Path(r"D:\Business Portal\300_Pricing\Legacy\Raw Import Data.xlsx")
STAGING_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Legacy_Catalog_Staging_Preview_v1.xlsx"
)
PROPOSAL_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
LABOR_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
SERVICE_ID_SOURCE_PATH = CANONICAL_PATH
SERVICE_ID_SOURCE_SHEET = "34 Master Services"
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Services_Catalog_v1.xlsx"
)

EXPECTED_SERVICE_ROWS = 314
SHEET_NAMES = [
    "00 - Instructions",
    "01 - Master Services",
    "02 - Service Categories",
    "03 - Repair Types",
    "04 - Device Families",
    "05 - Manufacturers",
    "06 - Labor Standards",
    "07 - Labor Tiers",
    "08 - Difficulty Levels",
    "09 - Skill Levels",
    "10 - Turnaround Times",
    "11 - Warranty Options",
    "12 - Status Values",
    "13 - Validation Summary",
    "14 - Revision History",
    "15 - Import Metadata",
]
MASTER_HEADERS = [
    "Service ID", "Legacy Service SKU", "Active", "Status", "Manufacturer ID",
    "Manufacturer Name", "Device Family Code", "Device Family Name", "Device Series",
    "Device Model", "Service Category ID", "Service Category", "Repair Type ID",
    "Repair Type", "Service Name", "Service Display Name", "Service Description",
    "Labor Standard ID", "Standard Minutes", "Minimum Minutes", "Maximum Minutes",
    "Labor Tier", "Repair Difficulty", "Skill Level", "Turnaround Time",
    "Requires Parts", "Requires Labor", "Diagnostic Required", "Warranty Eligible",
    "Default Warranty", "Mobile Service Eligible", "Mail-In Eligible", "Pricing Status",
    "Legacy Retail Price", "Legacy Cost", "Source Record Number", "Source Workbook",
    "Source Worksheet", "Import Batch ID", "Review Status", "Reviewer Notes",
    "Effective Date", "Last Reviewed", "Created At", "Updated At",
]
YES_NO_FIELDS = {
    "Active", "Requires Parts", "Requires Labor", "Diagnostic Required",
    "Warranty Eligible", "Mobile Service Eligible", "Mail-In Eligible",
}
PRICING_STATUSES = {
    "Pending Pricing Review", "Legacy Price Review", "No Pricing Exceptions",
    "Archive Candidate",
}
DEFINED_NAME_BY_HEADER = {
    "Active": "DV_YesNo",
    "Status": "DV_ServiceStatuses",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Service Category ID": "DV_ServiceCategoryIDs",
    "Repair Type ID": "DV_RepairTypeIDs",
    "Labor Standard ID": "DV_LaborStandardIDs",
    "Labor Tier": "DV_LaborTiers",
    "Repair Difficulty": "DV_DifficultyLevels",
    "Skill Level": "DV_SkillLevels",
    "Turnaround Time": "DV_TurnaroundTimes",
    "Requires Parts": "DV_YesNo",
    "Requires Labor": "DV_YesNo",
    "Diagnostic Required": "DV_YesNo",
    "Warranty Eligible": "DV_YesNo",
    "Default Warranty": "DV_WarrantyOptions",
    "Mobile Service Eligible": "DV_YesNo",
    "Mail-In Eligible": "DV_YesNo",
    "Pricing Status": "DV_PricingStatuses",
    "Review Status": "DV_ReviewStatuses",
}
REQUIRED_DEFINED_NAMES = set(DEFINED_NAME_BY_HEADER.values())
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
REQUIRED_FIELDS = {
    "Service ID", "Active", "Status", "Service Category ID", "Service Category",
    "Repair Type ID", "Repair Type", "Service Name", "Service Display Name",
    "Service Description", "Turnaround Time", "Requires Parts", "Requires Labor",
    "Diagnostic Required", "Warranty Eligible", "Default Warranty",
    "Mobile Service Eligible", "Mail-In Eligible", "Pricing Status",
    "Source Record Number", "Source Workbook", "Source Worksheet", "Import Batch ID",
    "Review Status",
}


class ValidationError(RuntimeError):
    """Raised when a validation rule fails."""


def text(value: Any) -> str:
    """Normalize a workbook scalar to stripped text."""
    return "" if value is None else str(value).strip()


def number(value: Any) -> Decimal | None:
    """Parse numeric workbook data for ordering checks."""
    if value is None or text(value) == "" or isinstance(value, bool):
        return None
    try:
        result = Decimal(text(value).replace("$", "").replace(",", ""))
    except InvalidOperation:
        return None
    return result if result.is_finite() else None


def sha256_file(path: Path) -> str:
    """Return a streaming SHA-256 hash."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require every validation input."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise ValidationError(f"Required file(s) missing: {', '.join(missing)}")


def worksheet_records(worksheet: Worksheet) -> list[dict[str, Any]]:
    """Read a worksheet's header-row dataset."""
    headers = [text(cell.value) for cell in worksheet[1]]
    return [
        dict(zip(headers, values, strict=False))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None and text(value) != "" for value in values)
    ]


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read records from a named table located anywhere on a worksheet."""
    if table_name not in worksheet.tables:
        raise ValidationError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = [text(value) for value in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:] if row[0]]


def external_sheet_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read a protected reference sheet without saving it."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Missing source worksheet {sheet_name!r}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        headers = [text(value) for value in next(rows)]
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    finally:
        workbook.close()


def read_existing_service_ids() -> tuple[set[str], list[str]]:
    """Read canonical Service IDs and report malformed values separately."""
    if not SERVICE_ID_SOURCE_PATH.is_file():
        raise ValidationError(
            f"Canonical Service ID source does not exist: {SERVICE_ID_SOURCE_PATH}"
        )
    if SERVICE_ID_SOURCE_PATH.stat().st_size <= 0:
        raise ValidationError(
            f"Canonical Service ID source is empty: {SERVICE_ID_SOURCE_PATH}"
        )
    if not zipfile.is_zipfile(SERVICE_ID_SOURCE_PATH):
        raise ValidationError(
            "Canonical Service ID source is not a valid Excel ZIP container: "
            f"{SERVICE_ID_SOURCE_PATH}"
        )
    try:
        workbook = load_workbook(
            SERVICE_ID_SOURCE_PATH,
            read_only=True,
            data_only=True,
            keep_vba=True,
        )
    except Exception as exc:
        raise ValidationError(
            f"Unable to open canonical Service ID source "
            f"{SERVICE_ID_SOURCE_PATH}: {exc}"
        ) from exc
    try:
        if SERVICE_ID_SOURCE_SHEET not in workbook.sheetnames:
            raise ValidationError(
                f"Missing canonical worksheet {SERVICE_ID_SOURCE_SHEET!r}"
            )
        worksheet = workbook[SERVICE_ID_SOURCE_SHEET]
        header_row = 0
        service_id_column = 0
        for row in worksheet.iter_rows(min_row=1, max_row=min(50, worksheet.max_row)):
            for cell in row:
                if text(cell.value) == "Service ID":
                    header_row = cell.row
                    service_id_column = cell.column
                    break
            if service_id_column:
                break
        if not service_id_column:
            raise ValidationError("Canonical Service ID header was not found")
        valid: set[str] = set()
        malformed: list[str] = []
        for (value,) in worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=worksheet.max_row,
            min_col=service_id_column,
            max_col=service_id_column,
            values_only=True,
        ):
            identifier = text(value)
            if not identifier:
                continue
            if SERVICE_ID_PATTERN.fullmatch(identifier):
                valid.add(identifier)
            else:
                malformed.append(identifier)
        if not valid:
            raise ValidationError(
                "No valid canonical Service IDs found; sequence cannot be validated"
            )
        return valid, malformed
    finally:
        workbook.close()


def lookup_values(workbook: Any, sheet_name: str, column_name: str) -> set[str]:
    """Return populated values from a lookup worksheet column."""
    records = worksheet_records(workbook[sheet_name])
    return {text(record.get(column_name)) for record in records if text(record.get(column_name))}


def lookup_map(
    workbook: Any, sheet_name: str, id_field: str, name_field: str
) -> dict[str, str]:
    """Return a unique ID/name map from a generated lookup worksheet."""
    result: dict[str, str] = {}
    for record in worksheet_records(workbook[sheet_name]):
        identifier = text(record.get(id_field))
        name = text(record.get(name_field))
        if not identifier:
            continue
        if identifier in result and result[identifier] != name:
            raise ValidationError(f"Lookup ID {identifier!r} has inconsistent names")
        result[identifier] = name
    return result


def validate_defined_names_and_validations(workbook: Any) -> list[str]:
    """Require named validation lists and prohibit direct cross-sheet formulas."""
    missing = sorted(REQUIRED_DEFINED_NAMES - set(workbook.defined_names.keys()))
    if missing:
        raise ValidationError(f"Required defined names missing: {missing}")
    master = workbook[SHEET_NAMES[1]]
    headers = [text(cell.value) for cell in master[1]]
    actual_by_header: dict[str, str] = {}
    for validation in master.data_validations.dataValidation:
        if validation.type != "list":
            continue
        formula = text(validation.formula1)
        if "!" in formula:
            raise ValidationError(
                f"Direct cross-sheet list validation is prohibited: {formula}"
            )
        ranges = str(validation.sqref).split()
        if not ranges:
            raise ValidationError("List validation has no target range")
        min_col, _min_row, _max_col, _max_row = range_boundaries(ranges[0])
        actual_by_header[headers[min_col - 1]] = formula
    for header, defined_name in DEFINED_NAME_BY_HEADER.items():
        expected = f"={defined_name}"
        if actual_by_header.get(header) != expected:
            raise ValidationError(
                f"Validation for {header!r} must use {expected}; found "
                f"{actual_by_header.get(header)!r}"
            )
    return ["Defined names and list-validation references: PASS"]


def validate_structure(workbook: Any) -> list[str]:
    """Validate sheets, tables, schema, and reopenability."""
    if workbook.sheetnames != SHEET_NAMES:
        raise ValidationError("Worksheet names/order do not match specification")
    if len(set(workbook.sheetnames)) != 16 or any(len(name) > 31 for name in workbook.sheetnames):
        raise ValidationError("Worksheet names are not unique or exceed 31 characters")
    master = workbook[SHEET_NAMES[1]]
    if "tblMasterServicesCatalog" not in master.tables:
        raise ValidationError("tblMasterServicesCatalog is missing")
    headers = [cell.value for cell in master[1]]
    if headers != MASTER_HEADERS:
        raise ValidationError("Master Services columns do not match specification")
    forbidden = {"Final Customer Price", "Customer Price", "Final Price", "Final Cost"}
    if forbidden & set(headers):
        raise ValidationError("Prohibited final price or cost field exists")
    table_names = [name for sheet in workbook.worksheets for name in sheet.tables]
    if len(table_names) != len(set(table_names)):
        raise ValidationError("Excel Table names are not unique")
    if any(not sheet.tables for sheet in workbook.worksheets):
        raise ValidationError("Every populated worksheet must contain an Excel Table")
    return ["Workbook structure: PASS", f"Excel Tables: PASS ({len(table_names)} unique)"]


def validate_services(
    workbook: Any, existing_ids: set[str]
) -> tuple[list[dict[str, Any]], list[str]]:
    """Validate service identities, controlled values, lookups, and labor bounds."""
    records = worksheet_records(workbook[SHEET_NAMES[1]])
    if len(records) != EXPECTED_SERVICE_ROWS:
        raise ValidationError(f"Expected {EXPECTED_SERVICE_ROWS} services, found {len(records)}")
    ids = [text(record["Service ID"]) for record in records]
    if any(not SERVICE_ID_PATTERN.fullmatch(value) for value in ids):
        raise ValidationError("A Service ID does not match ^SVC\\d{6}$")
    if len(ids) != len(set(ids)):
        raise ValidationError("Generated Service IDs are duplicated")
    overlap = set(ids) & existing_ids
    if overlap:
        raise ValidationError(
            f"Generated Service IDs already exist: {sorted(overlap)[:10]}"
        )
    numbers = [int(identifier[3:]) for identifier in ids]
    expected_first = max(int(identifier[3:]) for identifier in existing_ids) + 1
    expected_numbers = list(range(expected_first, expected_first + len(records)))
    if numbers != expected_numbers:
        raise ValidationError("Generated Service IDs are not one continuous sequence")
    if numbers[0] != expected_first:
        raise ValidationError("First Service ID is not highest existing plus one")
    expected_final = expected_first + len(records) - 1
    if numbers[-1] != expected_final:
        raise ValidationError("Final Service ID does not reconcile to row count")
    source_rows = [int(record["Source Record Number"]) for record in records]
    if len(source_rows) != len(set(source_rows)):
        raise ValidationError("Source row provenance is not unique")
    for row_number, record in enumerate(records, start=2):
        missing = sorted(field for field in REQUIRED_FIELDS if text(record.get(field)) == "")
        if missing:
            raise ValidationError(f"Required fields blank in row {row_number}: {missing}")
        for field in YES_NO_FIELDS:
            if record.get(field) not in {"Yes", "No"}:
                raise ValidationError(f"Invalid Yes/No value in row {row_number}, {field}")
        if record.get("Pricing Status") not in PRICING_STATUSES:
            raise ValidationError(f"Invalid Pricing Status in row {row_number}")
        review_status = text(record.get("Review Status"))
        manufacturer_id = text(record.get("Manufacturer ID"))
        family_code = text(record.get("Device Family Code"))
        labor_id = text(record.get("Labor Standard ID"))
        if not manufacturer_id and review_status != "Pending Review":
            raise ValidationError(
                f"Blank Manufacturer ID requires Pending Review in row {row_number}"
            )
        if not family_code and review_status != "Pending Review":
            raise ValidationError(
                f"Blank Device Family Code requires Pending Review in row {row_number}"
            )
        if not labor_id and review_status not in {
            "Pending Labor Mapping",
            "Pending Review",
        }:
            raise ValidationError(
                f"Blank Labor Standard ID has invalid status in row {row_number}"
            )
        if review_status == "Ready for Approval" and not all(
            (manufacturer_id, family_code, labor_id)
        ):
            raise ValidationError(
                f"Ready for Approval row {row_number} has a blank relationship"
            )

    lookup_rules = {
        "Manufacturer ID": ("05 - Manufacturers", "Manufacturer ID", True),
        "Device Family Code": ("04 - Device Families", "Device Family Code", True),
        "Service Category ID": ("02 - Service Categories", "Service Category ID", False),
        "Repair Type ID": ("03 - Repair Types", "Repair Type ID", False),
        "Labor Standard ID": ("06 - Labor Standards", "Labor ID", True),
        "Labor Tier": ("07 - Labor Tiers", "Labor Rate Tier", True),
        "Repair Difficulty": ("08 - Difficulty Levels", "Difficulty Level", True),
        "Skill Level": ("09 - Skill Levels", "Skill Level", True),
        "Turnaround Time": ("10 - Turnaround Times", "Turnaround Time", False),
        "Default Warranty": ("11 - Warranty Options", "Warranty Option", False),
    }
    for field, (sheet, lookup_column, optional) in lookup_rules.items():
        valid = lookup_values(workbook, sheet, lookup_column)
        for row_number, record in enumerate(records, start=2):
            value = text(record.get(field))
            if not value and optional:
                continue
            if value not in valid:
                raise ValidationError(f"Invalid {field} reference in row {row_number}: {value!r}")

    relationship_rules = {
        ("Manufacturer ID", "Manufacturer Name"): lookup_map(
            workbook, "05 - Manufacturers", "Manufacturer ID", "Manufacturer Name"
        ),
        ("Device Family Code", "Device Family Name"): lookup_map(
            workbook,
            "04 - Device Families",
            "Device Family Code",
            "Device Family Name",
        ),
        ("Repair Type ID", "Repair Type"): lookup_map(
            workbook, "03 - Repair Types", "Repair Type ID", "Repair Type"
        ),
        ("Service Category ID", "Service Category"): lookup_map(
            workbook,
            "02 - Service Categories",
            "Service Category ID",
            "Service Category",
        ),
    }
    for row_number, record in enumerate(records, start=2):
        for (id_field, name_field), mapping in relationship_rules.items():
            identifier = text(record.get(id_field))
            name = text(record.get(name_field))
            if identifier and mapping.get(identifier) != name:
                raise ValidationError(
                    f"{id_field}/{name_field} mismatch in row {row_number}: "
                    f"{identifier!r}/{name!r}"
                )

    warranty_values = lookup_values(
        workbook, "11 - Warranty Options", "Warranty Option"
    )
    if "N/A" not in warranty_values:
        raise ValidationError("Warranty Options does not contain N/A")
    if any(text(record.get("Default Warranty")) not in warranty_values for record in records):
        raise ValidationError("A service Default Warranty is not a valid lookup")

    for row_number, record in enumerate(records, start=2):
        labor_id = text(record.get("Labor Standard ID"))
        if not labor_id:
            continue
        minimum = number(record.get("Minimum Minutes"))
        standard = number(record.get("Standard Minutes"))
        maximum = number(record.get("Maximum Minutes"))
        if standard is None or standard <= 0:
            raise ValidationError(f"Invalid Standard Minutes in row {row_number}")
        if minimum is None or maximum is None or not minimum <= standard <= maximum:
            raise ValidationError(f"Invalid labor minute bounds in row {row_number}")
    return records, [
        f"Service population: PASS ({len(records)})",
        f"Service ID sequence: PASS ({ids[0]} through {ids[-1]})",
        "Conditional statuses and ID/name relationships: PASS",
        "Warranty lookup: PASS (N/A available)",
        "Labor mappings and bounds: PASS",
    ]


def validate_provenance(records: Sequence[dict[str, Any]]) -> list[str]:
    """Reconcile source rows/SKUs and reject excluded proposal rows."""
    retained = external_sheet_records(PROPOSAL_PATH, "01 - Retained")
    retained_repairs = {
        int(row["Source Row Number"]): text(row.get("Legacy SKU"))
        for row in retained
        if text(row.get("Record Category")) == "Repair"
    }
    exclusions = external_sheet_records(PROPOSAL_PATH, "02 - Duplicate Exclusions")
    excluded_rows = {
        int(row.get("Excluded Source Row Number", row.get("Source Row Number")))
        for row in exclusions
    }
    imported_rows = {int(row["Source Record Number"]) for row in records}
    if imported_rows & excluded_rows:
        raise ValidationError("A proposed duplicate exclusion was imported")
    if imported_rows != set(retained_repairs):
        raise ValidationError("Retained Repair source population was not preserved")
    for record in records:
        source_row = int(record["Source Record Number"])
        if text(record.get("Legacy Service SKU")) != retained_repairs[source_row]:
            raise ValidationError(f"Legacy Service SKU changed for source row {source_row}")
        if text(record.get("Review Status")) in {"Rejected", "Archived"}:
            raise ValidationError(f"Rejected/archived source row imported: {source_row}")
    return ["Legacy SKU and provenance reconciliation: PASS", "Excluded/rejected rows absent: PASS"]


def validate_labor_source(records: Sequence[dict[str, Any]]) -> list[str]:
    """Ensure populated labor IDs exist in the protected labor workbook."""
    labor = external_sheet_records(LABOR_PATH, "01 - Labor Standards")
    labor_ids = {text(row.get("Labor ID")) for row in labor if text(row.get("Labor ID"))}
    populated = {text(row.get("Labor Standard ID")) for row in records if text(row.get("Labor Standard ID"))}
    missing = sorted(populated - labor_ids)
    if missing:
        raise ValidationError(f"Labor IDs absent from labor catalog: {missing[:10]}")
    return [f"Labor source references: PASS ({len(populated)} distinct IDs)"]


def validate_labor_audit(workbook: Any, records: Sequence[dict[str, Any]]) -> list[str]:
    """Validate complete, conservative labor-match audit coverage."""
    audit = table_records(workbook[SHEET_NAMES[13]], "tblLaborMatchAudit")
    if len(audit) != EXPECTED_SERVICE_ROWS:
        raise ValidationError("Labor match audit does not contain 314 rows")
    service_ids = {text(record.get("Service ID")) for record in records}
    audit_ids = {text(record.get("Service ID")) for record in audit}
    if audit_ids != service_ids:
        raise ValidationError("Labor match audit Service IDs do not reconcile")
    allowed = {"Mapped", "Pending Labor Mapping", "Ambiguous"}
    services_by_id = {text(record.get("Service ID")): record for record in records}
    for row_number, record in enumerate(audit, start=2):
        result = text(record.get("Mapping Result"))
        labor_id = text(record.get("Proposed Labor Standard ID"))
        service_id = text(record.get("Service ID"))
        service = services_by_id[service_id]
        service_labor_id = text(service.get("Labor Standard ID"))
        match_score = number(record.get("Match Score")) or Decimal("0")
        score_margin = number(record.get("Score Margin")) or Decimal("0")
        if result not in allowed:
            raise ValidationError(f"Invalid labor audit result in row {row_number}")
        if result != "Mapped" and labor_id:
            raise ValidationError(
                f"Unresolved labor audit row {row_number} invents a labor match"
            )
        if result == "Mapped" and (not labor_id or labor_id != service_labor_id):
            raise ValidationError(
                f"Mapped labor audit row {row_number} does not reconcile to service"
            )
        if result != "Mapped" and service_labor_id:
            raise ValidationError(
                f"Unresolved labor audit row {row_number} populated the service labor ID"
            )
        if result == "Mapped" and (
            match_score < Decimal("0.82") or score_margin <= Decimal("0.03")
        ):
            raise ValidationError(
                f"Mapped labor audit row {row_number} is below or tied at threshold"
            )
        if result == "Ambiguous" and (
            match_score < Decimal("0.82") or score_margin > Decimal("0.03")
        ):
            raise ValidationError(
                f"Ambiguous labor audit row {row_number} has inconsistent scores"
            )
    return ["Labor match audit: PASS (314 rows)"]


def validate_hashes(workbook: Any) -> list[str]:
    """Compare protected files to hashes embedded at generation time."""
    metadata = worksheet_records(workbook[SHEET_NAMES[15]])
    values = {text(row.get("Metadata Field")): text(row.get("Value")) for row in metadata}
    protected = (
        RAW_PATH,
        STAGING_PATH,
        PROPOSAL_PATH,
        LABOR_PATH,
        CANONICAL_PATH,
    )
    for path in protected:
        key = f"SHA-256: {path.name}"
        expected = values.get(key)
        if not expected:
            raise ValidationError(f"Missing protected hash metadata: {key}")
        actual = sha256_file(path)
        if actual != expected:
            raise ValidationError(f"Protected file hash changed: {path}")
    return ["Protected input and canonical database hashes: PASS (unchanged)"]


def main() -> int:
    """Run independent workbook and source-integrity validation."""
    workbook = None
    try:
        existing_ids, malformed_existing_ids = read_existing_service_ids()
        require_files(
            (
                OUTPUT_PATH,
                RAW_PATH,
                STAGING_PATH,
                PROPOSAL_PATH,
                LABOR_PATH,
                CANONICAL_PATH,
            )
        )
        workbook = load_workbook(OUTPUT_PATH, read_only=False, data_only=False)
        messages = validate_structure(workbook)
        messages.extend(validate_defined_names_and_validations(workbook))
        records, service_messages = validate_services(workbook, existing_ids)
        messages.extend(service_messages)
        messages.append(
            "Malformed existing Service IDs excluded from sequence calculation: "
            f"{len(malformed_existing_ids)}"
        )
        messages.extend(validate_provenance(records))
        messages.extend(validate_labor_source(records))
        messages.extend(validate_labor_audit(workbook, records))
        messages.extend(validate_hashes(workbook))
        print(f"Validated: {OUTPUT_PATH}")
        for message in messages:
            print(message)
        return 0
    except (ValidationError, OSError, ValueError, KeyError, IndexError, StopIteration) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
