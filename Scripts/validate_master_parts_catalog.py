"""Independently validate the Nocturnix Master Parts Catalog v1 workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROPOSAL_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
MASTER_SERVICES_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Services_Catalog_v1.xlsx"
)
MASTER_PRICING_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Pricing_Catalog_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Parts_Catalog_v1.xlsx"
)

RETAINED_SHEET = "01 - Retained"
EXCLUSIONS_SHEET = "02 - Duplicate Exclusions"
EXPECTED_PART_ROWS = 48
IMPORT_BATCH_ID = "MASTER-PARTS-V1-REVIEW"
NAMESPACE_AUTHORITY = "ADR-008"
PART_ID_PATTERN = re.compile(r"^PRT\d{6}$")
PART_ID_SHEETS = ("39.4 Repair Parts", "41 Parts")
SHEET_NAMES = [
    "00 - Instructions",
    "01 - Master Parts",
    "02 - Part Categories",
    "03 - Part Types",
    "04 - Manufacturers",
    "05 - Device Families",
    "06 - Device References",
    "07 - Suppliers",
    "08 - Part Qualities",
    "09 - Conditions",
    "10 - Compatibility Review",
    "11 - Sourcing Review",
    "12 - Validation Summary",
    "13 - Revision History",
    "14 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterPartsInstructions",
    "01 - Master Parts": "tblMasterPartsCatalog",
    "02 - Part Categories": "tblPartCategories",
    "03 - Part Types": "tblPartTypes",
    "04 - Manufacturers": "tblPartsManufacturers",
    "05 - Device Families": "tblPartsDeviceFamilies",
    "06 - Device References": "tblPartsDeviceReferences",
    "07 - Suppliers": "tblPartsSuppliers",
    "08 - Part Qualities": "tblPartQualities",
    "09 - Conditions": "tblPartConditions",
    "10 - Compatibility Review": "tblPartsCompatibilityReview",
    "11 - Sourcing Review": "tblPartsSourcingReview",
    "12 - Validation Summary": "tblMasterPartsValidation",
    "13 - Revision History": "tblMasterPartsRevisionHistory",
    "14 - Import Metadata": "tblMasterPartsImportMetadata",
}
PART_HEADERS = [
    "Part ID", "Legacy Part SKU", "Active", "Status", "Part Name",
    "Part Display Name", "Part Description", "Part Category ID", "Part Category",
    "Part Type ID", "Part Type", "Manufacturer ID", "Manufacturer Name",
    "Device Family Code", "Device Family Name", "Device ID", "Device Name",
    "Compatibility Scope", "Part Quality", "Condition", "OEM Status", "Color",
    "Capacity", "Model Number", "Supplier ID", "Supplier Name",
    "Supplier Part Number", "Preferred Supplier", "Legacy Retail Price",
    "Legacy Cost", "Currency", "Cost Status", "Pricing Status",
    "Inventory Tracked", "Serialized", "Warranty Eligible", "Default Warranty",
    "Compatibility Status", "Sourcing Status", "Review Status",
    "Source Record Number", "Source Workbook", "Source Worksheet",
    "Import Batch ID", "Reviewer", "Reviewer Notes", "Effective Date",
    "Last Reviewed", "Created At", "Updated At",
]
COMPATIBILITY_HEADERS = [
    "Part ID", "Legacy Part SKU", "Part Name", "Manufacturer Name",
    "Device Family Name", "Device ID", "Device Name",
    "Proposed Compatibility Scope", "Missing Compatibility Inputs",
    "Required Action", "Review Status", "Reviewer Notes",
]
SOURCING_HEADERS = [
    "Part ID", "Legacy Part SKU", "Part Name", "Supplier ID", "Supplier Name",
    "Supplier Part Number", "Legacy Cost", "Cost Status", "Sourcing Status",
    "Missing Sourcing Inputs", "Required Action", "Review Status",
    "Reviewer Notes",
]
PART_STATUSES = {"Draft", "Active", "Planned", "Future", "Archived", "Rejected"}
REVIEW_STATUSES = {
    "Pending Review", "Pending Manufacturer Review", "Pending Device Mapping",
    "Pending Compatibility Review", "Pending Sourcing Review",
    "Pending Cost Review", "Ready for Approval", "Approved", "Rejected",
    "Archived",
}
COMPATIBILITY_STATUSES = {
    "Pending Compatibility Review", "Device Family Only", "Device Specific",
    "Universal", "Not Applicable", "Approved", "Rejected",
}
SOURCING_STATUSES = {
    "Pending Sourcing Review", "Supplier Unknown", "Supplier Observed",
    "Multiple Suppliers", "Preferred Supplier Proposed", "Approved", "Rejected",
}
COST_STATUSES = {
    "Pending Cost Review", "Legacy Cost Only", "Supplier Cost Required",
    "Landed Cost Required", "Ready for Cost Approval", "Approved", "Rejected",
}
PRICING_STATUSES = {
    "Pending Pricing Review", "Legacy Price Review", "Not for Direct Sale",
    "Ready for Pricing Review", "Approved", "Rejected",
}
YES_NO_FIELDS = {
    "Active", "Preferred Supplier", "Inventory Tracked", "Serialized",
    "Warranty Eligible",
}
GENERATED_BLANK_FIELDS = {
    "Device ID", "Device Name", "Compatibility Scope", "Part Quality",
    "OEM Status", "Color", "Capacity", "Model Number", "Supplier ID",
    "Supplier Part Number", "Currency", "Default Warranty", "Reviewer",
    "Effective Date", "Last Reviewed",
}
DEFINED_NAME_BY_HEADER = {
    "Active": "DV_YesNo",
    "Status": "DV_PartStatuses",
    "Part Category ID": "DV_PartCategoryIDs",
    "Part Type ID": "DV_PartTypeIDs",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Device ID": "DV_DeviceIDs",
    "Part Quality": "DV_PartQualities",
    "Condition": "DV_Conditions",
    "Supplier ID": "DV_SupplierIDs",
    "Preferred Supplier": "DV_YesNo",
    "Cost Status": "DV_CostStatuses",
    "Pricing Status": "DV_PricingStatuses",
    "Inventory Tracked": "DV_YesNo",
    "Serialized": "DV_YesNo",
    "Warranty Eligible": "DV_YesNo",
    "Default Warranty": "DV_WarrantyOptions",
    "Compatibility Status": "DV_CompatibilityStatuses",
    "Sourcing Status": "DV_SourcingStatuses",
    "Review Status": "DV_ReviewStatuses",
}
DEFINED_NAME_SHEET = {
    "DV_PartStatuses": "08 - Part Qualities",
    "DV_ReviewStatuses": "08 - Part Qualities",
    "DV_CompatibilityStatuses": "08 - Part Qualities",
    "DV_SourcingStatuses": "08 - Part Qualities",
    "DV_CostStatuses": "08 - Part Qualities",
    "DV_PricingStatuses": "08 - Part Qualities",
    "DV_PartCategoryIDs": "02 - Part Categories",
    "DV_PartTypeIDs": "03 - Part Types",
    "DV_ManufacturerIDs": "04 - Manufacturers",
    "DV_DeviceFamilyCodes": "05 - Device Families",
    "DV_DeviceIDs": "06 - Device References",
    "DV_SupplierIDs": "07 - Suppliers",
    "DV_PartQualities": "08 - Part Qualities",
    "DV_Conditions": "09 - Conditions",
    "DV_WarrantyOptions": "08 - Part Qualities",
    "DV_YesNo": "08 - Part Qualities",
}
PROHIBITED_HEADERS = {
    "Final Cost", "Landed Cost", "Markup", "Margin", "Final Customer Price",
    "Stock", "Stock Quantity", "Bin", "Location", "Serial Number",
}


class PartsValidationError(RuntimeError):
    """Raised when the Master Parts review workbook violates its contract."""


def text(value: Any) -> str:
    """Normalize a scalar to stripped text."""
    return "" if value is None else str(value).strip()


def excel_safe_value(value: Any) -> Any:
    """Normalize temporal values using the generator persistence policy."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def assert_excel_safe_value_contract() -> None:
    """Exercise temporal normalization without workbook I/O."""
    aware_utc = datetime(2026, 7, 23, 12, 30, tzinfo=UTC)
    assert excel_safe_value(aware_utc) == datetime(2026, 7, 23, 12, 30)
    eastern = timezone(timedelta(hours=-4))
    aware_eastern = datetime(2026, 7, 23, 8, 30, tzinfo=eastern)
    assert excel_safe_value(aware_eastern) == datetime(2026, 7, 23, 12, 30)
    naive = datetime(2026, 7, 23, 12, 30)
    assert excel_safe_value(naive) is naive
    calendar_date = date(2026, 7, 23)
    assert excel_safe_value(calendar_date) is calendar_date
    assert excel_safe_value("") == ""
    assert excel_safe_value(None) is None
    stamp = "2026-07-23T12:30:00Z"
    assert excel_safe_value(stamp) == stamp


def persisted_equal(expected: Any, actual: Any) -> bool:
    """Compare source and Excel-persisted values type-aware."""
    expected = excel_safe_value(expected)
    actual = excel_safe_value(actual)
    if expected is None or text(expected) == "":
        return actual is None or text(actual) == ""
    temporal = (date, datetime, time)
    if isinstance(expected, temporal) or isinstance(actual, temporal):
        left = expected.isoformat() if isinstance(expected, temporal) else text(expected)
        right = actual.isoformat() if isinstance(actual, temporal) else text(actual)
        return left == right
    if isinstance(expected, (int, float, Decimal)) and not isinstance(expected, bool):
        try:
            return decimal_value(expected, "expected") == decimal_value(actual, "actual")
        except PartsValidationError:
            return False
    return text(expected) == text(actual)


def decimal_value(value: Any, field: str, *, allow_text: bool = False) -> Decimal | None:
    """Parse a finite Decimal while preserving blank and invalid legacy text."""
    if value is None or text(value) == "":
        return None
    if isinstance(value, bool):
        raise PartsValidationError(f"{field} contains Boolean value {value!r}")
    try:
        result = Decimal(text(value))
    except InvalidOperation:
        if allow_text:
            return None
        raise PartsValidationError(f"{field} is not numeric: {value!r}") from None
    if not result.is_finite():
        raise PartsValidationError(f"{field} is not finite: {value!r}")
    return result


def file_hash(path: Path) -> str:
    """Calculate SHA-256 without modifying a source."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require each protected workbook exactly once and as a valid ZIP."""
    seen: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved in seen:
            raise PartsValidationError(f"Protected input configured twice: {path}")
        seen.add(resolved)
        if not path.is_file():
            raise PartsValidationError(f"Required workbook does not exist: {path}")
        if path.stat().st_size <= 0:
            raise PartsValidationError(f"Required workbook is empty: {path}")
        if not zipfile.is_zipfile(path):
            raise PartsValidationError(f"Required workbook is not a valid Excel ZIP: {path}")


def require_generated_workbook(path: Path) -> None:
    """Reject missing or incomplete output before invoking openpyxl."""
    message = (
        "Generated Parts workbook is missing or invalid; "
        "rerun the generator successfully."
    )
    if (
        not path.is_file()
        or path.stat().st_size <= 0
        or not zipfile.is_zipfile(path)
    ):
        raise PartsValidationError(message)
    try:
        with zipfile.ZipFile(path) as archive:
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(archive.namelist()) or archive.testzip() is not None:
                raise PartsValidationError(message)
    except (OSError, zipfile.BadZipFile) as exc:
        raise PartsValidationError(message) from exc


def read_records(
    path: Path,
    sheet_name: str,
    *,
    header_label: str | None = None,
    keep_vba: bool = False,
) -> list[dict[str, Any]]:
    """Read a source dataset from a located header row."""
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_vba=keep_vba,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise PartsValidationError(f"{path.name} lacks {sheet_name!r}")
        worksheet = workbook[sheet_name]
        header_row = 1
        if header_label is not None:
            header_row = 0
            for row in worksheet.iter_rows():
                if any(text(cell.value) == header_label for cell in row):
                    header_row = row[0].row
                    break
            if header_row == 0:
                raise PartsValidationError(
                    f"{path.name}:{sheet_name} lacks {header_label!r}"
                )
        rows = worksheet.iter_rows(min_row=header_row, values_only=True)
        headers = [text(value) for value in next(rows)]
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    except StopIteration as exc:
        raise PartsValidationError(f"{path.name}:{sheet_name} is empty") from exc
    finally:
        workbook.close()


def read_existing_part_ids() -> tuple[set[str], list[str], list[str]]:
    """Independently inspect canonical Part ID columns."""
    workbook = load_workbook(
        CANONICAL_PATH,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )
    valid: set[str] = set()
    malformed: list[str] = []
    locations: list[str] = []
    try:
        for sheet_name in PART_ID_SHEETS:
            if sheet_name not in workbook.sheetnames:
                raise PartsValidationError(
                    f"Canonical workbook lacks {sheet_name!r}"
                )
            worksheet = workbook[sheet_name]
            headers = [
                cell
                for row in worksheet.iter_rows()
                for cell in row
                if text(cell.value) == "Part ID"
            ]
            if not headers:
                locations.append(f"{sheet_name}: no Part ID column")
                continue
            for header in headers:
                locations.append(
                    f"{sheet_name}!{get_column_letter(header.column)}{header.row}"
                )
                for row_number in range(header.row + 1, worksheet.max_row + 1):
                    value = text(
                        worksheet.cell(
                            row=row_number,
                            column=header.column,
                        ).value
                    )
                    if not value:
                        continue
                    if PART_ID_PATTERN.fullmatch(value):
                        if value in valid:
                            raise PartsValidationError(
                                f"Duplicate canonical Part ID: {value}"
                            )
                        valid.add(value)
                    else:
                        malformed.append(value)
        return valid, malformed, locations
    finally:
        workbook.close()


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read records from a named Excel Table."""
    if table_name not in worksheet.tables:
        raise PartsValidationError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = [text(value) for value in rows[0]]
    return [
        dict(zip(headers, row, strict=False))
        for row in rows[1:]
        if any(value is not None and text(value) != "" for value in row)
    ]


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Return exact headers from a named table."""
    min_col, min_row, max_col, _ = range_boundaries(
        worksheet.tables[table_name].ref
    )
    return [
        text(worksheet.cell(row=min_row, column=column).value)
        for column in range(min_col, max_col + 1)
    ]


def source_parts(
    retained: Sequence[dict[str, Any]],
    exclusions: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Recalculate the retained Part population and exclusions."""
    parts = [row for row in retained if text(row.get("Record Category")) == "Part"]
    if len(parts) != EXPECTED_PART_ROWS:
        raise PartsValidationError(
            f"Expected {EXPECTED_PART_ROWS} retained Parts; found {len(parts)}"
        )
    try:
        parts.sort(key=lambda row: int(text(row["Source Row Number"])))
    except (KeyError, ValueError) as exc:
        raise PartsValidationError("Invalid Part Source Row Number") from exc
    source_numbers = [text(row["Source Row Number"]) for row in parts]
    if len(source_numbers) != len(set(source_numbers)):
        raise PartsValidationError("Duplicate retained Part Source Row Number")
    excluded = {
        text(row.get("Excluded Source Row Number") or row.get("Source Row Number"))
        for row in exclusions
    }
    overlap = sorted(set(source_numbers).intersection(excluded))
    if overlap:
        raise PartsValidationError(f"Retained Parts include exclusions: {overlap}")
    return parts


def validate_structure(workbook: Any) -> None:
    """Validate the exact workbook, table, and primary-schema contract."""
    if workbook.sheetnames != SHEET_NAMES:
        raise PartsValidationError(
            f"Worksheet order mismatch: {workbook.sheetnames}"
        )
    if len(set(workbook.sheetnames)) != len(workbook.sheetnames) or any(
        len(name) > 31 for name in workbook.sheetnames
    ):
        raise PartsValidationError("Worksheet names must be unique and <= 31 chars")
    table_names: list[str] = []
    for name in SHEET_NAMES:
        worksheet = workbook[name]
        expected = TABLE_NAMES[name]
        if expected not in worksheet.tables:
            raise PartsValidationError(f"{name} lacks table {expected}")
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise PartsValidationError(f"{name} lacks frozen header or filter")
        if worksheet.merged_cells.ranges:
            raise PartsValidationError(f"{name} contains merged cells")
        table_names.extend(worksheet.tables)
    if len(table_names) != len(set(table_names)):
        raise PartsValidationError("Excel Table names are not unique")
    if table_headers(workbook["01 - Master Parts"], "tblMasterPartsCatalog") != PART_HEADERS:
        raise PartsValidationError("Master Parts schema is not exactly 50 columns")
    if PROHIBITED_HEADERS.intersection(PART_HEADERS):
        raise PartsValidationError("Master Parts schema contains prohibited fields")


def validate_names_and_validations(workbook: Any) -> None:
    """Require workbook-scoped names and named-list formulas only."""
    required_names = set(DEFINED_NAME_BY_HEADER.values())
    missing = required_names - set(workbook.defined_names)
    if missing:
        raise PartsValidationError(f"Missing defined names: {sorted(missing)}")
    for name in required_names:
        defined = workbook.defined_names[name]
        if defined.localSheetId is not None:
            raise PartsValidationError(f"Defined name is not workbook scoped: {name}")
        destinations = list(defined.destinations)
        if (
            len(destinations) != 1
            or destinations[0][0] != DEFINED_NAME_SHEET[name]
        ):
            raise PartsValidationError(f"Defined name has invalid target: {name}")
    worksheet = workbook["01 - Master Parts"]
    validations = [
        validation
        for validation in worksheet.data_validations.dataValidation
        if validation.type == "list"
    ]
    formulas = {
        text(validation.formula1)
        for validation in validations
    }
    expected = {f"={name}" for name in required_names}
    if formulas != expected:
        raise PartsValidationError(
            f"List validation formulas differ: {sorted(formulas)}"
        )
    if any("!" in formula or "," in formula for formula in formulas):
        raise PartsValidationError("Direct or hard-coded list validation found")
    headers = [text(cell.value) for cell in worksheet[1]]
    for header, name in DEFINED_NAME_BY_HEADER.items():
        column = get_column_letter(headers.index(header) + 1)
        expected_range = f"{column}2:{column}{worksheet.max_row}"
        if not any(
            text(validation.formula1) == f"={name}"
            and expected_range in str(validation.sqref).split()
            for validation in validations
        ):
            raise PartsValidationError(
                f"{header} does not use defined-name validation {name}"
            )


def canonical_lookups(
    workbook: Any,
) -> tuple[
    dict[str, str],
    dict[str, str],
    dict[str, tuple[str, str, str]],
    dict[str, str],
]:
    """Read generated lookup relationships from their tables."""
    manufacturers = {
        text(row["Manufacturer ID"]): text(row["Manufacturer Name"])
        for row in table_records(
            workbook["04 - Manufacturers"],
            "tblPartsManufacturers",
        )
        if text(row["Manufacturer ID"])
    }
    families = {
        text(row["Device Family Code"]): text(row["Device Family Name"])
        for row in table_records(
            workbook["05 - Device Families"],
            "tblPartsDeviceFamilies",
        )
        if text(row["Device Family Code"])
    }
    devices = {
        text(row["Device ID"]): (
            text(row["Device Name"]),
            text(row["Manufacturer ID"]),
            text(row["Device Family Code"]),
        )
        for row in table_records(
            workbook["06 - Device References"],
            "tblPartsDeviceReferences",
        )
        if text(row["Device ID"])
    }
    suppliers = {
        text(row["Supplier ID"]): text(row["Supplier Name"])
        for row in table_records(
            workbook["07 - Suppliers"],
            "tblPartsSuppliers",
        )
        if text(row["Supplier ID"])
    }
    return manufacturers, families, devices, suppliers


def validate_lookup_snapshots(
    workbook: Any,
    manufacturer_source: Sequence[dict[str, Any]],
    family_source: Sequence[dict[str, Any]],
    device_source: Sequence[dict[str, Any]],
    supplier_source: Sequence[dict[str, Any]],
    sources: Sequence[dict[str, Any]],
) -> None:
    """Reconcile generated lookup snapshots to authoritative sources."""
    manufacturer_rows = table_records(
        workbook["04 - Manufacturers"],
        "tblPartsManufacturers",
    )
    output_manufacturers = [
        (text(row["Manufacturer ID"]), text(row["Manufacturer Name"]))
        for row in manufacturer_rows
        if text(row["Manufacturer ID"])
    ]
    expected_manufacturers = [
        (text(row["Manufacturer ID"]), text(row["Manufacturer"]))
        for row in manufacturer_source
        if text(row.get("Manufacturer ID")) and text(row.get("Manufacturer"))
    ]
    if output_manufacturers != expected_manufacturers or len(
        {identifier for identifier, _ in output_manufacturers}
    ) != len(output_manufacturers):
        raise PartsValidationError("Manufacturer lookup differs from canonical source")

    family_rows = table_records(
        workbook["05 - Device Families"],
        "tblPartsDeviceFamilies",
    )
    output_families = [
        (text(row["Device Family Code"]), text(row["Device Family Name"]))
        for row in family_rows
        if text(row["Device Family Code"])
    ]
    expected_families = [
        (text(row["Device Family Code"]), text(row["Device Family"]))
        for row in family_source
        if text(row.get("Device Family Code")) and text(row.get("Device Family"))
    ]
    if output_families != expected_families or len(
        {identifier for identifier, _ in output_families}
    ) != len(output_families):
        raise PartsValidationError("Device Family lookup differs from canonical source")

    device_rows = table_records(
        workbook["06 - Device References"],
        "tblPartsDeviceReferences",
    )
    output_devices = [
        (
            text(row["Device ID"]),
            text(row["Device Name"]),
            text(row["Manufacturer ID"]),
            text(row["Device Family Code"]),
            text(row["Model Number"]),
        )
        for row in device_rows
        if text(row["Device ID"])
    ]
    expected_devices = [
        (
            text(row["Device ID"]),
            text(row["Device Model"]),
            text(row["Manufacturer Code"]),
            text(row["Device Family Code"]),
            text(row["Model Number"]),
        )
        for row in device_source
        if text(row.get("Device ID")) and text(row.get("Device Model"))
    ]
    if output_devices != expected_devices or len(
        {identifier for identifier, *_ in output_devices}
    ) != len(output_devices):
        raise PartsValidationError("Device lookup differs from canonical source")

    supplier_rows = table_records(
        workbook["07 - Suppliers"],
        "tblPartsSuppliers",
    )
    if any(text(row["Supplier ID"]) for row in supplier_rows):
        raise PartsValidationError("Supplier IDs were invented")
    output_supplier_names = [
        text(row["Supplier Name"])
        for row in supplier_rows
        if text(row["Supplier Name"])
    ]
    expected_supplier_names = [
        text(row["Supplier"])
        for row in supplier_source
        if text(row.get("Supplier"))
    ]
    if output_supplier_names != expected_supplier_names:
        raise PartsValidationError("Supplier lookup differs from canonical source")

    category_rows = table_records(
        workbook["02 - Part Categories"],
        "tblPartCategories",
    )
    expected_categories = sorted(
        {text(row["Legacy Group"]) for row in sources if text(row["Legacy Group"])},
        key=str.casefold,
    )
    output_categories = [
        (text(row["Part Category ID"]), text(row["Part Category"]))
        for row in category_rows
    ]
    expected_category_rows = [
        (f"PC-{index:04d}", name)
        for index, name in enumerate(expected_categories, 1)
    ]
    if output_categories != expected_category_rows:
        raise PartsValidationError("Proposed Part Category lookup is not deterministic")

    type_rows = table_records(
        workbook["03 - Part Types"],
        "tblPartTypes",
    )
    expected_types = sorted(
        {text(row["Legacy Type"]) for row in sources if text(row["Legacy Type"])},
        key=str.casefold,
    )
    output_types = [
        (text(row["Part Type ID"]), text(row["Part Type"]))
        for row in type_rows
    ]
    expected_type_rows = [
        (f"PT-{index:04d}", name)
        for index, name in enumerate(expected_types, 1)
    ]
    if output_types != expected_type_rows:
        raise PartsValidationError("Proposed Part Type lookup is not deterministic")

    condition_rows = table_records(
        workbook["09 - Conditions"],
        "tblPartConditions",
    )
    output_conditions = [text(row["Condition"]) for row in condition_rows]
    expected_conditions = sorted(
        {
            text(row["Legacy Condition"])
            for row in sources
            if text(row["Legacy Condition"])
        },
        key=str.casefold,
    )
    if output_conditions != expected_conditions:
        raise PartsValidationError("Condition lookup differs from source observations")


def validate_identity(
    parts: Sequence[dict[str, Any]],
    sources: Sequence[dict[str, Any]],
    existing: set[str],
) -> None:
    """Validate population, Part IDs, and deterministic source ordering."""
    if len(parts) != EXPECTED_PART_ROWS or len(sources) != EXPECTED_PART_ROWS:
        raise PartsValidationError("Master Parts or source count is not 48")
    identifiers = [text(row["Part ID"]) for row in parts]
    if any(not PART_ID_PATTERN.fullmatch(value) for value in identifiers):
        raise PartsValidationError("Invalid generated Part ID")
    if len(identifiers) != len(set(identifiers)):
        raise PartsValidationError("Duplicate generated Part ID")
    if existing.intersection(identifiers):
        raise PartsValidationError("Generated Part ID overlaps canonical Part ID")
    expected_first = (
        max(int(value[3:]) for value in existing) + 1 if existing else 1
    )
    numbers = [int(value[3:]) for value in identifiers]
    if numbers != list(range(expected_first, expected_first + len(parts))):
        raise PartsValidationError("Part IDs are not the next continuous sequence")
    if not existing and identifiers[0] != "PRT000001":
        raise PartsValidationError("ADR-008 empty namespace must start PRT000001")
    if numbers[-1] != numbers[0] + len(parts) - 1:
        raise PartsValidationError("Final Part ID does not reconcile to row count")
    if not existing and (identifiers[0], identifiers[-1]) != (
        "PRT000001",
        "PRT000048",
    ):
        raise PartsValidationError(
            "Current empty-namespace range must be PRT000001 through PRT000048"
        )
    output_sources = [int(text(row["Source Record Number"])) for row in parts]
    expected_sources = [int(text(row["Source Row Number"])) for row in sources]
    if output_sources != expected_sources or output_sources != sorted(output_sources):
        raise PartsValidationError("Part source rows are missing, extra, or reordered")
    if len(output_sources) != len(set(output_sources)):
        raise PartsValidationError("Duplicate output Source Record Number")


def validate_source_and_status(
    parts: Sequence[dict[str, Any]],
    sources: Sequence[dict[str, Any]],
    manufacturers: dict[str, str],
    families: dict[str, str],
    devices: dict[str, tuple[str, str, str]],
    suppliers: dict[str, str],
    authoritative_manufacturers: dict[str, tuple[str, str]],
    authoritative_families: dict[str, tuple[str, str]],
) -> None:
    """Validate observations, relationships, blank rules, and statuses."""
    source_map = {
        "Legacy Part SKU": "Legacy SKU",
        "Part Name": "Legacy Name",
        "Part Display Name": "Legacy Name",
        "Part Description": "Legacy Note",
        "Part Category": "Legacy Group",
        "Part Type": "Legacy Type",
        "Manufacturer Name": "Legacy Manufacturer",
        "Condition": "Legacy Condition",
        "Supplier Name": "Legacy Supplier",
        "Legacy Retail Price": "Legacy Retail Price",
        "Legacy Cost": "Legacy Cost",
        "Reviewer Notes": "Legacy Note",
        "Source Record Number": "Source Row Number",
        "Source Workbook": "Source Workbook",
        "Source Worksheet": "Source Worksheet",
        "Created At": "Source Created At",
        "Updated At": "Source Updated At",
    }
    for row_number, (part, source) in enumerate(zip(parts, sources, strict=True), 2):
        part_id = text(part["Part ID"])
        for output_field, source_field in source_map.items():
            expected = source.get(source_field)
            if not persisted_equal(expected, part[output_field]):
                raise PartsValidationError(
                    f"Row {row_number} {part_id} changed {output_field}: "
                    f"{expected!r} -> {part[output_field]!r}"
                )
        manufacturer_id = text(part["Manufacturer ID"])
        manufacturer_observation = text(source["Legacy Manufacturer"])
        expected_manufacturer_id = ""
        manufacturer_candidate = authoritative_manufacturers.get(
            manufacturer_observation.casefold()
        )
        if (
            manufacturer_candidate is not None
            and manufacturer_candidate[1] == manufacturer_observation
        ):
            expected_manufacturer_id = manufacturer_candidate[0]
        if manufacturer_id != expected_manufacturer_id:
            raise PartsValidationError(
                f"{part_id} manufacturer mapping is not source-supported"
            )
        if manufacturer_id and manufacturers.get(manufacturer_id) != text(
            part["Manufacturer Name"]
        ):
            raise PartsValidationError(f"{part_id} manufacturer ID/name mismatch")
        family_code = text(part["Device Family Code"])
        legacy_type = text(source["Legacy Type"])
        expected_family = ("", "")
        if legacy_type.casefold().startswith("part - "):
            candidate = legacy_type.split("-", 1)[1].strip().casefold()
            if candidate not in {"", "other"}:
                expected_family = authoritative_families.get(candidate, ("", ""))
        if (family_code, text(part["Device Family Name"])) != expected_family:
            raise PartsValidationError(
                f"{part_id} device-family mapping is not source-supported"
            )
        if family_code and families.get(family_code) != text(part["Device Family Name"]):
            raise PartsValidationError(f"{part_id} device-family mismatch")
        device_id = text(part["Device ID"])
        if device_id:
            expected_device = devices.get(device_id)
            if expected_device is None or expected_device[0] != text(part["Device Name"]):
                raise PartsValidationError(f"{part_id} invalid Device ID/name")
            if family_code and expected_device[2] != family_code:
                raise PartsValidationError(f"{part_id} Device ID/family mismatch")
            if manufacturer_id and expected_device[1] != manufacturer_id:
                raise PartsValidationError(f"{part_id} Device ID/manufacturer mismatch")
        supplier_id = text(part["Supplier ID"])
        if supplier_id and suppliers.get(supplier_id) != text(part["Supplier Name"]):
            raise PartsValidationError(f"{part_id} supplier ID/name mismatch")
        for field in GENERATED_BLANK_FIELDS:
            if text(part[field]):
                raise PartsValidationError(
                    f"{part_id} generated unsupported {field}: {part[field]!r}"
                )
        for field in ("Legacy Retail Price", "Legacy Cost"):
            amount = decimal_value(part[field], field, allow_text=True)
            if amount is not None and amount < 0:
                raise PartsValidationError(f"{part_id} has negative {field}")
        if text(part["Status"]) != "Draft":
            raise PartsValidationError(f"{part_id} is not Draft")
        if text(part["Active"]) != "Yes":
            raise PartsValidationError(f"{part_id} Active default changed")
        if text(part["Preferred Supplier"]) != "No":
            raise PartsValidationError(f"{part_id} preferred supplier was inferred")
        if text(part["Inventory Tracked"]) != "Yes":
            raise PartsValidationError(f"{part_id} inventory review default changed")
        if text(part["Serialized"]) != "No" or text(part["Warranty Eligible"]) != "No":
            raise PartsValidationError(f"{part_id} unsupported flag was inferred")
        if text(part["Status"]) not in PART_STATUSES:
            raise PartsValidationError(f"{part_id} uncontrolled Part Status")
        if text(part["Review Status"]) not in REVIEW_STATUSES:
            raise PartsValidationError(f"{part_id} uncontrolled Review Status")
        if text(part["Compatibility Status"]) not in COMPATIBILITY_STATUSES:
            raise PartsValidationError(f"{part_id} uncontrolled compatibility")
        if text(part["Sourcing Status"]) not in SOURCING_STATUSES:
            raise PartsValidationError(f"{part_id} uncontrolled sourcing")
        if text(part["Cost Status"]) not in COST_STATUSES:
            raise PartsValidationError(f"{part_id} uncontrolled cost status")
        if text(part["Pricing Status"]) not in PRICING_STATUSES:
            raise PartsValidationError(f"{part_id} uncontrolled pricing status")
        governed = (
            text(part["Status"]),
            text(part["Review Status"]),
            text(part["Compatibility Status"]),
            text(part["Sourcing Status"]),
            text(part["Cost Status"]),
            text(part["Pricing Status"]),
        )
        if "Approved" in governed or "Ready for Approval" in governed:
            raise PartsValidationError(f"{part_id} was pre-approved")
        if not manufacturer_id and text(part["Review Status"]) != (
            "Pending Manufacturer Review"
        ):
            raise PartsValidationError(f"{part_id} missing manufacturer review")
        if not family_code and manufacturer_id and text(part["Review Status"]) != (
            "Pending Device Mapping"
        ):
            raise PartsValidationError(f"{part_id} missing device mapping review")
        expected_compatibility = (
            "Device Family Only" if family_code else "Pending Compatibility Review"
        )
        if text(part["Compatibility Status"]) != expected_compatibility:
            raise PartsValidationError(f"{part_id} compatibility status mismatch")
        expected_sourcing = (
            "Supplier Observed"
            if text(part["Supplier Name"])
            else "Supplier Unknown"
        )
        if text(part["Sourcing Status"]) != expected_sourcing:
            raise PartsValidationError(f"{part_id} sourcing status mismatch")
        expected_cost_status = (
            "Pending Cost Review"
            if source["Legacy Cost"] is None or text(source["Legacy Cost"]) == ""
            else "Legacy Cost Only"
        )
        if text(part["Cost Status"]) != expected_cost_status:
            raise PartsValidationError(f"{part_id} cost status mismatch")
        retail_amount = decimal_value(
            source["Legacy Retail Price"],
            "Legacy Retail Price",
            allow_text=True,
        )
        expected_pricing_status = (
            "Legacy Price Review"
            if retail_amount is not None and retail_amount > 0
            else "Pending Pricing Review"
        )
        if text(part["Pricing Status"]) != expected_pricing_status:
            raise PartsValidationError(f"{part_id} pricing status mismatch")
        if text(part["Currency"]):
            raise PartsValidationError(f"{part_id} defaulted currency")
        if text(part["Import Batch ID"]) != IMPORT_BATCH_ID:
            raise PartsValidationError(f"{part_id} wrong Import Batch ID")
        if any(text(part[field]) not in {"Yes", "No"} for field in YES_NO_FIELDS):
            raise PartsValidationError(f"{part_id} has uncontrolled Yes/No value")


def validate_review_tables(
    workbook: Any,
    parts: Sequence[dict[str, Any]],
) -> None:
    """Validate compatibility and sourcing queues row-for-row."""
    compatibility_sheet = workbook["10 - Compatibility Review"]
    sourcing_sheet = workbook["11 - Sourcing Review"]
    if table_headers(
        compatibility_sheet,
        "tblPartsCompatibilityReview",
    ) != COMPATIBILITY_HEADERS:
        raise PartsValidationError("Compatibility Review schema changed")
    if table_headers(
        sourcing_sheet,
        "tblPartsSourcingReview",
    ) != SOURCING_HEADERS:
        raise PartsValidationError("Sourcing Review schema changed")
    compatibility = table_records(
        compatibility_sheet,
        "tblPartsCompatibilityReview",
    )
    sourcing = table_records(sourcing_sheet, "tblPartsSourcingReview")
    if len(compatibility) != len(parts) or len(sourcing) != len(parts):
        raise PartsValidationError("Review table row count differs from Master Parts")
    for part, compatibility_row, sourcing_row in zip(
        parts,
        compatibility,
        sourcing,
        strict=True,
    ):
        for review_row, fields in (
            (
                compatibility_row,
                ("Part ID", "Legacy Part SKU", "Part Name", "Manufacturer Name"),
            ),
            (
                sourcing_row,
                ("Part ID", "Legacy Part SKU", "Part Name", "Supplier ID", "Supplier Name", "Legacy Cost"),
            ),
        ):
            for field in fields:
                if not persisted_equal(part[field], review_row[field]):
                    raise PartsValidationError(
                        f"Review table changed {field} for {part['Part ID']}"
                    )
        if not text(compatibility_row["Missing Compatibility Inputs"]) or not text(
            compatibility_row["Required Action"]
        ):
            raise PartsValidationError(
                f"Compatibility routing missing for {part['Part ID']}"
            )
        if text(compatibility_row["Review Status"]) != "Pending Compatibility Review":
            raise PartsValidationError(
                f"Compatibility row pre-reviewed for {part['Part ID']}"
            )
        if not text(sourcing_row["Missing Sourcing Inputs"]) or not text(
            sourcing_row["Required Action"]
        ):
            raise PartsValidationError(
                f"Sourcing routing missing for {part['Part ID']}"
            )
        if text(sourcing_row["Review Status"]) != "Pending Sourcing Review":
            raise PartsValidationError(
                f"Sourcing row pre-reviewed for {part['Part ID']}"
            )


def validate_metadata(
    workbook: Any,
    hashes: dict[Path, str],
    existing: set[str],
    malformed: Sequence[str],
    locations: Sequence[str],
    parts: Sequence[dict[str, Any]],
) -> None:
    """Validate embedded identity, schema, and protected-input metadata."""
    rows = table_records(
        workbook["14 - Import Metadata"],
        "tblMasterPartsImportMetadata",
    )
    metadata = {text(row["Metadata Field"]): text(row["Value"]) for row in rows}
    expected = {
        "Import Batch ID": IMPORT_BATCH_ID,
        "Schema Columns": str(len(PART_HEADERS)),
        "Namespace Authority": NAMESPACE_AUTHORITY,
        "Part ID Sources": "; ".join(locations),
        "Existing Part ID Count": str(len(existing)),
        "First Generated Part ID": text(parts[0]["Part ID"]),
        "Final Generated Part ID": text(parts[-1]["Part ID"]),
        "Generated Part Count": str(len(parts)),
        "Malformed Existing Part IDs": "; ".join(malformed),
    }
    for field, value in expected.items():
        if metadata.get(field) != value:
            raise PartsValidationError(
                f"Import Metadata {field!r} differs: {metadata.get(field)!r}"
            )
    for path, digest in hashes.items():
        if metadata.get(f"Protected Input Path: {path.name}") != str(path):
            raise PartsValidationError(f"Missing protected path for {path}")
        if metadata.get(f"SHA-256: {path.name}") != digest:
            raise PartsValidationError(f"Protected hash differs for {path}")


def main() -> int:
    """Run independent validation without writing any workbook."""
    protected = (
        PROPOSAL_PATH,
        MASTER_SERVICES_PATH,
        MASTER_PRICING_PATH,
        CANONICAL_PATH,
    )
    try:
        assert_excel_safe_value_contract()
        require_files(protected)
        require_generated_workbook(OUTPUT_PATH)
        hashes = {path: file_hash(path) for path in protected}
        retained = read_records(PROPOSAL_PATH, RETAINED_SHEET)
        exclusions = read_records(PROPOSAL_PATH, EXCLUSIONS_SHEET)
        sources = source_parts(retained, exclusions)
        existing, malformed, locations = read_existing_part_ids()
        manufacturer_source = read_records(
            CANONICAL_PATH,
            "30 Manufacturers",
            header_label="Manufacturer ID",
            keep_vba=True,
        )
        family_source = read_records(
            CANONICAL_PATH,
            "31 Device Families",
            header_label="Device Family Code",
            keep_vba=True,
        )
        device_source = read_records(
            CANONICAL_PATH,
            "32 Devices",
            header_label="Device ID",
            keep_vba=True,
        )
        supplier_source = read_records(
            CANONICAL_PATH,
            "43 Suppliers",
            header_label="Supplier",
            keep_vba=True,
        )
        workbook = load_workbook(OUTPUT_PATH, data_only=False)
        try:
            validate_structure(workbook)
            validate_names_and_validations(workbook)
            parts = table_records(
                workbook["01 - Master Parts"],
                "tblMasterPartsCatalog",
            )
            validate_identity(parts, sources, existing)
            validate_lookup_snapshots(
                workbook,
                manufacturer_source,
                family_source,
                device_source,
                supplier_source,
                sources,
            )
            manufacturers, families, devices, suppliers = canonical_lookups(workbook)
            authoritative_manufacturers = {
                text(row["Manufacturer"]).casefold(): (
                    text(row["Manufacturer ID"]),
                    text(row["Manufacturer"]),
                )
                for row in manufacturer_source
                if text(row.get("Manufacturer ID")) and text(row.get("Manufacturer"))
            }
            authoritative_families = {
                text(row["Device Family"]).casefold(): (
                    text(row["Device Family Code"]),
                    text(row["Device Family"]),
                )
                for row in family_source
                if text(row.get("Device Family Code"))
                and text(row.get("Device Family"))
            }
            validate_source_and_status(
                parts,
                sources,
                manufacturers,
                families,
                devices,
                suppliers,
                authoritative_manufacturers,
                authoritative_families,
            )
            validate_review_tables(workbook, parts)
            validate_metadata(
                workbook,
                hashes,
                existing,
                malformed,
                locations,
                parts,
            )
        finally:
            workbook.close()
        after_hashes = {path: file_hash(path) for path in protected}
        if after_hashes != hashes:
            raise PartsValidationError("A protected input changed during validation")
    except (
        AssertionError,
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        PartsValidationError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Validated: {OUTPUT_PATH}")
    print(f"Namespace authority: {NAMESPACE_AUTHORITY}")
    print(f"Retained Part rows: {len(parts)}")
    print(f"Part ID range: {parts[0]['Part ID']} through {parts[-1]['Part ID']}")
    print(f"Malformed canonical Part IDs: {len(malformed)}")
    print("Workbook structure, tables, and defined names: PASS")
    print("Source preservation and review queues: PASS")
    print("Protected input hashes: PASS")
    print("Inventory/canonical/pricing writes: NOT PERFORMED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
