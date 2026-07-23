"""Generate the governed Service-to-Labor mapping review workbook.

The output contains one pending-review row per Service. It never approves a
mapping and never saves or changes Master Services or Master Labor.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
MASTER_SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
MASTER_LABOR_PATH = WORKING_DIR / "Nocturnix_Master_Labor_Catalog_v1.xlsx"
OUTPUT_PATH = WORKING_DIR / "Labor_Mapping_Review_v1.xlsx"
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(f".{OUTPUT_PATH.stem}.tmp.xlsx")
SERVICES_SHEET = "01 - Master Services"
LABOR_SHEET = "01 - Labor Standards"
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
LABOR_ID_PATTERN = re.compile(r"^LAB\d{6}$")
IMPORT_BATCH = "LABOR-MAPPING-V1-REVIEW"

REVIEW_HEADERS = [
    "Service ID",
    "Service Name",
    "Current Labor Standard",
    "Suggested Labor Standard",
    "Confidence",
    "Evidence",
    "Match Score",
    "Margin",
    "Review Status",
]
SHEET_NAMES = [
    "00 - Summary",
    "01 - Mapping Review",
    "02 - Controls",
    "03 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Summary": "tblLaborMappingSummary",
    "01 - Mapping Review": "tblLaborMappingReview",
    "02 - Controls": "tblLaborMappingControls",
    "03 - Import Metadata": "tblLaborMappingImportMetadata",
}
CONFIDENCE_VALUES = ["Unassessed", "Low", "Medium", "High"]
REVIEW_STATUS_VALUES = ["Pending Review", "Approved", "Rejected", "Archived"]
DEFINED_NAME_SPECS = {
    "DV_LaborMappingConfidence": ("02 - Controls", 1),
    "DV_LaborMappingReviewStatus": ("02 - Controls", 2),
}


class MappingReviewError(RuntimeError):
    """Raised when a mapping-review invariant fails."""


def text(value: Any) -> str:
    """Return trimmed text."""
    return "" if value is None else str(value).strip()


def normalized(value: Any) -> str:
    """Normalize explicit text for exact comparison only."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).casefold()))


def file_hash(path: Path) -> str:
    """Return a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require every protected input."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise MappingReviewError(f"Missing protected input: {', '.join(missing)}")


def locate_header(
    worksheet: Worksheet,
    required: set[str],
) -> tuple[int, list[str]]:
    """Locate a header row containing the required fields."""
    for row_number in range(1, min(worksheet.max_row, 50) + 1):
        headers = [text(cell.value) for cell in worksheet[row_number]]
        if required <= set(headers):
            return row_number, headers
    raise MappingReviewError(
        f"{worksheet.title} lacks required headers: {sorted(required)}"
    )


def read_records(
    path: Path,
    sheet_name: str,
    required: set[str],
) -> list[dict[str, Any]]:
    """Read populated worksheet records without saving the workbook."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise MappingReviewError(f"{path.name} lacks {sheet_name}")
        worksheet = workbook[sheet_name]
        header_row, headers = locate_header(worksheet, required)
        records = []
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            values = [
                worksheet.cell(row_number, column).value
                for column in range(1, len(headers) + 1)
            ]
            if not any(text(value) for value in values):
                continue
            record = dict(zip(headers, values, strict=True))
            record["_Source Row"] = row_number
            records.append(record)
        return records
    finally:
        workbook.close()


def confidence(score: float, margin: float, suggested: str) -> str:
    """Classify explicit-match evidence without implying approval."""
    if not suggested:
        return "Unassessed"
    if score == 1.0 and margin == 1.0:
        return "High"
    if score == 1.0 and margin > 0:
        return "Medium"
    return "Low"


def candidate_scores(
    service: dict[str, Any],
    labor_rows: Sequence[dict[str, Any]],
) -> list[tuple[float, str, str]]:
    """Return only direct-ID or exact-name candidates.

    No fuzzy match, duration inference, tier inference, or taxonomy inference is
    permitted. The returned candidate remains a review suggestion.
    """
    current = text(service.get("Labor Standard ID"))
    service_name = normalized(service.get("Service Name"))
    candidates: list[tuple[float, str, str]] = []
    for labor in labor_rows:
        labor_id = text(labor.get("Labor Standard ID"))
        legacy_id = text(labor.get("Legacy Labor ID"))
        labor_names = {
            normalized(labor.get("Labor Name")),
            normalized(labor.get("Repair Type")),
        } - {""}
        if current and current in {labor_id, legacy_id}:
            identity_kind = (
                "governed Labor Standard ID"
                if current == labor_id
                else f"Legacy Labor ID {legacy_id}"
            )
            candidates.append(
                (
                    1.0,
                    labor_id,
                    "Current Labor Standard resolves through "
                    f"{identity_kind} to governed ID {labor_id}.",
                )
            )
        elif service_name and service_name in labor_names:
            candidates.append(
                (
                    1.0,
                    labor_id,
                    "Exact normalized Service Name equals Labor Name or Repair Type.",
                )
            )
    candidates.sort(key=lambda item: (-item[0], item[1]))
    return candidates


def mapping_row(
    service: dict[str, Any],
    labor_rows: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Create one pending-review mapping record."""
    candidates = candidate_scores(service, labor_rows)
    if not candidates:
        suggested = ""
        score = 0.0
        margin = 0.0
        evidence = "No direct ID or exact-name evidence; no suggestion created."
    else:
        best_score, best_id, best_evidence = candidates[0]
        second_score = candidates[1][0] if len(candidates) > 1 else 0.0
        score = best_score
        margin = best_score - second_score
        unique_best = [
            candidate
            for candidate in candidates
            if candidate[0] == best_score
        ]
        if len(unique_best) == 1:
            suggested = best_id
            evidence = best_evidence
        else:
            suggested = ""
            evidence = (
                "Multiple equally supported exact candidates; "
                "no suggestion created."
            )
    return {
        "Service ID": text(service.get("Service ID")),
        "Service Name": text(service.get("Service Name")),
        "Current Labor Standard": text(service.get("Labor Standard ID")),
        "Suggested Labor Standard": suggested,
        "Confidence": confidence(score, margin, suggested),
        "Evidence": evidence,
        "Match Score": score,
        "Margin": margin,
        "Review Status": "Pending Review",
    }


def validate_source_identity(
    services: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
) -> None:
    """Require unique valid source identities."""
    service_ids = [text(record.get("Service ID")) for record in services]
    if any(not SERVICE_ID_PATTERN.fullmatch(value) for value in service_ids):
        raise MappingReviewError("Master Services contains malformed Service IDs")
    if len(service_ids) != len(set(service_ids)):
        raise MappingReviewError("Master Services contains duplicate Service IDs")
    labor_ids = [text(record.get("Labor Standard ID")) for record in labor_rows]
    if any(not LABOR_ID_PATTERN.fullmatch(value) for value in labor_ids):
        raise MappingReviewError("Master Labor contains malformed Labor IDs")
    if len(labor_ids) != len(set(labor_ids)):
        raise MappingReviewError("Master Labor contains duplicate Labor IDs")
    legacy_ids = [text(record.get("Legacy Labor ID")) for record in labor_rows]
    if any(not value for value in legacy_ids):
        raise MappingReviewError("Master Labor contains blank Legacy Labor IDs")
    if len(legacy_ids) != len(set(legacy_ids)):
        raise MappingReviewError("Master Labor contains duplicate Legacy Labor IDs")


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write one formatted Excel Table."""
    worksheet.append(list(headers))
    materialized = list(rows) or [{header: "" for header in headers}]
    for record in materialized:
        worksheet.append([record.get(header, "") for header in headers])
    end_column = get_column_letter(len(headers))
    table = Table(
        displayName=table_name,
        ref=f"A1:{end_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = table.ref
    for column, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(len(header) + 2, 12),
            58,
        )
        if header in {"Match Score", "Margin"}:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row_number, column).number_format = "0.0000"


def aligned_controls() -> list[dict[str, str]]:
    """Align confidence and review-status lists."""
    row_count = max(len(CONFIDENCE_VALUES), len(REVIEW_STATUS_VALUES))
    return [
        {
            "Confidence": (
                CONFIDENCE_VALUES[index]
                if index < len(CONFIDENCE_VALUES)
                else ""
            ),
            "Review Status": (
                REVIEW_STATUS_VALUES[index]
                if index < len(REVIEW_STATUS_VALUES)
                else ""
            ),
        }
        for index in range(row_count)
    ]


def add_defined_names_and_validations(workbook: Workbook) -> None:
    """Use workbook defined names for every list validation."""
    for name, (sheet_name, column) in DEFINED_NAME_SPECS.items():
        worksheet = workbook[sheet_name]
        letter = get_column_letter(column)
        last_row = max(
            row
            for row in range(2, worksheet.max_row + 1)
            if text(worksheet.cell(row, column).value)
        )
        workbook.defined_names.add(
            DefinedName(
                name,
                attr_text=(
                    f"'{sheet_name}'!${letter}$2:${letter}${last_row}"
                ),
            )
        )
    review = workbook["01 - Mapping Review"]
    formulas = {
        "Confidence": "DV_LaborMappingConfidence",
        "Review Status": "DV_LaborMappingReviewStatus",
    }
    for header, defined_name in formulas.items():
        column = REVIEW_HEADERS.index(header) + 1
        letter = get_column_letter(column)
        validation = DataValidation(
            type="list",
            formula1=f"={defined_name}",
            allow_blank=False,
        )
        review.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{review.max_row}")


def build_workbook(
    review_rows: Sequence[dict[str, Any]],
    hashes: dict[Path, str],
    generated_at: datetime,
) -> Workbook:
    """Build the standalone mapping-review artifact."""
    workbook = Workbook()
    workbook.active.title = SHEET_NAMES[0]
    for sheet_name in SHEET_NAMES[1:]:
        workbook.create_sheet(sheet_name)
    suggestions = sum(
        1 for record in review_rows if text(record["Suggested Labor Standard"])
    )
    append_table(
        workbook["00 - Summary"],
        ["Metric", "Value"],
        [
            {"Metric": "Master Services", "Value": len(review_rows)},
            {"Metric": "Review Suggestions", "Value": suggestions},
            {"Metric": "Approved Mappings", "Value": 0},
            {"Metric": "Master Services Modified", "Value": "No"},
            {"Metric": "Automatic Approval", "Value": "No"},
        ],
        TABLE_NAMES["00 - Summary"],
    )
    append_table(
        workbook["01 - Mapping Review"],
        REVIEW_HEADERS,
        review_rows,
        TABLE_NAMES["01 - Mapping Review"],
    )
    append_table(
        workbook["02 - Controls"],
        ["Confidence", "Review Status"],
        aligned_controls(),
        TABLE_NAMES["02 - Controls"],
    )
    metadata = [
        {"Metadata Field": "Import Batch", "Value": IMPORT_BATCH},
        {"Metadata Field": "Generated At UTC", "Value": generated_at},
        {"Metadata Field": "Service Count", "Value": len(review_rows)},
        {"Metadata Field": "Automatic Approval", "Value": "No"},
        {"Metadata Field": "Master Services Modified", "Value": "No"},
    ]
    for path, digest in hashes.items():
        metadata.extend(
            [
                {
                    "Metadata Field": f"Protected Input Path: {path.name}",
                    "Value": str(path),
                },
                {
                    "Metadata Field": f"SHA-256: {path.name}",
                    "Value": digest,
                },
            ]
        )
    append_table(
        workbook["03 - Import Metadata"],
        ["Metadata Field", "Value"],
        metadata,
        TABLE_NAMES["03 - Import Metadata"],
    )
    add_defined_names_and_validations(workbook)
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read nonblank records from a table."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    records = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def require_ooxml(path: Path) -> None:
    """Require a safe OOXML workbook."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        members = set(archive.namelist())
        corrupt_member = archive.testzip()
    if required - members or corrupt_member:
        raise MappingReviewError("Mapping review OOXML package is incomplete")
    if "xl/vbaProject.bin" in members:
        raise MappingReviewError("Mapping review unexpectedly contains macros")
    if any(member.startswith("xl/externalLinks/") for member in members):
        raise MappingReviewError("Mapping review contains external links")


def validate_reopened(
    path: Path,
    services: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
    expected: Sequence[dict[str, Any]],
) -> None:
    """Reopen and validate coverage, status, suggestions, names, and tables."""
    require_ooxml(path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise MappingReviewError("Reopened worksheet contract differs")
        for sheet_name in SHEET_NAMES:
            worksheet = workbook[sheet_name]
            if list(worksheet.tables) != [TABLE_NAMES[sheet_name]]:
                raise MappingReviewError(f"Table contract differs: {sheet_name}")
            if worksheet.freeze_panes != "A2":
                raise MappingReviewError(f"Frozen header missing: {sheet_name}")
        if set(workbook.defined_names) != set(DEFINED_NAME_SPECS):
            raise MappingReviewError("Defined-name contract differs")
        worksheet = workbook["01 - Mapping Review"]
        headers = [text(cell.value) for cell in worksheet[1]]
        if headers != REVIEW_HEADERS:
            raise MappingReviewError("Mapping review schema differs")
        actual = table_records(worksheet, TABLE_NAMES[worksheet.title])
        if len(actual) != len(services) or len(actual) != len(expected):
            raise MappingReviewError("Mapping review does not have one row per Service")
        labor_ids = {text(record["Labor Standard ID"]) for record in labor_rows}
        for row_number, (actual_row, service) in enumerate(
            zip(actual, services, strict=True),
            start=2,
        ):
            if text(actual_row["Service ID"]) != text(service["Service ID"]):
                raise MappingReviewError(
                    f"Service ID/order differs at row {row_number}"
                )
            if text(actual_row["Service Name"]) != text(service["Service Name"]):
                raise MappingReviewError(f"Service Name differs at row {row_number}")
            if text(actual_row["Current Labor Standard"]) != text(
                service.get("Labor Standard ID")
            ):
                raise MappingReviewError(
                    f"Current Labor Standard differs at row {row_number}"
                )
            suggested = text(actual_row["Suggested Labor Standard"])
            if suggested and suggested not in labor_ids:
                raise MappingReviewError(
                    f"Unknown suggestion at row {row_number}: {suggested}"
                )
            if suggested and not LABOR_ID_PATTERN.fullmatch(suggested):
                raise MappingReviewError(
                    f"Suggestion is not a governed LAB ID at row {row_number}"
                )
            if text(actual_row["Review Status"]) != "Pending Review":
                raise MappingReviewError(
                    f"Automatic approval/status found at row {row_number}"
                )
            score = float(actual_row["Match Score"])
            margin = float(actual_row["Margin"])
            if not 0 <= score <= 1 or not 0 <= margin <= 1:
                raise MappingReviewError(f"Invalid score at row {row_number}")
            if suggested and not text(actual_row["Evidence"]):
                raise MappingReviewError(f"Suggestion lacks evidence at row {row_number}")
    finally:
        workbook.close()


def main() -> int:
    """Generate, validate, and atomically publish the mapping review."""
    protected = [MASTER_SERVICES_PATH, MASTER_LABOR_PATH]
    temporary_created = False
    try:
        require_files(protected)
        if TEMP_OUTPUT_PATH.exists():
            raise MappingReviewError(
                f"Stale temporary output exists: {TEMP_OUTPUT_PATH}"
            )
        hashes = {path: file_hash(path) for path in protected}
        services = read_records(
            MASTER_SERVICES_PATH,
            SERVICES_SHEET,
            {"Service ID", "Service Name", "Labor Standard ID"},
        )
        labor_rows = read_records(
            MASTER_LABOR_PATH,
            LABOR_SHEET,
            {
                "Labor Standard ID",
                "Legacy Labor ID",
                "Labor Name",
                "Repair Type",
            },
        )
        validate_source_identity(services, labor_rows)
        review_rows = [mapping_row(service, labor_rows) for service in services]
        generated_at = datetime.now(UTC).replace(tzinfo=None)
        workbook = build_workbook(review_rows, hashes, generated_at)
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary_created = True
        try:
            workbook.save(TEMP_OUTPUT_PATH)
        finally:
            workbook.close()
        validate_reopened(
            TEMP_OUTPUT_PATH,
            services,
            labor_rows,
            review_rows,
        )
        changed = [
            str(path)
            for path, digest in hashes.items()
            if file_hash(path) != digest
        ]
        if changed:
            raise MappingReviewError(
                f"Protected input hash changed: {', '.join(changed)}"
            )
        os.replace(TEMP_OUTPUT_PATH, OUTPUT_PATH)
        print(f"Generated: {OUTPUT_PATH}")
        print(f"Service review rows: {len(review_rows)}")
        print("Approved mappings: 0")
        print("Master Services modified: No")
        return 0
    except (
        MappingReviewError,
        OSError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
    ) as exc:
        print(f"ERROR: {text(exc)}", file=sys.stderr)
        return 1
    finally:
        if temporary_created and TEMP_OUTPUT_PATH.exists():
            try:
                TEMP_OUTPUT_PATH.unlink()
            except OSError as cleanup_error:
                print(
                    f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}",
                    file=sys.stderr,
                )


if __name__ == "__main__":
    raise SystemExit(main())
