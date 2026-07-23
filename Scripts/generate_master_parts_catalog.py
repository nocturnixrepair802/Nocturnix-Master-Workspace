"""Generate the standalone Nocturnix Master Parts Catalog v1 review workbook.

The output is a proposed catalog artifact only. Inputs are opened read-only,
protected by SHA-256 checks, and never saved.
"""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROPOSAL_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
MASTER_SERVICES_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Services_Catalog_v1.xlsx"
)
MASTER_PRICING_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Pricing_Catalog_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Parts_Catalog_v1.xlsx"
)
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(
    f"{OUTPUT_PATH.stem}.tmp{OUTPUT_PATH.suffix}"
)

RETAINED_SHEET = "01 - Retained"
EXCLUSIONS_SHEET = "02 - Duplicate Exclusions"
EXPECTED_PART_ROWS = 48
IMPORT_BATCH_ID = "MASTER-PARTS-V1-REVIEW"
NAMESPACE_AUTHORITY = "ADR-008"
PART_ID_PATTERN = re.compile(r"^PRT\d{6}$")
PART_ID_SHEETS = ("39.4 Repair Parts", "41 Parts")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Master Parts",
    "02 - Part Categories",
    "03 - Part Types",
    "04 - Manufacturers",
    "05 - Device Families",
    "06 - Device References",
    "07 - Suppliers",
    "08 - Part Qualities",
    "09 - Conditions",
    "10 - Compatibility Review",
    "11 - Sourcing Review",
    "12 - Validation Summary",
    "13 - Revision History",
    "14 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterPartsInstructions",
    "01 - Master Parts": "tblMasterPartsCatalog",
    "02 - Part Categories": "tblPartCategories",
    "03 - Part Types": "tblPartTypes",
    "04 - Manufacturers": "tblPartsManufacturers",
    "05 - Device Families": "tblPartsDeviceFamilies",
    "06 - Device References": "tblPartsDeviceReferences",
    "07 - Suppliers": "tblPartsSuppliers",
    "08 - Part Qualities": "tblPartQualities",
    "09 - Conditions": "tblPartConditions",
    "10 - Compatibility Review": "tblPartsCompatibilityReview",
    "11 - Sourcing Review": "tblPartsSourcingReview",
    "12 - Validation Summary": "tblMasterPartsValidation",
    "13 - Revision History": "tblMasterPartsRevisionHistory",
    "14 - Import Metadata": "tblMasterPartsImportMetadata",
}
PART_HEADERS = [
    "Part ID",
    "Legacy Part SKU",
    "Active",
    "Status",
    "Part Name",
    "Part Display Name",
    "Part Description",
    "Part Category ID",
    "Part Category",
    "Part Type ID",
    "Part Type",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Device ID",
    "Device Name",
    "Compatibility Scope",
    "Part Quality",
    "Condition",
    "OEM Status",
    "Color",
    "Capacity",
    "Model Number",
    "Supplier ID",
    "Supplier Name",
    "Supplier Part Number",
    "Preferred Supplier",
    "Legacy Retail Price",
    "Legacy Cost",
    "Currency",
    "Cost Status",
    "Pricing Status",
    "Inventory Tracked",
    "Serialized",
    "Warranty Eligible",
    "Default Warranty",
    "Compatibility Status",
    "Sourcing Status",
    "Review Status",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Reviewer",
    "Reviewer Notes",
    "Effective Date",
    "Last Reviewed",
    "Created At",
    "Updated At",
]
COMPATIBILITY_HEADERS = [
    "Part ID",
    "Legacy Part SKU",
    "Part Name",
    "Manufacturer Name",
    "Device Family Name",
    "Device ID",
    "Device Name",
    "Proposed Compatibility Scope",
    "Missing Compatibility Inputs",
    "Required Action",
    "Review Status",
    "Reviewer Notes",
]
SOURCING_HEADERS = [
    "Part ID",
    "Legacy Part SKU",
    "Part Name",
    "Supplier ID",
    "Supplier Name",
    "Supplier Part Number",
    "Legacy Cost",
    "Cost Status",
    "Sourcing Status",
    "Missing Sourcing Inputs",
    "Required Action",
    "Review Status",
    "Reviewer Notes",
]
PART_STATUSES = ["Draft", "Active", "Planned", "Future", "Archived", "Rejected"]
REVIEW_STATUSES = [
    "Pending Review",
    "Pending Manufacturer Review",
    "Pending Device Mapping",
    "Pending Compatibility Review",
    "Pending Sourcing Review",
    "Pending Cost Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
COMPATIBILITY_STATUSES = [
    "Pending Compatibility Review",
    "Device Family Only",
    "Device Specific",
    "Universal",
    "Not Applicable",
    "Approved",
    "Rejected",
]
SOURCING_STATUSES = [
    "Pending Sourcing Review",
    "Supplier Unknown",
    "Supplier Observed",
    "Multiple Suppliers",
    "Preferred Supplier Proposed",
    "Approved",
    "Rejected",
]
COST_STATUSES = [
    "Pending Cost Review",
    "Legacy Cost Only",
    "Supplier Cost Required",
    "Landed Cost Required",
    "Ready for Cost Approval",
    "Approved",
    "Rejected",
]
PRICING_STATUSES = [
    "Pending Pricing Review",
    "Legacy Price Review",
    "Not for Direct Sale",
    "Ready for Pricing Review",
    "Approved",
    "Rejected",
]
YES_NO = ["Yes", "No"]
DEFINED_NAME_BY_HEADER = {
    "Active": "DV_YesNo",
    "Status": "DV_PartStatuses",
    "Part Category ID": "DV_PartCategoryIDs",
    "Part Type ID": "DV_PartTypeIDs",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Device ID": "DV_DeviceIDs",
    "Part Quality": "DV_PartQualities",
    "Condition": "DV_Conditions",
    "Supplier ID": "DV_SupplierIDs",
    "Preferred Supplier": "DV_YesNo",
    "Cost Status": "DV_CostStatuses",
    "Pricing Status": "DV_PricingStatuses",
    "Inventory Tracked": "DV_YesNo",
    "Serialized": "DV_YesNo",
    "Warranty Eligible": "DV_YesNo",
    "Default Warranty": "DV_WarrantyOptions",
    "Compatibility Status": "DV_CompatibilityStatuses",
    "Sourcing Status": "DV_SourcingStatuses",
    "Review Status": "DV_ReviewStatuses",
}
DEFINED_NAME_SHEET = {
    "DV_PartStatuses": "08 - Part Qualities",
    "DV_ReviewStatuses": "08 - Part Qualities",
    "DV_CompatibilityStatuses": "08 - Part Qualities",
    "DV_SourcingStatuses": "08 - Part Qualities",
    "DV_CostStatuses": "08 - Part Qualities",
    "DV_PricingStatuses": "08 - Part Qualities",
    "DV_PartCategoryIDs": "02 - Part Categories",
    "DV_PartTypeIDs": "03 - Part Types",
    "DV_ManufacturerIDs": "04 - Manufacturers",
    "DV_DeviceFamilyCodes": "05 - Device Families",
    "DV_DeviceIDs": "06 - Device References",
    "DV_SupplierIDs": "07 - Suppliers",
    "DV_PartQualities": "08 - Part Qualities",
    "DV_Conditions": "09 - Conditions",
    "DV_WarrantyOptions": "08 - Part Qualities",
    "DV_YesNo": "08 - Part Qualities",
}
MONEY_FIELDS = {"Legacy Retail Price", "Legacy Cost"}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


class PartsCatalogError(RuntimeError):
    """Raised when the Master Parts review artifact cannot be built safely."""


def text(value: Any) -> str:
    """Normalize a scalar to stripped text."""
    return "" if value is None else str(value).strip()


def ascii_value(value: Any) -> Any:
    """Return ASCII-safe strings without changing typed non-string values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return value.encode("ascii", "replace").decode("ascii")


def excel_safe_value(value: Any) -> Any:
    """Normalize workbook-bound date/time values for Excel."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def assert_excel_safe_value_contract() -> None:
    """Exercise the Excel persistence boundary without workbook I/O."""
    aware_utc = datetime(2026, 7, 23, 12, 30, tzinfo=UTC)
    assert excel_safe_value(aware_utc) == datetime(2026, 7, 23, 12, 30)
    eastern = timezone(timedelta(hours=-4))
    aware_eastern = datetime(2026, 7, 23, 8, 30, tzinfo=eastern)
    assert excel_safe_value(aware_eastern) == datetime(2026, 7, 23, 12, 30)
    naive = datetime(2026, 7, 23, 12, 30)
    assert excel_safe_value(naive) is naive
    calendar_date = date(2026, 7, 23)
    assert excel_safe_value(calendar_date) is calendar_date
    assert excel_safe_value("") == ""
    assert excel_safe_value(None) is None
    stamp = "2026-07-23T12:30:00Z"
    assert excel_safe_value(stamp) == stamp


def decimal_value(value: Any, field: str, *, allow_text: bool = False) -> Decimal | None:
    """Parse a finite Decimal while distinguishing blank from zero."""
    if value is None or text(value) == "":
        return None
    if isinstance(value, bool):
        raise PartsCatalogError(f"{field} contains Boolean value {value!r}")
    try:
        result = Decimal(text(value))
    except InvalidOperation:
        if allow_text:
            return None
        raise PartsCatalogError(f"{field} is not numeric: {value!r}") from None
    if not result.is_finite():
        raise PartsCatalogError(f"{field} is not finite: {value!r}")
    return result


def file_hash(path: Path) -> str:
    """Calculate SHA-256 without changing the file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require each protected workbook exactly once and as a valid ZIP."""
    seen: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved in seen:
            raise PartsCatalogError(f"Protected input configured twice: {path}")
        seen.add(resolved)
        if not path.is_file():
            raise PartsCatalogError(f"Required workbook does not exist: {path}")
        if path.stat().st_size <= 0:
            raise PartsCatalogError(f"Required workbook is empty: {path}")
        if not zipfile.is_zipfile(path):
            raise PartsCatalogError(f"Required workbook is not a valid Excel ZIP: {path}")


def read_records(
    path: Path,
    sheet_name: str,
    *,
    header_label: str | None = None,
    keep_vba: bool = False,
) -> list[dict[str, Any]]:
    """Read a worksheet from its first row or a located header label."""
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_vba=keep_vba,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise PartsCatalogError(f"{path.name} lacks worksheet {sheet_name!r}")
        worksheet = workbook[sheet_name]
        header_row = 1
        if header_label is not None:
            header_row = 0
            for row in worksheet.iter_rows():
                if any(text(cell.value) == header_label for cell in row):
                    header_row = row[0].row
                    break
            if header_row == 0:
                raise PartsCatalogError(
                    f"{path.name}:{sheet_name} lacks header {header_label!r}"
                )
        rows = worksheet.iter_rows(min_row=header_row, values_only=True)
        headers = [text(value) for value in next(rows)]
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    except StopIteration as exc:
        raise PartsCatalogError(f"{path.name}:{sheet_name} is empty") from exc
    finally:
        workbook.close()


def read_existing_part_ids() -> tuple[set[str], list[str], list[str]]:
    """Inspect canonical Part ID columns under ADR-008."""
    workbook = load_workbook(
        CANONICAL_PATH,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )
    valid: set[str] = set()
    malformed: list[str] = []
    locations: list[str] = []
    try:
        for sheet_name in PART_ID_SHEETS:
            if sheet_name not in workbook.sheetnames:
                raise PartsCatalogError(
                    f"Canonical workbook lacks worksheet {sheet_name!r}"
                )
            worksheet = workbook[sheet_name]
            headers = [
                cell
                for row in worksheet.iter_rows()
                for cell in row
                if text(cell.value) == "Part ID"
            ]
            if not headers:
                locations.append(f"{sheet_name}: no Part ID column")
                continue
            for header in headers:
                locations.append(
                    f"{sheet_name}!{get_column_letter(header.column)}{header.row}"
                )
                for row_number in range(header.row + 1, worksheet.max_row + 1):
                    candidate = text(
                        worksheet.cell(
                            row=row_number,
                            column=header.column,
                        ).value
                    )
                    if not candidate:
                        continue
                    if PART_ID_PATTERN.fullmatch(candidate):
                        if candidate in valid:
                            raise PartsCatalogError(
                                f"Duplicate canonical Part ID: {candidate}"
                            )
                        valid.add(candidate)
                    else:
                        malformed.append(candidate)
        return valid, malformed, locations
    finally:
        workbook.close()


def allocate_part_ids(existing: set[str], count: int) -> list[str]:
    """Allocate and validate the next continuous Part ID sequence."""
    highest = max((int(identifier[3:]) for identifier in existing), default=0)
    identifiers = [
        f"PRT{number:06d}"
        for number in range(highest + 1, highest + count + 1)
    ]
    if any(not PART_ID_PATTERN.fullmatch(value) for value in identifiers):
        raise PartsCatalogError("Generated Part ID has invalid format")
    if len(identifiers) != len(set(identifiers)):
        raise PartsCatalogError("Generated Part IDs are not unique")
    if existing.intersection(identifiers):
        raise PartsCatalogError("Generated Part ID overlaps a canonical ID")
    numbers = [int(value[3:]) for value in identifiers]
    if numbers != list(range(highest + 1, highest + count + 1)):
        raise PartsCatalogError("Generated Part IDs are not continuous")
    if identifiers and not existing and identifiers[0] != "PRT000001":
        raise PartsCatalogError("ADR-008 empty namespace must start at PRT000001")
    if identifiers and numbers[-1] != numbers[0] + count - 1:
        raise PartsCatalogError("Final Part ID does not reconcile to row count")
    return identifiers


def require_headers(
    rows: Sequence[dict[str, Any]],
    required: Iterable[str],
    label: str,
) -> None:
    """Require source columns and at least one record."""
    if not rows:
        raise PartsCatalogError(f"{label} contains no records")
    missing = sorted(set(required) - set(rows[0]))
    if missing:
        raise PartsCatalogError(f"{label} missing columns: {', '.join(missing)}")


def source_part_rows(
    retained: Sequence[dict[str, Any]],
    exclusions: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Filter, order, and validate the retained Part population."""
    required = {
        "Source Row Number",
        "Record Category",
        "Legacy SKU",
        "Legacy Group",
        "Legacy Type",
        "Legacy Manufacturer",
        "Legacy Name",
        "Legacy Retail Price",
        "Legacy Cost",
        "Legacy Condition",
        "Legacy Supplier",
        "Legacy Note",
        "Source Workbook",
        "Source Worksheet",
        "Source Created At",
        "Source Updated At",
    }
    require_headers(retained, required, "Retained proposal")
    parts = [row for row in retained if text(row["Record Category"]) == "Part"]
    if len(parts) != EXPECTED_PART_ROWS:
        raise PartsCatalogError(
            f"Expected {EXPECTED_PART_ROWS} retained Part rows; found {len(parts)}"
        )
    try:
        parts.sort(key=lambda row: int(text(row["Source Row Number"])))
    except ValueError as exc:
        raise PartsCatalogError("Part Source Row Number must be an integer") from exc
    source_numbers = [int(text(row["Source Row Number"])) for row in parts]
    if len(source_numbers) != len(set(source_numbers)):
        raise PartsCatalogError("Retained Part Source Row Numbers are duplicated")
    excluded = {
        text(row.get("Excluded Source Row Number") or row.get("Source Row Number"))
        for row in exclusions
    }
    overlap = [number for number in source_numbers if str(number) in excluded]
    if overlap:
        raise PartsCatalogError(
            f"Retained Parts include duplicate exclusions: {overlap}"
        )
    return parts


def unique_lookup_rows(
    rows: Sequence[dict[str, Any]],
    identifier_field: str,
    name_field: str,
) -> list[dict[str, Any]]:
    """Require unique nonblank lookup identifiers and names."""
    seen_ids: set[str] = set()
    output: list[dict[str, Any]] = []
    for row in rows:
        identifier = text(row.get(identifier_field))
        name = text(row.get(name_field))
        if not identifier or not name:
            continue
        if identifier in seen_ids:
            raise PartsCatalogError(
                f"Duplicate lookup {identifier_field}: {identifier}"
            )
        seen_ids.add(identifier)
        output.append(row)
    return output


def proposed_lookup(
    values: Iterable[str],
    prefix: str,
) -> tuple[list[dict[str, str]], dict[str, str]]:
    """Build deterministic proposal-local lookup IDs."""
    names = sorted({text(value) for value in values if text(value)}, key=str.casefold)
    rows = [
        {"ID": f"{prefix}-{index:04d}", "Name": name}
        for index, name in enumerate(names, 1)
    ]
    return rows, {row["Name"]: row["ID"] for row in rows}


def manufacturer_map(
    manufacturers: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, tuple[str, str]]]:
    """Build canonical manufacturer output and exact-name lookup."""
    canonical = unique_lookup_rows(
        manufacturers,
        "Manufacturer ID",
        "Manufacturer",
    )
    output = [
        {
            "Manufacturer ID": row["Manufacturer ID"],
            "Manufacturer Name": row["Manufacturer"],
            "Active": row.get("Active", ""),
            "Source": "Canonical 30 Manufacturers",
        }
        for row in canonical
    ]
    lookup = {
        text(row["Manufacturer"]).casefold(): (
            text(row["Manufacturer ID"]),
            text(row["Manufacturer"]),
        )
        for row in canonical
    }
    return output, lookup


def family_map(
    families: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, tuple[str, str]]]:
    """Build canonical family output and explicit type-name lookup."""
    canonical = unique_lookup_rows(
        families,
        "Device Family Code",
        "Device Family",
    )
    output = [
        {
            "Device Family Code": row["Device Family Code"],
            "Device Family Name": row["Device Family"],
            "Description": row.get("Description", ""),
            "Active": row.get("Active", ""),
        }
        for row in canonical
    ]
    lookup = {
        text(row["Device Family"]).casefold(): (
            text(row["Device Family Code"]),
            text(row["Device Family"]),
        )
        for row in canonical
    }
    return output, lookup


def map_family(
    legacy_type: Any,
    families: dict[str, tuple[str, str]],
) -> tuple[str, str]:
    """Map only an explicit `Part - <family>` source classification."""
    value = text(legacy_type)
    if not value.casefold().startswith("part - "):
        return "", ""
    candidate = value.split("-", 1)[1].strip().casefold()
    if candidate in {"other", ""}:
        return "", ""
    return families.get(candidate, ("", ""))


def observation_status(value: Any, *, positive_status: str, pending: str) -> str:
    """Classify a provisional monetary observation without calculating."""
    amount = decimal_value(value, "Legacy monetary observation", allow_text=True)
    if amount is not None and amount < 0:
        raise PartsCatalogError(f"Negative legacy monetary value: {value!r}")
    if value is None or text(value) == "":
        return pending
    if positive_status == "Legacy Price Review" and (amount is None or amount <= 0):
        return pending
    return positive_status


def build_part_rows(
    sources: Sequence[dict[str, Any]],
    part_ids: Sequence[str],
    category_ids: dict[str, str],
    type_ids: dict[str, str],
    manufacturers: dict[str, tuple[str, str]],
    families: dict[str, tuple[str, str]],
) -> list[dict[str, Any]]:
    """Build one conservative draft record per retained Part."""
    records: list[dict[str, Any]] = []
    for part_id, source in zip(part_ids, sources, strict=True):
        manufacturer_observation = text(source["Legacy Manufacturer"])
        manufacturer_id = ""
        manufacturer_name = manufacturer_observation
        if manufacturer_observation.casefold() in manufacturers:
            candidate_id, canonical_name = manufacturers[
                manufacturer_observation.casefold()
            ]
            if manufacturer_observation == canonical_name:
                manufacturer_id = candidate_id
        family_code, family_name = map_family(source["Legacy Type"], families)
        supplier_name = text(source["Legacy Supplier"])
        cost_status = observation_status(
            source["Legacy Cost"],
            positive_status="Legacy Cost Only",
            pending="Pending Cost Review",
        )
        pricing_status = observation_status(
            source["Legacy Retail Price"],
            positive_status="Legacy Price Review",
            pending="Pending Pricing Review",
        )
        compatibility_status = (
            "Device Family Only" if family_code else "Pending Compatibility Review"
        )
        sourcing_status = (
            "Supplier Observed" if supplier_name else "Supplier Unknown"
        )
        if not manufacturer_id:
            review_status = "Pending Manufacturer Review"
        elif not family_code:
            review_status = "Pending Device Mapping"
        else:
            review_status = "Pending Compatibility Review"
        category = text(source["Legacy Group"])
        part_type = text(source["Legacy Type"])
        record = {header: "" for header in PART_HEADERS}
        record.update(
            {
                "Part ID": part_id,
                "Legacy Part SKU": source["Legacy SKU"],
                "Active": "Yes",
                "Status": "Draft",
                "Part Name": source["Legacy Name"],
                "Part Display Name": source["Legacy Name"],
                "Part Description": source["Legacy Note"],
                "Part Category ID": category_ids[category],
                "Part Category": category,
                "Part Type ID": type_ids[part_type],
                "Part Type": part_type,
                "Manufacturer ID": manufacturer_id,
                "Manufacturer Name": manufacturer_name,
                "Device Family Code": family_code,
                "Device Family Name": family_name,
                "Condition": source["Legacy Condition"],
                "Supplier Name": supplier_name,
                "Preferred Supplier": "No",
                "Legacy Retail Price": source["Legacy Retail Price"],
                "Legacy Cost": source["Legacy Cost"],
                "Cost Status": cost_status,
                "Pricing Status": pricing_status,
                "Inventory Tracked": "Yes",
                "Serialized": "No",
                "Warranty Eligible": "No",
                "Compatibility Status": compatibility_status,
                "Sourcing Status": sourcing_status,
                "Review Status": review_status,
                "Source Record Number": source["Source Row Number"],
                "Source Workbook": source["Source Workbook"],
                "Source Worksheet": source["Source Worksheet"],
                "Import Batch ID": IMPORT_BATCH_ID,
                "Reviewer Notes": source["Legacy Note"],
                "Created At": source["Source Created At"],
                "Updated At": source["Source Updated At"],
            }
        )
        records.append(record)
    return records


def build_compatibility_rows(
    parts: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build the unresolved compatibility review queue."""
    rows: list[dict[str, Any]] = []
    for part in parts:
        missing = ["Device ID", "Device Name", "Compatibility Scope"]
        if not text(part["Device Family Code"]):
            missing.insert(0, "Device Family")
        rows.append(
            {
                "Part ID": part["Part ID"],
                "Legacy Part SKU": part["Legacy Part SKU"],
                "Part Name": part["Part Name"],
                "Manufacturer Name": part["Manufacturer Name"],
                "Device Family Name": part["Device Family Name"],
                "Device ID": part["Device ID"],
                "Device Name": part["Device Name"],
                "Proposed Compatibility Scope": part["Compatibility Scope"],
                "Missing Compatibility Inputs": "; ".join(missing),
                "Required Action": "Research and approve part applicability",
                "Review Status": "Pending Compatibility Review",
                "Reviewer Notes": "",
            }
        )
    return rows


def build_sourcing_rows(
    parts: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build the unresolved sourcing review queue."""
    rows: list[dict[str, Any]] = []
    for part in parts:
        missing = ["Supplier ID", "Supplier Part Number", "Preferred Supplier Review"]
        if not text(part["Supplier Name"]):
            missing.insert(0, "Supplier Name")
        if part["Legacy Cost"] is None or text(part["Legacy Cost"]) == "":
            missing.append("Legacy Cost")
        rows.append(
            {
                "Part ID": part["Part ID"],
                "Legacy Part SKU": part["Legacy Part SKU"],
                "Part Name": part["Part Name"],
                "Supplier ID": part["Supplier ID"],
                "Supplier Name": part["Supplier Name"],
                "Supplier Part Number": part["Supplier Part Number"],
                "Legacy Cost": part["Legacy Cost"],
                "Cost Status": part["Cost Status"],
                "Sourcing Status": part["Sourcing Status"],
                "Missing Sourcing Inputs": "; ".join(missing),
                "Required Action": "Research supplier relationship and cost evidence",
                "Review Status": "Pending Sourcing Review",
                "Reviewer Notes": "",
            }
        )
    return rows


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    records: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write one formatted Excel Table through the safe value boundary."""
    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    if records:
        for record in records:
            worksheet.append(
                [
                    excel_safe_value(ascii_value(record.get(header, "")))
                    for header in headers
                ]
            )
    else:
        worksheet.append(["" for _ in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"
    )
    table = Table(
        displayName=table_name,
        ref=f"A1:{get_column_letter(len(headers))}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    format_columns(worksheet, headers)


def format_columns(worksheet: Worksheet, headers: Sequence[str]) -> None:
    """Apply bounded widths and semantic number formats."""
    for column, header in enumerate(headers, 1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = (
            42
            if any(word in header for word in ("Notes", "Inputs", "Description"))
            else 22
        )
        if header in MONEY_FIELDS:
            number_format = CURRENCY_FORMAT
        elif header in {"Effective Date", "Last Reviewed", "Date"}:
            number_format = DATE_FORMAT
        elif header in {"Created At", "Updated At", "Generated At UTC"}:
            number_format = DATETIME_FORMAT
        else:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = number_format


def control_rows(warranties: Sequence[str]) -> list[dict[str, str]]:
    """Place all governed enumerations in one table-backed lookup range."""
    columns = {
        "Part Quality": ["Unresolved"],
        "Part Status": PART_STATUSES,
        "Review Status": REVIEW_STATUSES,
        "Compatibility Status": COMPATIBILITY_STATUSES,
        "Sourcing Status": SOURCING_STATUSES,
        "Cost Status": COST_STATUSES,
        "Pricing Status": PRICING_STATUSES,
        "Yes/No": YES_NO,
        "Warranty Option": list(warranties),
    }
    rows = []
    for values in zip_longest(*columns.values(), fillvalue=""):
        rows.append(dict(zip(columns, values, strict=True)))
    return rows


def add_defined_names(
    workbook: Workbook,
    lengths: dict[str, int],
) -> None:
    """Create every workbook-scoped validation name."""
    ranges = {
        "DV_PartCategoryIDs": (
            f"'02 - Part Categories'!$A$2:$A${lengths['categories'] + 1}"
        ),
        "DV_PartTypeIDs": f"'03 - Part Types'!$A$2:$A${lengths['types'] + 1}",
        "DV_ManufacturerIDs": (
            f"'04 - Manufacturers'!$A$2:$A${lengths['manufacturers'] + 1}"
        ),
        "DV_DeviceFamilyCodes": (
            f"'05 - Device Families'!$A$2:$A${lengths['families'] + 1}"
        ),
        "DV_DeviceIDs": (
            f"'06 - Device References'!$A$2:$A${lengths['devices'] + 1}"
        ),
        "DV_SupplierIDs": (
            f"'07 - Suppliers'!$A$2:$A${max(lengths['suppliers'], 1) + 1}"
        ),
        "DV_PartQualities": "'08 - Part Qualities'!$A$2:$A$2",
        "DV_PartStatuses": "'08 - Part Qualities'!$B$2:$B$7",
        "DV_ReviewStatuses": "'08 - Part Qualities'!$C$2:$C$11",
        "DV_CompatibilityStatuses": "'08 - Part Qualities'!$D$2:$D$8",
        "DV_SourcingStatuses": "'08 - Part Qualities'!$E$2:$E$8",
        "DV_CostStatuses": "'08 - Part Qualities'!$F$2:$F$8",
        "DV_PricingStatuses": "'08 - Part Qualities'!$G$2:$G$7",
        "DV_YesNo": "'08 - Part Qualities'!$H$2:$H$3",
        "DV_WarrantyOptions": (
            f"'08 - Part Qualities'!$I$2:$I${lengths['warranties'] + 1}"
        ),
        "DV_Conditions": (
            f"'09 - Conditions'!$A$2:$A${lengths['conditions'] + 1}"
        ),
    }
    for name, reference in ranges.items():
        workbook.defined_names.add(DefinedName(name, attr_text=reference))


def add_data_validations(workbook: Workbook) -> None:
    """Apply governed list validations only through defined names."""
    worksheet = workbook["01 - Master Parts"]
    headers = [text(cell.value) for cell in worksheet[1]]
    for header, name in DEFINED_NAME_BY_HEADER.items():
        column = headers.index(header) + 1
        letter = get_column_letter(column)
        validation = DataValidation(
            type="list",
            formula1=f"={name}",
            allow_blank=True,
        )
        validation.error = f"Select a controlled value for {header}."
        validation.errorTitle = "Invalid value"
        worksheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{worksheet.max_row}")


def add_conditional_formatting(worksheet: Worksheet) -> None:
    """Highlight pending review states and prohibit generated approval."""
    headers = [text(cell.value) for cell in worksheet[1]]
    for header in (
        "Review Status",
        "Compatibility Status",
        "Sourcing Status",
        "Cost Status",
        "Pricing Status",
    ):
        letter = get_column_letter(headers.index(header) + 1)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{worksheet.max_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Pending",{letter}2))'],
                fill=YELLOW_FILL,
            ),
        )
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{worksheet.max_row}",
            CellIsRule(operator="equal", formula=['"Approved"'], fill=RED_FILL),
        )


def build_workbook(
    parts: list[dict[str, Any]],
    categories: list[dict[str, str]],
    part_types: list[dict[str, str]],
    manufacturers: list[dict[str, Any]],
    families: list[dict[str, Any]],
    devices: list[dict[str, Any]],
    suppliers: list[dict[str, Any]],
    conditions: list[str],
    warranties: list[str],
    compatibility: list[dict[str, Any]],
    sourcing: list[dict[str, Any]],
    hashes: dict[Path, str],
    existing_ids: set[str],
    malformed_ids: Sequence[str],
    id_locations: Sequence[str],
) -> Workbook:
    """Build the complete 15-sheet review artifact in memory."""
    if len(SHEET_NAMES) != len(set(SHEET_NAMES)) or any(
        len(name) > 31 for name in SHEET_NAMES
    ):
        raise PartsCatalogError("Worksheet titles must be unique and <= 31 chars")
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}

    instructions = [
        {"Topic": "Purpose", "Guidance": "Proposed Master Parts review catalog."},
        {"Topic": "Boundary", "Guidance": "Not inventory, purchasing, or final pricing."},
        {"Topic": "Identity", "Guidance": "ADR-008 PRT###### draft identities."},
        {"Topic": "Legacy SKU", "Guidance": "Alias/reference only."},
        {"Topic": "Approval", "Guidance": "No generated row is approved."},
        {"Topic": "Import", "Guidance": "No canonical import is performed."},
    ]
    append_table(
        sheets[SHEET_NAMES[0]],
        ["Topic", "Guidance"],
        instructions,
        TABLE_NAMES[SHEET_NAMES[0]],
    )
    append_table(
        sheets[SHEET_NAMES[1]],
        PART_HEADERS,
        parts,
        TABLE_NAMES[SHEET_NAMES[1]],
    )
    category_output = [
        {
            "Part Category ID": row["ID"],
            "Part Category": row["Name"],
            "Source": "Legacy Group",
            "Status": "Proposed",
        }
        for row in categories
    ]
    append_table(
        sheets[SHEET_NAMES[2]],
        ["Part Category ID", "Part Category", "Source", "Status"],
        category_output,
        TABLE_NAMES[SHEET_NAMES[2]],
    )
    type_output = [
        {
            "Part Type ID": row["ID"],
            "Part Type": row["Name"],
            "Source": "Legacy Type",
            "Status": "Proposed",
        }
        for row in part_types
    ]
    append_table(
        sheets[SHEET_NAMES[3]],
        ["Part Type ID", "Part Type", "Source", "Status"],
        type_output,
        TABLE_NAMES[SHEET_NAMES[3]],
    )
    append_table(
        sheets[SHEET_NAMES[4]],
        ["Manufacturer ID", "Manufacturer Name", "Active", "Source"],
        manufacturers,
        TABLE_NAMES[SHEET_NAMES[4]],
    )
    append_table(
        sheets[SHEET_NAMES[5]],
        ["Device Family Code", "Device Family Name", "Description", "Active"],
        families,
        TABLE_NAMES[SHEET_NAMES[5]],
    )
    append_table(
        sheets[SHEET_NAMES[6]],
        ["Device ID", "Device Name", "Manufacturer ID", "Device Family Code", "Model Number", "Active"],
        devices,
        TABLE_NAMES[SHEET_NAMES[6]],
    )
    append_table(
        sheets[SHEET_NAMES[7]],
        ["Supplier ID", "Supplier Name", "Website", "Notes"],
        suppliers,
        TABLE_NAMES[SHEET_NAMES[7]],
    )
    controls = control_rows(warranties)
    append_table(
        sheets[SHEET_NAMES[8]],
        list(controls[0]),
        controls,
        TABLE_NAMES[SHEET_NAMES[8]],
    )
    condition_rows = [
        {"Condition": value, "Source": "Retained legacy Part observation"}
        for value in conditions
    ]
    append_table(
        sheets[SHEET_NAMES[9]],
        ["Condition", "Source"],
        condition_rows,
        TABLE_NAMES[SHEET_NAMES[9]],
    )
    append_table(
        sheets[SHEET_NAMES[10]],
        COMPATIBILITY_HEADERS,
        compatibility,
        TABLE_NAMES[SHEET_NAMES[10]],
    )
    append_table(
        sheets[SHEET_NAMES[11]],
        SOURCING_HEADERS,
        sourcing,
        TABLE_NAMES[SHEET_NAMES[11]],
    )
    validation = [
        {"Validation Check": "Retained Part population", "Result": "PASS", "Evidence": len(parts)},
        {"Validation Check": "Part ID sequence", "Result": "PASS", "Evidence": f"{parts[0]['Part ID']} through {parts[-1]['Part ID']}"},
        {"Validation Check": "Final pricing fields", "Result": "PASS", "Evidence": "Not present"},
        {"Validation Check": "Canonical import", "Result": "PASS", "Evidence": "Not performed"},
    ]
    append_table(
        sheets[SHEET_NAMES[12]],
        ["Validation Check", "Result", "Evidence"],
        validation,
        TABLE_NAMES[SHEET_NAMES[12]],
    )
    revision = [
        {
            "Version": "1.0 Draft",
            "Date": datetime.now(UTC).date(),
            "Description": "Initial Master Parts review framework",
            "Status": "Draft",
            "Approved By": "",
            "Notes": "No canonical import, inventory update, or pricing approval.",
        }
    ]
    append_table(
        sheets[SHEET_NAMES[13]],
        list(revision[0]),
        revision,
        TABLE_NAMES[SHEET_NAMES[13]],
    )
    highest = (
        max(existing_ids, key=lambda value: int(value[3:]))
        if existing_ids
        else ""
    )
    metadata = [
        {"Metadata Field": "Artifact", "Value": "Master Parts Catalog V1 Review"},
        {"Metadata Field": "Generated At UTC", "Value": datetime.now(UTC)},
        {"Metadata Field": "Import Batch ID", "Value": IMPORT_BATCH_ID},
        {"Metadata Field": "Schema Columns", "Value": len(PART_HEADERS)},
        {"Metadata Field": "Namespace Authority", "Value": NAMESPACE_AUTHORITY},
        {"Metadata Field": "Part ID Sources", "Value": "; ".join(id_locations)},
        {"Metadata Field": "Existing Part ID Count", "Value": len(existing_ids)},
        {"Metadata Field": "Highest Existing Part ID", "Value": highest},
        {"Metadata Field": "First Generated Part ID", "Value": parts[0]["Part ID"]},
        {"Metadata Field": "Final Generated Part ID", "Value": parts[-1]["Part ID"]},
        {"Metadata Field": "Generated Part Count", "Value": len(parts)},
        {"Metadata Field": "Malformed Existing Part IDs", "Value": "; ".join(malformed_ids)},
        {"Metadata Field": "Output", "Value": str(OUTPUT_PATH)},
    ]
    for path, digest in hashes.items():
        metadata.extend(
            [
                {"Metadata Field": f"Protected Input Path: {path.name}", "Value": str(path)},
                {"Metadata Field": f"SHA-256: {path.name}", "Value": digest},
            ]
        )
    append_table(
        sheets[SHEET_NAMES[14]],
        ["Metadata Field", "Value"],
        metadata,
        TABLE_NAMES[SHEET_NAMES[14]],
    )

    add_defined_names(
        workbook,
        {
            "categories": len(categories),
            "types": len(part_types),
            "manufacturers": len(manufacturers),
            "families": len(families),
            "devices": len(devices),
            "suppliers": len(suppliers),
            "warranties": len(warranties),
            "conditions": len(conditions),
        },
    )
    add_data_validations(workbook)
    add_conditional_formatting(sheets[SHEET_NAMES[1]])
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read records from a named generated table."""
    if table_name not in worksheet.tables:
        raise PartsCatalogError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = [text(value) for value in rows[0]]
    return [
        dict(zip(headers, row, strict=False))
        for row in rows[1:]
        if any(value is not None and text(value) != "" for value in row)
    ]


def require_excel_archive(path: Path) -> None:
    """Require a complete, readable OOXML workbook container."""
    if (
        not path.is_file()
        or path.stat().st_size <= 0
        or not zipfile.is_zipfile(path)
    ):
        raise PartsCatalogError("Temporary generated Parts workbook is invalid")
    with zipfile.ZipFile(path) as archive:
        required = {"[Content_Types].xml", "xl/workbook.xml"}
        if not required.issubset(archive.namelist()) or archive.testzip() is not None:
            raise PartsCatalogError("Temporary generated Parts workbook is invalid")


def validate_reopened(
    output_path: Path,
    expected_parts: Sequence[dict[str, Any]],
    expected_hashes: dict[Path, str],
) -> None:
    """Validate the temporary workbook before atomic publication."""
    require_excel_archive(output_path)
    workbook = load_workbook(output_path, data_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise PartsCatalogError("Reopened worksheet order changed")
        all_tables: list[str] = []
        for name in SHEET_NAMES:
            worksheet = workbook[name]
            expected_table = TABLE_NAMES[name]
            if expected_table not in worksheet.tables:
                raise PartsCatalogError(f"{name} lacks {expected_table}")
            if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
                raise PartsCatalogError(f"{name} lacks frozen header or filter")
            if worksheet.merged_cells.ranges:
                raise PartsCatalogError(f"{name} contains merged cells")
            all_tables.extend(worksheet.tables)
        if len(all_tables) != len(set(all_tables)):
            raise PartsCatalogError("Excel Table names are not unique")
        worksheet = workbook["01 - Master Parts"]
        headers = [text(cell.value) for cell in worksheet[1]]
        if headers != PART_HEADERS:
            raise PartsCatalogError("Reopened Master Parts schema changed")
        actual = table_records(worksheet, "tblMasterPartsCatalog")
        if len(actual) != EXPECTED_PART_ROWS:
            raise PartsCatalogError("Reopened Master Parts row count changed")
        for row_number, (expected, reopened) in enumerate(
            zip(expected_parts, actual, strict=True),
            2,
        ):
            for field in PART_HEADERS:
                left = excel_safe_value(expected[field])
                right = excel_safe_value(reopened[field])
                if text(left) != text(right):
                    raise PartsCatalogError(
                        f"Reopened row {row_number} changed {field}: "
                        f"{expected[field]!r} -> {reopened[field]!r}"
                    )
        required_names = set(DEFINED_NAME_BY_HEADER.values())
        missing_names = required_names - set(workbook.defined_names)
        if missing_names:
            raise PartsCatalogError(
                f"Missing defined names: {sorted(missing_names)}"
            )
        for name, expected_sheet in DEFINED_NAME_SHEET.items():
            destinations = list(workbook.defined_names[name].destinations)
            if len(destinations) != 1 or destinations[0][0] != expected_sheet:
                raise PartsCatalogError(
                    f"Defined name {name} does not target {expected_sheet}"
                )
        validations = worksheet.data_validations.dataValidation
        formulas = {
            text(validation.formula1)
            for validation in validations
            if validation.type == "list"
        }
        expected_formulas = {f"={name}" for name in required_names}
        if formulas != expected_formulas or any("!" in value for value in formulas):
            raise PartsCatalogError(
                f"Unexpected list-validation formulas: {sorted(formulas)}"
            )
        for header, name in DEFINED_NAME_BY_HEADER.items():
            column = get_column_letter(headers.index(header) + 1)
            expected_range = f"{column}2:{column}{worksheet.max_row}"
            if not any(
                text(validation.formula1) == f"={name}"
                and expected_range in str(validation.sqref).split()
                for validation in validations
            ):
                raise PartsCatalogError(
                    f"{header} does not use defined-name validation {name}"
                )
    finally:
        workbook.close()
    current_hashes = {path: file_hash(path) for path in expected_hashes}
    if current_hashes != expected_hashes:
        changed = [
            str(path)
            for path in expected_hashes
            if current_hashes[path] != expected_hashes[path]
        ]
        raise PartsCatalogError(f"Protected input hash changed: {changed}")


def main() -> int:
    """Generate, validate, and atomically publish the review workbook."""
    protected = (
        PROPOSAL_PATH,
        MASTER_SERVICES_PATH,
        MASTER_PRICING_PATH,
        CANONICAL_PATH,
    )
    workbook: Workbook | None = None
    published = False
    try:
        assert_excel_safe_value_contract()
        require_files(protected)
        hashes = {path: file_hash(path) for path in protected}
        retained = read_records(PROPOSAL_PATH, RETAINED_SHEET)
        exclusions = read_records(PROPOSAL_PATH, EXCLUSIONS_SHEET)
        sources = source_part_rows(retained, exclusions)
        existing_ids, malformed_ids, id_locations = read_existing_part_ids()
        part_ids = allocate_part_ids(existing_ids, len(sources))

        manufacturers_raw = read_records(
            CANONICAL_PATH,
            "30 Manufacturers",
            header_label="Manufacturer ID",
            keep_vba=True,
        )
        families_raw = read_records(
            CANONICAL_PATH,
            "31 Device Families",
            header_label="Device Family Code",
            keep_vba=True,
        )
        devices_raw = read_records(
            CANONICAL_PATH,
            "32 Devices",
            header_label="Device ID",
            keep_vba=True,
        )
        suppliers_raw = read_records(
            CANONICAL_PATH,
            "43 Suppliers",
            header_label="Supplier",
            keep_vba=True,
        )
        warranty_raw = read_records(
            MASTER_SERVICES_PATH,
            "11 - Warranty Options",
        )

        manufacturers, manufacturer_lookup = manufacturer_map(manufacturers_raw)
        families, family_lookup = family_map(families_raw)
        categories, category_ids = proposed_lookup(
            (text(row["Legacy Group"]) for row in sources),
            "PC",
        )
        part_types, type_ids = proposed_lookup(
            (text(row["Legacy Type"]) for row in sources),
            "PT",
        )
        devices = [
            {
                "Device ID": row["Device ID"],
                "Device Name": row["Device Model"],
                "Manufacturer ID": row["Manufacturer Code"],
                "Device Family Code": row["Device Family Code"],
                "Model Number": row["Model Number"],
                "Active": row["Active"],
            }
            for row in unique_lookup_rows(devices_raw, "Device ID", "Device Model")
        ]
        suppliers = [
            {
                "Supplier ID": "",
                "Supplier Name": row["Supplier"],
                "Website": row.get("Website", ""),
                "Notes": row.get("Notes", ""),
            }
            for row in suppliers_raw
            if text(row.get("Supplier"))
        ]
        conditions = sorted(
            {
                text(row["Legacy Condition"])
                for row in sources
                if text(row["Legacy Condition"])
            },
            key=str.casefold,
        )
        warranties = sorted(
            {"N/A"}
            | {
                text(row.get("Warranty Option"))
                for row in warranty_raw
                if text(row.get("Warranty Option"))
            },
            key=str.casefold,
        )
        parts = build_part_rows(
            sources,
            part_ids,
            category_ids,
            type_ids,
            manufacturer_lookup,
            family_lookup,
        )
        compatibility = build_compatibility_rows(parts)
        sourcing = build_sourcing_rows(parts)
        workbook = build_workbook(
            parts,
            categories,
            part_types,
            manufacturers,
            families,
            devices,
            suppliers,
            conditions,
            warranties,
            compatibility,
            sourcing,
            hashes,
            existing_ids,
            malformed_ids,
            id_locations,
        )
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        if TEMP_OUTPUT_PATH.exists():
            TEMP_OUTPUT_PATH.unlink()
        workbook.save(TEMP_OUTPUT_PATH)
        workbook.close()
        workbook = None
        validate_reopened(TEMP_OUTPUT_PATH, parts, hashes)
        TEMP_OUTPUT_PATH.replace(OUTPUT_PATH)
        published = True
    except (
        AssertionError,
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        PartsCatalogError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError as close_error:
                print(f"ERROR: Failed to close workbook: {close_error}", file=sys.stderr)
        if not published:
            try:
                if TEMP_OUTPUT_PATH.exists():
                    TEMP_OUTPUT_PATH.unlink()
            except OSError as cleanup_error:
                print(
                    f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}",
                    file=sys.stderr,
                )

    print(f"Generated: {OUTPUT_PATH}")
    print(f"Namespace authority: {NAMESPACE_AUTHORITY}")
    print(f"Retained Part rows: {len(parts)}")
    print(f"Part ID range: {parts[0]['Part ID']} through {parts[-1]['Part ID']}")
    print(f"Existing canonical Part IDs: {len(existing_ids)}")
    print(f"Malformed canonical Part IDs: {len(malformed_ids)}")
    print("Protected input hashes: PASS")
    print("Inventory/canonical/pricing writes: NOT PERFORMED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
