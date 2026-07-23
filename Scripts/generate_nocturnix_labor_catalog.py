from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

file_path = r"D:\Business Portal\300_Pricing\Nocturnix_Standard_Labor_Catalog_v1.xlsx"

wb = Workbook()
ws_ref = wb.active
ws_ref.title = "00 - Reference Lists"
ws_lab = wb.create_sheet("01 - Labor Standards")
ws_rate = wb.create_sheet("02 - Labor Rate Tiers")
ws_summary = wb.create_sheet("03 - Service Summary")
ws_history = wb.create_sheet("04 - Revision History")

# Reference lists content
reference_lists = [
    ("Device Categories", ["Phones", "Tablets", "Computers", "Gaming", "Wearables"]),
    (
        "Manufacturers",
        [
            "Apple",
            "Samsung",
            "Google",
            "Motorola",
            "OnePlus",
            "General Android",
            "Microsoft",
            "General Tablet",
            "Windows Laptop",
            "MacBook",
            "Chromebook",
            "Desktop",
            "All-in-One",
            "PlayStation",
            "Xbox",
            "Nintendo Switch",
            "Steam Deck",
            "General Controller",
            "Apple Watch",
            "Samsung Watch",
            "General Wearable",
        ],
    ),
    (
        "Services",
        [
            "Diagnostic",
            "Screen Replacement",
            "Battery Replacement",
            "Charging Port Replacement",
            "Charging Port Cleaning",
            "Camera Replacement",
            "Speaker Replacement",
            "Microphone Replacement",
            "Button Repair",
            "Back Glass Replacement",
            "Water Damage Assessment",
            "Internal Cleaning",
            "Data Backup",
            "Data Transfer",
            "Operating System Installation",
            "Driver Installation",
            "Software Setup",
            "SSD Upgrade",
            "RAM Upgrade",
            "Fan Replacement",
            "Thermal Paste Replacement",
            "Keyboard Replacement",
            "Laptop Screen Replacement",
            "Hinge Repair",
            "Console Cleaning",
            "HDMI Port Repair",
            "Storage Replacement",
            "Controller Cleaning",
            "Analog Stick Replacement",
            "Joy-Con Repair",
        ],
    ),
    ("Repair Difficulties", ["Easy", "Moderate", "Difficult", "Advanced"]),
    ("Skill Levels", ["Basic", "Standard", "Advanced", "Expert"]),
    ("Yes/No", ["Yes", "No"]),
    ("Labor Rate Tiers", ["L1 Basic", "L2 Standard", "L3 Advanced", "L4 Expert"]),
    (
        "Warranty Options",
        ["30-Day", "60-Day", "90-Day", "1-Year", "90-Day Limited", "N/A"],
    ),
    ("Status Values", ["Active", "Planned", "Future", "Draft"]),
    ("Confidence Values", ["High", "Medium", "Low"]),
    (
        "Source Types",
        [
            "Internal Estimate",
            "Historical Average",
            "Vendor Guidance",
            "Manufacturer Estimate",
            "Market Research",
        ],
    ),
]

for col, (title, values) in enumerate(reference_lists, start=1):
    ws_ref.cell(row=1, column=col, value=title)
    for row_offset, value in enumerate(values, start=2):
        ws_ref.cell(row=row_offset, column=col, value=value)

# Freeze top row and apply filter
ws_ref.freeze_panes = "A2"
ws_ref.auto_filter.ref = "A1:K1"

# Header layout for Labor Standards
headers = [
    "Labor ID",
    "Device Category",
    "Manufacturer",
    "Device Scope",
    "Service",
    "Repair Difficulty",
    "Skill Level",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Includes Intake",
    "Includes Pre-Test",
    "Includes Repair",
    "Includes Post-Test",
    "Includes Cleanup",
    "Labor Rate Tier",
    "Recommended Warranty",
    "Status",
    "Confidence",
    "Source Type",
    "Notes",
    "Effective Date",
    "Last Reviewed",
]
ws_lab.append(headers)
ws_lab.freeze_panes = "A2"

# Service definitions for generated entries
service_templates = {
    "Diagnostic": {
        "difficulty": "Easy",
        "skill": "Basic",
        "minutes": 25,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L1 Basic",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "General diagnostic and assessment; actual time varies by device condition.",
    },
    "Screen Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 70,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Historical Average",
        "notes": "General display replacement; variability depends on model design and parts access.",
    },
    "Battery Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 45,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Historical Average",
        "notes": "Battery replacement time varies with housing style and adhesive level.",
    },
    "Charging Port Replacement": {
        "difficulty": "Difficult",
        "skill": "Advanced",
        "minutes": 90,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Internal Estimate",
        "notes": "Port replacement is variable and may require disassembly; not a board-level solder service.",
    },
    "Charging Port Cleaning": {
        "difficulty": "Easy",
        "skill": "Basic",
        "minutes": 25,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L1 Basic",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Cleaning and inspection of the charging interface; times vary with debris condition.",
    },
    "Camera Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 65,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Historical Average",
        "notes": "Camera replacement depends on housing design and access complexity.",
    },
    "Speaker Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 55,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Speaker access may require partial disassembly; times are provisional.",
    },
    "Microphone Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 50,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Microphone replacement often requires careful disassembly and verification.",
    },
    "Button Repair": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 50,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Button repair varies with button type and housing access.",
    },
    "Back Glass Replacement": {
        "difficulty": "Difficult",
        "skill": "Advanced",
        "minutes": 110,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Internal Estimate",
        "notes": "Back glass changes are fragile and vary by adhesive and chassis design.",
    },
    "Water Damage Assessment": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 40,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Assessment and initial testing; follow-up repairs may be additional.",
    },
    "Internal Cleaning": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 50,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Cleaning of dust and debris inside the chassis; time varies with access and contamination.",
    },
    "Data Backup": {
        "difficulty": "Easy",
        "skill": "Basic",
        "minutes": 40,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L1 Basic",
        "warranty": "N/A",
        "status": "Active",
        "confidence": "Medium",
        "source": "Vendor Guidance",
        "notes": "Standard backup service; customer data size and device responsiveness affect timing.",
    },
    "Data Transfer": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 55,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "N/A",
        "status": "Active",
        "confidence": "Medium",
        "source": "Vendor Guidance",
        "notes": "Transfer times depend on source/destination data size and device performance.",
    },
    "Operating System Installation": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 90,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Manufacturer Estimate",
        "notes": "OS install time depends on device speed and update requirements.",
    },
    "Driver Installation": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 55,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Manufacturer Estimate",
        "notes": "Driver setup may vary by device model and ecosystem.",
    },
    "Software Setup": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 45,
        "includes": ("Yes", "Yes", "No", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "General software configuration and basic setup; add-on apps are extra.",
    },
    "SSD Upgrade": {
        "difficulty": "Moderate",
        "skill": "Advanced",
        "minutes": 85,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Historical Average",
        "notes": "SSD upgrades depend on device access and cloning requirements.",
    },
    "RAM Upgrade": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 70,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Historical Average",
        "notes": "RAM upgrades vary by slot access and device type.",
    },
    "Fan Replacement": {
        "difficulty": "Moderate",
        "skill": "Advanced",
        "minutes": 85,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Historical Average",
        "notes": "Fan replacement times vary with cooling system access.",
    },
    "Thermal Paste Replacement": {
        "difficulty": "Moderate",
        "skill": "Advanced",
        "minutes": 75,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Internal Estimate",
        "notes": "Thermal service is provisional and depends on heatsink access.",
    },
    "Keyboard Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 80,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Historical Average",
        "notes": "Keyboard replacement depends on internal access and connector layout.",
    },
    "Laptop Screen Replacement": {
        "difficulty": "Difficult",
        "skill": "Advanced",
        "minutes": 110,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Historical Average",
        "notes": "Laptop display replacement varies greatly by hinge and bezel design.",
    },
    "Hinge Repair": {
        "difficulty": "Difficult",
        "skill": "Advanced",
        "minutes": 120,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Internal Estimate",
        "notes": "Hinge repair can require detailed disassembly and parts alignment.",
    },
    "Console Cleaning": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 55,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Console cleaning removes dust and debris; internal access times vary.",
    },
    "HDMI Port Repair": {
        "difficulty": "Difficult",
        "skill": "Advanced",
        "minutes": 95,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L3 Advanced",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Low",
        "source": "Internal Estimate",
        "notes": "High variability due to port assembly and access.",
    },
    "Storage Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 70,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Historical Average",
        "notes": "Storage replacement includes transfer and configuration where applicable.",
    },
    "Controller Cleaning": {
        "difficulty": "Easy",
        "skill": "Basic",
        "minutes": 30,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L1 Basic",
        "warranty": "30-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Controller cleaning is general and depends on contamination level.",
    },
    "Analog Stick Replacement": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 55,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Replacement depends on controller model and disassembly complexity.",
    },
    "Joy-Con Repair": {
        "difficulty": "Moderate",
        "skill": "Standard",
        "minutes": 65,
        "includes": ("Yes", "Yes", "Yes", "Yes", "Yes"),
        "tier": "L2 Standard",
        "warranty": "90-Day",
        "status": "Active",
        "confidence": "Medium",
        "source": "Internal Estimate",
        "notes": "Joy-Con repairs vary by issue type and access complexity.",
    },
}

category_services = {
    "Phones": [
        "Diagnostic",
        "Screen Replacement",
        "Battery Replacement",
        "Charging Port Replacement",
        "Charging Port Cleaning",
        "Camera Replacement",
        "Speaker Replacement",
        "Microphone Replacement",
        "Button Repair",
        "Back Glass Replacement",
        "Water Damage Assessment",
        "Internal Cleaning",
        "Data Backup",
        "Data Transfer",
    ],
    "Tablets": [
        "Diagnostic",
        "Screen Replacement",
        "Battery Replacement",
        "Charging Port Replacement",
        "Charging Port Cleaning",
        "Camera Replacement",
        "Water Damage Assessment",
        "Internal Cleaning",
        "Data Backup",
        "Data Transfer",
        "Operating System Installation",
        "Driver Installation",
    ],
    "Computers": [
        "Diagnostic",
        "Battery Replacement",
        "Data Backup",
        "Data Transfer",
        "Operating System Installation",
        "Driver Installation",
        "Software Setup",
        "SSD Upgrade",
        "RAM Upgrade",
        "Fan Replacement",
        "Thermal Paste Replacement",
        "Keyboard Replacement",
        "Laptop Screen Replacement",
        "Hinge Repair",
        "Storage Replacement",
    ],
    "Gaming": [
        "Diagnostic",
        "Console Cleaning",
        "HDMI Port Repair",
        "Storage Replacement",
        "Controller Cleaning",
        "Analog Stick Replacement",
        "Joy-Con Repair",
        "Software Setup",
    ],
    "Wearables": [
        "Diagnostic",
        "Battery Replacement",
        "Charging Port Cleaning",
        "Internal Cleaning",
        "Software Setup",
        "Data Backup",
    ],
}

manufacturers_by_category = {
    "Phones": ["Apple", "Samsung", "Google", "Motorola", "OnePlus", "General Android"],
    "Tablets": ["Apple", "Samsung", "Microsoft", "General Tablet"],
    "Computers": ["Windows Laptop", "MacBook", "Chromebook", "Desktop", "All-in-One"],
    "Gaming": [
        "PlayStation",
        "Xbox",
        "Nintendo Switch",
        "Steam Deck",
        "General Controller",
    ],
    "Wearables": ["Apple Watch", "Samsung Watch", "General Wearable"],
}

device_scopes = {
    "Phones": "Smartphone",
    "Tablets": "Tablet",
    "Computers": "Computer",
    "Gaming": "Game Console / Accessory",
    "Wearables": "Wearable",
}

warranty_map = {
    "Diagnostic": "30-Day",
    "Screen Replacement": "90-Day",
    "Battery Replacement": "90-Day",
    "Charging Port Replacement": "90-Day",
    "Charging Port Cleaning": "30-Day",
    "Camera Replacement": "90-Day",
    "Speaker Replacement": "90-Day",
    "Microphone Replacement": "90-Day",
    "Button Repair": "90-Day",
    "Back Glass Replacement": "90-Day",
    "Water Damage Assessment": "30-Day",
    "Internal Cleaning": "30-Day",
    "Data Backup": "N/A",
    "Data Transfer": "N/A",
    "Operating System Installation": "30-Day",
    "Driver Installation": "30-Day",
    "Software Setup": "30-Day",
    "SSD Upgrade": "90-Day",
    "RAM Upgrade": "90-Day",
    "Fan Replacement": "90-Day",
    "Thermal Paste Replacement": "90-Day",
    "Keyboard Replacement": "90-Day",
    "Laptop Screen Replacement": "90-Day",
    "Hinge Repair": "90-Day",
    "Console Cleaning": "90-Day",
    "HDMI Port Repair": "90-Day",
    "Storage Replacement": "90-Day",
    "Controller Cleaning": "30-Day",
    "Analog Stick Replacement": "90-Day",
    "Joy-Con Repair": "90-Day",
}

rows = []
id_counter = 1
for category, services in category_services.items():
    for service in services:
        template = service_templates[service]
        for manufacturer in manufacturers_by_category[category]:
            labor_id = f"NSLC-{id_counter:03d}"
            id_counter += 1
            device_scope = device_scopes[category]
            std = template["minutes"]
            # Slight randomness-like adjustment based on device type and service complexity
            if category == "Computers" and service in [
                "SSD Upgrade",
                "RAM Upgrade",
                "Fan Replacement",
                "Thermal Paste Replacement",
                "Laptop Screen Replacement",
                "Hinge Repair",
            ]:
                std += 5
            if category == "Gaming" and service in [
                "HDMI Port Repair",
                "Console Cleaning",
            ]:
                std += 5
            if category == "Wearables" and service in [
                "Battery Replacement",
                "Charging Port Cleaning",
            ]:
                std -= 5
            minimum = max(5, std - 10)
            maximum = std + 15
            if service == "Diagnostic":
                maximum = std + 10
            if service in [
                "Screen Replacement",
                "Laptop Screen Replacement",
                "Back Glass Replacement",
                "Hinge Repair",
            ]:
                maximum = std + 20
            if service == "Battery Replacement" and category == "Wearables":
                minimum = max(5, std - 5)
                maximum = std + 10
            notes = template["notes"]
            if category in ["Tablets", "Computers"] and service == "Diagnostic":
                notes = "Standard diagnostic service for tablets and computers; actual time varies with startup and component verification."
            if category == "Gaming" and service in [
                "Controller Cleaning",
                "Analog Stick Replacement",
                "Joy-Con Repair",
            ]:
                notes = "General controller service; variability depends on controller model and disassembly requirements."
            row = [
                labor_id,
                category,
                manufacturer,
                device_scope,
                service,
                template["difficulty"],
                template["skill"],
                std,
                minimum,
                maximum,
                template["includes"][0],
                template["includes"][1],
                template["includes"][2],
                template["includes"][3],
                template["includes"][4],
                template["tier"],
                warranty_map.get(service, template["warranty"]),
                template["status"],
                template["confidence"],
                template["source"],
                notes,
                date.today(),
                date.today(),
            ]
            rows.append(row)

for row in rows:
    ws_lab.append(row)

# Add table to Labor Standards
end_row = ws_lab.max_row
end_col = ws_lab.max_column
table_range = f"A1:{get_column_letter(end_col)}{end_row}"
table = Table(displayName="tblLaborStandards", ref=table_range)
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style
ws_lab.add_table(table)
ws_lab.auto_filter.ref = table_range

# Add data validation rules based on Reference Lists
list_names = {
    "Device Category": (1, len(reference_lists[0][1]) + 1),
    "Manufacturer": (2, len(reference_lists[1][1]) + 1),
    "Service": (3, len(reference_lists[2][1]) + 1),
    "Repair Difficulty": (4, len(reference_lists[3][1]) + 1),
    "Skill Level": (5, len(reference_lists[4][1]) + 1),
    "Yes/No": (6, len(reference_lists[5][1]) + 1),
    "Labor Rate Tier": (7, len(reference_lists[6][1]) + 1),
    "Recommended Warranty": (8, len(reference_lists[7][1]) + 1),
    "Status": (9, len(reference_lists[8][1]) + 1),
    "Confidence": (10, len(reference_lists[9][1]) + 1),
    "Source Type": (11, len(reference_lists[10][1]) + 1),
}

validation_columns = {
    "B": "Device Category",
    "C": "Manufacturer",
    "E": "Service",
    "F": "Repair Difficulty",
    "G": "Skill Level",
    "K": "Yes/No",
    "L": "Yes/No",
    "M": "Yes/No",
    "N": "Yes/No",
    "O": "Yes/No",
    "P": "Labor Rate Tier",
    "Q": "Recommended Warranty",
    "R": "Status",
    "S": "Confidence",
    "T": "Source Type",
}

for col, name in validation_columns.items():
    col_letter = col
    header_index, end_row_ref = list_names[name]
    start = 2
    ref = f"'00 - Reference Lists'!${get_column_letter(header_index)}$2:${get_column_letter(header_index)}${end_row_ref}"
    dv = DataValidation(type="list", formula1=ref, showDropDown=True, allow_blank=False)
    dv.error = f"Select a valid {name} from the reference list."
    dv.errorTitle = "Invalid entry"
    ws_lab.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{end_row}")

# Format labor sheet columns
whole_number = "0"
date_format = "MM/DD/YYYY"
for col, header in enumerate(headers, start=1):
    col_letter = get_column_letter(col)
    if header in ["Standard Minutes", "Minimum Minutes", "Maximum Minutes"]:
        for row in range(2, end_row + 1):
            ws_lab[f"{col_letter}{row}"].number_format = whole_number
    if header in ["Effective Date", "Last Reviewed"]:
        for row in range(2, end_row + 1):
            ws_lab[f"{col_letter}{row}"].number_format = date_format
    ws_lab.column_dimensions[col_letter].width = 16

# Labor Rate Tiers sheet
rate_headers = [
    "Labor Rate Tier",
    "Skill Level",
    "Hourly Rate",
    "Description",
    "Active",
    "Effective Date",
]
ws_rate.append(rate_headers)
rate_rows = [
    [
        "L1 Basic",
        "Basic",
        75.00,
        "Basic labor rate for standard diagnostics and simple repairs.",
        "Yes",
        date.today(),
    ],
    [
        "L2 Standard",
        "Standard",
        85.00,
        "Standard labor rate for common repairs and service work.",
        "Yes",
        date.today(),
    ],
    [
        "L3 Advanced",
        "Advanced",
        100.00,
        "Advanced labor rate for more complex repairs and upgrades.",
        "Yes",
        date.today(),
    ],
    [
        "L4 Expert",
        "Expert",
        125.00,
        "Expert labor rate reserved for advanced diagnostics and planned future work.",
        "Yes",
        date.today(),
    ],
]
for row in rate_rows:
    ws_rate.append(row)
ws_rate.freeze_panes = "A2"
end_row_rate = ws_rate.max_row
end_col_rate = ws_rate.max_column
rate_table_ref = f"A1:{get_column_letter(end_col_rate)}{end_row_rate}"
rate_table = Table(displayName="tblLaborRateTiers", ref=rate_table_ref)
rate_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True
)
ws_rate.add_table(rate_table)
ws_rate.auto_filter.ref = rate_table_ref
for row in range(2, end_row_rate + 1):
    ws_rate[f"C{row}"].number_format = "$#,##0.00"
    ws_rate[f"F{row}"].number_format = date_format
for col_letter in ["A", "B", "C", "D", "E", "F"]:
    ws_rate.column_dimensions[col_letter].width = 18

# Summary sheet formulas
ws_summary.append(["Metric", "Value"])
ws_summary.append(["Total Labor Standards", "=COUNTA(tblLaborStandards[Labor ID])"])
ws_summary.append(
    ["Average Standard Minutes", "=AVERAGE(tblLaborStandards[Standard Minutes])"]
)
ws_summary.append(
    ["Minimum Standard Minutes", "=MIN(tblLaborStandards[Standard Minutes])"]
)
ws_summary.append(
    ["Maximum Standard Minutes", "=MAX(tblLaborStandards[Standard Minutes])"]
)
ws_summary.append([None, None])
ws_summary.append(["Device Category", "Count"])
summary_start = ws_summary.max_row + 1
for category in reference_lists[0][1]:
    ws_summary.append(
        [category, f'=COUNTIF(tblLaborStandards[Device Category], "{category}")']
    )
ws_summary.append([None, None])
ws_summary.append(["Skill Level", "Count"])
for level in reference_lists[4][1]:
    ws_summary.append([level, f'=COUNTIF(tblLaborStandards[Skill Level], "{level}")'])
ws_summary.append([None, None])
ws_summary.append(["Repair Difficulty", "Count"])
for difficulty in reference_lists[3][1]:
    ws_summary.append(
        [difficulty, f'=COUNTIF(tblLaborStandards[Repair Difficulty], "{difficulty}")']
    )
ws_summary.append([None, None])
ws_summary.append(["Status", "Count"])
for status in reference_lists[8][1]:
    ws_summary.append([status, f'=COUNTIF(tblLaborStandards[Status], "{status}")'])
ws_summary.append([None, None])
ws_summary.append(["Confidence", "Count"])
for confidence in reference_lists[9][1]:
    ws_summary.append(
        [confidence, f'=COUNTIF(tblLaborStandards[Confidence], "{confidence}")']
    )
ws_summary.freeze_panes = "A2"
for col_letter in ["A", "B"]:
    ws_summary.column_dimensions[col_letter].width = 30

# Revision History sheet
history_headers = [
    "Version",
    "Date",
    "Change Type",
    "Description",
    "Prepared By",
    "Approved By",
    "Notes",
]
ws_history.append(history_headers)
ws_history.append(
    [
        "1.0 Draft",
        date.today(),
        "Initial Creation",
        "Provisional labor standards catalog for initial review and validation.",
        "Tamara Grandoit",
        "",
        "Draft includes generalized time entries and planned summary formatting.",
    ]
)
ws_history.freeze_panes = "A2"
end_row_history = ws_history.max_row
end_col_history = ws_history.max_column
history_table_ref = f"A1:{get_column_letter(end_col_history)}{end_row_history}"
history_table = Table(
    displayName="tblLaborCatalogRevisionHistory", ref=history_table_ref
)
history_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True
)
ws_history.add_table(history_table)
ws_history.auto_filter.ref = history_table_ref
ws_history.column_dimensions["A"].width = 14
ws_history.column_dimensions["B"].width = 12
ws_history.column_dimensions["C"].width = 18
ws_history.column_dimensions["D"].width = 48
ws_history.column_dimensions["E"].width = 18
ws_history.column_dimensions["F"].width = 18
ws_history.column_dimensions["G"].width = 30
for row in range(2, end_row_history + 1):
    ws_history[f"B{row}"].number_format = date_format

# Save workbook
wb.save(file_path)

# Validation
wb2 = load_workbook(file_path, data_only=True)
sheets = wb2.sheetnames
assert len(sheets) == 5, f"Expected 5 sheets, found {len(sheets)}: {sheets}"
assert sheets == [
    "00 - Reference Lists",
    "01 - Labor Standards",
    "02 - Labor Rate Tiers",
    "03 - Service Summary",
    "04 - Revision History",
], f"Unexpected worksheet order or names: {sheets}"

# Confirm tables exist
expected_tables = {
    "tblLaborStandards",
    "tblLaborRateTiers",
    "tblLaborCatalogRevisionHistory",
}
table_names = set()
for ws in wb2.worksheets:
    if hasattr(ws, "tables"):
        table_names.update(ws.tables.keys())
assert expected_tables.issubset(
    table_names
), f"Missing table names: {expected_tables - table_names}"

ws_lab2 = wb2["01 - Labor Standards"]
ids = []
rows_checked = 0
for row in ws_lab2.iter_rows(
    min_row=2,
    max_row=ws_lab2.max_row,
    min_col=1,
    max_col=len(headers),
    values_only=True,
):
    if row[0] is None:
        continue
    rows_checked += 1
    ids.append(row[0])
    std = row[7]
    minimum = row[8]
    maximum = row[9]
    assert (
        std is not None and std > 0
    ), f"Standard Minutes invalid in row {rows_checked + 1}: {std}"
    assert (
        minimum is not None and maximum is not None
    ), f"Missing min/max in row {rows_checked + 1}"
    assert (
        minimum <= std <= maximum
    ), f"Time bounds invalid in row {rows_checked + 1}: min={minimum}, std={std}, max={maximum}"
    for idx, value in enumerate(row):
        if idx in [7, 8, 9, 21, 22]:
            continue
        assert value not in (
            None,
            "",
        ), f"Blank required cell at row {rows_checked + 1}, column {idx + 1}"
assert len(ids) == len(set(ids)), "Labor IDs are not unique"

# Confirm labor rate tiers exist
ws_rate2 = wb2["02 - Labor Rate Tiers"]
rate_tiers = {
    row[0] for row in ws_rate2.iter_rows(min_row=2, values_only=True) if row[0]
}
for tier in ["L1 Basic", "L2 Standard", "L3 Advanced", "L4 Expert"]:
    assert tier in rate_tiers, f"Missing labor rate tier: {tier}"

# Confirm every labor rate tier in standards exists in rate table
rate_tiers_upper = set(rate_tiers)
for row in ws_lab2.iter_rows(min_row=2, min_col=16, max_col=16, values_only=True):
    if row[0] is None:
        continue
    assert (
        row[0] in rate_tiers_upper
    ), f"Labor Rate Tier {row[0]} not found in tblLaborRateTiers"

print(f"Generated {file_path}")
print(f"Total Labor Standards rows: {rows_checked}")
print(f"Validated sheets: {sheets}")
print(f"Validated tables: {sorted(table_names)}")
print(f"Labor IDs unique: {len(ids)} entries")
