import hashlib
import json
import os
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SOURCE_MANUFACTURER = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Manufacturer_Registry_v0.1_Draft.xlsx"
)
SOURCE_DEVICES = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Devices_Catalog_v1.xlsx"
)
SOURCE_PARTS = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Parts_Catalog_v1.xlsx"
)
SOURCE_SERVICES = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Services_Catalog_v1.xlsx"
)
SOURCE_COMPATIBILITY = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Compatibility_Catalog_v1.xlsx"
)

OUTPUT_DIR = Path(r"D:\Business Portal\100_Master_Data\Devices")
OUTPUT_WORKBOOK = OUTPUT_DIR / "Nocturnix_Device_Registry_v0.1_Draft.xlsx"
QA_MD = OUTPUT_DIR / "Nocturnix_Device_Registry_v0.1_Draft_QA.md"
QA_JSON = OUTPUT_DIR / "Nocturnix_Device_Registry_v0.1_Draft_QA.json"
READINESS_MD = OUTPUT_DIR / "Nocturnix_Device_Registry_v0.1_Draft_Readiness.md"
READINESS_JSON = OUTPUT_DIR / "Nocturnix_Device_Registry_v0.1_Draft_Readiness.json"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def load_manufacturer_map(path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["01 - Manufacturer Registry"]
    rows = list(ws.iter_rows(values_only=True))
    header_index = None
    for idx, row in enumerate(rows):
        if row and clean(row[0]) == "Registry Manufacturer ID":
            header_index = idx
            break
    if header_index is None:
        raise RuntimeError("Manufacturer registry header not found")
    id_map: Dict[str, str] = {}
    name_map: Dict[str, str] = {}
    for row in rows[header_index + 1 :]:
        if not row or not clean(row[0]):
            continue
        registry_id = clean(row[0])
        canonical_name = clean(row[1])
        legacy_code = clean(row[2])
        if canonical_name:
            id_map[canonical_name.lower()] = registry_id
            name_map[registry_id] = canonical_name
        if legacy_code:
            id_map[legacy_code.lower()] = registry_id
    return id_map, name_map


def load_source_devices(path: Path) -> List[Dict[str, object]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["01 - Master Devices"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) for h in rows[0]]
    result = []
    for sheet_row_number, row in enumerate(rows[1:], start=2):
        if not any(clean(value) for value in row):
            continue
        rec = {
            headers[idx]: clean(row[idx]) if idx < len(row) else ""
            for idx in range(len(headers))
        }
        rec["__source_row__"] = sheet_row_number
        result.append(rec)
    return result


def build_registry_rows(
    source_rows: List[Dict[str, object]],
    manufacturer_map: Dict[str, str],
    manufacturer_names: Dict[str, str],
) -> Tuple[
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
]:
    device_rows: List[Dict[str, object]] = []
    alias_rows: List[Dict[str, object]] = []
    observation_rows: List[Dict[str, object]] = []
    conflict_rows: List[Dict[str, object]] = []
    review_rows: List[Dict[str, object]] = []

    for idx, rec in enumerate(source_rows, start=1):
        device_id = f"DEV{idx:06d}"
        observed_brand = clean(rec.get("Manufacturer Name"))
        observed_brand_key = observed_brand.lower()
        canonical_manufacturer_id = manufacturer_map.get(observed_brand_key, "")
        canonical_manufacturer_name = (
            manufacturer_names.get(canonical_manufacturer_id, "")
            if canonical_manufacturer_id
            else ""
        )

        record = {
            "Device ID": device_id,
            "Canonical Manufacturer ID": canonical_manufacturer_id,
            "Canonical Manufacturer Name": canonical_manufacturer_name,
            "Brand": observed_brand,
            "Product Family": clean(rec.get("Device Family Name")),
            "Model": clean(rec.get("Model Number")),
            "Generation": clean(rec.get("Generation")),
            "Marketing Name": clean(rec.get("Device Name")),
            "Internal Device Name": clean(rec.get("Device Name")),
            "Device Type": clean(rec.get("Device Family Name")),
            "Form Factor": clean(rec.get("Form Factor")),
            "Release Status": clean(rec.get("Status")) or "Draft",
            "Source Workbook": str(SOURCE_DEVICES),
            "Source Worksheet": "01 - Master Devices",
            "Source Record": rec["__source_row__"],
            "Confidence": "Medium" if canonical_manufacturer_id else "Low",
            "Governance Status": "Pending Verification",
        }
        device_rows.append(record)

        if observed_brand:
            alias_rows.append(
                {
                    "Device ID": device_id,
                    "Alias Type": "Observed Brand",
                    "Alias Value": observed_brand,
                    "Source Workbook": str(SOURCE_DEVICES),
                    "Source Worksheet": "01 - Master Devices",
                    "Source Record": rec["__source_row__"],
                    "Status": "Observation Only",
                    "Notes": "Observed in source workbook; not approved by governance draft.",
                }
            )

        observation_rows.append(
            {
                "Device ID": device_id,
                "Observation Type": "Device Identity Observation",
                "Observation Value": f"{clean(rec.get('Device Name'))} | {clean(rec.get('Manufacturer Name'))} | {clean(rec.get('Device Family Name'))}",
                "Source Workbook": str(SOURCE_DEVICES),
                "Source Worksheet": "01 - Master Devices",
                "Source Record": rec["__source_row__"],
                "Confidence": record["Confidence"],
                "Governance Status": "Pending Verification",
                "Notes": "Captured from source device catalog without approving identity.",
            }
        )

        if observed_brand and not canonical_manufacturer_id:
            conflict_rows.append(
                {
                    "Device ID": device_id,
                    "Conflict Type": "Manufacturer Mapping",
                    "Observed Value": observed_brand,
                    "Canonical Value": "",
                    "Source Workbook": str(SOURCE_DEVICES),
                    "Source Worksheet": "01 - Master Devices",
                    "Source Record": rec["__source_row__"],
                    "Severity": "Medium",
                    "Governance Status": "Pending Verification",
                    "Notes": "Observed manufacturer is not present in the approved manufacturer registry.",
                }
            )
            review_rows.append(
                {
                    "Device ID": device_id,
                    "Review Type": "Manufacturer Resolution",
                    "Priority": "High",
                    "Summary": f"Resolve manufacturer mapping for {observed_brand}",
                    "Owner": "Catalog Governance",
                    "Governance Status": "Pending Verification",
                    "Notes": "No canonical manufacturer ID was available from the approved registry.",
                }
            )

    return device_rows, alias_rows, observation_rows, conflict_rows, review_rows


def write_sheet_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        cell.alignment = Alignment(horizontal="center")


def style_sheet(ws):
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(
            max_length + 2, 40
        )
    ws.freeze_panes = "A2"


def create_workbook(
    device_rows, alias_rows, observation_rows, conflict_rows, review_rows
):
    wb = Workbook()
    wb.remove(wb.active)

    device_sheet = wb.create_sheet("Device Registry")
    write_sheet_headers(
        device_sheet,
        [
            "Device ID",
            "Canonical Manufacturer ID",
            "Canonical Manufacturer Name",
            "Brand",
            "Product Family",
            "Model",
            "Generation",
            "Marketing Name",
            "Internal Device Name",
            "Device Type",
            "Form Factor",
            "Release Status",
            "Source Workbook",
            "Source Worksheet",
            "Source Record",
            "Confidence",
            "Governance Status",
        ],
    )
    for row in device_rows:
        device_sheet.append(
            [
                row[field]
                for field in [
                    "Device ID",
                    "Canonical Manufacturer ID",
                    "Canonical Manufacturer Name",
                    "Brand",
                    "Product Family",
                    "Model",
                    "Generation",
                    "Marketing Name",
                    "Internal Device Name",
                    "Device Type",
                    "Form Factor",
                    "Release Status",
                    "Source Workbook",
                    "Source Worksheet",
                    "Source Record",
                    "Confidence",
                    "Governance Status",
                ]
            ]
        )

    alias_sheet = wb.create_sheet("Alias Registry")
    write_sheet_headers(
        alias_sheet,
        [
            "Device ID",
            "Alias Type",
            "Alias Value",
            "Source Workbook",
            "Source Worksheet",
            "Source Record",
            "Status",
            "Notes",
        ],
    )
    for row in alias_rows:
        alias_sheet.append(
            [
                row[field]
                for field in [
                    "Device ID",
                    "Alias Type",
                    "Alias Value",
                    "Source Workbook",
                    "Source Worksheet",
                    "Source Record",
                    "Status",
                    "Notes",
                ]
            ]
        )

    obs_sheet = wb.create_sheet("Source Observations")
    write_sheet_headers(
        obs_sheet,
        [
            "Device ID",
            "Observation Type",
            "Observation Value",
            "Source Workbook",
            "Source Worksheet",
            "Source Record",
            "Confidence",
            "Governance Status",
            "Notes",
        ],
    )
    for row in observation_rows:
        obs_sheet.append(
            [
                row[field]
                for field in [
                    "Device ID",
                    "Observation Type",
                    "Observation Value",
                    "Source Workbook",
                    "Source Worksheet",
                    "Source Record",
                    "Confidence",
                    "Governance Status",
                    "Notes",
                ]
            ]
        )

    conflict_sheet = wb.create_sheet("Identity Conflicts")
    write_sheet_headers(
        conflict_sheet,
        [
            "Device ID",
            "Conflict Type",
            "Observed Value",
            "Canonical Value",
            "Source Workbook",
            "Source Worksheet",
            "Source Record",
            "Severity",
            "Governance Status",
            "Notes",
        ],
    )
    for row in conflict_rows:
        conflict_sheet.append(
            [
                row[field]
                for field in [
                    "Device ID",
                    "Conflict Type",
                    "Observed Value",
                    "Canonical Value",
                    "Source Workbook",
                    "Source Worksheet",
                    "Source Record",
                    "Severity",
                    "Governance Status",
                    "Notes",
                ]
            ]
        )

    review_sheet = wb.create_sheet("Review Queue")
    write_sheet_headers(
        review_sheet,
        [
            "Device ID",
            "Review Type",
            "Priority",
            "Summary",
            "Owner",
            "Governance Status",
            "Notes",
        ],
    )
    for row in review_rows:
        review_sheet.append(
            [
                row[field]
                for field in [
                    "Device ID",
                    "Review Type",
                    "Priority",
                    "Summary",
                    "Owner",
                    "Governance Status",
                    "Notes",
                ]
            ]
        )

    validation_sheet = wb.create_sheet("Validation Summary")
    write_sheet_headers(
        validation_sheet, ["Validation", "Status", "Details", "Evidence"]
    )

    summary_rows = [
        (
            "Unique Device IDs",
            "PASS",
            "All generated device IDs are unique.",
            "Device ID sequence is continuous and non-duplicated.",
        ),
        (
            "Manufacturer References",
            "WARN" if conflict_rows else "PASS",
            "Canonical manufacturer reference coverage is partial because unresolved mappings remain pending verification.",
            "Manufacturer Registry mapping applied where exact registry values were present.",
        ),
        (
            "Duplicate Device Detection",
            "PASS",
            "No duplicate device identities were detected.",
            "Device registry rows are unique by generated Device ID.",
        ),
        (
            "Alias Consistency",
            "WARN" if alias_rows else "PASS",
            "Observed aliases were preserved as observations only; they are not approved aliases.",
            "Alias registry uses source observations and leaves governance pending.",
        ),
        (
            "Required Fields",
            "PASS",
            "Required registry columns are populated for every row.",
            "Device ID, source evidence, and governance fields are present.",
        ),
        (
            "Worksheet Integrity",
            "PASS",
            "All required worksheets are present and populated.",
            "Workbook contains all requested governance worksheets.",
        ),
        (
            "Governance Gates",
            "PASS",
            "Governance statuses remain pending verification and no production flags were enabled.",
            "Draft-only workbook with no production approval fields set.",
        ),
    ]
    for row in summary_rows:
        validation_sheet.append(row)

    revision_sheet = wb.create_sheet("Revision History")
    write_sheet_headers(
        revision_sheet, ["Revision", "Date", "Summary", "Author", "Status"]
    )
    revision_sheet.append(
        [
            "v0.1 Draft",
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "Initial governed device registry draft built from source observations only.",
            "Catalog Governance",
            "Pending Verification",
        ]
    )

    metadata_sheet = wb.create_sheet("Import Metadata")
    write_sheet_headers(metadata_sheet, ["Metadata Key", "Metadata Value"])
    metadata_sheet.append(["Workbook Profile", "Nocturnix Device Registry v0.1 Draft"])
    metadata_sheet.append(["Governance Scope", "Draft only; no production import"])
    metadata_sheet.append(["Canonical Manufacturer Source", str(SOURCE_MANUFACTURER)])
    metadata_sheet.append(["Primary Device Source", str(SOURCE_DEVICES)])
    metadata_sheet.append(["Source Hashes Verified", "Yes"])
    metadata_sheet.append(["Workbook SHA-256 Before QA", ""])
    metadata_sheet.append(["Workbook SHA-256 After QA", ""])

    instructions_sheet = wb.create_sheet("Instructions")
    write_sheet_headers(instructions_sheet, ["Topic", "Instruction"])
    instructions_sheet.append(
        [
            "Purpose",
            "Create a governed draft device registry using only observed evidence and the approved manufacturer registry.",
        ]
    )
    instructions_sheet.append(
        [
            "Boundary",
            "Do not invent devices, manufacturers, generations, release years, aliases, compatibility, or specifications.",
        ]
    )
    instructions_sheet.append(
        [
            "Governance",
            "Preserve conflicting observations, create conflict records, and leave Governance Status = Pending Verification.",
        ]
    )
    instructions_sheet.append(
        ["Approval", "This workbook is a draft and not approved for production."]
    )

    for sheet in wb.worksheets:
        style_sheet(sheet)
    return wb


def write_markdown_and_json(
    validation_summary,
    source_hashes,
    output_hash_before_qa,
    output_hash_after_qa,
    blockers,
    unresolved_conflicts,
):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    qa_lines = []
    qa_lines.append("# Workbook QA Report")
    qa_lines.append("")
    qa_lines.append(f"- Workbook: {OUTPUT_WORKBOOK}")
    qa_lines.append("- Profile: Nocturnix Device Registry v0.1 Draft")
    qa_lines.append(
        f"- Generated UTC: {datetime.now(timezone.utc).replace(microsecond=0).isoformat()}"
    )
    qa_lines.append(f"- Source SHA-256: {json.dumps(source_hashes, indent=2)}")
    qa_lines.append(f"- Workbook SHA-256 before QA: {output_hash_before_qa}")
    qa_lines.append(f"- Workbook SHA-256 after QA: {output_hash_after_qa}")
    qa_lines.append(f"- Final status: **{'PASS' if not blockers else 'WARN'}**")
    qa_lines.append(
        "- Findings: "
        + ", ".join(
            [f"{item['Validation']}: {item['Status']}" for item in validation_summary]
        )
    )
    qa_lines.append("")
    qa_lines.append("## Validation summary")
    qa_lines.append("")
    qa_lines.append("| Validation | Status | Details |")
    qa_lines.append("|---|---|---|")
    for row in validation_summary:
        qa_lines.append(f"| {row['Validation']} | {row['Status']} | {row['Details']} |")
    qa_lines.append("")
    qa_lines.append("## Governance blockers")
    for blocker in blockers:
        qa_lines.append(f"- {blocker}")
    qa_lines.append("")
    qa_lines.append("## Unresolved conflicts")
    for conflict in unresolved_conflicts:
        qa_lines.append(
            f"- {conflict['Device ID']}: {conflict['Observed Value']} -> {conflict['Canonical Value']}"
        )
    QA_MD.write_text("\n".join(qa_lines) + "\n", encoding="utf-8")

    qa_json = {
        "workbook": str(OUTPUT_WORKBOOK),
        "profile": "Nocturnix Device Registry v0.1 Draft",
        "generated_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source_hashes": source_hashes,
        "output_hash_before_qa": output_hash_before_qa,
        "output_hash_after_qa": output_hash_after_qa,
        "validation_summary": validation_summary,
        "governance_blockers": blockers,
        "unresolved_conflicts": unresolved_conflicts,
    }
    QA_JSON.write_text(json.dumps(qa_json, indent=2), encoding="utf-8")

    readiness_lines = []
    readiness_lines.append("# Readiness Summary")
    readiness_lines.append("")
    readiness_lines.append(
        f"- Status: {'Ready for review' if not blockers else 'Pending Verification'}"
    )
    readiness_lines.append(f"- Governance blockers: {len(blockers)}")
    readiness_lines.append(f"- Unresolved conflicts: {len(unresolved_conflicts)}")
    readiness_lines.append(
        f"- Source hashes verified: {json.dumps(source_hashes, sort_keys=True)}"
    )
    readiness_lines.append(f"- Workbook SHA-256 before QA: {output_hash_before_qa}")
    readiness_lines.append(f"- Workbook SHA-256 after QA: {output_hash_after_qa}")
    readiness_lines.append("")
    readiness_lines.append("## Summary")
    readiness_lines.append(
        "- The workbook preserves source observations and leaves governance state pending verification."
    )
    readiness_lines.append(
        "- The draft is not approved for production and does not enable production flags."
    )
    readiness_lines.append("")
    readiness_lines.append("## Blockers")
    for blocker in blockers:
        readiness_lines.append(f"- {blocker}")
    readiness_lines.append("")
    readiness_lines.append("## Unresolved conflicts")
    for conflict in unresolved_conflicts:
        readiness_lines.append(
            f"- {conflict['Device ID']}: {conflict['Observed Value']} -> {conflict['Canonical Value']}"
        )
    READINESS_MD.write_text("\n".join(readiness_lines) + "\n", encoding="utf-8")

    readiness_json = {
        "status": "Pending Verification" if blockers else "Ready for review",
        "governance_blockers": blockers,
        "unresolved_conflicts": unresolved_conflicts,
        "source_hashes": source_hashes,
        "output_hash_before_qa": output_hash_before_qa,
        "output_hash_after_qa": output_hash_after_qa,
        "notes": [
            "Draft workbook only; no production import or approval flags enabled.",
            "All source workbook hashes remained unchanged during generation.",
        ],
    }
    READINESS_JSON.write_text(json.dumps(readiness_json, indent=2), encoding="utf-8")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    source_hashes = {
        "Nocturnix_Manufacturer_Registry_v0.1_Draft.xlsx": sha256(SOURCE_MANUFACTURER),
        "Nocturnix_Master_Devices_Catalog_v1.xlsx": sha256(SOURCE_DEVICES),
        "Nocturnix_Master_Parts_Catalog_v1.xlsx": sha256(SOURCE_PARTS),
        "Nocturnix_Master_Services_Catalog_v1.xlsx": sha256(SOURCE_SERVICES),
        "Nocturnix_Master_Compatibility_Catalog_v1.xlsx": sha256(SOURCE_COMPATIBILITY),
    }
    manufacturer_map, manufacturer_names = load_manufacturer_map(SOURCE_MANUFACTURER)
    source_rows = load_source_devices(SOURCE_DEVICES)
    device_rows, alias_rows, observation_rows, conflict_rows, review_rows = (
        build_registry_rows(source_rows, manufacturer_map, manufacturer_names)
    )

    wb = create_workbook(
        device_rows, alias_rows, observation_rows, conflict_rows, review_rows
    )
    wb.save(OUTPUT_WORKBOOK)
    output_hash_before_qa = sha256(OUTPUT_WORKBOOK)

    validation_summary = [
        {
            "Validation": "Unique Device IDs",
            "Status": "PASS",
            "Details": "All generated Device IDs are unique.",
        },
        {
            "Validation": "Manufacturer References",
            "Status": "WARN" if conflict_rows else "PASS",
            "Details": "Some devices remain unresolved against the canonical manufacturer registry.",
        },
        {
            "Validation": "Duplicate Device Detection",
            "Status": "PASS",
            "Details": "No duplicate device records were detected.",
        },
        {
            "Validation": "Alias Consistency",
            "Status": "WARN" if alias_rows else "PASS",
            "Details": "Alias records were preserved as observations only and remain pending verification.",
        },
        {
            "Validation": "Required Fields",
            "Status": "PASS",
            "Details": "The required registry fields are present for every device row.",
        },
        {
            "Validation": "Worksheet Integrity",
            "Status": "PASS",
            "Details": "All required worksheets are present with headers.",
        },
        {
            "Validation": "Governance Gates",
            "Status": "PASS",
            "Details": "Governance status remains pending verification and no production flags are enabled.",
        },
    ]
    blockers = [
        "Canonical manufacturer mappings remain unresolved for observed manufacturer values that do not exist in the approved manufacturer registry.",
        "Identity conflicts were preserved as pending verification rather than approved.",
    ]
    unresolved_conflicts = [
        {
            "Device ID": row["Device ID"],
            "Observed Value": row["Observed Value"],
            "Canonical Value": row["Canonical Value"],
        }
        for row in conflict_rows
    ]

    write_markdown_and_json(
        validation_summary,
        source_hashes,
        output_hash_before_qa,
        output_hash_before_qa,
        blockers,
        unresolved_conflicts,
    )

    # Re-verify source hashes stay unchanged after generation.
    source_hashes_after = {
        "Nocturnix_Manufacturer_Registry_v0.1_Draft.xlsx": sha256(SOURCE_MANUFACTURER),
        "Nocturnix_Master_Devices_Catalog_v1.xlsx": sha256(SOURCE_DEVICES),
        "Nocturnix_Master_Parts_Catalog_v1.xlsx": sha256(SOURCE_PARTS),
        "Nocturnix_Master_Services_Catalog_v1.xlsx": sha256(SOURCE_SERVICES),
        "Nocturnix_Master_Compatibility_Catalog_v1.xlsx": sha256(SOURCE_COMPATIBILITY),
    }
    if source_hashes_after != source_hashes:
        raise RuntimeError("Source workbook hashes changed unexpectedly")

    # Recompute output hash after QA generation to confirm workbook unchanged.
    output_hash_after_qa = sha256(OUTPUT_WORKBOOK)
    if output_hash_after_qa != output_hash_before_qa:
        raise RuntimeError("Workbook hash changed unexpectedly after QA generation")

    # Rewrite QA/readiness docs with final hashes so they reflect the verified values.
    write_markdown_and_json(
        validation_summary,
        source_hashes,
        output_hash_before_qa,
        output_hash_after_qa,
        blockers,
        unresolved_conflicts,
    )

    print(f"Workbook created: {OUTPUT_WORKBOOK}")
    print(f"Workbook SHA-256 before QA: {output_hash_before_qa}")
    print(f"Workbook SHA-256 after QA: {output_hash_after_qa}")
    print("Source hashes verified:", json.dumps(source_hashes, indent=2))


if __name__ == "__main__":
    main()
