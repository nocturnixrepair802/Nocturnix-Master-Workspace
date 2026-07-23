"""Independently validate the Master Compatibility Catalog v1 workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
DEVICES_PATH = WORKING_DIR / "Nocturnix_Master_Devices_Catalog_v1.xlsx"
SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
PARTS_PATH = WORKING_DIR / "Nocturnix_Master_Parts_Catalog_v1.xlsx"
PRICING_PATH = WORKING_DIR / "Nocturnix_Master_Pricing_Catalog_v1.xlsx"
PROPOSAL_PATH = (
    WORKING_DIR / "Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = WORKING_DIR / "Nocturnix_Master_Compatibility_Catalog_v1.xlsx"
PROTECTED_PATHS = (
    DEVICES_PATH,
    SERVICES_PATH,
    PARTS_PATH,
    PRICING_PATH,
    PROPOSAL_PATH,
    CANONICAL_PATH,
)
CANONICAL_SHEET = "35 Compatibility Matrix"
IMPORT_BATCH_ID = "MASTER-COMPATIBILITY-V1-REVIEW"
COMPATIBILITY_ID_PATTERN = re.compile(r"^CMP\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Compatibility Records",
    "02 - Relationship Types",
    "03 - Compatibility Levels",
    "04 - Evidence Types",
    "05 - Devices",
    "06 - Services",
    "07 - Parts",
    "08 - Family Relationships",
    "09 - Model Relationships",
    "10 - Variant Relationships",
    "11 - Unresolved Review",
    "12 - Validation Summary",
    "13 - Revision History",
    "14 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterCompatibilityInstructions",
    "01 - Compatibility Records": "tblMasterCompatibilityCatalog",
    "02 - Relationship Types": "tblCompatibilityRelationshipTypes",
    "03 - Compatibility Levels": "tblCompatibilityLevels",
    "04 - Evidence Types": "tblCompatibilityEvidenceTypes",
    "05 - Devices": "tblCompatibilityDevices",
    "06 - Services": "tblCompatibilityServices",
    "07 - Parts": "tblCompatibilityParts",
    "08 - Family Relationships": "tblCompatibilityFamilyReview",
    "09 - Model Relationships": "tblCompatibilityModelReview",
    "10 - Variant Relationships": "tblCompatibilityVariantReview",
    "11 - Unresolved Review": "tblCompatibilityUnresolvedReview",
    "12 - Validation Summary": "tblMasterCompatibilityValidation",
    "13 - Revision History": "tblMasterCompatibilityRevisionHistory",
    "14 - Import Metadata": "tblMasterCompatibilityImportMetadata",
}
COMPATIBILITY_HEADERS = [
    "Compatibility ID",
    "Relationship Type",
    "Device ID",
    "Device Family Code",
    "Device Variant",
    "Service ID",
    "Part ID",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Name",
    "Service Name",
    "Part Name",
    "Compatibility Level",
    "Compatibility Status",
    "Evidence Type",
    "Evidence Source",
    "Evidence Detail",
    "Confidence",
    "Requires Manual Review",
    "Active",
    "Effective Date",
    "Expiration Date",
    "Review Status",
    "Reviewer",
    "Reviewer Notes",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Created At",
    "Updated At",
]
UNRESOLVED_HEADERS = [
    "Candidate Type",
    "Device ID",
    "Device Family Code",
    "Device Name",
    "Service ID",
    "Service Name",
    "Part ID",
    "Part Name",
    "Missing Evidence",
    "Ambiguity Reason",
    "Required Action",
    "Review Priority",
    "Review Status",
    "Reviewer Notes",
]
RELATIONSHIP_TYPES = {
    "Device to Service",
    "Device to Part",
    "Device Family to Service",
    "Device Family to Part",
    "Device Variant to Service",
    "Device Variant to Part",
}
COMPATIBILITY_LEVELS = {
    "Family Level",
    "Model Level",
    "Variant Level",
    "Universal",
    "Not Applicable",
    "Unresolved",
}
EVIDENCE_TYPES = {
    "Explicit Source Match",
    "Canonical Relationship",
    "Exact Model Match",
    "Exact Manufacturer and Model",
    "Family-Level Evidence",
    "Legacy Name Evidence",
    "Manual Research Required",
    "No Reliable Evidence",
}
COMPATIBILITY_STATUSES = {
    "Pending Review",
    "Proposed",
    "Confirmed",
    "Rejected",
    "Archived",
}
REVIEW_STATUSES = {
    "Pending Review",
    "Pending Evidence Review",
    "Pending Device Review",
    "Pending Service Review",
    "Pending Part Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
}
CONFIDENCE_VALUES = {"Unassessed", "Low", "Medium", "High"}
YES_NO_VALUES = {"Yes", "No"}
REQUIRED_DEFINED_NAMES = {
    "DV_RelationshipTypes",
    "DV_CompatibilityLevels",
    "DV_CompatibilityStatuses",
    "DV_EvidenceTypes",
    "DV_ReviewStatuses",
    "DV_DeviceIDs",
    "DV_DeviceFamilyCodes",
    "DV_ServiceIDs",
    "DV_PartIDs",
    "DV_ManufacturerIDs",
    "DV_ConfidenceValues",
    "DV_YesNo",
}
DEFINED_NAME_BY_HEADER = {
    "Relationship Type": "DV_RelationshipTypes",
    "Device ID": "DV_DeviceIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Service ID": "DV_ServiceIDs",
    "Part ID": "DV_PartIDs",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Compatibility Level": "DV_CompatibilityLevels",
    "Compatibility Status": "DV_CompatibilityStatuses",
    "Evidence Type": "DV_EvidenceTypes",
    "Confidence": "DV_ConfidenceValues",
    "Requires Manual Review": "DV_YesNo",
    "Active": "DV_YesNo",
    "Review Status": "DV_ReviewStatuses",
}
MODEL_NOISE = {
    "assembly",
    "battery",
    "camera",
    "charging",
    "digitizer",
    "front",
    "glass",
    "lcd",
    "oled",
    "port",
    "repair",
    "replacement",
    "screen",
    "service",
}
PLACEHOLDER_MANUFACTURERS = {
    "",
    "n a",
    "na",
    "none",
    "unknown",
    "tbd",
}
PROHIBITED_HEADERS = {
    "Price",
    "Cost",
    "Supplier Cost",
    "Stock",
    "Quantity",
    "Inventory",
    "Final Approval",
    "Final Customer Price",
}


class CompatibilityValidationError(RuntimeError):
    """Raised when the compatibility review workbook violates its contract."""


def text(value: Any) -> str:
    """Normalize a scalar to stripped text."""
    return "" if value is None else str(value).strip()


def file_hash(path: Path) -> str:
    """Return a SHA-256 digest without modifying a file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require each protected input to exist and be nonempty."""
    for path in paths:
        if not path.is_file():
            raise CompatibilityValidationError(f"Required input is missing: {path}")
        if path.stat().st_size == 0:
            raise CompatibilityValidationError(f"Required input is empty: {path}")


def require_output(path: Path) -> None:
    """Reject missing, empty, or structurally corrupt generated output."""
    if not path.is_file():
        raise CompatibilityValidationError(
            f"Generated workbook is missing: {path}. Run the generator first."
        )
    if path.stat().st_size == 0:
        raise CompatibilityValidationError(f"Generated workbook is empty: {path}")
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    try:
        with zipfile.ZipFile(path) as archive:
            missing = required - set(archive.namelist())
    except zipfile.BadZipFile as exc:
        raise CompatibilityValidationError(
            f"Generated workbook is not a valid Excel ZIP: {path}"
        ) from exc
    if missing:
        raise CompatibilityValidationError(
            f"Generated workbook lacks OOXML members: {sorted(missing)}"
        )


def read_records(
    path: Path,
    sheet_name: str,
    *,
    header_label: str | None = None,
    keep_vba: bool = False,
) -> list[dict[str, Any]]:
    """Read a protected worksheet without saving it."""
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=keep_vba,
    )
    try:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header_index = 0
    if header_label is not None:
        header_index = next(
            (
                index
                for index, row in enumerate(rows)
                if header_label in {text(value) for value in row}
            ),
            -1,
        )
        if header_index < 0:
            raise CompatibilityValidationError(
                f"{path.name}/{sheet_name} lacks {header_label!r}"
            )
    headers = [text(value) for value in rows[header_index]]
    return [
        {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        for row in rows[header_index + 1 :]
        if any(text(value) for value in row)
    ]


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Read table headers in physical order."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, _ = range_boundaries(table.ref)
    return [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read nonblank rows from an Excel Table."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    records: list[dict[str, Any]] = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def canonical_identity() -> tuple[set[str], list[str], int]:
    """Independently inspect canonical Compatibility IDs."""
    records = read_records(
        CANONICAL_PATH,
        CANONICAL_SHEET,
        header_label="Compatibility ID",
        keep_vba=True,
    )
    populated = [
        text(record.get("Compatibility ID"))
        for record in records
        if text(record.get("Compatibility ID"))
    ]
    valid = [value for value in populated if COMPATIBILITY_ID_PATTERN.fullmatch(value)]
    duplicates = sorted(
        value for value, count in Counter(valid).items() if count > 1
    )
    if duplicates:
        raise CompatibilityValidationError(
            f"Duplicate canonical Compatibility IDs: {duplicates}"
        )
    malformed = sorted(set(populated) - set(valid))
    return set(valid), malformed, max(
        (int(value[3:]) for value in valid),
        default=0,
    )


def normalized(value: Any) -> str:
    """Return conservative lowercase alphanumeric text."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).casefold()))


def model_signature(value: Any) -> tuple[str, ...]:
    """Return exact model tokens after removing repair-context noise."""
    return tuple(
        token
        for token in re.findall(r"[a-z0-9]+", text(value).casefold())
        if token not in MODEL_NOISE and token != "for"
    )


def manufacturer_agrees(
    device: dict[str, Any],
    target: dict[str, Any],
) -> bool:
    """Require explicit manufacturer agreement for precise matching."""
    device_id = normalized(device.get("Manufacturer ID"))
    target_id = normalized(target.get("Manufacturer ID"))
    if device_id and target_id:
        return device_id == target_id
    device_names = {
        normalized(device.get("Manufacturer Name")),
        normalized(device.get("Product Line")),
    } - PLACEHOLDER_MANUFACTURERS
    target_name = normalized(target.get("Manufacturer Name"))
    return target_name not in PLACEHOLDER_MANUFACTURERS and target_name in device_names


def expected_target(
    target_kind: str,
    target: dict[str, Any],
    devices: Sequence[dict[str, Any]],
    excluded_rows: set[str],
) -> tuple[tuple[str, ...] | None, str]:
    """Independently derive the expected relationship key or unresolved state."""
    target_id_field = f"{target_kind} ID"
    target_id = text(target.get(target_id_field))
    if not target_id:
        return None, "unresolved"
    if (
        text(target.get("Status")) in {"Rejected", "Archived"}
        or text(target.get("Source Record Number")) in excluded_rows
    ):
        return None, "unresolved"
    explicit_device_id = text(target.get("Device ID"))
    if explicit_device_id:
        matches = [
            device
            for device in devices
            if text(device.get("Device ID")) == explicit_device_id
        ]
        if len(matches) != 1:
            return None, "unresolved"
        device = matches[0]
        if text(device.get("Variant")):
            return (
                f"Device Variant to {target_kind}",
                explicit_device_id,
                target_id,
                text(device.get("Variant")),
            ), "Variant Level"
        return (
            f"Device to {target_kind}",
            explicit_device_id,
            target_id,
            "",
        ), "Model Level"
    source_model = (
        target.get("Device Model")
        or target.get("Device Name")
        or target.get("Model Number")
    )
    signature = model_signature(source_model)
    exact = [
        device
        for device in devices
        if signature
        and signature
        in {
            model_signature(device.get("Device Name")),
            model_signature(device.get("Device Display Name")),
            model_signature(device.get("Model Number")),
        }
        and text(target.get("Device Family Code"))
        == text(device.get("Device Family Code"))
        and manufacturer_agrees(device, target)
    ]
    if len(exact) > 1:
        return None, "unresolved"
    if len(exact) == 1:
        device = exact[0]
        device_id = text(device.get("Device ID"))
        variant = text(device.get("Variant"))
        if variant:
            return (
                f"Device Variant to {target_kind}",
                device_id,
                target_id,
                variant,
            ), "Variant Level"
        return (
            f"Device to {target_kind}",
            device_id,
            target_id,
            "",
        ), "Model Level"
    family = text(target.get("Device Family Code"))
    if family:
        return (
            f"Device Family to {target_kind}",
            family,
            target_id,
            "",
        ), "Family Level"
    return None, "unresolved"


def record_key(record: dict[str, Any]) -> tuple[str, ...]:
    """Return the governed relationship key."""
    return (
        text(record["Relationship Type"]),
        text(record["Device ID"]) or text(record["Device Family Code"]),
        text(record["Service ID"]) or text(record["Part ID"]),
        text(record["Device Variant"]),
    )


def validate_structure(workbook: Any) -> None:
    """Validate exact sheets, tables, filters, and primary schemas."""
    if workbook.sheetnames != SHEET_NAMES:
        raise CompatibilityValidationError(
            f"Worksheet contract differs: {workbook.sheetnames}"
        )
    if len(set(workbook.sheetnames)) != len(workbook.sheetnames):
        raise CompatibilityValidationError("Worksheet names are duplicated")
    if any(len(name) > 31 for name in workbook.sheetnames):
        raise CompatibilityValidationError("A worksheet name exceeds 31 characters")
    all_tables: list[str] = []
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        expected_table = TABLE_NAMES[sheet_name]
        if expected_table not in worksheet.tables:
            raise CompatibilityValidationError(
                f"{sheet_name} lacks {expected_table}"
            )
        if len(worksheet.tables) != 1:
            raise CompatibilityValidationError(
                f"{sheet_name} must contain exactly one table"
            )
        if worksheet.freeze_panes != "A2":
            raise CompatibilityValidationError(
                f"{sheet_name} does not freeze its header"
            )
        table = worksheet.tables[expected_table]
        if not table.autoFilter:
            raise CompatibilityValidationError(
                f"{sheet_name}/{expected_table} lacks filters"
            )
        all_tables.extend(worksheet.tables)
    if len(all_tables) != len(set(all_tables)):
        raise CompatibilityValidationError("Table names are not globally unique")
    if table_headers(
        workbook["01 - Compatibility Records"],
        "tblMasterCompatibilityCatalog",
    ) != COMPATIBILITY_HEADERS:
        raise CompatibilityValidationError("Primary 31-column schema differs")
    if table_headers(
        workbook["11 - Unresolved Review"],
        "tblCompatibilityUnresolvedReview",
    ) != UNRESOLVED_HEADERS:
        raise CompatibilityValidationError("Unresolved Review schema differs")
    if PROHIBITED_HEADERS & set(COMPATIBILITY_HEADERS):
        raise CompatibilityValidationError("Primary schema contains prohibited fields")


def validate_names_and_validations(workbook: Any) -> None:
    """Validate named lookup ranges and prohibit direct cross-sheet formulas."""
    names = set(workbook.defined_names)
    missing = REQUIRED_DEFINED_NAMES - names
    if missing:
        raise CompatibilityValidationError(
            f"Required defined names are missing: {sorted(missing)}"
        )
    for name in REQUIRED_DEFINED_NAMES:
        definition = workbook.defined_names[name]
        destinations = list(definition.destinations)
        if len(destinations) != 1:
            raise CompatibilityValidationError(
                f"Defined name {name} must have one destination"
            )
        sheet_name, coordinates = destinations[0]
        if sheet_name not in workbook.sheetnames or not coordinates:
            raise CompatibilityValidationError(
                f"Defined name {name} has an invalid destination"
            )
    worksheet = workbook["01 - Compatibility Records"]
    expected = {
        f"={name}" for name in set(DEFINED_NAME_BY_HEADER.values())
    }
    observed: set[str] = set()
    for validation in worksheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        formula = text(validation.formula1)
        if "!" in formula or "," in formula:
            raise CompatibilityValidationError(
                f"Direct or hard-coded list validation is prohibited: {formula}"
            )
        if formula not in expected:
            raise CompatibilityValidationError(
                f"Unexpected list validation formula: {formula}"
            )
        observed.add(formula)
    if observed != expected:
        raise CompatibilityValidationError(
            f"List validations differ; observed {sorted(observed)}"
        )


def validate_identity(
    records: Sequence[dict[str, Any]],
    existing: set[str],
    highest: int,
) -> None:
    """Validate Compatibility IDs, order, overlap, and reconciliation."""
    ids = [text(record["Compatibility ID"]) for record in records]
    if any(not COMPATIBILITY_ID_PATTERN.fullmatch(value) for value in ids):
        raise CompatibilityValidationError("A generated Compatibility ID is malformed")
    if len(ids) != len(set(ids)):
        raise CompatibilityValidationError("Generated Compatibility IDs duplicate")
    if set(ids) & existing:
        raise CompatibilityValidationError("Generated IDs overlap canonical IDs")
    expected = [
        f"CMP{highest + offset:06d}"
        for offset in range(1, len(records) + 1)
    ]
    if ids != expected:
        raise CompatibilityValidationError(
            "Compatibility IDs are not continuous from the correct first ID"
        )
    keys = [record_key(record) for record in records]
    if len(keys) != len(set(keys)):
        raise CompatibilityValidationError("Relationship keys duplicate")
    sort_keys = [
        (
            text(row["Relationship Type"]),
            text(row["Device ID"]),
            text(row["Service ID"]) or text(row["Part ID"]),
            text(row["Source Record Number"]).zfill(12),
            text(row["Device Family Code"]),
        )
        for row in records
    ]
    if sort_keys != sorted(sort_keys):
        raise CompatibilityValidationError(
            "Deterministic relationship ordering is not preserved"
        )


def validate_population(
    records: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
    devices: Sequence[dict[str, Any]],
    services: Sequence[dict[str, Any]],
    parts: Sequence[dict[str, Any]],
    excluded_rows: set[str],
) -> None:
    """Recalculate population and validate relationship and evidence integrity."""
    expected: dict[tuple[str, ...], str] = {}
    expected_unresolved = 0
    for kind, targets in (("Service", services), ("Part", parts)):
        for target in targets:
            key, level = expected_target(kind, target, devices, excluded_rows)
            if key is None:
                expected_unresolved += 1
            else:
                if key in expected:
                    raise CompatibilityValidationError(
                        f"Independent expected relationship duplicates: {key}"
                    )
                expected[key] = level
    actual = {record_key(record): record for record in records}
    if set(actual) != set(expected):
        missing = sorted(set(expected) - set(actual))
        extra = sorted(set(actual) - set(expected))
        raise CompatibilityValidationError(
            f"Relationship population differs; missing={missing}, extra={extra}"
        )
    if len(unresolved) != expected_unresolved:
        raise CompatibilityValidationError(
            "Unresolved Review count does not reconcile to source candidates"
        )
    device_map = {
        text(row["Device ID"]): row
        for row in devices
        if text(row.get("Device ID"))
    }
    service_map = {
        text(row["Service ID"]): row
        for row in services
        if text(row.get("Service ID"))
    }
    part_map = {
        text(row["Part ID"]): row
        for row in parts
        if text(row.get("Part ID"))
    }
    for key, row in actual.items():
        relationship_type = text(row["Relationship Type"])
        level = text(row["Compatibility Level"])
        if relationship_type not in RELATIONSHIP_TYPES:
            raise CompatibilityValidationError(f"Invalid relationship type: {key}")
        if level not in COMPATIBILITY_LEVELS or level != expected[key]:
            raise CompatibilityValidationError(f"Invalid relationship level: {key}")
        if text(row["Compatibility Status"]) != "Proposed":
            raise CompatibilityValidationError(f"Generated row is not Proposed: {key}")
        if text(row["Review Status"]) not in REVIEW_STATUSES:
            raise CompatibilityValidationError(f"Invalid review status: {key}")
        if text(row["Review Status"]) in {"Approved", "Ready for Approval"}:
            raise CompatibilityValidationError(f"Generated row is approved: {key}")
        if text(row["Active"]) != "No":
            raise CompatibilityValidationError(f"Generated row is active: {key}")
        if text(row["Requires Manual Review"]) != "Yes":
            raise CompatibilityValidationError(f"Review gate is absent: {key}")
        if text(row["Evidence Type"]) not in EVIDENCE_TYPES:
            raise CompatibilityValidationError(f"Invalid evidence type: {key}")
        if not text(row["Evidence Source"]) or not text(row["Evidence Detail"]):
            raise CompatibilityValidationError(f"Evidence is incomplete: {key}")
        if text(row["Confidence"]) not in CONFIDENCE_VALUES:
            raise CompatibilityValidationError(f"Invalid confidence: {key}")
        is_service = relationship_type.endswith("Service")
        if is_service:
            if text(row["Part ID"]) or text(row["Service ID"]) not in service_map:
                raise CompatibilityValidationError(f"Invalid Service target: {key}")
            if text(row["Service Name"]) != text(
                service_map[text(row["Service ID"])].get("Service Name")
            ):
                raise CompatibilityValidationError(f"Service name differs: {key}")
        else:
            if text(row["Service ID"]) or text(row["Part ID"]) not in part_map:
                raise CompatibilityValidationError(f"Invalid Part target: {key}")
            if text(row["Part Name"]) != text(
                part_map[text(row["Part ID"])].get("Part Name")
            ):
                raise CompatibilityValidationError(f"Part name differs: {key}")
        if level == "Family Level":
            if text(row["Device ID"]) or not text(row["Device Family Code"]):
                raise CompatibilityValidationError(f"Invalid family endpoint: {key}")
            if text(row["Evidence Type"]) != "Family-Level Evidence":
                raise CompatibilityValidationError(f"Family evidence differs: {key}")
        else:
            device_id = text(row["Device ID"])
            if device_id not in device_map:
                raise CompatibilityValidationError(f"Device ID is invalid: {key}")
            device = device_map[device_id]
            if text(row["Device Name"]) != text(device.get("Device Name")):
                raise CompatibilityValidationError(f"Device name differs: {key}")
            if text(row["Device Family Code"]) != text(
                device.get("Device Family Code")
            ):
                raise CompatibilityValidationError(f"Device family differs: {key}")
            if level == "Variant Level" and not text(row["Device Variant"]):
                raise CompatibilityValidationError(f"Variant is missing: {key}")
            if text(row["Evidence Type"]) == "Family-Level Evidence":
                raise CompatibilityValidationError(
                    f"Family-only evidence produced a precise record: {key}"
                )
        if text(row["Import Batch ID"]) != IMPORT_BATCH_ID:
            raise CompatibilityValidationError(f"Import Batch ID differs: {key}")
        source_number = text(row["Source Record Number"])
        if source_number and source_number in excluded_rows:
            raise CompatibilityValidationError(f"Excluded row re-entered: {key}")


def validate_queues(
    workbook: Any,
    records: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
) -> None:
    """Require one correct review-queue row per primary or unresolved row."""
    queue_specs = {
        "08 - Family Relationships": "Family Level",
        "09 - Model Relationships": "Model Level",
        "10 - Variant Relationships": "Variant Level",
    }
    seen: list[tuple[str, ...]] = []
    for sheet_name, level in queue_specs.items():
        rows = table_records(workbook[sheet_name], TABLE_NAMES[sheet_name])
        expected = [
            record_key(row)
            for row in records
            if text(row["Compatibility Level"]) == level
        ]
        actual = [record_key(row) for row in rows]
        if actual != expected:
            raise CompatibilityValidationError(
                f"{sheet_name} does not reconcile in source order"
            )
        seen.extend(actual)
    if Counter(seen) != Counter(record_key(row) for row in records):
        raise CompatibilityValidationError("Relationship queues duplicate or omit rows")
    unresolved_keys = [
        (
            text(row["Candidate Type"]),
            text(row["Device ID"]),
            text(row["Service ID"]) or text(row["Part ID"]),
        )
        for row in unresolved
    ]
    if len(unresolved_keys) != len(set(unresolved_keys)):
        raise CompatibilityValidationError("Unresolved Review contains duplicates")
    for row in unresolved:
        if text(row["Review Status"]) != "Pending Evidence Review":
            raise CompatibilityValidationError(
                "Unresolved Review contains a reviewed outcome"
            )


def validate_metadata(
    workbook: Any,
    hashes: dict[Path, str],
    existing: set[str],
    malformed: Sequence[str],
    highest: int,
    records: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
) -> None:
    """Validate embedded namespace and protected-source metadata."""
    rows = table_records(
        workbook["14 - Import Metadata"],
        "tblMasterCompatibilityImportMetadata",
    )
    metadata = {text(row["Metadata Field"]): text(row["Value"]) for row in rows}
    expected = {
        "Import Batch ID": IMPORT_BATCH_ID,
        "Namespace Authority": "ADR-010",
        "Schema Columns": str(len(COMPATIBILITY_HEADERS)),
        "Canonical Valid ID Count": str(len(existing)),
        "Highest Canonical Compatibility ID": (
            f"CMP{highest:06d}" if highest else ""
        ),
        "Malformed Canonical IDs": "; ".join(malformed),
        "First Generated ID": (
            text(records[0]["Compatibility ID"]) if records else ""
        ),
        "Final Generated ID": (
            text(records[-1]["Compatibility ID"]) if records else ""
        ),
        "Generated Relationship Count": str(len(records)),
        "Unresolved Candidate Count": str(len(unresolved)),
        "Canonical Import Authorized": "No",
    }
    for field, value in expected.items():
        if metadata.get(field) != value:
            raise CompatibilityValidationError(
                f"Import Metadata differs for {field}: {metadata.get(field)!r}"
            )
    for path, digest in hashes.items():
        if metadata.get(f"Protected Input Path: {path.name}") != str(path):
            raise CompatibilityValidationError(f"Protected path missing for {path}")
        if metadata.get(f"SHA-256: {path.name}") != digest:
            raise CompatibilityValidationError(f"Protected hash differs for {path}")


def load_sources() -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    set[str],
]:
    """Load protected source records independently."""
    devices = read_records(DEVICES_PATH, "01 - Master Devices")
    services = read_records(SERVICES_PATH, "01 - Master Services")
    parts = read_records(PARTS_PATH, "01 - Master Parts")
    exclusions = read_records(PROPOSAL_PATH, "02 - Duplicate Exclusions")
    excluded_rows = {
        text(row.get("Source Row Number"))
        for row in exclusions
        if text(row.get("Source Row Number"))
    }
    return devices, services, parts, excluded_rows


def main() -> int:
    """Run complete independent validation without writing any workbook."""
    try:
        require_files(PROTECTED_PATHS)
        require_output(OUTPUT_PATH)
        before_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        devices, services, parts, excluded_rows = load_sources()
        existing, malformed, highest = canonical_identity()
        workbook = load_workbook(OUTPUT_PATH, data_only=False)
        try:
            validate_structure(workbook)
            validate_names_and_validations(workbook)
            records = table_records(
                workbook["01 - Compatibility Records"],
                "tblMasterCompatibilityCatalog",
            )
            unresolved = table_records(
                workbook["11 - Unresolved Review"],
                "tblCompatibilityUnresolvedReview",
            )
            validate_identity(records, existing, highest)
            validate_population(
                records,
                unresolved,
                devices,
                services,
                parts,
                excluded_rows,
            )
            validate_queues(workbook, records, unresolved)
            validate_metadata(
                workbook,
                before_hashes,
                existing,
                malformed,
                highest,
                records,
                unresolved,
            )
        finally:
            workbook.close()
        after_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        if after_hashes != before_hashes:
            raise CompatibilityValidationError(
                "A protected input changed during validation"
            )
    except (
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        CompatibilityValidationError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    type_counts = Counter(text(row["Relationship Type"]) for row in records)
    level_counts = Counter(text(row["Compatibility Level"]) for row in records)
    print(f"Validated: {OUTPUT_PATH}")
    print(f"Schema columns: {len(COMPATIBILITY_HEADERS)}")
    print(f"Generated relationships: {len(records)}")
    print(f"Unresolved candidates: {len(unresolved)}")
    print(f"Relationship counts: {dict(sorted(type_counts.items()))}")
    print(f"Level counts: {dict(sorted(level_counts.items()))}")
    print(f"Malformed canonical Compatibility IDs: {len(malformed)}")
    print("Workbook structure, evidence, relationships, and queues: PASS")
    print("Protected input hashes: PASS")
    print("Canonical, pricing, inventory, and catalog writes: NOT PERFORMED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
