"""Generate the standalone Master Labor Mapping Engine v1 audit report.

The engine reads the proposed Master Services and labor standards workbooks,
scores candidate mappings, and writes a separate audit workbook. It never saves
either input workbook and never updates the canonical database.
"""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

MASTER_SERVICES_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Services_Catalog_v1.xlsx"
)
LABOR_CATALOG_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Labor_Mapping_Audit_v1.xlsx"
)

MASTER_SHEET = "01 - Master Services"
LABOR_SHEET = "01 - Labor Standards"
EXPECTED_SERVICE_ROWS = 314
APPROVED_THRESHOLD = 0.82
MATERIAL_TIE_MARGIN = 0.03
MAPPING_RESULTS = {"Mapped", "Pending Labor Mapping", "Ambiguous"}
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")

REQUIRED_SERVICE_HEADERS = {
    "Service ID",
    "Legacy Service SKU",
    "Manufacturer Name",
    "Device Family Name",
    "Device Model",
    "Repair Type",
    "Service Name",
    "Labor Standard ID",
    "Repair Difficulty",
    "Skill Level",
    "Source Record Number",
}
REQUIRED_LABOR_HEADERS = {
    "Labor ID",
    "Device Category",
    "Manufacturer",
    "Device Scope",
    "Service",
    "Repair Difficulty",
    "Skill Level",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Labor Rate Tier",
}

AUDIT_HEADERS = [
    "Source Record Number",
    "Service ID",
    "Legacy Service SKU",
    "Manufacturer",
    "Device Family",
    "Repair Type",
    "Service Name",
    "Device Model",
    "Existing Labor Standard ID",
    "Proposed Labor Standard ID",
    "Match Score",
    "Second Best Score",
    "Score Margin",
    "Match Evidence",
    "Mapping Result",
    "Matched Standard Minutes",
    "Matched Minimum Minutes",
    "Matched Maximum Minutes",
    "Matched Labor Tier",
    "Matched Repair Difficulty",
    "Matched Skill Level",
]

FEATURE_WEIGHTS = {
    "Manufacturer": 0.15,
    "Device Family": 0.12,
    "Repair Type": 0.15,
    "Service Name": 0.30,
    "Device Model Keywords": 0.08,
    "Labor Category": 0.08,
    "Repair Difficulty": 0.06,
    "Skill Level": 0.06,
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DATE_FORMAT = "yyyy-mm-dd hh:mm:ss"


class MappingError(RuntimeError):
    """Raised when a safe deterministic mapping report cannot be produced."""


@dataclass(frozen=True)
class CandidateScore:
    """A scored labor candidate with auditable feature evidence."""

    labor: dict[str, Any]
    score: float
    evidence: str


def ascii_value(value: Any) -> Any:
    """Return ASCII-safe strings while preserving typed non-string values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u00a0": " ",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return value.encode("ascii", "replace").decode("ascii")


def text(value: Any) -> str:
    """Return a stripped ASCII-safe representation."""
    return "" if value is None else str(ascii_value(value)).strip()


def normalized(value: Any) -> str:
    """Return lowercase ASCII words suitable for deterministic comparison."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).lower()))


def tokens(value: Any) -> set[str]:
    """Return meaningful normalized tokens."""
    ignored = {
        "a",
        "an",
        "and",
        "for",
        "of",
        "repair",
        "replacement",
        "service",
        "the",
        "to",
    }
    return {token for token in normalized(value).split() if token not in ignored}


def sha256_file(path: Path) -> str:
    """Return a streaming SHA-256 hash."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require nonempty Excel ZIP containers before analysis."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise MappingError(f"Required file(s) missing: {', '.join(missing)}")
    empty = [str(path) for path in paths if path.stat().st_size <= 0]
    if empty:
        raise MappingError(f"Required file(s) empty: {', '.join(empty)}")
    invalid = [str(path) for path in paths if not zipfile.is_zipfile(path)]
    if invalid:
        raise MappingError(
            f"Required file(s) are not valid Excel ZIP containers: {', '.join(invalid)}"
        )


def read_records(
    path: Path,
    sheet_name: str,
    required_headers: set[str],
) -> list[dict[str, Any]]:
    """Read a worksheet as dictionaries without saving its workbook."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise MappingError(f"Missing worksheet {sheet_name!r} in {path}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        try:
            headers = [text(value) for value in next(rows)]
        except StopIteration as exc:
            raise MappingError(f"Worksheet {sheet_name!r} is empty") from exc
        if not headers or any(not header for header in headers):
            raise MappingError(f"Worksheet {sheet_name!r} has invalid headers")
        missing_headers = sorted(required_headers - set(headers))
        if missing_headers:
            raise MappingError(
                f"Worksheet {sheet_name!r} is missing required columns: "
                f"{', '.join(missing_headers)}"
            )
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    finally:
        workbook.close()


def similarity(left: Any, right: Any) -> float:
    """Combine normalized text similarity with token overlap."""
    left_text = normalized(left)
    right_text = normalized(right)
    if not left_text or not right_text:
        return 0.0
    if left_text == right_text:
        return 1.0
    if left_text in right_text or right_text in left_text:
        return 0.9
    left_tokens = tokens(left_text)
    right_tokens = tokens(right_text)
    union = left_tokens | right_tokens
    token_score = len(left_tokens & right_tokens) / len(union) if union else 0.0
    sequence_score = SequenceMatcher(None, left_text, right_text).ratio()
    return max(token_score, sequence_score)


def expected_labor_category(service: dict[str, Any]) -> str:
    """Map a Master Services device family to the labor category vocabulary."""
    source = normalized(
        " ".join(
            (
                text(service.get("Device Family Name")),
                text(service.get("Device Model")),
                text(service.get("Device Series")),
            )
        )
    )
    mappings = [
        (("tablet", "ipad"), "Tablets"),
        (("computer", "laptop", "desktop", "macbook", "chromebook"), "Computers"),
        (("console", "gaming", "xbox", "playstation", "switch"), "Gaming"),
        (("watch", "wearable"), "Wearables"),
        (("phone", "iphone", "android", "samsung"), "Phones"),
    ]
    for keywords, category in mappings:
        if any(keyword in source for keyword in keywords):
            return category
    return ""


def feature_values(
    service: dict[str, Any], labor: dict[str, Any]
) -> dict[str, tuple[str, float]]:
    """Return available feature evidence and normalized component scores."""
    manufacturer = text(service.get("Manufacturer Name"))
    family = text(service.get("Device Family Name"))
    repair_type = text(service.get("Repair Type"))
    service_name = text(service.get("Service Name"))
    device_model = text(service.get("Device Model"))
    difficulty = text(service.get("Repair Difficulty"))
    skill = text(service.get("Skill Level"))
    category = expected_labor_category(service)

    model_keywords = tokens(device_model)
    labor_model_context = tokens(
        " ".join(
            (
                text(labor.get("Manufacturer")),
                text(labor.get("Device Scope")),
                text(labor.get("Device Category")),
            )
        )
    )
    model_union = model_keywords | labor_model_context
    model_score = (
        len(model_keywords & labor_model_context) / len(model_union)
        if model_union
        else 0.0
    )

    values: dict[str, tuple[str, float]] = {}
    if manufacturer:
        score = similarity(manufacturer, labor.get("Manufacturer"))
        values["Manufacturer"] = (
            f"{manufacturer!r} vs {text(labor.get('Manufacturer'))!r}",
            score,
        )
    if family:
        labor_family = " ".join(
            (
                text(labor.get("Device Scope")),
                text(labor.get("Device Category")),
            )
        )
        values["Device Family"] = (
            f"{family!r} vs {labor_family!r}",
            similarity(family, labor_family),
        )
    if repair_type:
        values["Repair Type"] = (
            f"{repair_type!r} vs {text(labor.get('Service'))!r}",
            similarity(repair_type, labor.get("Service")),
        )
    if service_name:
        values["Service Name"] = (
            f"{service_name!r} vs {text(labor.get('Service'))!r}",
            similarity(service_name, labor.get("Service")),
        )
    if device_model:
        values["Device Model Keywords"] = (
            f"overlap={sorted(model_keywords & labor_model_context)}",
            model_score,
        )
    if category:
        labor_category = text(labor.get("Device Category"))
        values["Labor Category"] = (
            f"{category!r} vs {labor_category!r}",
            1.0 if category == labor_category else 0.0,
        )
    if difficulty:
        values["Repair Difficulty"] = (
            f"{difficulty!r} vs {text(labor.get('Repair Difficulty'))!r}",
            1.0
            if normalized(difficulty) == normalized(labor.get("Repair Difficulty"))
            else 0.0,
        )
    if skill:
        values["Skill Level"] = (
            f"{skill!r} vs {text(labor.get('Skill Level'))!r}",
            1.0 if normalized(skill) == normalized(labor.get("Skill Level")) else 0.0,
        )
    return values


def score_candidate(service: dict[str, Any], labor: dict[str, Any]) -> CandidateScore:
    """Calculate a weighted score normalized over available service evidence."""
    values = feature_values(service, labor)
    available_weight = sum(FEATURE_WEIGHTS[name] for name in values)
    if available_weight == 0:
        return CandidateScore(labor=labor, score=0.0, evidence="No usable evidence")
    earned = sum(
        FEATURE_WEIGHTS[name] * component_score
        for name, (_description, component_score) in values.items()
    )
    score = earned / available_weight
    evidence = "; ".join(
        f"{name}={component_score:.3f} ({description})"
        for name, (description, component_score) in values.items()
    )
    return CandidateScore(labor=labor, score=score, evidence=evidence)


def score_service(
    service: dict[str, Any], labor_rows: Sequence[dict[str, Any]]
) -> dict[str, Any]:
    """Return one deterministic mapping-audit record for a service."""
    candidates = [score_candidate(service, labor) for labor in labor_rows]
    candidates.sort(
        key=lambda candidate: (
            -candidate.score,
            text(candidate.labor.get("Labor ID")),
        )
    )
    if not candidates:
        raise MappingError("Labor catalog contains no candidate rows")
    best = candidates[0]
    second_score = candidates[1].score if len(candidates) > 1 else 0.0
    margin = best.score - second_score
    best_labor_id = text(best.labor.get("Labor ID"))

    if not best_labor_id or best.score <= APPROVED_THRESHOLD:
        result = "Pending Labor Mapping"
        proposed_id = ""
    elif margin <= MATERIAL_TIE_MARGIN:
        result = "Ambiguous"
        proposed_id = ""
    else:
        result = "Mapped"
        proposed_id = best_labor_id

    mapped = result == "Mapped"
    return {
        "Source Record Number": service.get("Source Record Number"),
        "Service ID": text(service.get("Service ID")),
        "Legacy Service SKU": text(service.get("Legacy Service SKU")),
        "Manufacturer": text(service.get("Manufacturer Name")),
        "Device Family": text(service.get("Device Family Name")),
        "Repair Type": text(service.get("Repair Type")),
        "Service Name": text(service.get("Service Name")),
        "Device Model": text(service.get("Device Model")),
        "Existing Labor Standard ID": text(service.get("Labor Standard ID")),
        "Proposed Labor Standard ID": proposed_id,
        "Match Score": best.score,
        "Second Best Score": second_score,
        "Score Margin": margin,
        "Match Evidence": best.evidence,
        "Mapping Result": result,
        "Matched Standard Minutes": best.labor.get("Standard Minutes", "")
        if mapped
        else "",
        "Matched Minimum Minutes": best.labor.get("Minimum Minutes", "")
        if mapped
        else "",
        "Matched Maximum Minutes": best.labor.get("Maximum Minutes", "")
        if mapped
        else "",
        "Matched Labor Tier": text(best.labor.get("Labor Rate Tier")) if mapped else "",
        "Matched Repair Difficulty": text(best.labor.get("Repair Difficulty"))
        if mapped
        else "",
        "Matched Skill Level": text(best.labor.get("Skill Level")) if mapped else "",
    }


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    records: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write a styled, filtered Excel Table."""
    worksheet.append(list(headers))
    for record in records:
        worksheet.append([ascii_value(record.get(header, "")) for header in headers])
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    for column, header in enumerate(headers, start=1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = (
            60 if header == "Match Evidence" else max(12, min(28, len(header) + 3))
        )
        if header in {"Match Score", "Second Best Score", "Score Margin"}:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column).number_format = "0.0000"


def build_report(
    audit_rows: list[dict[str, Any]],
    hashes_before: dict[Path, str],
    generated_at: datetime,
) -> Workbook:
    """Build the standalone mapping audit workbook."""
    counts = Counter(text(row["Mapping Result"]) for row in audit_rows)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "00 - Summary"
    audit = workbook.create_sheet("01 - Mapping Audit")
    metadata = workbook.create_sheet("02 - Import Metadata")

    summary_rows = [
        {"Metric": "Total Master Services", "Value": len(audit_rows)},
        {"Metric": "Mapped", "Value": counts["Mapped"]},
        {"Metric": "Pending Labor Mapping", "Value": counts["Pending Labor Mapping"]},
        {"Metric": "Ambiguous", "Value": counts["Ambiguous"]},
        {"Metric": "Approved Threshold", "Value": APPROVED_THRESHOLD},
        {"Metric": "Material Tie Margin", "Value": MATERIAL_TIE_MARGIN},
        {"Metric": "Source Workbooks Modified", "Value": "No"},
    ]
    append_table(summary, ["Metric", "Value"], summary_rows, "tblLaborMappingSummary")
    append_table(audit, AUDIT_HEADERS, audit_rows, "tblMasterLaborMappingAudit")

    metadata_rows = [
        {"Metadata Field": "Artifact", "Value": "Master Labor Mapping Audit v1"},
        {"Metadata Field": "Generated At UTC", "Value": generated_at},
        {
            "Metadata Field": "Master Services Source",
            "Value": str(MASTER_SERVICES_PATH),
        },
        {"Metadata Field": "Labor Catalog Source", "Value": str(LABOR_CATALOG_PATH)},
        {"Metadata Field": "Output", "Value": str(OUTPUT_PATH)},
    ]
    for path, digest in hashes_before.items():
        metadata_rows.append(
            {"Metadata Field": f"SHA-256: {path.name}", "Value": digest}
        )
    append_table(
        metadata,
        ["Metadata Field", "Value"],
        metadata_rows,
        "tblLaborMappingImportMetadata",
    )
    metadata["B3"].number_format = DATE_FORMAT
    return workbook


def validate_audit_rows(
    services: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
    audit_rows: Sequence[dict[str, Any]],
) -> None:
    """Validate coverage, identity preservation, thresholds, and copied labor values."""
    if (
        len(services) != EXPECTED_SERVICE_ROWS
        or len(audit_rows) != EXPECTED_SERVICE_ROWS
    ):
        raise MappingError("Master Services or audit population is not 314")
    service_ids = [text(row.get("Service ID")) for row in services]
    audit_ids = [text(row.get("Service ID")) for row in audit_rows]
    if audit_ids != service_ids:
        raise MappingError(
            "Audit Service IDs do not preserve source order and identity"
        )
    invalid_service_ids = [
        identifier
        for identifier in service_ids
        if not SERVICE_ID_PATTERN.fullmatch(identifier)
    ]
    if invalid_service_ids:
        raise MappingError(
            f"Master Services contains invalid Service IDs: {invalid_service_ids[:10]}"
        )
    if len(service_ids) != len(set(service_ids)):
        raise MappingError("Master Services contains duplicate Service IDs")

    labor_ids = [text(row.get("Labor ID")) for row in labor_rows]
    if any(not identifier for identifier in labor_ids):
        raise MappingError("Labor catalog contains a blank Labor ID")
    if len(labor_ids) != len(set(labor_ids)):
        duplicates = sorted(
            identifier for identifier, count in Counter(labor_ids).items() if count > 1
        )
        raise MappingError(
            f"Labor catalog contains duplicate Labor IDs: {duplicates[:10]}"
        )
    labor_by_id = dict(zip(labor_ids, labor_rows, strict=True))

    source_numbers = [row.get("Source Record Number") for row in services]
    if any(value in (None, "") for value in source_numbers):
        raise MappingError("Master Services contains a blank Source Record Number")
    try:
        numeric_source_numbers = [int(value) for value in source_numbers]
    except (TypeError, ValueError) as exc:
        raise MappingError(
            "Master Services contains a nonnumeric Source Record Number"
        ) from exc
    if numeric_source_numbers != sorted(numeric_source_numbers):
        raise MappingError("Master Services is not ordered by Source Record Number")
    if len(numeric_source_numbers) != len(set(numeric_source_numbers)):
        raise MappingError("Master Services contains duplicate Source Record Numbers")

    for row_number, row in enumerate(audit_rows, start=2):
        result = text(row.get("Mapping Result"))
        proposed_id = text(row.get("Proposed Labor Standard ID"))
        score = float(row.get("Match Score") or 0)
        margin = float(row.get("Score Margin") or 0)
        if result not in MAPPING_RESULTS:
            raise MappingError(f"Invalid Mapping Result in audit row {row_number}")
        if result == "Mapped":
            if score <= APPROVED_THRESHOLD or margin <= MATERIAL_TIE_MARGIN:
                raise MappingError(f"Unsafe mapped result in audit row {row_number}")
            if proposed_id not in labor_by_id:
                raise MappingError(
                    f"Unknown Labor Standard ID in audit row {row_number}"
                )
            labor = labor_by_id[proposed_id]
            comparisons = {
                "Matched Standard Minutes": labor.get("Standard Minutes", ""),
                "Matched Minimum Minutes": labor.get("Minimum Minutes", ""),
                "Matched Maximum Minutes": labor.get("Maximum Minutes", ""),
                "Matched Labor Tier": text(labor.get("Labor Rate Tier")),
                "Matched Repair Difficulty": text(labor.get("Repair Difficulty")),
                "Matched Skill Level": text(labor.get("Skill Level")),
            }
            if any(row.get(field) != value for field, value in comparisons.items()):
                raise MappingError(
                    f"Mapped labor fields were not copied exactly in audit row {row_number}"
                )
        elif proposed_id:
            raise MappingError(
                f"Unresolved audit row {row_number} contains an invented labor ID"
            )
        if result == "Pending Labor Mapping" and (
            score > APPROVED_THRESHOLD and margin > MATERIAL_TIE_MARGIN
        ):
            raise MappingError(
                f"Eligible audit row {row_number} was left Pending Labor Mapping"
            )
        if result == "Ambiguous" and (
            score <= APPROVED_THRESHOLD or margin > MATERIAL_TIE_MARGIN
        ):
            raise MappingError(f"Invalid ambiguous result in audit row {row_number}")


def validate_reopened_report(
    path: Path,
    source_service_ids: Sequence[str],
    expected_audit_rows: Sequence[dict[str, Any]],
    hashes_before: dict[Path, str],
) -> list[str]:
    """Reopen the report and confirm its structure and protected inputs."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        expected_sheets = [
            "00 - Summary",
            "01 - Mapping Audit",
            "02 - Import Metadata",
        ]
        if workbook.sheetnames != expected_sheets:
            raise MappingError("Audit report worksheet structure is invalid")
        worksheet = workbook["01 - Mapping Audit"]
        if "tblMasterLaborMappingAudit" not in worksheet.tables:
            raise MappingError("tblMasterLaborMappingAudit is missing")
        records = list(worksheet.iter_rows(min_row=2, values_only=True))
        records = [row for row in records if row[0] is not None]
        if len(records) != EXPECTED_SERVICE_ROWS:
            raise MappingError("Reopened audit report does not contain 314 rows")
        reopened = [dict(zip(AUDIT_HEADERS, row, strict=False)) for row in records]
        service_id_column = AUDIT_HEADERS.index("Service ID")
        reopened_ids = [text(row[service_id_column]) for row in records]
        if reopened_ids != list(source_service_ids):
            raise MappingError("Reopened audit report changed Service IDs or order")
        comparison_fields = (
            "Proposed Labor Standard ID",
            "Match Score",
            "Second Best Score",
            "Score Margin",
            "Match Evidence",
            "Mapping Result",
            "Matched Standard Minutes",
            "Matched Minimum Minutes",
            "Matched Maximum Minutes",
            "Matched Labor Tier",
            "Matched Repair Difficulty",
            "Matched Skill Level",
        )
        for row_number, (actual, expected) in enumerate(
            zip(reopened, expected_audit_rows, strict=True),
            start=2,
        ):
            for field in comparison_fields:
                if actual.get(field) != expected.get(field):
                    raise MappingError(
                        f"Reopened audit row {row_number} changed {field}"
                    )
    finally:
        workbook.close()

    hashes_after = {source: sha256_file(source) for source in hashes_before}
    changed = [
        str(source)
        for source in hashes_before
        if hashes_before[source] != hashes_after[source]
    ]
    if changed:
        raise MappingError(f"Protected input hash changed: {', '.join(changed)}")
    return [
        "Audit report structure: PASS (314 rows)",
        "Service IDs and source order: PASS",
        "Protected input hashes: PASS (unchanged)",
    ]


def main() -> int:
    """Generate and validate the standalone labor mapping audit report."""
    try:
        protected = (MASTER_SERVICES_PATH, LABOR_CATALOG_PATH)
        require_files(protected)
        hashes_before = {path: sha256_file(path) for path in protected}
        services = read_records(
            MASTER_SERVICES_PATH,
            MASTER_SHEET,
            REQUIRED_SERVICE_HEADERS,
        )
        labor_rows = read_records(
            LABOR_CATALOG_PATH,
            LABOR_SHEET,
            REQUIRED_LABOR_HEADERS,
        )
        if len(services) != EXPECTED_SERVICE_ROWS:
            raise MappingError(
                f"Expected {EXPECTED_SERVICE_ROWS} Master Services, found {len(services)}"
            )
        if not labor_rows:
            raise MappingError("Labor standards workbook contains no records")

        audit_rows = [score_service(service, labor_rows) for service in services]
        validate_audit_rows(services, labor_rows, audit_rows)
        generated_at = datetime.now(UTC).replace(tzinfo=None)
        workbook = build_report(audit_rows, hashes_before, generated_at)
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(OUTPUT_PATH)
        workbook.close()

        source_ids = [text(row.get("Service ID")) for row in services]
        messages = validate_reopened_report(
            OUTPUT_PATH,
            source_ids,
            audit_rows,
            hashes_before,
        )
        counts = Counter(text(row["Mapping Result"]) for row in audit_rows)
        print(f"Generated: {OUTPUT_PATH}")
        print(f"Master Services preserved: {len(services)}")
        print(f"Mapped: {counts['Mapped']}")
        print(f"Pending Labor Mapping: {counts['Pending Labor Mapping']}")
        print(f"Ambiguous: {counts['Ambiguous']}")
        for message in messages:
            print(message)
        return 0
    except (MappingError, OSError, ValueError, KeyError, IndexError) as exc:
        print(f"ERROR: {ascii_value(str(exc))}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
