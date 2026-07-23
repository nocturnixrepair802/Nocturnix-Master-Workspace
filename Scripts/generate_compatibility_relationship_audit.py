"""Generate a standalone compatibility relationship evidence audit.

The audit is independent from the Master Compatibility workbook and never
writes audit results into it automatically.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
DEVICES_PATH = WORKING_DIR / "Nocturnix_Master_Devices_Catalog_v1.xlsx"
SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
PARTS_PATH = WORKING_DIR / "Nocturnix_Master_Parts_Catalog_v1.xlsx"
PROPOSAL_PATH = (
    WORKING_DIR / "Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = (
    WORKING_DIR / "Nocturnix_Compatibility_Relationship_Audit_v1.xlsx"
)
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(
    f"{OUTPUT_PATH.stem}.tmp{OUTPUT_PATH.suffix}"
)
PROTECTED_PATHS = (
    DEVICES_PATH,
    SERVICES_PATH,
    PARTS_PATH,
    PROPOSAL_PATH,
    CANONICAL_PATH,
)
SHEET_NAME = "01 - Relationship Audit"
TABLE_NAME = "tblCompatibilityRelationshipAudit"
AUDIT_HEADERS = [
    "Candidate Relationship Key",
    "Relationship Type",
    "Device ID",
    "Device Family Code",
    "Device Name",
    "Service ID",
    "Service Name",
    "Part ID",
    "Part Name",
    "Best Evidence Type",
    "Evidence Score",
    "Second Best Score",
    "Score Margin",
    "Evidence Coverage",
    "Match Evidence",
    "Proposed Compatibility Level",
    "Proposed Compatibility Status",
    "Mapping Result",
    "Failure Reason",
    "Requires Manual Review",
]
MAPPING_RESULTS = {
    "Proposed",
    "Ambiguous",
    "Insufficient Evidence",
    "Rejected Candidate",
}
MODEL_NOISE = {
    "assembly",
    "battery",
    "camera",
    "charging",
    "digitizer",
    "front",
    "glass",
    "lcd",
    "oled",
    "port",
    "repair",
    "replacement",
    "screen",
    "service",
}
PLACEHOLDER_MANUFACTURERS = {
    "",
    "n a",
    "na",
    "none",
    "unknown",
    "tbd",
}


class CompatibilityAuditError(RuntimeError):
    """Raised when the relationship audit violates its safety contract."""


def text(value: Any) -> str:
    """Return stripped text, treating None as blank."""
    return "" if value is None else str(value).strip()


def ascii_value(value: Any) -> Any:
    """Make generated text ASCII-safe while preserving typed values."""
    if not isinstance(value, str):
        return value
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
    }
    for source, replacement in replacements.items():
        value = value.replace(source, replacement)
    return value.encode("ascii", "replace").decode("ascii")


def excel_safe_value(value: Any) -> Any:
    """Normalize timezone-aware values before Excel persistence."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def file_hash(path: Path) -> str:
    """Return a SHA-256 digest without changing a protected input."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require each protected input to exist and be nonempty."""
    for path in paths:
        if not path.is_file():
            raise CompatibilityAuditError(f"Required input is missing: {path}")
        if path.stat().st_size == 0:
            raise CompatibilityAuditError(f"Required input is empty: {path}")


def read_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read a first-row-header worksheet without saving it."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    headers = [text(value) for value in rows[0]]
    return [
        {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        for row in rows[1:]
        if any(text(value) for value in row)
    ]


def normalized(value: Any) -> str:
    """Return conservative lowercase alphanumeric text."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).casefold()))


def model_signature(value: Any) -> tuple[str, ...]:
    """Remove only known repair-context tokens from model text."""
    return tuple(
        token
        for token in re.findall(r"[a-z0-9]+", text(value).casefold())
        if token not in MODEL_NOISE and token != "for"
    )


def manufacturer_agrees(
    device: dict[str, Any],
    target: dict[str, Any],
) -> bool:
    """Require explicit ID or nonplaceholder manufacturer-name agreement."""
    device_id = normalized(device.get("Manufacturer ID"))
    target_id = normalized(target.get("Manufacturer ID"))
    if device_id and target_id:
        return device_id == target_id
    device_names = {
        normalized(device.get("Manufacturer Name")),
        normalized(device.get("Product Line")),
    } - PLACEHOLDER_MANUFACTURERS
    target_name = normalized(target.get("Manufacturer Name"))
    return target_name not in PLACEHOLDER_MANUFACTURERS and target_name in device_names


def candidate_key(
    relationship_type: str,
    endpoint: str,
    target_id: str,
    variant: str = "",
) -> str:
    """Build an auditable deterministic candidate key."""
    return "|".join((relationship_type, endpoint, target_id, variant))


def target_audit(
    target_kind: str,
    target: dict[str, Any],
    devices: Sequence[dict[str, Any]],
    excluded_rows: set[str],
) -> dict[str, Any]:
    """Score one target conservatively and retain ambiguity."""
    is_service = target_kind == "Service"
    target_id = text(target.get(f"{target_kind} ID"))
    service_id = target_id if is_service else ""
    part_id = target_id if not is_service else ""
    common: dict[str, Any] = {
        "Device Family Code": text(target.get("Device Family Code")),
        "Service ID": service_id,
        "Service Name": text(target.get("Service Name")) if is_service else "",
        "Part ID": part_id,
        "Part Name": text(target.get("Part Name")) if not is_service else "",
        "Proposed Compatibility Status": "Proposed",
        "Requires Manual Review": "Yes",
    }
    rejected = (
        text(target.get("Status")) in {"Rejected", "Archived"}
        or text(target.get("Source Record Number")) in excluded_rows
    )
    if rejected:
        relationship_type = f"Device to {target_kind}"
        return common | {
            "Candidate Relationship Key": candidate_key(
                relationship_type,
                "",
                target_id,
            ),
            "Relationship Type": relationship_type,
            "Device ID": "",
            "Device Name": "",
            "Best Evidence Type": "No Reliable Evidence",
            "Evidence Score": 0.0,
            "Second Best Score": 0.0,
            "Score Margin": 0.0,
            "Evidence Coverage": "Rejected or excluded source",
            "Match Evidence": "",
            "Proposed Compatibility Level": "Unresolved",
            "Mapping Result": "Rejected Candidate",
            "Failure Reason": "Source record is rejected, archived, or excluded",
        }

    explicit_device_id = text(target.get("Device ID"))
    if explicit_device_id:
        matches = [
            device
            for device in devices
            if text(device.get("Device ID")) == explicit_device_id
        ]
        if len(matches) == 1:
            device = matches[0]
            variant = text(device.get("Variant"))
            level = "Variant Level" if variant else "Model Level"
            relationship_type = (
                f"Device Variant to {target_kind}"
                if variant
                else f"Device to {target_kind}"
            )
            return common | {
                "Candidate Relationship Key": candidate_key(
                    relationship_type,
                    explicit_device_id,
                    target_id,
                    variant,
                ),
                "Relationship Type": relationship_type,
                "Device ID": explicit_device_id,
                "Device Name": text(device.get("Device Name")),
                "Best Evidence Type": "Explicit Source Match",
                "Evidence Score": 1.0,
                "Second Best Score": 0.0,
                "Score Margin": 1.0,
                "Evidence Coverage": "Device ID",
                "Match Evidence": "Exact source Device ID resolves once",
                "Proposed Compatibility Level": level,
                "Mapping Result": "Proposed",
                "Failure Reason": "",
            }
        relationship_type = f"Device to {target_kind}"
        return common | {
            "Candidate Relationship Key": candidate_key(
                relationship_type,
                explicit_device_id,
                target_id,
            ),
            "Relationship Type": relationship_type,
            "Device ID": explicit_device_id,
            "Device Name": "",
            "Best Evidence Type": "Manual Research Required",
            "Evidence Score": 0.0,
            "Second Best Score": 0.0,
            "Score Margin": 0.0,
            "Evidence Coverage": "Unresolved Device ID",
            "Match Evidence": "",
            "Proposed Compatibility Level": "Unresolved",
            "Mapping Result": "Ambiguous",
            "Failure Reason": "Explicit Device ID is missing or duplicated",
        }

    source_model = (
        target.get("Device Model")
        or target.get("Device Name")
        or target.get("Model Number")
    )
    signature = model_signature(source_model)
    scored: list[tuple[float, dict[str, Any], str]] = []
    for device in devices:
        device_signatures = {
            model_signature(device.get("Device Name")),
            model_signature(device.get("Device Display Name")),
            model_signature(device.get("Model Number")),
        }
        if (
            signature
            and signature in device_signatures
            and text(target.get("Device Family Code"))
            == text(device.get("Device Family Code"))
            and manufacturer_agrees(device, target)
        ):
            scored.append(
                (
                    0.95,
                    device,
                    "Exact model tokens, manufacturer, and family",
                )
            )
    scored.sort(
        key=lambda item: (
            -item[0],
            text(item[1].get("Device ID")),
        )
    )
    if scored:
        best_score, device, evidence = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0
        variant = text(device.get("Variant"))
        level = "Variant Level" if variant else "Model Level"
        relationship_type = (
            f"Device Variant to {target_kind}"
            if variant
            else f"Device to {target_kind}"
        )
        ambiguous = len(scored) > 1 and second_score == best_score
        return common | {
            "Candidate Relationship Key": candidate_key(
                relationship_type,
                text(device.get("Device ID")),
                target_id,
                variant,
            ),
            "Relationship Type": relationship_type,
            "Device ID": text(device.get("Device ID")),
            "Device Name": text(device.get("Device Name")),
            "Best Evidence Type": "Exact Manufacturer and Model",
            "Evidence Score": best_score,
            "Second Best Score": second_score,
            "Score Margin": round(best_score - second_score, 4),
            "Evidence Coverage": "Manufacturer; Family; Model",
            "Match Evidence": evidence,
            "Proposed Compatibility Level": (
                "Unresolved" if ambiguous else level
            ),
            "Mapping Result": "Ambiguous" if ambiguous else "Proposed",
            "Failure Reason": (
                "Multiple devices share the same best exact evidence"
                if ambiguous
                else ""
            ),
        }

    family = text(target.get("Device Family Code"))
    if family:
        relationship_type = f"Device Family to {target_kind}"
        return common | {
            "Candidate Relationship Key": candidate_key(
                relationship_type,
                family,
                target_id,
            ),
            "Relationship Type": relationship_type,
            "Device ID": "",
            "Device Name": "",
            "Best Evidence Type": "Family-Level Evidence",
            "Evidence Score": 0.6,
            "Second Best Score": 0.0,
            "Score Margin": 0.6,
            "Evidence Coverage": "Explicit Device Family Code",
            "Match Evidence": "Source explicitly designates device family",
            "Proposed Compatibility Level": "Family Level",
            "Mapping Result": "Proposed",
            "Failure Reason": "",
        }
    relationship_type = f"Device to {target_kind}"
    return common | {
        "Candidate Relationship Key": candidate_key(
            relationship_type,
            "",
            target_id,
        ),
        "Relationship Type": relationship_type,
        "Device ID": "",
        "Device Name": "",
        "Best Evidence Type": "No Reliable Evidence",
        "Evidence Score": 0.0,
        "Second Best Score": 0.0,
        "Score Margin": 0.0,
        "Evidence Coverage": "None",
        "Match Evidence": "",
        "Proposed Compatibility Level": "Unresolved",
        "Mapping Result": "Insufficient Evidence",
        "Failure Reason": "No explicit device, exact model, or family evidence",
    }


def append_table(
    worksheet: Worksheet,
    rows: Sequence[dict[str, Any]],
) -> None:
    """Write the audit table with stable formatting."""
    worksheet.append(AUDIT_HEADERS)
    materialized = list(rows) or [{header: "" for header in AUDIT_HEADERS}]
    for row in materialized:
        worksheet.append(
            [
                excel_safe_value(ascii_value(row.get(header, "")))
                for header in AUDIT_HEADERS
            ]
        )
    end_column = get_column_letter(len(AUDIT_HEADERS))
    table = Table(
        displayName=TABLE_NAME,
        ref=f"A1:{end_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = table.ref
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
    for index, header in enumerate(AUDIT_HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(header) + 2, 12),
            38,
        )


def table_records(worksheet: Worksheet) -> list[dict[str, Any]]:
    """Read nonblank audit table records."""
    table = worksheet.tables[TABLE_NAME]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    rows: list[dict[str, Any]] = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            rows.append(dict(zip(headers, values, strict=True)))
    return rows


def validate_reopened(
    path: Path,
    expected_rows: Sequence[dict[str, Any]],
) -> None:
    """Validate the temporary audit before publication."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        missing = required - set(archive.namelist())
    if missing:
        raise CompatibilityAuditError(
            f"Temporary audit lacks OOXML members: {sorted(missing)}"
        )
    workbook = load_workbook(path, data_only=False)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise CompatibilityAuditError("Audit worksheet contract differs")
        worksheet = workbook[SHEET_NAME]
        if TABLE_NAME not in worksheet.tables:
            raise CompatibilityAuditError("Audit table is missing")
        if worksheet.freeze_panes != "A2":
            raise CompatibilityAuditError("Audit header is not frozen")
        rows = table_records(worksheet)
        if len(rows) != len(expected_rows):
            raise CompatibilityAuditError("Reopened audit row count differs")
        headers = [
            text(cell.value)
            for cell in worksheet[1][: len(AUDIT_HEADERS)]
        ]
        if headers != AUDIT_HEADERS:
            raise CompatibilityAuditError("Reopened audit schema differs")
        keys = [text(row["Candidate Relationship Key"]) for row in rows]
        if len(keys) != len(set(keys)):
            raise CompatibilityAuditError("Audit candidate keys duplicate")
        for row in rows:
            if text(row["Mapping Result"]) not in MAPPING_RESULTS:
                raise CompatibilityAuditError("Audit mapping result is invalid")
            if text(row["Requires Manual Review"]) != "Yes":
                raise CompatibilityAuditError("Audit row bypasses manual review")
            if text(row["Proposed Compatibility Status"]) != "Proposed":
                raise CompatibilityAuditError("Audit row invents approval")
    finally:
        workbook.close()


def build_rows() -> list[dict[str, Any]]:
    """Load protected inputs and calculate deterministic audit rows."""
    devices = read_records(DEVICES_PATH, "01 - Master Devices")
    services = read_records(SERVICES_PATH, "01 - Master Services")
    parts = read_records(PARTS_PATH, "01 - Master Parts")
    exclusions = read_records(PROPOSAL_PATH, "02 - Duplicate Exclusions")
    excluded_rows = {
        text(row.get("Source Row Number"))
        for row in exclusions
        if text(row.get("Source Row Number"))
    }
    rows = [
        target_audit(kind, target, devices, excluded_rows)
        for kind, targets in (("Service", services), ("Part", parts))
        for target in targets
    ]
    rows.sort(
        key=lambda row: (
            text(row["Relationship Type"]),
            text(row["Device ID"]),
            text(row["Service ID"]) or text(row["Part ID"]),
            text(row["Candidate Relationship Key"]),
        )
    )
    keys = [text(row["Candidate Relationship Key"]) for row in rows]
    if len(keys) != len(set(keys)):
        raise CompatibilityAuditError("Calculated audit candidate keys duplicate")
    return rows


def main() -> int:
    """Generate, validate, and atomically publish the standalone audit."""
    workbook: Workbook | None = None
    try:
        require_files(PROTECTED_PATHS)
        before_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        rows = build_rows()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        append_table(worksheet, rows)
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        if TEMP_OUTPUT_PATH.exists():
            TEMP_OUTPUT_PATH.unlink()
        workbook.save(TEMP_OUTPUT_PATH)
        workbook.close()
        workbook = None
        validate_reopened(TEMP_OUTPUT_PATH, rows)
        after_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        if after_hashes != before_hashes:
            raise CompatibilityAuditError(
                "A protected input changed during audit generation"
            )
        os.replace(TEMP_OUTPUT_PATH, OUTPUT_PATH)
        results: dict[str, int] = {}
        for row in rows:
            value = text(row["Mapping Result"])
            results[value] = results.get(value, 0) + 1
        print("Compatibility Relationship Audit V1 validation: PASS")
        print(f"Generated workbook: {OUTPUT_PATH}")
        print(f"Audit rows: {len(rows)}")
        print(f"Mapping results: {dict(sorted(results.items()))}")
        print("Master Compatibility workbook writes: NONE")
        print("Protected input hashes: PASS")
        return 0
    except (
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        CompatibilityAuditError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError as close_error:
                print(f"ERROR: Failed to close workbook: {close_error}", file=sys.stderr)
        try:
            if TEMP_OUTPUT_PATH.exists():
                TEMP_OUTPUT_PATH.unlink()
        except OSError as cleanup_error:
            print(
                f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}",
                file=sys.stderr,
            )


if __name__ == "__main__":
    raise SystemExit(main())
