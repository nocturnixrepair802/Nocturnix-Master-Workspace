"""Workbook validation helpers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .audit import ValidationIssue


def scan_ref_errors(wb: Workbook) -> list[ValidationIssue]:
    """Scan formulas for #REF! references."""
    issues: list[ValidationIssue] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "#REF!" in cell.value:
                    issues.append(ValidationIssue("Formula #REF!", "", cell.coordinate, "", "FAIL", f"{ws.title}!{cell.coordinate} contains #REF!"))
    return issues


def validate_reopen(path: Path, expected_sheet_names: list[str], expected_tables: set[str]) -> list[ValidationIssue]:
    """Save/reopen validation for the output workbook."""
    issues: list[ValidationIssue] = []
    try:
        wb = load_workbook(path, data_only=False)
        if wb.sheetnames != expected_sheet_names:
            issues.append(ValidationIssue("Worksheet names", "", "", "", "FAIL", "Reopened sheet names differ"))
        table_names = {table.name for ws in wb.worksheets for table in ws.tables.values()}
        missing = expected_tables - table_names
        if missing:
            issues.append(ValidationIssue("Source tables", "", "", "", "FAIL", f"Missing tables after reopen: {sorted(missing)}"))
        wb.close()
    except Exception as exc:
        issues.append(ValidationIssue("Workbook reopen", "", "", "", "FAIL", str(exc)))
    return issues


def validation_result(issues: list[ValidationIssue], unresolved: int, duplicates: int) -> str:
    """Return PASS, PASS WITH WARNINGS, or FAIL."""
    if any(issue.result == "FAIL" for issue in issues):
        return "FAIL"
    if unresolved or duplicates:
        return "PASS WITH WARNINGS"
    return "PASS"
