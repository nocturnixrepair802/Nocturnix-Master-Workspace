"""Plan and apply deterministic updates to generated device-related sheets."""

from __future__ import annotations

from datetime import UTC, datetime
from typing import Any

from openpyxl.cell.cell import Cell
from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook

from .audit import CellChange
from .config import ALIASES, AUDIT_SHEETS
from .normalization import format_bool_like, normalize_header, trim_text, valid_id
from .table_reader import table_headers

ENTITY_COLUMNS = {
    "DeviceModel": ("devicemodelid", "DeviceModel"),
    "DeviceFamily": ("devicefamilyid", "DeviceFamily"),
    "Manufacturer": ("manufacturerid", "Manufacturer"),
    "DeviceType": ("devicetypeid", "DeviceType"),
}


def alias_columns(columns: dict[str, int]) -> dict[str, int]:
    """Map normalized workbook headers to canonical columns."""
    mapped = dict(columns)
    normalized_aliases = {target: {normalize_header(a) for a in aliases} for target, aliases in ALIASES.items()}
    for target, aliases in normalized_aliases.items():
        for key, col in columns.items():
            if key in aliases and target not in mapped:
                mapped[target] = col
    return mapped


def device_related(columns: dict[str, int]) -> bool:
    """Return whether columns are device-related."""
    related = {"devicetypeid", "devicetype", "manufacturerid", "manufacturer", "devicefamilyid", "devicefamily", "devicemodelid", "devicemodel", "isactive"}
    return bool(related.intersection(alias_columns(columns)))


def plan_updates(wb: Workbook, maps: dict[str, Any], source_sheet: str) -> tuple[list[CellChange], list[list[Any]], set[str], int, int]:
    """Plan updates for all non-authoritative/generated worksheets."""
    changes: list[CellChange] = []
    exceptions: list[list[Any]] = []
    updated_sheets: set[str] = set()
    matched_by_id = 0
    unknown_ids = 0
    timestamp = datetime.now(UTC).isoformat()
    records = maps["records"]
    for ws in wb.worksheets:
        if ws.title in AUDIT_SHEETS or ws.title == source_sheet:
            continue
        table_items = list(ws.tables.values()) or [None]
        for table in table_items:
            if table is not None:
                header_row, columns, _ = table_headers(ws, table.ref)
                _, _, _, max_row = range_boundaries(table.ref)
                table_name = table.name
            else:
                header_row, columns = discover_header_row(ws)
                max_row = ws.max_row
                table_name = ""
            cols = alias_columns(columns)
            if not cols or not device_related(cols):
                continue
            for row in range(header_row + 1, max_row + 1):
                row_changed = False
                entity, record_id, canonical, match_method = resolve_row(ws, row, cols, records)
                if record_id and canonical is None:
                    unknown_ids += 1
                    exceptions.append(exception_row(ws.title, table_name, row, entity, record_id, ws, cols, "UNKNOWN_ID", f"{record_id} not found in authoritative tables"))
                    continue
                if canonical is None:
                    if any(ws.cell(row, c).value not in (None, "") for c in cols.values()):
                        exceptions.append(exception_row(ws.title, table_name, row, "", "", ws, cols, "UNRESOLVED", "No unique authoritative ID match"))
                    continue
                matched_by_id += 1 if match_method == "ID" else 0
                proposed = canonical_values(entity, canonical, records)
                for key, new_value in proposed.items():
                    if key not in cols or new_value in (None, ""):
                        continue
                    cell = ws.cell(row, cols[key])
                    final_value = format_bool_like(new_value, cell.value) if key == "isactive" else trim_text(new_value)
                    if should_change(cell, final_value):
                        changes.append(CellChange(timestamp, ws.title, table_name, cell.coordinate, row, entity, record_id, key, cell.value, final_value, match_method, "Synchronized from authoritative ID"))
                        row_changed = True
                for key in ("devicetype", "manufacturer", "devicefamily", "devicemodel"):
                    if key in cols:
                        cell = ws.cell(row, cols[key])
                        trimmed = trim_text(cell.value)
                        if should_change(cell, trimmed):
                            changes.append(CellChange(timestamp, ws.title, table_name, cell.coordinate, row, entity, record_id, key, cell.value, trimmed, match_method, "Trimmed text whitespace"))
                            row_changed = True
                if row_changed:
                    updated_sheets.add(ws.title)
    return changes, exceptions, updated_sheets, matched_by_id, unknown_ids


def discover_header_row(ws) -> tuple[int, dict[str, int]]:
    """Find the first likely header row on a plain worksheet."""
    for row in range(1, min(ws.max_row, 20) + 1):
        columns = {}
        for col in range(1, min(ws.max_column, 80) + 1):
            key = normalize_header(ws.cell(row, col).value)
            if key:
                columns[key] = col
        if device_related(columns):
            return row, columns
    return 1, {}


def resolve_row(ws, row: int, cols: dict[str, int], records: dict[str, dict[str, Any]]) -> tuple[str, str, Any, str]:
    """Resolve a target row to an authoritative record by existing valid ID."""
    for key, entity in ENTITY_COLUMNS.values():
        if key in cols:
            value = ws.cell(row, cols[key]).value
            record_id = str(value or "").strip()
            if record_id:
                if valid_id(entity, record_id):
                    return entity, record_id, records.get(entity, {}).get(record_id), "ID"
                return entity, record_id, None, "ID"
    return "", "", None, ""


def canonical_values(entity: str, record: Any, records: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Return values implied by an authoritative record and relationships."""
    allowed = {
        "devicetypeid",
        "devicetype",
        "manufacturerid",
        "manufacturer",
        "devicefamilyid",
        "devicefamily",
        "devicemodelid",
        "devicemodel",
        "modelnumber",
        "isactive",
    }
    vals = {key: value for key, value in record.values.items() if key in allowed}
    if entity == "DeviceModel":
        family = records.get("DeviceFamily", {}).get(str(vals.get("devicefamilyid") or "").strip())
        if family:
            vals.setdefault("manufacturerid", family.values.get("manufacturerid"))
            vals.setdefault("devicetypeid", family.values.get("devicetypeid"))
            vals["devicefamily"] = family.values.get("devicefamily")
    if vals.get("manufacturerid"):
        manufacturer = records.get("Manufacturer", {}).get(str(vals.get("manufacturerid")).strip())
        if manufacturer:
            vals["manufacturer"] = manufacturer.values.get("manufacturer")
    if vals.get("devicetypeid"):
        device_type = records.get("DeviceType", {}).get(str(vals.get("devicetypeid")).strip())
        if device_type:
            vals["devicetype"] = device_type.values.get("devicetype")
    return vals


def should_change(cell: Cell, new_value: Any) -> bool:
    """Return whether a cell can be safely changed."""
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return False
    if new_value in (None, ""):
        return False
    return cell.value != new_value


def apply_changes(wb: Workbook, changes: list[CellChange]) -> None:
    """Apply planned safe changes."""
    for change in changes:
        wb[change.worksheet][change.cell] = change.new_value


def exception_row(sheet: str, table: str, row: int, entity: str, record_id: str, ws, cols: dict[str, int], kind: str, details: str) -> list[Any]:
    """Build an exception report row."""
    def val(key: str) -> Any:
        return ws.cell(row, cols[key]).value if key in cols else ""
    return [sheet, table, row, entity, record_id, val("manufacturer"), val("devicefamily"), val("devicemodel"), val("modelnumber"), kind, details, "Review manually; no automatic merge or reassignment performed"]
