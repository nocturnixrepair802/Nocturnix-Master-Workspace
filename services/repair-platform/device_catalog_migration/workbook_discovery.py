"""Workbook and authoritative table discovery."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook

from .audit import SourceTable
from .config import ENTITY_SPECS
from .normalization import normalize_header
from .table_reader import table_headers


def load(path: Path) -> Workbook:
    """Load workbook preserving formulas."""
    return load_workbook(path, data_only=False)


def inventory_tables(wb: Workbook) -> list[SourceTable]:
    """Inventory every Excel table in the workbook."""
    tables: list[SourceTable] = []
    for ws in wb.worksheets:
        for table in ws.tables.values():
            header_row, columns, _ = table_headers(ws, table.ref)
            _, _, _, max_row = range_boundaries(table.ref)
            entity, confidence, reasons = classify_table(table.name, columns, ws.title)
            tables.append(
                SourceTable(
                    entity_type=entity,
                    worksheet=ws.title,
                    table_name=table.name,
                    ref=table.ref,
                    header_row=header_row,
                    first_data_row=header_row + 1,
                    last_data_row=max_row,
                    row_count=max(0, max_row - header_row),
                    columns=columns,
                    confidence=confidence,
                    reasons=reasons,
                )
            )
    return tables


def classify_table(table_name: str, columns: dict[str, int], worksheet: str) -> tuple[str, float, list[str]]:
    """Classify a table by normalized headers and table context."""
    name = normalize_header(table_name)
    sheet = normalize_header(worksheet)
    best = ("", 0.0, [])
    for entity, spec in ENTITY_SPECS.items():
        required = spec["required"]
        score = len(required.intersection(columns)) / len(required)
        reasons = [f"required headers matched {sorted(required.intersection(columns))}"]
        id_key = spec["id"]
        if id_key in columns:
            score += 0.2
            reasons.append(f"contains {id_key}")
        if normalize_header(entity) in name or normalize_header(entity) in sheet:
            score += 0.1
            reasons.append("table/sheet name context matched")
        if score > best[1]:
            best = (entity if score >= 0.55 else "", min(score, 1.0), reasons)
    return best


def find_authoritative(tables: list[SourceTable], preferred_sheet: str | None = None) -> dict[str, SourceTable]:
    """Pick the four authoritative source tables from discovered tables."""
    result: dict[str, SourceTable] = {}
    for entity in ENTITY_SPECS:
        candidates = [t for t in tables if t.entity_type == entity]
        if preferred_sheet:
            normalized_preferred = normalize_header(preferred_sheet)
            candidates.sort(key=lambda t: (normalize_header(t.worksheet) == normalized_preferred, t.confidence, t.row_count), reverse=True)
        else:
            candidates.sort(key=lambda t: (t.confidence, t.row_count), reverse=True)
        if candidates:
            candidates[0].authoritative = True
            result[entity] = candidates[0]
    return result


def print_inspection(wb: Workbook, tables: list[SourceTable], authoritative: dict[str, SourceTable]) -> None:
    """Print the required read-only workbook inspection."""
    print("Worksheet names:")
    for name in wb.sheetnames:
        print(f"- {name}")
    print("\nExcel tables:")
    for table in tables:
        print(f"- {table.table_name} | worksheet={table.worksheet} | range={table.ref}")
        print(f"  normalized_headers={list(table.columns.keys())}")
    print("\nProposed authoritative tables:")
    for entity, table in authoritative.items():
        print(f"- {entity}: {table.table_name} on {table.worksheet} {table.ref}")
        print(f"  confidence={table.confidence:.2f}; reasons={'; '.join(table.reasons)}")
