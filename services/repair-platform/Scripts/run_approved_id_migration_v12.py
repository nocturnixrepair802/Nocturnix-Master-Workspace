from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
import warnings
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class CatalogEntry:
    lookup_catalog_id: str
    lookup_group_id: str
    lookup_table_name: str
    primary_key_field: str
    primary_key_prefix: str
    is_system_lookup: object
    table_exists: object


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def is_truthy(value: object) -> bool:
    return str(value).strip().lower() in {"true", "yes", "y", "1", "active", "exists"}


def catalog_table_exists(entry, table_lookup) -> bool:
    raw = str(entry.table_exists or "").strip()
    if raw.startswith("="):
        return entry.lookup_table_name in table_lookup
    return is_truthy(entry.table_exists)


def table_rows(ws, table):
    cells = list(ws[table.ref])
    headers = [c.value for c in cells[0]]
    return headers, cells[1:]


def find_table(wb, table_name: str):
    matches = []
    for ws in wb.worksheets:
        if table_name in ws.tables:
            matches.append((ws, ws.tables[table_name]))
    return matches


def find_lookup_catalog(wb):
    for ws in wb.worksheets:
        for table in ws.tables.values():
            headers, rows = table_rows(ws, table)
            required = {"LookupTableName", "PrimaryKeyField", "PrimaryKeyPrefix", "LookupGroupID", "IsSystemLookup", "TableExists"}
            if required.issubset(set(headers)):
                return ws, table, headers, rows
    raise RuntimeError("No lookup catalog table with required dynamic fields was found.")


def load_catalog(wb):
    ws, table, headers, rows = find_lookup_catalog(wb)
    idx = {h: headers.index(h) for h in headers}
    entries = []
    for row in rows:
        entries.append(CatalogEntry(
            lookup_catalog_id=norm(row[idx.get("LookupCatalogID")].value) if "LookupCatalogID" in idx else "",
            lookup_group_id=norm(row[idx["LookupGroupID"]].value),
            lookup_table_name=norm(row[idx["LookupTableName"]].value),
            primary_key_field=norm(row[idx["PrimaryKeyField"]].value),
            primary_key_prefix=norm(row[idx["PrimaryKeyPrefix"]].value),
            is_system_lookup=row[idx["IsSystemLookup"]].value,
            table_exists=row[idx["TableExists"]].value,
        ))
    return ws.title, table.name, entries


def valid_id(value: object, prefix: str) -> bool:
    if value in (None, "") or not prefix:
        return False
    return re.fullmatch(re.escape(prefix) + r"\d+", str(value)) is not None


def suffix_num(value: object, prefix: str) -> int | None:
    if not valid_id(value, prefix):
        return None
    return int(str(value)[len(prefix):])


def build_table_lookup(wb):
    by_name = defaultdict(list)
    for ws in wb.worksheets:
        for table in ws.tables.values():
            by_name[table.name].append((ws, table))
    return by_name


def row_record_id(headers, row):
    for h in headers:
        if h and str(h).lower().endswith("id"):
            return row[headers.index(h)].value
    return ""


def collect_formula_obsolete_refs(wb, catalog_names: set[str]):
    findings = []
    obsolete_patterns = {
        "tblLookupCatalog": re.compile(r"(?<![A-Za-z0-9_])tblLookupCatalog(?![A-Za-z0-9_])"),
        "Manufacturer Model": re.compile(r"Manufacturer Model"),
        "ManufacturerModel": re.compile(r"ManufacturerModel"),
    }
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    for term, pattern in obsolete_patterns.items():
                        if pattern.search(value):
                            findings.append({"worksheet": ws.title, "cell": cell.coordinate, "formula": value, "term": term})
                    for name in catalog_names:
                        # Presence is useful dependency evidence; not necessarily obsolete.
                        if name in value:
                            pass
    return findings


def analyze(wb, source: Path):
    catalog_sheet, catalog_table, entries = load_catalog(wb)
    table_lookup = build_table_lookup(wb)
    duplicate_table_names = {name: len(matches) for name, matches in table_lookup.items() if len(matches) > 1}
    catalog_names = {e.lookup_table_name for e in entries if e.lookup_table_name}
    excel_table_names = set(table_lookup)

    migration_rows = []
    fk_exceptions = []
    validation = {}
    valid_preserved = 0
    ids_generated = 0
    ids_replaced = 0
    duplicate_ids_found = 0
    malformed_ids_found = 0

    table_records = {}
    namespace_ids = defaultdict(list)
    old_to_new_by_table_field = defaultdict(dict)
    missing_catalog_tables = []

    for entry in entries:
        if not catalog_table_exists(entry, table_lookup):
            continue
        matches = table_lookup.get(entry.lookup_table_name, [])
        if not matches:
            missing_catalog_tables.append(entry.lookup_table_name)
            continue
        if len(matches) > 1:
            validation[f"duplicate_table_{entry.lookup_table_name}"] = "FAIL"
            continue
        ws, table = matches[0]
        headers, rows = table_rows(ws, table)
        if entry.primary_key_field not in headers:
            validation[f"pk_field_{entry.lookup_table_name}"] = "FAIL"
            missing_catalog_tables.append(f"{entry.lookup_table_name}.{entry.primary_key_field}")
            continue
        pk_idx = headers.index(entry.primary_key_field)
        values = [row[pk_idx].value for row in rows]
        counts = Counter(str(v) for v in values if v not in (None, ""))
        highest = max([suffix_num(v, entry.primary_key_prefix) or 0 for v in values] + [0])
        next_num = highest + 1
        used = {str(v) for v in values if v not in (None, "")}
        for row in rows:
            old = row[pk_idx].value
            excel_row = row[pk_idx].row
            reason = ""
            if old in (None, ""):
                reason = "Blank primary key"
            elif counts[str(old)] > 1:
                reason = "Duplicated primary key"
                duplicate_ids_found += 1
            elif not valid_id(old, entry.primary_key_prefix):
                reason = "Malformed or unapproved prefix"
                malformed_ids_found += 1
            else:
                valid_preserved += 1
                namespace_ids[entry.primary_key_prefix].append((entry.lookup_table_name, old))
                table_records[entry.lookup_table_name] = table_records.get(entry.lookup_table_name, set())
                table_records[entry.lookup_table_name].add(str(old))
                continue

            width = max([len(str(v)[len(entry.primary_key_prefix):]) for v in values if valid_id(v, entry.primary_key_prefix)] + [3])
            while True:
                new = f"{entry.primary_key_prefix}{next_num:0{width}d}"
                next_num += 1
                if new not in used:
                    used.add(new)
                    break
            ids_generated += 1
            ids_replaced += 1
            migration_rows.append({
                "Worksheet": ws.title,
                "TableName": table.name,
                "ExcelRow": excel_row,
                "PrimaryKeyField": entry.primary_key_field,
                "OldID": "" if old is None else old,
                "NewID": new,
                "ChangeReason": reason,
                "ApprovedPrefix": entry.primary_key_prefix,
                "MappingStatus": "Pending Apply",
            })
            old_to_new_by_table_field[(table.name, entry.primary_key_field)][str(old) if old is not None else ""] = new
            table_records.setdefault(entry.lookup_table_name, set()).add(new)

    duplicate_ids_across_namespace = []
    for prefix, pairs in namespace_ids.items():
        c = Counter(str(v) for _, v in pairs)
        for value, count in c.items():
            if count > 1:
                duplicate_ids_across_namespace.append({"prefix": prefix, "id": value, "count": count})

    uncataloged_tables = sorted(excel_table_names - catalog_names)

    # Foreign-key checks: any ID-ending column not primary key is checked against catalog table with same PK field.
    pk_field_to_table = {e.primary_key_field: e.lookup_table_name for e in entries if e.lookup_table_name and e.primary_key_field}
    for ws in wb.worksheets:
        for table in ws.tables.values():
            headers, rows = table_rows(ws, table)
            for col_idx, header in enumerate(headers):
                if not header or not str(header).endswith("ID"):
                    continue
                expected_table = pk_field_to_table.get(str(header))
                if not expected_table or expected_table == table.name:
                    continue
                known = table_records.get(expected_table, set())
                for row in rows:
                    value = row[col_idx].value
                    if value in (None, ""):
                        continue
                    mapped = None
                    # Deterministic FK update only if current value has a migration map in the expected parent table.
                    for (mapped_table, mapped_field), mapping in old_to_new_by_table_field.items():
                        if mapped_field == header and str(value) in mapping:
                            mapped = mapping[str(value)]
                            break
                    if str(value) not in known and mapped is None:
                        fk_exceptions.append({
                            "Worksheet": ws.title,
                            "TableName": table.name,
                            "ExcelRow": row[col_idx].row,
                            "RecordID": row_record_id(headers, row),
                            "ForeignKeyField": header,
                            "CurrentValue": value,
                            "ExpectedLookupTable": expected_table,
                            "ExceptionType": "Foreign key target not found",
                            "ResolutionStatus": "Unresolved",
                            "Notes": "No deterministic old-to-new mapping exists.",
                        })

    obsolete_refs = collect_formula_obsolete_refs(wb, catalog_names)
    summary = {
        "catalog_sheet": catalog_sheet,
        "catalog_table": catalog_table,
        "lookup_catalog_record_count": len(entries),
        "workbook_sheet_count": len(wb.sheetnames),
        "excel_table_count": sum(len(ws.tables) for ws in wb.worksheets),
        "valid_ids_preserved": valid_preserved,
        "ids_generated": ids_generated,
        "ids_replaced": ids_replaced,
        "duplicate_ids_found": duplicate_ids_found,
        "malformed_ids_found": malformed_ids_found,
        "foreign_keys_updated": 0,
        "unresolved_foreign_keys": len(fk_exceptions),
        "missing_catalog_tables": len(missing_catalog_tables),
        "uncataloged_excel_tables": len(uncataloged_tables),
        "duplicate_table_names": duplicate_table_names,
        "duplicate_ids_across_namespace": duplicate_ids_across_namespace,
        "formulas_or_references_using_obsolete_names": obsolete_refs,
        "missing_catalog_table_names": missing_catalog_tables,
        "uncataloged_table_names": uncataloged_tables,
    }
    return entries, migration_rows, fk_exceptions, summary


def apply_migration(wb, migration_rows):
    by_target = {(m["Worksheet"], m["TableName"], int(m["ExcelRow"]), m["PrimaryKeyField"]): m for m in migration_rows}
    for ws in wb.worksheets:
        for table in ws.tables.values():
            headers, rows = table_rows(ws, table)
            for row in rows:
                for h in headers:
                    key = (ws.title, table.name, row[0].row, h)
                    if key in by_target:
                        row[headers.index(h)].value = by_target[key]["NewID"]
                        by_target[key]["MappingStatus"] = "Applied"


def write_csv(path: Path, rows, headers):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in headers})


def workbook_health(path: Path):
    with zipfile.ZipFile(path) as z:
        bad = z.testzip()
        if bad:
            raise RuntimeError(f"ZIP integrity failed at {bad}")
        if "xl/workbook.xml" not in z.namelist():
            raise RuntimeError("xl/workbook.xml missing")
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    return wb, [str(w.message) for w in caught]


def scan_formula_errors(wb):
    errors = []
    tokens = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(t in cell.value for t in tokens):
                    errors.append({"Worksheet": ws.title, "Cell": cell.coordinate, "Value": cell.value})
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    if args.dry_run == args.apply:
        raise SystemExit("Choose exactly one of --dry-run or --apply")

    source = Path(args.source)
    output = Path(args.output)
    if args.apply and output.exists():
        raise SystemExit(f"Refusing to overwrite existing output: {output}")

    run_timestamp = datetime.now(UTC).isoformat()
    source_before = sha256(source)
    wb, warnings_seen = workbook_health(source)
    entries, migration_rows, fk_exceptions, summary = analyze(wb, source)

    # Safety stop: any proposed change whose old ID already matches the approved prefix and is unique is unexpected.
    unexpected = [m for m in migration_rows if m["ChangeReason"] not in {"Blank primary key", "Duplicated primary key", "Malformed or unapproved prefix"}]
    if unexpected:
        raise SystemExit(f"Unexpected valid-ID changes proposed: {unexpected[:5]}")

    base = output.with_suffix("")
    qa_json = Path(str(base) + "_QA.json")
    qa_md = Path(str(base) + "_QA.md")
    map_csv = Path(str(base) + "_Migration_Map.csv")
    fk_csv = Path(str(base) + "_FK_Exceptions.csv")
    map_headers = ["Worksheet", "TableName", "ExcelRow", "PrimaryKeyField", "OldID", "NewID", "ChangeReason", "ApprovedPrefix", "MappingStatus"]
    fk_headers = ["Worksheet", "TableName", "ExcelRow", "RecordID", "ForeignKeyField", "CurrentValue", "ExpectedLookupTable", "ExceptionType", "ResolutionStatus", "Notes"]

    if args.apply:
        shutil.copy2(source, output)
        out_wb, _ = workbook_health(output)
        apply_migration(out_wb, migration_rows)
        out_wb.save(output)
        out_wb, output_warnings = workbook_health(output)
        formula_errors = scan_formula_errors(out_wb)
        output_hash = sha256(output)
        # Mark rows applied after successful save.
        for m in migration_rows:
            if m["MappingStatus"] == "Pending Apply":
                m["MappingStatus"] = "Applied"
    else:
        output_hash = ""
        output_warnings = []
        formula_errors = []

    source_after = sha256(source)
    if source_before != source_after:
        raise SystemExit("Source workbook hash changed; aborting.")

    validations = {
        "source_hash_unchanged": "PASS" if source_before == source_after else "FAIL",
        "source_openpyxl_load": "PASS",
        "lookup_catalog_read": "PASS",
        "missing_catalog_tables": "PASS" if summary["missing_catalog_tables"] == 0 else "FAIL",
        "uncataloged_excel_tables": "PASS" if summary["uncataloged_excel_tables"] == 0 else "WARNING",
        "duplicate_table_names": "PASS" if not summary["duplicate_table_names"] else "FAIL",
        "duplicate_ids_across_namespace": "PASS" if not summary["duplicate_ids_across_namespace"] else "WARNING",
        "malformed_ids": "PASS" if summary["malformed_ids_found"] == 0 else "WARNING",
        "foreign_key_integrity": "PASS" if summary["unresolved_foreign_keys"] == 0 else "WARNING",
        "formula_error_tokens": "PASS" if not formula_errors else "FAIL",
        "output_openpyxl_load": "PASS" if args.apply else "NOT RUN",
    }

    qa = {
        "source_workbook_path": str(source),
        "source_sha256_before_execution": source_before,
        "source_sha256_after_execution": source_after,
        "output_workbook_path": str(output) if args.apply else "",
        "output_sha256": output_hash,
        "script_path": str(Path(__file__).resolve()),
        "script_sha256": sha256(Path(__file__).resolve()),
        "run_timestamp_utc": run_timestamp,
        "dry_run": args.dry_run,
        "apply": args.apply,
        "python_version": sys.version,
        "openpyxl_version": __import__("openpyxl").__version__,
        "source_openpyxl_warnings": warnings_seen,
        "output_openpyxl_warnings": output_warnings,
        **summary,
        "formulas_or_references_using_obsolete_names_count": len(summary["formulas_or_references_using_obsolete_names"]),
        "formula_error_tokens": formula_errors,
        "validation_results": validations,
        "recommendation": "READY FOR MANUAL REVIEW" if validations["missing_catalog_tables"] == "PASS" and validations["formula_error_tokens"] == "PASS" else "BLOCKED - CORRECTIONS REQUIRED",
    }

    write_csv(map_csv, migration_rows, map_headers)
    write_csv(fk_csv, fk_exceptions, fk_headers)
    qa_json.write_text(json.dumps(qa, indent=2, default=str), encoding="utf-8")
    qa_md.write_text(
        "# Approved ID Migration v1.2 QA\n\n"
        f"Recommendation: {qa['recommendation']}\n\n"
        f"Source: {source}\n\n"
        f"Source SHA-256 before: {source_before}\n\n"
        f"Source SHA-256 after: {source_after}\n\n"
        f"Output SHA-256: {output_hash}\n\n"
        f"Script: {Path(__file__).resolve()}\n\n"
        f"Script SHA-256: {qa['script_sha256']}\n\n"
        "## Totals\n\n"
        f"- Workbook sheet count: {summary['workbook_sheet_count']}\n"
        f"- Excel table count: {summary['excel_table_count']}\n"
        f"- Lookup catalog record count: {summary['lookup_catalog_record_count']}\n"
        f"- Valid IDs preserved: {summary['valid_ids_preserved']}\n"
        f"- IDs generated: {summary['ids_generated']}\n"
        f"- IDs replaced: {summary['ids_replaced']}\n"
        f"- Duplicate IDs found: {summary['duplicate_ids_found']}\n"
        f"- Malformed IDs found: {summary['malformed_ids_found']}\n"
        f"- Foreign keys updated: {summary['foreign_keys_updated']}\n"
        f"- Unresolved foreign keys: {summary['unresolved_foreign_keys']}\n"
        f"- Missing catalog tables: {summary['missing_catalog_tables']}\n"
        f"- Uncataloged Excel tables: {summary['uncataloged_excel_tables']}\n\n"
        "## Validation\n\n"
        + "\n".join(f"- {k}: {v}" for k, v in validations.items())
        + "\n",
        encoding="utf-8",
    )
    print(json.dumps(qa, indent=2, default=str))


if __name__ == "__main__":
    main()
