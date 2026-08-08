from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(
    r"D:\Business Portal\01_Nocturnix_Business_Portal\100_Master_Data\Devices"
    r"\Nocturnix_Master_Devices_Catalog_v1.5_Approved_ID_Migration_Draft_v1"
)
SOURCE = BASE.with_name(BASE.name + ".2.xlsx")
PROPOSED_MAPPINGS = BASE.with_name(BASE.name + ".2_Proposed_FK_Mappings.csv")
ROOT_CAUSE = BASE.with_name(BASE.name + ".2_FK_Root_Cause_Summary.csv")
SOURCE_FK_EXCEPTIONS = BASE.with_name(BASE.name + ".2_FK_Exceptions.csv")

OUTPUT = BASE.with_name(BASE.name + ".3.xlsx")
QA_MD = BASE.with_name(BASE.name + ".3_QA.md")
QA_JSON = BASE.with_name(BASE.name + ".3_QA.json")
CHANGE_LOG = BASE.with_name(BASE.name + ".3_FK_Change_Log.csv")
FK_EXCEPTIONS = BASE.with_name(BASE.name + ".3_FK_Exceptions.csv")
DEFERRED = BASE.with_name(BASE.name + ".3_Deferred_Missing_Lookups.csv")
APPROVED = BASE.with_name(BASE.name + ".3_Approved_FK_Mappings.csv")

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
EXPECTED_CHANGES = 1372
EXPECTED_DEFERRED = 575


class Blocked(RuntimeError):
    pass


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, object]], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def table_cells(ws, table):
    cells = list(ws[table.ref])
    headers = [str(c.value or "") for c in cells[0]]
    return headers, cells[1:]


def table_lookup(wb):
    found = {}
    for ws in wb.worksheets:
        for table in ws.tables.values():
            found[(ws.title, table.name)] = (ws, table)
            found[(None, table.name)] = (ws, table)
    return found


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def id_parts(value: object) -> tuple[str | None, str | None]:
    text = str(value or "").strip().upper().replace("-", "")
    match = re.match(r"^([A-Z]+)(\d+)$", text)
    if not match:
        return None, None
    return match.group(1), match.group(2)


def valid_id(value: object, prefix: str) -> bool:
    p, digits = id_parts(value)
    return p == prefix and bool(digits)


def find_catalog(wb):
    required = {"LookupTableName", "PrimaryKeyField", "PrimaryKeyPrefix", "TableExists"}
    for ws in wb.worksheets:
        for table in ws.tables.values():
            headers, rows = table_cells(ws, table)
            if required.issubset(headers):
                catalog = {}
                for row in rows:
                    values = {h: row[i].value for i, h in enumerate(headers)}
                    name = normalize_text(values.get("LookupTableName"))
                    if name:
                        catalog[name] = {
                            "PrimaryKeyField": normalize_text(values.get("PrimaryKeyField")),
                            "PrimaryKeyPrefix": normalize_text(values.get("PrimaryKeyPrefix")),
                            "DisplayName": normalize_text(values.get("DisplayName")),
                        }
                return ws.title, table.name, catalog
    raise Blocked("Lookup catalog table was not found.")


def likely_name_field(headers: list[str], pk: str) -> str | None:
    for candidate in ["Name", "DisplayName", "Manufacturer", "Device Family", "Device Type", "ManufacturerName", "DeviceFamilyName"]:
        if candidate in headers and candidate != pk:
            return candidate
    for header in headers:
        lower = header.lower()
        if header != pk and ("name" in lower or "family" in lower or "type" in lower):
            return header
    return None


def lookup_targets(wb, catalog):
    lookups = {}
    tables = table_lookup(wb)
    for table_name, meta in catalog.items():
        match = tables.get((None, table_name))
        if not match:
            continue
        ws, table = match
        headers, rows = table_cells(ws, table)
        pk = meta["PrimaryKeyField"]
        if pk not in headers:
            continue
        name_field = likely_name_field(headers, pk)
        active_field = "IsActive" if "IsActive" in headers else None
        pk_idx = headers.index(pk)
        by_id = defaultdict(list)
        for row in rows:
            value = normalize_text(row[pk_idx].value)
            if value:
                values = {h: row[i].value for i, h in enumerate(headers)}
                by_id[value].append({
                    "worksheet": ws.title,
                    "table": table.name,
                    "row": row[0].row,
                    "values": values,
                    "name": normalize_text(values.get(name_field)) if name_field else "",
                    "active": normalize_text(values.get(active_field)) if active_field else "",
                })
        lookups[table_name] = {"pk": pk, "prefix": meta["PrimaryKeyPrefix"], "by_id": by_id}
    return lookups


def load_approved_mappings(proposed_rows, root_rows, lookups):
    root_by_key = {
        (
            r["SourceWorksheet"],
            r["SourceTable"],
            r["ForeignKeyField"],
            r["ExpectedLookupTable"],
            r["DistinctInvalidValues"],
        ): r
        for r in root_rows
    }
    approved = []
    for row in proposed_rows:
        key = (
            row["SourceWorksheet"],
            row["SourceTable"],
            row["ForeignKeyField"],
            row["TargetLookupTable"],
            row["OldValue"],
        )
        root = root_by_key.get(key)
        basis = row["MappingBasis"].lower()
        confidence = row["Confidence"].lower()
        if not root:
            continue
        if root["RootCause"] != "legacy shortened IDs":
            continue
        if root["DeterministicMappingAvailable"] != "Yes":
            continue
        if "prefix" not in basis and "zero-padding" not in basis:
            continue
        if "high" not in confidence and "deterministic" not in confidence:
            continue
        if not row["ProposedNewValue"] or row["OldValue"] == row["ProposedNewValue"]:
            continue
        target = lookups.get(row["TargetLookupTable"], {}).get("by_id", {}).get(row["ProposedNewValue"], [])
        if len(target) != 1:
            raise Blocked(f"Mapping target missing or duplicated: {row}")
        row = dict(row)
        row["ApprovalStatus"] = "Approved"
        row["RootCause"] = root["RootCause"]
        approved.append(row)
    return approved


def source_rows_for_mapping(wb, mapping):
    tables = table_lookup(wb)
    match = tables.get((mapping["SourceWorksheet"], mapping["SourceTable"]))
    if not match:
        raise Blocked(f"Source table missing: {mapping['SourceWorksheet']} {mapping['SourceTable']}")
    ws, table = match
    headers, rows = table_cells(ws, table)
    fk = mapping["ForeignKeyField"]
    if fk not in headers:
        raise Blocked(f"Foreign-key column missing: {mapping}")
    fk_idx = headers.index(fk)
    record_idx = next((i for i, h in enumerate(headers) if h.lower().endswith("id")), None)
    hits = []
    for row in rows:
        if normalize_text(row[fk_idx].value) == mapping["OldValue"]:
            record_id = normalize_text(row[record_idx].value) if record_idx is not None else ""
            hits.append((ws, row[fk_idx], record_id))
    return hits


def validate_mappings(wb, approved, lookups, catalog):
    expected_count = 0
    seen = set()
    for mapping in approved:
        if not mapping["OldValue"]:
            raise Blocked(f"Blank mapping value is not allowed: {mapping}")
        target_meta = lookups[mapping["TargetLookupTable"]]
        target_rows = target_meta["by_id"].get(mapping["ProposedNewValue"], [])
        if len(target_rows) != 1:
            raise Blocked(f"Mapping target missing or duplicated: {mapping}")
        target = target_rows[0]
        if target["active"].lower() in {"no", "false", "inactive", "0"}:
            raise Blocked(f"Mapping points to inactive target: {mapping}")
        if not valid_id(mapping["ProposedNewValue"], target_meta["prefix"]):
            raise Blocked(f"Mapping target does not match catalog prefix: {mapping}")
        if catalog[mapping["TargetLookupTable"]]["PrimaryKeyField"] != target_meta["pk"]:
            raise Blocked(f"Catalog PK mismatch for target: {mapping}")
        if valid_id(mapping["OldValue"], target_meta["prefix"]) and mapping["OldValue"] in target_meta["by_id"]:
            raise Blocked(f"Mapping would change an already-valid FK value: {mapping}")
        hits = source_rows_for_mapping(wb, mapping)
        if not hits:
            raise Blocked(f"OldValue does not exist in eligible source rows: {mapping}")
        for _, cell, _ in hits:
            key = (mapping["SourceWorksheet"], mapping["SourceTable"], cell.row, mapping["ForeignKeyField"])
            if key in seen:
                raise Blocked(f"Cell matched more than one mapping: {key}")
            seen.add(key)
        expected_count += len(hits)
    if expected_count != EXPECTED_CHANGES:
        raise Blocked(f"Expected {EXPECTED_CHANGES} FK changes, validated {expected_count}.")
    return expected_count


def pk_snapshot(wb, catalog):
    tables = table_lookup(wb)
    snapshot = {}
    for table_name, meta in catalog.items():
        match = tables.get((None, table_name))
        if not match:
            continue
        ws, table = match
        headers, rows = table_cells(ws, table)
        pk = meta["PrimaryKeyField"]
        if pk not in headers:
            continue
        idx = headers.index(pk)
        snapshot[table_name] = [normalize_text(row[idx].value) for row in rows]
    return snapshot


def collect_fk_exceptions(wb, catalog):
    lookups = lookup_targets(wb, catalog)
    tables = table_lookup(wb)
    exceptions = []
    for (ws_name, table_name), (ws, table) in list(tables.items()):
        if ws_name is None:
            continue
        headers, rows = table_cells(ws, table)
        record_idx = next((i for i, h in enumerate(headers) if h.lower().endswith("id")), None)
        for idx, header in enumerate(headers):
            if not header.endswith("ID"):
                continue
            expected_table = next((name for name, meta in catalog.items() if meta["PrimaryKeyField"] == header), None)
            if not expected_table:
                continue
            target_ids = lookups.get(expected_table, {}).get("by_id", {})
            for row in rows:
                value = normalize_text(row[idx].value)
                if value and value not in target_ids:
                    exceptions.append({
                        "Worksheet": ws.title,
                        "TableName": table.name,
                        "ExcelRow": row[idx].row,
                        "RecordID": normalize_text(row[record_idx].value) if record_idx is not None else "",
                        "ForeignKeyField": header,
                        "CurrentValue": value,
                        "ExpectedLookupTable": expected_table,
                        "ExceptionType": "Foreign key target not found",
                        "ResolutionStatus": "Unresolved",
                        "Notes": "No deterministic old-to-new mapping exists in this run.",
                    })
    return exceptions


def deferred_missing(source_exceptions, root_rows):
    missing_keys = {
        (
            r["SourceWorksheet"],
            r["SourceTable"],
            r["ForeignKeyField"],
            r["DistinctInvalidValues"],
        ): r
        for r in root_rows
        if r["RootCause"] == "missing lookup record"
    }
    rows = []
    for exception in source_exceptions:
        key = (
            exception["Worksheet"],
            exception["TableName"],
            exception["ForeignKeyField"],
            exception["CurrentValue"],
        )
        root = missing_keys.get(key)
        if not root:
            continue
        prefix, digits = id_parts(exception["CurrentValue"])
        width = 6
        normalized = f"{root['ExpectedPrefix']}{int(digits):0{width}d}" if prefix and digits else ""
        rows.append({
            "Worksheet": exception["Worksheet"],
            "TableName": exception["TableName"],
            "ExcelRow": exception["ExcelRow"],
            "RecordID": exception["RecordID"],
            "ForeignKeyField": exception["ForeignKeyField"],
            "CurrentValue": exception["CurrentValue"],
            "NormalizedCandidateID": normalized,
            "ExpectedLookupTable": exception["ExpectedLookupTable"],
            "RequiredLookupRecordName": "",
            "AffectedRowCount": root["AffectedRowCount"],
            "Status": "Pending Lookup Approval",
            "RecommendedAction": "Create or approve the missing lookup record before FK remediation.",
        })
    return rows


def scan_formula_errors(wb):
    findings = []
    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if any(token in value for token in ERROR_TOKENS):
                        findings.append({"worksheet": ws.title, "cell": cell.coordinate, "formula": value})
    return formula_count, findings


def duplicate_and_malformed_pks(wb, catalog):
    dupes = []
    malformed = []
    tables = table_lookup(wb)
    for table_name, meta in catalog.items():
        match = tables.get((None, table_name))
        if not match:
            continue
        ws, table = match
        headers, rows = table_cells(ws, table)
        pk = meta["PrimaryKeyField"]
        if pk not in headers:
            continue
        idx = headers.index(pk)
        values = [normalize_text(row[idx].value) for row in rows if normalize_text(row[idx].value)]
        for value, count in Counter(values).items():
            if count > 1:
                dupes.append({"table": table_name, "id": value, "count": count})
        for row in rows:
            value = normalize_text(row[idx].value)
            if value and not valid_id(value, meta["PrimaryKeyPrefix"]):
                malformed.append({"table": table_name, "row": row[idx].row, "id": value})
    return dupes, malformed


def main() -> int:
    if OUTPUT.exists():
        raise Blocked(f"Output already exists and will not be overwritten: {OUTPUT}")
    source_hash_before = sha256(SOURCE)
    script_hash = sha256(Path(__file__))
    proposed_rows = read_csv(PROPOSED_MAPPINGS)
    root_rows = read_csv(ROOT_CAUSE)
    source_fk_rows = read_csv(SOURCE_FK_EXCEPTIONS)

    source_wb = load_workbook(SOURCE, read_only=False, data_only=False, keep_links=True)
    _, _, catalog = find_catalog(source_wb)
    source_lookups = lookup_targets(source_wb, catalog)
    approved = load_approved_mappings(proposed_rows, root_rows, source_lookups)
    if len(approved) != 48:
        raise Blocked(f"Expected 48 approved mapping rows, found {len(approved)}.")
    expected_count = validate_mappings(source_wb, approved, source_lookups, catalog)
    source_pk_snapshot = pk_snapshot(source_wb, catalog)
    source_formula_count, source_formula_errors = scan_formula_errors(source_wb)

    write_csv(APPROVED, approved, [
        "SourceWorksheet", "SourceTable", "ForeignKeyField", "OldValue", "ProposedNewValue",
        "TargetLookupTable", "TargetRecordName", "MappingBasis", "EvidenceLocation", "AffectedRowCount",
        "Confidence", "ApprovalStatus",
    ])

    shutil.copy2(SOURCE, OUTPUT)
    out_wb = load_workbook(OUTPUT, read_only=False, data_only=False, keep_links=True)
    changed_at = datetime.now(timezone.utc).isoformat()
    change_rows = []
    actual_count = 0
    for mapping in approved:
        hits = source_rows_for_mapping(out_wb, mapping)
        target = source_lookups[mapping["TargetLookupTable"]]["by_id"][mapping["ProposedNewValue"]][0]
        for _, cell, record_id in hits:
            old = normalize_text(cell.value)
            if old != mapping["OldValue"]:
                raise Blocked(f"Unexpected source value while applying mapping at {cell.coordinate}.")
            cell.value = mapping["ProposedNewValue"]
            actual_count += 1
            change_rows.append({
                "Worksheet": mapping["SourceWorksheet"],
                "TableName": mapping["SourceTable"],
                "ExcelRow": cell.row,
                "RecordID": record_id,
                "ForeignKeyField": mapping["ForeignKeyField"],
                "OldValue": old,
                "NewValue": mapping["ProposedNewValue"],
                "TargetLookupTable": mapping["TargetLookupTable"],
                "TargetRecordName": target["name"],
                "MappingBasis": mapping["MappingBasis"],
                "MappingApproval": mapping["ApprovalStatus"],
                "ChangedAtUTC": changed_at,
            })
    if actual_count != expected_count or actual_count != EXPECTED_CHANGES:
        raise Blocked(f"Actual changes {actual_count} did not match approved expected {expected_count}.")
    out_wb.save(OUTPUT)

    reopened = load_workbook(OUTPUT, read_only=False, data_only=False, keep_links=True)
    _, _, out_catalog = find_catalog(reopened)
    if pk_snapshot(reopened, out_catalog) != source_pk_snapshot:
        raise Blocked("Primary-key snapshot changed.")
    output_formula_count, formula_errors = scan_formula_errors(reopened)
    if output_formula_count != source_formula_count:
        raise Blocked(f"Formula count changed from {source_formula_count} to {output_formula_count}.")
    remaining_fk = collect_fk_exceptions(reopened, out_catalog)
    deferred = deferred_missing(source_fk_rows, root_rows)
    if len(deferred) != EXPECTED_DEFERRED:
        raise Blocked(f"Deferred missing lookup row count {len(deferred)} did not match {EXPECTED_DEFERRED}.")
    deferred_keys = {(r["Worksheet"], r["TableName"], int(r["ExcelRow"]), r["ForeignKeyField"], r["CurrentValue"]) for r in deferred}
    remaining_keys = {(r["Worksheet"], r["TableName"], int(r["ExcelRow"]), r["ForeignKeyField"], r["CurrentValue"]) for r in remaining_fk}
    if not deferred_keys.issubset(remaining_keys):
        raise Blocked("One or more deferred missing-lookup rows was modified or no longer appears unresolved.")

    dupes, malformed = duplicate_and_malformed_pks(reopened, out_catalog)
    missing_catalog = []
    table_names = {table.name for ws in reopened.worksheets for table in ws.tables.values()}
    for table_name in out_catalog:
        if table_name not in table_names:
            missing_catalog.append(table_name)
    uncataloged = sorted(table_names - set(out_catalog))
    with zipfile.ZipFile(OUTPUT) as zf:
        zip_bad = zf.testzip()
        has_workbook_xml = "xl/workbook.xml" in zf.namelist()

    source_hash_after = sha256(SOURCE)
    output_hash = sha256(OUTPUT)
    recommendation = "READY FOR LOOKUP-RECORD REVIEW"
    if source_hash_after != source_hash_before or zip_bad is not None or formula_errors:
        recommendation = "BLOCKED - REMEDIATION VALIDATION FAILED"

    write_csv(CHANGE_LOG, change_rows, [
        "Worksheet", "TableName", "ExcelRow", "RecordID", "ForeignKeyField", "OldValue", "NewValue",
        "TargetLookupTable", "TargetRecordName", "MappingBasis", "MappingApproval", "ChangedAtUTC",
    ])
    write_csv(FK_EXCEPTIONS, remaining_fk, [
        "Worksheet", "TableName", "ExcelRow", "RecordID", "ForeignKeyField", "CurrentValue",
        "ExpectedLookupTable", "ExceptionType", "ResolutionStatus", "Notes",
    ])
    write_csv(DEFERRED, deferred, [
        "Worksheet", "TableName", "ExcelRow", "RecordID", "ForeignKeyField", "CurrentValue",
        "NormalizedCandidateID", "ExpectedLookupTable", "RequiredLookupRecordName", "AffectedRowCount",
        "Status", "RecommendedAction",
    ])
    qa = {
        "source_workbook_path": str(SOURCE),
        "output_workbook_path": str(OUTPUT),
        "source_sha256_before": source_hash_before,
        "source_sha256_after": source_hash_after,
        "source_hash_unchanged": source_hash_before == source_hash_after,
        "output_sha256": output_hash,
        "remediation_script_path": str(Path(__file__).resolve()),
        "remediation_script_sha256": script_hash,
        "run_timestamp_utc": changed_at,
        "approved_mapping_count": len(approved),
        "expected_fk_cell_changes": expected_count,
        "actual_fk_cell_changes": actual_count,
        "unchanged_missing_lookup_rows": len(deferred),
        "remaining_unresolved_fk_count": len(remaining_fk),
        "duplicate_primary_keys": len(dupes),
        "malformed_primary_keys": len(malformed),
        "invalid_foreign_keys": len(remaining_fk),
        "missing_catalog_tables": len(missing_catalog),
        "uncataloged_tables": len(uncataloged),
        "formula_error_tokens": len(formula_errors),
        "xlsx_zip_integrity": "PASS" if zip_bad is None and has_workbook_xml else "FAIL",
        "workbook_reopen_status": "PASS",
        "worksheet_count": len(reopened.worksheets),
        "excel_table_count": sum(len(ws.tables) for ws in reopened.worksheets),
        "primary_key_changes": 0,
        "lookup_records_created": 0,
        "guessed_mappings": 0,
        "recommendation": recommendation,
        "remaining_fk_by_field": Counter(f"{r['Worksheet']}|{r['TableName']}|{r['ForeignKeyField']}" for r in remaining_fk),
        "missing_catalog_table_names": missing_catalog,
        "uncataloged_table_names": uncataloged,
    }
    QA_JSON.write_text(json.dumps(qa, indent=2, default=dict) + "\n", encoding="utf-8")
    QA_MD.write_text(
        "\n".join([
            "# Controlled FK Remediation v1.3 QA",
            "",
            f"Recommendation: {recommendation}",
            "",
            f"Source SHA-256 before: {source_hash_before}",
            f"Source SHA-256 after: {source_hash_after}",
            f"Output SHA-256: {output_hash}",
            f"Remediation script SHA-256: {script_hash}",
            "",
            "## Totals",
            "",
            f"- Approved mapping count: {len(approved)}",
            f"- Expected FK cell changes: {expected_count}",
            f"- Actual FK cell changes: {actual_count}",
            f"- Unchanged missing-lookup rows: {len(deferred)}",
            f"- Remaining unresolved FK count: {len(remaining_fk)}",
            f"- Duplicate primary keys: {len(dupes)}",
            f"- Malformed primary keys: {len(malformed)}",
            f"- Invalid foreign keys: {len(remaining_fk)}",
            f"- Missing catalog tables: {len(missing_catalog)}",
            f"- Uncataloged tables: {len(uncataloged)}",
            f"- Formula error tokens: {len(formula_errors)}",
            f"- XLSX ZIP integrity: {qa['xlsx_zip_integrity']}",
            f"- Workbook reopen status: {qa['workbook_reopen_status']}",
            f"- Worksheet count: {qa['worksheet_count']}",
            f"- Excel table count: {qa['excel_table_count']}",
            "",
            "No primary keys were changed, no lookup records were created, and no mappings were inferred during execution.",
        ])
        + "\n",
        encoding="utf-8",
    )
    print(json.dumps(qa, indent=2, default=dict))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Blocked as exc:
        print(f"BLOCKED: {exc}", file=sys.stderr)
        raise SystemExit(2)
