"""Generate the standalone Nocturnix Master Devices Catalog v1 review workbook.

The generator reads protected inputs without modifying them, writes a
transactional temporary workbook, validates the reopened artifact, and only
then atomically replaces the planned output.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
PROPOSAL_PATH = WORKING_DIR / "Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
MASTER_SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
MASTER_PARTS_PATH = WORKING_DIR / "Nocturnix_Master_Parts_Catalog_v1.xlsx"
MASTER_PRICING_PATH = WORKING_DIR / "Nocturnix_Master_Pricing_Catalog_v1.xlsx"
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
OUTPUT_PATH = WORKING_DIR / "Nocturnix_Master_Devices_Catalog_v1.xlsx"
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(f"{OUTPUT_PATH.stem}.tmp{OUTPUT_PATH.suffix}")

PROTECTED_PATHS = (
    PROPOSAL_PATH,
    MASTER_SERVICES_PATH,
    MASTER_PARTS_PATH,
    MASTER_PRICING_PATH,
    CANONICAL_PATH,
)
RETAINED_SHEET = "01 - Retained"
EXCLUSIONS_SHEET = "02 - Duplicate Exclusions"
CANONICAL_DEVICE_SHEET = "32 Devices"
EXPECTED_DEVICE_ROWS = 46
IMPORT_BATCH_ID = "MASTER-DEVICES-V1-REVIEW"
NAMESPACE_AUTHORITY = "ADR-009"
DEVICE_ID_PATTERN = re.compile(r"^DEV\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Master Devices",
    "02 - Manufacturers",
    "03 - Device Families",
    "04 - Product Lines",
    "05 - Device Series",
    "06 - Form Factors",
    "07 - Operating Systems",
    "08 - Network Types",
    "09 - Device Statuses",
    "10 - Identity Review",
    "11 - Relationship Review",
    "12 - Validation Summary",
    "13 - Revision History",
    "14 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterDevicesInstructions",
    "01 - Master Devices": "tblMasterDevicesCatalog",
    "02 - Manufacturers": "tblDeviceManufacturers",
    "03 - Device Families": "tblDeviceFamilies",
    "04 - Product Lines": "tblProductLines",
    "05 - Device Series": "tblDeviceSeries",
    "06 - Form Factors": "tblFormFactors",
    "07 - Operating Systems": "tblOperatingSystems",
    "08 - Network Types": "tblNetworkTypes",
    "09 - Device Statuses": "tblDeviceStatuses",
    "10 - Identity Review": "tblDeviceIdentityReview",
    "11 - Relationship Review": "tblDeviceRelationshipReview",
    "12 - Validation Summary": "tblMasterDevicesValidation",
    "13 - Revision History": "tblMasterDevicesRevisionHistory",
    "14 - Import Metadata": "tblMasterDevicesImportMetadata",
}
DEVICE_HEADERS = [
    "Device ID",
    "Legacy Device SKU",
    "Active",
    "Status",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Product Line",
    "Device Series",
    "Device Name",
    "Device Display Name",
    "Device Description",
    "Model Number",
    "Variant",
    "Generation",
    "Release Year",
    "Form Factor",
    "Operating System Family",
    "Network Type",
    "Storage Capacity",
    "Memory Capacity",
    "Color",
    "Region",
    "Carrier",
    "Wi-Fi Only",
    "Cellular Capable",
    "Repair Supported",
    "Parts Supported",
    "Mail-In Eligible",
    "Mobile Service Eligible",
    "Compatibility Status",
    "Service Mapping Status",
    "Parts Mapping Status",
    "Review Status",
    "Legacy Retail Price",
    "Legacy Cost",
    "Currency",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch ID",
    "Reviewer",
    "Reviewer Notes",
    "Effective Date",
    "Last Reviewed",
    "Created At",
    "Updated At",
]
IDENTITY_HEADERS = [
    "Device ID",
    "Legacy Device SKU",
    "Manufacturer Name",
    "Device Family Name",
    "Product Line",
    "Device Series",
    "Device Name",
    "Model Number",
    "Variant",
    "Missing Identity Inputs",
    "Identity Concern",
    "Required Action",
    "Review Status",
    "Reviewer Notes",
]
RELATIONSHIP_HEADERS = [
    "Device ID",
    "Device Name",
    "Manufacturer ID",
    "Device Family Code",
    "Compatibility Status",
    "Service Mapping Status",
    "Parts Mapping Status",
    "Missing Relationships",
    "Required Action",
    "Review Status",
    "Reviewer Notes",
]
DEVICE_STATUSES = ["Draft", "Active", "Planned", "Future", "Discontinued", "Archived", "Rejected"]
REVIEW_STATUSES = [
    "Pending Review",
    "Pending Manufacturer Review",
    "Pending Family Review",
    "Pending Identity Review",
    "Pending Relationship Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
COMPATIBILITY_STATUSES = [
    "Pending Compatibility Review",
    "Family-Level Only",
    "Model-Level Proposed",
    "Variant-Level Proposed",
    "Not Applicable",
    "Approved",
    "Rejected",
]
SERVICE_MAPPING_STATUSES = [
    "Pending Service Mapping",
    "Family Services Available",
    "Model Services Proposed",
    "Complete",
    "Not Applicable",
    "Approved",
    "Rejected",
]
PARTS_MAPPING_STATUSES = [
    "Pending Parts Mapping",
    "Family Parts Available",
    "Model Parts Proposed",
    "Complete",
    "Not Applicable",
    "Approved",
    "Rejected",
]
YES_NO = ["Yes", "No"]
CURRENCIES = ["USD"]
DEFINED_NAME_BY_HEADER = {
    "Active": "DV_YesNo",
    "Status": "DV_DeviceStatuses",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Product Line": "DV_ProductLines",
    "Device Series": "DV_DeviceSeries",
    "Form Factor": "DV_FormFactors",
    "Operating System Family": "DV_OperatingSystemFamilies",
    "Network Type": "DV_NetworkTypes",
    "Wi-Fi Only": "DV_YesNo",
    "Cellular Capable": "DV_YesNo",
    "Repair Supported": "DV_YesNo",
    "Parts Supported": "DV_YesNo",
    "Mail-In Eligible": "DV_YesNo",
    "Mobile Service Eligible": "DV_YesNo",
    "Compatibility Status": "DV_CompatibilityStatuses",
    "Service Mapping Status": "DV_ServiceMappingStatuses",
    "Parts Mapping Status": "DV_PartsMappingStatuses",
    "Review Status": "DV_ReviewStatuses",
    "Currency": "DV_Currencies",
}
DEFINED_NAME_SHEET = {
    "DV_YesNo": "09 - Device Statuses",
    "DV_DeviceStatuses": "09 - Device Statuses",
    "DV_ReviewStatuses": "09 - Device Statuses",
    "DV_CompatibilityStatuses": "09 - Device Statuses",
    "DV_ServiceMappingStatuses": "09 - Device Statuses",
    "DV_PartsMappingStatuses": "09 - Device Statuses",
    "DV_Currencies": "09 - Device Statuses",
    "DV_ManufacturerIDs": "02 - Manufacturers",
    "DV_DeviceFamilyCodes": "03 - Device Families",
    "DV_ProductLines": "04 - Product Lines",
    "DV_DeviceSeries": "05 - Device Series",
    "DV_FormFactors": "06 - Form Factors",
    "DV_OperatingSystemFamilies": "07 - Operating Systems",
    "DV_NetworkTypes": "08 - Network Types",
}
GENERATED_BLANK_FIELDS = {
    "Device Series",
    "Model Number",
    "Variant",
    "Generation",
    "Release Year",
    "Form Factor",
    "Operating System Family",
    "Network Type",
    "Storage Capacity",
    "Memory Capacity",
    "Color",
    "Region",
    "Carrier",
    "Currency",
    "Reviewer",
    "Effective Date",
    "Last Reviewed",
}
PROHIBITED_HEADERS = {
    "Serial Number",
    "IMEI",
    "Stock",
    "Stock Quantity",
    "Bin",
    "Location",
    "Inventory Quantity",
    "Final Cost",
    "Markup",
    "Margin",
    "Final Customer Price",
}


class DevicesCatalogError(RuntimeError):
    """Raised when the Master Devices review artifact violates its contract."""


def text(value: Any) -> str:
    """Return stripped text, treating None as blank."""
    return "" if value is None else str(value).strip()


def ascii_value(value: Any) -> Any:
    """Make generated text ASCII-safe without changing typed non-text values."""
    if not isinstance(value, str):
        return value
    replacements = {"\u2013": "-", "\u2014": "-", "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"'}
    for source, replacement in replacements.items():
        value = value.replace(source, replacement)
    return value.encode("ascii", "replace").decode("ascii")


def excel_safe_value(value: Any) -> Any:
    """Return a value openpyxl can persist without timezone errors."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def assert_excel_safe_value_contract() -> None:
    """Exercise temporal normalization without workbook I/O."""
    aware_utc = datetime(2026, 7, 23, 12, 30, tzinfo=UTC)
    eastern = timezone(timedelta(hours=-4))
    assert excel_safe_value(aware_utc) == datetime(2026, 7, 23, 12, 30)
    assert excel_safe_value(datetime(2026, 7, 23, 8, 30, tzinfo=eastern)) == datetime(2026, 7, 23, 12, 30)
    naive = datetime(2026, 7, 23, 12, 30)
    calendar_date = date(2026, 7, 23)
    assert excel_safe_value(naive) is naive
    assert excel_safe_value(calendar_date) is calendar_date
    assert excel_safe_value("") == ""
    assert excel_safe_value(None) is None
    assert excel_safe_value("2026-07-23T12:30:00Z") == "2026-07-23T12:30:00Z"


def decimal_value(value: Any, field: str) -> Decimal | None:
    """Parse a monetary source value while preserving blank versus zero."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        result = Decimal(str(value).replace("$", "").replace(",", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise DevicesCatalogError(f"{field} contains a nonnumeric value: {value!r}") from exc
    if not result.is_finite() or result < 0:
        raise DevicesCatalogError(f"{field} must be finite and nonnegative: {value!r}")
    return result


def file_hash(path: Path) -> str:
    """Return a file's SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Validate protected Office inputs without modifying them."""
    for path in paths:
        if not path.exists():
            raise DevicesCatalogError(f"Required source workbook does not exist: {path}")
        if path.stat().st_size <= 0:
            raise DevicesCatalogError(f"Required source workbook is empty: {path}")
        if not zipfile.is_zipfile(path):
            raise DevicesCatalogError(f"Required source workbook is not a valid ZIP-based Office file: {path}")


def read_records(path: Path, sheet: str, header_row: int, *, keep_vba: bool = False) -> list[dict[str, Any]]:
    """Read a worksheet into dictionaries using an exact header row."""
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    try:
        if sheet not in workbook.sheetnames:
            raise DevicesCatalogError(f"Required worksheet {sheet!r} is missing from {path}")
        worksheet = workbook[sheet]
        headers = [text(cell.value) for cell in worksheet[header_row]]
        if not any(headers):
            raise DevicesCatalogError(f"Header row {header_row} is blank in {path}!{sheet}")
        records = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            record = {header: value for header, value in zip_longest(headers, values, fillvalue=None) if header}
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def require_headers(records: Sequence[dict[str, Any]], headers: Sequence[str], context: str) -> None:
    """Require fields even when the source table has no data rows."""
    if not records:
        raise DevicesCatalogError(f"{context} contains no records")
    missing = [header for header in headers if header not in records[0]]
    if missing:
        raise DevicesCatalogError(f"{context} is missing required headers: {', '.join(missing)}")


def source_device_rows() -> list[dict[str, Any]]:
    """Return the complete retained Device population in source-row order."""
    records = read_records(PROPOSAL_PATH, RETAINED_SHEET, 1)
    required = [
        "Record Category",
        "Source Row Number",
        "Legacy SKU",
        "Legacy Group",
        "Legacy Type",
        "Legacy Manufacturer",
        "Legacy Name",
        "Legacy Retail Price",
        "Legacy Cost",
        "Legacy Note",
        "Source Workbook",
        "Source Worksheet",
        "Import Batch ID",
        "Source Created At",
        "Source Updated At",
    ]
    require_headers(records, required, f"{PROPOSAL_PATH}!{RETAINED_SHEET}")
    devices = [record for record in records if text(record["Record Category"]) == "Device"]
    if len(devices) != EXPECTED_DEVICE_ROWS:
        raise DevicesCatalogError(f"Expected {EXPECTED_DEVICE_ROWS} retained Device rows; found {len(devices)}")
    try:
        devices.sort(key=lambda record: int(record["Source Row Number"]))
    except (TypeError, ValueError) as exc:
        raise DevicesCatalogError("Every Device source row requires an integer Source Row Number") from exc
    source_numbers = [int(record["Source Row Number"]) for record in devices]
    if len(source_numbers) != len(set(source_numbers)):
        raise DevicesCatalogError("Retained Device Source Row Numbers are not unique")
    exclusion_records = read_records(PROPOSAL_PATH, EXCLUSIONS_SHEET, 1)
    excluded_numbers = {
        int(record["Excluded Source Row Number"])
        for record in exclusion_records
        if record.get("Excluded Source Row Number") not in (None, "")
    }
    overlap = sorted(set(source_numbers) & excluded_numbers)
    if overlap:
        raise DevicesCatalogError(f"Retained Device rows also occur in duplicate exclusions: {overlap}")
    return devices


def read_existing_device_ids() -> tuple[set[str], list[str], int]:
    """Read and validate the authoritative canonical Device ID namespace."""
    workbook = load_workbook(CANONICAL_PATH, read_only=True, data_only=True, keep_vba=True)
    try:
        if CANONICAL_DEVICE_SHEET not in workbook.sheetnames:
            raise DevicesCatalogError(f"Canonical worksheet {CANONICAL_DEVICE_SHEET!r} does not exist")
        worksheet = workbook[CANONICAL_DEVICE_SHEET]
        headers = [text(cell.value) for cell in worksheet[1]]
        if "Device ID" not in headers:
            raise DevicesCatalogError(f"Canonical {CANONICAL_DEVICE_SHEET!r} has no Device ID header")
        column = headers.index("Device ID")
        populated = [
            str(row[column])
            for row in worksheet.iter_rows(min_row=2, values_only=True)
            if row[column] is not None and str(row[column]).strip()
        ]
    finally:
        workbook.close()
    valid_list = [value for value in populated if DEVICE_ID_PATTERN.fullmatch(value)]
    malformed = [value for value in populated if not DEVICE_ID_PATTERN.fullmatch(value)]
    duplicates = sorted(value for value, count in Counter(valid_list).items() if count > 1)
    if duplicates:
        raise DevicesCatalogError(f"Canonical Device ID source contains duplicates: {duplicates}")
    highest = max((int(value[3:]) for value in valid_list), default=0)
    return set(valid_list), malformed, highest


def allocate_device_ids(existing: set[str], highest: int, count: int) -> list[str]:
    """Allocate one collision-free continuous Device ID sequence."""
    generated = [f"DEV{number:06d}" for number in range(highest + 1, highest + count + 1)]
    if any(not DEVICE_ID_PATTERN.fullmatch(value) for value in generated):
        raise DevicesCatalogError("Generated Device ID exceeds the six-digit namespace")
    if set(generated) & existing:
        raise DevicesCatalogError("Generated Device IDs collide with canonical IDs")
    return generated


def manufacturer_rows() -> list[dict[str, Any]]:
    """Read the canonical manufacturer lookup snapshot."""
    rows = read_records(CANONICAL_PATH, "30 Manufacturers", 3, keep_vba=True)
    require_headers(rows, ["Manufacturer ID", "Manufacturer"], "canonical Manufacturers")
    result = []
    for row in rows:
        identifier = text(row["Manufacturer ID"])
        name = text(row["Manufacturer"])
        if identifier and name:
            result.append(
                {
                    "Manufacturer ID": identifier,
                    "Manufacturer Name": name,
                    "Website": row.get("Website"),
                    "Active": "Yes" if bool(row.get("Active")) else "No",
                    "Notes": row.get("Notes"),
                }
            )
    if not result:
        raise DevicesCatalogError("Canonical manufacturer lookup is empty")
    return result


def family_rows() -> list[dict[str, Any]]:
    """Read the canonical Device Family lookup snapshot."""
    rows = read_records(CANONICAL_PATH, "31 Device Families", 3, keep_vba=True)
    require_headers(rows, ["Device Family Code", "Device Family"], "canonical Device Families")
    result = []
    for row in rows:
        code = text(row["Device Family Code"])
        name = text(row["Device Family"])
        if code and name:
            result.append(
                {
                    "Device Family Code": code,
                    "Device Family Name": name,
                    "Description": row.get("Description"),
                    "Active": "Yes" if bool(row.get("Active")) else "No",
                }
            )
    if not result:
        raise DevicesCatalogError("Canonical Device Family lookup is empty")
    return result


def exact_name_map(rows: Sequence[dict[str, Any]], id_field: str, name_field: str) -> dict[str, tuple[str, str]]:
    """Build a case-insensitive lookup while rejecting ambiguous names."""
    grouped: dict[str, list[tuple[str, str]]] = {}
    for row in rows:
        identifier = text(row[id_field])
        name = text(row[name_field])
        grouped.setdefault(name.casefold(), []).append((identifier, name))
    return {key: values[0] for key, values in grouped.items() if len(values) == 1}


def map_family(legacy_type: Any, family_map: dict[str, tuple[str, str]]) -> tuple[str, str]:
    """Map only an explicit `Device - <family>` source type."""
    value = text(legacy_type)
    if not value.casefold().startswith("device - "):
        return "", ""
    candidate = value.split("-", 1)[1].strip()
    return family_map.get(candidate.casefold(), ("", ""))


def build_device_rows(
    sources: Sequence[dict[str, Any]],
    device_ids: Sequence[str],
    manufacturers: Sequence[dict[str, Any]],
    families: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Create conservative proposed device identities."""
    manufacturer_map = exact_name_map(manufacturers, "Manufacturer ID", "Manufacturer Name")
    family_map = exact_name_map(families, "Device Family Code", "Device Family Name")
    rows = []
    for source, device_id in zip(sources, device_ids, strict=True):
        manufacturer_name = text(source["Legacy Manufacturer"])
        manufacturer_id, _canonical_name = manufacturer_map.get(
            manufacturer_name.casefold(), ("", "")
        )
        family_code, family_name = map_family(source["Legacy Type"], family_map)
        if not manufacturer_id:
            review_status = "Pending Manufacturer Review"
        elif not family_code:
            review_status = "Pending Family Review"
        else:
            review_status = "Pending Identity Review"
        row = {
            "Device ID": device_id,
            "Legacy Device SKU": text(source["Legacy SKU"]),
            "Active": "Yes",
            "Status": "Draft",
            "Manufacturer ID": manufacturer_id,
            "Manufacturer Name": manufacturer_name,
            "Device Family Code": family_code,
            "Device Family Name": family_name,
            "Product Line": text(source["Legacy Group"]),
            "Device Series": "",
            "Device Name": text(source["Legacy Name"]),
            "Device Display Name": text(source["Legacy Name"]),
            "Device Description": text(source["Legacy Note"]),
            "Model Number": "",
            "Variant": "",
            "Generation": "",
            "Release Year": "",
            "Form Factor": "",
            "Operating System Family": "",
            "Network Type": "",
            "Storage Capacity": "",
            "Memory Capacity": "",
            "Color": "",
            "Region": "",
            "Carrier": "",
            "Wi-Fi Only": "No",
            "Cellular Capable": "No",
            "Repair Supported": "No",
            "Parts Supported": "No",
            "Mail-In Eligible": "No",
            "Mobile Service Eligible": "No",
            "Compatibility Status": "Pending Compatibility Review",
            "Service Mapping Status": "Pending Service Mapping",
            "Parts Mapping Status": "Pending Parts Mapping",
            "Review Status": review_status,
            "Legacy Retail Price": decimal_value(
                source["Legacy Retail Price"],
                f"Source row {source['Source Row Number']} Legacy Retail Price",
            ),
            "Legacy Cost": decimal_value(
                source["Legacy Cost"],
                f"Source row {source['Source Row Number']} Legacy Cost",
            ),
            "Currency": "",
            "Source Record Number": int(source["Source Row Number"]),
            "Source Workbook": text(source["Source Workbook"]),
            "Source Worksheet": text(source["Source Worksheet"]),
            "Import Batch ID": IMPORT_BATCH_ID,
            "Reviewer": "",
            "Reviewer Notes": text(source["Legacy Note"]),
            "Effective Date": "",
            "Last Reviewed": "",
            "Created At": source["Source Created At"],
            "Updated At": source["Source Updated At"],
        }
        if list(row) != DEVICE_HEADERS:
            raise DevicesCatalogError("Internal Device row order does not match the approved schema")
        rows.append(row)
    return rows


def build_identity_rows(devices: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build a complete identity review queue."""
    rows = []
    for device in devices:
        missing = [
            field
            for field in (
                "Manufacturer ID",
                "Device Family Code",
                "Device Series",
                "Model Number",
                "Variant",
            )
            if not text(device[field])
        ]
        rows.append(
            {
                "Device ID": device["Device ID"],
                "Legacy Device SKU": device["Legacy Device SKU"],
                "Manufacturer Name": device["Manufacturer Name"],
                "Device Family Name": device["Device Family Name"],
                "Product Line": device["Product Line"],
                "Device Series": device["Device Series"],
                "Device Name": device["Device Name"],
                "Model Number": device["Model Number"],
                "Variant": device["Variant"],
                "Missing Identity Inputs": "; ".join(missing),
                "Identity Concern": "Legacy identity requires canonical model and variant review",
                "Required Action": "Confirm manufacturer, family, product line, model, and variant",
                "Review Status": device["Review Status"],
                "Reviewer Notes": "",
            }
        )
    return rows


def build_relationship_rows(devices: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build a complete unresolved relationship review queue."""
    return [
        {
            "Device ID": device["Device ID"],
            "Device Name": device["Device Name"],
            "Manufacturer ID": device["Manufacturer ID"],
            "Device Family Code": device["Device Family Code"],
            "Compatibility Status": device["Compatibility Status"],
            "Service Mapping Status": device["Service Mapping Status"],
            "Parts Mapping Status": device["Parts Mapping Status"],
            "Missing Relationships": "Compatibility; Service mapping; Parts mapping",
            "Required Action": "Review relationships after device identity approval",
            "Review Status": "Pending Relationship Review",
            "Reviewer Notes": "",
        }
        for device in devices
    ]


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write an ASCII-safe table and apply baseline formatting."""
    worksheet.append(list(headers))
    materialized = list(rows) or [{header: "" for header in headers}]
    for row in materialized:
        worksheet.append(
            [excel_safe_value(ascii_value(row.get(header, ""))) for header in headers]
        )
    end_column = get_column_letter(len(headers))
    table = Table(displayName=table_name, ref=f"A1:{end_column}{worksheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = table.ref
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
    for index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 34)
        for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for value_cell in cell:
                if value_cell.value is not None:
                    width = min(max(width, len(str(value_cell.value)) + 2), 34)
        worksheet.column_dimensions[get_column_letter(index)].width = width
        if header in {"Legacy Retail Price", "Legacy Cost"}:
            for cell in worksheet.iter_cols(
                min_col=index, max_col=index, min_row=2, max_row=worksheet.max_row
            ):
                for value_cell in cell:
                    value_cell.number_format = '"$"#,##0.00'
        if header in {"Effective Date"}:
            for row_cells in worksheet.iter_rows(
                min_col=index, max_col=index, min_row=2, max_row=worksheet.max_row
            ):
                row_cells[0].number_format = "yyyy-mm-dd"
        if header in {"Last Reviewed", "Created At", "Updated At", "Generated At UTC", "Revision Date"}:
            for row_cells in worksheet.iter_rows(
                min_col=index, max_col=index, min_row=2, max_row=worksheet.max_row
            ):
                row_cells[0].number_format = "yyyy-mm-dd hh:mm:ss"


def control_rows() -> list[dict[str, str]]:
    """Build aligned controlled lists for the status worksheet."""
    headers = [
        "Device Status",
        "Review Status",
        "Compatibility Status",
        "Service Mapping Status",
        "Parts Mapping Status",
        "Yes/No",
        "Currency",
    ]
    lists = [
        DEVICE_STATUSES,
        REVIEW_STATUSES,
        COMPATIBILITY_STATUSES,
        SERVICE_MAPPING_STATUSES,
        PARTS_MAPPING_STATUSES,
        YES_NO,
        CURRENCIES,
    ]
    return [
        {header: values[index] if index < len(values) else "" for header, values in zip(headers, lists, strict=True)}
        for index in range(max(map(len, lists)))
    ]


def add_defined_name(workbook: Workbook, name: str, sheet: str, column: int, row_count: int) -> None:
    """Add one workbook-scoped lookup name."""
    end_row = max(2, row_count + 1)
    column_letter = get_column_letter(column)
    workbook.defined_names.add(
        DefinedName(name, attr_text=f"'{sheet}'!${column_letter}$2:${column_letter}${end_row}")
    )


def add_defined_names(workbook: Workbook) -> None:
    """Create every approved list-validation name."""
    specs = {
        "DV_ManufacturerIDs": ("02 - Manufacturers", 1),
        "DV_DeviceFamilyCodes": ("03 - Device Families", 1),
        "DV_ProductLines": ("04 - Product Lines", 1),
        "DV_DeviceSeries": ("05 - Device Series", 1),
        "DV_FormFactors": ("06 - Form Factors", 1),
        "DV_OperatingSystemFamilies": ("07 - Operating Systems", 1),
        "DV_NetworkTypes": ("08 - Network Types", 1),
        "DV_DeviceStatuses": ("09 - Device Statuses", 1),
        "DV_ReviewStatuses": ("09 - Device Statuses", 2),
        "DV_CompatibilityStatuses": ("09 - Device Statuses", 3),
        "DV_ServiceMappingStatuses": ("09 - Device Statuses", 4),
        "DV_PartsMappingStatuses": ("09 - Device Statuses", 5),
        "DV_YesNo": ("09 - Device Statuses", 6),
        "DV_Currencies": ("09 - Device Statuses", 7),
    }
    for name, (sheet, column) in specs.items():
        worksheet = workbook[sheet]
        nonblank = sum(
            1 for row in range(2, worksheet.max_row + 1) if text(worksheet.cell(row, column).value)
        )
        add_defined_name(workbook, name, sheet, column, nonblank)


def add_data_validations(workbook: Workbook) -> None:
    """Apply named-list validations to every controlled primary field."""
    worksheet = workbook["01 - Master Devices"]
    for header, defined_name in DEFINED_NAME_BY_HEADER.items():
        column = DEVICE_HEADERS.index(header) + 1
        validation = DataValidation(type="list", formula1=f"={defined_name}", allow_blank=True)
        validation.error = f"Select a value from {defined_name}."
        validation.errorTitle = "Invalid controlled value"
        validation.showErrorMessage = True
        worksheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}2:{get_column_letter(column)}{worksheet.max_row}")


def add_conditional_formatting(worksheet: Worksheet) -> None:
    """Highlight pending states and unresolved relationships."""
    if worksheet.max_row < 2:
        return
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="F4CCCC")
    headers = [text(cell.value) for cell in worksheet[1]]
    for header in (
        "Review Status",
        "Compatibility Status",
        "Service Mapping Status",
        "Parts Mapping Status",
    ):
        if header not in headers:
            continue
        column = headers.index(header) + 1
        letter = get_column_letter(column)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{worksheet.max_row}",
            FormulaRule(formula=[f'ISNUMBER(SEARCH("Pending",{letter}2))'], fill=yellow),
        )
    for header in ("Missing Identity Inputs", "Missing Relationships"):
        if header not in headers:
            continue
        column = headers.index(header) + 1
        letter = get_column_letter(column)
        worksheet.conditional_formatting.add(
            f"{letter}2:{letter}{worksheet.max_row}",
            FormulaRule(formula=[f'LEN(TRIM({letter}2))>0'], fill=red),
        )


def validation_rows(
    devices: Sequence[dict[str, Any]],
    highest_existing: int,
    malformed_ids: Sequence[str],
) -> list[dict[str, str]]:
    """Create the visible generator validation summary."""
    first_id = text(devices[0]["Device ID"]) if devices else ""
    final_id = text(devices[-1]["Device ID"]) if devices else ""
    highest_label = (
        f"DEV{highest_existing:06d}"
        if highest_existing
        else "None - ADR-009 empty namespace"
    )
    return [
        {"Validation": "Retained Device rows", "Result": "PASS", "Details": str(len(devices))},
        {"Validation": "Master Devices schema", "Result": "PASS", "Details": f"{len(DEVICE_HEADERS)} columns"},
        {"Validation": "Worksheet contract", "Result": "PASS", "Details": f"{len(SHEET_NAMES)} worksheets"},
        {
            "Validation": "Canonical Device ID source",
            "Result": "PASS",
            "Details": f"Highest valid ID {highest_label}",
        },
        {
            "Validation": "Generated Device ID sequence",
            "Result": "PASS",
            "Details": f"{first_id} through {final_id}",
        },
        {
            "Validation": "Malformed canonical Device IDs",
            "Result": "PASS",
            "Details": "; ".join(malformed_ids) if malformed_ids else "None",
        },
        {
            "Validation": "Protected input hashes",
            "Result": "PASS",
            "Details": "Verified before and after generation",
        },
        {
            "Validation": "Canonical import",
            "Result": "PASS",
            "Details": "Not performed; review artifact only",
        },
    ]


def build_workbook(
    devices: Sequence[dict[str, Any]],
    manufacturers: Sequence[dict[str, Any]],
    families: Sequence[dict[str, Any]],
    protected_hashes: dict[Path, str],
    highest_existing: int,
    malformed_ids: Sequence[str],
) -> Workbook:
    """Build the complete review workbook in memory."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in SHEET_NAMES:
        workbook.create_sheet(sheet_name)

    instructions = [
        {"Step": 1, "Instruction": "Review identity rows before relationship rows.", "Owner": "Catalog Governance"},
        {"Step": 2, "Instruction": "Resolve manufacturer aliases without changing the legacy observation.", "Owner": "Catalog Governance"},
        {"Step": 3, "Instruction": "Confirm family, product line, model, and variant using reliable evidence.", "Owner": "Product Catalog"},
        {"Step": 4, "Instruction": "Review compatibility, service mapping, and parts mapping after identity approval.", "Owner": "Repair Engineering"},
        {"Step": 5, "Instruction": "Do not import this workbook or publish legacy monetary observations.", "Owner": "Data Engineering"},
    ]
    append_table(
        workbook["00 - Instructions"],
        ["Step", "Instruction", "Owner"],
        instructions,
        TABLE_NAMES["00 - Instructions"],
    )
    append_table(
        workbook["01 - Master Devices"],
        DEVICE_HEADERS,
        devices,
        TABLE_NAMES["01 - Master Devices"],
    )
    append_table(
        workbook["02 - Manufacturers"],
        ["Manufacturer ID", "Manufacturer Name", "Website", "Active", "Notes"],
        manufacturers,
        TABLE_NAMES["02 - Manufacturers"],
    )
    append_table(
        workbook["03 - Device Families"],
        ["Device Family Code", "Device Family Name", "Description", "Active"],
        families,
        TABLE_NAMES["03 - Device Families"],
    )
    product_lines = sorted({text(row["Product Line"]) for row in devices if text(row["Product Line"])}, key=str.casefold)
    append_table(
        workbook["04 - Product Lines"],
        ["Product Line", "Status", "Notes"],
        [{"Product Line": value, "Status": "Proposed", "Notes": "Legacy Group observation"} for value in product_lines],
        TABLE_NAMES["04 - Product Lines"],
    )
    append_table(
        workbook["05 - Device Series"],
        ["Device Series", "Status", "Notes"],
        [{"Device Series": "Unresolved", "Status": "Review Required", "Notes": "No series is generated automatically"}],
        TABLE_NAMES["05 - Device Series"],
    )
    append_table(
        workbook["06 - Form Factors"],
        ["Form Factor", "Status", "Notes"],
        [{"Form Factor": "Unresolved", "Status": "Review Required", "Notes": "No form factor is generated automatically"}],
        TABLE_NAMES["06 - Form Factors"],
    )
    append_table(
        workbook["07 - Operating Systems"],
        ["Operating System Family", "Status", "Notes"],
        [{"Operating System Family": "Unresolved", "Status": "Review Required", "Notes": "No OS is generated automatically"}],
        TABLE_NAMES["07 - Operating Systems"],
    )
    append_table(
        workbook["08 - Network Types"],
        ["Network Type", "Status", "Notes"],
        [{"Network Type": "Unresolved", "Status": "Review Required", "Notes": "No network type is generated automatically"}],
        TABLE_NAMES["08 - Network Types"],
    )
    status_headers = [
        "Device Status",
        "Review Status",
        "Compatibility Status",
        "Service Mapping Status",
        "Parts Mapping Status",
        "Yes/No",
        "Currency",
    ]
    append_table(
        workbook["09 - Device Statuses"],
        status_headers,
        control_rows(),
        TABLE_NAMES["09 - Device Statuses"],
    )
    append_table(
        workbook["10 - Identity Review"],
        IDENTITY_HEADERS,
        build_identity_rows(devices),
        TABLE_NAMES["10 - Identity Review"],
    )
    append_table(
        workbook["11 - Relationship Review"],
        RELATIONSHIP_HEADERS,
        build_relationship_rows(devices),
        TABLE_NAMES["11 - Relationship Review"],
    )
    append_table(
        workbook["12 - Validation Summary"],
        ["Validation", "Result", "Details"],
        validation_rows(devices, highest_existing, malformed_ids),
        TABLE_NAMES["12 - Validation Summary"],
    )
    append_table(
        workbook["13 - Revision History"],
        ["Revision", "Revision Date", "Author", "Status", "Notes"],
        [
            {
                "Revision": "v1",
                "Revision Date": datetime.now(UTC),
                "Author": "Nocturnix Catalog Generator",
                "Status": "Review Draft",
                "Notes": "Initial Master Devices review package",
            }
        ],
        TABLE_NAMES["13 - Revision History"],
    )
    metadata = [
        {"Metadata Key": "Artifact Type", "Metadata Value": "Master Devices Catalog V1 review workbook"},
        {"Metadata Key": "Generated At UTC", "Metadata Value": datetime.now(UTC)},
        {"Metadata Key": "Import Batch ID", "Metadata Value": IMPORT_BATCH_ID},
        {"Metadata Key": "Namespace Authority", "Metadata Value": NAMESPACE_AUTHORITY},
        {"Metadata Key": "Expected Device Rows", "Metadata Value": len(devices)},
        {"Metadata Key": "Schema Columns", "Metadata Value": len(DEVICE_HEADERS)},
        {
            "Metadata Key": "Highest Existing Device ID",
            "Metadata Value": (
                f"DEV{highest_existing:06d}"
                if highest_existing
                else "None - ADR-009 empty namespace"
            ),
        },
        {"Metadata Key": "First Proposed Device ID", "Metadata Value": devices[0]["Device ID"]},
        {"Metadata Key": "Final Proposed Device ID", "Metadata Value": devices[-1]["Device ID"]},
        {"Metadata Key": "Malformed Existing Device IDs", "Metadata Value": "; ".join(malformed_ids)},
        {"Metadata Key": "Canonical Write Performed", "Metadata Value": "No"},
    ]
    metadata.extend(
        {
            "Metadata Key": f"SHA-256 {path.name}",
            "Metadata Value": digest,
        }
        for path, digest in protected_hashes.items()
    )
    append_table(
        workbook["14 - Import Metadata"],
        ["Metadata Key", "Metadata Value"],
        metadata,
        TABLE_NAMES["14 - Import Metadata"],
    )
    metadata_sheet = workbook["14 - Import Metadata"]
    for row in range(2, metadata_sheet.max_row + 1):
        if text(metadata_sheet.cell(row, 1).value) == "Generated At UTC":
            metadata_sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
    add_defined_names(workbook)
    add_data_validations(workbook)
    for worksheet in workbook.worksheets:
        add_conditional_formatting(worksheet)
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read one Excel Table into dictionaries."""
    if table_name not in worksheet.tables:
        raise DevicesCatalogError(f"Missing table {table_name} on {worksheet.title}")
    min_column, min_row, max_column, max_row = range_boundaries(worksheet.tables[table_name].ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    return [
        {
            header: worksheet.cell(row, column).value
            for header, column in zip(headers, range(min_column, max_column + 1), strict=True)
        }
        for row in range(min_row + 1, max_row + 1)
        if any(
            worksheet.cell(row, column).value not in (None, "")
            for column in range(min_column, max_column + 1)
        )
    ]


def require_excel_archive(path: Path) -> None:
    """Reject missing, empty, or corrupt generated output."""
    if not path.exists() or path.stat().st_size <= 0 or not zipfile.is_zipfile(path):
        raise DevicesCatalogError("Generated devices workbook is missing or invalid; rerun the generator successfully.")
    with zipfile.ZipFile(path) as archive:
        required_members = {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
        missing = sorted(required_members - set(archive.namelist()))
    if missing:
        raise DevicesCatalogError(f"Generated devices workbook is missing OOXML members: {missing}")


def validate_structure(workbook: Workbook) -> None:
    """Validate workbook, worksheet, table, and schema structure."""
    if workbook.sheetnames != SHEET_NAMES:
        raise DevicesCatalogError(f"Reopened worksheet order differs: {workbook.sheetnames!r}")
    if len(set(SHEET_NAMES)) != len(SHEET_NAMES) or any(len(name) > 31 for name in SHEET_NAMES):
        raise DevicesCatalogError("Configured worksheet titles must be unique and at most 31 characters")
    table_names = []
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        expected_table = TABLE_NAMES[sheet_name]
        if expected_table not in worksheet.tables:
            raise DevicesCatalogError(f"Missing {expected_table} on {sheet_name}")
        if worksheet.freeze_panes != "A2":
            raise DevicesCatalogError(f"{sheet_name} does not freeze the header row")
        table_names.extend(table.name for table in worksheet.tables.values())
    if len(table_names) != len(set(table_names)):
        raise DevicesCatalogError("Excel Table names are not unique")
    primary = workbook["01 - Master Devices"]
    headers = [text(cell.value) for cell in primary[1]]
    if headers != DEVICE_HEADERS:
        raise DevicesCatalogError("Reopened Master Devices schema differs from the approved 48 columns")
    if PROHIBITED_HEADERS & set(headers):
        raise DevicesCatalogError("Master Devices includes prohibited inventory or final-pricing fields")


def validate_names_and_validations(workbook: Workbook) -> None:
    """Validate all defined names and primary list formulas."""
    required = set(DEFINED_NAME_SHEET)
    actual = set(workbook.defined_names)
    missing = sorted(required - actual)
    if missing:
        raise DevicesCatalogError(f"Required defined names are missing: {missing}")
    for name, expected_sheet in DEFINED_NAME_SHEET.items():
        defined_name = workbook.defined_names[name]
        destinations = list(defined_name.destinations)
        if len(destinations) != 1 or destinations[0][0] != expected_sheet:
            raise DevicesCatalogError(f"Defined name {name} does not resolve to {expected_sheet}")
        try:
            range_boundaries(destinations[0][1].replace("$", ""))
        except ValueError as exc:
            raise DevicesCatalogError(f"Defined name {name} has an invalid destination") from exc
    worksheet = workbook["01 - Master Devices"]
    validations = list(worksheet.data_validations.dataValidation)
    formulas = Counter(text(validation.formula1) for validation in validations)
    expected_formulas = Counter(
        f"={defined_name}" for defined_name in DEFINED_NAME_BY_HEADER.values()
    )
    if formulas != expected_formulas:
        raise DevicesCatalogError("Reopened list-validation formulas differ from the approved defined names")
    for validation in validations:
        if "!" in text(validation.formula1):
            raise DevicesCatalogError("A direct cross-sheet list validation formula was found")
        ranges = list(validation.ranges.ranges)
        if len(ranges) != 1:
            raise DevicesCatalogError(f"{validation.formula1} has unexpected validation ranges")
        min_column, min_row, max_column, max_row = range_boundaries(str(ranges[0]))
        if (
            min_column != max_column
            or min_row != 2
            or max_row != worksheet.max_row
        ):
            raise DevicesCatalogError(f"{validation.formula1} has incomplete row coverage")
        header = DEVICE_HEADERS[min_column - 1]
        if text(validation.formula1) != f"={DEFINED_NAME_BY_HEADER[header]}":
            raise DevicesCatalogError(f"{header} uses the wrong defined-name validation")


def validate_reopened(
    path: Path,
    expected_devices: Sequence[dict[str, Any]],
    existing_ids: set[str],
    highest_existing: int,
) -> None:
    """Reopen and validate the generated review artifact before replacement."""
    require_excel_archive(path)
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        validate_structure(workbook)
        validate_names_and_validations(workbook)
        actual = table_records(workbook["01 - Master Devices"], TABLE_NAMES["01 - Master Devices"])
        if len(actual) != EXPECTED_DEVICE_ROWS:
            raise DevicesCatalogError(f"Reopened workbook contains {len(actual)} Device rows")
        actual_ids = [text(row["Device ID"]) for row in actual]
        expected_ids = [text(row["Device ID"]) for row in expected_devices]
        if actual_ids != expected_ids:
            raise DevicesCatalogError("Reopened Device IDs or source ordering changed")
        if len(actual_ids) != len(set(actual_ids)) or any(not DEVICE_ID_PATTERN.fullmatch(value) for value in actual_ids):
            raise DevicesCatalogError("Reopened Device IDs are duplicate or malformed")
        if set(actual_ids) & existing_ids:
            raise DevicesCatalogError("Reopened Device IDs collide with canonical IDs")
        expected_numbers = list(range(highest_existing + 1, highest_existing + 1 + len(actual_ids)))
        if [int(value[3:]) for value in actual_ids] != expected_numbers:
            raise DevicesCatalogError("Reopened Device IDs do not form the required continuous sequence")
        source_numbers = [int(row["Source Record Number"]) for row in actual]
        if source_numbers != sorted(source_numbers) or len(source_numbers) != len(set(source_numbers)):
            raise DevicesCatalogError("Reopened source provenance is reordered or duplicated")
        for row_number, (expected, reopened) in enumerate(zip(expected_devices, actual, strict=True), start=2):
            for field in DEVICE_HEADERS:
                expected_value = expected[field]
                actual_value = reopened[field]
                if field in {"Legacy Retail Price", "Legacy Cost"}:
                    if decimal_value(expected_value, field) != decimal_value(actual_value, field):
                        raise DevicesCatalogError(f"Reopened row {row_number} changed {field}")
                elif excel_safe_value(expected_value) != excel_safe_value(actual_value):
                    if text(excel_safe_value(expected_value)) != text(excel_safe_value(actual_value)):
                        raise DevicesCatalogError(f"Reopened row {row_number} changed {field}")
            if text(reopened["Status"]) != "Draft" or text(reopened["Review Status"]) in {"Approved", "Ready for Approval"}:
                raise DevicesCatalogError(f"Reopened row {row_number} has an impermissible approval status")
            if any(text(reopened[field]) for field in GENERATED_BLANK_FIELDS):
                raise DevicesCatalogError(f"Reopened row {row_number} contains an inferred V1 field")
        for sheet_name, table_name in (
            ("10 - Identity Review", "tblDeviceIdentityReview"),
            ("11 - Relationship Review", "tblDeviceRelationshipReview"),
        ):
            ids = [text(row["Device ID"]) for row in table_records(workbook[sheet_name], table_name)]
            if ids != actual_ids:
                raise DevicesCatalogError(f"{sheet_name} does not contain every Device exactly once in order")
    finally:
        workbook.close()


def main() -> int:
    """Generate, validate, and atomically publish the review workbook."""
    workbook: Workbook | None = None
    try:
        assert_excel_safe_value_contract()
        require_files(PROTECTED_PATHS)
        before_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        sources = source_device_rows()
        existing_ids, malformed_ids, highest_existing = read_existing_device_ids()
        manufacturers = manufacturer_rows()
        families = family_rows()
        device_ids = allocate_device_ids(existing_ids, highest_existing, len(sources))
        devices = build_device_rows(sources, device_ids, manufacturers, families)
        workbook = build_workbook(
            devices,
            manufacturers,
            families,
            before_hashes,
            highest_existing,
            malformed_ids,
        )
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        if TEMP_OUTPUT_PATH.exists():
            TEMP_OUTPUT_PATH.unlink()
        workbook.save(TEMP_OUTPUT_PATH)
        workbook.close()
        workbook = None
        validate_reopened(TEMP_OUTPUT_PATH, devices, existing_ids, highest_existing)
        after_hashes = {path: file_hash(path) for path in PROTECTED_PATHS}
        if before_hashes != after_hashes:
            raise DevicesCatalogError("A protected source workbook hash changed during generation")
        os.replace(TEMP_OUTPUT_PATH, OUTPUT_PATH)
        print("Master Devices Catalog V1 validation: PASS")
        print(f"Generated workbook: {OUTPUT_PATH}")
        print(f"Retained Device rows: {len(devices)}")
        print(f"Schema columns: {len(DEVICE_HEADERS)}")
        print(f"Device ID range: {device_ids[0]} through {device_ids[-1]}")
        print(f"Identity review rows: {len(devices)}")
        print(f"Relationship review rows: {len(devices)}")
        print("Protected input hashes: PASS")
        print("Canonical database writes: NONE")
        return 0
    except (
        AssertionError,
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        DevicesCatalogError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except OSError as close_error:
                print(f"ERROR: Failed to close workbook: {close_error}", file=sys.stderr)
        try:
            if TEMP_OUTPUT_PATH.exists():
                TEMP_OUTPUT_PATH.unlink()
        except OSError as cleanup_error:
            print(f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}", file=sys.stderr)


if __name__ == "__main__":
    raise SystemExit(main())
