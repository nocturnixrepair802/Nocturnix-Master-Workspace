"""Read-only cross-source validator and readiness reporter for Pilot v0.1."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from generate_competitive_pricing_pilot import (
    CANONICAL_SHA256,
    SHEETS,
    validate_workbook,
)
from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def table_rows(ws: Any) -> list[dict[str, Any]]:
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, values, strict=False))
        for values in ws.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in values)
    ]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    parser.add_argument("--json", dest="json_path", type=Path, required=True)
    args = parser.parse_args()
    before = sha256(args.workbook)
    canonical_hash = sha256(args.canonical)
    errors = validate_workbook(args.workbook)
    if canonical_hash != CANONICAL_SHA256:
        errors.append("Canonical source SHA-256 mismatch")
    wb = load_workbook(args.workbook, read_only=False, data_only=False)
    try:
        pricing = table_rows(wb["01 - Pilot Pricing Records"])
        canonical = table_rows(wb["05 - Canonical References"])
        costs = table_rows(wb["03 - Supplier Cost Evidence"])
        labor = table_rows(wb["04 - Labor References"])
        policies = table_rows(wb["06 - Pricing Policy"])
        calculations = table_rows(wb["07 - Calculation Results"])
        if wb.sheetnames != SHEETS:
            errors.append("Worksheet order differs from contract")
        canonical_map = {
            row["Pricing Record ID"]: (
                row["Service ID"],
                row["Canonical Service Type ID"],
                row["Canonical Service Type"],
            )
            for row in canonical
        }
        for row in pricing:
            expected = canonical_map.get(row["Pricing Record ID"])
            actual = (
                row["Service ID"],
                row["Canonical Service Type ID"],
                row["Canonical Service Type"],
            )
            if expected != actual:
                errors.append(f"Canonical mapping mismatch: {row['Pricing Record ID']}")
        cost_by_pricing = {row["Pricing Record ID"]: row for row in costs}
        labor_by_pricing = {row["Pricing Record ID"]: row for row in labor}
        missing_cost = [
            row["Pricing Record ID"]
            for row in pricing
            if not cost_by_pricing.get(row["Pricing Record ID"], {}).get("Part Cost")
        ]
        missing_labor = [
            row["Pricing Record ID"]
            for row in pricing
            if not labor_by_pricing.get(row["Pricing Record ID"], {}).get("Labor Standard ID")
        ]
        pending_policies = [
            row["Policy Key"]
            for row in policies
            if row["Approval Status"] != "Approved"
        ]
        ready_for_research = [
            row["Pricing Record ID"]
            for row in pricing
            if row["Candidate Status"] == "Pending Research"
        ]
        if any(row["Final Approved Price"] not in (None, "") for row in calculations):
            errors.append("Final Approved Price must be blank for every record")
    finally:
        wb.close()
    after = sha256(args.workbook)
    if before != after:
        errors.append("Read-only validator changed the workbook")
    result = {
        "profile": "Nocturnix Competitive Pricing Pilot v0.1 cross-source QA",
        "validated_at_utc": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "workbook": str(args.workbook.resolve()),
        "workbook_sha256_before": before,
        "workbook_sha256_after": after,
        "validator_read_only": before == after,
        "canonical_sha256": canonical_hash,
        "status": "PASS" if not errors else "FAIL",
        "errors": errors,
        "counts": {
            "pilot_records": len(pricing),
            "ready_for_competitor_research": len(ready_for_research),
            "missing_main_component_cost_evidence": len(missing_cost),
            "missing_labor_mappings": len(missing_labor),
            "pending_policy_decisions": len(pending_policies),
            "final_approved_prices": 0,
        },
        "records_ready_for_competitor_research": ready_for_research,
        "records_missing_main_component_cost_evidence": missing_cost,
        "records_missing_labor_mappings": missing_labor,
        "pending_policy_decisions": pending_policies,
        "exact_fields_to_populate_next": [
            "Governed Manufacturer ID and Device Model ID",
            "Verified Part Cost, Shipping Cost, Currency, evidence reference/date/status/reviewer",
            "Approved LAB###### mapping, Labor Minutes, Labor Rate and approval version",
            "Approved currency, market, competitor categories, observation sufficiency/freshness",
            "Approved processing fee, overhead, warranty/risk, target margin, rounding and minimum-profit policies",
            "Verified comparable competitor observations after policy approval",
        ],
    }
    args.json_path.parent.mkdir(parents=True, exist_ok=True)
    args.markdown.parent.mkdir(parents=True, exist_ok=True)
    args.json_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    lines = [
        "# Nocturnix Competitive Pricing Pilot v0.1 Readiness Report",
        "",
        f"Status: **{result['status']}**",
        "",
        f"Workbook SHA-256 before validation: `{before}`",
        "",
        f"Workbook SHA-256 after validation: `{after}`",
        "",
        f"Read-only proof: **{'PASS' if before == after else 'FAIL'}**",
        "",
        "## Readiness summary",
        "",
        f"- Records ready for competitor research workflow: {len(ready_for_research)}",
        f"- Records missing verified main-component cost evidence: {len(missing_cost)}",
        f"- Records missing governed labor mappings: {len(missing_labor)}",
        f"- Pending pricing-policy decisions: {len(pending_policies)}",
        "- Final Approved Price populated: 0",
        "",
        "“Ready for competitor research” means the approved cohort row is present and "
        "queued; competitor-price collection remains blocked until the market and "
        "observation policies are explicitly approved.",
        "",
        "## Records ready for competitor research",
        "",
        ", ".join(f"`{value}`" for value in ready_for_research),
        "",
        "## Records missing verified main-component cost evidence",
        "",
        ", ".join(f"`{value}`" for value in missing_cost),
        "",
        "## Records missing labor mappings",
        "",
        ", ".join(f"`{value}`" for value in missing_labor),
        "",
        "## Pending policy decisions",
        "",
        *[f"- `{value}`" for value in pending_policies],
        "",
        "## Exact fields to populate next",
        "",
        *[f"- {value}" for value in result["exact_fields_to_populate_next"]],
        "",
        "## Validation errors",
        "",
        *(["None."] if not errors else [f"- {value}" for value in errors]),
    ]
    args.markdown.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(result["counts"], indent=2))
    print(f"Status: {result['status']}")
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
