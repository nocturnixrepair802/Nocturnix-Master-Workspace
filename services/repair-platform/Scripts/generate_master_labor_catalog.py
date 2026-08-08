"""Generate the governed Master Labor Catalog V1 review workbook.

The script reads the protected Labor Standards workbook, creates a standalone
review artifact through a temporary sibling file, validates the persisted
artifact, and never saves a source workbook.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
CANONICAL_LABOR_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards"
    r"\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
OUTPUT_PATH = WORKING_DIR / "Nocturnix_Master_Labor_Catalog_v1.xlsx"
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(f".{OUTPUT_PATH.stem}.tmp.xlsx")
SOURCE_SHEET = "01 - Labor Standards"
IMPORT_BATCH = "MASTER-LABOR-V1-REVIEW"
LABOR_ID_PATTERN = re.compile(r"^LAB\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Labor Standards",
    "02 - Repair Categories",
    "03 - Repair Types",
    "04 - Device Families",
    "05 - Manufacturers",
    "06 - Labor Tiers",
    "07 - Skill Levels",
    "08 - Difficulty",
    "09 - Warranty Options",
    "10 - Review Queue",
    "11 - Validation Summary",
    "12 - Revision History",
    "13 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterLaborInstructions",
    "01 - Labor Standards": "tblMasterLaborCatalog",
    "02 - Repair Categories": "tblLaborRepairCategories",
    "03 - Repair Types": "tblLaborRepairTypes",
    "04 - Device Families": "tblLaborDeviceFamilies",
    "05 - Manufacturers": "tblLaborManufacturers",
    "06 - Labor Tiers": "tblLaborTiers",
    "07 - Skill Levels": "tblLaborSkillLevels",
    "08 - Difficulty": "tblLaborDifficulties",
    "09 - Warranty Options": "tblLaborWarrantyOptions",
    "10 - Review Queue": "tblLaborReviewQueue",
    "11 - Validation Summary": "tblMasterLaborValidation",
    "12 - Revision History": "tblMasterLaborRevisionHistory",
    "13 - Import Metadata": "tblMasterLaborImportMetadata",
}
LABOR_HEADERS = [
    "Labor Standard ID",
    "Legacy Labor ID",
    "Labor Name",
    "Repair Category",
    "Repair Type",
    "Device Family",
    "Manufacturer",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Labor Tier",
    "Skill Level",
    "Difficulty",
    "Warranty Option",
    "Requires Calibration",
    "Requires Waterproof Test",
    "Requires Programming",
    "Requires Pairing",
    "Requires Board Repair",
    "Special Tools Required",
    "Technician Certification",
    "Review Status",
    "Confidence",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch",
    "Reviewer",
    "Reviewer Notes",
    "Created At",
    "Updated At",
]
QUEUE_HEADERS = [
    "Labor Standard ID",
    "Labor Name",
    "Missing Evidence",
    "Relationship Issue",
    "Required Action",
    "Review Status",
    "Reviewer",
    "Reviewer Notes",
]
REVIEW_STATUSES = [
    "Pending Review",
    "Pending Evidence Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
CONFIDENCE_VALUES = ["Unassessed", "Low", "Medium", "High"]
YES_NO_VALUES = ["Yes", "No"]
FLAG_HEADERS = [
    "Requires Calibration",
    "Requires Waterproof Test",
    "Requires Programming",
    "Requires Pairing",
    "Requires Board Repair",
]
SOURCE_ALIASES: dict[str, tuple[str, ...]] = {
    "Labor Name": ("Labor Name", "Service", "Labor Standard", "Name"),
    "Repair Category": ("Repair Category", "Service Category"),
    "Repair Type": ("Repair Type", "Service"),
    "Device Family": ("Device Family", "Device Scope", "Device Category"),
    "Manufacturer": ("Manufacturer",),
    "Standard Minutes": ("Standard Minutes",),
    "Minimum Minutes": ("Minimum Minutes",),
    "Maximum Minutes": ("Maximum Minutes",),
    "Labor Tier": ("Labor Tier", "Labor Rate Tier"),
    "Skill Level": ("Skill Level",),
    "Difficulty": ("Difficulty", "Repair Difficulty"),
    "Warranty Option": ("Warranty Option", "Recommended Warranty"),
    "Requires Calibration": ("Requires Calibration",),
    "Requires Waterproof Test": ("Requires Waterproof Test",),
    "Requires Programming": ("Requires Programming",),
    "Requires Pairing": ("Requires Pairing",),
    "Requires Board Repair": ("Requires Board Repair",),
    "Special Tools Required": ("Special Tools Required",),
    "Technician Certification": ("Technician Certification",),
    "Confidence": ("Confidence",),
    "Reviewer Notes": ("Reviewer Notes", "Notes"),
    "Created At": ("Created At",),
    "Updated At": ("Updated At", "Last Reviewed"),
}
DEFINED_NAME_BY_HEADER = {
    "Repair Category": "DV_LaborRepairCategories",
    "Repair Type": "DV_LaborRepairTypes",
    "Device Family": "DV_LaborDeviceFamilies",
    "Manufacturer": "DV_LaborManufacturers",
    "Labor Tier": "DV_LaborTiers",
    "Skill Level": "DV_LaborSkillLevels",
    "Difficulty": "DV_LaborDifficulties",
    "Warranty Option": "DV_LaborWarrantyOptions",
    "Review Status": "DV_LaborReviewStatuses",
    "Confidence": "DV_LaborConfidence",
    **{header: "DV_YesNo" for header in FLAG_HEADERS},
}
DEFINED_NAME_SPECS = {
    "DV_LaborRepairCategories": ("02 - Repair Categories", 1),
    "DV_LaborRepairTypes": ("03 - Repair Types", 1),
    "DV_LaborDeviceFamilies": ("04 - Device Families", 1),
    "DV_LaborManufacturers": ("05 - Manufacturers", 1),
    "DV_LaborTiers": ("06 - Labor Tiers", 1),
    "DV_LaborSkillLevels": ("07 - Skill Levels", 1),
    "DV_LaborDifficulties": ("08 - Difficulty", 1),
    "DV_LaborWarrantyOptions": ("09 - Warranty Options", 1),
    "DV_LaborReviewStatuses": ("09 - Warranty Options", 2),
    "DV_LaborConfidence": ("09 - Warranty Options", 3),
    "DV_YesNo": ("09 - Warranty Options", 4),
}
PROHIBITED_HEADERS = {
    "Labor Rate",
    "Labor Cost",
    "Customer Price",
    "Payroll",
    "Technician Schedule",
    "Clock In",
    "Clock Out",
    "Inventory",
    "Automatic Approval",
}


class LaborCatalogError(RuntimeError):
    """Raised when a governed catalog invariant fails."""


def text(value: Any) -> str:
    """Return a trimmed display value."""
    return "" if value is None else str(value).strip()


def excel_value(value: Any) -> Any:
    """Return an Excel-safe scalar without timezone-aware datetimes."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(UTC).replace(tzinfo=None)
    if isinstance(value, str):
        return value.encode("ascii", "replace").decode("ascii")
    return value


def file_hash(path: Path) -> str:
    """Return a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require every protected source."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise LaborCatalogError(f"Missing protected input: {', '.join(missing)}")


def locate_source_header(worksheet: Worksheet) -> tuple[int, list[str]]:
    """Locate the labor source header without assuming a fixed row."""
    for row_number in range(1, min(worksheet.max_row, 50) + 1):
        headers = [text(cell.value) for cell in worksheet[row_number]]
        if (
            "Labor ID" in headers or "Legacy Labor ID" in headers
        ) and any(alias in headers for alias in SOURCE_ALIASES["Labor Name"]):
            return row_number, headers
    raise LaborCatalogError(
        f"{SOURCE_SHEET} lacks a legacy Labor ID and labor-name header contract"
    )


def read_source() -> tuple[list[dict[str, Any]], set[str], int]:
    """Read source rows and validate the protected ID namespace."""
    workbook = load_workbook(
        CANONICAL_LABOR_PATH,
        read_only=True,
        data_only=False,
    )
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise LaborCatalogError(f"Missing source worksheet: {SOURCE_SHEET}")
        worksheet = workbook[SOURCE_SHEET]
        header_row, headers = locate_source_header(worksheet)
        governed_id_header = (
            "Labor Standard ID" if "Labor Standard ID" in headers else None
        )
        rows: list[dict[str, Any]] = []
        namespace_order: list[str] = []
        for source_row in range(header_row + 1, worksheet.max_row + 1):
            values = [
                worksheet.cell(source_row, column).value
                for column in range(1, len(headers) + 1)
            ]
            if not any(text(value) for value in values):
                continue
            record = dict(zip(headers, values, strict=True))
            record["_Source Row"] = source_row
            rows.append(record)
            identifier = (
                text(record.get(governed_id_header))
                if governed_id_header is not None
                else ""
            )
            if identifier:
                if not LABOR_ID_PATTERN.fullmatch(identifier):
                    raise LaborCatalogError(
                        f"Malformed canonical Labor ID at row {source_row}: "
                        f"{identifier}"
                    )
                namespace_order.append(identifier)
    finally:
        workbook.close()

    duplicates = sorted(
        value for value, count in Counter(namespace_order).items() if count > 1
    )
    if duplicates:
        raise LaborCatalogError(f"Duplicate canonical Labor IDs: {duplicates[:10]}")
    numbers = [int(identifier[3:]) for identifier in namespace_order]
    if numbers != sorted(numbers):
        raise LaborCatalogError("Canonical Labor IDs are out of order")
    legacy_ids = [
        text(record.get("Legacy Labor ID") or record.get("Labor ID"))
        for record in rows
    ]
    if any(not identifier for identifier in legacy_ids):
        raise LaborCatalogError("Protected source contains a blank Legacy Labor ID")
    legacy_duplicates = sorted(
        value for value, count in Counter(legacy_ids).items() if count > 1
    )
    if legacy_duplicates:
        raise LaborCatalogError(
            f"Duplicate Legacy Labor IDs: {legacy_duplicates[:10]}"
        )
    existing = set(namespace_order)
    return rows, existing, max(numbers, default=0)


def first_value(record: dict[str, Any], aliases: Sequence[str]) -> Any:
    """Return the first explicitly populated aliased value."""
    for alias in aliases:
        if alias in record and text(record.get(alias)):
            return record[alias]
    return ""


def source_record_number(record: dict[str, Any]) -> Any:
    """Use explicit lineage when present, otherwise the physical source row."""
    for header in ("Source Record Number", "Record Number"):
        if text(record.get(header)):
            return record[header]
    return record["_Source Row"]


def sortable_record_number(value: Any) -> tuple[int, int | str]:
    """Sort numeric lineage first and retain deterministic text lineage."""
    try:
        return (0, int(text(value)))
    except ValueError:
        return (1, text(value).casefold())


def build_records(
    source_rows: Sequence[dict[str, Any]],
    existing: set[str],
    highest: int,
) -> list[dict[str, Any]]:
    """Copy explicit observations and allocate a governed review identity."""
    prepared: list[dict[str, Any]] = []
    for source in source_rows:
        name = first_value(source, SOURCE_ALIASES["Labor Name"])
        if not text(name):
            raise LaborCatalogError(
                f"Source row {source['_Source Row']} has no Labor Name"
            )
        confidence = text(first_value(source, SOURCE_ALIASES["Confidence"]))
        if confidence and confidence not in CONFIDENCE_VALUES:
            raise LaborCatalogError(
                f"Source row {source['_Source Row']} has invalid Confidence: "
                f"{confidence}"
            )
        legacy_value = (
            source.get("Legacy Labor ID")
            if text(source.get("Legacy Labor ID"))
            else source.get("Labor ID")
        )
        identifier = text(legacy_value)
        if not identifier:
            raise LaborCatalogError(
                f"Source row {source['_Source Row']} has no Legacy Labor ID"
            )
        record: dict[str, Any] = {header: "" for header in LABOR_HEADERS}
        for target, aliases in SOURCE_ALIASES.items():
            record[target] = first_value(source, aliases)
        record.update(
            {
                "Legacy Labor ID": legacy_value,
                "Labor Name": name,
                "Review Status": "Pending Review",
                "Confidence": confidence or "Unassessed",
                "Source Record Number": source_record_number(source),
                "Source Workbook": str(CANONICAL_LABOR_PATH),
                "Source Worksheet": SOURCE_SHEET,
                "Import Batch": IMPORT_BATCH,
                "Reviewer": "",
            }
        )
        prepared.append(record)
    prepared.sort(
        key=lambda row: (
            sortable_record_number(row["Source Record Number"]),
            text(row["Labor Name"]).casefold(),
        )
    )
    lineage = [text(row["Source Record Number"]) for row in prepared]
    if len(lineage) != len(set(lineage)):
        raise LaborCatalogError("Duplicate Source Record Number in labor source")
    for offset, record in enumerate(prepared, start=1):
        identifier = f"LAB{highest + offset:06d}"
        if identifier in existing:
            raise LaborCatalogError(f"Generated Labor ID overlaps source: {identifier}")
        record["Labor Standard ID"] = identifier
    return prepared


def distinct_values(
    records: Sequence[dict[str, Any]],
    header: str,
) -> list[str]:
    """Return stable, distinct, nonblank source values."""
    return sorted(
        {text(record.get(header)) for record in records if text(record.get(header))},
        key=str.casefold,
    )


def aligned_rows(
    headers: Sequence[str],
    columns: Sequence[Sequence[str]],
) -> list[dict[str, str]]:
    """Align control columns in a single lookup table."""
    count = max((len(values) for values in columns), default=0)
    return [
        {
            header: values[index] if index < len(values) else ""
            for header, values in zip(headers, columns, strict=True)
        }
        for index in range(count)
    ]


def minute_issue(record: dict[str, Any]) -> str:
    """Describe duration relationship defects without inventing values."""
    values: dict[str, int] = {}
    for header in ("Minimum Minutes", "Standard Minutes", "Maximum Minutes"):
        raw = record.get(header)
        if not text(raw):
            continue
        if isinstance(raw, bool):
            return f"{header} is not a whole number"
        try:
            number = int(raw)
        except (TypeError, ValueError):
            return f"{header} is not a whole number"
        if number != raw and text(number) != text(raw):
            return f"{header} is not a whole number"
        if header == "Minimum Minutes" and number < 0:
            return "Minimum Minutes is negative"
        if header != "Minimum Minutes" and number <= 0:
            return f"{header} is not positive"
        values[header] = number
    minimum = values.get("Minimum Minutes")
    standard = values.get("Standard Minutes")
    maximum = values.get("Maximum Minutes")
    if minimum is not None and standard is not None and minimum > standard:
        return "Minimum Minutes exceeds Standard Minutes"
    if standard is not None and maximum is not None and standard > maximum:
        return "Standard Minutes exceeds Maximum Minutes"
    if minimum is not None and maximum is not None and minimum > maximum:
        return "Minimum Minutes exceeds Maximum Minutes"
    return ""


def validate_record_relationships(records: Sequence[dict[str, Any]]) -> None:
    """Reject invalid explicit controlled values and minute relationships."""
    for row_number, record in enumerate(records, start=2):
        issue = minute_issue(record)
        if issue:
            raise LaborCatalogError(f"Labor row {row_number}: {issue}")
        for header in FLAG_HEADERS:
            value = text(record.get(header))
            if value and value not in YES_NO_VALUES:
                raise LaborCatalogError(
                    f"Labor row {row_number} has invalid {header}: {value}"
                )


def review_queue(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    """Create a nonauthoritative queue for every pending labor record."""
    evidence_fields = [
        "Repair Category",
        "Repair Type",
        "Device Family",
        "Standard Minutes",
        "Labor Tier",
        "Skill Level",
        "Difficulty",
    ]
    return [
        {
            "Labor Standard ID": record["Labor Standard ID"],
            "Labor Name": record["Labor Name"],
            "Missing Evidence": ", ".join(
                header for header in evidence_fields if not text(record.get(header))
            ),
            "Relationship Issue": minute_issue(record),
            "Required Action": "Review explicit source evidence",
            "Review Status": "Pending Review",
            "Reviewer": "",
            "Reviewer Notes": "",
        }
        for record in records
    ]


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write one formatted Excel Table."""
    worksheet.append(list(headers))
    materialized = list(rows) or [{header: "" for header in headers}]
    for record in materialized:
        worksheet.append([excel_value(record.get(header, "")) for header in headers])
    end_column = get_column_letter(len(headers))
    table = Table(
        displayName=table_name,
        ref=f"A1:{end_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = table.ref
    for column, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(len(header) + 2, 12),
            42,
        )


def add_defined_names(workbook: Workbook) -> None:
    """Create workbook-scoped lookup names."""
    for name, (sheet_name, column) in DEFINED_NAME_SPECS.items():
        worksheet = workbook[sheet_name]
        last_row = max(
            (
                row
                for row in range(2, worksheet.max_row + 1)
                if text(worksheet.cell(row, column).value)
            ),
            default=2,
        )
        letter = get_column_letter(column)
        workbook.defined_names.add(
            DefinedName(
                name,
                attr_text=(
                    f"'{sheet_name}'!${letter}$2:${letter}${last_row}"
                ),
            )
        )


def add_validations(workbook: Workbook) -> None:
    """Apply only defined-name-based list validation."""
    worksheet = workbook["01 - Labor Standards"]
    for header, defined_name in DEFINED_NAME_BY_HEADER.items():
        column = LABOR_HEADERS.index(header) + 1
        letter = get_column_letter(column)
        validation = DataValidation(
            type="list",
            formula1=f"={defined_name}",
            allow_blank=True,
        )
        validation.showErrorMessage = True
        validation.errorTitle = "Invalid governed value"
        validation.error = f"Select a value from {defined_name}."
        worksheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{worksheet.max_row}")


def build_workbook(
    records: Sequence[dict[str, Any]],
    source_hash: str,
    existing: set[str],
    highest: int,
    generated_at: datetime,
) -> Workbook:
    """Build the complete governed review workbook."""
    workbook = Workbook()
    workbook.active.title = SHEET_NAMES[0]
    for sheet_name in SHEET_NAMES[1:]:
        workbook.create_sheet(sheet_name)

    instructions = [
        {
            "Topic": "Purpose",
            "Instruction": "Governed labor review artifact; not scheduling or payroll.",
        },
        {
            "Topic": "Approval",
            "Instruction": "No generated row or mapping is automatically approved.",
        },
        {
            "Topic": "Import",
            "Instruction": "Canonical import is not authorized by this workbook.",
        },
    ]
    append_table(
        workbook["00 - Instructions"],
        ["Topic", "Instruction"],
        instructions,
        TABLE_NAMES["00 - Instructions"],
    )
    append_table(
        workbook["01 - Labor Standards"],
        LABOR_HEADERS,
        records,
        TABLE_NAMES["01 - Labor Standards"],
    )
    lookup_contract = {
        "02 - Repair Categories": ("Repair Category", "Repair Category"),
        "03 - Repair Types": ("Repair Type", "Repair Type"),
        "04 - Device Families": ("Device Family", "Device Family"),
        "05 - Manufacturers": ("Manufacturer", "Manufacturer"),
        "06 - Labor Tiers": ("Labor Tier", "Labor Tier"),
        "07 - Skill Levels": ("Skill Level", "Skill Level"),
        "08 - Difficulty": ("Difficulty", "Difficulty"),
    }
    for sheet_name, (header, source_header) in lookup_contract.items():
        append_table(
            workbook[sheet_name],
            [header],
            [{header: value} for value in distinct_values(records, source_header)],
            TABLE_NAMES[sheet_name],
        )
    warranty_headers = [
        "Warranty Option",
        "Review Status",
        "Confidence",
        "Yes No",
    ]
    warranty_rows = aligned_rows(
        warranty_headers,
        [
            distinct_values(records, "Warranty Option"),
            REVIEW_STATUSES,
            CONFIDENCE_VALUES,
            YES_NO_VALUES,
        ],
    )
    append_table(
        workbook["09 - Warranty Options"],
        warranty_headers,
        warranty_rows,
        TABLE_NAMES["09 - Warranty Options"],
    )
    queue = review_queue(records)
    append_table(
        workbook["10 - Review Queue"],
        QUEUE_HEADERS,
        queue,
        TABLE_NAMES["10 - Review Queue"],
    )
    validation_rows = [
        {"Check": "Schema columns", "Result": len(LABOR_HEADERS), "Status": "PASS"},
        {"Check": "Generated records", "Result": len(records), "Status": "PASS"},
        {"Check": "Canonical import authorized", "Result": "No", "Status": "PASS"},
        {"Check": "Protected source modified", "Result": "No", "Status": "PASS"},
    ]
    append_table(
        workbook["11 - Validation Summary"],
        ["Check", "Result", "Status"],
        validation_rows,
        TABLE_NAMES["11 - Validation Summary"],
    )
    append_table(
        workbook["12 - Revision History"],
        ["Version", "Revision Date", "Change", "Author"],
        [
            {
                "Version": "1.0.1",
                "Revision Date": generated_at,
                "Change": "Initial governed Master Labor review catalog",
                "Author": "Generator",
            }
        ],
        TABLE_NAMES["12 - Revision History"],
    )
    first_id = text(records[0]["Labor Standard ID"]) if records else ""
    final_id = text(records[-1]["Labor Standard ID"]) if records else ""
    metadata = [
        {"Metadata Field": "Import Batch", "Value": IMPORT_BATCH},
        {"Metadata Field": "Namespace Authority", "Value": "ADR-011"},
        {"Metadata Field": "Schema Columns", "Value": len(LABOR_HEADERS)},
        {"Metadata Field": "Canonical Valid ID Count", "Value": len(existing)},
        {
            "Metadata Field": "Highest Canonical Labor ID",
            "Value": f"LAB{highest:06d}" if highest else "",
        },
        {"Metadata Field": "First Generated ID", "Value": first_id},
        {"Metadata Field": "Final Generated ID", "Value": final_id},
        {"Metadata Field": "Generated Record Count", "Value": len(records)},
        {"Metadata Field": "Canonical Import Authorized", "Value": "No"},
        {
            "Metadata Field": f"Protected Input Path: {CANONICAL_LABOR_PATH.name}",
            "Value": str(CANONICAL_LABOR_PATH),
        },
        {
            "Metadata Field": f"SHA-256: {CANONICAL_LABOR_PATH.name}",
            "Value": source_hash,
        },
        {"Metadata Field": "Generated At UTC", "Value": generated_at},
    ]
    append_table(
        workbook["13 - Import Metadata"],
        ["Metadata Field", "Value"],
        metadata,
        TABLE_NAMES["13 - Import Metadata"],
    )
    add_defined_names(workbook)
    add_validations(workbook)
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read nonblank records from an Excel Table."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    records = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def require_ooxml(path: Path) -> None:
    """Require core OOXML members and prohibit macros/external links."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        members = set(archive.namelist())
    missing = required - members
    if missing:
        raise LaborCatalogError(f"Missing OOXML members: {sorted(missing)}")
    if "xl/vbaProject.bin" in members:
        raise LaborCatalogError("Review workbook unexpectedly contains macros")
    if any(name.startswith("xl/externalLinks/") for name in members):
        raise LaborCatalogError("Review workbook contains external links")


def validate_reopened(
    path: Path,
    expected: Sequence[dict[str, Any]],
    existing: set[str],
    highest: int,
) -> None:
    """Validate persisted structure, names, validations, and identity."""
    require_ooxml(path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise LaborCatalogError("Reopened worksheet contract differs")
        tables: list[str] = []
        for sheet_name in SHEET_NAMES:
            worksheet = workbook[sheet_name]
            table_name = TABLE_NAMES[sheet_name]
            if list(worksheet.tables) != [table_name]:
                raise LaborCatalogError(f"Invalid table contract: {sheet_name}")
            if worksheet.freeze_panes != "A2":
                raise LaborCatalogError(f"Header not frozen: {sheet_name}")
            tables.append(table_name)
        if len(tables) != len(set(tables)):
            raise LaborCatalogError("Table names are not globally unique")
        missing_names = set(DEFINED_NAME_SPECS) - set(workbook.defined_names)
        if missing_names:
            raise LaborCatalogError(f"Missing defined names: {sorted(missing_names)}")
        worksheet = workbook["01 - Labor Standards"]
        actual = table_records(worksheet, TABLE_NAMES[worksheet.title])
        actual_headers = [text(cell.value) for cell in worksheet[1]]
        if actual_headers != LABOR_HEADERS:
            raise LaborCatalogError("Reopened Labor schema differs")
        if len(actual) != len(expected):
            raise LaborCatalogError("Reopened Labor population differs")
        if [row["Legacy Labor ID"] for row in actual] != [
            row["Legacy Labor ID"] for row in expected
        ]:
            raise LaborCatalogError("Reopened Legacy Labor IDs differ from source")
        if [
            text(row["Source Record Number"])
            for row in actual
        ] != [
            text(row["Source Record Number"])
            for row in expected
        ]:
            raise LaborCatalogError("Reopened source-row lineage differs")
        identifiers = [text(row["Labor Standard ID"]) for row in actual]
        expected_ids = [
            f"LAB{highest + offset:06d}"
            for offset in range(1, len(actual) + 1)
        ]
        if identifiers != expected_ids:
            raise LaborCatalogError("Generated Labor IDs are not continuous")
        if len(identifiers) != len(set(identifiers)) or set(identifiers) & existing:
            raise LaborCatalogError("Generated Labor IDs duplicate or overlap")
        if any(
            text(row["Review Status"]) != "Pending Review"
            or text(row["Reviewer"])
            for row in actual
        ):
            raise LaborCatalogError("Generated labor was approved or reviewed")
        validate_record_relationships(actual)
        formulas = {
            validation.formula1
            for validation in worksheet.data_validations.dataValidation
            if validation.type == "list"
        }
        expected_formulas = {
            f"={defined_name}"
            for defined_name in set(DEFINED_NAME_BY_HEADER.values())
        }
        if formulas != expected_formulas:
            raise LaborCatalogError("Data validations do not use defined names")
        if any("!" in formula or "," in formula for formula in formulas):
            raise LaborCatalogError("Direct or hard-coded list validation found")
        if PROHIBITED_HEADERS & set(actual_headers):
            raise LaborCatalogError("Prohibited field found in primary schema")
    finally:
        workbook.close()


def main() -> int:
    """Generate, validate, and atomically publish the review workbook."""
    temporary_created = False
    try:
        require_files([CANONICAL_LABOR_PATH])
        if TEMP_OUTPUT_PATH.exists():
            raise LaborCatalogError(
                f"Stale temporary output exists: {TEMP_OUTPUT_PATH}"
            )
        source_hash = file_hash(CANONICAL_LABOR_PATH)
        source_rows, existing, highest = read_source()
        records = build_records(source_rows, existing, highest)
        validate_record_relationships(records)
        generated_at = datetime.now(UTC).replace(tzinfo=None)
        workbook = build_workbook(
            records,
            source_hash,
            existing,
            highest,
            generated_at,
        )
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary_created = True
        try:
            workbook.save(TEMP_OUTPUT_PATH)
        finally:
            workbook.close()
        validate_reopened(TEMP_OUTPUT_PATH, records, existing, highest)
        if file_hash(CANONICAL_LABOR_PATH) != source_hash:
            raise LaborCatalogError("Protected Labor Standards hash changed")
        os.replace(TEMP_OUTPUT_PATH, OUTPUT_PATH)
        print(f"Generated: {OUTPUT_PATH}")
        print(f"Labor review rows: {len(records)}")
        print("Protected Labor Standards hash: PASS (unchanged)")
        print("Canonical import: NOT AUTHORIZED")
        return 0
    except (LaborCatalogError, OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        print(f"ERROR: {text(exc)}", file=sys.stderr)
        return 1
    finally:
        if temporary_created and TEMP_OUTPUT_PATH.exists():
            try:
                TEMP_OUTPUT_PATH.unlink()
            except OSError as cleanup_error:
                print(
                    f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}",
                    file=sys.stderr,
                )


if __name__ == "__main__":
    raise SystemExit(main())
