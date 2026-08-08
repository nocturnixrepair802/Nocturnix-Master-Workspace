"""Independently validate the Service Type normalization review workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
MASTER_SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
MASTER_LABOR_PATH = WORKING_DIR / "Nocturnix_Master_Labor_Catalog_v1.xlsx"
LEGACY_MAPPING_PATH = WORKING_DIR / "Labor_Mapping_Review_v1.xlsx"
OUTPUT_PATH = (
    WORKING_DIR / "Nocturnix_Service_Type_Normalization_Review_v1.xlsx"
)
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(f".{OUTPUT_PATH.stem}.tmp.xlsx")

CANONICAL_SHEET = "33 Service Types"
SERVICES_SHEET = "01 - Master Services"
LABOR_SHEET = "01 - Labor Standards"
CANONICAL_HEADER_ROW = 4
CANONICAL_MIN_COLUMN = 12
CANONICAL_MAX_COLUMN = 20
EXPECTED_CANONICAL_ROWS = 70
EXPECTED_SERVICE_ROWS = 314
EXPECTED_LABOR_ROWS = 265

SERVICE_TYPE_ID_PATTERN = re.compile(r"^STY\d{6}$")
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
LABOR_ID_PATTERN = re.compile(r"^LAB\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Canonical Service Types",
    "02 - Service Type Aliases",
    "03 - Service Normalization",
    "04 - Labor Normalization",
    "05 - Service Labor Candidates",
    "06 - Unresolved Review",
    "07 - Validation Summary",
    "08 - Revision History",
    "09 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblSTNInstructions",
    "01 - Canonical Service Types": "tblCanonicalServiceTypes",
    "02 - Service Type Aliases": "tblServiceTypeAliases",
    "03 - Service Normalization": "tblServiceNormalization",
    "04 - Labor Normalization": "tblLaborNormalization",
    "05 - Service Labor Candidates": "tblServiceLaborCandidates",
    "06 - Unresolved Review": "tblSTNUnresolvedReview",
    "07 - Validation Summary": "tblSTNValidationSummary",
    "08 - Revision History": "tblSTNRevisionHistory",
    "09 - Import Metadata": "tblSTNImportMetadata",
}
CANONICAL_HEADERS = [
    "Proposed Canonical Service Type ID",
    "Service Category",
    "Service Type",
    "Service Description",
    "Applies To",
    "Estimated Time (Min)",
    "Default Warranty (Days)",
    "Taxable",
    "Active",
    "Internal Notes",
    "Identity Authority",
    "Review Status",
    "Reviewer Notes",
]
ALIAS_HEADERS = [
    "Alias ID",
    "Source System",
    "Source Field",
    "Source Value",
    "Normalized Source Value",
    "Proposed Canonical Service Type ID",
    "Proposed Canonical Service Type",
    "Alias Rule Type",
    "Evidence",
    "Confidence",
    "Review Status",
    "Reviewer",
    "Reviewer Notes",
]
SERVICE_HEADERS = [
    "Service ID",
    "Service Name",
    "Current Repair Type ID",
    "Current Repair Type",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Proposed Canonical Service Type ID",
    "Proposed Canonical Service Type",
    "Mapping Method",
    "Mapping Evidence",
    "Confidence",
    "Review Status",
    "Reviewer Notes",
]
LABOR_HEADERS = [
    "Labor Standard ID",
    "Legacy Labor ID",
    "Labor Name",
    "Current Repair Type",
    "Device Family",
    "Manufacturer",
    "Proposed Canonical Service Type ID",
    "Proposed Canonical Service Type",
    "Mapping Method",
    "Mapping Evidence",
    "Confidence",
    "Review Status",
    "Reviewer Notes",
]
CANDIDATE_HEADERS = [
    "Service ID",
    "Service Name",
    "Canonical Service Type ID",
    "Canonical Service Type",
    "Device Family Code",
    "Manufacturer ID",
    "Suggested Labor Standard ID",
    "Legacy Labor ID",
    "Labor Name",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Candidate Method",
    "Evidence",
    "Confidence",
    "Ambiguity Count",
    "Review Status",
    "Reviewer Notes",
]
UNRESOLVED_HEADERS = [
    "Record Type",
    "Source Record ID",
    "Source Name",
    "Current Type",
    "Candidate Canonical Types",
    "Candidate Labor Standards",
    "Ambiguity Reason",
    "Missing Evidence",
    "Required Action",
    "Review Priority",
    "Review Status",
    "Reviewer Notes",
]
EXPECTED_HEADERS = {
    "01 - Canonical Service Types": CANONICAL_HEADERS,
    "02 - Service Type Aliases": ALIAS_HEADERS,
    "03 - Service Normalization": SERVICE_HEADERS,
    "04 - Labor Normalization": LABOR_HEADERS,
    "05 - Service Labor Candidates": CANDIDATE_HEADERS,
    "06 - Unresolved Review": UNRESOLVED_HEADERS,
}
DEFINED_NAME_SPECS = {
    "DV_CanonicalServiceTypeIDs": ("01 - Canonical Service Types", 1),
    "DV_CanonicalServiceTypes": ("01 - Canonical Service Types", 3),
    "DV_AliasRuleTypes": ("00 - Instructions", 3),
    "DV_MappingMethods": ("00 - Instructions", 4),
    "DV_ConfidenceValues": ("00 - Instructions", 5),
    "DV_ReviewStatuses": ("00 - Instructions", 6),
    "DV_DeviceFamilyCodes": ("00 - Instructions", 7),
    "DV_ManufacturerIDs": ("00 - Instructions", 8),
    "DV_LaborStandardIDs": ("00 - Instructions", 9),
    "DV_YesNoValues": ("00 - Instructions", 10),
}
ALIAS_RULE_TYPES = {
    "Exact Match",
    "Approved Synonym Candidate",
    "Broader-to-Specific Review",
    "Device-Family-Specific",
    "Manufacturer-Specific",
    "No Safe Mapping",
}
MAPPING_METHODS = {
    "Exact Match",
    "Explicit Alias Candidate",
    "Service Name Rule Candidate",
    "Device-Family Constraint",
    "No Safe Mapping",
}
CONFIDENCE_VALUES = {"Unassessed", "Low", "Medium", "High"}
REVIEW_STATUSES = {
    "Pending Review",
    "Pending Evidence Review",
    "Pending Service Review",
    "Pending Labor Review",
    "Unresolved",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
}
PENDING_REVIEW_STATUSES = {
    "Pending Review",
    "Pending Evidence Review",
    "Pending Service Review",
    "Pending Labor Review",
    "Unresolved",
}
PROHIBITED_GENERATED_STATUSES = {
    "approved",
    "confirmed",
    "accepted",
    "ready for import",
}
STATUS_AUDIT_SPECS = {
    "00 - Instructions": (None, "Topic", "Status Vocabulary"),
    "01 - Canonical Service Types": (
        "Review Status",
        "Proposed Canonical Service Type ID",
        "Canonical Service Type",
    ),
    "02 - Service Type Aliases": ("Review Status", "Alias ID", "Service Type Alias"),
    "03 - Service Normalization": (
        "Review Status",
        "Service ID",
        "Service Normalization",
    ),
    "04 - Labor Normalization": (
        "Review Status",
        "Labor Standard ID",
        "Labor Normalization",
    ),
    "05 - Service Labor Candidates": (
        "Review Status",
        "Service ID",
        "Service Labor Candidate",
    ),
    "06 - Unresolved Review": (
        "Review Status",
        "Source Record ID",
        "Unresolved Review",
    ),
    "07 - Validation Summary": ("Result", "Validation", "Validation Summary"),
    "08 - Revision History": ("Status", "Version", "Revision History"),
    "09 - Import Metadata": (None, "Metadata Field", "Import Metadata"),
}
UNIVERSAL_FAMILIES = {"all", "all devices", "any", "universal"}


class NormalizationValidationError(RuntimeError):
    """Raised when the review artifact violates its independent contract."""


def text(value: Any) -> str:
    """Return trimmed display text."""
    return "" if value is None else str(value).strip()


def normalized(value: Any) -> str:
    """Normalize text for exact comparison only."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).casefold()))


def expected_generated_review_status(
    record_type: str,
    *,
    resolved: bool,
    ambiguous: bool = False,
) -> str:
    """Return the required documented pending status for generated records."""
    if ambiguous:
        return "Pending Evidence Review"
    if resolved:
        return "Pending Review"
    if record_type == "Service":
        return "Pending Service Review"
    if record_type == "Labor":
        return "Pending Labor Review"
    return "Unresolved"


def reject_prohibited_generated_statuses(
    records_by_sheet: dict[str, Sequence[dict[str, Any]]],
) -> None:
    """Reject prohibited values only in governed status fields."""
    for sheet_name in SHEET_NAMES:
        status_column, id_column, default_record_type = STATUS_AUDIT_SPECS[sheet_name]
        records = records_by_sheet.get(sheet_name, ())
        if status_column is None:
            continue
        for excel_row, record in enumerate(records, start=2):
            status = text(record.get(status_column))
            if normalized(status) not in PROHIBITED_GENERATED_STATUSES:
                continue
            record_type = text(record.get("Record Type")) or default_record_type
            source_id = text(record.get(id_column)) or "(none)"
            raise NormalizationValidationError(
                "Prohibited generated status: "
                f"worksheet={sheet_name}; table/record type={record_type}; "
                f"Excel row={excel_row}; source record ID={source_id}; "
                f"prohibited status={status}"
            )


def assert_generated_status_contract() -> None:
    """Exercise pending assignment and exact governed-field prohibition."""
    assert expected_generated_review_status(
        "Service", resolved=True
    ) == "Pending Review"
    assert expected_generated_review_status("Alias", resolved=True) == "Pending Review"
    assert expected_generated_review_status(
        "Relationship", resolved=False, ambiguous=True
    ) in PENDING_REVIEW_STATUSES
    try:
        reject_prohibited_generated_statuses(
            {"03 - Service Normalization": [{"Service ID": "SVC000001", "Review Status": "Approved"}]}
        )
    except NormalizationValidationError as exc:
        assert "SVC000001" in text(exc)
        assert "Approved" in text(exc)
    else:
        raise AssertionError("Literal Approved generated Review Status was accepted")
    reject_prohibited_generated_statuses(
        {
            "09 - Import Metadata": [
                {"Metadata Field": "Approval Status", "Value": "Not Approved"}
            ]
        }
    )


def file_hash(path: Path) -> str:
    """Return a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require all validator inputs."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise NormalizationValidationError(
            f"Missing required file: {', '.join(missing)}"
        )


def locate_header(
    worksheet: Worksheet,
    required: set[str],
) -> tuple[int, list[str]]:
    """Locate a source header within the first 50 rows."""
    for row_number, headers in bounded_header_rows(worksheet):
        if required <= set(headers):
            return row_number, headers
    raise NormalizationValidationError(
        f"{worksheet.title} lacks required headers: {sorted(required)}"
    )


def bounded_header_rows(
    worksheet: Worksheet,
    *,
    scan_limit: int = 50,
) -> Iterable[tuple[int, list[str]]]:
    """Yield populated candidate header rows strictly within sheet bounds."""
    first_row = max(1, worksheet.min_row)
    last_row = min(worksheet.max_row, scan_limit)
    if first_row > last_row:
        return
    rows = worksheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=max(1, worksheet.min_column),
        max_col=max(1, worksheet.max_column),
    )
    for row_number, row in enumerate(rows, start=first_row):
        headers = [text(cell.value) for cell in row]
        if any(headers):
            yield row_number, headers


def assert_bounded_header_scan_contract() -> None:
    """Exercise short sheets, blank candidate rows, and absent headers."""
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Only one populated row"])
    try:
        assert list(bounded_header_rows(worksheet)) == [
            (1, ["Only one populated row"])
        ]
        try:
            locate_header(worksheet, {"Service ID"})
        except NormalizationValidationError as exc:
            assert "lacks required headers" in text(exc)
        else:
            raise AssertionError("Missing header was accepted")
        worksheet.append([])
        worksheet.append(["Service ID", "Service Name"])
        assert locate_header(
            worksheet,
            {"Service ID", "Service Name"},
        ) == (3, ["Service ID", "Service Name"])
    finally:
        workbook.close()


def sequential_records(
    worksheet: Worksheet,
    required: set[str],
    *,
    header_row: int | None = None,
    start_row: int | None = None,
    min_column: int | None = None,
    max_column: int | None = None,
) -> list[dict[str, Any]]:
    """Return records from one bounded, sequential worksheet pass."""
    first_column = min_column or max(1, worksheet.min_column)
    last_column = max_column or max(1, worksheet.max_column)
    first_row = header_row or max(1, worksheet.min_row)
    rows = worksheet.iter_rows(
        min_row=first_row,
        max_row=worksheet.max_row,
        min_col=first_column,
        max_col=last_column,
        values_only=True,
    )
    records: list[dict[str, Any]] = []
    headers: list[str] | None = None
    header_indexes: list[tuple[int, str]] = []
    data_start = start_row
    for row_number, values_tuple in enumerate(rows, start=first_row):
        values = list(values_tuple)
        if headers is None:
            candidate_headers = [text(value) for value in values]
            is_header = (
                row_number == header_row
                if header_row is not None
                else row_number <= 50 and required <= set(candidate_headers)
            )
            if not is_header:
                continue
            if not required <= set(candidate_headers):
                raise NormalizationValidationError(
                    f"{worksheet.title} lacks required headers: {sorted(required)}"
                )
            headers = candidate_headers
            header_indexes = list(enumerate(headers))
            data_start = data_start or row_number + 1
            continue
        if row_number < data_start:
            continue
        if not any(text(value) for value in values):
            continue
        records.append(
            {header: values[index] for index, header in header_indexes}
        )
    if headers is None:
        raise NormalizationValidationError(
            f"{worksheet.title} lacks required headers: {sorted(required)}"
        )
    return records


def assert_sequential_read_contract() -> None:
    """Exercise ranges, blank rows, value fidelity, counts, and order."""
    from openpyxl import Workbook

    marker = datetime(2025, 1, 2, 3, 4, 5)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["preamble"])
    worksheet.append([])
    worksheet.append([None, None, "ID", "Value", "When"])
    worksheet.append([None, None, "001", 0, marker])
    worksheet.append([None, None, None, None, None])
    worksheet.append([None, None, "002", "", "text"])
    try:
        records = sequential_records(
            worksheet,
            {"ID", "Value", "When"},
            header_row=3,
            start_row=4,
            min_column=3,
            max_column=5,
        )
        assert records == [
            {"ID": "001", "Value": 0, "When": marker},
            {"ID": "002", "Value": "", "When": "text"},
        ]
        service_sheet = workbook.create_sheet("Services")
        service_sheet.append(["Service ID"])
        for index in range(1, EXPECTED_SERVICE_ROWS + 1):
            service_sheet.append([f"SVC{index:06d}"])
        labor_sheet = workbook.create_sheet("Labor")
        labor_sheet.append(["Labor Standard ID"])
        for index in range(1, EXPECTED_LABOR_ROWS + 1):
            labor_sheet.append([f"LAB{index:06d}"])
        service_records = sequential_records(service_sheet, {"Service ID"})
        labor_records = sequential_records(labor_sheet, {"Labor Standard ID"})
        assert len(service_records) == 314
        assert len(labor_records) == 265
        assert service_records[0]["Service ID"] == "SVC000001"
        assert service_records[-1]["Service ID"] == "SVC000314"
        assert labor_records[0]["Labor Standard ID"] == "LAB000001"
        assert labor_records[-1]["Labor Standard ID"] == "LAB000265"
    finally:
        workbook.close()


def read_source_records(
    path: Path,
    sheet_name: str,
    required: set[str],
) -> list[dict[str, Any]]:
    """Read protected source records without saving a workbook."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise NormalizationValidationError(f"{path.name} lacks {sheet_name}")
        return sequential_records(workbook[sheet_name], required)
    finally:
        workbook.close()


def canonical_source_records() -> list[dict[str, Any]]:
    """Read the detailed L:T canonical taxonomy directly."""
    headers = [
        "Service Category",
        "Service Type",
        "Service Description",
        "Applies To",
        "Estimated Time (Min)",
        "Default Warranty (Days)",
        "Taxable",
        "Active",
        "Internal Notes",
    ]
    workbook = load_workbook(
        CANONICAL_PATH,
        read_only=True,
        data_only=False,
        keep_vba=True,
    )
    try:
        if CANONICAL_SHEET not in workbook.sheetnames:
            raise NormalizationValidationError(
                f"Canonical workbook lacks {CANONICAL_SHEET}"
            )
        worksheet = workbook[CANONICAL_SHEET]
        if not worksheet.min_row <= CANONICAL_HEADER_ROW <= worksheet.max_row:
            raise NormalizationValidationError(
                f"{CANONICAL_SHEET} lacks canonical header row "
                f"{CANONICAL_HEADER_ROW}"
            )
        if (
            worksheet.min_column > CANONICAL_MIN_COLUMN
            or worksheet.max_column < CANONICAL_MAX_COLUMN
        ):
            raise NormalizationValidationError(
                f"{CANONICAL_SHEET} lacks canonical columns L:T"
            )
        rows = worksheet.iter_rows(
            min_row=CANONICAL_HEADER_ROW,
            max_row=worksheet.max_row,
            min_col=CANONICAL_MIN_COLUMN,
            max_col=CANONICAL_MAX_COLUMN,
            values_only=True,
        )
        actual_headers = [text(value) for value in next(rows)]
        if actual_headers != headers:
            raise NormalizationValidationError(
                f"Detailed canonical headers differ: {actual_headers}"
            )
        records = []
        for values_tuple in rows:
            values = list(values_tuple)
            if text(values[1]):
                records.append(dict(zip(headers, values, strict=True)))
        return records
    finally:
        workbook.close()


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Return an Excel Table header list."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, _ = range_boundaries(table.ref)
    return [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Return nonblank Excel Table records."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    records = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def require_ooxml(path: Path) -> None:
    """Require a safe, complete, macro-free OOXML package."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        members = set(archive.namelist())
        corrupt_member = archive.testzip()
    if required - members or corrupt_member:
        raise NormalizationValidationError("Review OOXML package is incomplete")
    if "xl/vbaProject.bin" in members:
        raise NormalizationValidationError("Review workbook contains macros")
    if any(member.startswith("xl/externalLinks/") for member in members):
        raise NormalizationValidationError("Review workbook has external links")


def metadata_map(workbook: Any) -> dict[str, str]:
    """Return unique import metadata fields."""
    records = table_records(
        workbook["09 - Import Metadata"],
        TABLE_NAMES["09 - Import Metadata"],
    )
    keys = [text(record["Metadata Field"]) for record in records]
    if len(keys) != len(set(keys)):
        raise NormalizationValidationError("Import metadata keys are duplicated")
    return {key: text(record["Value"]) for key, record in zip(keys, records)}


def validate_structure(workbook: Any) -> None:
    """Validate exact sheet order, unique names, schemas, and table contract."""
    if workbook.sheetnames != SHEET_NAMES:
        raise NormalizationValidationError(
            f"Worksheet contract differs: {workbook.sheetnames}"
        )
    if len(set(workbook.sheetnames)) != 10:
        raise NormalizationValidationError("Worksheet names are not unique")
    if any(len(name) > 31 for name in workbook.sheetnames):
        raise NormalizationValidationError("Worksheet name exceeds 31 characters")
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        if list(worksheet.tables) != [TABLE_NAMES[sheet_name]]:
            raise NormalizationValidationError(
                f"Table contract differs: {sheet_name}"
            )
        if worksheet.freeze_panes != "A2":
            raise NormalizationValidationError(
                f"Frozen header is missing: {sheet_name}"
            )
    for sheet_name, headers in EXPECTED_HEADERS.items():
        if table_headers(workbook[sheet_name], TABLE_NAMES[sheet_name]) != headers:
            raise NormalizationValidationError(f"Schema differs: {sheet_name}")


def validate_names_and_validations(workbook: Any) -> None:
    """Require defined-name-only list validation formulas."""
    if set(workbook.defined_names) != set(DEFINED_NAME_SPECS):
        raise NormalizationValidationError("Defined-name contract differs")
    for name, (sheet_name, column) in DEFINED_NAME_SPECS.items():
        defined = workbook.defined_names[name]
        destinations = list(defined.destinations)
        if len(destinations) != 1:
            raise NormalizationValidationError(
                f"Defined name has invalid destinations: {name}"
            )
        actual_sheet, coordinate = destinations[0]
        if actual_sheet != sheet_name:
            raise NormalizationValidationError(
                f"Defined name points to wrong sheet: {name}"
            )
        min_column, min_row, max_column, max_row = range_boundaries(coordinate)
        if min_column != column or max_column != column or min_row != 2:
            raise NormalizationValidationError(
                f"Defined name has invalid range: {name}"
            )
        if max_row < 2:
            raise NormalizationValidationError(
                f"Defined name source is empty: {name}"
            )
    for worksheet in workbook.worksheets:
        for validation in worksheet.data_validations.dataValidation:
            formula = text(validation.formula1)
            if validation.type != "list" or not formula.startswith("="):
                raise NormalizationValidationError(
                    f"Invalid validation in {worksheet.title}"
                )
            if formula[1:] not in DEFINED_NAME_SPECS:
                raise NormalizationValidationError(
                    f"Direct cross-sheet validation in {worksheet.title}"
                )


def validate_canonical(
    workbook: Any,
    source: Sequence[dict[str, Any]],
) -> tuple[set[str], dict[str, str]]:
    """Validate exact canonical snapshot preservation and ID integrity."""
    records = table_records(
        workbook["01 - Canonical Service Types"],
        TABLE_NAMES["01 - Canonical Service Types"],
    )
    if len(records) != len(source):
        raise NormalizationValidationError(
            "Canonical review count differs from runtime source"
        )
    if len(source) != EXPECTED_CANONICAL_ROWS:
        metadata = metadata_map(workbook)
        if int(metadata["Canonical Detailed Row Count"]) != len(source):
            raise NormalizationValidationError(
                "Runtime canonical change is not recorded in metadata"
            )
    ids = [text(record["Proposed Canonical Service Type ID"]) for record in records]
    if any(not SERVICE_TYPE_ID_PATTERN.fullmatch(value) for value in ids):
        raise NormalizationValidationError("Invalid proposed Service Type ID")
    if len(ids) != len(set(ids)):
        raise NormalizationValidationError("Duplicate proposed Service Type ID")
    by_id: dict[str, str] = {}
    source_fields = CANONICAL_HEADERS[1:10]
    for actual, expected in zip(records, source, strict=True):
        for field in source_fields:
            if text(actual[field]) != text(expected[field]):
                raise NormalizationValidationError(
                    f"Canonical source value changed: {field}"
                )
        by_id[text(actual["Proposed Canonical Service Type ID"])] = text(
            actual["Service Type"]
        )
        if text(actual["Review Status"]) != "Pending Review":
            raise NormalizationValidationError(
                "Canonical snapshot is not Pending Review"
            )
    return set(ids), by_id


def validate_source_preservation(
    workbook: Any,
    services: Sequence[dict[str, Any]],
    labor: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Validate one-to-one coverage and exact source-field preservation."""
    service_rows = table_records(
        workbook["03 - Service Normalization"],
        TABLE_NAMES["03 - Service Normalization"],
    )
    labor_rows = table_records(
        workbook["04 - Labor Normalization"],
        TABLE_NAMES["04 - Labor Normalization"],
    )
    if len(service_rows) != len(services):
        raise NormalizationValidationError(
            "Service normalization does not preserve runtime row count"
        )
    if len(labor_rows) != len(labor):
        raise NormalizationValidationError(
            "Labor normalization does not preserve runtime row count"
        )
    if len(services) != EXPECTED_SERVICE_ROWS:
        metadata = metadata_map(workbook)
        if int(metadata["Master Service Row Count"]) != len(services):
            raise NormalizationValidationError(
                "Runtime Service count change is not recorded"
            )
    if len(labor) != EXPECTED_LABOR_ROWS:
        metadata = metadata_map(workbook)
        if int(metadata["Master Labor Row Count"]) != len(labor):
            raise NormalizationValidationError(
                "Runtime Labor count change is not recorded"
            )
    service_map = (
        ("Service ID", "Service ID"),
        ("Service Name", "Service Name"),
        ("Current Repair Type ID", "Repair Type ID"),
        ("Current Repair Type", "Repair Type"),
        ("Manufacturer ID", "Manufacturer ID"),
        ("Manufacturer Name", "Manufacturer Name"),
        ("Device Family Code", "Device Family Code"),
        ("Device Family Name", "Device Family Name"),
    )
    labor_map = (
        ("Labor Standard ID", "Labor Standard ID"),
        ("Legacy Labor ID", "Legacy Labor ID"),
        ("Labor Name", "Labor Name"),
        ("Current Repair Type", "Repair Type"),
        ("Device Family", "Device Family"),
        ("Manufacturer", "Manufacturer"),
    )
    for actual, source in zip(service_rows, services, strict=True):
        for actual_field, source_field in service_map:
            if text(actual[actual_field]) != text(source.get(source_field)):
                raise NormalizationValidationError(
                    f"Service source value changed: {actual_field}"
                )
    for actual, source in zip(labor_rows, labor, strict=True):
        for actual_field, source_field in labor_map:
            if text(actual[actual_field]) != text(source.get(source_field)):
                raise NormalizationValidationError(
                    f"Labor source value changed: {actual_field}"
                )
    return service_rows, labor_rows


def validate_normalization_references(
    service_rows: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
    canonical_ids: set[str],
    canonical_by_id: dict[str, str],
) -> None:
    """Validate proposed pairs, methods, confidence, and non-approval."""
    for record in [*service_rows, *labor_rows]:
        proposed_id = text(record["Proposed Canonical Service Type ID"])
        proposed_type = text(record["Proposed Canonical Service Type"])
        if bool(proposed_id) != bool(proposed_type):
            raise NormalizationValidationError(
                "Partial proposed canonical reference found"
            )
        if proposed_id:
            if proposed_id not in canonical_ids:
                raise NormalizationValidationError(
                    f"Unknown canonical Service Type ID: {proposed_id}"
                )
            if canonical_by_id[proposed_id] != proposed_type:
                raise NormalizationValidationError(
                    f"Canonical ID/type mismatch: {proposed_id}"
                )
        if text(record["Mapping Method"]) not in MAPPING_METHODS:
            raise NormalizationValidationError("Invalid Mapping Method")
        if text(record["Confidence"]) not in CONFIDENCE_VALUES:
            raise NormalizationValidationError("Invalid Confidence")
        expected_status = expected_generated_review_status(
            "Service" if "Service ID" in record else "Labor",
            resolved=bool(proposed_id),
        )
        if text(record["Review Status"]) != expected_status:
            raise NormalizationValidationError(
                f"Generated normalization status is not {expected_status}"
            )


def validate_aliases(
    workbook: Any,
    canonical_ids: set[str],
    canonical_by_id: dict[str, str],
) -> None:
    """Validate alias identity, canonical references, and review state."""
    records = table_records(
        workbook["02 - Service Type Aliases"],
        TABLE_NAMES["02 - Service Type Aliases"],
    )
    alias_ids = [text(record["Alias ID"]) for record in records]
    if len(alias_ids) != len(set(alias_ids)) or any(
        not re.fullmatch(r"STA\d{6}", value) for value in alias_ids
    ):
        raise NormalizationValidationError("Alias IDs are invalid or duplicated")
    for record in records:
        proposed_id = text(record["Proposed Canonical Service Type ID"])
        proposed_type = text(record["Proposed Canonical Service Type"])
        if bool(proposed_id) != bool(proposed_type):
            raise NormalizationValidationError(
                "Alias has a partial canonical reference"
            )
        if proposed_id:
            if proposed_id not in canonical_ids:
                raise NormalizationValidationError(
                    f"Alias references unknown canonical ID: {proposed_id}"
                )
            if canonical_by_id[proposed_id] != proposed_type:
                raise NormalizationValidationError(
                    "Alias canonical ID/type pair differs"
                )
        if text(record["Alias Rule Type"]) not in ALIAS_RULE_TYPES:
            raise NormalizationValidationError("Invalid Alias Rule Type")
        if text(record["Review Status"]) != "Pending Review":
            raise NormalizationValidationError("Generated alias is not pending")


def family_matches(service: dict[str, Any], labor: dict[str, Any]) -> bool:
    """Return whether the family constraint is satisfied."""
    labor_family = normalized(labor["Device Family"])
    if labor_family in UNIVERSAL_FAMILIES:
        return True
    return labor_family in {
        normalized(service["Device Family Code"]),
        normalized(service["Device Family Name"]),
    } - {""}


def manufacturer_matches(
    service: dict[str, Any],
    labor: dict[str, Any],
) -> bool:
    """Return whether a manufacturer-specific Labor Standard is eligible."""
    labor_manufacturer = normalized(labor["Manufacturer"])
    if not labor_manufacturer:
        return True
    return labor_manufacturer in {
        normalized(service["Manufacturer ID"]),
        normalized(service["Manufacturer Name"]),
    } - {""}


def validate_candidates(
    workbook: Any,
    service_rows: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
) -> None:
    """Validate Service Type, family, manufacturer, and ambiguity constraints."""
    candidates = table_records(
        workbook["05 - Service Labor Candidates"],
        TABLE_NAMES["05 - Service Labor Candidates"],
    )
    services = {text(row["Service ID"]): row for row in service_rows}
    labor = {text(row["Labor Standard ID"]): row for row in labor_rows}
    expected_by_service: dict[str, list[dict[str, Any]]] = {}
    for service_id, service in services.items():
        canonical_id = text(service["Proposed Canonical Service Type ID"])
        if not canonical_id:
            continue
        eligible = [
            labor_record
            for labor_record in labor_rows
            if text(labor_record["Proposed Canonical Service Type ID"])
            == canonical_id
            and family_matches(service, labor_record)
            and manufacturer_matches(service, labor_record)
        ]
        if eligible:
            expected_by_service[service_id] = eligible
    actual_service_ids = [text(record["Service ID"]) for record in candidates]
    if len(actual_service_ids) != len(set(actual_service_ids)):
        raise NormalizationValidationError(
            "Service Labor Candidates contains duplicate Services"
        )
    if set(actual_service_ids) != set(expected_by_service):
        raise NormalizationValidationError(
            "Service Labor Candidate coverage differs from governed constraints"
        )
    for record in candidates:
        service_id = text(record["Service ID"])
        if service_id not in services:
            raise NormalizationValidationError(
                f"Candidate references unknown Service: {service_id}"
            )
        service = services[service_id]
        if text(record["Canonical Service Type ID"]) != text(
            service["Proposed Canonical Service Type ID"]
        ):
            raise NormalizationValidationError(
                "Candidate Service Type differs from Service normalization"
            )
        ambiguity_count = int(record["Ambiguity Count"])
        expected_eligible = expected_by_service[service_id]
        if ambiguity_count != len(expected_eligible):
            raise NormalizationValidationError(
                "Labor candidate Ambiguity Count was not independently reproduced"
            )
        suggested = text(record["Suggested Labor Standard ID"])
        if ambiguity_count > 1 and suggested:
            raise NormalizationValidationError(
                "Ambiguous relationship has an assigned Labor Standard ID"
            )
        if ambiguity_count == 1 and not suggested:
            raise NormalizationValidationError(
                "Unique relationship lacks its Labor Standard ID"
            )
        if suggested:
            if suggested not in labor:
                raise NormalizationValidationError(
                    f"Candidate references unknown Labor Standard: {suggested}"
                )
            labor_record = labor[suggested]
            if text(labor_record["Proposed Canonical Service Type ID"]) != text(
                record["Canonical Service Type ID"]
            ):
                raise NormalizationValidationError(
                    "Candidate canonical Service Type constraint failed"
                )
            if not family_matches(service, labor_record):
                raise NormalizationValidationError(
                    "Candidate Device Family constraint failed"
                )
            if not manufacturer_matches(service, labor_record):
                raise NormalizationValidationError(
                    "Candidate Manufacturer constraint failed"
                )
            if suggested != text(expected_eligible[0]["Labor Standard ID"]):
                raise NormalizationValidationError(
                    "Suggested Labor Standard is not the unique eligible row"
                )
        expected_status = expected_generated_review_status(
            "Relationship",
            resolved=bool(suggested),
            ambiguous=int(record["Ambiguity Count"]) > 1,
        )
        if text(record["Review Status"]) != expected_status:
            raise NormalizationValidationError(
                f"Generated labor candidate status is not {expected_status}"
            )


def validate_statuses_and_timezones(workbook: Any) -> None:
    """Reject generated approval and timezone-aware persisted values."""
    records_by_sheet = {
        worksheet.title: table_records(worksheet, TABLE_NAMES[worksheet.title])
        for worksheet in workbook.worksheets
    }
    reject_prohibited_generated_statuses(records_by_sheet)
    for worksheet in workbook.worksheets:
        records = records_by_sheet[worksheet.title]
        for record in records:
            status = text(record.get("Review Status"))
            if status and status not in REVIEW_STATUSES:
                raise NormalizationValidationError(
                    f"Invalid Review Status in {worksheet.title}: {status}"
                )
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime) and cell.value.tzinfo:
                    raise NormalizationValidationError(
                        f"Timezone-aware value in {worksheet.title}"
                    )


def validate_hashes(workbook: Any) -> None:
    """Compare protected inputs with recorded generation-time state."""
    metadata = metadata_map(workbook)
    for path in (
        CANONICAL_PATH,
        MASTER_SERVICES_PATH,
        MASTER_LABOR_PATH,
        LEGACY_MAPPING_PATH,
    ):
        key = f"SHA-256: {path.name}"
        if key not in metadata:
            raise NormalizationValidationError(
                f"Protected hash metadata is missing: {path.name}"
            )
        actual = file_hash(path) if path.is_file() else "ABSENT"
        if metadata[key] != actual:
            raise NormalizationValidationError(
                f"Protected input hash differs: {path}"
            )


def main() -> int:
    """Validate the review artifact without invoking its generator."""
    try:
        assert_bounded_header_scan_contract()
        assert_sequential_read_contract()
        assert_generated_status_contract()
        require_files(
            [
                CANONICAL_PATH,
                MASTER_SERVICES_PATH,
                MASTER_LABOR_PATH,
                OUTPUT_PATH,
            ]
        )
        if TEMP_OUTPUT_PATH.exists():
            raise NormalizationValidationError(
                f"Stale transactional output exists: {TEMP_OUTPUT_PATH}"
            )
        require_ooxml(OUTPUT_PATH)
        print("Reading canonical Service Types...")
        canonical_source = canonical_source_records()
        print("Reading canonical Service Types complete.")
        print("Reading Master Services...")
        service_source = read_source_records(
            MASTER_SERVICES_PATH,
            SERVICES_SHEET,
            {
                "Service ID",
                "Service Name",
                "Repair Type ID",
                "Repair Type",
                "Manufacturer ID",
                "Manufacturer Name",
                "Device Family Code",
                "Device Family Name",
            },
        )
        print("Reading Master Services complete.")
        print("Reading Master Labor...")
        labor_source = read_source_records(
            MASTER_LABOR_PATH,
            LABOR_SHEET,
            {
                "Labor Standard ID",
                "Legacy Labor ID",
                "Labor Name",
                "Repair Type",
                "Device Family",
                "Manufacturer",
            },
        )
        print("Reading Master Labor complete.")
        print("Reopening and validating...")
        workbook = load_workbook(OUTPUT_PATH, data_only=False, read_only=False)
        try:
            validate_structure(workbook)
            validate_names_and_validations(workbook)
            canonical_ids, canonical_by_id = validate_canonical(
                workbook,
                canonical_source,
            )
            service_rows, labor_rows = validate_source_preservation(
                workbook,
                service_source,
                labor_source,
            )
            validate_normalization_references(
                service_rows,
                labor_rows,
                canonical_ids,
                canonical_by_id,
            )
            validate_aliases(workbook, canonical_ids, canonical_by_id)
            validate_candidates(workbook, service_rows, labor_rows)
            validate_statuses_and_timezones(workbook)
            validate_hashes(workbook)
        finally:
            workbook.close()
        print("Reopening and validating complete.")
        print(f"Validated: {OUTPUT_PATH}")
        print(f"Canonical Service Types: {len(canonical_source)}")
        print(f"Service normalization rows: {len(service_source)}")
        print(f"Labor normalization rows: {len(labor_source)}")
        print("Generated Approved statuses: 0")
        return 0
    except (
        NormalizationValidationError,
        OSError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
    ) as exc:
        print(f"ERROR: {text(exc)}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
