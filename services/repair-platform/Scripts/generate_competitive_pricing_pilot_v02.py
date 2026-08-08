"""Create the governed Nocturnix Competitive Pricing Pilot v0.2 Draft.

The locked v0.1 checkpoint is hash-verified before it is opened and is never
saved. The transformed workbook is validated as a temporary sibling and then
atomically published.
"""

from __future__ import annotations

import argparse
import hashlib
from copy import copy
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

V01_SHA256 = "5D52F9B2438C9C7B9DD3A7E803C5607F968336AFD65A9BE91F214700432F08FB"
CANONICAL_SHA256 = "DE0F0957F687DF4866A2D06C4DF85A542FF58B61897481741EB1E6A04D825FBA"
GENERATOR_VERSION = "v0.2-draft-generator-1.0"
QA_PROFILE_VERSION = "v0.2-draft-qa-1.0"
SHEETS = [
    "00 - Instructions",
    "01 - Pilot Pricing Records",
    "02 - Competitor Observations",
    "03 - Supplier Cost Evidence",
    "04 - Labor References",
    "05 - Canonical References",
    "06 - Pricing Policy",
    "07 - Calculation Results",
    "08 - Review Queue",
    "09 - Validation Summary",
    "10 - Revision History",
    "11 - Import Metadata",
]
AMBER = "F4B183"
PALE_AMBER = "FFF2CC"
RED = "C00000"
PALE_RED = "F4CCCC"
NAVY = "17233B"
WHITE = "FFFFFF"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def headers(ws: Any) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def table_rows(ws: Any) -> list[dict[str, Any]]:
    names = [cell.value for cell in ws[1]]
    return [
        dict(zip(names, values, strict=False))
        for values in ws.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in values)
    ]


def set_table_ref(ws: Any, table_name: str) -> None:
    ws.tables[table_name].ref = (
        f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
    )
    ws.auto_filter.ref = ws.dimensions


def copy_header_style(ws: Any, source_column: int, target_column: int) -> None:
    source = ws.cell(1, source_column)
    target = ws.cell(1, target_column)
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format


def append_columns(ws: Any, names: list[str]) -> dict[str, int]:
    first = ws.max_column + 1
    for offset, name in enumerate(names):
        column = first + offset
        ws.cell(1, column, name)
        copy_header_style(ws, first - 1, column)
        ws.column_dimensions[get_column_letter(column)].width = min(
            max(len(name) + 2, 16), 34
        )
    return headers(ws)


def add_list_validation(
    ws: Any, column: int, values: list[str], start_row: int, end_row: int
) -> None:
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
    )
    validation.error = "Select an approved controlled value."
    validation.errorTitle = "Invalid value"
    validation.showErrorMessage = True
    ws.add_data_validation(validation)
    validation.add(
        f"{get_column_letter(column)}{start_row}:"
        f"{get_column_letter(column)}{end_row}"
    )


def banner_workbook(workbook: Any) -> None:
    text = "v0.2 DRAFT | NOT PRICING-READY | CALCULATIONS DISABLED | PRODUCTION PROHIBITED"
    for ws in workbook.worksheets:
        ws.oddHeader.center.text = text
        ws.oddHeader.center.size = 10
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.color = RED
        ws.sheet_properties.tabColor = AMBER
        ws["A1"].comment = Comment(text, "Nocturnix Governance")


def update_instructions(workbook: Any) -> None:
    ws = workbook["00 - Instructions"]
    additions = [
        ["Lifecycle banner", "v0.2 Draft working revision."],
        [
            "Readiness banner",
            "Structurally validated; not pricing-ready while any governance gate is blocked.",
        ],
        ["Calculation banner", "Calculation Status is Disabled for all 25 records."],
        ["Production banner", "Production activation is Prohibited."],
        [
            "Final approval banner",
            "Final Approved Price is blank, protected, and not authorized for this revision.",
        ],
    ]
    for row in additions:
        ws.append(row)
    for row in range(ws.max_row - len(additions) + 1, ws.max_row + 1):
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=PALE_AMBER)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=PALE_AMBER)
        ws.cell(row, 1).font = Font(color=RED, bold=True)
        ws.cell(row, 2).font = Font(color=RED, bold=True)
    set_table_ref(ws, "tblInstructions")


def update_identity_controls(workbook: Any) -> None:
    ws = workbook["01 - Pilot Pricing Records"]
    mapping = append_columns(
        ws,
        [
            "Identity Resolution Status",
            "Identity Evidence Reference",
            "Identity Reviewer",
            "Identity Verified At",
            "Identity Dataset Version",
        ],
    )
    for row in range(2, 27):
        ws.cell(row, mapping["Identity Resolution Status"], "Pending Governance")
        ws.cell(row, mapping["Identity Verified At"]).number_format = (
            "yyyy-mm-dd hh:mm"
        )
    add_list_validation(
        ws,
        mapping["Identity Resolution Status"],
        ["Pending Governance", "In Review", "Governed", "Rejected"],
        2,
        26,
    )
    set_table_ref(ws, "tblPilotPricingRecords")


def update_competitor_controls(workbook: Any) -> None:
    ws = workbook["02 - Competitor Observations"]
    mapping = append_columns(
        ws,
        [
            "Record Role",
            "Evidence Package Status",
            "Competitor Dataset Version",
            "Verification Date",
        ],
    )
    ws.cell(2, mapping["Record Role"], "Template - Do Not Count")
    ws.cell(2, mapping["Evidence Package Status"], "Template Only")
    ws.cell(2, mapping["Verification Date"]).number_format = "yyyy-mm-dd"
    add_list_validation(
        ws,
        mapping["Record Role"],
        ["Template - Do Not Count", "Evidence"],
        2,
        1001,
    )
    set_table_ref(ws, "tblCompetitorObservations")


def update_supplier_controls(workbook: Any) -> None:
    ws = workbook["03 - Supplier Cost Evidence"]
    pilot = workbook["01 - Pilot Pricing Records"]
    pilot_map = headers(pilot)
    pilot_pairs = [
        (
            pilot.cell(row, pilot_map["Pricing Record ID"]).value,
            pilot.cell(row, pilot_map["Service ID"]).value,
        )
        for row in range(2, 27)
    ]
    existing = table_rows(ws)
    existing_ids = {row["Pricing Record ID"] for row in existing}
    missing = [pair for pair in pilot_pairs if pair[0] not in existing_ids]
    if len(existing) != 22 or len(missing) != 3:
        raise ValueError(
            f"Supplier reference reconciliation failed: existing={len(existing)}, "
            f"missing={len(missing)}"
        )
    for pricing_id, service_id in missing:
        row = [None] * ws.max_column
        row[1] = pricing_id
        row[2] = service_id
        row[13] = "No Evidence Captured"
        ws.append(row)
    mapping = append_columns(
        ws,
        [
            "Evidence Record Status",
            "Evidence Package Status",
            "Supplier Dataset Version",
            "Verified At",
            "Evidence Completeness",
            "Missing Evidence Reason",
        ],
    )
    for row in range(2, 27):
        pricing_id = ws.cell(row, 2).value
        if pricing_id in existing_ids:
            ws.cell(row, mapping["Evidence Record Status"], "Legacy Reference Preserved")
            ws.cell(row, mapping["Evidence Package Status"], "Incomplete")
            reason = (
                "Verified current part cost, currency, source date, verification, "
                "and reviewer are required."
            )
        else:
            ws.cell(row, mapping["Evidence Record Status"], "No Evidence Captured")
            ws.cell(row, mapping["Evidence Package Status"], "Missing")
            reason = "No related legacy part reference; verified evidence is required."
        ws.cell(
            row,
            mapping["Evidence Completeness"],
            (
                f'=IF(T{row}="No Evidence Captured","Missing",'
                f'IF(AND(I{row}<>"",K{row}<>"",L{row}<>"",M{row}<>"",'
                f'N{row}="Verified",O{row}<>""),"Complete","Incomplete"))'
            ),
        )
        ws.cell(row, mapping["Missing Evidence Reason"], reason)
        ws.cell(row, mapping["Verified At"]).number_format = "yyyy-mm-dd hh:mm"
    for validation in ws.data_validations.dataValidation:
        if validation.type == "list" and "N2:N1001" in str(validation.sqref):
            validation.formula1 = (
                '"No Evidence Captured,Pending Verification,Verified,Rejected,'
                'Stale,Superseded"'
            )
    add_list_validation(
        ws,
        mapping["Evidence Record Status"],
        ["Legacy Reference Preserved", "No Evidence Captured", "New Evidence"],
        2,
        1001,
    )
    add_list_validation(
        ws,
        mapping["Evidence Package Status"],
        ["Missing", "Incomplete", "Complete", "Rejected", "Stale"],
        2,
        1001,
    )
    set_table_ref(ws, "tblSupplierCostEvidence")


def update_labor_controls(workbook: Any) -> None:
    ws = workbook["04 - Labor References"]
    mapping = append_columns(
        ws,
        [
            "Approval Status",
            "Approver",
            "Approval Date",
            "Labor Policy Version",
            "Evidence Reference",
            "Labor Dataset Version",
            "Labor Readiness Status",
        ],
    )
    for row in range(2, 27):
        ws.cell(row, mapping["Approval Status"], "Pending Approval")
        ws.cell(row, mapping["Approval Date"]).number_format = "yyyy-mm-dd"
        ws.cell(
            row,
            mapping["Labor Readiness Status"],
            (
                f'=IF(AND(C{row}<>"",E{row}<>"",K{row}="Approved",'
                f'N{row}<>"",O{row}<>"",P{row}<>""),"Ready","Blocked")'
            ),
        )
    add_list_validation(
        ws,
        mapping["Approval Status"],
        ["Pending Approval", "Approved", "Rejected", "Superseded"],
        2,
        26,
    )
    set_table_ref(ws, "tblLaborReferences")


def update_policy_controls(workbook: Any) -> None:
    ws = workbook["06 - Pricing Policy"]
    mapping = append_columns(
        ws, ["Decision Evidence Reference", "Policy Gate Status"]
    )
    for row in range(2, 15):
        ws.cell(row, mapping["Policy Gate Status"], "Blocked - Pending Approval")
    add_list_validation(
        ws,
        mapping["Policy Gate Status"],
        ["Blocked - Pending Approval", "Ready", "Rejected", "Superseded"],
        2,
        14,
    )
    set_table_ref(ws, "tblPricingPolicy")


def update_calculation_controls(workbook: Any) -> None:
    ws = workbook["07 - Calculation Results"]
    mapping = append_columns(
        ws,
        [
            "Policy Version",
            "Supplier Dataset Version",
            "Labor Dataset Version",
            "Competitor Dataset Version",
            "Gate Status",
            "Gate Blocking Reason",
        ],
    )
    for row in range(2, 27):
        gate = f"$AF{row}"
        ws.cell(
            row,
            9,
            f'=IF({gate}<>"Ready","",IF(OR(G{row}="",H{row}=""),"",G{row}/60*H{row}))',
        )
        ws.cell(
            row,
            13,
            f'=IF({gate}<>"Ready","",IF(COUNT(C{row}:F{row},I{row}:L{row})<8,"",'
            f"SUM(C{row}:F{row},I{row}:L{row})))",
        )
        ws.cell(
            row,
            15,
            f'=IF({gate}<>"Ready","",IF(OR(M{row}="",N{row}=""),"",M{row}+N{row}))',
        )
        ws.cell(
            row,
            17,
            f'=IF({gate}<>"Ready","",IF(OR(M{row}="",P{row}="",P{row}<=0,'
            f'P{row}>=1),"",M{row}/(1-P{row})))',
        )
        ws.cell(
            row,
            18,
            f'=IF({gate}<>"Ready","",COUNTIFS(\'02 - Competitor Observations\'!$B:$B,'
            f'A{row},\'02 - Competitor Observations\'!$X:$X,"Verified Comparable",'
            f'\'02 - Competitor Observations\'!$Z:$Z,"Evidence",'
            f'\'02 - Competitor Observations\'!$AA:$AA,"Complete"))',
        )
        ws.cell(
            row,
            19,
            f'=IF(OR({gate}<>"Ready",R{row}=0),"",MINIFS('
            f'\'02 - Competitor Observations\'!$Q:$Q,'
            f'\'02 - Competitor Observations\'!$B:$B,A{row},'
            f'\'02 - Competitor Observations\'!$X:$X,"Verified Comparable",'
            f'\'02 - Competitor Observations\'!$Z:$Z,"Evidence",'
            f'\'02 - Competitor Observations\'!$AA:$AA,"Complete"))',
        )
        ws.cell(
            row,
            20,
            f'=IF(OR({gate}<>"Ready",R{row}=0),"",MEDIAN(FILTER('
            f'\'02 - Competitor Observations\'!$Q:$Q,'
            f'(\'02 - Competitor Observations\'!$B:$B=A{row})*'
            f'(\'02 - Competitor Observations\'!$X:$X="Verified Comparable")*'
            f'(\'02 - Competitor Observations\'!$Z:$Z="Evidence")*'
            f'(\'02 - Competitor Observations\'!$AA:$AA="Complete"))))',
        )
        ws.cell(
            row,
            21,
            f'=IF(OR({gate}<>"Ready",R{row}=0),"",AVERAGEIFS('
            f'\'02 - Competitor Observations\'!$Q:$Q,'
            f'\'02 - Competitor Observations\'!$B:$B,A{row},'
            f'\'02 - Competitor Observations\'!$X:$X,"Verified Comparable",'
            f'\'02 - Competitor Observations\'!$Z:$Z,"Evidence",'
            f'\'02 - Competitor Observations\'!$AA:$AA,"Complete"))',
        )
        ws.cell(
            row,
            24,
            f'=IF({gate}<>"Ready","",IF(OR(W{row}="",M{row}=""),"",W{row}-M{row}))',
        )
        ws.cell(
            row,
            25,
            f'=IF({gate}<>"Ready","",IF(OR(W{row}="",W{row}=0,X{row}=""),"",X{row}/W{row}))',
        )
        ws.cell(row, 26, "Disabled")
        ws.cell(
            row,
            27,
            "Identity, policy, supplier, labor, competitor, review, and approval gates remain incomplete.",
        )
        ws.cell(row, mapping["Gate Status"], "Blocked")
        ws.cell(
            row,
            mapping["Gate Blocking Reason"],
            "Required governed inputs and version provenance are incomplete.",
        )
        ws.cell(row, 23, None)
        ws.cell(row, 23).protection = Protection(locked=True)
    for validation in ws.data_validations.dataValidation:
        if validation.type == "list" and "Z2:Z26" in str(validation.sqref):
            validation.formula1 = '"Disabled,Ready for Review,Calculated,Approved"'
    add_list_validation(
        ws, mapping["Gate Status"], ["Blocked", "Ready"], 2, 26
    )
    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True
    ws["W1"].comment = Comment(
        "Final Approved Price is blank and protected. Population requires a separate explicitly authorized human-approval workflow.",
        "Nocturnix Governance",
    )
    ws.conditional_formatting.add(
        "AF2:AF26",
        FormulaRule(
            formula=['AF2="Blocked"'],
            fill=PatternFill("solid", fgColor=PALE_RED),
        ),
    )
    set_table_ref(ws, "tblCalculationResults")


def update_review_controls(workbook: Any) -> None:
    ws = workbook["08 - Review Queue"]
    mapping = append_columns(
        ws,
        ["Due Date", "Resolution Reviewer", "Closure Date", "Approval State"],
    )
    for row in range(2, 127):
        ws.cell(row, mapping["Approval State"], "Not Approved")
        ws.cell(row, mapping["Due Date"]).number_format = "yyyy-mm-dd"
        ws.cell(row, mapping["Closure Date"]).number_format = "yyyy-mm-dd"
    add_list_validation(
        ws,
        mapping["Approval State"],
        ["Not Approved", "Pending Approval", "Approved", "Rejected"],
        2,
        126,
    )
    set_table_ref(ws, "tblReviewQueue")


def replace_validation_summary(workbook: Any) -> None:
    ws = workbook["09 - Validation Summary"]
    for table_name in list(ws.tables):
        del ws.tables[table_name]
    ws.delete_rows(1, ws.max_row)
    rows = [
        ["GATE-001", "Structural Validity", "PASS", "PASS", "PASS", 0, "Workbook contract validated."],
        ["GATE-002", "Identity Readiness", "25 governed", "0 governed", "BLOCKED", 25, "Governed manufacturer/model identities required."],
        ["GATE-003", "Policy Readiness", "13 approved", "0 approved", "BLOCKED", 13, "All policy decisions remain Pending Approval."],
        ["GATE-004", "Supplier-Cost Evidence Readiness", "25 complete", "0 complete", "BLOCKED", 25, "22 legacy references preserved; 3 No Evidence Captured rows."],
        ["GATE-005", "Labor Readiness", "25 ready", "0 ready", "BLOCKED", 25, "Governed LAB mappings and approval provenance required."],
        ["GATE-006", "Competitor-Evidence Readiness", "Approved observation rule satisfied", "0 evidence rows", "BLOCKED", 25, "Template row is excluded from evidence."],
        ["GATE-007", "Calculation Readiness", "Enabled", "Disabled", "BLOCKED", 25, "Calculations remain disabled."],
        ["GATE-008", "Review Readiness", "0 blocking items", "125 open", "BLOCKED", 125, "Review queue remains open."],
        ["GATE-009", "Final-Approval Readiness", "Authorized approval workflow", "Not Authorized; 25 blank", "BLOCKED", 25, "Final Approved Price remains blank and protected."],
        ["GATE-010", "Production-Activation Readiness", "Separately authorized", "Prohibited", "BLOCKED", 1, "No production activation authorized."],
    ]
    ws.append(
        [
            "Gate ID",
            "Gate Category",
            "Required State",
            "Actual State",
            "Gate Status",
            "Blocking Count",
            "Notes",
        ]
    )
    for row in rows:
        ws.append(row)
    from openpyxl.worksheet.table import Table, TableStyleInfo

    table = Table(displayName="tblValidationSummary", ref="A1:G11")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    add_list_validation(
        ws, 5, ["PASS", "BLOCKED", "FAIL", "NOT APPLICABLE"], 2, 11
    )
    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True


def update_revision_and_metadata(
    workbook: Any,
    source: Path,
    source_hashes: dict[str, str],
    generated: datetime,
) -> None:
    history = workbook["10 - Revision History"]
    history.append(
        [
            "v0.2 Draft",
            generated.isoformat(),
            "Pre-evidence design revision",
            (
                "Created from locked v0.1; added evidence, provenance, readiness, "
                "calculation-gate, final-price protection, and non-production controls. "
                "No pricing inputs or final prices populated."
            ),
            "Nocturnix",
            "Draft",
        ]
    )
    set_table_ref(history, "tblRevisionHistory")
    history.protection.sheet = True

    metadata = workbook["11 - Import Metadata"]
    rows = [
        ["Workbook Version", "v0.2 Draft", "Working revision; not production-active"],
        ["Lifecycle Status", "Draft / Pre-Evidence", "Not pricing-ready"],
        ["Parent Workbook Path", str(source), "Locked v0.1 checkpoint"],
        ["Parent Workbook SHA-256", V01_SHA256, "Verified before reading"],
        ["Generated Timestamp UTC", generated.isoformat(), "System-generated"],
        ["Generator Version", GENERATOR_VERSION, "Local governed generator"],
        ["QA Profile Version", QA_PROFILE_VERSION, "Dedicated v0.2 Draft profile"],
        ["Canonical Source Path", source_hashes["canonical_path"], "Read-only approved source"],
        ["Canonical Source Version", "v1.0", "Approved normalization release"],
        ["Canonical Source SHA-256", source_hashes["canonical"], "Verified source hash"],
        ["Pricing Source SHA-256", source_hashes["pricing"], "Read-only source"],
        ["Services Source SHA-256", source_hashes["services"], "Read-only source"],
        ["Parts Source SHA-256", source_hashes["parts"], "Read-only source"],
        ["Policy Version", "Not Approved / Not Assigned", "All 13 policies Pending Approval"],
        ["Supplier Dataset Version", "Not Assigned", "Evidence collection not started"],
        ["Labor Dataset Version", "Not Assigned", "Governed mappings not approved"],
        ["Competitor Dataset Version", "Not Assigned", "Evidence collection not started"],
        ["Calculation Status", "Disabled", "All calculation gates blocked"],
        ["Production Activation", "Prohibited", "Separate authorization required"],
        ["Pilot Records", 25, "Approved cohort unchanged"],
        ["Final Approved Price Populated", 0, "No final-price approval authorized"],
        ["Pricing Inputs Populated by v0.2 Transformation", 0, "No pricing inputs invented or populated"],
    ]
    metadata.delete_rows(2, metadata.max_row - 1)
    for row in rows:
        metadata.append(row)
    set_table_ref(metadata, "tblImportMetadata")
    metadata.protection.sheet = True


def validate_v02(path: Path) -> list[str]:
    errors: list[str] = []
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    try:
        if workbook.sheetnames != SHEETS:
            errors.append("Worksheet order mismatch")
        expected_tables = {
            "00 - Instructions": "tblInstructions",
            "01 - Pilot Pricing Records": "tblPilotPricingRecords",
            "02 - Competitor Observations": "tblCompetitorObservations",
            "03 - Supplier Cost Evidence": "tblSupplierCostEvidence",
            "04 - Labor References": "tblLaborReferences",
            "05 - Canonical References": "tblCanonicalReferences",
            "06 - Pricing Policy": "tblPricingPolicy",
            "07 - Calculation Results": "tblCalculationResults",
            "08 - Review Queue": "tblReviewQueue",
            "09 - Validation Summary": "tblValidationSummary",
            "10 - Revision History": "tblRevisionHistory",
            "11 - Import Metadata": "tblImportMetadata",
        }
        for sheet, table in expected_tables.items():
            ws = workbook[sheet]
            if table not in ws.tables:
                errors.append(f"{sheet}: missing table {table}")
            if ws.freeze_panes != "A2":
                errors.append(f"{sheet}: freeze panes differ from A2")
            if "v0.2 DRAFT" not in (ws.oddHeader.center.text or ""):
                errors.append(f"{sheet}: draft banner missing")
        pilot = workbook["01 - Pilot Pricing Records"]
        pm = headers(pilot)
        pricing_ids = [pilot.cell(row, pm["Pricing Record ID"]).value for row in range(2, 27)]
        service_ids = [pilot.cell(row, pm["Service ID"]).value for row in range(2, 27)]
        if len(set(pricing_ids)) != 25:
            errors.append("Pricing Record IDs are not 25 unique values")
        if len(set(service_ids)) != 25:
            errors.append("Service IDs are not 25 unique values")
        if "SVC000343" in service_ids:
            errors.append("SVC000343 is present in the cohort")
        supplier = workbook["03 - Supplier Cost Evidence"]
        sm = headers(supplier)
        supplier_rows = table_rows(supplier)
        if len(supplier_rows) != 25:
            errors.append("Supplier evidence does not contain exactly 25 status rows")
        if sum(row["Evidence Record Status"] == "Legacy Reference Preserved" for row in supplier_rows) != 22:
            errors.append("Supplier legacy-reference preservation count is not 22")
        if sum(row["Evidence Record Status"] == "No Evidence Captured" for row in supplier_rows) != 3:
            errors.append("No Evidence Captured count is not 3")
        competitor = workbook["02 - Competitor Observations"]
        cm = headers(competitor)
        if competitor.cell(2, cm["Record Role"]).value != "Template - Do Not Count":
            errors.append("Competitor template row is not separated from evidence")
        calc = workbook["07 - Calculation Results"]
        xm = headers(calc)
        for row in range(2, 27):
            if calc.cell(row, xm["Calculation Status"]).value != "Disabled":
                errors.append(f"Calculation status is not Disabled at row {row}")
            if calc.cell(row, xm["Gate Status"]).value != "Blocked":
                errors.append(f"Calculation gate is not Blocked at row {row}")
            if calc.cell(row, xm["Final Approved Price"]).value not in (None, ""):
                errors.append(f"Final Approved Price is populated at row {row}")
            if not calc.cell(row, xm["Final Approved Price"]).protection.locked:
                errors.append(f"Final Approved Price is not locked at row {row}")
        if not calc.protection.sheet:
            errors.append("Calculation worksheet protection is not enabled")
        policies = table_rows(workbook["06 - Pricing Policy"])
        if len(policies) != 13 or any(row["Approval Status"] != "Pending Approval" for row in policies):
            errors.append("Policy rows are not exactly 13 Pending Approval items")
        gates = table_rows(workbook["09 - Validation Summary"])
        if len(gates) != 10:
            errors.append("Validation Summary does not contain 10 gates")
        if gates[0]["Gate Status"] != "PASS" or any(
            row["Gate Status"] != "BLOCKED" for row in gates[1:]
        ):
            errors.append("Validation Summary gate states are incorrect")
        metadata = {row["Metadata Field"]: row["Value"] for row in table_rows(workbook["11 - Import Metadata"])}
        if metadata.get("Calculation Status") != "Disabled":
            errors.append("Metadata Calculation Status is not Disabled")
        if metadata.get("Production Activation") != "Prohibited":
            errors.append("Metadata Production Activation is not Prohibited")
        if metadata.get("Parent Workbook SHA-256") != V01_SHA256:
            errors.append("Parent workbook hash metadata mismatch")
        if metadata.get("Final Approved Price Populated") != 0:
            errors.append("Final price metadata is not zero")
        formula_errors = {
            "#REF!",
            "#DIV/0!",
            "#VALUE!",
            "#NAME?",
            "#N/A",
        }
        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in formula_errors:
                        errors.append(f"Formula error literal at {ws.title}!{cell.coordinate}")
    finally:
        workbook.close()
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--target", type=Path, required=True)
    parser.add_argument("--pricing", type=Path, required=True)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--services", type=Path, required=True)
    parser.add_argument("--parts", type=Path, required=True)
    args = parser.parse_args()
    source = args.source.resolve()
    if sha256(source) != V01_SHA256:
        raise ValueError("Locked v0.1 SHA-256 mismatch")
    source_hashes = {
        "pricing": sha256(args.pricing),
        "canonical": sha256(args.canonical),
        "services": sha256(args.services),
        "parts": sha256(args.parts),
        "canonical_path": str(args.canonical.resolve()),
    }
    if source_hashes["canonical"] != CANONICAL_SHA256:
        raise ValueError("Canonical source SHA-256 mismatch")
    workbook = load_workbook(source, read_only=False, data_only=False, keep_links=True)
    generated = datetime.now(UTC).replace(microsecond=0)
    try:
        workbook.properties.title = "Nocturnix Competitive Pricing Pilot v0.2 Draft"
        workbook.properties.subject = "Governed pre-evidence working revision"
        workbook.properties.description = (
            "Draft, calculation-disabled, non-production workbook derived from locked v0.1."
        )
        banner_workbook(workbook)
        update_instructions(workbook)
        update_identity_controls(workbook)
        update_competitor_controls(workbook)
        update_supplier_controls(workbook)
        update_labor_controls(workbook)
        update_policy_controls(workbook)
        update_calculation_controls(workbook)
        update_review_controls(workbook)
        replace_validation_summary(workbook)
        update_revision_and_metadata(
            workbook, source, source_hashes, generated
        )
        target = args.target.resolve()
        target.parent.mkdir(parents=True, exist_ok=True)
        temporary = target.with_suffix(".tmp.xlsx")
        workbook.save(temporary)
    finally:
        workbook.close()
    errors = validate_v02(temporary)
    if errors:
        temporary.unlink(missing_ok=True)
        raise ValueError("v0.2 structural validation failed:\n- " + "\n- ".join(errors))
    temporary.replace(target)
    print(f"Created: {target}")
    print(f"SHA-256: {sha256(target)}")
    print("Structural validation: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
