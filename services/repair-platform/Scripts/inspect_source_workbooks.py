from pathlib import Path

import openpyxl

paths = [
    Path(
        r"D:\Business Portal\300_Pricing\Working\Nocturnix_Manufacturer_Registry_v0.1_Draft.xlsx"
    ),
    Path(
        r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Devices_Catalog_v1.xlsx"
    ),
    Path(
        r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Parts_Catalog_v1.xlsx"
    ),
    Path(
        r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Services_Catalog_v1.xlsx"
    ),
    Path(
        r"D:\Business Portal\300_Pricing\Working\Nocturnix_Master_Compatibility_Catalog_v1.xlsx"
    ),
]

for path in paths:
    print(f"=== {path.name} ===")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames[:5]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"-- {sheet_name} --")
        for row in rows[:8]:
            print(row)
        print()
    print()
