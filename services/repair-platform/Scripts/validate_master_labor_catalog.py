"""Validate the governed Master Labor Catalog V1 review workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
CANONICAL_LABOR_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards"
    r"\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
OUTPUT_PATH = WORKING_DIR / "Nocturnix_Master_Labor_Catalog_v1.xlsx"
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(f".{OUTPUT_PATH.stem}.tmp.xlsx")
SOURCE_SHEET = "01 - Labor Standards"
IMPORT_BATCH = "MASTER-LABOR-V1-REVIEW"
LABOR_ID_PATTERN = re.compile(r"^LAB\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Labor Standards",
    "02 - Repair Categories",
    "03 - Repair Types",
    "04 - Device Families",
    "05 - Manufacturers",
    "06 - Labor Tiers",
    "07 - Skill Levels",
    "08 - Difficulty",
    "09 - Warranty Options",
    "10 - Review Queue",
    "11 - Validation Summary",
    "12 - Revision History",
    "13 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterLaborInstructions",
    "01 - Labor Standards": "tblMasterLaborCatalog",
    "02 - Repair Categories": "tblLaborRepairCategories",
    "03 - Repair Types": "tblLaborRepairTypes",
    "04 - Device Families": "tblLaborDeviceFamilies",
    "05 - Manufacturers": "tblLaborManufacturers",
    "06 - Labor Tiers": "tblLaborTiers",
    "07 - Skill Levels": "tblLaborSkillLevels",
    "08 - Difficulty": "tblLaborDifficulties",
    "09 - Warranty Options": "tblLaborWarrantyOptions",
    "10 - Review Queue": "tblLaborReviewQueue",
    "11 - Validation Summary": "tblMasterLaborValidation",
    "12 - Revision History": "tblMasterLaborRevisionHistory",
    "13 - Import Metadata": "tblMasterLaborImportMetadata",
}
LABOR_HEADERS = [
    "Labor Standard ID",
    "Legacy Labor ID",
    "Labor Name",
    "Repair Category",
    "Repair Type",
    "Device Family",
    "Manufacturer",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Labor Tier",
    "Skill Level",
    "Difficulty",
    "Warranty Option",
    "Requires Calibration",
    "Requires Waterproof Test",
    "Requires Programming",
    "Requires Pairing",
    "Requires Board Repair",
    "Special Tools Required",
    "Technician Certification",
    "Review Status",
    "Confidence",
    "Source Record Number",
    "Source Workbook",
    "Source Worksheet",
    "Import Batch",
    "Reviewer",
    "Reviewer Notes",
    "Created At",
    "Updated At",
]
QUEUE_HEADERS = [
    "Labor Standard ID",
    "Labor Name",
    "Missing Evidence",
    "Relationship Issue",
    "Required Action",
    "Review Status",
    "Reviewer",
    "Reviewer Notes",
]
REVIEW_STATUSES = {
    "Pending Review",
    "Pending Evidence Review",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
}
CONFIDENCE_VALUES = {"Unassessed", "Low", "Medium", "High"}
YES_NO_VALUES = {"Yes", "No"}
FLAG_HEADERS = {
    "Requires Calibration",
    "Requires Waterproof Test",
    "Requires Programming",
    "Requires Pairing",
    "Requires Board Repair",
}
DEFINED_NAME_BY_HEADER = {
    "Repair Category": "DV_LaborRepairCategories",
    "Repair Type": "DV_LaborRepairTypes",
    "Device Family": "DV_LaborDeviceFamilies",
    "Manufacturer": "DV_LaborManufacturers",
    "Labor Tier": "DV_LaborTiers",
    "Skill Level": "DV_LaborSkillLevels",
    "Difficulty": "DV_LaborDifficulties",
    "Warranty Option": "DV_LaborWarrantyOptions",
    "Review Status": "DV_LaborReviewStatuses",
    "Confidence": "DV_LaborConfidence",
    **{header: "DV_YesNo" for header in FLAG_HEADERS},
}
DEFINED_NAME_SPECS = {
    "DV_LaborRepairCategories": ("02 - Repair Categories", 1),
    "DV_LaborRepairTypes": ("03 - Repair Types", 1),
    "DV_LaborDeviceFamilies": ("04 - Device Families", 1),
    "DV_LaborManufacturers": ("05 - Manufacturers", 1),
    "DV_LaborTiers": ("06 - Labor Tiers", 1),
    "DV_LaborSkillLevels": ("07 - Skill Levels", 1),
    "DV_LaborDifficulties": ("08 - Difficulty", 1),
    "DV_LaborWarrantyOptions": ("09 - Warranty Options", 1),
    "DV_LaborReviewStatuses": ("09 - Warranty Options", 2),
    "DV_LaborConfidence": ("09 - Warranty Options", 3),
    "DV_YesNo": ("09 - Warranty Options", 4),
}
LOOKUP_FIELD_SPECS = {
    "Repair Category": ("02 - Repair Categories", "Repair Category"),
    "Repair Type": ("03 - Repair Types", "Repair Type"),
    "Device Family": ("04 - Device Families", "Device Family"),
    "Manufacturer": ("05 - Manufacturers", "Manufacturer"),
    "Labor Tier": ("06 - Labor Tiers", "Labor Tier"),
    "Skill Level": ("07 - Skill Levels", "Skill Level"),
    "Difficulty": ("08 - Difficulty", "Difficulty"),
    "Warranty Option": ("09 - Warranty Options", "Warranty Option"),
}
PROHIBITED_HEADERS = {
    "Labor Rate",
    "Labor Cost",
    "Customer Price",
    "Payroll",
    "Technician Schedule",
    "Clock In",
    "Clock Out",
    "Inventory",
    "Automatic Approval",
}


class LaborValidationError(RuntimeError):
    """Raised when the review artifact violates its governance contract."""


def text(value: Any) -> str:
    """Return a trimmed display value."""
    return "" if value is None else str(value).strip()


def file_hash(path: Path) -> str:
    """Return a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require all protected and review inputs."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise LaborValidationError(f"Missing required file: {', '.join(missing)}")


def require_ooxml(path: Path) -> None:
    """Validate core OOXML members and package safety."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    try:
        with zipfile.ZipFile(path) as archive:
            members = set(archive.namelist())
            bad_member = archive.testzip()
    except zipfile.BadZipFile as exc:
        raise LaborValidationError("Output is not a valid OOXML ZIP") from exc
    missing = required - members
    if missing:
        raise LaborValidationError(f"Missing OOXML members: {sorted(missing)}")
    if bad_member:
        raise LaborValidationError(f"Corrupt OOXML member: {bad_member}")
    worksheets = {
        member
        for member in members
        if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", member)
    }
    tables = {
        member
        for member in members
        if re.fullmatch(r"xl/tables/table\d+\.xml", member)
    }
    if len(worksheets) != len(SHEET_NAMES) or len(tables) != len(TABLE_NAMES):
        raise LaborValidationError("OOXML worksheet/table member count differs")
    if "xl/vbaProject.bin" in members:
        raise LaborValidationError("Review workbook contains a macro project")
    if any(member.startswith("xl/externalLinks/") for member in members):
        raise LaborValidationError("Review workbook contains external links")


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Read table headers."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, _ = range_boundaries(table.ref)
    return [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read nonblank records from a table."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = table_headers(worksheet, table_name)
    records = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def locate_source_header(worksheet: Worksheet) -> tuple[int, list[str]]:
    """Locate the protected legacy Labor ID header."""
    for row_number in range(1, min(worksheet.max_row, 50) + 1):
        headers = [text(cell.value) for cell in worksheet[row_number]]
        if "Labor ID" in headers or "Legacy Labor ID" in headers:
            return row_number, headers
    raise LaborValidationError(f"{SOURCE_SHEET} lacks a legacy Labor ID header")


def canonical_identity() -> tuple[set[str], int, dict[str, Any]]:
    """Return the governed namespace and source lineage-to-legacy-ID map."""
    workbook = load_workbook(
        CANONICAL_LABOR_PATH,
        read_only=True,
        data_only=False,
    )
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise LaborValidationError(f"Missing source sheet: {SOURCE_SHEET}")
        worksheet = workbook[SOURCE_SHEET]
        header_row, headers = locate_source_header(worksheet)
        governed_id_column = (
            headers.index("Labor Standard ID") + 1
            if "Labor Standard ID" in headers
            else None
        )
        legacy_id_header = (
            "Legacy Labor ID" if "Legacy Labor ID" in headers else "Labor ID"
        )
        legacy_id_column = headers.index(legacy_id_header) + 1
        source_number_column = (
            headers.index("Source Record Number") + 1
            if "Source Record Number" in headers
            else (
                headers.index("Record Number") + 1
                if "Record Number" in headers
                else None
            )
        )
        identifiers: list[str] = []
        source_legacy_by_lineage: dict[str, Any] = {}
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            row_values = [
                worksheet.cell(row_number, column).value
                for column in range(1, len(headers) + 1)
            ]
            if not any(text(value) for value in row_values):
                continue
            legacy_value = worksheet.cell(row_number, legacy_id_column).value
            legacy_id = text(legacy_value)
            if not legacy_id:
                raise LaborValidationError(
                    f"Blank Legacy Labor ID at source row {row_number}"
                )
            lineage = (
                text(worksheet.cell(row_number, source_number_column).value)
                if source_number_column is not None
                else str(row_number)
            )
            if not lineage:
                raise LaborValidationError(
                    f"Blank source-row lineage at source row {row_number}"
                )
            if lineage in source_legacy_by_lineage:
                raise LaborValidationError(
                    f"Duplicate source-row lineage: {lineage}"
                )
            source_legacy_by_lineage[lineage] = legacy_value
            identifier = (
                text(worksheet.cell(row_number, governed_id_column).value)
                if governed_id_column is not None
                else ""
            )
            if not identifier:
                continue
            if not LABOR_ID_PATTERN.fullmatch(identifier):
                raise LaborValidationError(
                    f"Malformed canonical Labor ID at row {row_number}: "
                    f"{identifier}"
                )
            identifiers.append(identifier)
    finally:
        workbook.close()
    duplicates = sorted(
        value for value, count in Counter(identifiers).items() if count > 1
    )
    if duplicates:
        raise LaborValidationError(
            f"Duplicate canonical Labor IDs: {duplicates[:10]}"
        )
    numbers = [int(identifier[3:]) for identifier in identifiers]
    if numbers != sorted(numbers):
        raise LaborValidationError("Canonical Labor IDs are out of order")
    legacy_ids = [text(value) for value in source_legacy_by_lineage.values()]
    legacy_duplicates = sorted(
        value for value, count in Counter(legacy_ids).items() if count > 1
    )
    if legacy_duplicates:
        raise LaborValidationError(
            f"Duplicate Legacy Labor IDs: {legacy_duplicates[:10]}"
        )
    return set(identifiers), max(numbers, default=0), source_legacy_by_lineage


def validate_structure(workbook: Any) -> None:
    """Validate worksheet and table contracts."""
    if workbook.sheetnames != SHEET_NAMES:
        raise LaborValidationError("Worksheet contract or order differs")
    if len(set(workbook.sheetnames)) != len(SHEET_NAMES):
        raise LaborValidationError("Worksheet names are duplicated")
    if any(len(name) > 31 for name in workbook.sheetnames):
        raise LaborValidationError("Worksheet name exceeds 31 characters")
    all_tables: list[str] = []
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        expected = TABLE_NAMES[sheet_name]
        if list(worksheet.tables) != [expected]:
            raise LaborValidationError(f"Table contract differs: {sheet_name}")
        table = worksheet.tables[expected]
        if table.autoFilter is None:
            raise LaborValidationError(f"Table filter missing: {sheet_name}")
        if worksheet.freeze_panes != "A2":
            raise LaborValidationError(f"Frozen header missing: {sheet_name}")
        all_tables.append(expected)
    if len(all_tables) != len(set(all_tables)):
        raise LaborValidationError("Table names are not globally unique")
    primary = workbook["01 - Labor Standards"]
    headers = table_headers(primary, TABLE_NAMES[primary.title])
    if headers != LABOR_HEADERS:
        raise LaborValidationError("Primary Labor schema differs")
    if PROHIBITED_HEADERS & set(headers):
        raise LaborValidationError("Primary Labor schema has a prohibited field")
    queue = workbook["10 - Review Queue"]
    if table_headers(queue, TABLE_NAMES[queue.title]) != QUEUE_HEADERS:
        raise LaborValidationError("Review Queue schema differs")


def validate_names_and_validations(workbook: Any) -> None:
    """Validate name destinations and primary data validations."""
    actual_names = set(workbook.defined_names)
    if actual_names != set(DEFINED_NAME_SPECS):
        raise LaborValidationError(
            "Defined-name contract differs: "
            f"expected {sorted(DEFINED_NAME_SPECS)}, got {sorted(actual_names)}"
        )
    for name, (expected_sheet, expected_column) in DEFINED_NAME_SPECS.items():
        defined_name = workbook.defined_names[name]
        destinations = list(defined_name.destinations)
        if len(destinations) != 1:
            raise LaborValidationError(f"Defined name {name} is not singular")
        sheet_name, coordinates = destinations[0]
        min_column, min_row, max_column, _ = range_boundaries(coordinates)
        if (
            sheet_name != expected_sheet
            or min_column != expected_column
            or max_column != expected_column
            or min_row != 2
        ):
            raise LaborValidationError(f"Defined name {name} targets wrong range")

    worksheet = workbook["01 - Labor Standards"]
    validation_by_column: dict[int, str] = {}
    for validation in worksheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        formula = text(validation.formula1)
        if not formula.startswith("=") or "!" in formula or "," in formula:
            raise LaborValidationError("Invalid list validation formula")
        defined_name = formula.removeprefix("=")
        if defined_name not in DEFINED_NAME_SPECS:
            raise LaborValidationError(
                f"Validation references unknown name: {defined_name}"
            )
        for cell_range in validation.sqref.ranges:
            min_column, min_row, max_column, max_row = range_boundaries(
                str(cell_range)
            )
            if min_column != max_column or min_row != 2:
                raise LaborValidationError("Validation destination is invalid")
            if max_row != worksheet.max_row:
                raise LaborValidationError("Validation does not cover table rows")
            validation_by_column[min_column] = defined_name
    expected = {
        LABOR_HEADERS.index(header) + 1: defined_name
        for header, defined_name in DEFINED_NAME_BY_HEADER.items()
    }
    if validation_by_column != expected:
        raise LaborValidationError("Controlled-field validation contract differs")


def whole_number(value: Any, *, allow_zero: bool) -> int | None:
    """Parse an optional whole number under the labor minute policy."""
    if not text(value):
        return None
    if isinstance(value, bool):
        raise ValueError
    number = int(value)
    if number != value and text(number) != text(value):
        raise ValueError
    if number < 0 or (number == 0 and not allow_zero):
        raise ValueError
    return number


def lookup_values(
    workbook: Any,
    sheet_name: str,
    header: str,
) -> set[str]:
    """Read a governed lookup column."""
    records = table_records(workbook[sheet_name], TABLE_NAMES[sheet_name])
    return {text(record.get(header)) for record in records if text(record.get(header))}


def validate_records(
    workbook: Any,
    existing: set[str],
    highest: int,
    source_legacy_by_lineage: dict[str, Any],
) -> list[dict[str, Any]]:
    """Validate identity, population, provenance, relationships, and status."""
    worksheet = workbook["01 - Labor Standards"]
    records = table_records(worksheet, TABLE_NAMES[worksheet.title])
    if len(records) != len(source_legacy_by_lineage):
        raise LaborValidationError(
            "Population differs: expected "
            f"{len(source_legacy_by_lineage)}, got {len(records)}"
        )
    identifiers = [text(record["Labor Standard ID"]) for record in records]
    if any(not LABOR_ID_PATTERN.fullmatch(value) for value in identifiers):
        raise LaborValidationError("Generated Labor ID is malformed")
    if len(identifiers) != len(set(identifiers)):
        raise LaborValidationError("Generated Labor IDs are duplicated")
    if set(identifiers) & existing:
        raise LaborValidationError("Generated Labor IDs overlap protected IDs")
    expected_ids = [
        f"LAB{highest + offset:06d}"
        for offset in range(1, len(records) + 1)
    ]
    if identifiers != expected_ids:
        raise LaborValidationError(
            "Generated Labor IDs are discontinuous or out of order"
        )
    lineage = [text(record["Source Record Number"]) for record in records]
    if any(not value for value in lineage) or len(lineage) != len(set(lineage)):
        raise LaborValidationError("Source lineage is blank or duplicated")
    if set(lineage) != set(source_legacy_by_lineage):
        raise LaborValidationError("Generated source lineage differs from source")

    lookups = {
        field: lookup_values(workbook, sheet_name, header)
        for field, (sheet_name, header) in LOOKUP_FIELD_SPECS.items()
    }
    for row_number, record in enumerate(records, start=2):
        if not text(record["Labor Name"]):
            raise LaborValidationError(f"Labor Name blank at row {row_number}")
        legacy_value = record["Legacy Labor ID"]
        legacy_id = text(legacy_value)
        if not legacy_id:
            raise LaborValidationError(
                f"Legacy Labor ID blank at row {row_number}"
            )
        expected_legacy_id = source_legacy_by_lineage[
            text(record["Source Record Number"])
        ]
        if legacy_value != expected_legacy_id:
            raise LaborValidationError(
                f"Legacy Labor ID differs from source at row {row_number}"
            )
        if text(record["Source Workbook"]) != str(CANONICAL_LABOR_PATH):
            raise LaborValidationError(f"Source Workbook differs at row {row_number}")
        if text(record["Source Worksheet"]) != SOURCE_SHEET:
            raise LaborValidationError(f"Source Worksheet differs at row {row_number}")
        if text(record["Import Batch"]) != IMPORT_BATCH:
            raise LaborValidationError(f"Import Batch differs at row {row_number}")
        if text(record["Review Status"]) != "Pending Review":
            raise LaborValidationError(
                f"Generated status is not Pending Review at row {row_number}"
            )
        if text(record["Reviewer"]):
            raise LaborValidationError(f"Generated Reviewer populated at row {row_number}")
        confidence = text(record["Confidence"])
        if confidence not in CONFIDENCE_VALUES:
            raise LaborValidationError(f"Invalid Confidence at row {row_number}")
        for header in FLAG_HEADERS:
            value = text(record[header])
            if value and value not in YES_NO_VALUES:
                raise LaborValidationError(
                    f"Invalid {header} at row {row_number}: {value}"
                )
        for field, allowed in lookups.items():
            value = text(record[field])
            if value and value not in allowed:
                raise LaborValidationError(
                    f"{field} does not resolve at row {row_number}: {value}"
                )
        try:
            minimum = whole_number(record["Minimum Minutes"], allow_zero=True)
            standard = whole_number(record["Standard Minutes"], allow_zero=False)
            maximum = whole_number(record["Maximum Minutes"], allow_zero=False)
        except (TypeError, ValueError):
            raise LaborValidationError(
                f"Invalid whole-number minutes at row {row_number}"
            ) from None
        if minimum is not None and standard is not None and minimum > standard:
            raise LaborValidationError(f"Minimum exceeds Standard at row {row_number}")
        if standard is not None and maximum is not None and standard > maximum:
            raise LaborValidationError(f"Standard exceeds Maximum at row {row_number}")
        if minimum is not None and maximum is not None and minimum > maximum:
            raise LaborValidationError(f"Minimum exceeds Maximum at row {row_number}")
    return records


def validate_queue(workbook: Any, records: Sequence[dict[str, Any]]) -> None:
    """Require one pending review queue row per labor record."""
    worksheet = workbook["10 - Review Queue"]
    queue = table_records(worksheet, TABLE_NAMES[worksheet.title])
    primary_ids = [text(record["Labor Standard ID"]) for record in records]
    queue_ids = [text(record["Labor Standard ID"]) for record in queue]
    if queue_ids != primary_ids or len(queue_ids) != len(set(queue_ids)):
        raise LaborValidationError("Review Queue does not reconcile to Labor records")
    if any(text(record["Review Status"]) != "Pending Review" for record in queue):
        raise LaborValidationError("Review Queue contains a nonpending status")


def metadata_map(workbook: Any) -> dict[str, str]:
    """Read unique metadata keys."""
    worksheet = workbook["13 - Import Metadata"]
    records = table_records(worksheet, TABLE_NAMES[worksheet.title])
    keys = [text(record["Metadata Field"]) for record in records]
    if len(keys) != len(set(keys)):
        raise LaborValidationError("Import Metadata contains duplicate keys")
    return {
        text(record["Metadata Field"]): text(record["Value"])
        for record in records
    }


def validate_metadata(
    workbook: Any,
    records: Sequence[dict[str, Any]],
    existing: set[str],
    highest: int,
) -> None:
    """Reconcile metadata and the protected source digest."""
    metadata = metadata_map(workbook)
    expected = {
        "Import Batch": IMPORT_BATCH,
        "Namespace Authority": "ADR-011",
        "Schema Columns": str(len(LABOR_HEADERS)),
        "Canonical Valid ID Count": str(len(existing)),
        "Highest Canonical Labor ID": f"LAB{highest:06d}" if highest else "",
        "Generated Record Count": str(len(records)),
        "Canonical Import Authorized": "No",
        f"Protected Input Path: {CANONICAL_LABOR_PATH.name}": str(
            CANONICAL_LABOR_PATH
        ),
        f"SHA-256: {CANONICAL_LABOR_PATH.name}": file_hash(
            CANONICAL_LABOR_PATH
        ),
    }
    if records:
        expected["First Generated ID"] = text(records[0]["Labor Standard ID"])
        expected["Final Generated ID"] = text(records[-1]["Labor Standard ID"])
    for key, expected_value in expected.items():
        if metadata.get(key, "") != expected_value:
            raise LaborValidationError(f"Metadata mismatch for {key}")


def validate_reopen(path: Path) -> None:
    """Perform a second independent reopen integrity check."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise LaborValidationError("Second reopen changed worksheet contract")
    finally:
        workbook.close()


def main() -> int:
    """Run all Master Labor validation rules."""
    try:
        require_files([CANONICAL_LABOR_PATH, OUTPUT_PATH])
        if TEMP_OUTPUT_PATH.exists():
            raise LaborValidationError(
                f"Stale temporary output exists: {TEMP_OUTPUT_PATH}"
            )
        require_ooxml(OUTPUT_PATH)
        existing, highest, source_legacy_by_lineage = canonical_identity()
        source_hash_before = file_hash(CANONICAL_LABOR_PATH)
        workbook = load_workbook(OUTPUT_PATH, data_only=False, read_only=False)
        try:
            validate_structure(workbook)
            validate_names_and_validations(workbook)
            records = validate_records(
                workbook,
                existing,
                highest,
                source_legacy_by_lineage,
            )
            validate_queue(workbook, records)
            validate_metadata(workbook, records, existing, highest)
        finally:
            workbook.close()
        validate_reopen(OUTPUT_PATH)
        if file_hash(CANONICAL_LABOR_PATH) != source_hash_before:
            raise LaborValidationError("Protected Labor Standards hash changed")
        print(f"Validated: {OUTPUT_PATH}")
        print(f"Labor review rows: {len(records)}")
        print("Worksheet/table/defined-name contract: PASS")
        print("Identity/provenance/relationships: PASS")
        print("Protected source hash: PASS (unchanged)")
        print("Canonical import: NOT AUTHORIZED")
        return 0
    except (
        LaborValidationError,
        OSError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
    ) as exc:
        print(f"ERROR: {text(exc)}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
