"""Independently validate the Nocturnix Master Pricing Catalog v1 workbook."""

from __future__ import annotations

import hashlib
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
MASTER_SERVICES_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Services_Catalog_v1.xlsx"
)
LABOR_CATALOG_PATH = Path(
    r"D:\Business Portal\300_Pricing\Labor Standards"
    r"\Nocturnix_Standard_Labor_Catalog_v1.xlsx"
)
PROPOSAL_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Legacy_Catalog_Deduplication_Proposal_v1.xlsx"
)
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
PRICING_ID_SOURCE_PATH = CANONICAL_PATH
OUTPUT_PATH = Path(
    r"D:\Business Portal\300_Pricing\Working"
    r"\Nocturnix_Master_Pricing_Catalog_v1.xlsx"
)

MASTER_SERVICES_SHEET = "01 - Master Services"
LABOR_SHEET = "01 - Labor Standards"
EXCLUSIONS_SHEET = "02 - Duplicate Exclusions"
PRICING_ID_SOURCE_SHEET = "50 Pricing"
EXPECTED_ROWS = 314
IMPORT_BATCH_ID = "MASTER-PRICING-V1-REVIEW"
PRICING_ID_PATTERN = re.compile(r"^PRC\d{6}$")
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
NAMESPACE_AUTHORITY = "ADR-007"
SHEET_NAMES = [
    "00 - Instructions",
    "01 - Pricing Records",
    "02 - Pricing Statuses",
    "03 - Cost Components",
    "04 - Labor References",
    "05 - Service References",
    "06 - Margin Targets",
    "07 - Regional Markets",
    "08 - Pricing Methods",
    "09 - Review Queue",
    "10 - Validation Summary",
    "11 - Revision History",
    "12 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblMasterPricingInstructions",
    "01 - Pricing Records": "tblMasterPricingCatalog",
    "02 - Pricing Statuses": "tblPricingStatuses",
    "03 - Cost Components": "tblCostComponents",
    "04 - Labor References": "tblPricingLaborReferences",
    "05 - Service References": "tblPricingServiceReferences",
    "06 - Margin Targets": "tblMarginTargets",
    "07 - Regional Markets": "tblRegionalMarkets",
    "08 - Pricing Methods": "tblPricingMethods",
    "09 - Review Queue": "tblPricingReviewQueue",
    "10 - Validation Summary": "tblMasterPricingValidation",
    "11 - Revision History": "tblMasterPricingRevisionHistory",
    "12 - Import Metadata": "tblMasterPricingImportMetadata",
}
PRICING_HEADERS = [
    "Pricing Record ID", "Service ID", "Legacy Service SKU", "Service Name",
    "Manufacturer ID", "Manufacturer Name", "Device Family Code",
    "Device Family Name", "Legacy Pricing Status", "Pricing Status",
    "Pricing Method", "Currency", "Legacy Retail Price", "Legacy Cost",
    "Part Cost", "Shipping Cost", "Consumables Cost", "Testing Cost",
    "Labor Standard ID", "Standard Labor Minutes", "Labor Rate Tier",
    "Labor Rate", "Labor Cost", "Overhead Allocation", "Processing Fees",
    "Warranty Allowance", "Risk Allowance", "Total Internal Cost",
    "Target Margin Percent", "Market Adjustment", "Recommended Price",
    "Minimum Approved Price", "Maximum Approved Price", "Final Customer Price",
    "Regional Market", "Effective Date", "Expiration Date", "Review Status",
    "Pricing Confidence", "Source Record Number", "Source Workbook",
    "Source Worksheet", "Import Batch ID", "Reviewer", "Reviewer Notes",
    "Created At", "Updated At",
]
PRICING_STATUSES = {
    "Pending Pricing Review", "Legacy Price Review", "Cost Research Required",
    "Labor Mapping Required", "Market Research Required",
    "Ready for Pricing Calculation", "Ready for Approval", "Approved",
    "Rejected", "Archived",
}
PRICING_METHODS = {
    "Cost Plus", "Market Aligned", "Fixed Service Fee", "Diagnostic Fee",
    "Labor Only", "Parts and Labor", "Manual Review", "Not Yet Determined",
}
REVIEW_STATUSES = {
    "Pending Review", "In Review", "Ready for Approval", "Approved", "Rejected",
    "Archived",
}
CONFIDENCE_VALUES = {"Unassessed", "Low", "Medium", "High"}
COST_FIELDS = {
    "Part Cost", "Shipping Cost", "Consumables Cost", "Testing Cost", "Labor Rate",
    "Labor Cost", "Overhead Allocation", "Processing Fees", "Warranty Allowance",
    "Risk Allowance",
}
MONEY_FIELDS = {
    "Legacy Retail Price", "Legacy Cost", *COST_FIELDS, "Total Internal Cost",
    "Market Adjustment", "Recommended Price", "Minimum Approved Price",
    "Maximum Approved Price", "Final Customer Price",
}
GENERATED_BLANK_FIELDS = {
    *COST_FIELDS,
    "Total Internal Cost",
    "Target Margin Percent",
    "Market Adjustment",
    "Recommended Price",
    "Minimum Approved Price",
    "Maximum Approved Price",
    "Final Customer Price",
    "Regional Market",
    "Effective Date",
    "Expiration Date",
    "Reviewer",
    "Reviewer Notes",
}
DEFINED_NAME_BY_HEADER = {
    "Pricing Status": "DV_PricingStatuses",
    "Pricing Method": "DV_PricingMethods",
    "Currency": "DV_Currencies",
    "Service ID": "DV_ServiceIDs",
    "Labor Standard ID": "DV_LaborStandardIDs",
    "Labor Rate Tier": "DV_LaborRateTiers",
    "Target Margin Percent": "DV_MarginTargets",
    "Regional Market": "DV_RegionalMarkets",
    "Review Status": "DV_ReviewStatuses",
    "Pricing Confidence": "DV_PricingConfidenceValues",
}


class PricingValidationError(RuntimeError):
    """Raised when the pricing review artifact violates its contract."""


def text(value: Any) -> str:
    """Normalize workbook text without changing substantive content."""
    return "" if value is None else str(value).strip()


def excel_safe_value(value: Any) -> Any:
    """Normalize persisted date/time values using the generator's policy."""
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if isinstance(value, time):
        if value.tzinfo is not None and value.utcoffset() is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    return value


def assert_excel_safe_value_contract() -> None:
    """Exercise the persisted date/time comparison boundary."""
    aware_utc = datetime(2026, 7, 23, 12, 30, tzinfo=UTC)
    assert excel_safe_value(aware_utc) == datetime(2026, 7, 23, 12, 30)

    eastern = timezone(timedelta(hours=-4))
    aware_eastern = datetime(2026, 7, 23, 8, 30, tzinfo=eastern)
    assert excel_safe_value(aware_eastern) == datetime(2026, 7, 23, 12, 30)

    naive = datetime(2026, 7, 23, 12, 30)
    assert excel_safe_value(naive) is naive

    calendar_date = date(2026, 7, 23)
    assert excel_safe_value(calendar_date) is calendar_date

    aware_clock = time(8, 30, tzinfo=eastern)
    normalized_clock = excel_safe_value(aware_clock)
    assert normalized_clock == time(8, 30)
    assert normalized_clock.tzinfo is None

    assert excel_safe_value("") == ""
    assert excel_safe_value(None) is None
    timestamp_text = "2026-07-23T12:30:00Z"
    assert excel_safe_value(timestamp_text) == timestamp_text


def decimal_value(value: Any, field: str, *, allow_text: bool = False) -> Decimal | None:
    """Parse a finite Decimal while retaining blank-versus-zero semantics."""
    if value is None or text(value) == "":
        return None
    if isinstance(value, bool):
        raise PricingValidationError(f"{field} contains Boolean value {value!r}")
    try:
        result = Decimal(text(value))
    except InvalidOperation:
        if allow_text:
            return None
        raise PricingValidationError(f"{field} is not numeric: {value!r}") from None
    if not result.is_finite():
        raise PricingValidationError(f"{field} is not finite: {value!r}")
    return result


def persisted_equal(expected: Any, actual: Any) -> bool:
    """Compare source values with Excel-safe blank and date normalization."""
    expected = excel_safe_value(expected)
    actual = excel_safe_value(actual)
    if expected is None or text(expected) == "":
        return actual is None or text(actual) == ""
    temporal_types = (date, datetime, time)
    if isinstance(expected, temporal_types) or isinstance(actual, temporal_types):
        left = expected.isoformat() if isinstance(expected, temporal_types) else text(expected)
        right = actual.isoformat() if isinstance(actual, temporal_types) else text(actual)
        return left == right
    if isinstance(expected, (int, float, Decimal)) and not isinstance(expected, bool):
        try:
            return decimal_value(expected, "source") == decimal_value(actual, "output")
        except PricingValidationError:
            return False
    return text(expected) == text(actual)


def file_hash(path: Path) -> str:
    """Calculate SHA-256 without modifying the file."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_files(paths: Iterable[Path]) -> None:
    """Require each unique protected workbook and the output to be valid."""
    seen: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved in seen:
            raise PricingValidationError(f"Workbook configured twice: {path}")
        seen.add(resolved)
        if not path.is_file():
            raise PricingValidationError(f"Required workbook does not exist: {path}")
        if path.stat().st_size <= 0:
            raise PricingValidationError(f"Required workbook is empty: {path}")
        if not zipfile.is_zipfile(path):
            raise PricingValidationError(f"Workbook is not a valid Excel ZIP: {path}")


def require_generated_workbook(path: Path) -> None:
    """Fail clearly when generation left no valid Excel workbook."""
    message = (
        "Generated pricing workbook is missing or invalid; "
        "rerun the generator successfully."
    )
    if (
        not path.is_file()
        or path.stat().st_size <= 0
        or not zipfile.is_zipfile(path)
    ):
        raise PricingValidationError(message)
    try:
        with zipfile.ZipFile(path) as archive:
            required_members = {"[Content_Types].xml", "xl/workbook.xml"}
            if (
                not required_members.issubset(archive.namelist())
                or archive.testzip() is not None
            ):
                raise PricingValidationError(message)
    except (OSError, zipfile.BadZipFile) as exc:
        raise PricingValidationError(message) from exc


def read_records(
    path: Path,
    sheet_name: str,
    *,
    keep_vba: bool = False,
) -> list[dict[str, Any]]:
    """Read a header-row dataset without saving the source."""
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_vba=keep_vba,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise PricingValidationError(f"{path.name} lacks {sheet_name!r}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        try:
            headers = [text(value) for value in next(rows)]
        except StopIteration as exc:
            raise PricingValidationError(f"{path.name}:{sheet_name} is empty") from exc
        return [
            dict(zip(headers, values, strict=False))
            for values in rows
            if any(value is not None and text(value) != "" for value in values)
        ]
    finally:
        workbook.close()


def read_existing_pricing_ids() -> tuple[set[str], list[str]]:
    """Read canonical IDs, permitting ADR-007's approved empty namespace."""
    workbook = load_workbook(
        PRICING_ID_SOURCE_PATH,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )
    try:
        if PRICING_ID_SOURCE_SHEET not in workbook.sheetnames:
            raise PricingValidationError(
                f"Canonical workbook lacks {PRICING_ID_SOURCE_SHEET!r}"
            )
        worksheet = workbook[PRICING_ID_SOURCE_SHEET]
        header = next(
            (
                cell
                for row in worksheet.iter_rows()
                for cell in row
                if text(cell.value) == "Pricing Record ID"
            ),
            None,
        )
        if header is None:
            return set(), []
        valid: set[str] = set()
        malformed: list[str] = []
        for row in range(header.row + 1, worksheet.max_row + 1):
            value = text(worksheet.cell(row=row, column=header.column).value)
            if not value:
                continue
            if PRICING_ID_PATTERN.fullmatch(value):
                if value in valid:
                    raise PricingValidationError(
                        f"Duplicate canonical Pricing Record ID: {value}"
                    )
                valid.add(value)
            else:
                malformed.append(value)
        return valid, malformed
    finally:
        workbook.close()


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Return records from a named Excel Table."""
    if table_name not in worksheet.tables:
        raise PricingValidationError(f"Required table missing: {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        worksheet.tables[table_name].ref
    )
    rows = list(
        worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = [text(value) for value in rows[0]]
    return [
        dict(zip(headers, row, strict=False))
        for row in rows[1:]
        if any(value is not None and text(value) != "" for value in row)
    ]


def table_headers(worksheet: Worksheet, table_name: str) -> list[str]:
    """Return the exact header sequence from a named table."""
    min_col, min_row, max_col, _ = range_boundaries(
        worksheet.tables[table_name].ref
    )
    return [
        text(worksheet.cell(row=min_row, column=column).value)
        for column in range(min_col, max_col + 1)
    ]


def validate_structure(workbook: Any) -> None:
    """Validate worksheet, table, filter, freeze, and schema contracts."""
    if workbook.sheetnames != SHEET_NAMES:
        raise PricingValidationError(
            f"Worksheet order mismatch: expected {SHEET_NAMES}; got {workbook.sheetnames}"
        )
    if len(set(workbook.sheetnames)) != len(workbook.sheetnames) or any(
        len(name) > 31 for name in workbook.sheetnames
    ):
        raise PricingValidationError("Worksheet titles must be unique and <= 31 chars")
    table_names: list[str] = []
    for name in SHEET_NAMES:
        worksheet = workbook[name]
        expected = TABLE_NAMES[name]
        if expected not in worksheet.tables:
            raise PricingValidationError(f"{name} lacks table {expected}")
        if worksheet.freeze_panes != "A2" or not worksheet.auto_filter.ref:
            raise PricingValidationError(f"{name} lacks frozen header or filter")
        if worksheet.merged_cells.ranges:
            raise PricingValidationError(f"{name} contains merged cells")
        table_names.extend(worksheet.tables)
    if len(table_names) != len(set(table_names)):
        raise PricingValidationError("Excel Table names are not unique")
    if table_headers(
        workbook["01 - Pricing Records"],
        "tblMasterPricingCatalog",
    ) != PRICING_HEADERS:
        raise PricingValidationError("Pricing Records schema is not the exact 47 columns")


def validate_names_and_data_validations(workbook: Any) -> None:
    """Require governed validation to use workbook-defined names only."""
    required_names = set(DEFINED_NAME_BY_HEADER.values())
    actual_names = set(workbook.defined_names)
    missing = required_names - actual_names
    if missing:
        raise PricingValidationError(f"Missing defined names: {sorted(missing)}")
    for name in required_names:
        defined = workbook.defined_names[name]
        if defined.localSheetId is not None:
            raise PricingValidationError(f"Defined name is not workbook scoped: {name}")
        destinations = list(defined.destinations)
        if len(destinations) != 1:
            raise PricingValidationError(f"Defined name has invalid target: {name}")
        sheet_name, _coordinate = destinations[0]
        if sheet_name not in SHEET_NAMES:
            raise PricingValidationError(f"Defined name {name} targets {sheet_name}")
    worksheet = workbook["01 - Pricing Records"]
    validations = [
        item
        for item in worksheet.data_validations.dataValidation
        if item.type == "list"
    ]
    formulas = {text(item.formula1) for item in validations}
    expected = {f"={name}" for name in required_names}
    if formulas != expected:
        raise PricingValidationError(
            f"List validation formulas differ: expected {sorted(expected)}; "
            f"got {sorted(formulas)}"
        )
    if any("!" in formula or "," in formula for formula in formulas):
        raise PricingValidationError("Direct cross-sheet or hard-coded list validation found")


def validate_identity(
    records: Sequence[dict[str, Any]],
    services: Sequence[dict[str, Any]],
    existing_ids: set[str],
) -> None:
    """Validate pricing/service identities, counts, ordering, and sequence."""
    if len(records) != EXPECTED_ROWS or len(services) != EXPECTED_ROWS:
        raise PricingValidationError(
            f"Expected {EXPECTED_ROWS} pricing and service rows; "
            f"found {len(records)} and {len(services)}"
        )
    pricing_ids = [text(row["Pricing Record ID"]) for row in records]
    if any(not PRICING_ID_PATTERN.fullmatch(value) for value in pricing_ids):
        raise PricingValidationError("One or more Pricing Record IDs are invalid")
    if len(set(pricing_ids)) != len(pricing_ids):
        raise PricingValidationError("Duplicate Pricing Record ID")
    if existing_ids.intersection(pricing_ids):
        raise PricingValidationError("Generated Pricing Record ID reuses canonical ID")
    numbers = [int(value[3:]) for value in pricing_ids]
    expected_first = (
        max(int(value[3:]) for value in existing_ids) + 1
        if existing_ids
        else 1
    )
    if numbers != list(range(expected_first, expected_first + len(records))):
        raise PricingValidationError("Pricing Record IDs are not the next continuous sequence")
    if not existing_ids and pricing_ids[0] != "PRC000001":
        raise PricingValidationError(
            "ADR-007 empty namespace must begin at PRC000001"
        )
    expected_final = expected_first + len(records) - 1
    if numbers[-1] != expected_final:
        raise PricingValidationError(
            "Final Pricing Record ID does not reconcile to generated row count"
        )
    if (
        not existing_ids
        and len(records) == EXPECTED_ROWS
        and (pricing_ids[0], pricing_ids[-1]) != ("PRC000001", "PRC000314")
    ):
        raise PricingValidationError(
            "Current empty-namespace range must be PRC000001 through PRC000314"
        )
    output_service_ids = [text(row["Service ID"]) for row in records]
    source_service_ids = [text(row["Service ID"]) for row in services]
    if output_service_ids != source_service_ids:
        raise PricingValidationError("Service IDs are missing, extra, or reordered")
    if len(set(output_service_ids)) != len(output_service_ids):
        raise PricingValidationError("Duplicate Service ID in Pricing Records")
    if any(not SERVICE_ID_PATTERN.fullmatch(value) for value in output_service_ids):
        raise PricingValidationError("Invalid Service ID in Pricing Records")
    source_numbers = [text(row["Source Record Number"]) for row in records]
    if "" in source_numbers or len(set(source_numbers)) != len(source_numbers):
        raise PricingValidationError("Source Record Numbers are blank or duplicated")
    try:
        numeric_source_numbers = [int(value) for value in source_numbers]
    except ValueError as exc:
        raise PricingValidationError(
            "Source Record Numbers must be integer values"
        ) from exc
    if numeric_source_numbers != sorted(numeric_source_numbers):
        raise PricingValidationError(
            "Pricing Records are not in ascending source-record order"
        )


def validate_source_preservation(
    records: Sequence[dict[str, Any]],
    services: Sequence[dict[str, Any]],
    labor: Sequence[dict[str, Any]],
    exclusions: Sequence[dict[str, Any]],
) -> None:
    """Validate protected observations, relationships, and status routing."""
    labor_by_id = {
        text(row["Labor ID"]): row for row in labor if text(row.get("Labor ID"))
    }
    if len(labor_by_id) != sum(1 for row in labor if text(row.get("Labor ID"))):
        raise PricingValidationError("Duplicate Labor ID in protected labor catalog")
    excluded = {
        text(row.get("Excluded Source Row Number") or row.get("Source Row Number"))
        for row in exclusions
    }
    source_fields = {
        "Service ID": "Service ID",
        "Legacy Service SKU": "Legacy Service SKU",
        "Service Name": "Service Name",
        "Manufacturer ID": "Manufacturer ID",
        "Manufacturer Name": "Manufacturer Name",
        "Device Family Code": "Device Family Code",
        "Device Family Name": "Device Family Name",
        "Legacy Pricing Status": "Pricing Status",
        "Legacy Retail Price": "Legacy Retail Price",
        "Legacy Cost": "Legacy Cost",
        "Labor Standard ID": "Labor Standard ID",
        "Standard Labor Minutes": "Standard Minutes",
        "Labor Rate Tier": "Labor Tier",
        "Source Record Number": "Source Record Number",
        "Source Workbook": "Source Workbook",
        "Source Worksheet": "Source Worksheet",
        "Created At": "Created At",
        "Updated At": "Updated At",
    }
    for row_number, (record, service) in enumerate(
        zip(records, services, strict=True),
        2,
    ):
        service_id = text(record["Service ID"])
        for output_field, source_field in source_fields.items():
            if not persisted_equal(service[source_field], record[output_field]):
                raise PricingValidationError(
                    f"Row {row_number} {service_id} changed {output_field}: "
                    f"{service[source_field]!r} -> {record[output_field]!r}"
                )
        if text(record["Source Record Number"]) in excluded:
            raise PricingValidationError(
                f"{service_id} reintroduces an excluded source row"
            )
        labor_id = text(record["Labor Standard ID"])
        expected_status = (
            "Labor Mapping Required" if not labor_id else "Cost Research Required"
        )
        if text(record["Pricing Status"]) != expected_status:
            raise PricingValidationError(
                f"{service_id} expected Pricing Status {expected_status!r}"
            )
        if labor_id:
            if labor_id not in labor_by_id:
                raise PricingValidationError(f"{service_id} has invented Labor ID {labor_id}")
            source_labor = labor_by_id[labor_id]
            if not persisted_equal(
                source_labor["Standard Minutes"],
                record["Standard Labor Minutes"],
            ) or text(source_labor["Labor Rate Tier"]) != text(record["Labor Rate Tier"]):
                raise PricingValidationError(
                    f"{service_id} labor attributes differ from {labor_id}"
                )
        if text(record["Pricing Method"]) != "Not Yet Determined":
            raise PricingValidationError(f"{service_id} has nondefault Pricing Method")
        if text(record["Currency"]):
            raise PricingValidationError(f"{service_id} defaults an unconfirmed currency")
        if text(record["Review Status"]) != "Pending Review":
            raise PricingValidationError(f"{service_id} was pre-reviewed")
        if text(record["Pricing Confidence"]) != "Unassessed":
            raise PricingValidationError(f"{service_id} was assigned confidence")
        if text(record["Import Batch ID"]) != IMPORT_BATCH_ID:
            raise PricingValidationError(f"{service_id} has wrong Import Batch ID")
        for field in GENERATED_BLANK_FIELDS:
            if text(record[field]):
                raise PricingValidationError(
                    f"{service_id} generated prohibited/unresolved {field}: "
                    f"{record[field]!r}"
                )
        for field in MONEY_FIELDS:
            value = decimal_value(
                record[field],
                field,
                allow_text=field.startswith("Legacy"),
            )
            if value is not None and value < 0:
                raise PricingValidationError(
                    f"{service_id} has negative {field}: {record[field]!r}"
                )
        if text(record["Pricing Status"]) not in PRICING_STATUSES:
            raise PricingValidationError(f"{service_id} has uncontrolled Pricing Status")
        if text(record["Pricing Method"]) not in PRICING_METHODS:
            raise PricingValidationError(f"{service_id} has uncontrolled Pricing Method")
        if text(record["Review Status"]) not in REVIEW_STATUSES:
            raise PricingValidationError(f"{service_id} has uncontrolled Review Status")
        if text(record["Pricing Confidence"]) not in CONFIDENCE_VALUES:
            raise PricingValidationError(f"{service_id} has uncontrolled confidence")
        if any(
            isinstance(record[field], str) and record[field].startswith("=")
            for field in MONEY_FIELDS
        ):
            raise PricingValidationError(f"{service_id} contains a pricing formula")


def validate_review_queue(
    workbook: Any,
    records: Sequence[dict[str, Any]],
) -> None:
    """Require one unresolved queue row per pricing record."""
    queue = table_records(workbook["09 - Review Queue"], "tblPricingReviewQueue")
    if len(queue) != len(records):
        raise PricingValidationError("Review Queue does not contain one row per record")
    for record, item in zip(records, queue, strict=True):
        for field in ("Pricing Record ID", "Service ID", "Service Name", "Pricing Status"):
            if text(record[field]) != text(item[field]):
                raise PricingValidationError(
                    f"Review Queue changed {field} for {record['Service ID']}"
                )
        if not text(item["Missing Inputs"]) or not text(item["Required Action"]):
            raise PricingValidationError(
                f"Review Queue lacks routing for {record['Service ID']}"
            )
        if text(item["Review Status"]) != "Pending Review" or text(item["Reviewer Notes"]):
            raise PricingValidationError(
                f"Review Queue pre-resolved {record['Service ID']}"
            )


def validate_metadata(
    workbook: Any,
    hashes: dict[Path, str],
    existing_ids: set[str],
    records: Sequence[dict[str, Any]],
) -> None:
    """Validate embedded provenance and protected hashes."""
    rows = table_records(
        workbook["12 - Import Metadata"],
        "tblMasterPricingImportMetadata",
    )
    metadata = {text(row["Metadata Field"]): text(row["Value"]) for row in rows}
    if metadata.get("Import Batch ID") != IMPORT_BATCH_ID:
        raise PricingValidationError("Import Metadata has wrong batch ID")
    if metadata.get("Schema Columns") != str(len(PRICING_HEADERS)):
        raise PricingValidationError("Import Metadata has wrong schema count")
    if metadata.get("Namespace Authority") != NAMESPACE_AUTHORITY:
        raise PricingValidationError("Import Metadata lacks ADR-007 authority")
    if metadata.get("Existing Pricing Record Count") != str(len(existing_ids)):
        raise PricingValidationError("Existing Pricing Record Count is incorrect")
    if metadata.get("Generated Pricing Record Count") != str(len(records)):
        raise PricingValidationError("Generated Pricing Record Count is incorrect")
    if metadata.get("First Generated Pricing Record ID") != text(
        records[0]["Pricing Record ID"]
    ):
        raise PricingValidationError("First generated Pricing Record ID metadata differs")
    if metadata.get("Final Generated Pricing Record ID") != text(
        records[-1]["Pricing Record ID"]
    ):
        raise PricingValidationError("Final generated Pricing Record ID metadata differs")
    for path, digest in hashes.items():
        if metadata.get(f"Protected Input Path: {path.name}") != str(path):
            raise PricingValidationError(f"Missing protected path metadata for {path}")
        if metadata.get(f"SHA-256: {path.name}") != digest:
            raise PricingValidationError(f"Protected hash mismatch for {path}")


def main() -> int:
    """Run independent validation without writing any workbook."""
    protected = (
        MASTER_SERVICES_PATH,
        LABOR_CATALOG_PATH,
        PROPOSAL_PATH,
        CANONICAL_PATH,
    )
    try:
        assert_excel_safe_value_contract()
        require_files(protected)
        require_generated_workbook(OUTPUT_PATH)
        hashes = {path: file_hash(path) for path in protected}
        existing_ids, malformed_ids = read_existing_pricing_ids()
        services = read_records(MASTER_SERVICES_PATH, MASTER_SERVICES_SHEET)
        labor = read_records(LABOR_CATALOG_PATH, LABOR_SHEET)
        exclusions = read_records(PROPOSAL_PATH, EXCLUSIONS_SHEET)
        workbook = load_workbook(OUTPUT_PATH, data_only=False)
        try:
            validate_structure(workbook)
            validate_names_and_data_validations(workbook)
            records = table_records(
                workbook["01 - Pricing Records"],
                "tblMasterPricingCatalog",
            )
            validate_identity(records, services, existing_ids)
            validate_source_preservation(records, services, labor, exclusions)
            validate_review_queue(workbook, records)
            validate_metadata(workbook, hashes, existing_ids, records)
        finally:
            workbook.close()
        after_hashes = {path: file_hash(path) for path in protected}
        if after_hashes != hashes:
            raise PricingValidationError("A protected input changed during validation")
    except (
        AssertionError,
        OSError,
        TypeError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        PricingValidationError,
    ) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Validated: {OUTPUT_PATH}")
    print(f"Namespace authority: {NAMESPACE_AUTHORITY}")
    print(
        "Pricing Record ID range: "
        f"{records[0]['Pricing Record ID']} through "
        f"{records[-1]['Pricing Record ID']}"
    )
    print(f"Pricing records: {len(records)}")
    print(f"Existing malformed Pricing IDs reported: {len(malformed_ids)}")
    print("Workbook structure and tables: PASS")
    print("Defined names and list validations: PASS")
    print("Source relationships and provenance: PASS")
    print("Monetary safeguards and blank/zero rules: PASS")
    print("Protected input hashes: PASS")
    print("Final pricing approval/import: NOT PERFORMED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
