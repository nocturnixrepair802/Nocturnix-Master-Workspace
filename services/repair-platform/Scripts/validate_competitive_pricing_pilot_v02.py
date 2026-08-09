"""Read-only cross-source and readiness validator for Pilot v0.2 Draft."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path

from generate_competitive_pricing_pilot_v02 import (
    CANONICAL_SHA256,
    QA_PROFILE_VERSION,
    V01_SHA256,
    headers,
    table_rows,
    validate_v02,
)
from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--parent", type=Path, required=True)
    parser.add_argument("--canonical", type=Path, required=True)
    parser.add_argument("--markdown", type=Path, required=True)
    parser.add_argument("--json", dest="json_path", type=Path, required=True)
    args = parser.parse_args()
    before = sha256(args.workbook)
    errors = validate_v02(args.workbook)
    parent_hash = sha256(args.parent)
    canonical_hash = sha256(args.canonical)
    if parent_hash != V01_SHA256:
        errors.append("Locked parent v0.1 SHA-256 mismatch")
    if canonical_hash != CANONICAL_SHA256:
        errors.append("Canonical source SHA-256 mismatch")
    workbook = load_workbook(
        args.workbook, read_only=False, data_only=False, keep_links=True
    )
    try:
        pilot = table_rows(workbook["01 - Pilot Pricing Records"])
        supplier = table_rows(workbook["03 - Supplier Cost Evidence"])
        labor = table_rows(workbook["04 - Labor References"])
        policies = table_rows(workbook["06 - Pricing Policy"])
        calculations = table_rows(workbook["07 - Calculation Results"])
        observations = table_rows(workbook["02 - Competitor Observations"])
        reviews = table_rows(workbook["08 - Review Queue"])
        gates = table_rows(workbook["09 - Validation Summary"])
        canonical = table_rows(workbook["05 - Canonical References"])
        canonical_book = load_workbook(
            args.canonical, read_only=True, data_only=False
        )
        try:
            source_sheet = canonical_book["03 - Service Normalization"]
            source_headers = {
                cell.value: cell.column
                for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))
            }
            source_map = {
                source_sheet.cell(row, source_headers["Service ID"]).value: (
                    source_sheet.cell(
                        row,
                        source_headers["Proposed Canonical Service Type ID"],
                    ).value,
                    source_sheet.cell(
                        row,
                        source_headers["Proposed Canonical Service Type"],
                    ).value,
                )
                for row in range(2, source_sheet.max_row + 1)
            }
        finally:
            canonical_book.close()
        target_map = {
            row["Service ID"]: (
                row["Canonical Service Type ID"],
                row["Canonical Service Type"],
            )
            for row in canonical
        }
        mapping_mismatches = [
            row["Service ID"]
            for row in pilot
            if target_map.get(row["Service ID"]) != source_map.get(row["Service ID"])
        ]
        if mapping_mismatches:
            errors.append(f"Canonical mapping mismatches: {mapping_mismatches}")
        categories = {row["Gate Category"]: row for row in gates}
        readiness = {
            "structural_validity": categories["Structural Validity"]["Gate Status"],
            "identity_readiness": categories["Identity Readiness"]["Gate Status"],
            "policy_readiness": categories["Policy Readiness"]["Gate Status"],
            "supplier_cost_evidence_readiness": categories[
                "Supplier-Cost Evidence Readiness"
            ]["Gate Status"],
            "labor_readiness": categories["Labor Readiness"]["Gate Status"],
            "competitor_evidence_readiness": categories[
                "Competitor-Evidence Readiness"
            ]["Gate Status"],
            "calculation_readiness": categories["Calculation Readiness"][
                "Gate Status"
            ],
            "review_readiness": categories["Review Readiness"]["Gate Status"],
            "final_approval_readiness": categories["Final-Approval Readiness"][
                "Gate Status"
            ],
            "production_activation_readiness": categories[
                "Production-Activation Readiness"
            ]["Gate Status"],
        }
        counts = {
            "pilot_records": len(pilot),
            "unique_pricing_record_ids": len(
                {row["Pricing Record ID"] for row in pilot}
            ),
            "unique_service_ids": len({row["Service ID"] for row in pilot}),
            "canonical_mappings_valid": len(pilot) - len(mapping_mismatches),
            "supplier_status_rows": len(supplier),
            "legacy_supplier_references_preserved": sum(
                row["Evidence Record Status"] == "Legacy Reference Preserved"
                for row in supplier
            ),
            "no_evidence_captured_rows": sum(
                row["Evidence Record Status"] == "No Evidence Captured"
                for row in supplier
            ),
            "complete_supplier_evidence_packages": sum(
                row["Evidence Package Status"] == "Complete" for row in supplier
            ),
            "governed_labor_mappings": sum(
                row["Labor Readiness Status"] == "Ready" for row in labor
            ),
            "approved_policies": sum(
                row["Approval Status"] == "Approved" for row in policies
            ),
            "pending_policies": sum(
                row["Approval Status"] == "Pending Approval" for row in policies
            ),
            "competitor_template_rows": sum(
                row["Record Role"] == "Template - Do Not Count"
                for row in observations
            ),
            "competitor_evidence_rows": sum(
                row["Record Role"] == "Evidence" for row in observations
            ),
            "open_review_items": sum(
                row["Current Status"] != "Resolved" for row in reviews
            ),
            "disabled_calculation_rows": sum(
                row["Calculation Status"] == "Disabled" for row in calculations
            ),
            "blocked_calculation_gates": sum(
                row["Gate Status"] == "Blocked" for row in calculations
            ),
            "final_approved_prices_populated": sum(
                row["Final Approved Price"] not in (None, "")
                for row in calculations
            ),
        }
        table_ranges = {
            ws.title: {
                table_name: ws.tables[table_name].ref
                for table_name in ws.tables
            }
            for ws in workbook.worksheets
        }
        row_counts = {
            ws.title: sum(
                1
                for row in ws.iter_rows(min_row=2, values_only=True)
                if any(value is not None for value in row)
            )
            for ws in workbook.worksheets
        }
        worksheet_headers = {
            ws.title: [cell.value for cell in ws[1] if cell.value]
            for ws in workbook.worksheets
        }
        data_validations = {
            ws.title: [
                {
                    "type": validation.type,
                    "ranges": str(validation.sqref),
                    "formula1": validation.formula1,
                }
                for validation in ws.data_validations.dataValidation
            ]
            for ws in workbook.worksheets
        }
        protections = {
            ws.title: {
                "sheet_protected": bool(ws.protection.sheet),
                "auto_filter_allowed": bool(ws.protection.autoFilter),
                "sort_allowed": bool(ws.protection.sort),
            }
            for ws in workbook.worksheets
        }
        calc_ws = workbook["07 - Calculation Results"]
        calc_headers = headers(calc_ws)
        formula_fields = [
            "Labor Cost",
            "Total Internal Cost",
            "Minimum Profitable Price",
            "Target-Margin Price",
            "Verified Observation Count",
            "Competitor Low",
            "Competitor Median",
            "Competitor Average",
            "Profit Dollars",
            "Gross Margin Percentage",
        ]
        formula_validation = {
            field: {
                "formula_cells": sum(
                    isinstance(
                        calc_ws.cell(row, calc_headers[field]).value, str
                    )
                    and calc_ws.cell(
                        row, calc_headers[field]
                    ).value.startswith("=")
                    for row in range(2, 27)
                ),
                "gate_wrapped_cells": sum(
                    "$AF" in str(
                        calc_ws.cell(row, calc_headers[field]).value
                    )
                    for row in range(2, 27)
                ),
            }
            for field in formula_fields
        }
        final_column = calc_headers["Final Approved Price"]
        protections["07 - Calculation Results"][
            "final_approved_price_locked_cells"
        ] = sum(
            calc_ws.cell(row, final_column).protection.locked
            for row in range(2, 27)
        )
        protections["07 - Calculation Results"][
            "final_approved_price_formula_cells"
        ] = sum(
            isinstance(calc_ws.cell(row, final_column).value, str)
            and calc_ws.cell(row, final_column).value.startswith("=")
            for row in range(2, 27)
        )
        number_formats = {
            "01 - Pilot Pricing Records.Identity Verified At": workbook[
                "01 - Pilot Pricing Records"
            ].cell(2, headers(workbook["01 - Pilot Pricing Records"])["Identity Verified At"]).number_format,
            "02 - Competitor Observations.Verification Date": workbook[
                "02 - Competitor Observations"
            ].cell(2, headers(workbook["02 - Competitor Observations"])["Verification Date"]).number_format,
            "03 - Supplier Cost Evidence.Part Cost": workbook[
                "03 - Supplier Cost Evidence"
            ].cell(2, headers(workbook["03 - Supplier Cost Evidence"])["Part Cost"]).number_format,
            "03 - Supplier Cost Evidence.Verified At": workbook[
                "03 - Supplier Cost Evidence"
            ].cell(2, headers(workbook["03 - Supplier Cost Evidence"])["Verified At"]).number_format,
            "04 - Labor References.Approval Date": workbook[
                "04 - Labor References"
            ].cell(2, headers(workbook["04 - Labor References"])["Approval Date"]).number_format,
            "07 - Calculation Results.Final Approved Price": calc_ws.cell(
                2, final_column
            ).number_format,
            "08 - Review Queue.Due Date": workbook["08 - Review Queue"].cell(
                2, headers(workbook["08 - Review Queue"])["Due Date"]
            ).number_format,
        }
        supplier_missing_ids = [
            row["Pricing Record ID"]
            for row in supplier
            if row["Evidence Record Status"] == "No Evidence Captured"
        ]
        validation_gates = {
            row["Gate Category"]: {
                "required": row["Required State"],
                "actual": row["Actual State"],
                "status": row["Gate Status"],
                "blocking_count": row["Blocking Count"],
                "notes": row["Notes"],
            }
            for row in gates
        }
        revision_history = table_rows(workbook["10 - Revision History"])
        metadata_rows = table_rows(workbook["11 - Import Metadata"])
        metadata = {
            row["Metadata Field"]: row["Value"] for row in metadata_rows
        }
        banners = {
            ws.title: ws.oddHeader.center.text for ws in workbook.worksheets
        }
        audit = {
            "worksheet_order": workbook.sheetnames,
            "table_ranges": table_ranges,
            "row_counts": row_counts,
            "headers": worksheet_headers,
            "data_validations": data_validations,
            "protections": protections,
            "number_formats": number_formats,
            "formula_validation": formula_validation,
            "supplier_no_evidence_pricing_record_ids": supplier_missing_ids,
            "validation_gates": validation_gates,
            "revision_history": revision_history,
            "import_metadata": metadata,
            "banners": banners,
        }
    finally:
        workbook.close()
    after = sha256(args.workbook)
    if before != after:
        errors.append("Read-only validator changed the workbook")
    result = {
        "profile": "Nocturnix Competitive Pricing Pilot v0.2 Draft cross-source/readiness QA",
        "qa_profile_version": QA_PROFILE_VERSION,
        "validated_at_utc": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "workbook": str(args.workbook.resolve()),
        "workbook_sha256_before": before,
        "workbook_sha256_after": after,
        "validator_read_only": before == after,
        "parent_v01_sha256": parent_hash,
        "canonical_sha256": canonical_hash,
        "status": "PASS" if not errors else "FAIL",
        "classification": "STRUCTURALLY VALIDATED / CALCULATION-BLOCKED / NON-PRODUCTION",
        "errors": errors,
        "readiness": readiness,
        "counts": counts,
        "audit": audit,
        "unresolved_blocking_items": [
            "25 governed manufacturer/device identity packages",
            "13 approved and versioned pricing policies",
            "25 complete verified supplier-cost evidence packages",
            "25 governed and approved labor mappings",
            "Approved competitor research policy and verified evidence dataset",
            "125 review-queue items",
            "Separate final-price approval workflow",
            "Separate production-activation authorization",
        ],
    }
    args.json_path.parent.mkdir(parents=True, exist_ok=True)
    args.markdown.parent.mkdir(parents=True, exist_ok=True)
    args.json_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    lines = [
        "# Nocturnix Competitive Pricing Pilot v0.2 Draft Readiness Report",
        "",
        f"Status: **{result['status']}**",
        "",
        f"Classification: **{result['classification']}**",
        "",
        f"Workbook SHA-256 before validation: `{before}`",
        "",
        f"Workbook SHA-256 after validation: `{after}`",
        "",
        f"Read-only proof: **{'PASS' if before == after else 'FAIL'}**",
        "",
        "## Readiness gates",
        "",
        "| Gate | Status |",
        "|---|---|",
        *[
            f"| {name.replace('_', ' ').title()} | {status} |"
            for name, status in readiness.items()
        ],
        "",
        "## Counts",
        "",
        *[f"- {name.replace('_', ' ').title()}: {value}" for name, value in counts.items()],
        "",
        "## Unresolved blocking items",
        "",
        *[f"- {item}" for item in result["unresolved_blocking_items"]],
        "",
        "## Validation errors",
        "",
        *(["None."] if not errors else [f"- {error}" for error in errors]),
    ]
    args.markdown.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "readiness": readiness, "counts": counts}, indent=2))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
