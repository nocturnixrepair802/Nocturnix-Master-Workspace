"""Generate the governed Service Type normalization and labor review package.

The generator reads protected workbooks without saving them, writes a standalone
review artifact through a temporary sibling file, and never approves a mapping.
Fuzzy similarity is deliberately not used as mapping evidence.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import zipfile
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKING_DIR = Path(r"D:\Business Portal\300_Pricing\Working")
CANONICAL_PATH = PROJECT_ROOT / "Data" / "Nocturnix_Master_Database.xlsm"
MASTER_SERVICES_PATH = WORKING_DIR / "Nocturnix_Master_Services_Catalog_v1.xlsx"
MASTER_LABOR_PATH = WORKING_DIR / "Nocturnix_Master_Labor_Catalog_v1.xlsx"
LEGACY_MAPPING_PATH = WORKING_DIR / "Labor_Mapping_Review_v1.xlsx"
OUTPUT_PATH = (
    WORKING_DIR / "Nocturnix_Service_Type_Normalization_Review_v1.xlsx"
)
TEMP_OUTPUT_PATH = OUTPUT_PATH.with_name(f".{OUTPUT_PATH.stem}.tmp.xlsx")

CANONICAL_SHEET = "33 Service Types"
SERVICES_SHEET = "01 - Master Services"
LABOR_SHEET = "01 - Labor Standards"
CANONICAL_HEADER_ROW = 4
CANONICAL_MIN_COLUMN = 12
CANONICAL_MAX_COLUMN = 20
EXPECTED_CANONICAL_ROWS = 70
EXPECTED_SERVICE_ROWS = 314
EXPECTED_LABOR_ROWS = 265
IMPORT_BATCH = "SERVICE-TYPE-NORMALIZATION-V1.0.2-REVIEW"

SERVICE_TYPE_ID_PATTERN = re.compile(r"^STY\d{6}$")
SERVICE_ID_PATTERN = re.compile(r"^SVC\d{6}$")
LABOR_ID_PATTERN = re.compile(r"^LAB\d{6}$")

SHEET_NAMES = [
    "00 - Instructions",
    "01 - Canonical Service Types",
    "02 - Service Type Aliases",
    "03 - Service Normalization",
    "04 - Labor Normalization",
    "05 - Service Labor Candidates",
    "06 - Unresolved Review",
    "07 - Validation Summary",
    "08 - Revision History",
    "09 - Import Metadata",
]
TABLE_NAMES = {
    "00 - Instructions": "tblSTNInstructions",
    "01 - Canonical Service Types": "tblCanonicalServiceTypes",
    "02 - Service Type Aliases": "tblServiceTypeAliases",
    "03 - Service Normalization": "tblServiceNormalization",
    "04 - Labor Normalization": "tblLaborNormalization",
    "05 - Service Labor Candidates": "tblServiceLaborCandidates",
    "06 - Unresolved Review": "tblSTNUnresolvedReview",
    "07 - Validation Summary": "tblSTNValidationSummary",
    "08 - Revision History": "tblSTNRevisionHistory",
    "09 - Import Metadata": "tblSTNImportMetadata",
}

CANONICAL_SOURCE_HEADERS = [
    "Service Category",
    "Service Type",
    "Service Description",
    "Applies To",
    "Estimated Time (Min)",
    "Default Warranty (Days)",
    "Taxable",
    "Active",
    "Internal Notes",
]
CANONICAL_HEADERS = [
    "Proposed Canonical Service Type ID",
    *CANONICAL_SOURCE_HEADERS,
    "Identity Authority",
    "Review Status",
    "Reviewer Notes",
]
ALIAS_HEADERS = [
    "Alias ID",
    "Source System",
    "Source Field",
    "Source Value",
    "Normalized Source Value",
    "Proposed Canonical Service Type ID",
    "Proposed Canonical Service Type",
    "Alias Rule Type",
    "Evidence",
    "Confidence",
    "Review Status",
    "Reviewer",
    "Reviewer Notes",
]
SERVICE_HEADERS = [
    "Service ID",
    "Service Name",
    "Current Repair Type ID",
    "Current Repair Type",
    "Manufacturer ID",
    "Manufacturer Name",
    "Device Family Code",
    "Device Family Name",
    "Proposed Canonical Service Type ID",
    "Proposed Canonical Service Type",
    "Mapping Method",
    "Mapping Evidence",
    "Confidence",
    "Review Status",
    "Reviewer Notes",
]
LABOR_HEADERS = [
    "Labor Standard ID",
    "Legacy Labor ID",
    "Labor Name",
    "Current Repair Type",
    "Device Family",
    "Manufacturer",
    "Proposed Canonical Service Type ID",
    "Proposed Canonical Service Type",
    "Mapping Method",
    "Mapping Evidence",
    "Confidence",
    "Review Status",
    "Reviewer Notes",
]
CANDIDATE_HEADERS = [
    "Service ID",
    "Service Name",
    "Canonical Service Type ID",
    "Canonical Service Type",
    "Device Family Code",
    "Manufacturer ID",
    "Suggested Labor Standard ID",
    "Legacy Labor ID",
    "Labor Name",
    "Standard Minutes",
    "Minimum Minutes",
    "Maximum Minutes",
    "Candidate Method",
    "Evidence",
    "Confidence",
    "Ambiguity Count",
    "Review Status",
    "Reviewer Notes",
]
UNRESOLVED_HEADERS = [
    "Record Type",
    "Source Record ID",
    "Source Name",
    "Current Type",
    "Candidate Canonical Types",
    "Candidate Labor Standards",
    "Ambiguity Reason",
    "Missing Evidence",
    "Required Action",
    "Review Priority",
    "Review Status",
    "Reviewer Notes",
]

ALIAS_RULE_TYPES = [
    "Exact Match",
    "Approved Synonym Candidate",
    "Broader-to-Specific Review",
    "Device-Family-Specific",
    "Manufacturer-Specific",
    "No Safe Mapping",
]
MAPPING_METHODS = [
    "Exact Match",
    "Explicit Alias Candidate",
    "Service Name Rule Candidate",
    "Device-Family Constraint",
    "No Safe Mapping",
]
CONFIDENCE_VALUES = ["Unassessed", "Low", "Medium", "High"]
REVIEW_STATUSES = [
    "Pending Review",
    "Pending Evidence Review",
    "Pending Service Review",
    "Pending Labor Review",
    "Unresolved",
    "Ready for Approval",
    "Approved",
    "Rejected",
    "Archived",
]
PENDING_REVIEW_STATUSES = {
    "Pending Review",
    "Pending Evidence Review",
    "Pending Service Review",
    "Pending Labor Review",
    "Unresolved",
}
PROHIBITED_GENERATED_STATUSES = {
    "approved",
    "confirmed",
    "accepted",
    "ready for import",
}
STATUS_AUDIT_SPECS = {
    "00 - Instructions": (None, "Topic", "Status Vocabulary"),
    "01 - Canonical Service Types": (
        "Review Status",
        "Proposed Canonical Service Type ID",
        "Canonical Service Type",
    ),
    "02 - Service Type Aliases": ("Review Status", "Alias ID", "Service Type Alias"),
    "03 - Service Normalization": (
        "Review Status",
        "Service ID",
        "Service Normalization",
    ),
    "04 - Labor Normalization": (
        "Review Status",
        "Labor Standard ID",
        "Labor Normalization",
    ),
    "05 - Service Labor Candidates": (
        "Review Status",
        "Service ID",
        "Service Labor Candidate",
    ),
    "06 - Unresolved Review": (
        "Review Status",
        "Source Record ID",
        "Unresolved Review",
    ),
    "07 - Validation Summary": ("Result", "Validation", "Validation Summary"),
    "08 - Revision History": ("Status", "Version", "Revision History"),
    "09 - Import Metadata": (None, "Metadata Field", "Import Metadata"),
}
YES_NO_VALUES = ["Yes", "No"]
UNIVERSAL_FAMILIES = {"all", "all devices", "any", "universal"}
GOVERNED_ID_HEADER_KEYS = {
    "service type id",
    "canonical service type id",
    "governed service type id",
}
SERVICE_TYPE_HEADER_KEYS = {
    "service type",
    "canonical service type",
    "governed service type",
}

# Each entry is a review candidate, never an approval. A tuple of targets means
# the business meaning is ambiguous and the normalization row remains blank.
ALIAS_CANDIDATES: dict[str, tuple[tuple[str, ...], str, str]] = {
    "Screen Repair": (
        ("Screen Replacement",),
        "Approved Synonym Candidate",
        "Explicit ADR-012 candidate; equivalence requires review.",
    ),
    "iPhone Screen Repair": (
        ("Screen Replacement",),
        "Manufacturer-Specific",
        "Screen candidate constrained by the source Apple/iPhone context.",
    ),
    "Samsung Screen Repair": (
        ("Screen Replacement",),
        "Manufacturer-Specific",
        "Screen candidate constrained by the source Samsung context.",
    ),
    "Camera Replacement": (
        ("Rear Camera Replacement", "Front Camera Replacement"),
        "Broader-to-Specific Review",
        "Source does not identify front versus rear camera.",
    ),
    "Data Backup": (
        ("Backup Service",),
        "Approved Synonym Candidate",
        "Explicit ADR-012 synonym candidate.",
    ),
    "Operating System Installation": (
        ("Operating System Reinstallation",),
        "Approved Synonym Candidate",
        "Explicit ADR-012 candidate; installation and reinstallation may differ.",
    ),
    "Software Setup": (
        ("Device Setup",),
        "Approved Synonym Candidate",
        "Explicit ADR-012 candidate; scope requires review.",
    ),
    "Storage Replacement": (
        ("Storage Upgrade",),
        "Broader-to-Specific Review",
        "Replacement is not necessarily an upgrade.",
    ),
    "Water Damage Assessment": (
        ("Liquid Damage Inspection",),
        "Approved Synonym Candidate",
        "Explicit ADR-012 candidate; terminology requires review.",
    ),
    "Charging Port Cleaning": (
        (),
        "No Safe Mapping",
        "Review whether a new canonical Service Type is required.",
    ),
    "Console Cleaning": (
        ("Internal Cleaning",),
        "Device-Family-Specific",
        "Candidate only for a matching console Device Family.",
    ),
    "Controller Cleaning": (
        ("Internal Cleaning",),
        "Device-Family-Specific",
        "Candidate only for a matching controller Device Family.",
    ),
    "Laptop Screen Replacement": (
        ("Screen Replacement",),
        "Device-Family-Specific",
        "Candidate requires a Laptop Device Family constraint.",
    ),
    "Keyboard Replacement": (
        (),
        "No Safe Mapping",
        "Review whether a new canonical Service Type is required.",
    ),
    "Hinge Repair": (
        (),
        "No Safe Mapping",
        "Review whether a new canonical Service Type is required.",
    ),
    "Analog Stick Replacement": (
        (),
        "No Safe Mapping",
        "Review whether a new canonical Service Type is required.",
    ),
    "Joy-Con Repair": (
        (),
        "No Safe Mapping",
        "Review whether a new canonical Service Type is required.",
    ),
}

# Explicit token rules classify Miscellaneous Repair only as review candidates.
SERVICE_NAME_RULES: tuple[tuple[tuple[str, ...], str], ...] = (
    (("back", "glass"), "Back Glass Replacement"),
    (("battery",), "Battery Replacement"),
    (("charging", "port"), "Charging Port Replacement"),
    (("data", "transfer"), "Data Transfer"),
    (("diagnostic",), "Diagnostic"),
    (("fan",), "Fan Replacement"),
    (("hdmi", "port"), "HDMI Port Repair"),
    (("microphone",), "Microphone Replacement"),
    (("speaker",), "Speaker Replacement"),
    (("thermal", "paste"), "Thermal Paste Replacement"),
    (("screen",), "Screen Replacement"),
)

DEFINED_NAME_SPECS = {
    "DV_CanonicalServiceTypeIDs": ("01 - Canonical Service Types", 1),
    "DV_CanonicalServiceTypes": ("01 - Canonical Service Types", 3),
    "DV_AliasRuleTypes": ("00 - Instructions", 3),
    "DV_MappingMethods": ("00 - Instructions", 4),
    "DV_ConfidenceValues": ("00 - Instructions", 5),
    "DV_ReviewStatuses": ("00 - Instructions", 6),
    "DV_DeviceFamilyCodes": ("00 - Instructions", 7),
    "DV_ManufacturerIDs": ("00 - Instructions", 8),
    "DV_LaborStandardIDs": ("00 - Instructions", 9),
    "DV_YesNoValues": ("00 - Instructions", 10),
}
VALIDATION_NAMES_BY_HEADER = {
    "Proposed Canonical Service Type ID": "DV_CanonicalServiceTypeIDs",
    "Proposed Canonical Service Type": "DV_CanonicalServiceTypes",
    "Alias Rule Type": "DV_AliasRuleTypes",
    "Mapping Method": "DV_MappingMethods",
    "Confidence": "DV_ConfidenceValues",
    "Review Status": "DV_ReviewStatuses",
    "Device Family Code": "DV_DeviceFamilyCodes",
    "Manufacturer ID": "DV_ManufacturerIDs",
    "Labor Standard ID": "DV_LaborStandardIDs",
    "Suggested Labor Standard ID": "DV_LaborStandardIDs",
    "Taxable": "DV_YesNoValues",
    "Active": "DV_YesNoValues",
}


class NormalizationReviewError(RuntimeError):
    """Raised when a governed normalization invariant cannot be satisfied."""


def text(value: Any) -> str:
    """Return trimmed display text."""
    return "" if value is None else str(value).strip()


def normalized(value: Any) -> str:
    """Normalize text for exact comparison, never fuzzy approval."""
    return " ".join(re.findall(r"[a-z0-9]+", text(value).casefold()))


def generated_review_status(
    record_type: str,
    *,
    resolved: bool,
    ambiguous: bool = False,
) -> str:
    """Return a documented review-pending status for a generated record."""
    if ambiguous:
        return "Pending Evidence Review"
    if resolved:
        return "Pending Review"
    if record_type == "Service":
        return "Pending Service Review"
    if record_type == "Labor":
        return "Pending Labor Review"
    return "Unresolved"


def assert_generated_status_contract() -> None:
    """Exercise pending assignment and exact governed-field prohibition."""
    assert generated_review_status("Service", resolved=True) == "Pending Review"
    assert generated_review_status("Alias", resolved=True) == "Pending Review"
    assert generated_review_status(
        "Relationship", resolved=False, ambiguous=True
    ) in PENDING_REVIEW_STATUSES
    try:
        reject_prohibited_generated_statuses(
            {"03 - Service Normalization": [{"Service ID": "SVC000001", "Review Status": "Approved"}]}
        )
    except NormalizationReviewError as exc:
        assert "SVC000001" in text(exc)
        assert "Approved" in text(exc)
    else:
        raise AssertionError("Literal Approved generated Review Status was accepted")
    reject_prohibited_generated_statuses(
        {
            "09 - Import Metadata": [
                {"Metadata Field": "Approval Status", "Value": "Not Approved"}
            ]
        }
    )


def excel_value(value: Any) -> Any:
    """Return an OOXML-safe scalar without timezone-aware datetimes."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.astimezone(UTC).replace(tzinfo=None)
    return value


def file_hash(path: Path) -> str:
    """Return a streaming SHA-256 digest."""
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def protected_state(paths: Iterable[Path]) -> dict[Path, str]:
    """Hash present protected files and record intentionally absent files."""
    return {
        path: file_hash(path) if path.is_file() else "ABSENT"
        for path in paths
    }


def require_files(paths: Iterable[Path]) -> None:
    """Require every mandatory protected input."""
    missing = [str(path) for path in paths if not path.is_file()]
    if missing:
        raise NormalizationReviewError(
            f"Missing protected input: {', '.join(missing)}"
        )


def locate_header(
    worksheet: Worksheet,
    required: set[str],
) -> tuple[int, list[str]]:
    """Locate a source header row within the first 50 rows."""
    for row_number, headers in bounded_header_rows(worksheet):
        if required <= set(headers):
            return row_number, headers
    raise NormalizationReviewError(
        f"{worksheet.title} lacks required headers: {sorted(required)}"
    )


def bounded_header_rows(
    worksheet: Worksheet,
    *,
    scan_limit: int = 50,
) -> Iterable[tuple[int, list[str]]]:
    """Yield populated candidate header rows strictly within sheet bounds."""
    first_row = max(1, worksheet.min_row)
    last_row = min(worksheet.max_row, scan_limit)
    if first_row > last_row:
        return
    rows = worksheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=max(1, worksheet.min_column),
        max_col=max(1, worksheet.max_column),
    )
    for row_number, row in enumerate(rows, start=first_row):
        headers = [text(cell.value) for cell in row]
        if any(headers):
            yield row_number, headers


def sequential_records(
    worksheet: Worksheet,
    required: set[str],
    *,
    header_row: int | None = None,
    start_row: int | None = None,
    min_column: int | None = None,
    max_column: int | None = None,
    include_source_row: bool = False,
) -> list[dict[str, Any]]:
    """Return records from one bounded, sequential worksheet pass."""
    first_column = min_column or max(1, worksheet.min_column)
    last_column = max_column or max(1, worksheet.max_column)
    first_row = header_row or max(1, worksheet.min_row)
    rows = worksheet.iter_rows(
        min_row=first_row,
        max_row=worksheet.max_row,
        min_col=first_column,
        max_col=last_column,
        values_only=True,
    )
    records: list[dict[str, Any]] = []
    headers: list[str] | None = None
    header_indexes: list[tuple[int, str]] = []
    data_start = start_row
    for row_number, values_tuple in enumerate(rows, start=first_row):
        values = list(values_tuple)
        if headers is None:
            candidate_headers = [text(value) for value in values]
            is_header = (
                row_number == header_row
                if header_row is not None
                else row_number <= 50 and required <= set(candidate_headers)
            )
            if not is_header:
                continue
            if not required <= set(candidate_headers):
                raise NormalizationReviewError(
                    f"{worksheet.title} lacks required headers: {sorted(required)}"
                )
            headers = candidate_headers
            header_indexes = list(enumerate(headers))
            data_start = data_start or row_number + 1
            continue
        if row_number < data_start:
            continue
        if not any(text(value) for value in values):
            continue
        record = {header: values[index] for index, header in header_indexes}
        if include_source_row:
            record["_Source Row"] = row_number
        records.append(record)
    if headers is None:
        raise NormalizationReviewError(
            f"{worksheet.title} lacks required headers: {sorted(required)}"
        )
    return records


def read_records(
    path: Path,
    sheet_name: str,
    required: set[str],
    *,
    header_row: int | None = None,
    start_row: int | None = None,
    min_column: int | None = None,
    max_column: int | None = None,
) -> list[dict[str, Any]]:
    """Read populated records in one sequential pass without saving."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise NormalizationReviewError(f"{path.name} lacks {sheet_name}")
        return sequential_records(
            workbook[sheet_name],
            required,
            header_row=header_row,
            start_row=start_row,
            min_column=min_column,
            max_column=max_column,
            include_source_row=True,
        )
    finally:
        workbook.close()


def assert_sequential_read_contract() -> None:
    """Exercise bounded sequential source reading without source workbooks."""
    marker = datetime(2025, 1, 2, 3, 4, 5)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Source"
    worksheet.append(["preamble"])
    worksheet.append([])
    worksheet.append([None, None, "ID", "Value", "When"])
    worksheet.append([None, None, "001", 0, marker])
    worksheet.append([None, None, None, None, None])
    worksheet.append([None, None, "002", "", "text"])
    try:
        records = sequential_records(
            worksheet,
            {"ID", "Value", "When"},
            header_row=3,
            start_row=4,
            min_column=3,
            max_column=5,
        )
        assert records == [
            {"ID": "001", "Value": 0, "When": marker},
            {"ID": "002", "Value": "", "When": "text"},
        ]
        assert [record["ID"] for record in records] == ["001", "002"]
        assert records[0]["Value"] == 0
        assert records[1]["Value"] == ""
        assert records[0]["When"] is marker
        service_sheet = workbook.create_sheet("Services")
        service_sheet.append(["Service ID"])
        for index in range(1, EXPECTED_SERVICE_ROWS + 1):
            service_sheet.append([f"SVC{index:06d}"])
        labor_sheet = workbook.create_sheet("Labor")
        labor_sheet.append(["Labor Standard ID"])
        for index in range(1, EXPECTED_LABOR_ROWS + 1):
            labor_sheet.append([f"LAB{index:06d}"])
        service_records = sequential_records(service_sheet, {"Service ID"})
        labor_records = sequential_records(labor_sheet, {"Labor Standard ID"})
        assert len(service_records) == 314
        assert len(labor_records) == 265
        assert service_records[0]["Service ID"] == "SVC000001"
        assert service_records[-1]["Service ID"] == "SVC000314"
        assert labor_records[0]["Labor Standard ID"] == "LAB000001"
        assert labor_records[-1]["Labor Standard ID"] == "LAB000265"
    finally:
        workbook.close()


def governed_service_type_ids(
    workbook: Any,
) -> tuple[dict[str, str], list[str], int]:
    """Inspect explicit Service Type ID columns throughout the canonical file."""
    by_type: dict[str, str] = {}
    findings: list[str] = []
    seen_ids: set[str] = set()
    total_populated = 0
    for worksheet in workbook.worksheets:
        for row_number, headers in bounded_header_rows(worksheet):
            normalized_headers = [normalized(header) for header in headers]
            id_columns = [
                index + max(1, worksheet.min_column)
                for index, header in enumerate(normalized_headers)
                if header in GOVERNED_ID_HEADER_KEYS
            ]
            if not id_columns:
                continue
            if len(id_columns) > 1:
                raise NormalizationReviewError(
                    "Multiple explicit governed Service Type ID columns in "
                    f"{worksheet.title}!{row_number}"
                )
            id_column = id_columns[0]
            type_columns = [
                index + max(1, worksheet.min_column)
                for index, header in enumerate(normalized_headers)
                if header in SERVICE_TYPE_HEADER_KEYS
            ]
            type_column = type_columns[0] if len(type_columns) == 1 else None
            populated = 0
            first_data_row = max(row_number + 1, worksheet.min_row)
            selected_columns = [id_column]
            if type_column is not None:
                selected_columns.append(type_column)
            first_column = min(selected_columns)
            last_column = max(selected_columns)
            if first_data_row <= worksheet.max_row:
                data_rows = worksheet.iter_rows(
                    min_row=first_data_row,
                    max_row=worksheet.max_row,
                    min_col=first_column,
                    max_col=last_column,
                    values_only=True,
                )
            else:
                data_rows = ()
            for source_row, source_values in enumerate(
                data_rows,
                start=first_data_row,
            ):
                value = text(source_values[id_column - first_column])
                if not value:
                    continue
                populated += 1
                total_populated += 1
                if not SERVICE_TYPE_ID_PATTERN.fullmatch(value):
                    raise NormalizationReviewError(
                        "Malformed governed Service Type ID "
                        f"{value!r} in {worksheet.title}!{source_row}"
                    )
                if value in seen_ids:
                    raise NormalizationReviewError(
                        f"Duplicate governed Service Type ID: {value}"
                    )
                seen_ids.add(value)
                if type_column is not None:
                    service_type = text(source_values[type_column - first_column])
                    if service_type:
                        key = normalized(service_type)
                        if key in by_type:
                            raise NormalizationReviewError(
                                "Duplicate governed Service Type name: "
                                f"{service_type}"
                            )
                        by_type[key] = value
            findings.append(
                f"{worksheet.title}: explicit Service Type ID column, "
                f"{populated} populated value(s)"
            )
            break
    return by_type, findings, total_populated


def assert_governed_service_type_id_scan_contract() -> None:
    """Exercise bounded scanning and governed-namespace failure modes."""

    def scan_rows(rows: Sequence[Sequence[Any]]) -> tuple[
        dict[str, str],
        list[str],
        int,
    ]:
        workbook = Workbook()
        worksheet = workbook.active
        for row in rows:
            worksheet.append(list(row))
        try:
            return governed_service_type_ids(workbook)
        finally:
            workbook.close()

    assert scan_rows([["Only one populated row"]]) == ({}, [], 0)
    assert scan_rows([[], [], ["Unrelated ID", "Name"]]) == ({}, [], 0)
    blank_namespace = [
        ["Canonical Service Type ID", "Canonical Service Type"],
        *(["", f"Type {index}"] for index in range(1, 71)),
    ]
    blank_ids, blank_findings, blank_count = scan_rows(blank_namespace)
    assert blank_ids == {}
    assert blank_count == 0
    assert blank_findings and "0 populated" in blank_findings[0]
    complete_namespace = [
        ["Governed Service Type ID", "Governed Service Type"],
        *(
            [f"STY{index:06d}", f"Type {index}"]
            for index in range(1, 71)
        ),
    ]
    complete_ids, _, complete_count = scan_rows(complete_namespace)
    assert complete_count == 70
    assert len(complete_ids) == 70
    assert complete_ids["type 1"] == "STY000001"
    try:
        scan_rows(
            [
                ["Service Type ID", "Service Type"],
                ["INVALID", "Type 1"],
            ]
        )
    except NormalizationReviewError as exc:
        assert "Malformed governed Service Type ID" in text(exc)
    else:
        raise AssertionError("Malformed governed Service Type ID was accepted")
    try:
        scan_rows(
            [
                ["Service Type ID", "Service Type"],
                ["STY000001", "Type 1"],
                ["STY000001", "Type 2"],
            ]
        )
    except NormalizationReviewError as exc:
        assert "Duplicate governed Service Type ID" in text(exc)
    else:
        raise AssertionError("Duplicate governed Service Type ID was accepted")


def canonical_snapshot() -> tuple[list[dict[str, Any]], list[str], str]:
    """Read detailed L:T taxonomy and assign governed or review-local IDs."""
    workbook = load_workbook(
        CANONICAL_PATH,
        read_only=True,
        data_only=False,
        keep_vba=True,
    )
    try:
        if CANONICAL_SHEET not in workbook.sheetnames:
            raise NormalizationReviewError(
                f"Canonical workbook lacks {CANONICAL_SHEET}"
            )
        worksheet = workbook[CANONICAL_SHEET]
        if not worksheet.min_row <= CANONICAL_HEADER_ROW <= worksheet.max_row:
            raise NormalizationReviewError(
                f"{CANONICAL_SHEET} lacks canonical header row "
                f"{CANONICAL_HEADER_ROW}"
            )
        if (
            worksheet.min_column > CANONICAL_MIN_COLUMN
            or worksheet.max_column < CANONICAL_MAX_COLUMN
        ):
            raise NormalizationReviewError(
                f"{CANONICAL_SHEET} lacks canonical columns L:T"
            )
        rows = worksheet.iter_rows(
            min_row=CANONICAL_HEADER_ROW,
            max_row=worksheet.max_row,
            min_col=CANONICAL_MIN_COLUMN,
            max_col=CANONICAL_MAX_COLUMN,
            values_only=True,
        )
        header_values = next(rows)
        headers = [text(value) for value in header_values]
        if headers != CANONICAL_SOURCE_HEADERS:
            raise NormalizationReviewError(
                f"Detailed canonical headers differ: {headers}"
            )
        source_rows: list[dict[str, Any]] = []
        for values_tuple in rows:
            values = list(values_tuple)
            if not text(values[1]):
                continue
            source_rows.append(
                dict(zip(CANONICAL_SOURCE_HEADERS, values, strict=True))
            )
        id_by_type, findings, populated_id_count = governed_service_type_ids(
            workbook
        )
    finally:
        workbook.close()

    names = [normalized(record["Service Type"]) for record in source_rows]
    if len(names) != len(set(names)):
        raise NormalizationReviewError(
            "Detailed canonical Service Types are not unique after normalization"
        )
    if populated_id_count:
        missing = [
            text(record["Service Type"])
            for record in source_rows
            if normalized(record["Service Type"]) not in id_by_type
        ]
        if missing:
            raise NormalizationReviewError(
                "Governed Service Type namespace is incomplete for detailed "
                f"taxonomy: {', '.join(missing)}"
            )
        authority = "Existing governed canonical Service Type ID"
    else:
        id_by_type = {
            normalized(record["Service Type"]): f"STY{index:06d}"
            for index, record in enumerate(source_rows, start=1)
        }
        authority = "ADR-012 review-local identity"
        if not findings:
            findings.append(
                "No explicit Service Type ID column found in canonical workbook"
            )
        else:
            findings.append(
                "No populated governed Service Type IDs found; ADR-012 "
                "review-local namespace authorized"
            )

    result = []
    for record in source_rows:
        service_type = text(record["Service Type"])
        result.append(
            {
                "Proposed Canonical Service Type ID": id_by_type[
                    normalized(service_type)
                ],
                **record,
                "Identity Authority": authority,
                "Review Status": "Pending Review",
                "Reviewer Notes": "",
            }
        )
    return result, findings, authority


def canonical_indexes(
    canonical: Sequence[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    """Index canonical records by normalized type and proposed ID."""
    by_type = {
        normalized(record["Service Type"]): record for record in canonical
    }
    by_id = {
        text(record["Proposed Canonical Service Type ID"]): record
        for record in canonical
    }
    return by_type, by_id


def alias_rows(
    canonical_by_type: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Create the explicit governed alias candidate register."""
    rows: list[dict[str, Any]] = []
    for index, (source_value, spec) in enumerate(
        ALIAS_CANDIDATES.items(),
        start=1,
    ):
        targets, rule_type, evidence = spec
        resolved = [
            canonical_by_type[normalized(target)]
            for target in targets
            if normalized(target) in canonical_by_type
        ]
        unique_target = resolved[0] if len(targets) == 1 and len(resolved) == 1 else None
        missing_targets = [
            target
            for target in targets
            if normalized(target) not in canonical_by_type
        ]
        detail = evidence
        if len(targets) > 1:
            detail += f" Candidate types: {', '.join(targets)}."
        if missing_targets:
            detail += (
                " Runtime canonical target(s) not found: "
                f"{', '.join(missing_targets)}."
            )
        rows.append(
            {
                "Alias ID": f"STA{index:06d}",
                "Source System": "Governed Initial Candidate Register",
                "Source Field": "Repair Type",
                "Source Value": source_value,
                "Normalized Source Value": normalized(source_value),
                "Proposed Canonical Service Type ID": (
                    text(unique_target["Proposed Canonical Service Type ID"])
                    if unique_target
                    else ""
                ),
                "Proposed Canonical Service Type": (
                    text(unique_target["Service Type"]) if unique_target else ""
                ),
                "Alias Rule Type": rule_type,
                "Evidence": detail,
                "Confidence": "Medium" if unique_target else "Unassessed",
                "Review Status": "Pending Review",
                "Reviewer": "",
                "Reviewer Notes": "",
            }
        )
    return rows


def exact_or_alias_candidate(
    source_type: Any,
    canonical_by_type: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any] | None, str, str, str]:
    """Return an exact or single-target explicit candidate."""
    source = text(source_type)
    exact = canonical_by_type.get(normalized(source))
    if exact:
        return (
            exact,
            "Exact Match",
            "Exact normalized source Repair Type equals canonical Service Type.",
            "High",
        )
    alias = ALIAS_CANDIDATES.get(source)
    if alias is None:
        return None, "No Safe Mapping", "No explicit mapping rule.", "Unassessed"
    targets, _, evidence = alias
    if len(targets) != 1:
        return None, "No Safe Mapping", evidence, "Unassessed"
    candidate = canonical_by_type.get(normalized(targets[0]))
    if candidate is None:
        return (
            None,
            "No Safe Mapping",
            f"{evidence} Runtime canonical target was not found.",
            "Unassessed",
        )
    return candidate, "Explicit Alias Candidate", evidence, "Medium"


def service_name_candidate(
    service_name: Any,
    canonical_by_type: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any] | None, str]:
    """Apply deterministic explicit token rules to Miscellaneous Repair."""
    tokens = set(normalized(service_name).split())
    matches: list[dict[str, Any]] = []
    for required_tokens, target in SERVICE_NAME_RULES:
        if set(required_tokens) <= tokens:
            candidate = canonical_by_type.get(normalized(target))
            if candidate is not None:
                matches.append(candidate)
    unique = {
        text(record["Proposed Canonical Service Type ID"]): record
        for record in matches
    }
    if len(unique) == 1:
        return next(iter(unique.values())), (
            "Explicit service-name token rule produced one candidate; "
            "business review is required."
        )
    if len(unique) > 1:
        names = sorted(text(record["Service Type"]) for record in unique.values())
        return None, f"Multiple explicit service-name rules matched: {', '.join(names)}"
    return None, "No explicit service-name rule produced a safe candidate."


def service_normalization_rows(
    services: Sequence[dict[str, Any]],
    canonical_by_type: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Create one pending-review normalization row per Master Service."""
    rows = []
    for service in services:
        current = text(service.get("Repair Type"))
        if current == "Miscellaneous Repair":
            candidate, evidence = service_name_candidate(
                service.get("Service Name"),
                canonical_by_type,
            )
            method = "Service Name Rule Candidate" if candidate else "No Safe Mapping"
            confidence = "Low" if candidate else "Unassessed"
        else:
            candidate, method, evidence, confidence = exact_or_alias_candidate(
                current,
                canonical_by_type,
            )
        rows.append(
            {
                "Service ID": text(service.get("Service ID")),
                "Service Name": text(service.get("Service Name")),
                "Current Repair Type ID": text(service.get("Repair Type ID")),
                "Current Repair Type": current,
                "Manufacturer ID": text(service.get("Manufacturer ID")),
                "Manufacturer Name": text(service.get("Manufacturer Name")),
                "Device Family Code": text(service.get("Device Family Code")),
                "Device Family Name": text(service.get("Device Family Name")),
                "Proposed Canonical Service Type ID": (
                    text(candidate["Proposed Canonical Service Type ID"])
                    if candidate
                    else ""
                ),
                "Proposed Canonical Service Type": (
                    text(candidate["Service Type"]) if candidate else ""
                ),
                "Mapping Method": method,
                "Mapping Evidence": evidence,
                "Confidence": confidence,
                "Review Status": generated_review_status(
                    "Service", resolved=candidate is not None
                ),
                "Reviewer Notes": "",
            }
        )
    return rows


def labor_normalization_rows(
    labor: Sequence[dict[str, Any]],
    canonical_by_type: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Create one pending-review normalization row per Master Labor record."""
    rows = []
    for record in labor:
        candidate, method, evidence, confidence = exact_or_alias_candidate(
            record.get("Repair Type"),
            canonical_by_type,
        )
        rows.append(
            {
                "Labor Standard ID": text(record.get("Labor Standard ID")),
                "Legacy Labor ID": text(record.get("Legacy Labor ID")),
                "Labor Name": text(record.get("Labor Name")),
                "Current Repair Type": text(record.get("Repair Type")),
                "Device Family": text(record.get("Device Family")),
                "Manufacturer": text(record.get("Manufacturer")),
                "Proposed Canonical Service Type ID": (
                    text(candidate["Proposed Canonical Service Type ID"])
                    if candidate
                    else ""
                ),
                "Proposed Canonical Service Type": (
                    text(candidate["Service Type"]) if candidate else ""
                ),
                "Mapping Method": method,
                "Mapping Evidence": evidence,
                "Confidence": confidence,
                "Review Status": generated_review_status(
                    "Labor", resolved=candidate is not None
                ),
                "Reviewer Notes": "",
            }
        )
    return rows


def family_matches(service: dict[str, Any], labor: dict[str, Any]) -> bool:
    """Require an exact family match or an explicitly universal labor row."""
    labor_family = normalized(labor.get("Device Family"))
    if labor_family in UNIVERSAL_FAMILIES:
        return True
    service_families = {
        normalized(service.get("Device Family Code")),
        normalized(service.get("Device Family Name")),
    } - {""}
    return bool(labor_family and labor_family in service_families)


def manufacturer_matches(service: dict[str, Any], labor: dict[str, Any]) -> bool:
    """Constrain manufacturer-specific labor rows."""
    labor_manufacturer = normalized(labor.get("Manufacturer"))
    if not labor_manufacturer:
        return True
    service_manufacturers = {
        normalized(service.get("Manufacturer ID")),
        normalized(service.get("Manufacturer Name")),
    } - {""}
    return labor_manufacturer in service_manufacturers


def service_labor_candidates(
    service_rows: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
    source_labor: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Create at most one governed relationship suggestion per Service."""
    source_by_id = {
        text(record.get("Labor Standard ID")): record for record in source_labor
    }
    candidates = []
    for service in service_rows:
        canonical_id = text(service["Proposed Canonical Service Type ID"])
        if not canonical_id:
            continue
        valid = [
            labor
            for labor in labor_rows
            if text(labor["Proposed Canonical Service Type ID"]) == canonical_id
            and family_matches(service, labor)
            and manufacturer_matches(service, labor)
        ]
        if not valid:
            continue
        unique = valid[0] if len(valid) == 1 else None
        source = (
            source_by_id[text(unique["Labor Standard ID"])] if unique else {}
        )
        candidates.append(
            {
                "Service ID": text(service["Service ID"]),
                "Service Name": text(service["Service Name"]),
                "Canonical Service Type ID": canonical_id,
                "Canonical Service Type": text(
                    service["Proposed Canonical Service Type"]
                ),
                "Device Family Code": text(service["Device Family Code"]),
                "Manufacturer ID": text(service["Manufacturer ID"]),
                "Suggested Labor Standard ID": (
                    text(unique["Labor Standard ID"]) if unique else ""
                ),
                "Legacy Labor ID": (
                    text(unique["Legacy Labor ID"]) if unique else ""
                ),
                "Labor Name": text(unique["Labor Name"]) if unique else "",
                "Standard Minutes": source.get("Standard Minutes", ""),
                "Minimum Minutes": source.get("Minimum Minutes", ""),
                "Maximum Minutes": source.get("Maximum Minutes", ""),
                "Candidate Method": (
                    "Canonical Service Type + family/manufacturer constraints"
                ),
                "Evidence": (
                    "One labor row satisfies canonical Service Type, Device "
                    "Family, and Manufacturer constraints."
                    if unique
                    else "Multiple labor rows satisfy all governed constraints: "
                    + ", ".join(
                        sorted(text(row["Labor Standard ID"]) for row in valid)
                    )
                ),
                "Confidence": "High" if unique else "Unassessed",
                "Ambiguity Count": len(valid),
                "Review Status": generated_review_status(
                    "Relationship",
                    resolved=unique is not None,
                    ambiguous=unique is None,
                ),
                "Reviewer Notes": "",
            }
        )
    return candidates


def unresolved_rows(
    service_rows: Sequence[dict[str, Any]],
    labor_rows: Sequence[dict[str, Any]],
    candidate_rows: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Route unsafe normalization and ambiguous labor choices to review."""
    rows: list[dict[str, Any]] = []
    for record in service_rows:
        if text(record["Proposed Canonical Service Type ID"]):
            continue
        rows.append(
            {
                "Record Type": "Service Normalization",
                "Source Record ID": text(record["Service ID"]),
                "Source Name": text(record["Service Name"]),
                "Current Type": text(record["Current Repair Type"]),
                "Candidate Canonical Types": "",
                "Candidate Labor Standards": "",
                "Ambiguity Reason": text(record["Mapping Evidence"]),
                "Missing Evidence": "Canonical operation classification",
                "Required Action": "Select a canonical type or document no mapping.",
                "Review Priority": "High",
                "Review Status": generated_review_status("Service", resolved=False),
                "Reviewer Notes": "",
            }
        )
    for record in labor_rows:
        if text(record["Proposed Canonical Service Type ID"]):
            continue
        alias = ALIAS_CANDIDATES.get(text(record["Current Repair Type"]))
        candidate_types = ", ".join(alias[0]) if alias else ""
        rows.append(
            {
                "Record Type": "Labor Normalization",
                "Source Record ID": text(record["Labor Standard ID"]),
                "Source Name": text(record["Labor Name"]),
                "Current Type": text(record["Current Repair Type"]),
                "Candidate Canonical Types": candidate_types,
                "Candidate Labor Standards": "",
                "Ambiguity Reason": text(record["Mapping Evidence"]),
                "Missing Evidence": "Approved alias or canonical operation",
                "Required Action": "Review alias and canonical taxonomy coverage.",
                "Review Priority": "High",
                "Review Status": generated_review_status("Labor", resolved=False),
                "Reviewer Notes": "",
            }
        )
    for record in candidate_rows:
        if int(record["Ambiguity Count"]) <= 1:
            continue
        evidence = text(record["Evidence"])
        labor_ids = evidence.partition(": ")[2]
        rows.append(
            {
                "Record Type": "Service-Labor Candidate",
                "Source Record ID": text(record["Service ID"]),
                "Source Name": text(record["Service Name"]),
                "Current Type": text(record["Canonical Service Type"]),
                "Candidate Canonical Types": text(
                    record["Canonical Service Type"]
                ),
                "Candidate Labor Standards": labor_ids,
                "Ambiguity Reason": (
                    f"{record['Ambiguity Count']} labor standards remain valid."
                ),
                "Missing Evidence": "Labor-standard selection evidence",
                "Required Action": "Select one labor standard or refine constraints.",
                "Review Priority": "High",
                "Review Status": generated_review_status(
                    "Relationship", resolved=False, ambiguous=True
                ),
                "Reviewer Notes": "",
            }
        )
    return rows


def validate_source_identity(
    services: Sequence[dict[str, Any]],
    labor: Sequence[dict[str, Any]],
) -> None:
    """Require stable unique source identities."""
    service_ids = [text(record.get("Service ID")) for record in services]
    labor_ids = [text(record.get("Labor Standard ID")) for record in labor]
    legacy_ids = [text(record.get("Legacy Labor ID")) for record in labor]
    if any(not SERVICE_ID_PATTERN.fullmatch(value) for value in service_ids):
        raise NormalizationReviewError("Master Services has malformed Service IDs")
    if len(service_ids) != len(set(service_ids)):
        raise NormalizationReviewError("Master Services has duplicate Service IDs")
    if any(not LABOR_ID_PATTERN.fullmatch(value) for value in labor_ids):
        raise NormalizationReviewError("Master Labor has malformed Labor IDs")
    if len(labor_ids) != len(set(labor_ids)):
        raise NormalizationReviewError("Master Labor has duplicate Labor IDs")
    if any(not value for value in legacy_ids):
        raise NormalizationReviewError("Master Labor has blank Legacy Labor IDs")
    if len(legacy_ids) != len(set(legacy_ids)):
        raise NormalizationReviewError("Master Labor has duplicate Legacy Labor IDs")


def append_table(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
    table_name: str,
) -> None:
    """Write a formatted Excel Table, retaining a blank template row if needed."""
    worksheet.append(list(headers))
    materialized = list(rows) or [{header: "" for header in headers}]
    for record in materialized:
        worksheet.append([excel_value(record.get(header, "")) for header in headers])
    end_column = get_column_letter(len(headers))
    table = Table(
        displayName=table_name,
        ref=f"A1:{end_column}{worksheet.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"
    for column, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(len(header) + 2, 12),
            58,
        )


def instruction_rows(
    services: Sequence[dict[str, Any]],
    labor: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Align guidance and validation lookup values in one table."""
    guidance = [
        (
            "Purpose",
            "Review-only normalization and labor candidate package; no import.",
        ),
        (
            "Authority",
            "Detailed L:T taxonomy on 33 Service Types is authoritative for review.",
        ),
        ("Approval", "Generated rows remain Pending Review."),
        ("Fuzzy Matching", "Similarity may not approve or assign a mapping."),
        ("Device Family", "Labor selection requires exact family or universal scope."),
        (
            "Manufacturer",
            "Manufacturer constrains selection only when labor is specific.",
        ),
    ]
    family_codes = sorted(
        {
            text(record.get("Device Family Code"))
            for record in services
            if text(record.get("Device Family Code"))
        }
    )
    manufacturer_ids = sorted(
        {
            text(record.get("Manufacturer ID"))
            for record in services
            if text(record.get("Manufacturer ID"))
        }
    )
    labor_ids = sorted(text(record.get("Labor Standard ID")) for record in labor)
    row_count = max(
        len(guidance),
        len(ALIAS_RULE_TYPES),
        len(MAPPING_METHODS),
        len(CONFIDENCE_VALUES),
        len(REVIEW_STATUSES),
        len(family_codes),
        len(manufacturer_ids),
        len(labor_ids),
        len(YES_NO_VALUES),
    )
    rows = []
    for index in range(row_count):
        topic, detail = guidance[index] if index < len(guidance) else ("", "")
        rows.append(
            {
                "Topic": topic,
                "Guidance": detail,
                "Alias Rule Type": (
                    ALIAS_RULE_TYPES[index]
                    if index < len(ALIAS_RULE_TYPES)
                    else ""
                ),
                "Mapping Method": (
                    MAPPING_METHODS[index]
                    if index < len(MAPPING_METHODS)
                    else ""
                ),
                "Confidence": (
                    CONFIDENCE_VALUES[index]
                    if index < len(CONFIDENCE_VALUES)
                    else ""
                ),
                "Review Status": (
                    REVIEW_STATUSES[index]
                    if index < len(REVIEW_STATUSES)
                    else ""
                ),
                "Device Family Code": (
                    family_codes[index] if index < len(family_codes) else ""
                ),
                "Manufacturer ID": (
                    manufacturer_ids[index]
                    if index < len(manufacturer_ids)
                    else ""
                ),
                "Labor Standard ID": (
                    labor_ids[index] if index < len(labor_ids) else ""
                ),
                "Yes/No Value": (
                    YES_NO_VALUES[index] if index < len(YES_NO_VALUES) else ""
                ),
            }
        )
    return rows


def add_defined_names_and_validations(workbook: Workbook) -> None:
    """Create workbook-defined validation lists and apply only named formulas."""
    for name, (sheet_name, column) in DEFINED_NAME_SPECS.items():
        worksheet = workbook[sheet_name]
        letter = get_column_letter(column)
        populated_rows = [
            row
            for row in range(2, worksheet.max_row + 1)
            if text(worksheet.cell(row, column).value)
        ]
        if not populated_rows:
            raise NormalizationReviewError(
                f"Defined-name source is empty: {name}"
            )
        workbook.defined_names.add(
            DefinedName(
                name,
                attr_text=(
                    f"'{sheet_name}'!${letter}$2:${letter}${max(populated_rows)}"
                ),
            )
        )
    for worksheet in workbook.worksheets:
        header_cells = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                min_col=1,
                max_col=worksheet.max_column,
            )
        )
        headers = [text(cell.value) for cell in header_cells]
        for column, header in enumerate(headers, start=1):
            defined_name = VALIDATION_NAMES_BY_HEADER.get(header)
            if defined_name is None or worksheet.max_row < 2:
                continue
            letter = get_column_letter(column)
            validation = DataValidation(
                type="list",
                formula1=f"={defined_name}",
                allow_blank=True,
            )
            worksheet.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}{worksheet.max_row}")


def build_workbook(
    canonical: Sequence[dict[str, Any]],
    aliases: Sequence[dict[str, Any]],
    services: Sequence[dict[str, Any]],
    labor: Sequence[dict[str, Any]],
    candidates: Sequence[dict[str, Any]],
    unresolved: Sequence[dict[str, Any]],
    protected: dict[Path, str],
    id_findings: Sequence[str],
    identity_authority: str,
    generated_at: datetime,
) -> Workbook:
    """Build the exact ten-sheet governed review artifact."""
    workbook = Workbook()
    workbook.active.title = SHEET_NAMES[0]
    for sheet_name in SHEET_NAMES[1:]:
        workbook.create_sheet(sheet_name)
    append_table(
        workbook["00 - Instructions"],
        [
            "Topic",
            "Guidance",
            "Alias Rule Type",
            "Mapping Method",
            "Confidence",
            "Review Status",
            "Device Family Code",
            "Manufacturer ID",
            "Labor Standard ID",
            "Yes/No Value",
        ],
        instruction_rows(services, labor),
        TABLE_NAMES["00 - Instructions"],
    )
    sheet_rows = {
        "01 - Canonical Service Types": (CANONICAL_HEADERS, canonical),
        "02 - Service Type Aliases": (ALIAS_HEADERS, aliases),
        "03 - Service Normalization": (SERVICE_HEADERS, services),
        "04 - Labor Normalization": (LABOR_HEADERS, labor),
        "05 - Service Labor Candidates": (CANDIDATE_HEADERS, candidates),
        "06 - Unresolved Review": (UNRESOLVED_HEADERS, unresolved),
    }
    for sheet_name, (headers, rows) in sheet_rows.items():
        append_table(
            workbook[sheet_name],
            headers,
            rows,
            TABLE_NAMES[sheet_name],
        )
    exact_services = sum(
        1 for row in services if row["Mapping Method"] == "Exact Match"
    )
    exact_labor = sum(
        1 for row in labor if row["Mapping Method"] == "Exact Match"
    )
    validation_rows = [
        {"Validation": "Worksheet contract", "Result": "Pass", "Count": 10},
        {
            "Validation": "Detailed canonical records",
            "Result": (
                "Pass"
                if len(canonical) == EXPECTED_CANONICAL_ROWS
                else "Runtime Source Changed"
            ),
            "Count": len(canonical),
        },
        {
            "Validation": "Service normalization records",
            "Result": (
                "Pass"
                if len(services) == EXPECTED_SERVICE_ROWS
                else "Runtime Source Changed"
            ),
            "Count": len(services),
        },
        {
            "Validation": "Labor normalization records",
            "Result": (
                "Pass"
                if len(labor) == EXPECTED_LABOR_ROWS
                else "Runtime Source Changed"
            ),
            "Count": len(labor),
        },
        {
            "Validation": "Exact Service candidates",
            "Result": "Review Only",
            "Count": exact_services,
        },
        {
            "Validation": "Exact Labor candidates",
            "Result": "Review Only",
            "Count": exact_labor,
        },
        {
            "Validation": "Unresolved records",
            "Result": "Pending Review",
            "Count": len(unresolved),
        },
        {"Validation": "Generated Approved statuses", "Result": "Pass", "Count": 0},
    ]
    append_table(
        workbook["07 - Validation Summary"],
        ["Validation", "Result", "Count"],
        validation_rows,
        TABLE_NAMES["07 - Validation Summary"],
    )
    append_table(
        workbook["08 - Revision History"],
        ["Version", "Date", "Change", "Status"],
        [
            {
                "Version": "1.0.2",
                "Date": generated_at.date(),
                "Change": (
                    "Initial Service Type normalization and labor mapping "
                    "governance review package."
                ),
                "Status": "Pending Review",
            }
        ],
        TABLE_NAMES["08 - Revision History"],
    )
    metadata = [
        {"Metadata Field": "Import Batch", "Value": IMPORT_BATCH},
        {"Metadata Field": "Generated At UTC", "Value": generated_at},
        {"Metadata Field": "Canonical Worksheet", "Value": CANONICAL_SHEET},
        {"Metadata Field": "Canonical Detailed Row Count", "Value": len(canonical)},
        {"Metadata Field": "Master Service Row Count", "Value": len(services)},
        {"Metadata Field": "Master Labor Row Count", "Value": len(labor)},
        {"Metadata Field": "Alias Candidate Count", "Value": len(aliases)},
        {"Metadata Field": "Unresolved Row Count", "Value": len(unresolved)},
        {"Metadata Field": "Identity Authority", "Value": identity_authority},
        {"Metadata Field": "Automatic Approval", "Value": "No"},
        {"Metadata Field": "Canonical Import Authorized", "Value": "No"},
    ]
    for index, finding in enumerate(id_findings, start=1):
        metadata.append(
            {
                "Metadata Field": f"Service Type ID Finding {index}",
                "Value": finding,
            }
        )
    for path, digest in protected.items():
        metadata.extend(
            [
                {
                    "Metadata Field": f"Protected Path: {path.name}",
                    "Value": str(path),
                },
                {
                    "Metadata Field": f"SHA-256: {path.name}",
                    "Value": digest,
                },
            ]
        )
    append_table(
        workbook["09 - Import Metadata"],
        ["Metadata Field", "Value"],
        metadata,
        TABLE_NAMES["09 - Import Metadata"],
    )
    add_defined_names_and_validations(workbook)
    return workbook


def table_records(worksheet: Worksheet, table_name: str) -> list[dict[str, Any]]:
    """Read nonblank records from an Excel Table."""
    table = worksheet.tables[table_name]
    min_column, min_row, max_column, max_row = range_boundaries(table.ref)
    headers = [
        text(worksheet.cell(min_row, column).value)
        for column in range(min_column, max_column + 1)
    ]
    records = []
    for row_number in range(min_row + 1, max_row + 1):
        values = [
            worksheet.cell(row_number, column).value
            for column in range(min_column, max_column + 1)
        ]
        if any(text(value) for value in values):
            records.append(dict(zip(headers, values, strict=True)))
    return records


def reject_prohibited_generated_statuses(
    records_by_sheet: dict[str, Sequence[dict[str, Any]]],
) -> None:
    """Reject prohibited values only in governed status fields."""
    for sheet_name in SHEET_NAMES:
        status_column, id_column, default_record_type = STATUS_AUDIT_SPECS[sheet_name]
        records = records_by_sheet.get(sheet_name, ())
        if status_column is None:
            continue
        for excel_row, record in enumerate(records, start=2):
            status = text(record.get(status_column))
            if normalized(status) not in PROHIBITED_GENERATED_STATUSES:
                continue
            record_type = text(record.get("Record Type")) or default_record_type
            source_id = text(record.get(id_column)) or "(none)"
            raise NormalizationReviewError(
                "Prohibited generated status: "
                f"worksheet={sheet_name}; table/record type={record_type}; "
                f"Excel row={excel_row}; source record ID={source_id}; "
                f"prohibited status={status}"
            )


def require_ooxml(path: Path) -> None:
    """Require a macro-free OOXML workbook without external links."""
    required = {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
    with zipfile.ZipFile(path) as archive:
        members = set(archive.namelist())
        corrupt_member = archive.testzip()
    if required - members or corrupt_member:
        raise NormalizationReviewError("Output OOXML package is incomplete")
    if "xl/vbaProject.bin" in members:
        raise NormalizationReviewError("Output unexpectedly contains macros")
    if any(member.startswith("xl/externalLinks/") for member in members):
        raise NormalizationReviewError("Output contains external links")


def validate_reopened(
    path: Path,
    services: Sequence[dict[str, Any]],
    labor: Sequence[dict[str, Any]],
) -> None:
    """Validate the persisted schema, source preservation, IDs, and statuses."""
    require_ooxml(path)
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        if workbook.sheetnames != SHEET_NAMES:
            raise NormalizationReviewError("Reopened worksheet contract differs")
        for sheet_name in SHEET_NAMES:
            if list(workbook[sheet_name].tables) != [TABLE_NAMES[sheet_name]]:
                raise NormalizationReviewError(
                    f"Reopened table contract differs: {sheet_name}"
                )
        if set(workbook.defined_names) != set(DEFINED_NAME_SPECS):
            raise NormalizationReviewError("Reopened defined-name contract differs")
        service_actual = table_records(
            workbook["03 - Service Normalization"],
            TABLE_NAMES["03 - Service Normalization"],
        )
        labor_actual = table_records(
            workbook["04 - Labor Normalization"],
            TABLE_NAMES["04 - Labor Normalization"],
        )
        if len(service_actual) != len(services):
            raise NormalizationReviewError("Service row count changed after save")
        if len(labor_actual) != len(labor):
            raise NormalizationReviewError("Labor row count changed after save")
        for actual, expected in zip(service_actual, services, strict=True):
            for field in (
                "Service ID",
                "Service Name",
                "Current Repair Type ID",
                "Current Repair Type",
                "Manufacturer ID",
                "Manufacturer Name",
                "Device Family Code",
                "Device Family Name",
            ):
                if text(actual[field]) != text(expected[field]):
                    raise NormalizationReviewError(
                        f"Service source value changed: {field}"
                    )
        for actual, expected in zip(labor_actual, labor, strict=True):
            for field in (
                "Labor Standard ID",
                "Legacy Labor ID",
                "Labor Name",
                "Current Repair Type",
                "Device Family",
                "Manufacturer",
            ):
                if text(actual[field]) != text(expected[field]):
                    raise NormalizationReviewError(
                        f"Labor source value changed: {field}"
                    )
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, datetime) and cell.value.tzinfo:
                        raise NormalizationReviewError(
                            f"Timezone-aware value persisted in {worksheet.title}"
                        )
            for validation in worksheet.data_validations.dataValidation:
                if not text(validation.formula1).startswith("="):
                    raise NormalizationReviewError(
                        f"Non-name validation formula in {worksheet.title}"
                    )
                if text(validation.formula1)[1:] not in DEFINED_NAME_SPECS:
                    raise NormalizationReviewError(
                        f"Direct cross-sheet validation in {worksheet.title}"
                    )
        records_by_sheet = {
            sheet_name: table_records(
                workbook[sheet_name],
                TABLE_NAMES[sheet_name],
            )
            for sheet_name in SHEET_NAMES
        }
        reject_prohibited_generated_statuses(records_by_sheet)
    finally:
        workbook.close()


def main() -> int:
    """Generate, self-validate, and atomically publish the review workbook."""
    mandatory = [CANONICAL_PATH, MASTER_SERVICES_PATH, MASTER_LABOR_PATH]
    protected_paths = [*mandatory, LEGACY_MAPPING_PATH]
    temporary_created = False
    try:
        assert_sequential_read_contract()
        assert_governed_service_type_id_scan_contract()
        assert_generated_status_contract()
        require_files(mandatory)
        if TEMP_OUTPUT_PATH.exists():
            raise NormalizationReviewError(
                f"Stale temporary output exists: {TEMP_OUTPUT_PATH}"
            )
        before = protected_state(protected_paths)
        print("Reading canonical Service Types...")
        canonical, id_findings, authority = canonical_snapshot()
        canonical_by_type, _ = canonical_indexes(canonical)
        print("Reading canonical Service Types complete.")
        print("Reading Master Services...")
        services_source = read_records(
            MASTER_SERVICES_PATH,
            SERVICES_SHEET,
            {
                "Service ID",
                "Service Name",
                "Repair Type ID",
                "Repair Type",
                "Manufacturer ID",
                "Manufacturer Name",
                "Device Family Code",
                "Device Family Name",
            },
        )
        print("Reading Master Services complete.")
        print("Reading Master Labor...")
        labor_source = read_records(
            MASTER_LABOR_PATH,
            LABOR_SHEET,
            {
                "Labor Standard ID",
                "Legacy Labor ID",
                "Labor Name",
                "Repair Type",
                "Device Family",
                "Manufacturer",
                "Standard Minutes",
                "Minimum Minutes",
                "Maximum Minutes",
            },
        )
        print("Reading Master Labor complete.")
        print("Building normalization records...")
        validate_source_identity(services_source, labor_source)
        aliases = alias_rows(canonical_by_type)
        services = service_normalization_rows(
            services_source,
            canonical_by_type,
        )
        labor = labor_normalization_rows(labor_source, canonical_by_type)
        candidates = service_labor_candidates(
            services,
            labor,
            labor_source,
        )
        unresolved = unresolved_rows(services, labor, candidates)
        print("Building normalization records complete.")
        generated_at = datetime.now(UTC).replace(tzinfo=None)
        workbook = build_workbook(
            canonical,
            aliases,
            services,
            labor,
            candidates,
            unresolved,
            before,
            id_findings,
            authority,
            generated_at,
        )
        print("Writing review workbook...")
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary_created = True
        try:
            workbook.save(TEMP_OUTPUT_PATH)
        finally:
            workbook.close()
        print("Writing review workbook complete.")
        print("Reopening and validating...")
        validate_reopened(TEMP_OUTPUT_PATH, services, labor)
        print("Reopening and validating complete.")
        after = protected_state(protected_paths)
        if after != before:
            changed = [
                str(path)
                for path in protected_paths
                if before[path] != after[path]
            ]
            raise NormalizationReviewError(
                f"Protected input state changed: {', '.join(changed)}"
            )
        os.replace(TEMP_OUTPUT_PATH, OUTPUT_PATH)
        exact_service_count = sum(
            row["Mapping Method"] == "Exact Match" for row in services
        )
        exact_labor_count = sum(
            row["Mapping Method"] == "Exact Match" for row in labor
        )
        unresolved_types = sorted(
            {
                text(row["Current Type"])
                for row in unresolved
                if text(row["Current Type"])
            }
        )
        print(f"Generated: {OUTPUT_PATH}")
        print(f"Canonical Service Types: {len(canonical)}")
        print(f"Service normalization rows: {len(services)}")
        print(f"Labor normalization rows: {len(labor)}")
        print(f"Exact Service candidates: {exact_service_count}")
        print(f"Exact Labor candidates: {exact_labor_count}")
        print(f"Alias candidates: {len(aliases)}")
        print(f"Unresolved records: {len(unresolved)}")
        print(f"Unresolved types: {', '.join(unresolved_types)}")
        print("Approved mappings: 0")
        return 0
    except (
        NormalizationReviewError,
        OSError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
    ) as exc:
        print(f"ERROR: {text(exc)}", file=sys.stderr)
        return 1
    finally:
        if temporary_created and TEMP_OUTPUT_PATH.exists():
            try:
                TEMP_OUTPUT_PATH.unlink()
            except OSError as cleanup_error:
                print(
                    f"ERROR: Failed to remove {TEMP_OUTPUT_PATH}: {cleanup_error}",
                    file=sys.stderr,
                )


if __name__ == "__main__":
    raise SystemExit(main())
