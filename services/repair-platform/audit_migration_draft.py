from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

SEVERITIES = ("PASS", "INFO", "WARNING", "ERROR", "BLOCKER")
STATUS_VALUES = {"Draft", "Pending Review", "Approved", "Active", "Retired", "Deprecated"}
CROSSWALK_STATUS_VALUES = {"Draft", "Pending Review", "Approved", "Retired"}
BOOLEAN_VALUES = {"TRUE", "FALSE", "True", "False", "Yes", "No", "1", "0", True, False, 1, 0}
EXPECTED_GOVERNANCE_SHEET = "15 LL_Catalog"
EXPECTED_CROSSWALK_SHEET = "19 - ID Crosswalks"
CONTROL_SHEETS = {"12 - Validation Summary", "14 - Import Metadata", "16 - Lookup Audit", "20 - Lookup Registry"}
VOLATILE_FUNCS = ("INDIRECT", "OFFSET", "TODAY", "NOW", "RAND", "RANDBETWEEN")
COMPAT_FUNCS = ("XLOOKUP", "FILTER", "SORT", "UNIQUE", "LET", "LAMBDA")
DYNAMIC_ARRAY_FUNCS = ("FILTER", "SORT", "UNIQUE")
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")


@dataclass
class Finding:
    FindingID: str
    Severity: str
    Category: str
    Worksheet: str = ""
    TableName: str = ""
    CellOrRange: str = ""
    FieldName: str = ""
    RecordID: str = ""
    Message: str = ""
    Expected: str = ""
    Actual: str = ""
    RecommendedAction: str = ""


@dataclass
class TableInfo:
    name: str
    worksheet: str
    ref: str
    headers: list[str]
    rows: list[dict[str, object]]
    row_numbers: list[int]
    style: str


class Audit:
    def __init__(self, workbook_path: Path, output_dir: Path | None = None):
        self.workbook_path = workbook_path.resolve()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_dir = output_dir or Path(__file__).resolve().parent / f"migration_audit_{stamp}"
        self.findings: list[Finding] = []
        self.finding_counter = 0
        self.wb = None
        self.cached_wb = None
        self.source_hash_before = ""
        self.source_hash_after = ""
        self.worksheet_inventory: list[dict[str, object]] = []
        self.table_inventory: list[dict[str, object]] = []
        self.tables: dict[str, TableInfo] = {}
        self.lookup_catalog_rows: list[dict[str, object]] = []
        self.lookup_group_rows: list[dict[str, object]] = []
        self.crosswalk_catalog_rows: list[dict[str, object]] = []
        self.lookup_catalog_audit: list[dict[str, object]] = []
        self.crosswalk_catalog_audit: list[dict[str, object]] = []
        self.primary_key_audit: list[dict[str, object]] = []
        self.foreign_key_audit: list[dict[str, object]] = []
        self.crosswalk_completeness: list[dict[str, object]] = []
        self.formula_audit: list[dict[str, object]] = []
        self.cached_tables: dict[str, TableInfo] = {}

    def add(self, severity: str, category: str, **kwargs) -> None:
        self.finding_counter += 1
        self.findings.append(Finding(f"F{self.finding_counter:05d}", severity, category, **kwargs))


def calculate_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def norm_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def is_true(value: object) -> bool:
    return str(value).strip().lower() in {"true", "yes", "1", "active", "approved"}


def is_blankish(value: object) -> bool:
    return clean(value) == ""


def id_matches(value: object, prefix: str) -> bool:
    return bool(re.fullmatch(re.escape(prefix) + r"\d{6}", clean(value)))


def registered_crosswalk_names(audit: Audit) -> set[str]:
    return {clean(r.get("CrosswalkTableName")) for r in audit.crosswalk_catalog_rows if clean(r.get("CrosswalkTableName"))}


def is_physical_crosswalk_table(audit: Audit, table: TableInfo) -> bool:
    if table.name == "tblCrosswalkCatalog":
        return False
    if table.worksheet == EXPECTED_CROSSWALK_SHEET:
        return True
    if table.name in registered_crosswalk_names(audit):
        return True
    return table.name.endswith("IDCrosswalk") or table.name.endswith("Crosswalk")


def classify_identifier_field(audit: Audit, table: TableInfo, field: str) -> str:
    for row in audit.lookup_catalog_rows:
        if clean(row.get("LookupTableName")) == table.name and clean(row.get("PrimaryKeyField")) == field:
            return "PRIMARY_KEY"
    if is_physical_crosswalk_table(audit, table):
        if field.startswith("Legacy") and field.endswith("ID"):
            return "CROSSWALK_LEGACY_KEY"
        if field.startswith("Canonical") and field.endswith("ID"):
            return "CROSSWALK_CANONICAL_KEY"
    pk_fields = {clean(r.get("PrimaryKeyField")) for r in audit.lookup_catalog_rows}
    if field in pk_fields:
        return "FOREIGN_KEY"
    if table.name in {"tblLookupCatalogID", "tblLookupGroupName", "tblCrosswalkCatalog"}:
        return "GOVERNANCE_METADATA_ID"
    if field.endswith("ID"):
        return "OTHER_ID_LIKE_FIELD"
    return ""


def load_workbook_safely(audit: Audit):
    if not audit.workbook_path.exists():
        raise FileNotFoundError(audit.workbook_path)
    with zipfile.ZipFile(audit.workbook_path) as zf:
        bad = zf.testzip()
        if bad:
            audit.add("BLOCKER", "Workbook Identity", Message="XLSX package integrity failed.", Actual=bad)
        if "xl/workbook.xml" not in zf.namelist():
            audit.add("BLOCKER", "Workbook Identity", Message="xl/workbook.xml missing.")
    audit.wb = load_workbook(audit.workbook_path, data_only=False, read_only=False)
    audit.cached_wb = load_workbook(audit.workbook_path, data_only=True, read_only=False)
    return audit.wb


def read_excel_table(ws, table) -> TableInfo:
    cells = list(ws[table.ref])
    headers = [clean(c.value) for c in cells[0]]
    rows, row_numbers = [], []
    for row in cells[1:]:
        rows.append({h: row[i].value for i, h in enumerate(headers)})
        row_numbers.append(row[0].row)
    style = table.tableStyleInfo.name if table.tableStyleInfo else ""
    return TableInfo(table.name, ws.title, table.ref, headers, rows, row_numbers, style)


def locate_table(audit: Audit, table_name: str) -> TableInfo | None:
    return audit.tables.get(table_name)


def inventory_worksheets(audit: Audit):
    seen_trimmed = Counter(ws.title.strip() for ws in audit.wb.worksheets)
    for ws in audit.wb.worksheets:
        tables = list(ws.tables.keys())
        row = {
            "Worksheet": ws.title,
            "Visibility": ws.sheet_state,
            "MaxRow": ws.max_row,
            "MaxColumn": ws.max_column,
            "TableCount": len(tables),
            "TableNames": "; ".join(tables),
            "LeadingSpaces": ws.title != ws.title.lstrip(),
            "TrailingSpaces": ws.title != ws.title.rstrip(),
            "DuplicateAfterTrim": seen_trimmed[ws.title.strip()] > 1,
            "NamingPatternIssue": ws.title.startswith("15 LL") and not re.match(r"^15 LL_[A-Za-z0-9]+$", ws.title),
        }
        audit.worksheet_inventory.append(row)
        if row["TrailingSpaces"] or row["LeadingSpaces"]:
            audit.add("WARNING", "Worksheet Inventory", Worksheet=ws.title, Message="Worksheet name has leading/trailing spaces.")
        if row["DuplicateAfterTrim"]:
            audit.add("ERROR", "Worksheet Inventory", Worksheet=ws.title, Message="Duplicate worksheet name after trimming whitespace.")
        if ws.sheet_state != "visible" and ("LL_" in ws.title or "Crosswalk" in ws.title or "Catalog" in ws.title):
            audit.add("WARNING", "Workbook Identity", Worksheet=ws.title, Message="Governance or migration sheet is hidden.")


def ranges_overlap(ref1: str, ref2: str) -> bool:
    a1, b1, c1, d1 = range_boundaries(ref1)
    a2, b2, c2, d2 = range_boundaries(ref2)
    return not (c1 < a2 or c2 < a1 or d1 < b2 or d2 < b1)


def inventory_tables(audit: Audit):
    by_name = Counter()
    temp = []
    for ws in audit.wb.worksheets:
        for table in ws.tables.values():
            info = read_excel_table(ws, table)
            audit.tables[info.name] = info
            if audit.cached_wb and ws.title in audit.cached_wb.sheetnames and table.name in audit.cached_wb[ws.title].tables:
                audit.cached_tables[info.name] = read_excel_table(audit.cached_wb[ws.title], audit.cached_wb[ws.title].tables[table.name])
            temp.append(info)
            by_name[info.name] += 1
    for info in temp:
        blank_headers = [i + 1 for i, h in enumerate(info.headers) if not h]
        dup_headers = [h for h, c in Counter(info.headers).items() if h and c > 1]
        min_col, min_row, max_col, max_row = range_boundaries(info.ref)
        ws = audit.wb[info.worksheet]
        below_values = [ws.cell(max_row + 1, col).value for col in range(min_col, max_col + 1)] if max_row + 1 <= ws.max_row else []
        blank_rows = sum(1 for r in info.rows if all(v in (None, "") for v in r.values()))
        overlaps = [o.name for o in temp if o.name != info.name and o.worksheet == info.worksheet and ranges_overlap(info.ref, o.ref)]
        formulas_inconsistent = []
        for h in info.headers:
            patterns = set()
            for r in info.rows:
                v = r.get(h)
                if isinstance(v, str) and v.startswith("="):
                    patterns.add(re.sub(r"\d+", "#", v))
            if len(patterns) > 1:
                formulas_inconsistent.append(h)
        row = {
            "TableName": info.name,
            "Worksheet": info.worksheet,
            "Range": info.ref,
            "HeaderRow": min_row,
            "ColumnCount": len(info.headers),
            "DataRowCount": len(info.rows),
            "TableStyle": info.style,
            "OverlapsAnotherTable": "; ".join(overlaps),
            "BlankHeaders": "; ".join(map(str, blank_headers)),
            "DuplicateHeaders": "; ".join(dup_headers),
            "DuplicateTableName": by_name[info.name] > 1,
            "BlankRowsInsideRange": blank_rows,
            "DataImmediatelyBelow": any(v not in (None, "") for v in below_values),
            "InconsistentFormulaColumns": "; ".join(formulas_inconsistent),
        }
        audit.table_inventory.append(row)
        if not info.name.startswith("tbl") or " " in info.name or re.search(r"[^A-Za-z0-9_]", info.name):
            audit.add("ERROR", "Excel Table Inventory", Worksheet=info.worksheet, TableName=info.name, Message="Table name violates tbl/no-space/alphanumeric convention.")
        if blank_headers or dup_headers:
            audit.add("ERROR", "Excel Table Inventory", Worksheet=info.worksheet, TableName=info.name, Message="Table has blank or duplicate headers.", Actual=f"blank={blank_headers}; duplicate={dup_headers}")
        if overlaps:
            audit.add("BLOCKER", "Excel Table Inventory", Worksheet=info.worksheet, TableName=info.name, Message="Table range overlaps another table.", Actual="; ".join(overlaps))


def validate_lookup_groups(audit: Audit):
    table = locate_table(audit, "tblLookupGroupName")
    if not table:
        audit.add("BLOCKER", "Lookup Group Validation", Message="tblLookupGroupName was not found.")
        return
    audit.lookup_group_rows = table.rows
    ids = [clean(r.get("LookupGroupID")) for r in table.rows]
    names = [clean(r.get("GroupName")) for r in table.rows]
    for group_id, count in Counter(ids).items():
        if group_id and count > 1:
            audit.add("ERROR", "Lookup Group Validation", TableName=table.name, FieldName="LookupGroupID", RecordID=group_id, Message="Duplicate LookupGroupID.")
        if group_id and not re.fullmatch(r"LKG\d{6}", group_id):
            audit.add("ERROR", "Lookup Group Validation", TableName=table.name, FieldName="LookupGroupID", RecordID=group_id, Message="LookupGroupID does not match LKG000000 format.")
    for name, count in Counter(names).items():
        if not name:
            audit.add("ERROR", "Lookup Group Validation", TableName=table.name, FieldName="GroupName", Message="Blank group name.")
        elif count > 1:
            audit.add("WARNING", "Lookup Group Validation", TableName=table.name, FieldName="GroupName", Message="Duplicate group name.", Actual=name)


def validate_lookup_catalog(audit: Audit):
    table = locate_table(audit, "tblLookupCatalogID")
    if not table:
        audit.add("BLOCKER", "Lookup Catalog Validation", Message="tblLookupCatalogID was not found.")
        return
    audit.lookup_catalog_rows = table.rows
    expected_cols = ["LookupCatalogID", "LookupGroupID", "LookupTableName", "DisplayName", "PrimaryKeyField", "PrimaryKeyPrefix", "Description", "DisplayOrder", "Status", "IsSystemLookup", "TableExists", "LastReviewed"]
    missing_cols = [c for c in expected_cols if c not in table.headers]
    if missing_cols:
        audit.add("WARNING", "Lookup Catalog Validation", TableName=table.name, Message="Expected metadata columns are absent.", Actual="; ".join(missing_cols))
    group_ids = {clean(r.get("LookupGroupID")) for r in audit.lookup_group_rows}
    table_names = set(audit.tables)
    catalog_names = []
    display_orders = defaultdict(list)
    for r in table.rows:
        catalog_id = clean(r.get("LookupCatalogID"))
        lookup_table = clean(r.get("LookupTableName"))
        pk = clean(r.get("PrimaryKeyField"))
        prefix = clean(r.get("PrimaryKeyPrefix"))
        status = clean(r.get("Status"))
        group_id = clean(r.get("LookupGroupID"))
        order = r.get("DisplayOrder")
        exists_actual = lookup_table in table_names
        catalog_names.append(lookup_table)
        display_orders[group_id].append(order)
        row = {
            "LookupCatalogID": catalog_id,
            "LookupGroupID": group_id,
            "LookupTableName": lookup_table,
            "DisplayName": clean(r.get("DisplayName")),
            "PrimaryKeyField": pk,
            "PrimaryKeyPrefix": prefix,
            "Status": status,
            "TableExistsValue": clean(r.get("TableExists")),
            "ActualTableExists": exists_actual,
            "Issue": "",
        }
        if catalog_id and not re.fullmatch(r"LKC\d{6}", catalog_id):
            row["Issue"] += "Bad LookupCatalogID format; "
            audit.add("ERROR", "Lookup Catalog Validation", TableName=table.name, RecordID=catalog_id, Message="LookupCatalogID does not match LKC000000 format.")
        if status and status not in STATUS_VALUES:
            row["Issue"] += "Unknown status; "
            audit.add("WARNING", "Lookup Catalog Validation", TableName=table.name, RecordID=catalog_id, FieldName="Status", Message="Unknown lookup status.", Actual=status)
        if group_id and group_id not in group_ids:
            row["Issue"] += "Missing group; "
            audit.add("ERROR", "Lookup Catalog Validation", TableName=table.name, RecordID=catalog_id, FieldName="LookupGroupID", Message="LookupGroupID not found in tblLookupGroupName.", Actual=group_id)
        if status in {"Active", "Approved"} and not exists_actual:
            row["Issue"] += "Active/Approved table missing; "
            audit.add("ERROR", "Lookup Catalog Validation", TableName=table.name, RecordID=catalog_id, Message="Active/Approved catalog row points to missing physical table.", Actual=lookup_table)
        if exists_actual:
            target = audit.tables[lookup_table]
            if pk not in target.headers:
                row["Issue"] += "PK field missing; "
                audit.add("BLOCKER", "Lookup Catalog Validation", Worksheet=target.worksheet, TableName=lookup_table, FieldName=pk, Message="Configured primary-key field does not exist.")
        audit.lookup_catalog_audit.append(row)
    for name, count in Counter(catalog_names).items():
        if name and count > 1:
            audit.add("ERROR", "Lookup Catalog Validation", FieldName="LookupTableName", Message="Duplicate LookupTableName in catalog.", Actual=name)
    for group_id, values in display_orders.items():
        numeric = [v for v in values if isinstance(v, (int, float))]
        if len(numeric) != len([v for v in values if v not in (None, "")]):
            audit.add("WARNING", "Lookup Catalog Validation", FieldName="DisplayOrder", RecordID=group_id, Message="Non-numeric DisplayOrder values found.")
        for v, c in Counter(numeric).items():
            if c > 1:
                audit.add("WARNING", "Lookup Catalog Validation", FieldName="DisplayOrder", RecordID=group_id, Message="Duplicate DisplayOrder within group.", Actual=str(v))


def validate_crosswalk_catalog(audit: Audit):
    table = locate_table(audit, "tblCrosswalkCatalog")
    if not table:
        audit.add("WARNING", "Crosswalk Catalog Validation", Message="tblCrosswalkCatalog was not found.")
        return
    audit.crosswalk_catalog_rows = table.rows
    for r in table.rows:
        cid = clean(r.get("CrosswalkCatalogID"))
        tname = clean(r.get("CrosswalkTableName"))
        status = clean(r.get("Status"))
        auth = r.get("IsProductionAuthorized")
        legacy = clean(r.get("LegacyKeyField"))
        canon = clean(r.get("CanonicalKeyField"))
        exists = tname in audit.tables
        issue = []
        if cid and not re.fullmatch(r"XWC\d{6}", cid):
            issue.append("Bad CrosswalkCatalogID format")
            audit.add("ERROR", "Crosswalk Catalog Validation", RecordID=cid, Message="CrosswalkCatalogID does not match XWC000000 format.")
        if status and status not in CROSSWALK_STATUS_VALUES:
            issue.append("Unknown status")
        if status == "Draft" and is_true(auth):
            issue.append("Draft production-authorized")
            audit.add("ERROR", "Crosswalk Catalog Validation", TableName=tname, RecordID=cid, Message="Draft crosswalk cannot be production-authorized.")
        if (status == "Approved" or is_true(auth)) and not exists:
            issue.append("Approved/authorized physical table missing")
            audit.add("ERROR", "Crosswalk Catalog Validation", TableName=tname, RecordID=cid, Message="Approved or production-authorized crosswalk points to missing table.")
        if exists:
            headers = audit.tables[tname].headers
            if legacy and legacy not in headers:
                issue.append("Legacy field missing")
            if canon and canon not in headers:
                issue.append("Canonical field missing")
        audit.crosswalk_catalog_audit.append({"CrosswalkCatalogID": cid, "CrosswalkTableName": tname, "Status": status, "IsProductionAuthorized": clean(auth), "PhysicalTableExists": exists, "Issue": "; ".join(issue)})


def build_primary_key_index(audit: Audit):
    index = defaultdict(dict)
    catalog_by_table = {}
    for r in audit.lookup_catalog_rows:
        tname, pk, prefix = clean(r.get("LookupTableName")), clean(r.get("PrimaryKeyField")), clean(r.get("PrimaryKeyPrefix"))
        if tname in audit.tables:
            catalog_by_table[tname] = (pk, prefix)
            table = audit.tables[tname]
            if pk in table.headers:
                for row, rownum in zip(table.rows, table.row_numbers):
                    val = clean(row.get(pk))
                    if val:
                        index[pk][val] = {"table": tname, "worksheet": table.worksheet, "row": rownum, "prefix": prefix}
    return index, catalog_by_table


def validate_primary_keys(audit: Audit):
    for r in audit.lookup_catalog_rows:
        tname, pk, prefix = clean(r.get("LookupTableName")), clean(r.get("PrimaryKeyField")), clean(r.get("PrimaryKeyPrefix"))
        table = audit.tables.get(tname)
        if not table or pk not in table.headers:
            continue
        values = [clean(row.get(pk)) for row in table.rows]
        blanks = sum(1 for v in values if not v)
        duplicates = sum(c for v, c in Counter(values).items() if v and c > 1)
        malformed = [v for v in values if v and not id_matches(v, prefix)]
        audit.primary_key_audit.append({"TableName": tname, "Worksheet": table.worksheet, "PrimaryKeyField": pk, "Prefix": prefix, "Rows": len(values), "BlankIDs": blanks, "DuplicateIDs": duplicates, "MalformedIDs": len(malformed), "Examples": "; ".join(malformed[:10])})
        if blanks:
            audit.add("BLOCKER", "Primary-Key Validation", Worksheet=table.worksheet, TableName=tname, FieldName=pk, Message="Blank primary keys found.", Actual=str(blanks))
        if duplicates:
            audit.add("BLOCKER", "Primary-Key Validation", Worksheet=table.worksheet, TableName=tname, FieldName=pk, Message="Duplicate primary keys found.", Actual=str(duplicates))
        if malformed:
            audit.add("ERROR", "Primary-Key Validation", Worksheet=table.worksheet, TableName=tname, FieldName=pk, Message="Malformed primary keys found.", Expected=f"{prefix} + exactly six digits", Actual="; ".join(malformed[:10]))


def validate_foreign_keys(audit: Audit):
    pk_index, catalog_by_table = build_primary_key_index(audit)
    own_pk_by_table = {t: pk for t, (pk, _) in catalog_by_table.items()}
    for table in audit.tables.values():
        for field in table.headers:
            if field not in pk_index or own_pk_by_table.get(table.name) == field:
                continue
            invalid = multi = legacy = 0
            examples = []
            for row, rownum in zip(table.rows, table.row_numbers):
                val = clean(row.get(field))
                if not val:
                    continue
                if any(d in val for d in [";", ",", "|", "\n"]):
                    multi += 1
                    examples.append(val)
                    continue
                if val not in pk_index[field]:
                    invalid += 1
                    examples.append(val)
                    prefix = next(iter(pk_index[field].values()))["prefix"] if pk_index[field] else ""
                    if prefix and re.fullmatch(re.escape(prefix) + r"-?\d{1,5}", val):
                        legacy += 1
            audit.foreign_key_audit.append({"Worksheet": table.worksheet, "TableName": table.name, "ForeignKeyField": field, "InvalidReferences": invalid, "LegacyFormatReferences": legacy, "MultiValueCells": multi, "Examples": "; ".join(examples[:10])})
            if invalid:
                audit.add("ERROR", "Foreign-Key Validation", Worksheet=table.worksheet, TableName=table.name, FieldName=field, Message="Foreign-key values not found in referenced table.", Actual=f"{invalid}; examples={examples[:5]}")
            if multi:
                audit.add("ERROR", "Foreign-Key Validation", Worksheet=table.worksheet, TableName=table.name, FieldName=field, Message="Potential single-FK column contains multi-value cells.", Actual=f"{multi}; examples={examples[:5]}")


def validate_crosswalk_tables(audit: Audit):
    physical = [t for t in audit.tables.values() if is_physical_crosswalk_table(audit, t)]
    for table in physical:
        legacy = next((h for h in table.headers if h.startswith("Legacy") and h.endswith("ID")), "")
        canon = next((h for h in table.headers if h.startswith("Canonical") and h.endswith("ID")), "")
        blanks_legacy = blanks_canon = conflicts = 0
        placeholder_rows = 0
        mapping = defaultdict(set)
        for row in table.rows:
            relevant = [row.get(h) for h in table.headers if h not in {"ReviewedBy", "ReviewedDate", "ApprovedBy", "ApprovedDate", "Notes"}]
            if all(is_blankish(v) for v in relevant):
                placeholder_rows += 1
                continue
            lval, cval = clean(row.get(legacy)), clean(row.get(canon))
            if not lval:
                blanks_legacy += 1
            if not cval:
                blanks_canon += 1
            if lval and cval:
                mapping[lval].add(cval)
        conflicts = sum(1 for vals in mapping.values() if len(vals) > 1)
        audit.crosswalk_completeness.append({"CrosswalkTable": table.name, "Worksheet": table.worksheet, "Rows": len(table.rows), "LegacyKeyField": legacy, "CanonicalKeyField": canon, "BlankLegacyKeys": blanks_legacy, "BlankCanonicalKeys": blanks_canon, "EmptyPlaceholderRows": placeholder_rows, "OneToManyConflicts": conflicts})
        if conflicts:
            audit.add("BLOCKER", "Crosswalk Table Validation", Worksheet=table.worksheet, TableName=table.name, Message="One legacy key maps to multiple canonical IDs.", Actual=str(conflicts))
        if blanks_legacy or blanks_canon:
            audit.add("ERROR", "Crosswalk Table Validation", Worksheet=table.worksheet, TableName=table.name, Message="Crosswalk contains blank legacy or canonical keys.", Actual=f"legacy={blanks_legacy}; canonical={blanks_canon}")


def evaluate_crosswalk_completeness(audit: Audit):
    # Basic completeness is recorded by validate_crosswalk_tables; this function reserves a PASS/INFO summary.
    count = sum(1 for t in audit.tables.values() if is_physical_crosswalk_table(audit, t))
    audit.add("INFO", "Crosswalk Completeness", Message="Physical crosswalk tables discovered dynamically.", Actual=str(count))


def validate_device_model_crosswalk(audit: Audit):
    table = locate_table(audit, "tblDeviceModelIDCrosswalk")
    if not table:
        audit.add("INFO", "Device Model Crosswalk Readiness", Message="tblDeviceModelIDCrosswalk is not physically present.")
        return
    required_groups = {
        "LegacyDeviceModelID": ["LegacyDeviceModelID"],
        "CanonicalDeviceModelID": ["CanonicalDeviceModelID"],
        "DeviceModel": ["DeviceModel", "DeviceModelName", "ModelName"],
        "DeviceFamilyID": ["DeviceFamilyID"],
        "ManufacturerID": ["ManufacturerID"],
        "LegacySource": ["LegacySource"],
        "CanonicalSource": ["CanonicalSource"],
        "MatchMethod": ["MatchMethod", "MappingMethod"],
        "Confidence": ["Confidence"],
        "MappingStatus": ["MappingStatus"],
        "EvidenceSource": ["EvidenceSource"],
        "ReviewedBy": ["ReviewedBy"],
        "ReviewedDate": ["ReviewedDate"],
        "ApprovedBy": ["ApprovedBy"],
        "ApprovedDate": ["ApprovedDate"],
        "Notes": ["Notes"],
    }
    missing = [name for name, aliases in required_groups.items() if not any(a in table.headers for a in aliases)]
    if missing:
        audit.add("WARNING", "Device Model Crosswalk Readiness", Worksheet=table.worksheet, TableName=table.name, Message="Preferred columns missing.", Actual="; ".join(missing))


def validate_display_orders(audit: Audit):
    for table in audit.tables.values():
        if "DisplayOrder" not in table.headers:
            continue
        parent_field = "LookupGroupID" if table.name == "tblLookupCatalogID" and "LookupGroupID" in table.headers else ""
        groups = defaultdict(list)
        for row in table.rows:
            groups[clean(row.get(parent_field)) if parent_field else "__ALL__"].append(row.get("DisplayOrder"))
        for parent, vals in groups.items():
            bad = [v for v in vals if v in (None, "") or not isinstance(v, (int, float)) or int(v) <= 0]
            numeric = [int(v) for v in vals if isinstance(v, (int, float)) and int(v) > 0]
            dup = [v for v, c in Counter(numeric).items() if c > 1]
            gaps = []
            if numeric:
                expected = set(range(min(numeric), max(numeric) + 1))
                gaps = sorted(expected - set(numeric))
            if bad or dup or gaps:
                audit.add("WARNING", "Display Order Validation", Worksheet=table.worksheet, TableName=table.name, FieldName="DisplayOrder", RecordID=parent, Message="DisplayOrder has blank, nonnumeric, duplicate, nonpositive, or gap values within parent group.", Actual=f"bad={bad[:5]}; duplicates={dup[:5]}; gaps={gaps[:10]}")


def validate_boolean_and_status(audit: Audit):
    for table in audit.tables.values():
        cached_table = audit.cached_tables.get(table.name)
        for h in table.headers:
            hl = h.lower()
            source_rows = cached_table.rows if cached_table and h in cached_table.headers else table.rows
            vals = [r.get(h) for r in source_rows if r.get(h) not in (None, "")]
            if h in {"IsSystemLookup", "IsProductionAuthorized", "TableExists"} or "active" in hl:
                for formula_row, cached_row in zip(table.rows, source_rows):
                    if isinstance(formula_row.get(h), str) and str(formula_row.get(h)).startswith("=") and cached_row.get(h) in (None, ""):
                        audit.add("WARNING", "Boolean and Status Normalization", Worksheet=table.worksheet, TableName=table.name, FieldName=h, Message="Formula-backed boolean has no cached value.")
                invalid = [clean(v) for v in vals if v not in BOOLEAN_VALUES and clean(v).upper() not in {"TRUE", "FALSE"}]
                reps = {clean(v) for v in vals}
                if invalid:
                    audit.add("WARNING", "Boolean and Status Normalization", Worksheet=table.worksheet, TableName=table.name, FieldName=h, Message="Boolean-like field has invalid values.", Actual="; ".join(invalid[:10]))
                if len(reps) > 2:
                    audit.add("INFO", "Boolean and Status Normalization", Worksheet=table.worksheet, TableName=table.name, FieldName=h, Message="Boolean-like field uses multiple representations.", Actual="; ".join(sorted(reps)))
            if "status" in hl:
                stripped = [clean(v) for v in vals]
                if len(set(stripped)) != len(set(map(str, vals))):
                    audit.add("WARNING", "Boolean and Status Normalization", Worksheet=table.worksheet, TableName=table.name, FieldName=h, Message="Status values differ by whitespace/casing.")


def audit_formulas(audit: Audit):
    table_by_cell = {}
    for table in audit.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                table_by_cell[(table.worksheet, r, c)] = table.name
    by_ws = Counter()
    by_table = Counter()
    volatile = compat = dynamic = broken = external = 0
    identifier_counts = Counter()
    for ws in audit.wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    by_ws[ws.title] += 1
                    tname = table_by_cell.get((ws.title, cell.row, cell.column), "")
                    if tname:
                        by_table[tname] += 1
                    upper = v.upper()
                    volatile += any(f + "(" in upper for f in VOLATILE_FUNCS)
                    compat += any(f + "(" in upper or f"_XLFN.{f}" in upper for f in COMPAT_FUNCS)
                    dynamic += any(f + "(" in upper or f"_XLFN.{f}" in upper for f in DYNAMIC_ARRAY_FUNCS)
                    broken += "#REF!" in v
                    external += bool(re.search(r"\[[^\]]+\.xls[a-zm]*\]", v, re.IGNORECASE))
                    if tname:
                        table = audit.tables[tname]
                        min_col, _, _, _ = range_boundaries(table.ref)
                        header = table.headers[cell.column - min_col] if 0 <= cell.column - min_col < len(table.headers) else ""
                        category = classify_identifier_field(audit, table, header)
                        if category:
                            identifier_counts[category] += 1
    for ws, count in by_ws.items():
        audit.formula_audit.append({"Scope": "Worksheet", "Name": ws, "FormulaCount": count, "VolatileFormulaUses": "", "CompatibilitySensitiveUses": "", "FormulaErrors": ""})
    for table, count in by_table.items():
        audit.formula_audit.append({"Scope": "Table", "Name": table, "FormulaCount": count, "VolatileFormulaUses": "", "CompatibilitySensitiveUses": "", "FormulaErrors": ""})
    audit.formula_audit.append({"Scope": "Workbook", "Name": "Totals", "FormulaCount": sum(by_ws.values()), "VolatileFormulaUses": volatile, "CompatibilitySensitiveUses": compat, "DynamicArrayFormulaUses": dynamic, "BrokenReferenceFormulaUses": broken, "ExternalReferenceFormulaUses": external, **{f"IdentifierFormula_{k}": v for k, v in identifier_counts.items()}})
    if broken:
        audit.add("ERROR", "Formula Audit", Message="Broken-reference formulas found.", Actual=str(broken))
    if volatile:
        audit.add("WARNING", "Formula Audit", Message="Volatile formulas found.", Actual=str(volatile))
    if compat:
        audit.add("WARNING", "Formula Audit", Message="Compatibility-sensitive formulas found.", Actual=str(compat))
    if dynamic:
        audit.add("WARNING", "Formula Audit", Message="Dynamic-array formulas found.", Actual=str(dynamic))
    if external:
        audit.add("WARNING", "Formula Audit", Message="External-reference formulas found.", Actual=str(external))
    severity_by_category = {
        "PRIMARY_KEY": "ERROR",
        "CROSSWALK_LEGACY_KEY": "ERROR",
        "CROSSWALK_CANONICAL_KEY": "ERROR",
        "FOREIGN_KEY": "WARNING",
        "GOVERNANCE_METADATA_ID": "WARNING",
        "OTHER_ID_LIKE_FIELD": "INFO",
    }
    for category, count in identifier_counts.items():
        audit.add(severity_by_category.get(category, "INFO"), "Formula Audit", FieldName=category, Message=f"Formulas found in {category} fields.", Actual=str(count))


def validate_table_specific_schemas(audit: Audit):
    checks = {
        "tblServiceCategories": ("ServiceCategoryID", "SER"),
        "tblChargingPortTypes": ("ChargingPortTypeID", ""),
        "tblToolCategories": ("ToolCategoryID", "TCAT"),
    }
    for name, (pk, prefix) in checks.items():
        t = locate_table(audit, name)
        if not t:
            continue
        if pk not in t.headers:
            audit.add("WARNING", "Table-Specific Schema Checks", Worksheet=t.worksheet, TableName=name, FieldName=pk, Message="Expected primary key field is absent.")
        if prefix and pk in t.headers:
            bad = [clean(r.get(pk)) for r in t.rows if clean(r.get(pk)) and not id_matches(r.get(pk), prefix)]
            if bad:
                audit.add("ERROR", "Table-Specific Schema Checks", Worksheet=t.worksheet, TableName=name, FieldName=pk, Message="Noncanonical table-specific ID format.", Expected=f"{prefix}000001", Actual="; ".join(bad[:10]))
    t = locate_table(audit, "tblChargingPortType")
    if t and "ChargePortTypeID" in t.headers:
        audit.add("WARNING", "Table-Specific Schema Checks", Worksheet=t.worksheet, TableName=t.name, FieldName="ChargePortTypeID", Message="ChargePortTypeID differs from expected ChargingPortTypeID.")
    dm = locate_table(audit, "tblDeviceModels")
    if dm:
        for col in ["DeviceModelID", "DeviceFamilyID", "ManufacturerID"]:
            if col not in dm.headers:
                audit.add("ERROR", "Table-Specific Schema Checks", Worksheet=dm.worksheet, TableName=dm.name, FieldName=col, Message="Expected relationship column is missing.")


def audit_data_quality(audit: Audit):
    for table in audit.tables.values():
        full_rows = [tuple(clean(r.get(h)) for h in table.headers) for r in table.rows]
        dup_rows = sum(c for row, c in Counter(full_rows).items() if any(row) and c > 1)
        whitespace = 0
        control = 0
        for r in table.rows:
            for v in r.values():
                if isinstance(v, str):
                    if v != v.strip() or "\u00a0" in v:
                        whitespace += 1
                    if re.search(r"[\x00-\x08\x0b-\x1f]", v):
                        control += 1
        if dup_rows:
            audit.add("WARNING", "Data Quality", Worksheet=table.worksheet, TableName=table.name, Message="Duplicate complete rows found.", Actual=str(dup_rows))
        if whitespace or control:
            audit.add("WARNING", "Data Quality", Worksheet=table.worksheet, TableName=table.name, Message="Whitespace or hidden control characters found.", Actual=f"whitespace={whitespace}; control={control}")


def validate_table_placement(audit: Audit):
    for table in audit.tables.values():
        expected = ""
        if table.name in {"tblLookupGroupName", "tblLookupCatalogID", "tblCrosswalkCatalog"}:
            expected = EXPECTED_GOVERNANCE_SHEET
        elif is_physical_crosswalk_table(audit, table):
            expected = EXPECTED_CROSSWALK_SHEET
        if expected and table.worksheet != expected:
            audit.add("WARNING", "Table Placement", Worksheet=table.worksheet, TableName=table.name, Message="Table is not on recommended worksheet.", Expected=expected, Actual=table.worksheet)


def detect_stale_control_sheets(audit: Audit):
    current_tables = set(audit.tables)
    for ws in audit.wb.worksheets:
        if ws.title not in CONTROL_SHEETS:
            continue
        text = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    text.append(clean(cell.value))
        joined = "\n".join(text)
        referenced = set(re.findall(r"tbl[A-Za-z0-9_]+", joined))
        removed = sorted(referenced - current_tables)
        missing_current = sorted(current_tables - referenced)[:25] if referenced else []
        if removed or missing_current:
            audit.add("WARNING", "Control-Sheet Staleness", Worksheet=ws.title, Message="Control sheet may be stale compared with current table inventory.", Expected="Current workbook tables", Actual=f"removed_refs={removed[:10]}; missing_current_sample={missing_current[:10]}")


def workbook_identity(audit: Audit):
    wb = audit.wb
    props = wb.calculation
    hidden = [ws.title for ws in wb.worksheets if ws.sheet_state == "hidden"]
    very_hidden = [ws.title for ws in wb.worksheets if ws.sheet_state == "veryHidden"]
    defined_names = list(wb.defined_names)
    dup_names = [n for n, c in Counter(str(n) for n in defined_names).items() if c > 1]
    audit.identity = {
        "FullPath": str(audit.workbook_path),
        "Filename": audit.workbook_path.name,
        "FileSize": audit.workbook_path.stat().st_size,
        "LastModified": datetime.fromtimestamp(audit.workbook_path.stat().st_mtime).isoformat(),
        "SHA256": audit.source_hash_before,
        "CalculationMode": getattr(props, "calcMode", ""),
        "WorksheetCount": len(wb.worksheets),
        "ExcelTableCount": sum(len(ws.tables) for ws in wb.worksheets),
        "HiddenWorksheets": hidden,
        "VeryHiddenWorksheets": very_hidden,
        "DefinedNameCount": len(defined_names),
        "DuplicateDefinedNames": dup_names,
        "ExternalLinksDetected": len(getattr(wb, "_external_links", []) or []),
    }
    if dup_names:
        audit.add("ERROR", "Workbook Identity", Message="Duplicate defined names detected.", Actual="; ".join(dup_names))


def determine_readiness(audit: Audit) -> str:
    counts = Counter(f.Severity for f in audit.findings)
    if counts["BLOCKER"]:
        return "NOT READY"
    if counts["ERROR"]:
        return "NOT READY"
    if counts["WARNING"]:
        return "READY WITH WARNINGS"
    return "READY"


def write_reports(audit: Audit, readiness: str):
    audit.output_dir.mkdir(parents=True, exist_ok=True)
    def dump_csv(name, rows, headers=None):
        path = audit.output_dir / name
        rows = list(rows)
        headers = headers or (list(rows[0].keys()) if rows else [])
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    finding_headers = list(Finding.__dataclass_fields__)
    dump_csv("migration_audit_findings.csv", [asdict(f) for f in audit.findings], finding_headers)
    dump_csv("worksheet_inventory.csv", audit.worksheet_inventory)
    dump_csv("table_inventory.csv", audit.table_inventory)
    dump_csv("lookup_catalog_audit.csv", audit.lookup_catalog_audit)
    dump_csv("crosswalk_catalog_audit.csv", audit.crosswalk_catalog_audit)
    dump_csv("primary_key_audit.csv", audit.primary_key_audit)
    dump_csv("foreign_key_audit.csv", audit.foreign_key_audit)
    dump_csv("crosswalk_completeness.csv", audit.crosswalk_completeness)
    dump_csv("formula_audit.csv", audit.formula_audit)

    counts = Counter(f.Severity for f in audit.findings)
    report = {
        "identity": audit.identity,
        "readiness": readiness,
        "severity_counts": {s: counts[s] for s in SEVERITIES},
        "source_hash_before": audit.source_hash_before,
        "source_hash_after": audit.source_hash_after,
        "source_hash_unchanged": audit.source_hash_before == audit.source_hash_after,
        "findings": [asdict(f) for f in audit.findings],
    }
    (audit.output_dir / "migration_audit_report.json").write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    summary = [
        f"Workbook: {audit.workbook_path}",
        f"SHA-256: {audit.source_hash_before}",
        f"Worksheets: {audit.identity['WorksheetCount']}",
        f"Excel Tables: {audit.identity['ExcelTableCount']}",
        f"Lookup Catalog Rows: {len(audit.lookup_catalog_rows)}",
        f"Crosswalk Catalog Rows: {len(audit.crosswalk_catalog_rows)}",
        f"Physical Crosswalk Tables: {sum(1 for t in audit.tables.values() if is_physical_crosswalk_table(audit, t))}",
        *(f"{s} Count: {counts[s]}" for s in SEVERITIES),
        f"Migration Readiness: {readiness}",
        f"Output Folder: {audit.output_dir}",
    ]
    (audit.output_dir / "migration_audit_summary.txt").write_text("\n".join(summary) + "\n", encoding="utf-8")
    md = ["# Migration Readiness Audit", "", f"**Readiness:** {readiness}", "", "## Severity Counts", ""]
    md.extend(f"- {s}: {counts[s]}" for s in SEVERITIES)
    md.extend(["", "## Key Workbook Facts", "", f"- Worksheets: {audit.identity['WorksheetCount']}", f"- Excel tables: {audit.identity['ExcelTableCount']}", f"- Lookup catalog rows: {len(audit.lookup_catalog_rows)}", f"- Crosswalk catalog rows: {len(audit.crosswalk_catalog_rows)}", "", "See CSV files for detailed inventories and findings."])
    (audit.output_dir / "migration_readiness.md").write_text("\n".join(md) + "\n", encoding="utf-8")


def run_audit(workbook_path: Path, output_dir: Path | None) -> tuple[Audit, str]:
    audit = Audit(workbook_path, output_dir)
    audit.source_hash_before = calculate_sha256(audit.workbook_path)
    load_workbook_safely(audit)
    workbook_identity(audit)
    inventory_worksheets(audit)
    inventory_tables(audit)
    validate_lookup_groups(audit)
    validate_lookup_catalog(audit)
    validate_crosswalk_catalog(audit)
    validate_primary_keys(audit)
    validate_table_specific_schemas(audit)
    validate_foreign_keys(audit)
    validate_crosswalk_tables(audit)
    evaluate_crosswalk_completeness(audit)
    validate_device_model_crosswalk(audit)
    validate_display_orders(audit)
    validate_boolean_and_status(audit)
    audit_formulas(audit)
    audit_data_quality(audit)
    validate_table_placement(audit)
    detect_stale_control_sheets(audit)
    audit.source_hash_after = calculate_sha256(audit.workbook_path)
    if audit.source_hash_after != audit.source_hash_before:
        audit.add("BLOCKER", "Read-Only Safety", Message="Source workbook SHA-256 changed during audit.", Expected=audit.source_hash_before, Actual=audit.source_hash_after)
    readiness = determine_readiness(audit)
    write_reports(audit, readiness)
    return audit, readiness


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Read-only migration-readiness audit for Nocturnix workbook drafts.")
    parser.add_argument("workbook", type=Path, help="Path to the workbook to audit.")
    parser.add_argument("--output-dir", type=Path, help="Optional output directory for audit reports.")
    args = parser.parse_args(argv)
    try:
        audit, readiness = run_audit(args.workbook, args.output_dir)
        counts = Counter(f.Severity for f in audit.findings)
        lines = [
            f"Workbook: {audit.workbook_path}",
            f"SHA-256: {audit.source_hash_before}",
            f"Worksheets: {audit.identity['WorksheetCount']}",
            f"Excel Tables: {audit.identity['ExcelTableCount']}",
            f"Lookup Catalog Rows: {len(audit.lookup_catalog_rows)}",
            f"Crosswalk Catalog Rows: {len(audit.crosswalk_catalog_rows)}",
            f"Physical Crosswalk Tables: {sum(1 for t in audit.tables.values() if is_physical_crosswalk_table(audit, t))}",
            *(f"{s} Count: {counts[s]}" for s in SEVERITIES),
            f"Migration Readiness: {readiness}",
            f"Output Folder: {audit.output_dir}",
        ]
        print("\n".join(lines))
        return {"READY": 0, "READY WITH WARNINGS": 1, "NOT READY": 2}[readiness]
    except Exception as exc:
        print(f"Script execution failure: {exc}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
