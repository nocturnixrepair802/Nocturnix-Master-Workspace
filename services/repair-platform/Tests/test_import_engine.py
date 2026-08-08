"""Phase 2 tests for the isolated, shadow-only Import Engine."""

from __future__ import annotations

import sqlite3
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from import_engine import ImportEngine, ImportEngineError, ImportState, ShadowStore
from import_engine.contract import WORKSHEET_CONTRACTS
from import_engine.workbook import file_sha256

FIXED_TIME = datetime(2026, 7, 23, 12, 0, tzinfo=UTC)


def _row_for(sheet: str, index: int, headers: tuple[str, ...]) -> list[object]:
    values: dict[str, object] = {header: None for header in headers}
    if sheet == "00 - Instructions":
        values["Topic"] = f"Topic {index}"
        values["Guidance"] = f"Guidance {index}"
    elif sheet == "01 - Canonical Service Types":
        values.update(
            {
                "Proposed Canonical Service Type ID": f"STY{index:06d}",
                "Service Category": "Repair",
                "Service Type": f"Canonical Type {index}",
                "Identity Authority": "Approved release",
                "Review Status": "Pending Review",
            }
        )
    elif sheet == "02 - Service Type Aliases":
        values.update(
            {
                "Alias ID": f"STA{index:06d}",
                "Source System": "Master Services",
                "Source Field": "Repair Type",
                "Source Value": f"Alias {index}",
                "Normalized Source Value": f"alias {index}",
                "Proposed Canonical Service Type ID": "STY000001",
                "Proposed Canonical Service Type": "Canonical Type 1",
                "Alias Rule Type": "Exact Match",
                "Evidence": "Test evidence",
                "Confidence": "High",
                "Review Status": "Ready for Approval",
            }
        )
    elif sheet == "03 - Service Normalization":
        values.update(
            {
                "Service ID": f"SVC{index:06d}",
                "Service Name": f"Service {index}",
                "Current Repair Type": "Legacy Type",
                "Proposed Canonical Service Type ID": "STY000001",
                "Proposed Canonical Service Type": "Canonical Type 1",
                "Mapping Method": "Exact Match",
                "Mapping Evidence": "Test evidence",
                "Confidence": "High",
                "Review Status": "Pending Labor Review",
            }
        )
    elif sheet == "04 - Labor Normalization":
        if index <= 167:
            status = "Pending Evidence Review"
        elif index <= 178:
            status = "Pending Labor Review"
        elif index <= 263:
            status = "Pending Review"
        else:
            status = "Unresolved"
        values.update(
            {
                "Labor Standard ID": f"LAB{index:06d}",
                "Labor Name": f"Labor {index}",
                "Current Repair Type": "Legacy Type",
                "Proposed Canonical Service Type ID": "STY000001",
                "Proposed Canonical Service Type": "Canonical Type 1",
                "Mapping Method": "Exact Match",
                "Mapping Evidence": "Test evidence",
                "Confidence": "High",
                "Review Status": status,
            }
        )
    elif sheet == "06 - Unresolved Review":
        values.update(
            {
                "Record Type": "Service Normalization",
                "Source Record ID": (
                    "SVC000343" if index == 1 else f"UNR{index:06d}"
                ),
                "Source Name": "Incomplete placeholder",
                "Ambiguity Reason": "Missing evidence",
                "Missing Evidence": "Canonical operation classification",
                "Required Action": "Review",
                "Review Priority": "High" if index == 1 else "Medium",
                "Review Status": "Pending Evidence Review",
            }
        )
    elif sheet == "07 - Validation Summary":
        values.update(
            {"Validation": f"Check {index}", "Result": "PASS", "Count": 0}
        )
    elif sheet == "08 - Revision History":
        values.update(
            {
                "Version": f"review-{index}",
                "Date": FIXED_TIME.replace(tzinfo=None),
                "Change": "Test revision",
                "Status": "Pending Review",
            }
        )
    elif sheet == "09 - Import Metadata":
        values.update(
            {"Metadata Field": f"Field {index}", "Value": f"Value {index}"}
        )
    return [values[header] for header in headers]


def write_contract_workbook(path: Path) -> None:
    """Create a compact-value fixture with the exact v1.0 workbook shape."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for contract in WORKSHEET_CONTRACTS:
        worksheet = workbook.create_sheet(contract.worksheet)
        worksheet.append(list(contract.headers))
        for index in range(1, contract.expected_rows + 1):
            worksheet.append(_row_for(contract.worksheet, index, contract.headers))
        max_column = worksheet.max_column
        final_row = max(1, contract.expected_rows + 1)
        end_column = worksheet.cell(row=1, column=max_column).column_letter
        worksheet.add_table(
            Table(
                displayName=contract.table,
                ref=f"A1:{end_column}{final_row}",
            )
        )
    workbook.save(path)
    workbook.close()


def run_fixture_import(workbook_path: Path, database_path: Path):
    return ImportEngine(
        workbook_path=workbook_path,
        shadow_store=ShadowStore(database_path),
        expected_sha256=file_sha256(workbook_path),
        clock=lambda: FIXED_TIME,
    ).run()


def test_imports_exact_populations_to_shadow_tables_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)

    manifest = run_fixture_import(workbook_path, database_path)

    assert manifest.import_status is ImportState.COMPLETED
    assert manifest.activation_allowed is False
    assert manifest.runtime_records_activated == 0
    assert manifest.reconciliation_counts["service_reconciliation_total"] == 314
    with sqlite3.connect(database_path) as connection:
        counts = {
            table: connection.execute(
                f"SELECT COUNT(*) FROM {table}"  # noqa: S608 - fixed test names
            ).fetchone()[0]
            for table in (
                "shadow_canonical_service_types",
                "shadow_service_type_aliases",
                "shadow_service_normalization",
                "shadow_labor_normalization",
            )
        }
        assert counts == {
            "shadow_canonical_service_types": 77,
            "shadow_service_type_aliases": 17,
            "shadow_service_normalization": 313,
            "shadow_labor_normalization": 265,
        }
        runtime_count = connection.execute(
            """
            SELECT SUM(runtime_active)
            FROM (
                SELECT runtime_active FROM shadow_canonical_service_types
                UNION ALL
                SELECT runtime_active FROM shadow_service_type_aliases
                UNION ALL
                SELECT runtime_active FROM shadow_service_normalization
                UNION ALL
                SELECT runtime_active FROM shadow_labor_normalization
            )
            """
        ).fetchone()[0]
        assert runtime_count == 0
        production_tables = connection.execute(
            """
            SELECT name FROM sqlite_master
            WHERE type = 'table' AND name NOT LIKE 'shadow_%'
            """
        ).fetchall()
        assert production_tables == []


def test_every_shadow_row_has_provenance_and_is_inactive(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)
    manifest = run_fixture_import(workbook_path, database_path)

    with sqlite3.connect(database_path) as connection:
        row = connection.execute(
            "SELECT * FROM shadow_service_normalization WHERE service_id = ?",
            ("SVC000001",),
        ).fetchone()
        columns = [
            description[0] for description in connection.execute(
                "SELECT * FROM shadow_service_normalization LIMIT 0"
            ).description
        ]
        payload = dict(zip(columns, row, strict=True))

    assert payload["source_workbook"] == manifest.workbook_path
    assert payload["approved_version"] == "v1.0"
    assert payload["source_sha256"] == manifest.workbook_sha256
    assert payload["source_worksheet"] == "03 - Service Normalization"
    assert payload["source_row_identifier"].endswith("|SVC000001")
    assert payload["imported_at_utc"] == FIXED_TIME.isoformat()
    assert payload["source_review_status"] == "Pending Labor Review"
    assert payload["imported_status"] == "SHADOW_REFERENCE"
    assert payload["effective_at_utc"] is None
    assert payload["runtime_active"] == 0
    assert len(payload["source_row_sha256"]) == 64
    assert '"Service ID":"SVC000001"' in payload["raw_payload"]


def test_same_release_is_idempotent(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)

    first = run_fixture_import(workbook_path, database_path)
    second = run_fixture_import(workbook_path, database_path)

    assert second == first
    with sqlite3.connect(database_path) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM shadow_import_releases"
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM shadow_service_normalization"
        ).fetchone()[0] == 313


def test_idempotent_lookup_still_verifies_current_workbook_hash(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)
    approved_hash = file_sha256(workbook_path)
    run_fixture_import(workbook_path, database_path)

    with workbook_path.open("ab") as workbook:
        workbook.write(b"tampered")
    engine = ImportEngine(
        workbook_path=workbook_path,
        shadow_store=ShadowStore(database_path),
        expected_sha256=approved_hash,
    )

    with pytest.raises(ImportEngineError, match="SHA-256 mismatch"):
        engine.run()

    assert engine.state is ImportState.FAILED


def test_hash_mismatch_fails_without_creating_shadow_database(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)
    engine = ImportEngine(
        workbook_path=workbook_path,
        shadow_store=ShadowStore(database_path),
        expected_sha256="0" * 64,
    )

    with pytest.raises(ImportEngineError, match="SHA-256 mismatch"):
        engine.run()

    assert engine.state is ImportState.FAILED
    assert not database_path.exists()


def test_unknown_status_fails_closed(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["02 - Service Type Aliases"]["K2"] = "approved"
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(ImportEngineError, match="status"):
        run_fixture_import(workbook_path, database_path)

    assert not database_path.exists()


def test_svc000343_cannot_enter_service_normalization(tmp_path: Path) -> None:
    workbook_path = tmp_path / "approved.xlsx"
    database_path = tmp_path / "shadow.sqlite3"
    write_contract_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["03 - Service Normalization"]["A2"] = "SVC000343"
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(ImportEngineError, match="SVC000343"):
        run_fixture_import(workbook_path, database_path)

    assert not database_path.exists()


def test_shadow_constraints_reject_activation(tmp_path: Path) -> None:
    database_path = tmp_path / "shadow.sqlite3"
    store = ShadowStore(database_path)
    store.install_schema()
    with store.connect() as connection:
        with pytest.raises(sqlite3.IntegrityError):
            connection.execute(
                """
                INSERT INTO shadow_import_releases (
                    import_release_id, contract_version, workbook_version,
                    workbook_sha256, source_workbook, imported_at_utc,
                    import_status, activation_allowed,
                    runtime_records_activated, rollback_token
                ) VALUES ('x', '0.1', 'v1.0', ?, 'source.xlsx', ?, 'COMPLETED',
                          1, 0, 'rollback')
                """,
                ("0" * 64, FIXED_TIME.isoformat()),
            )
