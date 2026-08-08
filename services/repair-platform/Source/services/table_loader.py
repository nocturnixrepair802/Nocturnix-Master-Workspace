"""
============================================================
Nocturnix Repair Platform
Table Loader
============================================================

Author: Nocturnix Mobile Repair
Version: 1.0.0 Alpha

Purpose:
    Loads all Excel Tables from the Nocturnix Master
    Database into pandas DataFrames for use throughout
    the application.

============================================================
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table

from config.database import TABLES


class TableLoader:
    """
    Loads Excel Tables into pandas DataFrames.
    """

    def __init__(self, workbook_path: str | Path):

        self.workbook_path = Path(workbook_path)

        if not self.workbook_path.exists():

            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        self.workbook = load_workbook(self.workbook_path, data_only=True)

    # ======================================================
    # Information
    # ======================================================

    def list_tables(self) -> None:
        """
        Prints every Excel Table found in the workbook.
        """

        print()
        print("=" * 70)
        print("Workbook Tables")
        print("=" * 70)

        for worksheet in self.workbook.worksheets:

            if worksheet.tables:

                print(f"\n{worksheet.title}")

                for table in worksheet.tables.values():

                    print(f"  - {table.name}")

    # ======================================================
    # Load Single Table
    # ======================================================

    def load_table(
        self,
        table_name: str,
    ) -> pd.DataFrame:
        """
        Loads a single Excel Table into a DataFrame.
        """

        for worksheet in self.workbook.worksheets:

            if table_name not in worksheet.tables:
                continue

            table: Table = worksheet.tables[table_name]

            min_col, min_row, max_col, max_row = range_boundaries(table.ref)

            rows = list(
                worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                )
            )

            if not rows:

                return pd.DataFrame()

            headers = [cell.value for cell in rows[0]]

            records = [[cell.value for cell in row] for row in rows[1:]]

            return pd.DataFrame(
                records,
                columns=headers,
            )

        raise ValueError(
            f"Excel table '{table_name}' was not found "
            f"in workbook '{self.workbook_path.name}'."
        )

    # ======================================================
    # Load Entire Database
    # ======================================================

    def load_all_tables(
        self,
    ) -> dict[str, pd.DataFrame]:
        """
        Loads every configured table from the workbook.
        """

        data: dict[str, pd.DataFrame] = {}

        print()
        print("=" * 70)
        print("Loading Database Tables")
        print("=" * 70)

        for key, table_name in TABLES.items():

            try:

                dataframe = self.load_table(table_name)

                data[key] = dataframe

                print(f"OK  {table_name:<35}" f"{len(dataframe):>6} rows")

            except (KeyError, TypeError, ValueError) as error:

                print(f"ERR {table_name}")
                print(f"   {error}")

        print()
        print(f"Loaded {len(data)} of " f"{len(TABLES)} tables.")

        return data
