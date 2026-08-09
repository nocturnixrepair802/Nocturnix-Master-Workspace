from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table


class WorkbookWriter:
    """Persists records into configured Excel tables."""

    def __init__(
        self,
        workbook_path: str | Path,
    ) -> None:
        self.workbook_path = Path(workbook_path)

        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

    def append_to_table(
        self,
        table_name: str,
        record: dict[str, Any],
    ) -> None:
        workbook = load_workbook(
            self.workbook_path,
            keep_vba=True,
        )

        try:
            worksheet, table = self._find_table(
                workbook,
                table_name,
            )

            (
                min_col,
                min_row,
                max_col,
                max_row,
            ) = range_boundaries(table.ref)

            if min_col is None or min_row is None or max_col is None or max_row is None:
                raise ValueError(f"Table {table_name!r} has an invalid cell range.")

            headers: list[str] = []

            for column in range(
                min_col,
                max_col + 1,
            ):
                value = worksheet.cell(
                    row=min_row,
                    column=column,
                ).value

                if value is None:
                    raise ValueError(f"Table {table_name!r} contains a blank header.")

                headers.append(str(value))

            row_number = max_row + 1

            for offset, header in enumerate(headers):
                worksheet.cell(
                    row=row_number,
                    column=min_col + offset,
                    value=record.get(header),
                )

            first_cell = worksheet.cell(
                row=min_row,
                column=min_col,
            ).coordinate

            last_cell = worksheet.cell(
                row=row_number,
                column=max_col,
            ).coordinate

            table.ref = f"{first_cell}:{last_cell}"

            workbook.save(self.workbook_path)

        finally:
            workbook.close()

    @staticmethod
    def _find_table(
        workbook: Any,
        table_name: str,
    ) -> tuple[Any, Table]:
        for worksheet in workbook.worksheets:
            if table_name not in worksheet.tables:
                continue

            table = worksheet.tables[table_name]

            return worksheet, table

        raise ValueError(f"Excel table {table_name!r} was not found.")
