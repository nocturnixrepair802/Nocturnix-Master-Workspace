"""
============================================================
Nocturnix Repair Platform
Workbook Manager
============================================================

Author: Nocturnix Mobile Repair
Version: 1.0.0 Alpha

Purpose:
    Central manager responsible for opening, saving,
    closing and tracking the Nocturnix Master Database.

============================================================
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class WorkbookManager:
    """
    Central workbook manager.

    This class is the only class that should directly
    interact with the Excel workbook.
    """

    def __init__(self):

        self.workbook: Workbook | None = None

        self.file_path: Path | None = None

        self.loaded: bool = False

        self.modified: bool = False

    # ======================================================
    # Workbook Operations
    # ======================================================

    def open(
        self,
        file_path: str | Path,
    ) -> bool:
        """
        Opens an Excel workbook.
        """

        path = Path(file_path)

        if not path.exists():

            raise FileNotFoundError(path)

        self.workbook = load_workbook(path)

        self.file_path = path

        self.loaded = True

        self.modified = False

        return True

    def save(self) -> None:

        if self.workbook is None or self.file_path is None:
            return

        self.workbook.save(self.file_path)

        self.modified = False

    def save_as(
        self,
        file_path: str | Path,
    ) -> None:

        if self.workbook is None:
            return

        path = Path(file_path)

        self.workbook.save(path)

        self.file_path = path

        self.modified = False

    def close(self) -> None:

        if self.workbook is not None:

            self.workbook.close()

        self.workbook = None

        self.loaded = False

        self.modified = False

    # ======================================================
    # Workbook Information
    # ======================================================

    @property
    def worksheets(self) -> list[str]:

        workbook = self.workbook_instance

        return workbook.sheetnames

    @property
    def workbook_instance(self) -> Workbook:
        """
        Returns the loaded workbook.

        Raises
        ------
        RuntimeError
            If no workbook has been loaded.
        """

        if self.workbook is None:

            raise RuntimeError("Workbook has not been loaded.")

        return self.workbook

    @property
    def worksheet_count(self):

        return len(self.worksheets)

    # ======================================================
    # Utilities
    # ======================================================

    def mark_modified(self) -> None:

        self.modified = True

    @staticmethod
    def timestamp():

        return datetime.now()
