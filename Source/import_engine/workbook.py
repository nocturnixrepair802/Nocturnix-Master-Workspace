"""Read-only workbook validation and extraction."""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from import_engine.contract import (
    RECOGNIZED_STATUSES,
    SVC000343,
    WORKSHEET_CONTRACTS,
)
from import_engine.manifest import ValidationSummary

STY_ID = re.compile(r"^STY\d{6}$")
ALIAS_ID = re.compile(r"^STA\d{6}$")
SERVICE_ID = re.compile(r"^SVC\d{6}$")
LABOR_ID = re.compile(r"^LAB\d{6}$")


class WorkbookContractError(ValueError):
    """Raised when a workbook violates the readiness contract."""


@dataclass(frozen=True, slots=True)
class SourceRow:
    """One lossless workbook row plus deterministic provenance."""

    worksheet: str
    table: str
    excel_row: int
    source_key: str
    source_row_identifier: str
    source_status: str
    status_class: str
    imported_status: str
    row_sha256: str
    values: dict[str, Any]


@dataclass(frozen=True, slots=True)
class ValidatedWorkbook:
    """Validated workbook contents ready for manifest creation."""

    workbook_sha256: str
    worksheet_rows: dict[str, tuple[dict[str, Any], ...]]
    import_rows: dict[str, tuple[SourceRow, ...]]
    row_counts: dict[str, int]
    reconciliation_counts: dict[str, int]
    validation_summary: ValidationSummary


def file_sha256(path: Path) -> str:
    """Hash a file without loading it all into memory."""
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def canonical_row_json(values: dict[str, Any]) -> str:
    """Serialize a row deterministically while preserving blanks and field names."""
    return json.dumps(
        {key: _json_value(value) for key, value in values.items()},
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )


def _status_class(status: str) -> str:
    if status == "Approved":
        return "ACTIVE_IMPORTABLE"
    if status == "Archived":
        return "INACTIVE_REFERENCE"
    if status in {"Rejected", "Unresolved"}:
        return "REJECTED_UNRESOLVED"
    return "BLOCKED"


class ApprovedWorkbookValidator:
    """Validate and extract an approved workbook without ever saving it."""

    def __init__(self, workbook_path: str | Path, expected_sha256: str):
        self.workbook_path = Path(workbook_path)
        self.expected_sha256 = expected_sha256.upper()

    def validate(self) -> ValidatedWorkbook:
        """Return validated source data or fail closed."""
        if not self.workbook_path.is_file():
            raise WorkbookContractError(
                f"Approved workbook not found: {self.workbook_path}"
            )
        before_hash = file_sha256(self.workbook_path)
        if before_hash != self.expected_sha256:
            raise WorkbookContractError(
                "Workbook SHA-256 mismatch: "
                f"expected {self.expected_sha256}, observed {before_hash}"
            )

        table_refs = self._validate_ooxml_structure()
        worksheet_rows = self._read_table_rows(table_refs)
        checks = self._validate_contract_rows(worksheet_rows)

        after_hash = file_sha256(self.workbook_path)
        if after_hash != before_hash:
            raise WorkbookContractError("Workbook changed while it was being read")

        import_rows = self._build_import_rows(worksheet_rows)
        row_counts = {
            contract.worksheet: len(worksheet_rows[contract.worksheet])
            for contract in WORKSHEET_CONTRACTS
        }
        reconciliation = {
            "source_service_records": 314,
            "service_normalization_rows": row_counts[
                "03 - Service Normalization"
            ],
            "explicit_service_exclusions": 1,
            "service_reconciliation_total": (
                row_counts["03 - Service Normalization"] + 1
            ),
            "runtime_records_activated": 0,
        }
        return ValidatedWorkbook(
            workbook_sha256=before_hash,
            worksheet_rows=worksheet_rows,
            import_rows=import_rows,
            row_counts=row_counts,
            reconciliation_counts=reconciliation,
            validation_summary=ValidationSummary(
                result="PASS",
                checks_passed=checks,
                checks_failed=0,
                messages=(
                    "Workbook hash, structure, counts, statuses, keys, and "
                    "SVC000343 exclusion reconciled.",
                    "All behavior-bearing rows are classified for shadow import only.",
                ),
            ),
        )

    def _validate_ooxml_structure(self) -> dict[str, str]:
        try:
            with ZipFile(self.workbook_path) as archive:
                names = set(archive.namelist())
                if "xl/vbaProject.bin" in names:
                    raise WorkbookContractError("Macros are not allowed")
                if any(name.startswith("xl/externalLinks/") for name in names):
                    raise WorkbookContractError("External links are not allowed")
                return self._table_refs_from_ooxml(archive)
        except BadZipFile as error:
            raise WorkbookContractError("Workbook is not valid OOXML") from error

    @staticmethod
    def _relationships(
        archive: ZipFile, path: str, base: str
    ) -> dict[str, str]:
        package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        root = ElementTree.fromstring(archive.read(path))
        result: dict[str, str] = {}
        for relationship in root.findall(f"{{{package_ns}}}Relationship"):
            target = relationship.attrib["Target"].replace("\\", "/")
            if target.startswith("/"):
                resolved = target.removeprefix("/")
            else:
                parts = base.rstrip("/").split("/") if base else []
                for component in target.split("/"):
                    if component == "..":
                        parts.pop()
                    elif component not in {"", "."}:
                        parts.append(component)
                resolved = "/".join(parts)
            result[relationship.attrib["Id"]] = resolved
        return result

    def _table_refs_from_ooxml(self, archive: ZipFile) -> dict[str, str]:
        spreadsheet_ns = (
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        )
        document_rel_ns = (
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        )
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels = self._relationships(
            archive, "xl/_rels/workbook.xml.rels", "xl"
        )
        sheet_nodes = workbook_root.findall(
            f".//{{{spreadsheet_ns}}}sheet"
        )
        sheet_names = [node.attrib["name"] for node in sheet_nodes]
        expected_sheets = [
            contract.worksheet for contract in WORKSHEET_CONTRACTS
        ]
        if sheet_names != expected_sheets:
            raise WorkbookContractError(
                "Worksheet names or order do not match the contract"
            )

        table_refs: dict[str, str] = {}
        for node, contract in zip(
            sheet_nodes, WORKSHEET_CONTRACTS, strict=True
        ):
            relationship_id = node.attrib[f"{{{document_rel_ns}}}id"]
            sheet_path = workbook_rels[relationship_id]
            sheet_directory, sheet_filename = sheet_path.rsplit("/", 1)
            sheet_rels_path = (
                f"{sheet_directory}/_rels/{sheet_filename}.rels"
            )
            sheet_rels = self._relationships(
                archive, sheet_rels_path, sheet_directory
            )
            sheet_root = ElementTree.fromstring(archive.read(sheet_path))
            table_parts = sheet_root.findall(
                f".//{{{spreadsheet_ns}}}tablePart"
            )
            if len(table_parts) != 1:
                raise WorkbookContractError(
                    f"{contract.worksheet} must contain exactly one table"
                )
            table_relationship_id = table_parts[0].attrib[
                f"{{{document_rel_ns}}}id"
            ]
            table_path = sheet_rels[table_relationship_id]
            table_root = ElementTree.fromstring(archive.read(table_path))
            if table_root.attrib.get("name") != contract.table:
                raise WorkbookContractError(
                    f"{contract.worksheet} must contain only {contract.table}"
                )
            table_refs[contract.worksheet] = table_root.attrib["ref"]
        return table_refs

    def _read_table_rows(
        self, table_refs: dict[str, str]
    ) -> dict[str, tuple[dict[str, Any], ...]]:
        workbook = load_workbook(
            self.workbook_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            result: dict[str, tuple[dict[str, Any], ...]] = {}
            for contract in WORKSHEET_CONTRACTS:
                worksheet = workbook[contract.worksheet]
                min_col, min_row, max_col, max_row = range_boundaries(
                    table_refs[contract.worksheet]
                )
                values = worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
                headers = tuple(next(values))
                if headers != contract.headers:
                    raise WorkbookContractError(
                        f"{contract.worksheet} headers do not match the contract"
                    )
                rows: tuple[dict[str, Any], ...] = tuple(
                    dict(zip(contract.headers, row, strict=True))
                    for row in values
                    if any(value is not None for value in row)
                )
                result[contract.worksheet] = rows
            return result
        finally:
            workbook.close()

    def _validate_contract_rows(
        self, worksheet_rows: dict[str, tuple[dict[str, Any], ...]]
    ) -> int:
        checks = 0
        canonical_pairs: dict[str, str] = {}
        for contract in WORKSHEET_CONTRACTS:
            rows = worksheet_rows[contract.worksheet]
            if len(rows) != contract.expected_rows:
                raise WorkbookContractError(
                    f"{contract.worksheet} row count is {len(rows)}; "
                    f"expected {contract.expected_rows}"
                )
            checks += 1
            if contract.status_field:
                statuses = [row[contract.status_field] for row in rows]
                if any(
                    not isinstance(status, str)
                    or status not in RECOGNIZED_STATUSES
                    for status in statuses
                ):
                    raise WorkbookContractError(
                        f"{contract.worksheet} contains an unknown or blank status"
                    )
                expected = dict(contract.expected_status_counts)
                if expected and Counter(statuses) != Counter(expected):
                    raise WorkbookContractError(
                        f"{contract.worksheet} status counts do not match the contract"
                    )
                checks += 1

        canonical = worksheet_rows["01 - Canonical Service Types"]
        self._validate_unique_keys(
            canonical, "Proposed Canonical Service Type ID", STY_ID
        )
        canonical_pairs = {
            str(row["Proposed Canonical Service Type ID"]): str(row["Service Type"])
            for row in canonical
        }
        self._validate_unique_keys(
            worksheet_rows["02 - Service Type Aliases"], "Alias ID", ALIAS_ID
        )
        self._validate_unique_keys(
            worksheet_rows["03 - Service Normalization"],
            "Service ID",
            SERVICE_ID,
        )
        self._validate_unique_keys(
            worksheet_rows["04 - Labor Normalization"],
            "Labor Standard ID",
            LABOR_ID,
        )
        checks += 4

        for sheet in (
            "02 - Service Type Aliases",
            "03 - Service Normalization",
            "04 - Labor Normalization",
        ):
            for row in worksheet_rows[sheet]:
                proposed_id = row["Proposed Canonical Service Type ID"]
                proposed_type = row["Proposed Canonical Service Type"]
                if (proposed_id is None) != (proposed_type is None):
                    raise WorkbookContractError(
                        f"{sheet} contains an incomplete canonical ID/type pair"
                    )
                if proposed_id is not None and canonical_pairs.get(
                    str(proposed_id)
                ) != str(proposed_type):
                    raise WorkbookContractError(
                        f"{sheet} contains a broken canonical reference"
                    )
        checks += 1

        services = worksheet_rows["03 - Service Normalization"]
        if any(row["Service ID"] == SVC000343 for row in services):
            raise WorkbookContractError(
                "SVC000343 must be excluded from Service Normalization"
            )
        unresolved = [
            row
            for row in worksheet_rows["06 - Unresolved Review"]
            if row["Record Type"] == "Service Normalization"
            and row["Source Record ID"] == SVC000343
        ]
        if len(unresolved) != 1:
            raise WorkbookContractError(
                "SVC000343 must appear exactly once in Unresolved Review"
            )
        svc_row = unresolved[0]
        if (
            svc_row["Review Priority"] != "High"
            or svc_row["Review Status"] != "Pending Evidence Review"
        ):
            raise WorkbookContractError(
                "SVC000343 unresolved-review disposition is invalid"
            )
        checks += 1
        return checks

    @staticmethod
    def _validate_unique_keys(
        rows: tuple[dict[str, Any], ...],
        field: str,
        pattern: re.Pattern[str],
    ) -> None:
        keys = [row[field] for row in rows]
        if any(not isinstance(key, str) or not pattern.fullmatch(key) for key in keys):
            raise WorkbookContractError(f"{field} contains an invalid identifier")
        if len(keys) != len(set(keys)):
            raise WorkbookContractError(f"{field} contains duplicate identifiers")

    def _build_import_rows(
        self, worksheet_rows: dict[str, tuple[dict[str, Any], ...]]
    ) -> dict[str, tuple[SourceRow, ...]]:
        result: dict[str, tuple[SourceRow, ...]] = {}
        for contract in WORKSHEET_CONTRACTS:
            if not contract.import_target or not contract.key_field:
                continue
            rows: list[SourceRow] = []
            for offset, values in enumerate(
                worksheet_rows[contract.worksheet], start=2
            ):
                source_key = str(values[contract.key_field])
                status = str(values[contract.status_field or "Review Status"])
                raw_json = canonical_row_json(values)
                rows.append(
                    SourceRow(
                        worksheet=contract.worksheet,
                        table=contract.table,
                        excel_row=offset,
                        source_key=source_key,
                        source_row_identifier=(
                            f"{contract.worksheet}|{source_key}"
                        ),
                        source_status=status,
                        status_class=_status_class(status),
                        imported_status="SHADOW_REFERENCE",
                        row_sha256=hashlib.sha256(
                            raw_json.encode("utf-8")
                        ).hexdigest().upper(),
                        values=values,
                    )
                )
            result[contract.worksheet] = tuple(rows)
        return result
