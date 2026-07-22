"""
============================================================
Nocturnix Repair Platform
Worksheet Manager
============================================================

Author: Nocturnix Mobile Repair
Version: 1.0.0 Alpha

Purpose:
    Handles worksheet operations within the
    Nocturnix Master Database workbook.
============================================================
"""

from openpyxl.worksheet.worksheet import Worksheet

from managers.workbook_manager import WorkbookManager


class WorksheetManager:
    """
    Handles worksheet operations.
    """

    def __init__(
        self,
        workbook_manager: WorkbookManager,
    ):

        self.workbook_manager = workbook_manager

    # ======================================================
    # Worksheet Retrieval
    # ======================================================

    def get(
        self,
        sheet_name: str,
    ) -> Worksheet:

        workbook = self.workbook_manager.workbook_instance

        return workbook[sheet_name]

    def exists(
        self,
        sheet_name: str,
    ) -> bool:

        workbook = self.workbook_manager.workbook_instance

        return sheet_name in workbook.sheetnames

    def names(self) -> list[str]:

        workbook = self.workbook_manager.workbook_instance

        return workbook.sheetnames

    def count(self) -> int:

        return len(self.names())

    # ======================================================
    # Worksheet Creation
    # ======================================================

    def create(
        self,
        sheet_name: str,
    ) -> None:

        workbook = self.workbook_manager.workbook_instance

        if self.exists(sheet_name):
            return

        workbook.create_sheet(sheet_name)

        self.workbook_manager.mark_modified()

    # ======================================================
    # Worksheet Deletion
    # ======================================================

    def delete(
        self,
        sheet_name: str,
    ) -> None:

        workbook = self.workbook_manager.workbook_instance

        if not self.exists(sheet_name):
            return

        sheet = self.get(sheet_name)

        workbook.remove(sheet)

        self.workbook_manager.mark_modified()

    # ======================================================
    # Worksheet Rename
    # ======================================================

    def rename(
        self,
        old_name: str,
        new_name: str,
    ) -> None:

        sheet = self.get(old_name)

        sheet.title = new_name

        self.workbook_manager.mark_modified()
