from openpyxl import load_workbook


class WorkbookInspector:

    def __init__(self, workbook_path):
        self.workbook_path = workbook_path

    def inspect(self):

        wb = load_workbook(
            self.workbook_path,
            read_only=True,
            data_only=True
        )

        print("=" * 70)
        print("Nocturnix Workbook Inspector")
        print("=" * 70)

        print(f"Workbook: {self.workbook_path.name}")
        print(f"Worksheets: {len(wb.sheetnames)}")

        print("\nWorksheet List")
        print("-" * 70)

        for sheet in wb.sheetnames:

            ws = wb[sheet]

            print(
                f"{sheet:<35}"
                f"Rows: {ws.max_row:<6}"
                f"Columns: {ws.max_column}"
            )