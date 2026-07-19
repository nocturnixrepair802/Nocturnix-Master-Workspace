from openpyxl import load_workbook
from config import MASTER_DATABASE

wb = load_workbook(MASTER_DATABASE)

print("Workbook loaded successfully!")
print()

print("Worksheets:")

for sheet in wb.sheetnames:
    print(f" - {sheet}")