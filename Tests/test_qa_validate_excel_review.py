"""Focused tests for the reusable Excel review QA validator."""

from __future__ import annotations

import importlib.util
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "Scripts"
    / "qa_validate_excel_review.py"
)
SPEC = importlib.util.spec_from_file_location("qa_validate_excel_review", SCRIPT_PATH)
assert SPEC is not None and SPEC.loader is not None
qa = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = qa
SPEC.loader.exec_module(qa)


def base_config() -> dict[str, object]:
    """Return a compact profile exercising the reusable rule engine."""
    return {
        "profile_name": "Test review",
        "expected_sheets": ["Review"],
        "thresholds": {
            "max_errors_for_pass": 0,
            "max_warnings_for_pass": 0,
            "max_errors_for_warning_status": 0,
        },
        "sheets": {
            "Review": {
                "header_row": 1,
                "expected_headers": ["ID", "Source", "Target", "Status"],
                "required_fields": ["ID", "Source", "Status"],
                "duplicate_keys": [["Source", "Target"]],
                "freeze_panes": "A2",
                "require_filter": True,
                "merged_cells": "error",
                "hidden": "warn",
                "validation_fields": [],
                "business_rules": [
                    {
                        "type": "regex",
                        "fields": ["ID"],
                        "pattern": "^REV\\d{3}$",
                        "code": "TEST_ID",
                        "message": "Invalid review ID.",
                    },
                    {
                        "type": "allowed_values",
                        "fields": ["Status"],
                        "values": ["Pending", "Approved"],
                        "code": "TEST_STATUS",
                        "message": "Invalid status.",
                    },
                ],
            }
        },
    }


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    """Create a temporary fixture workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Review"
    worksheet.append(["ID", "Source", "Target", "Status"])
    for row in rows:
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    worksheet.add_table(Table(displayName="tblReview", ref=f"A1:D{len(rows) + 1}"))
    workbook.save(path)
    workbook.close()


def test_clean_workbook_passes_and_remains_unchanged(tmp_path: Path) -> None:
    """A clean workbook passes and its bytes are not modified."""
    path = tmp_path / "clean.xlsx"
    write_workbook(
        path,
        [
            ["REV001", "Screen Repair", "Screen Replacement", "Pending"],
            ["REV002", "Battery", "Battery Replacement", "Approved"],
        ],
    )
    before = qa.file_hash(path)
    result = qa.validate_workbook(path, base_config())
    assert result.status == "PASS"
    assert result.counts == {"ERROR": 0, "WARNING": 0, "INFO": 0}
    assert qa.file_hash(path) == before


def test_errors_and_warnings_are_reported(tmp_path: Path) -> None:
    """Required, duplicate, rule, whitespace, hidden, and merge issues fail."""
    path = tmp_path / "issues.xlsx"
    write_workbook(
        path,
        [
            ["bad", " Screen Repair", "Screen Replacement", "Unknown"],
            ["REV002", "Screen Repair", "Screen Replacement", ""],
        ],
    )
    workbook = qa.load_workbook(path)
    worksheet = workbook["Review"]
    worksheet.row_dimensions[2].hidden = True
    worksheet.merge_cells("E1:F1")
    workbook.save(path)
    workbook.close()
    result = qa.validate_workbook(path, base_config())
    codes = {finding.code for finding in result.findings}
    assert result.status == "FAIL"
    assert "DATA_REQUIRED_BLANK" in codes
    assert "DATA_DUPLICATE_MAPPING" in codes
    assert "TEXT_SURROUNDING_WHITESPACE" in codes
    assert "STRUCTURE_HIDDEN" in codes
    assert "STRUCTURE_MERGED_CELLS" in codes
    assert "TEST_ID" in codes
    assert "TEST_STATUS" in codes


def test_warning_threshold_returns_pass_with_warnings(tmp_path: Path) -> None:
    """Warnings alone produce the intermediate automation status."""
    path = tmp_path / "warning.xlsx"
    write_workbook(
        path,
        [["REV001", " Screen Repair ", "Screen Replacement", "Pending"]],
    )
    result = qa.validate_workbook(path, base_config())
    assert result.status == "PASS WITH WARNINGS"
    assert result.counts["ERROR"] == 0
    assert result.counts["WARNING"] == 1


def test_markdown_and_json_payloads_are_automation_ready(tmp_path: Path) -> None:
    """Report renderers include final status and structured findings."""
    path = tmp_path / "report.xlsx"
    write_workbook(path, [["REV001", "Screen", "Screen", "Pending"]])
    result = qa.validate_workbook(path, base_config())
    markdown = qa.markdown_report(result)
    payload = result.to_dict()
    assert "# Workbook QA Report" in markdown
    assert "## Final status: PASS" in markdown
    assert payload["status"] == "PASS"
    assert payload["workbook_sha256"] == qa.file_hash(path)


def test_sheet_check_switches_disable_individual_validation_families(
    tmp_path: Path,
) -> None:
    """Each requested validation family can be disabled per worksheet."""
    path = tmp_path / "disabled.xlsx"
    write_workbook(path, [["bad", "", "Screen", "Unknown"]])
    config = base_config()
    sheet = config["sheets"]["Review"]
    sheet["expected_headers"] = ["Wrong"]
    sheet["expected_data_rows"] = 99
    sheet["freeze_panes"] = "B2"
    sheet["validation_fields"] = ["Status"]
    sheet["checks"] = {
        "required_fields": False,
        "headers": False,
        "row_count": False,
        "duplicates": False,
        "business_rules": False,
        "freeze_panes": False,
        "data_validations": False,
    }
    result = qa.validate_workbook(path, config)
    assert result.status == "PASS"
    assert result.findings == []


def test_table_scope_ignores_formatted_and_trailing_blank_rows(
    tmp_path: Path,
) -> None:
    """Only populated Table body rows are data unless worksheet scope is set."""
    path = tmp_path / "table_scope.xlsx"
    write_workbook(path, [["REV001", "Screen", "Screen", "Pending"]])
    workbook = qa.load_workbook(path)
    worksheet = workbook["Review"]
    worksheet["A20"].fill = copy(worksheet["A1"].fill)
    worksheet["A21"] = "bad"
    workbook.save(path)
    workbook.close()

    table_result = qa.validate_workbook(path, base_config())
    assert table_result.status == "PASS"
    assert table_result.metrics["sheets"]["Review"]["rows"] == 1

    worksheet_config = base_config()
    worksheet_config["sheets"]["Review"]["data_scope"] = "worksheet"
    worksheet_result = qa.validate_workbook(path, worksheet_config)
    assert worksheet_result.status == "FAIL"
    assert any(
        finding.code == "DATA_REQUIRED_BLANK"
        for finding in worksheet_result.findings
    )


def test_finding_summary_separates_issue_categories(tmp_path: Path) -> None:
    """Automation output separates profile, workbook, and rule findings."""
    path = tmp_path / "categories.xlsx"
    write_workbook(path, [["bad", " Screen ", "Target", "Pending"]])
    config = base_config()
    config["sheets"]["Review"]["checks"] = {"unknown": True}
    result = qa.validate_workbook(path, config)
    assert result.category_counts == {
        "CONFIGURATION": 1,
        "WORKBOOK": 1,
        "BUSINESS_RULE": 1,
    }
    markdown = qa.markdown_report(result)
    assert "## Configuration issues" in markdown
    assert "## Workbook issues" in markdown
    assert "## Business-rule issues" in markdown
