from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from audit_migration_draft import calculate_sha256
from remediate_migration_draft import BLOCKED, DECISION, SAFE, build_crosswalk_mapping, create_plan, run


class Args:
    def __init__(self, workbook_path, output_dir, apply=False):
        self.workbook_path = Path(workbook_path)
        self.output_dir = Path(output_dir)
        self.apply = apply
        self.plan_only = not apply
        self.refresh_control_sheets = False
        self.approve_schema_additions = False
        self.audit_script = Path("audit_migration_draft.py")
        self.verbose = False


def add_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def make_book(path: Path, duplicate_mapping=False, include_mapping=True, target_collision=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "15 LL_Geographic "
    ws.append(["GeoID"])
    ws.append(["GEO000001"])
    add_table(ws, "tblGeo", "A1:A2")
    cat = wb.create_sheet("15 LL_Catalog")
    cat.append(["LookupCatalogID", "LookupGroupID", "LookupTableName", "DisplayName", "PrimaryKeyField", "PrimaryKeyPrefix", "Description", "DisplayOrder", "Status", "IsSystemLookup", "TableExists", "LastReviewed"])
    cat.append(["LKC000001", "LKG000001", "tblInventoryLocations", "Inventory Locations", "InventoryLocationID", "ILOC", "", 1, "Active", "No", "No", ""])
    add_table(cat, "tblLookupCatalogID", "A1:L2")
    dev = wb.create_sheet("15 LL_Device")
    dev.append(["ManufacturerID", "Manufacturer"])
    dev.append(["MFG000001", "Apple"])
    dev.append(["MFG000002", "Samsung"])
    add_table(dev, "tblManufacturers", "A1:B3")
    dev["D1"] = "DeviceFamilyID"; dev["E1"] = "Device Family"
    dev["D2"] = "DF000001"; dev["E2"] = "iPhone"
    add_table(dev, "tblDeviceFamilies", "D1:E2")
    dev["G1"] = "DeviceModelID"; dev["H1"] = "ManufacturerID"; dev["I1"] = "DeviceFamilyID"
    dev["G2"] = "MOD000001"; dev["H2"] = "MFG001"; dev["I2"] = "DF001"
    add_table(dev, "tblDeviceModels", "G1:I2")
    inv = wb.create_sheet("15 LL_Inventory")
    inv.append(["InventoryStatusID", "PartStockStatusID"])
    inv.append(["INV000001", "PST000010; PST000011"])
    add_table(inv, "tblInventoryStatus", "A1:B2")
    inv["D1"] = "PartStockStatusID"; inv["E1"] = "Status"
    inv["D2"] = "PST000010"; inv["E2"] = "Available"
    inv["D3"] = "PST000011"; inv["E3"] = "Reserved"
    add_table(inv, "tblPartStockStatus", "D1:E3")
    xw = wb.create_sheet("19 - ID Crosswalks")
    xw.append(["LegacyManufacturerID", "CanonicalManufacturerID", "MappingStatus"])
    if include_mapping:
        xw.append(["MFG001", "MFG000001", "Approved"])
        if duplicate_mapping:
            xw.append(["MFG001", "MFG000002", "Approved"])
    else:
        xw.append(["MFG999", "MFG000001", "Approved"])
    add_table(xw, "tblManufacturerIDCrosswalk", f"A1:C{3 if duplicate_mapping else 2}")
    xw["E1"] = "LegacyDeviceFamilyID"; xw["F1"] = "CanonicalDeviceFamilyID"; xw["G1"] = "MappingStatus"
    xw["E2"] = "DF001"; xw["F2"] = "DF000001"; xw["G2"] = "Approved"
    add_table(xw, "tblDeviceFamilyIDCrosswalk", "E1:G2")
    wb.save(path)


def test_unique_approved_manufacturer_mapping_safe(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    wb = load_workbook(path)
    mapping, _ = build_crosswalk_mapping(wb, "tblManufacturerIDCrosswalk", ["LegacyManufacturerID"], ["CanonicalManufacturerID"], "tblManufacturers", "ManufacturerID")
    assert mapping["MFG001"]["classification"] == SAFE


def test_missing_manufacturer_mapping_blocked(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path, include_mapping=False)
    wb = load_workbook(path)
    actions, *_ = create_plan(wb)
    assert any(a.FindingID == "F00005" and a.Classification == BLOCKED for a in actions)


def test_multiple_manufacturer_mappings_blocked(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path, duplicate_mapping=True)
    wb = load_workbook(path)
    actions, *_ = create_plan(wb)
    assert any(a.FindingID == "F00005" and a.Classification == BLOCKED for a in actions)


def test_device_family_mapping_safe(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    wb = load_workbook(path)
    actions, *_ = create_plan(wb)
    assert any(a.FindingID == "F00006" and a.Classification == SAFE for a in actions)


def test_multivalue_part_stock_status_needs_decision(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    wb = load_workbook(path)
    actions, *_ = create_plan(wb)
    assert any(a.FindingID == "F00007" and a.Classification == DECISION for a in actions)


def test_plan_only_creates_no_workbook_and_source_unchanged(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    before = calculate_sha256(path)
    outdir, output, _, _, _ = run(Args(path, tmp_path / "plan", apply=False))
    assert output == ""
    assert not list(outdir.glob("*remediated*.xlsx"))
    assert calculate_sha256(path) == before


def test_refresh_control_sheets_flag_does_not_apply_without_apply(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    args = Args(path, tmp_path / "plan_flag", apply=False)
    args.refresh_control_sheets = True
    outdir, output, _, _, applied = run(args)
    assert output == ""
    assert applied == 0


def test_apply_writes_new_workbook_and_source_unchanged(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    before = calculate_sha256(path)
    outdir, output, _, _, applied = run(Args(path, tmp_path / "apply", apply=True))
    assert Path(output).exists()
    assert Path(output) != path
    assert applied >= 1
    assert calculate_sha256(path) == before


def test_worksheet_rename_refuses_collision(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    wb = load_workbook(path)
    wb.create_sheet("15 LL_Geographic")
    wb.save(path)
    actions, *_ = create_plan(load_workbook(path))
    assert any(a.FindingID == "F00001" and a.Classification == BLOCKED for a in actions)


def test_registry_refresh_requires_flag(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    wb = load_workbook(path)
    wb.create_sheet("20 - Lookup Registry")["A1"] = "tblOldTable"
    wb.save(path)
    actions, *_ = create_plan(load_workbook(path))
    assert any(a.FindingID == "F00020" and a.Classification == DECISION for a in actions)


def test_device_model_crosswalk_not_populated(tmp_path):
    path = tmp_path / "book.xlsx"
    make_book(path)
    outdir, output, _, _, _ = run(Args(path, tmp_path / "apply2", apply=True))
    if output:
        wb = load_workbook(output)
        assert "tblDeviceModelIDCrosswalk" not in {t.name for ws in wb.worksheets for t in ws.tables.values()}
