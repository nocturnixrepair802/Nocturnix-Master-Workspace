from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from audit_migration_draft import Audit, Finding, determine_readiness, run_audit


def add_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def save_minimal_governance_book(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "15 LL_Catalog"
    ws.append(["LookupGroupID", "GroupName", "DisplayOrder"])
    ws.append(["LKG000001", "Device", 1])
    ws.append(["LKG000002", "Other", 2])
    add_table(ws, "tblLookupGroupName", "A1:C3")
    ws.append([])
    ws.append(["LookupCatalogID", "LookupGroupID", "LookupTableName", "DisplayName", "PrimaryKeyField", "PrimaryKeyPrefix", "Description", "DisplayOrder", "Status", "IsSystemLookup", "TableExists", "LastReviewed"])
    ws.append(["LKC000001", "LKG000001", "tblDeviceTypes", "Device Types", "DeviceTypeID", "DT", "", 1, "Active", "No", "Yes", "2026-01-01"])
    ws.append(["LKC000002", "LKG000002", "tblOtherTypes", "Other Types", "OtherTypeID", "OT", "", 1, "Active", "No", "Yes", "2026-01-01"])
    add_table(ws, "tblLookupCatalogID", "A5:L7")
    ws.append([])
    ws.append(["CrosswalkCatalogID", "CrosswalkTableName", "Status", "IsProductionAuthorized", "LegacyKeyField", "CanonicalKeyField"])
    ws.append(["XWC000001", "tblDeviceTypeIDCrosswalk", "Draft", "No", "LegacyDeviceTypeID", "CanonicalDeviceTypeID"])
    add_table(ws, "tblCrosswalkCatalog", "A9:F10")
    dev = wb.create_sheet("15 LL_Device")
    dev.append(["DeviceTypeID", "Device Type", "DisplayOrder", "IsActive"])
    dev.append(["DT000001", "Phone", 1, "Yes"])
    add_table(dev, "tblDeviceTypes", "A1:D2")
    other = wb.create_sheet("15 LL_Other")
    other.append(["OtherTypeID", "Other Type", "DisplayOrder", "IsActive"])
    other.append(["OT000001", "Other", 1, "Yes"])
    add_table(other, "tblOtherTypes", "A1:D2")
    xw = wb.create_sheet("19 - ID Crosswalks")
    xw.append(["LegacyDeviceTypeID", "CanonicalDeviceTypeID", "MappingStatus"])
    xw.append(["DT001", "DT000001", "Approved"])
    xw.append([None, None, None])
    add_table(xw, "tblDeviceTypeIDCrosswalk", "A1:C3")
    wb.save(path)


def test_error_findings_produce_not_ready():
    audit = Audit(Path("dummy.xlsx"))
    audit.findings.append(Finding("F1", "ERROR", "x"))
    assert determine_readiness(audit) == "NOT READY"


def test_blocker_findings_produce_not_ready():
    audit = Audit(Path("dummy.xlsx"))
    audit.findings.append(Finding("F1", "BLOCKER", "x"))
    assert determine_readiness(audit) == "NOT READY"


def test_warning_only_produces_ready_with_warnings():
    audit = Audit(Path("dummy.xlsx"))
    audit.findings.append(Finding("F1", "WARNING", "x"))
    assert determine_readiness(audit) == "READY WITH WARNINGS"


def test_no_findings_produces_ready():
    audit = Audit(Path("dummy.xlsx"))
    assert determine_readiness(audit) == "READY"


def test_tbl_crosswalk_catalog_excluded_and_placement_passes(tmp_path):
    path = tmp_path / "book.xlsx"
    save_minimal_governance_book(path)
    audit, _ = run_audit(path, tmp_path / "out")
    messages = [f.Message for f in audit.findings if f.TableName == "tblCrosswalkCatalog"]
    assert not any("blank legacy or canonical" in m.lower() for m in messages)
    assert not any("not on recommended worksheet" in m.lower() for m in messages)


def test_display_order_duplicates_across_groups_pass(tmp_path):
    path = tmp_path / "book.xlsx"
    save_minimal_governance_book(path)
    audit, _ = run_audit(path, tmp_path / "out")
    assert not any(f.Category == "Display Order Validation" and "duplicates=[1]" in f.Actual for f in audit.findings)


def test_display_order_duplicates_within_group_fail(tmp_path):
    path = tmp_path / "book.xlsx"
    save_minimal_governance_book(path)
    wb = Workbook()
    save_minimal_governance_book(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["15 LL_Catalog"]
    ws["B7"] = "LKG000001"
    ws["H7"] = 1
    wb.save(path)
    audit, _ = run_audit(path, tmp_path / "out")
    assert any(f.Category == "Lookup Catalog Validation" and "Duplicate DisplayOrder" in f.Message for f in audit.findings)


def test_formula_backed_boolean_cached_values_do_not_flag_formula_text(tmp_path):
    path = tmp_path / "book.xlsx"
    save_minimal_governance_book(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["15 LL_Catalog"]
    ws["K6"] = '=IF(1=1,"Yes","No")'
    wb.save(path)
    audit, _ = run_audit(path, tmp_path / "out")
    assert not any(f.FieldName == "TableExists" and "invalid values" in f.Message and "IF(" in f.Actual for f in audit.findings)


def test_blank_crosswalk_rows_ignored_and_partial_rows_error(tmp_path):
    path = tmp_path / "book.xlsx"
    save_minimal_governance_book(path)
    audit, _ = run_audit(path, tmp_path / "out1")
    assert not any(f.TableName == "tblDeviceTypeIDCrosswalk" and "blank legacy" in f.Message.lower() for f in audit.findings)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["19 - ID Crosswalks"]
    ws["A3"] = "DT002"
    wb.save(path)
    audit, _ = run_audit(path, tmp_path / "out2")
    assert any(f.TableName == "tblDeviceTypeIDCrosswalk" and "blank legacy or canonical" in f.Message.lower() for f in audit.findings)
