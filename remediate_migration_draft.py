from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from audit_migration_draft import calculate_sha256, run_audit


SAFE = "SAFE_AUTOMATIC"
DECISION = "NEEDS_DECISION"
BLOCKED = "BLOCKED"
AUDIT_ONLY = "AUDIT_RULE_ONLY"


@dataclass
class PlanAction:
    ActionID: str
    FindingID: str
    Classification: str
    Category: str
    Worksheet: str = ""
    TableName: str = ""
    CellOrRange: str = ""
    RecordID: str = ""
    FieldName: str = ""
    OriginalValue: str = ""
    ProposedValue: str = ""
    MappingSource: str = ""
    Reason: str = ""
    Risk: str = ""
    WillApply: str = "No"
    Applied: str = "No"
    Result: str = "Planned"
    Notes: str = ""


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def sha256(path: Path) -> str:
    return calculate_sha256(path)


def read_table(ws, table):
    cells = list(ws[table.ref])
    headers = [clean(c.value) for c in cells[0]]
    rows = []
    for row in cells[1:]:
        rows.append({"__rownum": row[0].row, "__cells": row, **{h: row[i].value for i, h in enumerate(headers)}})
    return headers, rows


def table_map(wb):
    out = {}
    for ws in wb.worksheets:
        for t in ws.tables.values():
            out[t.name] = (ws, t)
    return out


def locate_catalog(wb):
    for ws, table in table_map(wb).values():
        headers, rows = read_table(ws, table)
        if {"LookupCatalogID", "LookupTableName", "PrimaryKeyField", "PrimaryKeyPrefix"}.issubset(headers):
            return ws, table, headers, rows
    return None, None, [], []


def find_col(headers, *candidates):
    lower = {h.lower(): h for h in headers}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    return ""


def build_crosswalk_mapping(wb, table_name, legacy_candidates, canonical_candidates, target_table, target_key):
    tables = table_map(wb)
    if table_name not in tables or target_table not in tables:
        return {}, {("*", "*"): BLOCKED}
    ws, t = tables[table_name]
    headers, rows = read_table(ws, t)
    legacy_col = find_col(headers, *legacy_candidates)
    canonical_col = find_col(headers, *canonical_candidates)
    status_col = find_col(headers, "MappingStatus", "Status")
    if not legacy_col or not canonical_col:
        return {}, {("*", "*"): BLOCKED}
    target_ws, target_t = tables[target_table]
    target_headers, target_rows = read_table(target_ws, target_t)
    if target_key not in target_headers:
        return {}, {("*", "*"): BLOCKED}
    target_ids = {clean(r.get(target_key)) for r in target_rows if clean(r.get(target_key))}
    raw = defaultdict(list)
    for r in rows:
        legacy = clean(r.get(legacy_col))
        canon = clean(r.get(canonical_col))
        status = clean(r.get(status_col))
        if not legacy or not canon:
            continue
        if canon not in target_ids:
            raw[legacy].append((canon, BLOCKED, status))
        elif status and status.lower() not in {"approved", "active", "pending review", "pending"}:
            raw[legacy].append((canon, BLOCKED, status))
        elif status.lower() in {"pending review", "pending"}:
            raw[legacy].append((canon, DECISION, status))
        else:
            raw[legacy].append((canon, SAFE, status or "Approved"))
    result = {}
    classifications = {}
    for legacy, vals in raw.items():
        unique = {v[0] for v in vals}
        statuses = {v[1] for v in vals}
        if len(unique) != 1:
            classifications[(legacy, table_name)] = BLOCKED
            continue
        canon = next(iter(unique))
        classification = BLOCKED if BLOCKED in statuses else DECISION if DECISION in statuses else SAFE
        result[legacy] = {"canonical": canon, "classification": classification, "status": "; ".join(sorted({v[2] for v in vals if v[2]}))}
        classifications[(legacy, table_name)] = classification
    return result, classifications


def next_id(actions):
    return f"A{len(actions) + 1:05d}"


def add_action(actions, finding, classification, category, **kwargs):
    will = "Yes" if classification == SAFE else "No"
    actions.append(PlanAction(next_id(actions), finding, classification, category, WillApply=will, **kwargs))


def analyze_worksheet_rename(wb, actions):
    old = "15 LL_Geographic "
    new = "15 LL_Geographic"
    if old not in wb.sheetnames:
        return
    formulas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and old in cell.value:
                    formulas.append(f"{ws.title}!{cell.coordinate}")
    if new in wb.sheetnames:
        add_action(actions, "F00001", BLOCKED, "Worksheet Rename", Worksheet=old, ProposedValue=new, Reason="Target worksheet already exists.", Risk="Rename would collide.")
    else:
        add_action(actions, "F00001", SAFE, "Worksheet Rename", Worksheet=old, OriginalValue=old, ProposedValue=new, Reason="Trailing-space worksheet name has a unique trimmed target.", Risk="Formula/defined-name references recorded before apply.", Notes="References: " + "; ".join(formulas[:20]))


def analyze_missing_catalog_tables(wb, actions, lookup_decisions):
    aliases = {
        "tblInventoryLocations": ["tblInventoryLocation", "InventoryLocations"],
        "tblEmployeeRoles": ["tblEmployeeRole", "EmployeeRoles"],
        "tblEmployeeStatus": ["tblEmployeeStatuses", "EmployeeStatus"],
    }
    _, _, headers, rows = locate_catalog(wb)
    existing = set(table_map(wb))
    for r in rows:
        name = clean(r.get("LookupTableName"))
        if name not in aliases:
            continue
        catalog_id = clean(r.get("LookupCatalogID"))
        status = clean(r.get("Status"))
        table_exists = clean(r.get("TableExists"))
        found_alias = next((a for a in aliases[name] if a in existing), "")
        if name in existing:
            continue
        if found_alias:
            proposed = f"Align catalog {name} to existing {found_alias}"
            reason = "A similar physical table exists."
        else:
            proposed = "Change Status to Draft and TableExists to FALSE, or create an empty governed table after approval."
            reason = "Catalog row appears active/planned but no physical table or data source exists."
        lookup_decisions.append({
            "LookupCatalogID": catalog_id,
            "LookupTableName": name,
            "CurrentStatus": status,
            "CurrentTableExists": table_exists,
            "ProposedAction": proposed,
            "Reason": reason,
            "UserDecision": "",
        })
        add_action(actions, f"Missing-{name}", DECISION, "Lookup Catalog Decision", TableName=name, RecordID=catalog_id, OriginalValue=f"{status}/{table_exists}", ProposedValue=proposed, Reason=reason, Risk="Changing catalog metadata or creating empty tables requires governance approval.")


def analyze_fk_mappings(wb, actions, mapping_report):
    tables = table_map(wb)
    mfg_map, _ = build_crosswalk_mapping(wb, "tblManufacturerIDCrosswalk", ["LegacyManufacturerID"], ["CanonicalManufacturerID"], "tblManufacturers", "ManufacturerID")
    fam_map, _ = build_crosswalk_mapping(wb, "tblDeviceFamilyIDCrosswalk", ["LegacyDeviceFamilyID"], ["CanonicalDeviceFamilyID"], "tblDeviceFamilies", "DeviceFamilyID")
    targets = [
        ("F00005", "tblDeviceModels", "ManufacturerID", mfg_map, "tblManufacturerIDCrosswalk"),
        ("F00008", "tblDeviceFamilyIDCrosswalk", "ManufacturerID", mfg_map, "tblManufacturerIDCrosswalk"),
        ("F00006", "tblDeviceModels", "DeviceFamilyID", fam_map, "tblDeviceFamilyIDCrosswalk"),
    ]
    for finding, table_name, field, mapping, source in targets:
        if table_name not in tables:
            continue
        ws, table = tables[table_name]
        headers, rows = read_table(ws, table)
        if field not in headers:
            continue
        idx = headers.index(field)
        pk_idx = next((i for i, h in enumerate(headers) if h.endswith("ID")), None)
        for r in rows:
            old = clean(r.get(field))
            if not old:
                continue
            if old in mapping:
                info = mapping[old]
                classification = info["classification"]
                new = info["canonical"]
            elif re.fullmatch(r"[A-Z]+-?\d{1,5}", old):
                classification = BLOCKED
                new = ""
            else:
                continue
            mapping_report.append({"Worksheet": ws.title, "TableName": table_name, "Row": r["__rownum"], "OriginalValue": old, "ProposedValue": new, "MappingSource": source, "MappingStatus": mapping.get(old, {}).get("status", "Missing"), "Classification": classification})
            add_action(actions, finding, classification, "Foreign Key Mapping", Worksheet=ws.title, TableName=table_name, CellOrRange=f"{r['__cells'][idx].coordinate}", RecordID=clean(r.get(headers[pk_idx])) if pk_idx is not None else "", FieldName=field, OriginalValue=old, ProposedValue=new, MappingSource=source, Reason="Authoritative crosswalk mapping." if classification != BLOCKED else "No unique approved crosswalk mapping exists.", Risk="Only SAFE_AUTOMATIC mappings are applied.")


def analyze_multivalue_fk(wb, actions, rows_out):
    tables = table_map(wb)
    if "tblInventoryStatus" not in tables:
        return
    ws, table = tables["tblInventoryStatus"]
    headers, rows = read_table(ws, table)
    if "PartStockStatusID" not in headers:
        return
    part_ids = set()
    if "tblPartStockStatus" in tables:
        pws, pt = tables["tblPartStockStatus"]
        ph, pr = read_table(pws, pt)
        pk = next((h for h in ph if h.endswith("ID")), "")
        part_ids = {clean(r.get(pk)) for r in pr}
    pk = next((h for h in headers if h.endswith("ID")), "")
    for r in rows:
        val = clean(r.get("PartStockStatusID"))
        if any(d in val for d in [";", ",", "|", "\n"]):
            parsed = [clean(x) for x in re.split(r"[;,|\n]+", val) if clean(x)]
            rows_out.append({"RowNumber": r["__rownum"], "RecordPrimaryKey": clean(r.get(pk)), "CurrentMultiValueCell": val, "ParsedIDs": "; ".join(parsed), "EachIDExists": "; ".join(f"{p}={p in part_ids}" for p in parsed), "RecommendedOptions": "select one status; create a junction table; redesign relationship", "UserDecision": ""})
            add_action(actions, "F00007", DECISION, "Multi-value Foreign Key", Worksheet=ws.title, TableName="tblInventoryStatus", RecordID=clean(r.get(pk)), FieldName="PartStockStatusID", OriginalValue=val, Reason="Single FK field contains multiple IDs.", Risk="Requires relationship-design decision.")


def analyze_device_model_crosswalk(wb, actions):
    tables = table_map(wb)
    readiness = ["# Device Model Crosswalk Population Readiness", "", "NOT READY", ""]
    if "tblDeviceModelIDCrosswalk" not in tables:
        readiness.append("- tblDeviceModelIDCrosswalk is not present.")
        return "\n".join(readiness) + "\n"
    ws, table = tables["tblDeviceModelIDCrosswalk"]
    headers, rows = read_table(ws, table)
    model_alias = find_col(headers, "DeviceModel", "DeviceModelName", "ModelName")
    method_alias = find_col(headers, "MatchMethod", "MappingMethod")
    if not model_alias:
        add_action(actions, "F00012", DECISION, "Schema Addition", Worksheet=ws.title, TableName=table.name, FieldName="DeviceModel", ProposedValue="Add DeviceModel column", Reason="No recognized model-name alias exists.", Risk="Schema changes require explicit approval.")
    if not method_alias:
        add_action(actions, "F00012", DECISION, "Schema Addition", Worksheet=ws.title, TableName=table.name, FieldName="MatchMethod", ProposedValue="Add MatchMethod column", Reason="No recognized mapping-method alias exists.", Risk="Schema changes require explicit approval.")
    readiness.append("- Device model crosswalk population is intentionally disabled by this remediation tool.")
    return "\n".join(readiness) + "\n"


def analyze_registry(wb, actions, registry_rows):
    if "20 - Lookup Registry" not in wb.sheetnames:
        return
    ws = wb["20 - Lookup Registry"]
    text = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                text.append(clean(cell.value))
    referenced = set(re.findall(r"tbl[A-Za-z0-9_]+", "\n".join(text)))
    current = set(table_map(wb))
    for name in sorted(current - referenced)[:200]:
        registry_rows.append({"IssueType": "Missing current table", "TableName": name, "CurrentLocation": table_map(wb)[name][0].title, "ProposedAction": "Review registry refresh after safe corrections.", "RequiresFlag": "--refresh-control-sheets"})
    if referenced - current:
        for name in sorted(referenced - current):
            registry_rows.append({"IssueType": "Removed reference", "TableName": name, "CurrentLocation": "", "ProposedAction": "Remove or mark stale if registry refresh is approved.", "RequiresFlag": "--refresh-control-sheets"})
    if registry_rows:
        add_action(actions, "F00020", DECISION, "Control Sheet Refresh", Worksheet="20 - Lookup Registry", Reason="Registry appears stale.", Risk="Refresh requires explicit --refresh-control-sheets flag.")


def create_plan(wb):
    actions: list[PlanAction] = []
    mapping_report = []
    lookup_decisions = []
    multivalue = []
    registry = []
    analyze_worksheet_rename(wb, actions)
    analyze_missing_catalog_tables(wb, actions, lookup_decisions)
    analyze_fk_mappings(wb, actions, mapping_report)
    analyze_multivalue_fk(wb, actions, multivalue)
    readiness_md = analyze_device_model_crosswalk(wb, actions)
    analyze_registry(wb, actions, registry)
    return actions, mapping_report, lookup_decisions, multivalue, readiness_md, registry


def write_csv(path, rows, headers):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def apply_safe_actions(source, out_wb_path, actions):
    shutil.copy2(source, out_wb_path)
    wb = load_workbook(out_wb_path, read_only=False, data_only=False)
    tables = table_map(wb)
    applied = 0
    for a in actions:
        if a.Classification == SAFE and a.Category == "Worksheet Rename":
            rename_action = a
            break
    else:
        rename_action = None
    if rename_action:
        wb["15 LL_Geographic "].title = "15 LL_Geographic"
        rename_action.Applied = "Yes"
        rename_action.Result = "Applied"
        applied += 1
    for a in actions:
        if a.Classification != SAFE or a.Category != "Foreign Key Mapping":
            continue
        ws, table = tables[a.TableName]
        headers, rows = read_table(ws, table)
        idx = headers.index(a.FieldName)
        for r in rows:
            cell = r["__cells"][idx]
            if cell.coordinate == a.CellOrRange and clean(cell.value) == a.OriginalValue:
                cell.value = a.ProposedValue
                a.Applied = "Yes"
                a.Result = "Applied"
                applied += 1
                break
    wb.save(out_wb_path)
    return applied


def run(args):
    source = args.workbook_path.resolve()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outdir = (args.output_dir or Path(f"migration_remediation_{stamp}")).resolve()
    outdir.mkdir(parents=True, exist_ok=True)
    log = []
    source_hash = sha256(source)
    (outdir / "source_workbook_sha256.txt").write_text(source_hash + "\n", encoding="utf-8")
    wb = load_workbook(source, read_only=False, data_only=False)
    actions, mapping_report, lookup_decisions, multivalue, readiness_md, registry = create_plan(wb)
    for a in actions:
        if args.plan_only or a.Classification != SAFE:
            a.WillApply = "No" if args.plan_only else a.WillApply
    write_csv(outdir / "remediation_plan.csv", [asdict(a) for a in actions], list(PlanAction.__dataclass_fields__))
    (outdir / "remediation_plan.json").write_text(json.dumps([asdict(a) for a in actions], indent=2), encoding="utf-8")
    write_csv(outdir / "safe_mapping_changes.csv", mapping_report, ["Worksheet", "TableName", "Row", "OriginalValue", "ProposedValue", "MappingSource", "MappingStatus", "Classification"])
    write_csv(outdir / "blocked_mappings.csv", [r for r in mapping_report if r["Classification"] == BLOCKED], ["Worksheet", "TableName", "Row", "OriginalValue", "ProposedValue", "MappingSource", "MappingStatus", "Classification"])
    write_csv(outdir / "decisions_required.csv", [asdict(a) for a in actions if a.Classification == DECISION], list(PlanAction.__dataclass_fields__))
    write_csv(outdir / "worksheet_rename_analysis.csv", [asdict(a) for a in actions if a.Category == "Worksheet Rename"], list(PlanAction.__dataclass_fields__))
    write_csv(outdir / "lookup_catalog_decisions.csv", lookup_decisions, ["LookupCatalogID", "LookupTableName", "CurrentStatus", "CurrentTableExists", "ProposedAction", "Reason", "UserDecision"])
    write_csv(outdir / "multi_value_fk_decisions.csv", multivalue, ["RowNumber", "RecordPrimaryKey", "CurrentMultiValueCell", "ParsedIDs", "EachIDExists", "RecommendedOptions", "UserDecision"])
    (outdir / "device_model_crosswalk_readiness.md").write_text(readiness_md, encoding="utf-8")
    write_csv(outdir / "registry_update_plan.csv", registry, ["IssueType", "TableName", "CurrentLocation", "ProposedAction", "RequiresFlag"])
    output_workbook = ""
    post_readiness = "NOT RUN"
    applied_count = 0
    if args.apply:
        output_name = f"{source.stem}_remediated_{stamp}{source.suffix}"
        output_path = outdir / output_name
        if output_path.resolve() == source:
            raise RuntimeError("Refusing to overwrite source workbook.")
        with tempfile.TemporaryDirectory() as tmp:
            temp_path = Path(tmp) / output_name
            applied_count = apply_safe_actions(source, temp_path, actions)
            out_wb = load_workbook(temp_path, read_only=False, data_only=False)
            if any("#REF!" in str(c.value) for ws in out_wb.worksheets for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")):
                raise RuntimeError("Validation failed: new workbook contains #REF! formulas.")
            shutil.move(str(temp_path), output_path)
        output_workbook = str(output_path)
        (outdir / "remediated_workbook_sha256.txt").write_text(sha256(output_path) + "\n", encoding="utf-8")
        post_dir = outdir / "post_remediation_audit"
        _, post_readiness = run_audit(output_path, post_dir)
        write_csv(outdir / "remediation_plan.csv", [asdict(a) for a in actions], list(PlanAction.__dataclass_fields__))
        (outdir / "remediation_plan.json").write_text(json.dumps([asdict(a) for a in actions], indent=2), encoding="utf-8")
    source_after = sha256(source)
    if source_after != source_hash:
        raise RuntimeError("Source workbook hash changed.")
    summary = [
        "# Migration Remediation Summary",
        "",
        f"Mode: {'apply' if args.apply else 'plan-only'}",
        f"Source: {source}",
        f"Source SHA-256 before: {source_hash}",
        f"Source SHA-256 after: {source_after}",
        f"Safe actions identified: {sum(1 for a in actions if a.Classification == SAFE)}",
        f"Safe actions applied: {applied_count}",
        f"Decisions required: {sum(1 for a in actions if a.Classification == DECISION)}",
        f"Blocked actions: {sum(1 for a in actions if a.Classification == BLOCKED)}",
        f"Output workbook: {output_workbook}",
        f"Post-remediation readiness: {post_readiness}",
    ]
    (outdir / "remediation_summary.md").write_text("\n".join(summary) + "\n", encoding="utf-8")
    (outdir / "execution_log.txt").write_text("\n".join(log + summary) + "\n", encoding="utf-8")
    return outdir, output_workbook, post_readiness, actions, applied_count


def main(argv=None):
    parser = argparse.ArgumentParser(description="Controlled Nocturnix migration draft remediation planner/apply tool.")
    parser.add_argument("workbook_path", type=Path)
    parser.add_argument("--output-dir", type=Path)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--plan-only", action="store_true")
    mode.add_argument("--apply", action="store_true")
    parser.add_argument("--refresh-control-sheets", action="store_true")
    parser.add_argument("--approve-schema-additions", action="store_true")
    parser.add_argument("--audit-script", type=Path, default=Path("audit_migration_draft.py"))
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args(argv)
    if not args.apply:
        args.plan_only = True
    try:
        outdir, output_workbook, post_readiness, actions, applied = run(args)
        print(f"Mode: {'apply' if args.apply else 'plan-only'}")
        print(f"Output Folder: {outdir}")
        print(f"Safe actions identified: {sum(1 for a in actions if a.Classification == SAFE)}")
        print(f"Safe actions applied: {applied}")
        print(f"Decisions required: {sum(1 for a in actions if a.Classification == DECISION)}")
        print(f"Blocked actions: {sum(1 for a in actions if a.Classification == BLOCKED)}")
        print(f"Output workbook: {output_workbook}")
        print(f"Post-remediation readiness: {post_readiness}")
        return 0
    except Exception as exc:
        print(f"Remediation failed: {exc}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
